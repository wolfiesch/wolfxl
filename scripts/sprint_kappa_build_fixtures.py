"""Sprint Κ Pod-γ — build pre-built .xlsb / .xls parity fixtures.

This script is run ONCE on a developer machine and the resulting binaries
are committed to the repo. It is *not* invoked by pytest — fixtures are
committed for stability and reproducibility across CI runners.

Two-phase generation:

1. ``.xls`` fixtures are produced by writing a source ``.xlsx`` with openpyxl
   and converting it via LibreOffice headless (filter ``MS Excel 97``). The
   set: ``numbers``, ``strings``, ``dates``, ``formulas``, ``multisheet``.

2. ``.xlsb`` fixtures: LibreOffice 26.x does NOT expose a working xlsb export
   filter (the named filter exists but write returns ``Error Area:Io
   Class:Parameter Code:26``). xlsxwriter / openpyxl don't write xlsb either.
   The remaining sanctioned route (per Sprint Κ spec) is to vendor real-world
   xlsb fixtures from the upstream calamine project, which uses the MIT
   license. We copy 5 of calamine's ``tests/*.xlsb`` files with attribution,
   chosen to roughly span numbers / dates / multi-sheet / mixed content.

3. Optional long-tail ``.xlsb`` fixtures are vendored from ExcelGen's
   MIT-licensed ``samples/*.xlsb`` corpus. These cover features that the
   compact calamine set does not, including tables, data validations,
   conditional formatting, merged ranges, drawing-backed images, shared
   formulas that reference defined names, and lookup / hyperlink formulas.

Usage::

    # one-time clone of upstream calamine for its xlsb test corpus
    git clone https://github.com/tafia/calamine.git /tmp/calamine
    git clone https://github.com/mbleron/ExcelGen.git /tmp/ExcelGen
    python3 scripts/sprint_kappa_build_fixtures.py /tmp/calamine /tmp/ExcelGen

Requires:
    - openpyxl (writes the source xlsx for the xls path)
    - LibreOffice (``soffice`` on PATH, or ``/Applications/...``)
    - The path to a local clone of https://github.com/tafia/calamine, passed
      as the first arg, supplying the compact xlsb corpus.
    - Optional path to https://github.com/mbleron/ExcelGen, passed as the
      second arg, supplying long-tail xlsb samples.
"""

from __future__ import annotations

import datetime as dt
import os
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

import openpyxl  # noqa: F401 — verifies it's installed
from openpyxl import Workbook

REPO_ROOT = Path(__file__).resolve().parent.parent
XLSB_DIR = REPO_ROOT / "tests" / "parity" / "fixtures" / "xlsb"
XLS_DIR = REPO_ROOT / "tests" / "parity" / "fixtures" / "xls"

# Calamine fixture name -> wolfxl fixture name. Picked to cover a variety
# of xlsb content shapes (dates, mixed types, multi-sheet, edge cases).
CALAMINE_XLSB_PICKS: tuple[tuple[str, str], ...] = (
    ("date.xlsb", "dates.xlsb"),
    ("any_sheets.xlsb", "multisheet.xlsb"),
    ("issue_419.xlsb", "numbers.xlsb"),
    ("issue_186.xlsb", "strings.xlsb"),
    ("issue127.xlsb", "formulas.xlsb"),
)

EXCELGEN_XLSB_PICKS: tuple[tuple[str, str], ...] = (
    ("test-dataval.xlsb", "data-validations-and-tables.xlsb"),
    ("cond-formatting.xlsb", "conditional-formatting.xlsb"),
    ("merged-cells.xlsb", "merged-cells.xlsb"),
    ("style-showcase.xlsb", "style-showcase.xlsb"),
    ("test-image-2.xlsb", "image-drawing.xlsb"),
    ("test-formula.xlsb", "formulas-and-names.xlsb"),
    ("test-links.xlsb", "links-and-hyperlink-formulas.xlsb"),
    ("test-image-1.xlsb", "lookup-formulas-and-table.xlsb"),
)


def _soffice() -> str:
    """Locate the LibreOffice binary."""
    for cand in (
        "soffice",
        "/Applications/LibreOffice.app/Contents/MacOS/soffice",
        "/usr/bin/soffice",
        "/usr/local/bin/soffice",
    ):
        path = shutil.which(cand) or (cand if os.path.exists(cand) else None)
        if path:
            return path
    raise SystemExit(
        "LibreOffice (soffice) not found. Install via "
        "`brew install --cask libreoffice` on macOS."
    )


def _build_numbers(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 1
    ws["A2"] = 2.5
    ws["A3"] = 1e10
    ws["A4"] = 0.5
    ws["A4"].number_format = "0%"
    ws["A5"] = 100
    ws["A5"].number_format = "$#,##0"
    wb.save(path)


def _build_strings(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "hello"
    ws["A2"] = "münchen"
    ws["A3"] = "line1\nline2"
    ws["A4"] = ""
    wb.save(path)


def _build_dates(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = dt.date(2024, 1, 15)
    ws["A2"] = dt.datetime(2024, 1, 15, 14, 30, 0)
    ws["A3"] = dt.time(14, 30, 0)
    wb.save(path)


def _build_formulas(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 1
    ws["A2"] = 2
    ws["A3"] = "=A1+A2"
    ws["B1"] = "=SUM(A1:A2)"
    wb.save(path)


def _build_multisheet(path: Path) -> None:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1["A1"] = 1
    ws1["A2"] = 2
    ws1["A3"] = 3

    ws2 = wb.create_sheet("Sheet2")
    ws2["A1"] = "alpha"
    ws2["A2"] = "beta"
    ws2["A3"] = "gamma"

    wb.create_sheet("Sheet3")  # empty
    wb.save(path)


XLS_BUILDERS: tuple[tuple[str, callable], ...] = (
    ("numbers", _build_numbers),
    ("strings", _build_strings),
    ("dates", _build_dates),
    ("formulas", _build_formulas),
    ("multisheet", _build_multisheet),
)


def _convert_xls(soffice: str, src_xlsx: Path, dst_dir: Path) -> Path:
    """LibreOffice headless: xlsx -> xls (BIFF8)."""
    dst_dir.mkdir(parents=True, exist_ok=True)
    subprocess.run(
        [
            soffice,
            "--headless",
            "--convert-to",
            "xls",
            "--outdir",
            str(dst_dir),
            str(src_xlsx),
        ],
        check=True,
        capture_output=True,
    )
    produced = dst_dir / f"{src_xlsx.stem}.xls"
    if not produced.exists():
        raise RuntimeError(f"LibreOffice did not produce {produced}")
    return produced


def _copy_xlsb_from_calamine(calamine_root: Path) -> None:
    """Vendor 5 xlsb fixtures from the upstream calamine MIT-licensed corpus."""
    src_dir = calamine_root / "tests"
    if not src_dir.exists():
        raise SystemExit(
            f"calamine tests/ dir not found at {src_dir}. "
            f"Pass the path to a local clone of github.com/tafia/calamine "
            f"as argv[1]."
        )
    XLSB_DIR.mkdir(parents=True, exist_ok=True)
    for src_name, dst_name in CALAMINE_XLSB_PICKS:
        src = src_dir / src_name
        if not src.exists():
            raise SystemExit(f"missing upstream fixture {src}")
        dst = XLSB_DIR / dst_name
        shutil.copy2(src, dst)
        print(f"  {src.name} -> xlsb/{dst.name} ({dst.stat().st_size} bytes)")


def _copy_xlsb_from_excelgen(excelgen_root: Path) -> None:
    """Vendor long-tail xlsb fixtures from ExcelGen's MIT-licensed samples."""
    src_dir = excelgen_root / "samples"
    if not src_dir.exists():
        raise SystemExit(
            f"ExcelGen samples/ dir not found at {src_dir}. "
            f"Pass the path to a local clone of github.com/mbleron/ExcelGen "
            f"as argv[2]."
        )
    dst_dir = XLSB_DIR / "excelgen"
    dst_dir.mkdir(parents=True, exist_ok=True)
    for src_name, dst_name in EXCELGEN_XLSB_PICKS:
        src = src_dir / src_name
        if not src.exists():
            raise SystemExit(f"missing upstream ExcelGen fixture {src}")
        dst = dst_dir / dst_name
        shutil.copy2(src, dst)
        print(
            f"  {src.name} -> xlsb/excelgen/{dst.name} "
            f"({dst.stat().st_size} bytes)"
        )


def main(argv: list[str]) -> int:
    soffice = _soffice()
    XLSB_DIR.mkdir(parents=True, exist_ok=True)
    XLS_DIR.mkdir(parents=True, exist_ok=True)

    # ----- xls via LibreOffice -----
    print("== Building .xls fixtures via LibreOffice ==")
    with tempfile.TemporaryDirectory(prefix="wolfxl-kappa-") as tmpdir:
        tmp = Path(tmpdir)
        for name, builder in XLS_BUILDERS:
            src = tmp / f"{name}.xlsx"
            builder(src)
            xls = _convert_xls(soffice, src, tmp)
            shutil.move(str(xls), str(XLS_DIR / f"{name}.xls"))
            print(f"  {name}.xls ({(XLS_DIR / f'{name}.xls').stat().st_size} bytes)")

    # ----- xlsb vendored from calamine -----
    if len(argv) > 1:
        calamine_root = Path(argv[1]).expanduser().resolve()
        print(f"\n== Vendoring .xlsb fixtures from {calamine_root} ==")
        _copy_xlsb_from_calamine(calamine_root)
    else:
        print(
            "\n[skip] no calamine path passed; xlsb fixtures already on "
            "disk will not be replaced. Pass /path/to/calamine to refresh."
        )

    if len(argv) > 2:
        excelgen_root = Path(argv[2]).expanduser().resolve()
        print(f"\n== Vendoring long-tail .xlsb fixtures from {excelgen_root} ==")
        _copy_xlsb_from_excelgen(excelgen_root)
    else:
        print(
            "\n[skip] no ExcelGen path passed; long-tail xlsb fixtures already "
            "on disk will not be replaced. Pass /path/to/ExcelGen to refresh."
        )

    total = sum(
        p.stat().st_size
        for p in (*XLSB_DIR.rglob("*.xlsb"), *XLS_DIR.glob("*.xls"))
    )
    print(f"\nTotal fixture bytes: {total}")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
