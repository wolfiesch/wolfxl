#!/usr/bin/env python3
"""Generate focused P1 OOXML fidelity fixtures.

These fixtures are intentionally small and feature-dense. They cover P1
surfaces that are easy to miss when only exercising charts, pivots, slicers,
and external links.
"""

from __future__ import annotations

import argparse
import tempfile
import zipfile
from pathlib import Path
from xml.etree import ElementTree

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

CT_NS = "http://schemas.openxmlformats.org/package/2006/content-types"
REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"


def generate_openpyxl_p1_fixture(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "P1 Surface"

    rows = [
        ("Region", "Amount", "Status", "Reviewed"),
        ("North", 10, "Open", ""),
        ("South", 20, "Closed", ""),
        ("West", 30, "Open", ""),
    ]
    for row in rows:
        ws.append(row)

    table = Table(displayName="P1Table", ref="A1:D4")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    ws["F1"] = "Structured total"
    ws["F2"] = "=SUM(P1Table[Amount])"

    validation = DataValidation(
        type="list",
        formula1='"Open,Closed"',
        allow_blank=False,
        showErrorMessage=True,
    )
    validation.error = "Choose Open or Closed"
    validation.errorTitle = "Invalid status"
    ws.add_data_validation(validation)
    validation.add("C2:C4")

    ws["A1"].comment = Comment("P1 legacy comment", "wolfxl")
    ws.freeze_panes = "A2"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.print_options.horizontalCentered = True
    ws.oddHeader.center.text = "P1 fidelity fixture"
    ws.protection.sheet = True

    wb.security.lockStructure = True
    wb.security.workbookPassword = "DAA7"
    wb.create_named_range("P1ReviewRange", ws, "$A$1:$D$4")
    wb.save(path)
    _inject_custom_xml(path)


def _inject_custom_xml(path: Path) -> None:
    replacements = {
        "customXml/item1.xml": """<?xml version="1.0" encoding="UTF-8"?>
<wolfxl:p1Fidelity xmlns:wolfxl="https://wolfxl.local/fidelity">
  <wolfxl:source>openpyxl-targeted-fixture</wolfxl:source>
</wolfxl:p1Fidelity>""".encode(),
        "customXml/itemProps1.xml": """<?xml version="1.0" encoding="UTF-8"?>
<ds:datastoreItem ds:itemID="{11111111-2222-3333-4444-555555555555}"
  xmlns:ds="http://schemas.openxmlformats.org/officeDocument/2006/customXml">
  <ds:schemaRefs/>
</ds:datastoreItem>""".encode(),
        "customXml/_rels/item1.xml.rels": """<?xml version="1.0" encoding="UTF-8"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1"
    Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/customXmlProps"
    Target="itemProps1.xml"/>
</Relationships>""".encode(),
    }
    with zipfile.ZipFile(path) as archive:
        names = set(archive.namelist())
        if "[Content_Types].xml" in names:
            replacements["[Content_Types].xml"] = _with_custom_xml_content_types(
                archive.read("[Content_Types].xml")
            )
        if "_rels/.rels" in names:
            replacements["_rels/.rels"] = _with_custom_xml_root_rel(
                archive.read("_rels/.rels")
            )
    _replace_zip_entries(path, replacements)


def _replace_zip_entries(path: Path, replacements: dict[str, bytes]) -> None:
    with tempfile.NamedTemporaryFile(delete=False, suffix=path.suffix) as handle:
        tmp_path = Path(handle.name)
    try:
        with zipfile.ZipFile(path) as source, zipfile.ZipFile(
            tmp_path, "w", zipfile.ZIP_DEFLATED
        ) as dest:
            copied = set()
            for info in source.infolist():
                if info.filename in copied:
                    continue
                copied.add(info.filename)
                payload = replacements.get(info.filename)
                if payload is None:
                    payload = source.read(info.filename)
                dest.writestr(info, payload)
            for name, payload in replacements.items():
                if name not in copied:
                    dest.writestr(name, payload)
        tmp_path.replace(path)
    finally:
        tmp_path.unlink(missing_ok=True)


def _with_custom_xml_content_types(payload: bytes) -> bytes:
    root = ElementTree.fromstring(payload)
    existing = {
        node.attrib.get("PartName")
        for node in root
        if node.tag == f"{{{CT_NS}}}Override"
    }
    for part_name, content_type in (
        ("/customXml/item1.xml", "application/xml"),
        (
            "/customXml/itemProps1.xml",
            "application/vnd.openxmlformats-officedocument.customXmlProperties+xml",
        ),
    ):
        if part_name in existing:
            continue
        ElementTree.SubElement(
            root,
            f"{{{CT_NS}}}Override",
            {"PartName": part_name, "ContentType": content_type},
        )
    return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)


def _with_custom_xml_root_rel(payload: bytes) -> bytes:
    root = ElementTree.fromstring(payload)
    if any(
        node.attrib.get("Target") == "customXml/item1.xml"
        for node in root
        if node.tag == f"{{{REL_NS}}}Relationship"
    ):
        return payload
    rel_ids = [
        int(rel_id[3:])
        for node in root
        if (rel_id := node.attrib.get("Id", "")).startswith("rId")
        and rel_id[3:].isdigit()
    ]
    next_id = max(rel_ids, default=0) + 1
    ElementTree.SubElement(
        root,
        f"{{{REL_NS}}}Relationship",
        {
            "Id": f"rId{next_id}",
            "Type": "http://schemas.openxmlformats.org/officeDocument/2006/relationships/customXml",
            "Target": "customXml/item1.xml",
        },
    )
    return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()
    generate_openpyxl_p1_fixture(args.output)
    print(args.output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
