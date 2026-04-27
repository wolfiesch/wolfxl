"""Sprint Μ Pod-γ (RFC-046 §7) — RFC-035 §10 chart deep-clone lift.

Pre-Sprint Μ behavior: charts on a copied sheet were aliased — the
cloned drawing's nested rels file kept pointing at the SOURCE chart
XML, so any post-copy edit to the source's chart range bled through
to the copy.

Post-Sprint Μ behavior: ``Workbook.copy_worksheet(src)`` deep-clones
every chart referenced by the source. The cloned chart XML's
``<c:f>`` formulas have their sheet name re-pointed from the source
title to the destination title using the formula translator
(`wolfxl_formula::rename_sheet`).

The lift is **always-on** (no opt-in flag, unlike the Sprint Θ
``deep_copy_images`` knob).
"""
from __future__ import annotations

import re
import shutil
import zipfile
from pathlib import Path

import openpyxl
import pytest

from wolfxl import load_workbook


pytestmark = pytest.mark.rfc035 if hasattr(pytest.mark, "rfc035") else pytest.mark.usefixtures()


@pytest.fixture(autouse=True)
def _force_test_epoch(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("WOLFXL_TEST_EPOCH", "0")


# ---------------------------------------------------------------------------
# Fixture builders
# ---------------------------------------------------------------------------


def _make_chart_fixture(path: Path, sheet_title: str = "Template") -> None:
    """A1:B5 mini table + one BarChart with cell-range refs."""
    from openpyxl.chart import BarChart, Reference

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    for r in range(1, 6):
        ws.cell(r, 1, f"l{r}")
        ws.cell(r, 2, r * 10)
    ch = BarChart()
    ch.title = "Demo"
    data = Reference(ws, min_col=2, min_row=1, max_row=5)
    cats = Reference(ws, min_col=1, min_row=1, max_row=5)
    ch.add_data(data, titles_from_data=False)
    ch.set_categories(cats)
    ws.add_chart(ch, "D2")
    wb.save(path)


def _make_multi_chart_fixture(path: Path) -> None:
    """Source sheet with TWO bar charts — each must get a fresh chartN."""
    from openpyxl.chart import BarChart, Reference

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"
    for r in range(1, 6):
        ws.cell(r, 1, f"l{r}")
        ws.cell(r, 2, r * 10)
        ws.cell(r, 3, r * 100)
    ch1 = BarChart()
    ch1.add_data(Reference(ws, min_col=2, min_row=1, max_row=5))
    ws.add_chart(ch1, "E2")
    ch2 = BarChart()
    ch2.add_data(Reference(ws, min_col=3, min_row=1, max_row=5))
    ws.add_chart(ch2, "M2")
    wb.save(path)


def _make_cross_sheet_chart_fixture(path: Path) -> None:
    """Source has a chart referencing data on ANOTHER sheet (not the
    source). The cross-sheet ref must be PRESERVED after copy.
    """
    from openpyxl.chart import BarChart, Reference

    wb = openpyxl.Workbook()
    src = wb.active
    src.title = "Template"
    other = wb.create_sheet("Data")
    for r in range(1, 6):
        other.cell(r, 1, f"l{r}")
        other.cell(r, 2, r * 10)
    ch = BarChart()
    # Reference targets the OTHER sheet, not src.
    ch.add_data(Reference(other, min_col=2, min_row=1, max_row=5))
    src.add_chart(ch, "B2")
    wb.save(path)


def _make_image_and_chart_fixture(path: Path) -> None:
    """Source has BOTH an image and a chart in the same drawing — exercises
    Pod-β's image deep-clone interaction with Pod-γ's chart deep-clone.
    """
    from openpyxl.chart import BarChart, Reference
    from openpyxl.drawing.image import Image

    img_bytes = (
        b"\x89PNG\r\n\x1a\n"
        b"\x00\x00\x00\rIHDR"
        b"\x00\x00\x00\x01\x00\x00\x00\x01\x08\x06\x00\x00\x00\x1f\x15\xc4\x89"
        b"\x00\x00\x00\rIDATx\x9cc\xfc\x0f\x00\x00\x01\x01\x01\x00\x18\xdd\x8d\xb4"
        b"\x00\x00\x00\x00IEND\xaeB`\x82"
    )
    tmp_png = path.parent / "_tiny.png"
    tmp_png.write_bytes(img_bytes)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"
    for r in range(1, 6):
        ws.cell(r, 1, f"l{r}")
        ws.cell(r, 2, r * 10)
    img = Image(str(tmp_png))
    img.anchor = "F2"
    ws.add_image(img)
    ch = BarChart()
    ch.add_data(Reference(ws, min_col=2, min_row=1, max_row=5))
    ws.add_chart(ch, "M2")
    wb.save(path)
    tmp_png.unlink(missing_ok=True)


def _zip_listing(path: Path) -> list[str]:
    with zipfile.ZipFile(path, "r") as z:
        return sorted(z.namelist())


def _zip_read(path: Path, member: str) -> bytes:
    with zipfile.ZipFile(path, "r") as z:
        return z.read(member)


def _normalize_xml_text(s: str) -> str:
    """Replace XML entity-encoded apostrophes with bare ones so test
    assertions can compare against the human-readable form. Does NOT
    re-encode anything else.
    """
    return s.replace("&apos;", "'").replace("&quot;", '"')


# ---------------------------------------------------------------------------
# A — Single chart on copied sheet → deep-cloned, sheet-name re-pointed.
# ---------------------------------------------------------------------------


def test_copy_worksheet_with_bar_chart_deep_clones(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_chart_fixture(src, sheet_title="Template")

    wb = load_workbook(src, modify=True)
    wb.copy_worksheet(wb["Template"])
    wb.save(dst)

    entries = _zip_listing(dst)
    chart_files = sorted(e for e in entries if e.startswith("xl/charts/chart"))
    # Source had 1 chart; copy must produce a SECOND chart part.
    assert len(chart_files) == 2, f"expected 2 chart parts; got {chart_files}"

    # Read both chart bodies. The cloned one must reference the
    # destination sheet ("Template Copy"), the original one keeps the
    # source sheet ("Template"). XML-entity apostrophes (`&apos;`)
    # are re-emitted by quick-xml on write — normalize before
    # asserting.
    bodies = [_normalize_xml_text(_zip_read(dst, p).decode()) for p in chart_files]
    bodies_with_source = [
        b for b in bodies
        if "'Template'!" in b or "<f>Template!" in b
    ]
    bodies_with_copy = [
        b for b in bodies
        if "'Template Copy'!" in b or "<f>Template Copy!" in b
    ]
    assert bodies_with_source, (
        f"original chart should still reference 'Template'; got bodies={bodies}"
    )
    assert bodies_with_copy, (
        f"cloned chart should reference 'Template Copy'; got bodies={bodies}"
    )


# ---------------------------------------------------------------------------
# B — Modifying source after copy must NOT bleed through to the copy.
#     The behavioral test for deep-clone (vs. alias).
# ---------------------------------------------------------------------------


def test_copy_worksheet_with_chart_then_modify_source_does_not_affect_copy(
    tmp_path: Path,
) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_chart_fixture(src, sheet_title="Template")

    # Copy + mutate the source's data range in the SAME save.
    wb = load_workbook(src, modify=True)
    wb.copy_worksheet(wb["Template"])
    # Note: the original chart references 'Template'!$B$1:$B$5; the
    # cloned chart references 'Template Copy'!$B$1:$B$5. Both are
    # independent — mutating the source's B1 does not reach the copy
    # because deep-clone produced fresh chart XML for the new sheet.
    ws = wb["Template"]
    ws["B1"] = -7777
    wb.save(dst)

    # Sanity: source chart bytes contain the ORIGINAL sheet name; cloned
    # chart bytes contain the destination sheet name.
    chart_files = sorted(
        e for e in _zip_listing(dst) if e.startswith("xl/charts/chart")
    )
    assert len(chart_files) == 2
    bodies = [_normalize_xml_text(_zip_read(dst, p).decode()) for p in chart_files]
    sheets_referenced = []
    for b in bodies:
        if "'Template Copy'" in b:
            sheets_referenced.append("Template Copy")
        elif "'Template'" in b or "<f>Template!" in b:
            sheets_referenced.append("Template")
    assert sorted(sheets_referenced) == ["Template", "Template Copy"], (
        f"each chart should bind to a distinct sheet; got {sheets_referenced}"
    )


# ---------------------------------------------------------------------------
# C — Cross-sheet chart ref preserved (only source-sheet refs are
#     re-pointed; refs at any OTHER sheet pass through untouched).
# ---------------------------------------------------------------------------


def test_copy_worksheet_with_multi_sheet_referenced_chart(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_cross_sheet_chart_fixture(src)

    wb = load_workbook(src, modify=True)
    wb.copy_worksheet(wb["Template"])
    wb.save(dst)

    chart_files = sorted(
        e for e in _zip_listing(dst) if e.startswith("xl/charts/chart")
    )
    assert len(chart_files) == 2, f"expected 2 chart parts; got {chart_files}"
    bodies = [_normalize_xml_text(_zip_read(dst, p).decode()) for p in chart_files]
    # BOTH charts should still reference 'Data' (the unaffected
    # cross-sheet target). Neither should mention "Template".
    for b in bodies:
        assert "'Data'" in b or "<f>Data!" in b, (
            f"cross-sheet ref must be preserved; got body={b[:300]}"
        )
        assert "'Template'!" not in b and "<f>Template!" not in b, (
            f"chart should not reference source title; got body={b[:300]}"
        )


# ---------------------------------------------------------------------------
# D — Multiple charts on source → each gets a fresh chartN.
# ---------------------------------------------------------------------------


def test_copy_worksheet_with_multiple_charts_each_get_fresh_chartN(
    tmp_path: Path,
) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_multi_chart_fixture(src)

    wb = load_workbook(src, modify=True)
    wb.copy_worksheet(wb["Template"])
    wb.save(dst)

    chart_files = sorted(
        e for e in _zip_listing(dst) if e.startswith("xl/charts/chart")
    )
    # Source had 2 charts; the deep-clone added 2 more.
    assert len(chart_files) == 4, f"expected 4 chart parts; got {chart_files}"

    # All 4 chart suffixes must be distinct.
    suffixes = sorted(
        int(re.search(r"chart(\d+)\.xml$", p).group(1))
        for p in chart_files
    )
    assert len(set(suffixes)) == 4, suffixes


# ---------------------------------------------------------------------------
# E — Chart + image on same sheet: deep-clone both correctly.
#     deep_copy_images must be opt-in for images, but chart deep-clone
#     is always-on.
# ---------------------------------------------------------------------------


def test_copy_worksheet_with_image_and_chart_both_clone_correctly(
    tmp_path: Path,
) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_image_and_chart_fixture(src)

    wb = load_workbook(src, modify=True)
    # Default: deep_copy_images=False → images aliased, charts deep-cloned.
    wb.copy_worksheet(wb["Template"])
    wb.save(dst)

    entries = _zip_listing(dst)
    # Charts: 2 (1 source + 1 cloned).
    chart_files = [e for e in entries if e.startswith("xl/charts/chart")]
    assert len(chart_files) == 2, f"chart deep-clone failed; got {chart_files}"
    # Images: 1 (alias mode default).
    image_files = [e for e in entries if e.startswith("xl/media/image")]
    assert len(image_files) == 1, (
        f"alias mode should leave 1 image; got {image_files}"
    )
    # Drawings: 2 (one per sheet).
    drawing_files = [
        e for e in entries
        if re.match(r"^xl/drawings/drawing\d+\.xml$", e)
    ]
    assert len(drawing_files) == 2, (
        f"copy_worksheet must allocate a fresh drawing; got {drawing_files}"
    )


# ---------------------------------------------------------------------------
# F — Round-trip: emitted workbook structure is well-formed (ZIP-level
#     plus xml.etree parse). Full openpyxl load is gated by a separate,
#     pre-existing RFC-035 namespace-declaration limitation (workbook.xml's
#     `<sheet>` row needs `xmlns:r=` — tracked outside Pod-γ) so we
#     verify the chart-deep-clone outputs rather than the full load path.
# ---------------------------------------------------------------------------


def test_copy_worksheet_chart_deep_clone_emits_well_formed_outputs(
    tmp_path: Path,
) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_chart_fixture(src, sheet_title="Template")

    wb = load_workbook(src, modify=True)
    wb.copy_worksheet(wb["Template"])
    wb.save(dst)

    # Each chart XML must be parseable by xml.etree.
    import xml.etree.ElementTree as ET

    for entry in _zip_listing(dst):
        if entry.startswith("xl/charts/chart") and entry.endswith(".xml"):
            body = _zip_read(dst, entry)
            ET.fromstring(body)  # raises if malformed


# ===========================================================================
# Sprint Μ-prime Pod-γ′ — RFC-035 deep-clone for new chart families.
#
# The Pod-γ deep-clone logic in
# ``crates/wolfxl-structural/src/sheet_copy.rs`` operates on chart XML
# bytes opaquely (regex-style cell-range replacement on
# ``<c:f>...</c:f>`` text nodes), so it should be agnostic to chart
# kind. These tests pin that behaviour for the 8 new families
# Sprint Μ-prime ships, with three representative samples:
# Bar3D, Stock (4 OHLC refs → 4 deep-cloned series), and Surface.
# ===========================================================================


def _make_bar3d_chart_fixture(path: Path, sheet_title: str = "Template") -> None:
    """Source fixture with a single 3D bar chart referencing the source
    sheet's data range. We seed via openpyxl so the .xlsx is realistic."""
    from openpyxl.chart import BarChart3D, Reference

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    for r in range(1, 6):
        ws.cell(r, 1, f"l{r}")
        ws.cell(r, 2, r * 10)
    ch = BarChart3D()
    ch.add_data(Reference(ws, min_col=2, min_row=1, max_row=5))
    ws.add_chart(ch, "D2")
    wb.save(path)


def _make_stock_chart_fixture(path: Path, sheet_title: str = "Template") -> None:
    """Source fixture with a Stock (HLC/OHLC) chart — 4 OHLC series so
    the deep-clone touches multiple ``<c:f>`` refs on the same chart."""
    from openpyxl.chart import StockChart
    from openpyxl.chart.reference import Reference

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.cell(1, 1, "label")
    ws.cell(1, 2, "open")
    ws.cell(1, 3, "high")
    ws.cell(1, 4, "low")
    ws.cell(1, 5, "close")
    for r in range(2, 6):
        ws.cell(r, 1, f"l{r-1}")
        ws.cell(r, 2, 10 + r)
        ws.cell(r, 3, 12 + r)
        ws.cell(r, 4, 9 + r)
        ws.cell(r, 5, 11 + r)
    ch = StockChart()
    ch.add_data(
        Reference(ws, min_col=2, max_col=5, min_row=1, max_row=5),
        titles_from_data=True,
    )
    ch.set_categories(
        Reference(ws, min_col=1, max_col=1, min_row=2, max_row=5),
    )
    ws.add_chart(ch, "G2")
    wb.save(path)


def _make_surface_chart_fixture(path: Path, sheet_title: str = "Template") -> None:
    """Source fixture with a 2D SurfaceChart on a 4×4 numeric grid."""
    from openpyxl.chart import SurfaceChart
    from openpyxl.chart.reference import Reference

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    for r in range(1, 5):
        for c in range(1, 5):
            ws.cell(r, c, (r * c) + 1)
    ch = SurfaceChart()
    ch.add_data(
        Reference(ws, min_col=1, max_col=4, min_row=1, max_row=4),
    )
    ws.add_chart(ch, "F2")
    wb.save(path)


def test_copy_worksheet_with_bar3d_chart_deep_clones(tmp_path: Path) -> None:
    """RFC-035 deep-clone is chart-kind-agnostic — Bar3D charts must
    survive the lift like flat BarCharts do."""
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_bar3d_chart_fixture(src, sheet_title="Template")

    wb = load_workbook(src, modify=True)
    wb.copy_worksheet(wb["Template"])
    wb.save(dst)

    chart_files = sorted(
        e for e in _zip_listing(dst) if e.startswith("xl/charts/chart")
    )
    assert len(chart_files) == 2, f"expected 2 chart parts; got {chart_files}"
    bodies = [_normalize_xml_text(_zip_read(dst, p).decode()) for p in chart_files]
    sources = [b for b in bodies if "'Template'!" in b or "<f>Template!" in b]
    copies = [b for b in bodies if "'Template Copy'!" in b or "<f>Template Copy!" in b]
    assert sources, f"original Bar3D chart should still reference 'Template'; bodies={bodies}"
    assert copies, f"cloned Bar3D chart should reference 'Template Copy'; bodies={bodies}"


def test_copy_worksheet_with_stock_chart_deep_clones(tmp_path: Path) -> None:
    """4 OHLC series → 4 deep-cloned <c:f> refs on the cloned chart.
    Validates the sheet-name re-pointing handles multi-series charts."""
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_stock_chart_fixture(src, sheet_title="Template")

    wb = load_workbook(src, modify=True)
    wb.copy_worksheet(wb["Template"])
    wb.save(dst)

    chart_files = sorted(
        e for e in _zip_listing(dst) if e.startswith("xl/charts/chart")
    )
    assert len(chart_files) == 2, f"expected 2 chart parts; got {chart_files}"

    # Locate the cloned chart body (references "Template Copy").
    bodies = [_normalize_xml_text(_zip_read(dst, p).decode()) for p in chart_files]
    cloned = [b for b in bodies if "'Template Copy'!" in b or "<f>Template Copy!" in b]
    assert cloned, f"cloned chart should re-point to 'Template Copy'; bodies={[b[:200] for b in bodies]}"

    # The cloned StockChart should retain at least 4 ``<c:f>`` refs
    # (one per OHLC series). Some emitters embed the categories ref
    # in addition, so accept >= 4.
    cloned_body = cloned[0]
    n_refs = len(re.findall(r"<c:f>[^<]+</c:f>", cloned_body))
    if n_refs == 0:
        # Fallback for emitters that elide the c: prefix.
        n_refs = len(re.findall(r"<f>[^<]+</f>", cloned_body))
    assert n_refs >= 4, (
        f"expected >= 4 deep-cloned series refs in stock chart; "
        f"got {n_refs}; body[:500]={cloned_body[:500]}"
    )


def test_copy_worksheet_with_surface_chart_deep_clones(tmp_path: Path) -> None:
    """SurfaceChart on a 4x4 grid — deep-clone must produce an
    independent chart part with sheet refs re-pointed to the copy."""
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_surface_chart_fixture(src, sheet_title="Template")

    wb = load_workbook(src, modify=True)
    wb.copy_worksheet(wb["Template"])
    wb.save(dst)

    chart_files = sorted(
        e for e in _zip_listing(dst) if e.startswith("xl/charts/chart")
    )
    assert len(chart_files) == 2, f"expected 2 chart parts; got {chart_files}"
    bodies = [_normalize_xml_text(_zip_read(dst, p).decode()) for p in chart_files]
    sources = [b for b in bodies if "'Template'!" in b or "<f>Template!" in b]
    copies = [b for b in bodies if "'Template Copy'!" in b or "<f>Template Copy!" in b]
    assert sources, "original SurfaceChart must still reference 'Template'"
    assert copies, "cloned SurfaceChart must reference 'Template Copy'"
