"""Per-data-point chart override round-trip tests (G16).

These tests pin down the `<c:dPt>` (per-point fill / line / explosion
override) path end-to-end through the wolfxl writer and back through
openpyxl's reader. The dict bridge `Series.dPt` -> `data_points` ->
Rust `DataPoint` -> `<c:dPt><c:spPr>` is shared across chart kinds,
but each chart type has a slightly different series schema in
``attribute_mapping``, so we exercise bar, line, and a multi-series
case as separate fixtures.

Companion oracle probe: ``charts_pivot_chart_per_point`` in
``tests/test_openpyxl_compat_oracle.py`` covers the pivot-source path.
"""

from __future__ import annotations

import zipfile
from pathlib import Path

import openpyxl as _opx
import pytest

import wolfxl
from wolfxl.chart import BarChart, LineChart, Reference
from wolfxl.chart.marker import DataPoint
from wolfxl.chart.shapes import GraphicalProperties


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _seed_xy(ws, rows: list[tuple]) -> None:
    """Append ``[(header_x, header_y), (x1, y1), ...]`` rows to ``ws``."""
    for row in rows:
        ws.append(row)


def _read_chart(path: Path):
    wb = _opx.load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    charts = list(getattr(ws, "_charts", []))
    assert charts, f"openpyxl found no charts in {path}"
    return charts[0]


# ---------------------------------------------------------------------------
# 1. Bar chart, single series, single point overridden
# ---------------------------------------------------------------------------


def test_bar_chart_single_data_point_override(tmp_path: Path) -> None:
    wb = wolfxl.Workbook()
    ws = wb.active
    _seed_xy(ws, [("x", "y"), (1, 10), (2, 20), (3, 30)])

    chart = BarChart()
    chart.add_data(
        Reference(ws, min_col=2, min_row=1, max_row=4),
        titles_from_data=True,
    )

    chart.series[0].dPt = [
        DataPoint(idx=1, spPr=GraphicalProperties(solidFill="00FF00"))
    ]
    ws.add_chart(chart, "D2")

    out = tmp_path / "bar_per_point.xlsx"
    wb.save(out)

    # Raw XML sanity (cheap, anchors the failure if dict bridge breaks)
    chart_xml = zipfile.ZipFile(out).read("xl/charts/chart1.xml").decode()
    assert "<c:dPt>" in chart_xml
    assert '<c:idx val="1"/>' in chart_xml
    assert "00FF00" in chart_xml

    ref_chart = _read_chart(out)
    dpt = list(ref_chart.series[0].dPt)
    assert len(dpt) == 1
    assert dpt[0].idx == 1
    assert dpt[0].spPr is not None


# ---------------------------------------------------------------------------
# 2. Line chart, two points overridden on the same series
# ---------------------------------------------------------------------------


def test_line_chart_multiple_points_override(tmp_path: Path) -> None:
    wb = wolfxl.Workbook()
    ws = wb.active
    _seed_xy(
        ws,
        [
            ("x", "y"),
            (1, 5),
            (2, 10),
            (3, 15),
            (4, 20),
            (5, 25),
        ],
    )

    chart = LineChart()
    chart.add_data(
        Reference(ws, min_col=2, min_row=1, max_row=6),
        titles_from_data=True,
    )

    chart.series[0].dPt = [
        DataPoint(idx=0, spPr=GraphicalProperties(solidFill="FF0000")),
        DataPoint(idx=3, spPr=GraphicalProperties(solidFill="0000FF")),
    ]
    ws.add_chart(chart, "D2")

    out = tmp_path / "line_per_point.xlsx"
    wb.save(out)

    chart_xml = zipfile.ZipFile(out).read("xl/charts/chart1.xml").decode()
    assert chart_xml.count("<c:dPt>") == 2
    assert "FF0000" in chart_xml
    assert "0000FF" in chart_xml

    ref_chart = _read_chart(out)
    dpt = list(ref_chart.series[0].dPt)
    idxs = sorted(dp.idx for dp in dpt)
    assert idxs == [0, 3]


# ---------------------------------------------------------------------------
# 3. Multiple series, each with its own per-point override list
# ---------------------------------------------------------------------------


def test_bar_chart_multiseries_per_point_overrides(tmp_path: Path) -> None:
    wb = wolfxl.Workbook()
    ws = wb.active
    _seed_xy(
        ws,
        [
            ("x", "a", "b"),
            (1, 10, 100),
            (2, 20, 200),
            (3, 30, 300),
        ],
    )

    chart = BarChart()
    chart.add_data(
        Reference(ws, min_col=2, min_row=1, max_col=3, max_row=4),
        titles_from_data=True,
    )
    assert len(chart.series) == 2

    chart.series[0].dPt = [
        DataPoint(idx=2, spPr=GraphicalProperties(solidFill="AAAAAA"))
    ]
    chart.series[1].dPt = [
        DataPoint(idx=0, spPr=GraphicalProperties(solidFill="BBBBBB")),
        DataPoint(idx=1, spPr=GraphicalProperties(solidFill="CCCCCC")),
    ]
    ws.add_chart(chart, "E2")

    out = tmp_path / "multiseries_per_point.xlsx"
    wb.save(out)

    chart_xml = zipfile.ZipFile(out).read("xl/charts/chart1.xml").decode()
    # Three total <c:dPt> elements (one on series[0], two on series[1])
    assert chart_xml.count("<c:dPt>") == 3
    for hex_ in ("AAAAAA", "BBBBBB", "CCCCCC"):
        assert hex_ in chart_xml

    ref_chart = _read_chart(out)
    s0 = list(ref_chart.series[0].dPt)
    s1 = list(ref_chart.series[1].dPt)
    assert len(s0) == 1 and s0[0].idx == 2
    assert sorted(dp.idx for dp in s1) == [0, 1]


# ---------------------------------------------------------------------------
# 4. Bar chart with a pivot source — per-point still round-trips
# ---------------------------------------------------------------------------


def test_pivot_source_bar_chart_per_point(tmp_path: Path) -> None:
    """Smoke-test of the oracle probe contract — pivot vs non-pivot share
    the same per-point emit path; this guards against a regression where a
    pivot-source-specific code path drops `data_points`.
    """
    wb = wolfxl.Workbook()
    ws = wb.active
    _seed_xy(ws, [("x", "y"), (1, 10), (2, 20), (3, 30)])

    chart = BarChart()
    chart.add_data(
        Reference(ws, min_col=2, min_row=1, max_row=4),
        titles_from_data=True,
    )
    chart.pivot_source = ("PivotForChart", 0)
    chart.series[0].dPt = [
        DataPoint(idx=2, spPr=GraphicalProperties(solidFill="123456"))
    ]
    ws.add_chart(chart, "D2")

    out = tmp_path / "pivot_per_point.xlsx"
    wb.save(out)

    chart_xml = zipfile.ZipFile(out).read("xl/charts/chart1.xml").decode()
    assert "<c:pivotSource>" in chart_xml
    assert "<c:dPt>" in chart_xml
    assert "123456" in chart_xml
    # Pivot charts MUST carry <c:fmtId> on each series
    assert '<c:fmtId val="0"/>' in chart_xml

    ref_chart = _read_chart(out)
    dpt = list(ref_chart.series[0].dPt)
    assert len(dpt) == 1
    assert dpt[0].idx == 2


if __name__ == "__main__":  # pragma: no cover
    pytest.main([__file__, "-v"])
