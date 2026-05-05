"""Sprint Θ Pod-C3 — calcChain.xml rebuild on save.

Modify-mode and write-mode coverage for the new behaviour: on every
save that performs at least one operation (modify mode) or any save
at all (write mode), wolfxl writes a fresh ``xl/calcChain.xml`` whose
``<c>`` entries cover every formula cell in the workbook.

Edge cases:
- A workbook with NO formulas → ``xl/calcChain.xml`` is omitted (write
  mode) or removed if it was present in the source (modify mode with
  any op).
- Structural composition: ``insert_rows`` then save → calcChain
  references the SHIFTED cell coords, not the original.
- ``i`` correctly tracks the source sheet's tab position.
"""
from __future__ import annotations

import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path

import openpyxl
import pytest

from wolfxl import Workbook, load_workbook


pytestmark = pytest.mark.rfc035


@pytest.fixture(autouse=True)
def _force_test_epoch(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("WOLFXL_TEST_EPOCH", "0")


_NS = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}


def _read_calc_chain(path: Path) -> list[tuple[str, int]]:
    """Return the [(cell_ref, sheet_index_1based), ...] entries from
    ``xl/calcChain.xml`` inside the saved zip. Empty list if the file
    isn't there.
    """
    with zipfile.ZipFile(path, "r") as z:
        if "xl/calcChain.xml" not in z.namelist():
            return []
        bytes_ = z.read("xl/calcChain.xml")
    root = ET.fromstring(bytes_)
    out: list[tuple[str, int]] = []
    for c in root.findall("main:c", _NS):
        ref = c.get("r") or ""
        i = int(c.get("i") or "1")
        out.append((ref, i))
    return out


def _calc_chain_present(path: Path) -> bool:
    with zipfile.ZipFile(path, "r") as z:
        return "xl/calcChain.xml" in z.namelist()


def _inject_calc_chain(path: Path, calc_chain_xml: str) -> None:
    """Add or replace ``xl/calcChain.xml`` in an existing workbook zip."""
    rewritten = path.with_suffix(".rewritten.xlsx")
    with zipfile.ZipFile(path, "r") as src, zipfile.ZipFile(
        rewritten, "w", compression=zipfile.ZIP_DEFLATED
    ) as dst:
        for info in src.infolist():
            if info.filename == "xl/calcChain.xml":
                continue
            dst.writestr(info, src.read(info.filename))
        dst.writestr("xl/calcChain.xml", calc_chain_xml)
    rewritten.replace(path)


# ---------------------------------------------------------------------------
# Helpers — fixtures.
# ---------------------------------------------------------------------------


def _make_formula_fixture(path: Path) -> None:
    """Two sheets, formulas on each."""
    wb = openpyxl.Workbook()
    s1 = wb.active
    s1.title = "First"
    for r in range(1, 4):
        s1.cell(row=r, column=1, value=r * 10)
    s1["B1"] = "=SUM(A1:A3)"
    s1["B2"] = "=A1*2"

    s2 = wb.create_sheet("Second")
    s2["A1"] = 1
    s2["B5"] = "=A1+1"

    wb.save(path)


def _make_no_formula_fixture(path: Path) -> None:
    wb = openpyxl.Workbook()
    s = wb.active
    s.title = "Plain"
    s["A1"] = 1
    s["A2"] = 2
    wb.save(path)


# ---------------------------------------------------------------------------
# Modify-mode: load fixture with formulas → save → calcChain present.
# ---------------------------------------------------------------------------


def test_modify_mode_rebuilds_calc_chain(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_formula_fixture(src)

    wb = load_workbook(src, modify=True)
    # We need at least one op to bypass the no-op short-circuit.
    # Setting a value to its current value still queues a write —
    # use a proper edit.
    wb["First"]["C1"] = "edit"
    wb.save(dst)

    entries = _read_calc_chain(dst)
    assert entries, "calcChain.xml must exist with formulas after modify-mode save"
    refs_first = {ref for ref, i in entries if i == 1}
    refs_second = {ref for ref, i in entries if i == 2}
    assert "B1" in refs_first, f"missing B1 on First; entries={entries}"
    assert "B2" in refs_first, f"missing B2 on First; entries={entries}"
    assert "B5" in refs_second, f"missing B5 on Second; entries={entries}"


# ---------------------------------------------------------------------------
# Write-mode: build a workbook with formulas → save → calcChain present.
# ---------------------------------------------------------------------------


def test_write_mode_emits_calc_chain(tmp_path: Path) -> None:
    out = tmp_path / "out.xlsx"
    wb = Workbook()
    s = wb.active
    assert s is not None
    s.title = "S"
    s["A1"] = 1
    s["A2"] = 2
    s["B1"] = "=A1+A2"
    wb.save(out)

    entries = _read_calc_chain(out)
    assert entries, f"write-mode must emit calcChain; saw {entries}"
    refs = {ref for ref, _i in entries}
    assert "B1" in refs


# ---------------------------------------------------------------------------
# Edge case: workbook with NO formulas → calcChain.xml omitted.
# ---------------------------------------------------------------------------


def test_write_mode_no_formulas_omits_calc_chain(tmp_path: Path) -> None:
    """Document the contract: zero formulas → no calcChain.xml part."""
    out = tmp_path / "out.xlsx"
    wb = Workbook()
    s = wb.active
    assert s is not None
    s["A1"] = 1
    s["A2"] = "hello"
    wb.save(out)

    assert not _calc_chain_present(out), (
        "write-mode with zero formulas must omit calcChain.xml"
    )


def test_modify_mode_no_formulas_removes_calc_chain(tmp_path: Path) -> None:
    """Modify-mode + edit + zero-formula source → calcChain.xml removed.

    The fixture has no formulas; even with one edit, the saved file
    should not have calcChain.xml.
    """
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_no_formula_fixture(src)

    wb = load_workbook(src, modify=True)
    wb["Plain"]["B1"] = "edit"
    wb.save(dst)

    # Note: openpyxl-built fixtures don't ship with calcChain at all
    # for formula-less workbooks, so the assertion is "still absent".
    assert not _calc_chain_present(dst), (
        "modify-mode save with zero formulas must not emit calcChain.xml"
    )


# ---------------------------------------------------------------------------
# Sheet-index ``i`` correctly tracks tab position.
# ---------------------------------------------------------------------------


def test_sheet_index_matches_tab_position(tmp_path: Path) -> None:
    out = tmp_path / "out.xlsx"
    wb = Workbook()
    s1 = wb.active
    assert s1 is not None
    s1.title = "First"
    s1["A1"] = "=1+1"

    s2 = wb.create_sheet("Second")
    s2["B2"] = "=2+2"

    s3 = wb.create_sheet("Third")
    s3["C3"] = "=3+3"

    wb.save(out)
    entries = _read_calc_chain(out)
    by_ref = dict(entries)
    assert by_ref["A1"] == 1, f"A1 should map to sheet-index 1; entries={entries}"
    assert by_ref["B2"] == 2, f"B2 should map to sheet-index 2; entries={entries}"
    assert by_ref["C3"] == 3, f"C3 should map to sheet-index 3; entries={entries}"


# ---------------------------------------------------------------------------
# Structural composition: insert_rows then save → calcChain references
# the SHIFTED coords.
# ---------------------------------------------------------------------------


def test_insert_rows_then_save_shifts_calc_chain_refs(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"

    # Build a fixture where a formula sits at row 5; after inserting
    # 3 rows at row 2, the formula should now live at row 8.
    fwb = openpyxl.Workbook()
    fs = fwb.active
    fs.title = "Sheet"
    for r in range(1, 11):
        fs.cell(row=r, column=1, value=r)
    fs["B5"] = "=SUM(A1:A4)"
    fwb.save(src)

    wb = load_workbook(src, modify=True)
    wb["Sheet"].insert_rows(2, 3)
    wb.save(dst)

    entries = _read_calc_chain(dst)
    refs = {ref for ref, _ in entries}
    # Original B5 → after inserting 3 rows above, the shifted position
    # is B8.
    assert "B8" in refs, (
        f"insert_rows must shift calcChain refs; expected B8 in {refs}"
    )
    assert "B5" not in refs, (
        f"insert_rows must NOT leave the unshifted ref; got {refs}"
    )


def test_modify_mode_prunes_stale_calc_chain_and_preserves_ext_lst(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"

    fwb = openpyxl.Workbook()
    first = fwb.active
    first.title = "First"
    first["A1"] = 1
    first["A2"] = 2
    first["B1"] = "=SUM(A1:A2)"
    first["B4"] = "=Second!A1"
    second = fwb.create_sheet("Second")
    second["A1"] = 10
    second["B2"] = "=First!B1+A1"
    fwb.save(src)
    _inject_calc_chain(
        src,
        """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<calcChain xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <c r="B1" i="1"/>
  <c r="B4" i="1"/>
  <c r="X99" i="2"/>
  <extLst><ext uri="{wolfxl-test-calcchain-ext}"><x:test xmlns:x="urn:wolfxl:test">keep</x:test></ext></extLst>
</calcChain>""",
    )

    wb = load_workbook(src, modify=True)
    wb["First"].delete_rows(4)
    wb.save(dst)

    entries = _read_calc_chain(dst)
    assert ("B1", 1) in entries, f"same-sheet formula missing: {entries}"
    assert ("B2", 2) in entries, f"cross-sheet formula missing: {entries}"
    assert ("B4", 1) not in entries, f"deleted formula leaked into calcChain: {entries}"
    assert ("X99", 2) not in entries, f"stale source calcChain ref leaked: {entries}"
    with zipfile.ZipFile(dst, "r") as z:
        calc_xml = z.read("xl/calcChain.xml").decode("utf-8")
    assert "{wolfxl-test-calcchain-ext}" in calc_xml
    assert "urn:wolfxl:test" in calc_xml


# ---------------------------------------------------------------------------
# Smoke: re-opening with openpyxl works (calcChain doesn't break the file).
# ---------------------------------------------------------------------------


def test_calc_chain_roundtrip_via_openpyxl(tmp_path: Path) -> None:
    out = tmp_path / "out.xlsx"
    wb = Workbook()
    s = wb.active
    assert s is not None
    s.title = "S"
    s["A1"] = 1
    s["A2"] = 2
    s["A3"] = "=A1+A2"
    wb.save(out)

    rwb = openpyxl.load_workbook(out)
    assert rwb["S"]["A3"].value == "=A1+A2"
