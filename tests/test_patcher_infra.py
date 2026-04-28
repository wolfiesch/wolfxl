"""RFC-013 — patcher infrastructure extensions, end-to-end coverage.

The new primitives (``file_adds``, ``ContentTypesGraph``,
``AncillaryPartRegistry``, Phase-2.5c aggregation) all live behind the PyO3
boundary, so the only way to exercise them under realistic conditions is to
construct an :class:`~wolfxl.Workbook` via ``_from_patcher`` and drive the
test-only hooks (``_test_inject_file_add``, ``_test_queue_content_type_op``,
``_test_populate_ancillary``, …) on the underlying ``XlsxPatcher``.

These tests cover the contract of each primitive in isolation:

* `file_adds` — emits a brand-new ZIP entry (`emits_new_zip_entry`),
  hard-fails on collisions with source entries (`collision_*panics`),
  and stays a pure no-op when nothing is queued (`no_op_*byte_identical`).
* `[Content_Types].xml` aggregation — multiple sheets pushing ops collapse
  to one rewritten file (`aggregation_combines`), and the rewrite preserves
  source order for unchanged entries (`aggregation_preserves_source_order`).
* `AncillaryPartRegistry` — lazily populates from a sheet's
  ``_rels/sheetN.xml.rels`` and classifies entries by relationship type
  (`lazy_populate_on_demand`).

The contract these tests defend is the "ship dark" promise: RFC-013's
primitives MUST NOT change save output for existing modify-mode flows that
don't use them. The byte-identical no-op test is the hardest gate.
"""
from __future__ import annotations

import zipfile
from pathlib import Path

import openpyxl
import pytest

from wolfxl import Workbook

# ---------------------------------------------------------------------------
# Fixtures and helpers
# ---------------------------------------------------------------------------


@pytest.fixture(autouse=True)
def _force_test_epoch(monkeypatch: pytest.MonkeyPatch) -> None:
    """Determinism for byte-identical save assertions."""
    monkeypatch.setenv("WOLFXL_TEST_EPOCH", "0")


def _make_clean_fixture(path: Path, sheet_titles: tuple[str, ...] = ("Sheet1",)) -> None:
    """Single- or multi-sheet workbook with no extra parts beyond what
    openpyxl's ``Workbook()`` produces. Has a `[Content_Types].xml` with the
    standard set of overrides — a stable baseline for the aggregation tests.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_titles[0]
    ws["A1"] = "header"
    for title in sheet_titles[1:]:
        wb.create_sheet(title)
    wb.save(path)


def _make_fixture_with_table(path: Path) -> None:
    """Single-sheet workbook with one openpyxl-built table — guarantees the
    sheet's `_rels/sheet1.xml.rels` references a real `xl/tables/table1.xml`
    (so the ancillary registry has something to classify).
    """
    from openpyxl.worksheet.table import Table, TableStyleInfo

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Name", "Value"])
    ws.append(["a", 1])
    ws.append(["b", 2])
    table = Table(displayName="Tbl", ref="A1:B3")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
    )
    ws.add_table(table)
    wb.save(path)


def _read_zip_text(path: Path, entry: str) -> str:
    with zipfile.ZipFile(path) as zf:
        return zf.read(entry).decode("utf-8")


def _read_zip_bytes(path: Path, entry: str) -> bytes:
    with zipfile.ZipFile(path) as zf:
        return zf.read(entry)


def _zip_entries(path: Path) -> list[str]:
    with zipfile.ZipFile(path) as zf:
        return zf.namelist()


# ---------------------------------------------------------------------------
# file_adds
# ---------------------------------------------------------------------------


def test_file_adds_emits_new_zip_entry(tmp_path: Path) -> None:
    """A path injected into ``file_adds`` should appear as a new ZIP entry
    after save with the exact bytes the test supplied."""
    src = tmp_path / "src.xlsx"
    _make_clean_fixture(src)

    wb = Workbook._from_patcher(str(src))
    new_path = "xl/customWidgets/widget1.xml"
    new_bytes = b"<widget><name>foo</name></widget>"
    wb._rust_patcher._test_inject_file_add(new_path, new_bytes)

    out = tmp_path / "out.xlsx"
    wb.save(out)
    wb.close()

    entries = _zip_entries(out)
    assert new_path in entries, f"file_adds entry missing from saved ZIP; got {entries}"
    assert _read_zip_bytes(out, new_path) == new_bytes


def test_file_adds_collision_with_source_panics(tmp_path: Path) -> None:
    """Injecting an entry whose path already exists in the source ZIP is a
    caller bug. The patcher's collision check must surface as a Python
    exception (RFC-013 §8 risk #2)."""
    src = tmp_path / "src.xlsx"
    _make_clean_fixture(src)

    wb = Workbook._from_patcher(str(src))
    # `xl/workbook.xml` is guaranteed to be in any source xlsx.
    wb._rust_patcher._test_inject_file_add("xl/workbook.xml", b"<bogus/>")

    out = tmp_path / "out.xlsx"
    with pytest.raises(BaseException, match="(?i)collision|file_adds"):
        wb.save(out)
    wb.close()


def test_file_adds_no_op_save_byte_identical(tmp_path: Path) -> None:
    """Empty ``file_adds`` + empty ``file_deletes`` + no other queued state
    must produce a byte-identical save. This is the regression guard that
    keeps RFC-013's primitives from accidentally rewriting the source ZIP
    when no caller has asked for it."""
    src = tmp_path / "src.xlsx"
    _make_clean_fixture(src)
    src_bytes = src.read_bytes()

    wb = Workbook._from_patcher(str(src))
    out = tmp_path / "out.xlsx"
    wb.save(out)
    wb.close()

    assert out.read_bytes() == src_bytes, (
        "no-op modify-mode save must be byte-identical to source. "
        "If this fails, RFC-013's short-circuit predicate is missing a field."
    )


# ---------------------------------------------------------------------------
# [Content_Types].xml aggregation (Phase 2.5c)
# ---------------------------------------------------------------------------

CT_COMMENTS = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.comments+xml"
)
CT_TABLE = "application/vnd.openxmlformats-officedocument.spreadsheetml.table+xml"
CT_VML = "application/vnd.openxmlformats-officedocument.vmlDrawing"


def test_content_types_aggregation_combines_per_sheet_ops(tmp_path: Path) -> None:
    """Two sheets each pushing one Override op should collapse to one
    rewritten ``[Content_Types].xml`` containing both new entries."""
    src = tmp_path / "src.xlsx"
    _make_clean_fixture(src, sheet_titles=("Sheet1", "Sheet2"))

    wb = Workbook._from_patcher(str(src))
    wb._rust_patcher._test_queue_content_type_op(
        "Sheet1", "add_override", "/xl/comments1.xml", CT_COMMENTS
    )
    wb._rust_patcher._test_queue_content_type_op(
        "Sheet2", "add_override", "/xl/tables/table1.xml", CT_TABLE
    )
    out = tmp_path / "out.xlsx"
    wb.save(out)
    wb.close()

    ct_xml = _read_zip_text(out, "[Content_Types].xml")
    assert '/xl/comments1.xml' in ct_xml, ct_xml
    assert '/xl/tables/table1.xml' in ct_xml, ct_xml
    # Source overrides preserved (additive aggregation).
    assert '/xl/workbook.xml' in ct_xml
    assert '/xl/styles.xml' in ct_xml


def test_content_types_aggregation_preserves_source_order(tmp_path: Path) -> None:
    """The aggregation pass must not reorder source overrides — that would
    break byte-stable diffs against unmodified parts. After save, every
    source override must keep its relative position; new overrides land
    strictly after all source overrides."""
    import re

    src = tmp_path / "src.xlsx"
    _make_clean_fixture(src)

    # Capture every Override PartName in the source, in document order.
    src_ct = _read_zip_text(src, "[Content_Types].xml")
    src_overrides = re.findall(r'<Override PartName="([^"]+)"', src_ct)
    assert "/xl/workbook.xml" in src_overrides, "fixture lacks workbook override?"

    wb = Workbook._from_patcher(str(src))
    wb._rust_patcher._test_queue_content_type_op(
        "Sheet1", "add_override", "/xl/comments1.xml", CT_COMMENTS
    )
    out = tmp_path / "out.xlsx"
    wb.save(out)
    wb.close()

    out_ct = _read_zip_text(out, "[Content_Types].xml")
    out_overrides = re.findall(r'<Override PartName="([^"]+)"', out_ct)

    # Every source override must still be present, in the same relative order.
    assert out_overrides[: len(src_overrides)] == src_overrides, (
        f"source override order changed:\n  src: {src_overrides}\n  out: {out_overrides}"
    )
    # New override appended at the tail, not interleaved into source.
    assert out_overrides[-1] == "/xl/comments1.xml", (
        f"new override should be last, got: {out_overrides}"
    )


# ---------------------------------------------------------------------------
# AncillaryPartRegistry — lazy populate
# ---------------------------------------------------------------------------


def test_ancillary_registry_lazy_populate_on_demand(tmp_path: Path) -> None:
    """Pre-populate: registry is empty. Trigger populate for one sheet,
    then verify the cached SheetAncillary correctly classified the
    fixture's table relationship."""
    src = tmp_path / "src.xlsx"
    _make_fixture_with_table(src)

    wb = Workbook._from_patcher(str(src))
    patcher = wb._rust_patcher

    # Pre-populate gate: nothing cached yet.
    assert patcher._test_ancillary_is_populated("Sheet1") is False, (
        "registry must be empty before populate_for_sheet is called"
    )

    patcher._test_populate_ancillary("Sheet1")
    assert patcher._test_ancillary_is_populated("Sheet1") is True

    # Fixture has exactly one table at xl/tables/table1.xml.
    tables = patcher._test_ancillary_table_parts("Sheet1")
    assert tables == ["xl/tables/table1.xml"], (
        f"expected one table part, got {tables}"
    )
    # No comments on this fixture.
    assert patcher._test_ancillary_comments_part("Sheet1") is None
    assert patcher._test_ancillary_vml_drawing_part("Sheet1") is None
    # No hyperlinks either.
    assert patcher._test_ancillary_hyperlink_rids("Sheet1") == []

    # Idempotency: a second populate is a cheap no-op.
    patcher._test_populate_ancillary("Sheet1")
    assert patcher._test_ancillary_table_parts("Sheet1") == ["xl/tables/table1.xml"]

    wb.close()
