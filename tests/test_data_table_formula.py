"""RFC-057 (Sprint Ο Pod 1C) — DataTableFormula tests.

Covers the public Python API: constructor (with default-flag and 2D
permutations), equality, repr, ``cell.value = DataTableFormula(...)``
setter contract, and round-trip of the typed instance through both
write-mode and modify-mode flush paths.
"""

from __future__ import annotations

from pathlib import Path


import wolfxl
from wolfxl.cell.cell import DataTableFormula


# ----------------------------------------------------------------------
# 1) Class semantics
# ----------------------------------------------------------------------


def test_constructor_minimal() -> None:
    dt = DataTableFormula(ref="B2:F11")
    assert dt.ref == "B2:F11"
    assert dt.ca is False
    assert dt.dt2D is False
    assert dt.dtr is False
    assert dt.r1 is None
    assert dt.r2 is None


def test_constructor_2d() -> None:
    dt = DataTableFormula(
        ref="B2:F11",
        ca=False,
        dt2D=True,
        dtr=False,
        r1="A1",
        r2="A2",
    )
    assert dt.dt2D is True
    assert dt.r1 == "A1"
    assert dt.r2 == "A2"


def test_equality() -> None:
    a = DataTableFormula(ref="B2:F11", dt2D=True, r1="A1", r2="A2")
    b = DataTableFormula(ref="B2:F11", dt2D=True, r1="A1", r2="A2")
    assert a == b


def test_inequality_attribute_diff() -> None:
    a = DataTableFormula(ref="B2:F11", r1="A1")
    b = DataTableFormula(ref="B2:F11", r1="A2")
    assert a != b


def test_repr_includes_only_set_fields() -> None:
    dt = DataTableFormula(ref="B2:F11")
    # 1D, no inputs — ref only.
    assert "ref='B2:F11'" in repr(dt)
    assert "dt2D" not in repr(dt)


# ----------------------------------------------------------------------
# 2) Cell.value setter / getter contract
# ----------------------------------------------------------------------


def test_setter_populates_metadata() -> None:
    wb = wolfxl.Workbook()
    ws = wb.active
    dt = DataTableFormula(ref="B2:F11", dt2D=True, r1="A1", r2="A2")
    ws["B2"].value = dt
    cell = ws["B2"]
    assert cell._formula_type == "dataTable"
    assert cell._array_ref == "B2:F11"
    assert cell._dt_2d is True
    assert cell._dt_r1 == "A1"
    assert cell._dt_r2 == "A2"


def test_getter_returns_data_table_instance() -> None:
    wb = wolfxl.Workbook()
    ws = wb.active
    dt = DataTableFormula(ref="B2:F11", dt2D=True, r1="A1", r2="A2")
    ws["B2"] = dt
    val = ws["B2"].value
    assert isinstance(val, DataTableFormula)
    assert val.ref == "B2:F11"
    assert val.dt2D is True
    assert val.r1 == "A1"
    assert val.r2 == "A2"


# ----------------------------------------------------------------------
# 3) Round-trip
# ----------------------------------------------------------------------


def test_round_trip_write_mode(tmp_path: Path) -> None:
    p = tmp_path / "wm_dt.xlsx"
    wb = wolfxl.Workbook()
    wb.active["B2"] = DataTableFormula(
        ref="B2:F11", dt2D=True, r1="A1", r2="A2"
    )
    wb.save(str(p))

    reloaded = wolfxl.load_workbook(str(p))
    val = reloaded.active["B2"].value
    assert isinstance(val, DataTableFormula)
    assert val.ref == "B2:F11"
    assert val.dt2D is True
    assert val.r1 == "A1"
    assert val.r2 == "A2"


def test_round_trip_preserves_all_data_table_attrs(tmp_path: Path) -> None:
    import openpyxl as _opx
    from openpyxl.worksheet.formula import DataTableFormula as _OpxDataTableFormula

    p = tmp_path / "wm_dt_all_attrs.xlsx"
    wb = wolfxl.Workbook()
    wb.active["C1"] = DataTableFormula(
        ref="C1:C3",
        t="dataTable",
        ca=True,
        dt2D=False,
        dtr=True,
        r1="A1",
        r2="B1",
    )
    wb.save(str(p))

    reloaded = wolfxl.load_workbook(str(p))
    val = reloaded.active["C1"].value
    assert isinstance(val, DataTableFormula)
    assert val.ref == "C1:C3"
    assert val.ca is True
    assert val.dt2D is False
    assert val.dtr is True
    assert val.r1 == "A1"
    assert val.r2 == "B1"

    ref_wb = _opx.load_workbook(str(p))
    ref_val = ref_wb.active["C1"].value
    assert isinstance(ref_val, _OpxDataTableFormula)
    assert ref_val.ref == "C1:C3"
    assert str(ref_val.ca).lower() in {"1", "true"}
    assert str(ref_val.dt2D).lower() in {"0", "false"}
    assert str(ref_val.dtr).lower() in {"1", "true"}
    assert ref_val.r1 == "A1"
    assert ref_val.r2 == "B1"


def test_round_trip_modify_mode(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    wb = wolfxl.Workbook()
    wb.active["A1"] = "anchor"
    wb.save(str(src))

    wb = wolfxl.load_workbook(str(src), modify=True)
    wb.active["B2"] = DataTableFormula(
        ref="B2:F11", dt2D=True, r1="A1", r2="A2"
    )
    wb.save(str(dst))

    reloaded = wolfxl.load_workbook(str(dst))
    val = reloaded.active["B2"].value
    assert isinstance(val, DataTableFormula)
    assert val.ref == "B2:F11"
    assert val.dt2D is True


def test_xml_emits_data_table_formula(tmp_path: Path) -> None:
    """Sanity check: byte-level XML carries the right ``<f t="dataTable">``."""
    import zipfile

    p = tmp_path / "xml.xlsx"
    wb = wolfxl.Workbook()
    wb.active["B2"] = DataTableFormula(
        ref="B2:F11", dt2D=True, r1="A1", r2="A2"
    )
    wb.save(str(p))

    with zipfile.ZipFile(p) as z:
        sheet = z.read("xl/worksheets/sheet1.xml").decode()
    assert '<f t="dataTable"' in sheet
    assert 'ref="B2:F11"' in sheet
    assert 'dt2D="1"' in sheet
    assert 'r1="A1"' in sheet
    assert 'r2="A2"' in sheet
    # No formula body — the f element is self-closing.
    assert '<f t="dataTable" ref="B2:F11" dt2D="1" r1="A1" r2="A2"/>' in sheet
