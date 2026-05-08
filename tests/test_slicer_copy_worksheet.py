"""Sprint Ο Pod 3.5 (RFC-061 §6 / RFC-035) — slicer deep-clone on
``Workbook.copy_worksheet``.

Per RFC-061 §6:
  * Slicer **presentations** are sheet-scoped — when the source
    sheet is cloned, the destination sheet must get its own
    ``xl/slicers/slicer{N}.xml`` part (a fresh suffix) and its own
    sheet-rel of type ``SLICER`` pointing at the new part.
  * Real desktop Excel deep-clones workbook-scoped slicer caches when
    the copied slicer is tied to a pivot/table on the copied sheet. The
    cloned presentation must point at the fresh cache name.

These tests pin that contract so future RFC-035 edits can't
silently flip slicer semantics.
"""
from __future__ import annotations

import re
import zipfile
from pathlib import Path

import openpyxl
import pytest

from wolfxl import load_workbook
from wolfxl.chart.reference import Reference
from wolfxl.pivot import PivotCache, PivotTable, Slicer, SlicerCache


@pytest.fixture(autouse=True)
def _force_test_epoch(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("WOLFXL_TEST_EPOCH", "0")


def _make_slicer_workbook_fixture(path: Path) -> None:
    seed = path.parent / "_seed.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    rows = [
        ("region", "quarter", "revenue"),
        ("North", "Q1", 100.0),
        ("South", "Q1", 200.0),
        ("North", "Q2", 150.0),
        ("South", "Q2", 250.0),
    ]
    for r, row in enumerate(rows, start=1):
        for c, v in enumerate(row, start=1):
            ws.cell(r, c, v)
    wb.save(seed)

    wbm = load_workbook(seed, modify=True)
    wsm = wbm["Data"]
    ref = Reference(worksheet=wsm, min_col=1, min_row=1, max_col=3, max_row=5)
    cache = PivotCache(source=ref)
    pt = PivotTable(cache=cache, location="F2", rows=["region"], data=["revenue"])
    wbm.add_pivot_cache(cache)
    wsm.add_pivot_table(pt)
    sc = SlicerCache(name="Slicer_region", source_pivot_cache=cache, field="region")
    sl = Slicer(name="Slicer_region1", cache=sc, caption="Region")
    wbm.add_slicer_cache(sc)
    wsm.add_slicer(sl, anchor="H2")
    wbm.save(path)
    seed.unlink(missing_ok=True)


def _zip_listing(p: Path) -> list[str]:
    with zipfile.ZipFile(p, "r") as z:
        return sorted(z.namelist())


def _zip_read(p: Path, member: str) -> bytes:
    with zipfile.ZipFile(p, "r") as z:
        return z.read(member)


def _copy_first_sheet(src: Path, dst: Path, name: str = "Copied Sheet") -> None:
    wb = load_workbook(src, modify=True)
    wb.copy_worksheet(wb.worksheets[0], name=name)
    wb.save(dst)


def _copy_save_reopen_remove_copy(src: Path, dst: Path, name: str = "Copied Sheet") -> None:
    first_save = dst.with_name(f"{dst.stem}-with-copy{dst.suffix}")
    _copy_first_sheet(src, first_save, name=name)

    wb = load_workbook(first_save, modify=True)
    wb.remove(wb[name])
    wb.save(dst)


def _inject_shared_slicer_relationship(src: Path, dst: Path) -> None:
    rels_path = "xl/worksheets/_rels/sheet2.xml.rels"
    rel = (
        '<Relationship Id="rId99" '
        'Type="http://schemas.microsoft.com/office/2007/relationships/slicer" '
        'Target="../slicers/slicer1.xml"/>'
    )
    with zipfile.ZipFile(src, "r") as zin, zipfile.ZipFile(dst, "w", zipfile.ZIP_DEFLATED) as zout:
        for info in zin.infolist():
            data = zin.read(info.filename)
            if info.filename == rels_path:
                text = data.decode()
                data = text.replace("</Relationships>", f"{rel}</Relationships>").encode()
            zout.writestr(info, data)


# ---------------------------------------------------------------------------
# Case 1 — copy_worksheet allocates a FRESH slicer{N}.xml.
# ---------------------------------------------------------------------------


def test_copy_worksheet_with_slicer_allocates_new_slicer_part(
    tmp_path: Path,
) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_slicer_workbook_fixture(src)

    wb = load_workbook(src, modify=True)
    wb.copy_worksheet(wb["Data"], name="DataCopy")
    wb.save(dst)

    entries = _zip_listing(dst)
    slicer_parts = [
        e for e in entries
        if re.match(r"^xl/slicers/slicer\d+\.xml$", e)
    ]
    assert len(slicer_parts) == 2, (
        f"expected 2 slicer presentations (orig + clone), got {slicer_parts}"
    )


# ---------------------------------------------------------------------------
# Case 2 — Slicer cache is deep-cloned for desktop-Excel fidelity.
# ---------------------------------------------------------------------------


def test_copy_worksheet_with_slicer_clones_cache(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_slicer_workbook_fixture(src)

    wb = load_workbook(src, modify=True)
    wb.copy_worksheet(wb["Data"], name="DataCopy")
    wb.save(dst)

    entries = _zip_listing(dst)
    cache_parts = [
        e for e in entries
        if re.match(r"^xl/slicerCaches/slicerCache\d+\.xml$", e)
    ]
    assert len(cache_parts) == 2, f"expected cloned slicer cache, got {cache_parts}"


# ---------------------------------------------------------------------------
# Case 3 — Cloned sheet's rels file points at the NEW slicer part,
# not the source's.
# ---------------------------------------------------------------------------


def test_copy_worksheet_with_slicer_emits_distinct_sheet_rel(
    tmp_path: Path,
) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_slicer_workbook_fixture(src)

    wb = load_workbook(src, modify=True)
    wb.copy_worksheet(wb["Data"], name="DataCopy")
    wb.save(dst)

    entries = _zip_listing(dst)
    sheet_rels_files = [
        e for e in entries
        if re.match(r"^xl/worksheets/_rels/sheet\d+\.xml\.rels$", e)
    ]
    assert len(sheet_rels_files) == 2, (
        f"expected 2 sheet rels files, got {sheet_rels_files}"
    )

    # Each rels file must carry its OWN slicer rel target. They must
    # NOT both point at slicer1.xml.
    targets: list[str] = []
    for rels_path in sheet_rels_files:
        rels_xml = _zip_read(dst, rels_path).decode()
        if "office/2007/relationships/slicer" in rels_xml and "slicerCache" not in rels_xml:
            # Extract slicer{N}.xml target.
            m = re.search(r'Target="\.\./slicers/(slicer\d+\.xml)"', rels_xml)
            if m:
                targets.append(m.group(1))
    assert len(targets) == 2 and len(set(targets)) == 2, (
        f"cloned slicer rels must point at distinct slicer parts: {targets}"
    )


# ---------------------------------------------------------------------------
# Case 4 — Cloned slicer presentation references the cloned cache name.
# ---------------------------------------------------------------------------


def test_cloned_slicer_xml_uses_cloned_cache_name(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_slicer_workbook_fixture(src)

    wb = load_workbook(src, modify=True)
    wb.copy_worksheet(wb["Data"], name="DataCopy")
    wb.save(dst)

    entries = _zip_listing(dst)
    slicer_parts = [
        e for e in entries
        if re.match(r"^xl/slicers/slicer\d+\.xml$", e)
    ]
    cache_refs = []
    for sp in slicer_parts:
        xml = _zip_read(dst, sp).decode()
        m = re.search(r'cache="(\w+)"', xml)
        if m:
            cache_refs.append(m.group(1))
    assert cache_refs == ["Slicer_region", "Slicer_region1"]


# ---------------------------------------------------------------------------
# Case 5 — Content-Types Override added for the cloned slicer part.
# ---------------------------------------------------------------------------


def test_cloned_slicer_emits_content_type_override(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_slicer_workbook_fixture(src)

    wb = load_workbook(src, modify=True)
    wb.copy_worksheet(wb["Data"], name="DataCopy")
    wb.save(dst)

    ct = _zip_read(dst, "[Content_Types].xml").decode()
    # Must have an Override per slicer presentation file.
    overrides = re.findall(
        r'<Override PartName="/xl/slicers/slicer\d+\.xml"', ct
    )
    assert len(overrides) == 2, (
        f"expected 2 slicer Overrides post-copy, got {len(overrides)}: {ct}"
    )


def test_real_excel_pivot_slicer_copy_clones_cache_parts(tmp_path: Path) -> None:
    src = Path("tests/fixtures/external_oracle/real-excel-pivot-chart-slicers.xlsx")
    dst = tmp_path / "pivot-slicer-copy.xlsx"

    _copy_first_sheet(src, dst)

    entries = _zip_listing(dst)
    cache_parts = [
        e for e in entries
        if re.match(r"^xl/slicerCaches/slicerCache\d+\.xml$", e)
    ]
    assert len(cache_parts) == 4

    copied_slicer_xml = _zip_read(dst, "xl/slicers/slicer2.xml").decode()
    assert 'name="REGION 1"' in copied_slicer_xml
    assert 'cache="Slicer_REGION1"' in copied_slicer_xml
    assert 'cache="Slicer_YEAR1"' in copied_slicer_xml

    workbook_xml = _zip_read(dst, "xl/workbook.xml").decode()
    assert workbook_xml.count("<x14:slicerCache ") == 4
    assert '<definedName name="Slicer_REGION1">#N/A</definedName>' in workbook_xml

    copied_sheet_id = re.search(r'name="Copied Sheet" sheetId="(\d+)"', workbook_xml)
    assert copied_sheet_id is not None
    cache_xml = _zip_read(dst, "xl/slicerCaches/slicerCache3.xml").decode()
    assert f'tabId="{copied_sheet_id.group(1)}"' in cache_xml


def test_real_excel_pivot_slicer_remove_copy_prunes_orphaned_cache_parts(
    tmp_path: Path,
) -> None:
    src = Path("tests/fixtures/external_oracle/real-excel-pivot-chart-slicers.xlsx")
    dst = tmp_path / "pivot-slicer-copy-removed.xlsx"

    _copy_save_reopen_remove_copy(src, dst)

    entries = _zip_listing(dst)
    assert [
        e for e in entries
        if re.match(r"^xl/slicerCaches/slicerCache\d+\.xml$", e)
    ] == ["xl/slicerCaches/slicerCache1.xml", "xl/slicerCaches/slicerCache2.xml"]
    assert [
        e for e in entries
        if re.match(r"^xl/slicers/slicer\d+\.xml$", e)
    ] == ["xl/slicers/slicer1.xml"]

    workbook_xml = _zip_read(dst, "xl/workbook.xml").decode()
    workbook_rels = _zip_read(dst, "xl/_rels/workbook.xml.rels").decode()
    content_types = _zip_read(dst, "[Content_Types].xml").decode()
    assert workbook_xml.count("<x14:slicerCache ") == 2
    assert "Slicer_REGION1" not in workbook_xml
    assert "Slicer_YEAR1" not in workbook_xml
    assert "slicerCache3.xml" not in workbook_rels
    assert "slicerCache4.xml" not in workbook_rels
    assert "/xl/slicerCaches/slicerCache3.xml" not in content_types
    assert "/xl/slicerCaches/slicerCache4.xml" not in content_types


def test_remove_sheet_keeps_slicer_cache_used_by_retained_sheet(
    tmp_path: Path,
) -> None:
    src = Path("tests/fixtures/external_oracle/real-excel-pivot-chart-slicers.xlsx")
    shared_src = tmp_path / "pivot-slicer-shared.xlsx"
    dst = tmp_path / "pivot-slicer-shared-sheet-removed.xlsx"
    _inject_shared_slicer_relationship(src, shared_src)

    wb = load_workbook(shared_src, modify=True)
    wb.remove(wb["Pivot Table"])
    wb.save(dst)

    entries = _zip_listing(dst)
    assert "xl/slicers/slicer1.xml" in entries
    assert "xl/slicerCaches/slicerCache1.xml" in entries
    assert "xl/slicerCaches/slicerCache2.xml" in entries

    kept_sheet_rels = _zip_read(dst, "xl/worksheets/_rels/sheet2.xml.rels").decode()
    assert "Target=\"../slicers/slicer1.xml\"" in kept_sheet_rels

    workbook_xml = _zip_read(dst, "xl/workbook.xml").decode()
    workbook_rels = _zip_read(dst, "xl/_rels/workbook.xml.rels").decode()
    assert workbook_xml.count("<x14:slicerCache ") == 2
    assert 'name="Slicer_REGION">#N/A</definedName>' in workbook_xml
    assert 'name="Slicer_YEAR">#N/A</definedName>' in workbook_xml
    assert "slicerCaches/slicerCache1.xml" in workbook_rels
    assert "slicerCaches/slicerCache2.xml" in workbook_rels


def test_real_excel_timeline_copy_clones_timeline_cache_parts(tmp_path: Path) -> None:
    src = Path("tests/fixtures/external_oracle/real-excel-timeline-slicer.xlsx")
    dst = tmp_path / "timeline-copy.xlsx"

    _copy_first_sheet(src, dst)

    entries = _zip_listing(dst)
    timeline_parts = [
        e for e in entries
        if re.match(r"^xl/timelines/timeline\d+\.xml$", e)
    ]
    timeline_cache_parts = [
        e for e in entries
        if re.match(r"^xl/timelineCaches/timelineCache\d+\.xml$", e)
    ]
    assert timeline_parts == ["xl/timelines/timeline1.xml", "xl/timelines/timeline2.xml"]
    assert timeline_cache_parts == [
        "xl/timelineCaches/timelineCache1.xml",
        "xl/timelineCaches/timelineCache2.xml",
    ]

    workbook_xml = _zip_read(dst, "xl/workbook.xml").decode()
    copied_sheet_id = re.search(r'name="Copied Sheet" sheetId="(\d+)"', workbook_xml)
    assert copied_sheet_id is not None
    assert int(copied_sheet_id.group(1)) > 4
    assert workbook_xml.count("<x15:timelineCacheRef ") == 2
    assert '<definedName name="NativeTimeline_ORDER_DATE1">#N/A</definedName>' in workbook_xml

    copied_timeline_xml = _zip_read(dst, "xl/timelines/timeline2.xml").decode()
    assert 'name="ORDER DATE 1"' in copied_timeline_xml
    assert 'cache="NativeTimeline_ORDER_DATE1"' in copied_timeline_xml
    copied_cache_xml = _zip_read(dst, "xl/timelineCaches/timelineCache2.xml").decode()
    assert 'name="NativeTimeline_ORDER_DATE1"' in copied_cache_xml
    assert f'tabId="{copied_sheet_id.group(1)}"' in copied_cache_xml


def test_real_excel_timeline_remove_copy_prunes_orphaned_cache_parts(
    tmp_path: Path,
) -> None:
    src = Path("tests/fixtures/external_oracle/real-excel-timeline-slicer.xlsx")
    dst = tmp_path / "timeline-copy-removed.xlsx"

    _copy_save_reopen_remove_copy(src, dst)

    entries = _zip_listing(dst)
    assert [
        e for e in entries
        if re.match(r"^xl/timelineCaches/timelineCache\d+\.xml$", e)
    ] == ["xl/timelineCaches/timelineCache1.xml"]
    assert [
        e for e in entries
        if re.match(r"^xl/timelines/timeline\d+\.xml$", e)
    ] == ["xl/timelines/timeline1.xml"]

    workbook_xml = _zip_read(dst, "xl/workbook.xml").decode()
    workbook_rels = _zip_read(dst, "xl/_rels/workbook.xml.rels").decode()
    content_types = _zip_read(dst, "[Content_Types].xml").decode()
    assert workbook_xml.count("<x15:timelineCacheRef ") == 1
    assert "NativeTimeline_ORDER_DATE1" not in workbook_xml
    assert "timelineCache2.xml" not in workbook_rels
    assert "/xl/timelineCaches/timelineCache2.xml" not in content_types
