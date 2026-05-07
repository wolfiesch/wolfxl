"""G09 - Rich text in worksheet headers/footers.

Covers the openpyxl-shaped ``ws.oddHeader.center.text`` / ``.font`` /
``.size`` / ``.color`` API, plus ``evenHeader`` / ``firstFooter`` and
the four other slots. Checks both write-mode and modify-mode round
trips, and verifies that the inline OOXML format-code mini-grammar is
preserved across an open->edit->save cycle.
"""
from __future__ import annotations

import zipfile
from pathlib import Path

import pytest

import wolfxl
from wolfxl import load_workbook
from wolfxl.worksheet.header_footer import (
    HeaderFooterItem,
    _HeaderFooterPart,
)


# ---------------------------------------------------------------------------
# Mini-format parser/composer (Python-side, no IO).
# ---------------------------------------------------------------------------


class TestPartMiniFormat:
    def test_part_default_is_empty(self) -> None:
        part = _HeaderFooterPart()
        assert part.is_empty()
        assert not part
        assert part.to_format_string() is None

    def test_part_text_only(self) -> None:
        part = _HeaderFooterPart(text="Hello")
        assert part.to_format_string() == "Hello"
        assert part == "Hello"

    def test_part_text_font_size_color(self) -> None:
        part = _HeaderFooterPart(
            text="Title", font="Arial,Bold", size=14, color="FF0000"
        )
        s = part.to_format_string()
        assert s is not None
        assert '&"Arial,Bold"' in s
        assert "&14" in s
        assert "&KFF0000" in s
        assert s.endswith("Title")

    def test_part_color_must_be_rgb_hex(self) -> None:
        with pytest.raises(ValueError):
            _HeaderFooterPart(color="not-a-color")

    def test_part_color_uppercased(self) -> None:
        part = _HeaderFooterPart(color="ff00aa")
        assert part.color == "FF00AA"

    def test_part_from_format_string_round_trip(self) -> None:
        src = '&"Arial,Bold"&14 &KFF0000Title'
        part = _HeaderFooterPart.from_format_string(src)
        assert part.font == "Arial,Bold"
        assert part.size == 14
        assert part.color == "FF0000"
        assert part.text == "Title"

    def test_part_from_format_string_text_only(self) -> None:
        part = _HeaderFooterPart.from_format_string("Plain header")
        assert part.text == "Plain header"
        assert part.font is None
        assert part.size is None
        assert part.color is None


# ---------------------------------------------------------------------------
# HeaderFooterItem.from_str / .text - L/C/R round-trip.
# ---------------------------------------------------------------------------


class TestItemRoundTrip:
    def test_from_str_three_segments(self) -> None:
        item = HeaderFooterItem.from_str(
            '&L&"Arial,Bold"&14 Hello&CCenter&R&10 Right'
        )
        assert item.left.text == "Hello"
        assert item.left.font == "Arial,Bold"
        assert item.left.size == 14
        assert item.center.text == "Center"
        assert item.right.text == "Right"
        assert item.right.size == 10

    def test_from_str_empty_text_yields_empty_item(self) -> None:
        assert HeaderFooterItem.from_str("").is_empty()
        assert HeaderFooterItem.from_str(None).is_empty()

    def test_compose_back_to_text(self) -> None:
        item = HeaderFooterItem()
        item.left.text = "L"
        item.left.font = "Arial,Bold"
        item.left.size = 12
        item.center.text = "C"
        item.right.text = "R"
        item.right.color = "AA00FF"
        s = item.text
        assert s.startswith("&L")
        assert "&CC" in s
        assert "&R" in s
        assert "&KAA00FF" in s

    def test_setting_str_parses_inline_format(self) -> None:
        item = HeaderFooterItem()
        item.center = '&"Calibri,Bold"&18Confidential'
        assert item.center.font == "Calibri,Bold"
        assert item.center.size == 18
        assert item.center.text == "Confidential"

    def test_centre_alias(self) -> None:
        item = HeaderFooterItem()
        item.centre.text = "X"
        assert item.center.text == "X"


# ---------------------------------------------------------------------------
# Write-mode: native-writer round-trip.
# ---------------------------------------------------------------------------


def _emit_and_reload(tmp_path: Path, configure):
    """Write a workbook with a configured header/footer; load it back."""
    src = tmp_path / "hf.xlsx"
    wb = wolfxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws["A1"] = "data"
    configure(ws)
    wb.save(str(src))
    wb2 = load_workbook(str(src))
    return wb2.active


def test_write_odd_header_center_rich(tmp_path: Path) -> None:
    def configure(ws):
        ws.oddHeader.center.text = "Title"
        ws.oddHeader.center.font = "Arial,Bold"
        ws.oddHeader.center.size = 14
        ws.oddHeader.center.color = "FF0000"

    ws2 = _emit_and_reload(tmp_path, configure)
    assert ws2.oddHeader.center.text == "Title"
    assert ws2.oddHeader.center.font == "Arial,Bold"
    assert ws2.oddHeader.center.size == 14
    assert ws2.oddHeader.center.color == "FF0000"


def test_write_all_six_slots(tmp_path: Path) -> None:
    def configure(ws):
        ws.oddHeader.left.text = "OL"
        ws.oddFooter.right.text = "OR"
        ws.evenHeader.center.text = "EC"
        ws.evenHeader.center.size = 16
        ws.evenFooter.left.text = "EFL"
        ws.firstHeader.center.text = "FHC"
        ws.firstFooter.right.text = "FFR"
        ws.firstFooter.right.color = "00AAFF"
        ws.header_footer.different_odd_even = True
        ws.header_footer.different_first = True

    ws2 = _emit_and_reload(tmp_path, configure)
    assert ws2.oddHeader.left.text == "OL"
    assert ws2.oddFooter.right.text == "OR"
    assert ws2.evenHeader.center.text == "EC"
    assert ws2.evenHeader.center.size == 16
    assert ws2.evenFooter.left.text == "EFL"
    assert ws2.firstHeader.center.text == "FHC"
    assert ws2.firstFooter.right.text == "FFR"
    assert ws2.firstFooter.right.color == "00AAFF"
    assert ws2.header_footer.different_odd_even is True
    assert ws2.header_footer.different_first is True


def test_write_multiline_text_in_section(tmp_path: Path) -> None:
    def configure(ws):
        ws.oddHeader.center.text = "Line1\nLine2"

    ws2 = _emit_and_reload(tmp_path, configure)
    # Newline survives through OOXML text node.
    assert ws2.oddHeader.center.text == "Line1\nLine2"


def test_write_empty_sections_omitted_from_xml(tmp_path: Path) -> None:
    """Empty headers/footers should not produce headerFooter element."""
    src = tmp_path / "empty.xlsx"
    wb = wolfxl.Workbook()
    ws = wb.active
    ws["A1"] = "data"
    wb.save(str(src))
    with zipfile.ZipFile(src) as zf:
        sheet_xml = zf.read("xl/worksheets/sheet1.xml").decode("utf-8")
    # No headerFooter element when nothing is set.
    assert "<headerFooter" not in sheet_xml


def test_write_format_string_appears_in_oddheader_xml(tmp_path: Path) -> None:
    src = tmp_path / "rt.xlsx"
    wb = wolfxl.Workbook()
    ws = wb.active
    ws["A1"] = "data"
    ws.oddHeader.center.text = "Title"
    ws.oddHeader.center.font = "Arial,Bold"
    ws.oddHeader.center.size = 14
    ws.oddHeader.center.color = "FF0000"
    wb.save(str(src))
    with zipfile.ZipFile(src) as zf:
        sheet_xml = zf.read("xl/worksheets/sheet1.xml").decode("utf-8")
    assert "<oddHeader>" in sheet_xml
    # All format codes survive into the XML text node. Ampersands escape;
    # quote characters inside element text need no entity escape.
    assert "&amp;C" in sheet_xml
    assert '&amp;"Arial,Bold"' in sheet_xml
    assert "&amp;14" in sheet_xml
    assert "&amp;KFF0000" in sheet_xml
    assert "Title" in sheet_xml


# ---------------------------------------------------------------------------
# Modify-mode: open -> edit -> save.
# ---------------------------------------------------------------------------


def test_modify_mode_no_op_preserves_header(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    wb = wolfxl.Workbook()
    ws = wb.active
    ws["A1"] = "data"
    ws.oddHeader.center.text = "Annual Report"
    ws.oddHeader.center.font = "Calibri,Bold"
    ws.oddHeader.center.size = 16
    ws.oddHeader.center.color = "FF0000"
    wb.save(str(src))

    wb2 = load_workbook(str(src), modify=True)
    wb2.save(str(dst))

    wb3 = load_workbook(str(dst))
    ws3 = wb3.active
    assert ws3.oddHeader.center.text == "Annual Report"
    assert ws3.oddHeader.center.font == "Calibri,Bold"
    assert ws3.oddHeader.center.size == 16
    assert ws3.oddHeader.center.color == "FF0000"


def test_modify_mode_replaces_header(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    wb = wolfxl.Workbook()
    ws = wb.active
    ws["A1"] = "data"
    ws.oddHeader.center.text = "Old Title"
    wb.save(str(src))

    wb2 = load_workbook(str(src), modify=True)
    ws2 = wb2.active
    ws2.oddHeader.center.text = "New Title"
    ws2.oddHeader.center.font = "Times New Roman,Bold"
    ws2.oddHeader.center.size = 20
    ws2.oddHeader.right.text = "Page &P"
    wb2.save(str(dst))

    wb3 = load_workbook(str(dst))
    ws3 = wb3.active
    assert ws3.oddHeader.center.text == "New Title"
    assert ws3.oddHeader.center.font == "Times New Roman,Bold"
    assert ws3.oddHeader.center.size == 20
    assert ws3.oddHeader.right.text == "Page &P"


def test_modify_mode_adds_footer_to_existing_header(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    wb = wolfxl.Workbook()
    ws = wb.active
    ws["A1"] = "data"
    ws.oddHeader.center.text = "Doc"
    wb.save(str(src))

    wb2 = load_workbook(str(src), modify=True)
    ws2 = wb2.active
    ws2.oddFooter.center.text = "Page &P of &N"
    ws2.oddFooter.center.size = 10
    wb2.save(str(dst))

    wb3 = load_workbook(str(dst))
    ws3 = wb3.active
    assert ws3.oddHeader.center.text == "Doc"
    assert ws3.oddFooter.center.text == "Page &P of &N"
    assert ws3.oddFooter.center.size == 10


def test_modify_mode_first_and_even_slots(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    wb = wolfxl.Workbook()
    ws = wb.active
    ws["A1"] = "data"
    wb.save(str(src))

    wb2 = load_workbook(str(src), modify=True)
    ws2 = wb2.active
    ws2.firstHeader.center.text = "First only"
    ws2.firstHeader.center.size = 18
    ws2.evenFooter.left.text = "Even"
    ws2.evenFooter.left.color = "0000FF"
    ws2.header_footer.different_odd_even = True
    ws2.header_footer.different_first = True
    wb2.save(str(dst))

    wb3 = load_workbook(str(dst))
    ws3 = wb3.active
    assert ws3.firstHeader.center.text == "First only"
    assert ws3.firstHeader.center.size == 18
    assert ws3.evenFooter.left.text == "Even"
    assert ws3.evenFooter.left.color == "0000FF"
    assert ws3.header_footer.different_odd_even is True
    assert ws3.header_footer.different_first is True


# ---------------------------------------------------------------------------
# Cross-compat with openpyxl's native HeaderFooter.
# ---------------------------------------------------------------------------


def test_openpyxl_authored_header_reads_back(tmp_path: Path) -> None:
    """An openpyxl-authored workbook with rich-text header is parsed
    correctly by wolfxl with full attribute fidelity."""
    pytest.importorskip("openpyxl")
    import openpyxl

    src = tmp_path / "openpyxl_in.xlsx"
    wb_op = openpyxl.Workbook()
    ws_op = wb_op.active
    ws_op["A1"] = "data"
    ws_op.oddHeader.center.text = "Title"
    ws_op.oddHeader.center.font = "Arial,Bold"
    ws_op.oddHeader.center.size = 14
    ws_op.oddHeader.center.color = "FF0000"
    wb_op.save(str(src))

    wb = load_workbook(str(src))
    ws = wb.active
    assert ws.oddHeader.center.text == "Title"
    assert ws.oddHeader.center.font == "Arial,Bold"
    assert ws.oddHeader.center.size == 14
    assert ws.oddHeader.center.color == "FF0000"
