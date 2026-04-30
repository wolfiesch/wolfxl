"""Smoke tests for the native XLSX reader.

The native reader is the eager XLSX read path. Most tests still set the old
``WOLFXL_NATIVE_READER`` flag so older invocation patterns stay covered while
the implementation no longer depends on that opt-in.
"""

from __future__ import annotations

import datetime as dt
import re
import zipfile
from pathlib import Path

import pytest

openpyxl = pytest.importorskip("openpyxl")
openpyxl_datavalidation = pytest.importorskip("openpyxl.worksheet.datavalidation")
openpyxl_hyperlink = pytest.importorskip("openpyxl.worksheet.hyperlink")
openpyxl_image = pytest.importorskip("openpyxl.drawing.image")
wolfxl = pytest.importorskip("wolfxl")

DataValidation = openpyxl_datavalidation.DataValidation
Hyperlink = openpyxl_hyperlink.Hyperlink
OpenpyxlImage = openpyxl_image.Image
Border = openpyxl.styles.Border
Font = openpyxl.styles.Font
PatternFill = openpyxl.styles.PatternFill
Side = openpyxl.styles.Side
Alignment = openpyxl.styles.Alignment

FIXTURES = Path(__file__).parent / "fixtures"
PNG_PATH = FIXTURES / "images" / "tiny_red_dot.png"


def _make_basic_xlsx(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Label"
    ws["B1"] = 42
    ws["B1"].number_format = "#,##0.00"
    ws["B1"].font = Font(
        name="Arial",
        size=14,
        bold=True,
        italic=True,
        underline="single",
        strike=True,
        color="FF123456",
    )
    ws["B1"].fill = PatternFill(patternType="solid", fgColor="FFABCDEF")
    ws["B1"].alignment = Alignment(
        horizontal="center",
        vertical="top",
        wrap_text=True,
        text_rotation=45,
        indent=2,
    )
    ws["B1"].border = Border(
        left=Side(style="thin", color="FFFF0000"),
        right=Side(style="medium"),
        top=Side(style="double", color="FF00FF00"),
        bottom=Side(style="dashed", color="FF0000FF"),
    )
    ws["C1"] = True
    ws["A2"] = "Formula"
    ws["B2"] = "=B1*2"
    ws["A3"] = dt.datetime(2024, 1, 15, 12, 30)
    ws["B3"] = dt.date(2024, 6, 1)
    ws["D1"] = "Anchor"
    ws["D1"].number_format = "#,##0"
    ws.merge_cells("D1:E1")
    ws["A5"] = "External"
    ws["A5"].hyperlink = Hyperlink(
        ref="A5",
        target="https://example.com/report",
        tooltip="Example report",
    )
    ws["B5"] = "Internal"
    ws["B5"].hyperlink = Hyperlink(ref="B5", location="Data!A1", display="Jump")
    ws["A6"] = "Commented"
    ws["A6"].comment = openpyxl.comments.Comment("Native reader note", "Wolf")
    ws.row_dimensions[6].height = 24
    ws.column_dimensions["C"].width = 18
    dv = DataValidation(type="list", formula1='"Red,Blue"', allow_blank=True)
    dv.add("C2:C6")
    ws.add_data_validation(dv)
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = "A1:D6"
    ws.protection.sheet = True
    ws.protection.objects = True
    ws.protection.formatCells = False
    ws.protection.sort = False
    ws.protection.set_password("hunter2")
    wb.save(path)
    wb.close()
    _inject_workbook_security(path)
    _inject_auto_filter_details(path)
    _inject_merged_subordinate_style(path)


def _inject_workbook_security(path: Path) -> None:
    with zipfile.ZipFile(path, "r") as zin:
        entries = {name: zin.read(name) for name in zin.namelist()}
    workbook_name = "xl/workbook.xml"
    workbook_xml = entries[workbook_name].decode()
    workbook_xml = re.sub(r"<fileSharing\b[^>]*/>", "", workbook_xml)
    workbook_xml = re.sub(r"<workbookProtection\b[^>]*/>", "", workbook_xml)
    security_xml = (
        '<fileSharing readOnlyRecommended="1" userName="Wolf" algorithmName="SHA-512" '
        'hashValue="FILEHASH" saltValue="FILESALT" spinCount="100000"/>'
        '<workbookProtection lockStructure="1" workbookAlgorithmName="SHA-512" '
        'workbookHashValue="HASH" workbookSaltValue="SALT" workbookSpinCount="100000"/>'
    )
    workbook_xml, count = re.subn(r"(<workbookPr\b)", security_xml + r"\1", workbook_xml)
    assert count == 1
    entries[workbook_name] = workbook_xml.encode()
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in entries.items():
            zout.writestr(name, data)


def _inject_workbook_calc_properties(path: Path) -> None:
    with zipfile.ZipFile(path, "r") as zin:
        entries = {name: zin.read(name) for name in zin.namelist()}
    workbook_name = "xl/workbook.xml"
    workbook_xml = entries[workbook_name].decode()
    workbook_xml = re.sub(r"<calcPr\b[^>]*/>", "", workbook_xml)
    workbook_xml = re.sub(
        r"<workbookPr\b[^>]*/>",
        (
            '<workbookPr date1904="1" dateCompatibility="0" showObjects="none" '
            'filterPrivacy="1" backupFile="1" updateLinks="never" '
            'codeName="NativeWorkbook" refreshAllConnections="1" '
            'defaultThemeVersion="164011"/>'
        ),
        workbook_xml,
        count=1,
    )
    calc_xml = (
        '<calcPr calcId="191029" calcMode="manual" fullCalcOnLoad="1" '
        'refMode="R1C1" iterate="1" iterateCount="25" iterateDelta="0.01" '
        'fullPrecision="0" calcCompleted="0" calcOnSave="0" '
        'concurrentCalc="0" concurrentManualCount="4" forceFullCalc="1"/>'
    )
    workbook_xml, count = re.subn(
        r"(</workbook>)",
        calc_xml + r"\1",
        workbook_xml,
        count=1,
    )
    assert count == 1
    entries[workbook_name] = workbook_xml.encode()
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in entries.items():
            zout.writestr(name, data)


def _inject_workbook_views(path: Path) -> None:
    with zipfile.ZipFile(path, "r") as zin:
        entries = {name: zin.read(name) for name in zin.namelist()}
    workbook_name = "xl/workbook.xml"
    workbook_xml = entries[workbook_name].decode()
    book_views_xml = (
        "<bookViews>"
        '<workbookView visibility="hidden" minimized="1" showHorizontalScroll="0" '
        'showVerticalScroll="0" showSheetTabs="0" xWindow="10" yWindow="20" '
        'windowWidth="12000" windowHeight="8000" tabRatio="750" '
        'firstSheet="1" activeTab="2" autoFilterDateGrouping="0"/>'
        "</bookViews>"
    )
    workbook_xml, count = re.subn(
        r"<bookViews\b.*?</bookViews>",
        book_views_xml,
        workbook_xml,
        count=1,
    )
    assert count == 1
    entries[workbook_name] = workbook_xml.encode()
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in entries.items():
            zout.writestr(name, data)


def _inject_merged_subordinate_style(path: Path) -> None:
    """Add a styled blank subordinate cell to the merged D1:E1 range."""
    with zipfile.ZipFile(path, "r") as zin:
        entries = {name: zin.read(name) for name in zin.namelist()}
    sheet_name = "xl/worksheets/sheet1.xml"
    sheet_xml = entries[sheet_name].decode()
    if '<c r="E1"' in sheet_xml:
        return
    sheet_xml = sheet_xml.replace("</row>", '<c r="E1" s="1"/></row>', 1)
    entries[sheet_name] = sheet_xml.encode()
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in entries.items():
            zout.writestr(name, data)


def _inject_auto_filter_details(path: Path) -> None:
    with zipfile.ZipFile(path, "r") as zin:
        entries = {name: zin.read(name) for name in zin.namelist()}
    sheet_name = "xl/worksheets/sheet1.xml"
    sheet_xml = entries[sheet_name].decode()
    auto_filter_xml = (
        '<autoFilter ref="A1:D6">'
        '<filterColumn colId="0"><filters><filter val="Label"/></filters></filterColumn>'
        '<filterColumn colId="1"><customFilters and="1">'
        '<customFilter operator="greaterThan" val="10"/>'
        '</customFilters></filterColumn>'
        '<sortState ref="A2:D6"><sortCondition ref="B2:B6" descending="1"/></sortState>'
        "</autoFilter>"
    )
    sheet_xml, count = re.subn(r'<autoFilter ref="A1:D6"\s*/>', auto_filter_xml, sheet_xml)
    assert count == 1
    entries[sheet_name] = sheet_xml.encode()
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in entries.items():
            zout.writestr(name, data)


def _inject_sheet_properties(path: Path) -> None:
    with zipfile.ZipFile(path, "r") as zin:
        entries = {name: zin.read(name) for name in zin.namelist()}
    sheet_name = "xl/worksheets/sheet1.xml"
    sheet_xml = entries[sheet_name].decode()
    sheet_xml = re.sub(r"<sheetPr\b.*?</sheetPr>", "", sheet_xml)
    sheet_xml = re.sub(r"<sheetPr\b[^>]*/>", "", sheet_xml)
    sheet_pr_xml = (
        '<sheetPr codeName="NativeProps" enableFormatConditionsCalculation="0" '
        'filterMode="1" published="0" syncHorizontal="1" syncRef="B2" '
        'syncVertical="1" transitionEvaluation="1" transitionEntry="1">'
        '<tabColor rgb="FF33AA55"/>'
        '<outlinePr summaryBelow="0" summaryRight="0" applyStyles="1" '
        'showOutlineSymbols="0"/>'
        '<pageSetUpPr autoPageBreaks="0" fitToPage="1"/>'
        "</sheetPr>"
    )
    sheet_xml, count = re.subn(
        r"(<worksheet\b[^>]*>)",
        r"\1" + sheet_pr_xml,
        sheet_xml,
        count=1,
    )
    assert count == 1
    entries[sheet_name] = sheet_xml.encode()
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in entries.items():
            zout.writestr(name, data)


def _make_shared_formula_xlsx(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for row in range(1, 4):
        ws[f"A{row}"] = row
        ws[f"B{row}"] = f"=A{row}*2"
    wb.save(path)
    wb.close()

    with zipfile.ZipFile(path, "r") as zin:
        entries = {name: zin.read(name) for name in zin.namelist()}
    sheet_name = "xl/worksheets/sheet1.xml"
    sheet_xml = entries[sheet_name].decode()
    replacements = {
        r'<c r="B1"[^>]*>\s*<f>A1\*2</f>\s*(?:<v(?:>[^<]*</v>|\s*/>)\s*)?</c>': (
            '<c r="B1"><f t="shared" si="0" ref="B1:B3">A1*2</f><v/></c>'
        ),
        r'<c r="B2"[^>]*>\s*<f>A2\*2</f>\s*(?:<v(?:>[^<]*</v>|\s*/>)\s*)?</c>': (
            '<c r="B2"><f t="shared" si="0"/><v/></c>'
        ),
        r'<c r="B3"[^>]*>\s*<f>A3\*2</f>\s*(?:<v(?:>[^<]*</v>|\s*/>)\s*)?</c>': (
            '<c r="B3"><f t="shared" si="0"/><v/></c>'
        ),
    }
    for pattern, replacement in replacements.items():
        sheet_xml, count = re.subn(pattern, replacement, sheet_xml)
        assert count == 1
    entries[sheet_name] = sheet_xml.encode()
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in entries.items():
            zout.writestr(name, data)


def _make_image_xlsx(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Images"
    ws["A1"] = "with image"
    ws.add_image(OpenpyxlImage(PNG_PATH), "B5")
    wb.save(path)
    wb.close()


def _make_wolfxl_anchor_image_xlsx(path: Path) -> None:
    from wolfxl.drawing import Image as WolfImage
    from wolfxl.drawing.spreadsheet_drawing import (
        AbsoluteAnchor,
        AnchorMarker,
        TwoCellAnchor,
        XDRPoint2D,
        XDRPositiveSize2D,
    )

    wb = wolfxl.Workbook()
    ws = wb.active
    ws.title = "Anchors"
    ws.add_image(
        WolfImage(PNG_PATH),
        TwoCellAnchor(
            _from=AnchorMarker(col=1, row=1, colOff=10, rowOff=20),
            to=AnchorMarker(col=3, row=4, colOff=30, rowOff=40),
            editAs="twoCell",
        ),
    )
    ws.add_image(
        WolfImage(PNG_PATH),
        AbsoluteAnchor(
            pos=XDRPoint2D(x=1000, y=2000),
            ext=XDRPositiveSize2D(cx=3000, cy=4000),
        ),
    )
    wb.save(path)


def _make_sheet_state_xlsx(path: Path) -> None:
    wb = openpyxl.Workbook()
    wb.active.title = "Visible"
    hidden = wb.create_sheet("Hidden")
    hidden.sheet_state = "hidden"
    very_hidden = wb.create_sheet("VeryHidden")
    very_hidden.sheet_state = "veryHidden"
    wb.save(path)
    wb.close()


def _make_workbook_calc_properties_xlsx(path: Path) -> None:
    from openpyxl.packaging.custom import (
        BoolProperty,
        DateTimeProperty,
        FloatProperty,
        IntProperty,
        LinkProperty,
        StringProperty,
    )

    wb = openpyxl.Workbook()
    wb.active.title = "Visible"
    wb.create_sheet("Second")
    wb.create_sheet("Third")
    wb.custom_doc_props.append(StringProperty(name="Client", value="ACME"))
    wb.custom_doc_props.append(IntProperty(name="Count", value=42))
    wb.custom_doc_props.append(FloatProperty(name="Ratio", value=2.5))
    wb.custom_doc_props.append(BoolProperty(name="Reviewed", value=True))
    wb.custom_doc_props.append(
        DateTimeProperty(name="AsOf", value=dt.datetime(2024, 1, 2, 3, 4, 5))
    )
    wb.custom_doc_props.append(LinkProperty(name="LinkedCell", value="Visible!A1"))
    wb.save(path)
    wb.close()
    _inject_workbook_calc_properties(path)
    _inject_workbook_views(path)


def _make_print_area_xlsx(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Print"
    ws.print_area = "A1:D10"
    wb.save(path)
    wb.close()


def _make_print_titles_xlsx(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Titles"
    ws.print_title_rows = "1:2"
    ws.print_title_cols = "A:B"
    wb.save(path)
    wb.close()


def _make_print_setup_xlsx(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Setup"
    ws.page_margins.left = 1.1
    ws.page_margins.right = 1.2
    ws.page_margins.top = 1.3
    ws.page_margins.bottom = 1.4
    ws.page_margins.header = 0.5
    ws.page_margins.footer = 0.6
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 9
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.scale = 75
    ws.oddHeader.left.text = "Left"
    ws.oddHeader.center.text = "Center"
    ws.oddFooter.right.text = "Page &P"
    ws.HeaderFooter.differentFirst = True
    ws.HeaderFooter.alignWithMargins = False
    wb.save(path)
    wb.close()


def _make_page_breaks_xlsx(path: Path) -> None:
    from openpyxl.worksheet.pagebreak import Break

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Breaks"
    ws.row_breaks.append(Break(id=5, min=0, max=16383, man=True))
    ws.row_breaks.append(Break(id=10, min=0, max=16383, man=False))
    ws.col_breaks.append(Break(id=3, min=0, max=1048575, man=True))
    ws.sheet_format.defaultRowHeight = 22.0
    ws.sheet_format.defaultColWidth = 12.5
    ws.sheet_format.outlineLevelRow = 2
    ws.sheet_format.outlineLevelCol = 1
    ws.sheet_format.thickTop = True
    wb.save(path)
    wb.close()
    _inject_sheet_format_outline_col(path)


def _inject_sheet_format_outline_col(path: Path) -> None:
    with zipfile.ZipFile(path, "r") as zin:
        entries = {name: zin.read(name) for name in zin.namelist()}
    sheet_name = "xl/worksheets/sheet1.xml"
    sheet_xml = entries[sheet_name].decode()
    sheet_xml, count = re.subn(
        r"(<sheetFormatPr\b(?![^>]*outlineLevelCol)(?:[^>/]|/(?!>))*)/>",
        r'\1 outlineLevelCol="1"/>',
        sheet_xml,
        count=1,
    )
    assert count == 1
    entries[sheet_name] = sheet_xml.encode()
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in entries.items():
            zout.writestr(name, data)


def _make_sheet_view_xlsx(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Views"
    ws.sheet_view.zoomScale = 150
    ws.sheet_view.zoomScaleNormal = 120
    ws.sheet_view.view = "pageLayout"
    ws.sheet_view.showGridLines = False
    ws.sheet_view.showRowColHeaders = False
    ws.sheet_view.showZeros = False
    ws.sheet_view.rightToLeft = True
    ws.sheet_view.tabSelected = True
    ws.sheet_view.topLeftCell = "C3"
    ws.freeze_panes = "B2"
    ws.sheet_view.selection[0].activeCell = "C3"
    ws.sheet_view.selection[0].sqref = "C3:D4"
    wb.save(path)
    wb.close()


def _make_sheet_properties_xlsx(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet Properties"
    ws["A1"] = "metadata"
    wb.save(path)
    wb.close()
    _inject_sheet_properties(path)


def _make_chart_xlsx(path: Path) -> None:
    from openpyxl.chart import BarChart, Reference

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Charts"
    rows = [
        ("Month", "Sales"),
        ("Jan", 10),
        ("Feb", 20),
        ("Mar", 30),
    ]
    for row in rows:
        ws.append(row)
    chart = BarChart()
    chart.title = "Sales Trend"
    chart.x_axis.title = "Month"
    chart.y_axis.title = "Sales"
    chart.style = 10
    chart.type = "bar"
    chart.grouping = "stacked"
    chart.varyColors = True
    chart.legend.position = "t"
    data = Reference(ws, min_col=2, min_row=1, max_row=4)
    cats = Reference(ws, min_col=1, min_row=2, max_row=4)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "D5")
    wb.save(path)
    wb.close()


def _make_chart_family_xlsx(path: Path) -> None:
    from openpyxl.chart import (
        AreaChart,
        BubbleChart,
        LineChart,
        PieChart,
        RadarChart,
        Reference,
        ScatterChart,
        Series,
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Chart Families"
    rows = [
        ("Month", "Sales", "Expenses", "Size"),
        ("Jan", 10, 7, 3),
        ("Feb", 20, 12, 4),
        ("Mar", 30, 15, 5),
    ]
    for row in rows:
        ws.append(row)

    labels = Reference(ws, min_col=1, min_row=2, max_row=4)
    sales_with_title = Reference(ws, min_col=2, min_row=1, max_row=4)
    sales = Reference(ws, min_col=2, min_row=2, max_row=4)
    expenses = Reference(ws, min_col=3, min_row=2, max_row=4)
    bubble_sizes = Reference(ws, min_col=4, min_row=2, max_row=4)

    categorical_chart_specs = [
        (LineChart(), "Line Trend", "F2"),
        (PieChart(), "Pie Mix", "F14"),
        (AreaChart(), "Area Trend", "F26"),
        (RadarChart(), "Radar Trend", "F38"),
    ]
    for chart, title, anchor in categorical_chart_specs:
        chart.title = title
        chart.add_data(sales_with_title, titles_from_data=True)
        chart.set_categories(labels)
        ws.add_chart(chart, anchor)

    scatter = ScatterChart()
    scatter.title = "Scatter Trend"
    scatter.scatterStyle = "smoothMarker"
    scatter.series.append(Series(sales, labels, title="Scatter Sales"))
    ws.add_chart(scatter, "F50")

    bubble = BubbleChart()
    bubble.title = "Bubble Trend"
    bubble.series.append(
        Series(expenses, sales, bubble_sizes, title="Bubble Expenses")
    )
    ws.add_chart(bubble, "F62")

    wb.save(path)
    wb.close()


def test_native_reader_flag_loads_path_values(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    path = tmp_path / "native-smoke.xlsx"
    _make_basic_xlsx(path)

    monkeypatch.setenv("WOLFXL_NATIVE_READER", "1")
    wb = wolfxl.load_workbook(path, data_only=False)
    try:
        assert wb._rust_reader.__class__.__name__ == "NativeXlsxBook"  # noqa: SLF001
        assert wb.sheetnames == ["Data"]
        ws = wb["Data"]
        assert ws["A1"].value == "Label"
        assert ws["B1"].value == 42
        assert ws["B1"].number_format == "#,##0.00"
        assert ws["B1"].font.name == "Arial"
        assert ws["B1"].font.size == 14.0
        assert ws["B1"].font.bold is True
        assert ws["B1"].font.italic is True
        assert ws["B1"].font.underline == "single"
        assert ws["B1"].font.strike is True
        assert ws["B1"].font.color == "#123456"
        assert ws["B1"].fill.patternType == "solid"
        assert ws["B1"].fill.fgColor == "#ABCDEF"
        assert ws["B1"].alignment.horizontal == "center"
        assert ws["B1"].alignment.vertical == "top"
        assert ws["B1"].alignment.wrap_text is True
        assert ws["B1"].alignment.text_rotation == 45
        assert ws["B1"].alignment.indent == 2
        assert ws["B1"].border.left.style == "thin"
        assert ws["B1"].border.left.color == "#FF0000"
        assert ws["B1"].border.right.style == "medium"
        assert ws["B1"].border.right.color == "#000000"
        assert ws["B1"].border.top.style == "double"
        assert ws["B1"].border.top.color == "#00FF00"
        assert ws["B1"].border.bottom.style == "dashed"
        assert ws["B1"].border.bottom.color == "#0000FF"
        assert ws["A1"].border.left.style is None
        assert ws["C1"].value is True
        assert ws["B2"].value == "=B1*2"
        assert wb._rust_reader.read_cell_formula("Data", "B2") == {  # noqa: SLF001
            "type": "formula",
            "formula": "=B1*2",
            "value": "=B1*2",
        }
        assert wb._rust_reader.read_cell_formula("Data", "A1") is None  # noqa: SLF001
        assert ws["A3"].value == dt.datetime(2024, 1, 15, 12, 30)
        assert ws["B3"].value == dt.datetime(2024, 6, 1, 0, 0)
        assert ws["A5"].hyperlink is not None
        assert ws["A5"].hyperlink.target == "https://example.com/report"
        assert ws["A5"].hyperlink.display == "External"
        assert ws["A5"].hyperlink.tooltip == "Example report"
        assert ws["B5"].hyperlink is not None
        assert ws["B5"].hyperlink.target is None
        assert ws["B5"].hyperlink.location == "Data!A1"
        assert ws["B5"].hyperlink.display == "Jump"
        assert ws["A6"].comment is not None
        assert ws["A6"].comment.text == "Native reader note"
        assert ws["A6"].comment.author == "Wolf"
        assert ws["A6"].comment.parent is ws["A6"]
        assert ws["A6"].comment.height == 79
        assert ws["A6"].comment.width == 144
        assert ws.row_dimensions[6].height == 24
        assert ws.row_dimensions[6].index == 6
        assert ws.row_dimensions[6].r == 6
        assert ws.row_dimensions[6].ht == 24
        assert ws.row_dimensions[6].customHeight is True
        assert ws.column_dimensions["C"].width == 18
        assert ws.column_dimensions["C"].index == "C"
        assert ws.column_dimensions["C"].customWidth is True
        assert ws.column_dimensions["C"].min == 3
        assert ws.column_dimensions["C"].max == 3
        assert ws.column_dimensions["C"].range == "C:C"
        validations = list(ws.data_validations)
        assert len(validations) == 1
        assert validations[0].type == "list"
        assert validations[0].formula1 == '="Red,Blue"'
        assert validations[0].allowBlank is True
        assert validations[0].sqref == "C2:C6"
        assert ws.freeze_panes == "B2"
        assert ws.auto_filter.ref == "A1:D6"
        assert len(ws.auto_filter.filter_columns) == 2
        assert ws.auto_filter.filter_columns[0].col_id == 0
        assert ws.auto_filter.filter_columns[0].filter.values == ["Label"]
        assert ws.auto_filter.filter_columns[1].filter.and_ is True
        assert ws.auto_filter.filter_columns[1].filter.customFilter[0].operator == "greaterThan"
        assert ws.auto_filter.filter_columns[1].filter.customFilter[0].val == "10"
        assert ws.auto_filter.sort_state is not None
        assert ws.auto_filter.sort_state.ref == "A2:D6"
        assert ws.auto_filter.sort_state.sort_conditions[0].ref == "B2:B6"
        assert ws.auto_filter.sort_state.sort_conditions[0].descending is True
        assert ws.protection.sheet is True
        assert ws.protection.objects is True
        assert ws.protection.formatCells is False
        assert ws.protection.sort is False
        assert ws.protection.password == "C258"
        assert wb.security is not None
        assert wb.security.lock_structure is True
        assert wb.security.workbook_algorithm_name == "SHA-512"
        assert wb.security.workbook_hash_value == "HASH"
        assert wb.security.workbook_salt_value == "SALT"
        assert wb.security.workbook_spin_count == 100000
        assert wb.fileSharing is not None
        assert wb.fileSharing.read_only_recommended is True
        assert wb.fileSharing.user_name == "Wolf"
        assert wb.fileSharing.algorithm_name == "SHA-512"
        assert wb.fileSharing.hash_value == "FILEHASH"
        assert wb.fileSharing.salt_value == "FILESALT"
        assert wb.fileSharing.spin_count == 100000
        assert {str(r) for r in ws.merged_cells.ranges} == {"D1:E1"}
        assert ws["D1"].number_format == "#,##0"
        assert ws["E1"].number_format is None
        assert ws["E1"].font.name is None
        records = {
            record["coordinate"]: record
            for record in ws.cell_records(include_format=True, include_empty=True)
        }
        assert records["B1"]["number_format"] == "#,##0.00"
        assert "number_format" not in records["E1"]
        assert records["A3"]["data_type"] == "datetime"
    finally:
        wb.close()


def test_native_reader_loads_workbook_sheet_states(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    path = tmp_path / "native-sheet-states.xlsx"
    _make_sheet_state_xlsx(path)

    monkeypatch.setenv("WOLFXL_NATIVE_READER", "1")
    wb = wolfxl.load_workbook(path)
    try:
        assert wb.sheetnames == ["Visible", "Hidden", "VeryHidden"]
        assert wb["Visible"].sheet_state == "visible"
        assert wb["Hidden"].sheet_state == "hidden"
        assert wb["VeryHidden"].sheet_state == "veryHidden"
    finally:
        wb.close()


def test_native_reader_loads_workbook_calc_properties(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    path = tmp_path / "native-workbook-properties.xlsx"
    _make_workbook_calc_properties_xlsx(path)

    monkeypatch.setenv("WOLFXL_NATIVE_READER", "1")
    wb = wolfxl.load_workbook(path)
    try:
        properties = wb.workbook_properties
        assert properties.date1904 is True
        from wolfxl.utils.datetime import CALENDAR_MAC_1904

        assert wb.epoch == CALENDAR_MAC_1904
        assert wb.excel_base_date == CALENDAR_MAC_1904
        assert wb.iso_dates is False
        assert wb.template is False
        assert wb.is_template is False
        assert wb.encoding == "utf-8"
        assert wb.data_only is False
        assert wb.read_only is False
        assert wb.write_only is False
        assert wb.path == "/xl/workbook.xml"
        assert wb.rels == []
        assert wb.shared_strings == []
        assert wb.loaded_theme is None
        assert wb.vba_archive is None
        assert wb.style_names == ["Normal"]
        assert wb.custom_doc_props.names == [
            "Client",
            "Count",
            "Ratio",
            "Reviewed",
            "AsOf",
            "LinkedCell",
        ]
        assert wb.custom_doc_props["Client"].value == "ACME"
        assert wb.custom_doc_props["Count"].value == 42
        assert wb.custom_doc_props["Ratio"].value == 2.5
        assert wb.custom_doc_props["Reviewed"].value is True
        assert wb.custom_doc_props["AsOf"].value == dt.datetime(2024, 1, 2, 3, 4, 5)
        assert wb.custom_doc_props["LinkedCell"].value == "Visible!A1"
        assert (
            wb.mime_type
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"
        )
        assert wb.code_name == "NativeWorkbook"
        assert properties.dateCompatibility is False
        assert properties.showObjects == "none"
        assert properties.filterPrivacy is True
        assert properties.backupFile is True
        assert properties.updateLinks == "never"
        assert properties.codeName == "NativeWorkbook"
        assert properties.refreshAllConnections is True
        assert properties.defaultThemeVersion == 164011

        calculation = wb.calculation
        assert wb.calc_properties is calculation
        assert calculation.calcId == 191029
        assert calculation.calcMode == "manual"
        assert calculation.fullCalcOnLoad is True
        assert calculation.refMode == "R1C1"
        assert calculation.iterate is True
        assert calculation.iterateCount == 25
        assert calculation.iterateDelta == 0.01
        assert calculation.fullPrecision is False
        assert calculation.calcCompleted is False
        assert calculation.calcOnSave is False
        assert calculation.concurrentCalc is False
        assert calculation.concurrentManualCount == 4
        assert calculation.forceFullCalc is True

        assert len(wb.views) == 1
        view = wb.views[0]
        assert view.visibility == "hidden"
        assert view.minimized is True
        assert view.showHorizontalScroll is False
        assert view.showVerticalScroll is False
        assert view.showSheetTabs is False
        assert view.xWindow == 10
        assert view.yWindow == 20
        assert view.windowWidth == 12000
        assert view.windowHeight == 8000
        assert view.tabRatio == 750
        assert view.firstSheet == 1
        assert view.activeTab == 2
        assert view.autoFilterDateGrouping is False
    finally:
        wb.close()


def test_native_reader_loads_print_area(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    path = tmp_path / "native-print-area.xlsx"
    _make_print_area_xlsx(path)
    expected = openpyxl.load_workbook(path)
    try:
        expected_print_area = expected["Print"].print_area
    finally:
        expected.close()

    monkeypatch.setenv("WOLFXL_NATIVE_READER", "1")
    wb = wolfxl.load_workbook(path)
    try:
        assert wb["Print"].print_area == expected_print_area
    finally:
        wb.close()


def test_native_reader_loads_print_titles(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    path = tmp_path / "native-print-titles.xlsx"
    _make_print_titles_xlsx(path)

    monkeypatch.setenv("WOLFXL_NATIVE_READER", "1")
    wb = wolfxl.load_workbook(path)
    try:
        ws = wb["Titles"]
        assert ws.print_title_rows == "1:2"
        assert ws.print_title_cols == "A:B"
    finally:
        wb.close()


def test_native_reader_loads_print_setup_metadata(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    path = tmp_path / "native-print-setup.xlsx"
    _make_print_setup_xlsx(path)

    monkeypatch.setenv("WOLFXL_NATIVE_READER", "1")
    wb = wolfxl.load_workbook(path)
    try:
        ws = wb["Setup"]
        assert ws.page_margins.left == 1.1
        assert ws.page_margins.right == 1.2
        assert ws.page_margins.top == 1.3
        assert ws.page_margins.bottom == 1.4
        assert ws.page_margins.header == 0.5
        assert ws.page_margins.footer == 0.6
        assert ws.page_setup.orientation == "landscape"
        assert ws.page_setup.paperSize == 9
        assert ws.page_setup.fitToWidth == 1
        assert ws.page_setup.fitToHeight == 0
        assert ws.page_setup.scale == 75
        assert ws.header_footer.odd_header.left == "Left"
        assert ws.header_footer.odd_header.center == "Center"
        assert ws.header_footer.odd_footer.right == "Page &P"
        assert ws.header_footer.different_first is True
        assert ws.header_footer.align_with_margins is False
    finally:
        wb.close()


def test_native_reader_loads_page_breaks_and_sheet_format(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    path = tmp_path / "native-page-breaks.xlsx"
    _make_page_breaks_xlsx(path)

    monkeypatch.setenv("WOLFXL_NATIVE_READER", "1")
    wb = wolfxl.load_workbook(path)
    try:
        ws = wb["Breaks"]
        assert ws.row_breaks.count == 2
        assert ws.row_breaks.manualBreakCount == 2
        row_breaks = list(ws.row_breaks)
        assert [brk.id for brk in row_breaks] == [5, 10]
        assert row_breaks[0].max == 16383
        assert row_breaks[0].man is True
        assert row_breaks[1].man is False

        assert ws.col_breaks.count == 1
        col_break = list(ws.col_breaks)[0]
        assert col_break.id == 3
        assert col_break.max == 1048575

        assert ws.sheet_format.defaultRowHeight == 22.0
        assert ws.sheet_format.defaultColWidth == 12.5
        assert ws.sheet_format.outlineLevelRow == 2
        assert ws.sheet_format.outlineLevelCol == 1
        assert ws.sheet_format.thickTop is True
    finally:
        wb.close()


def test_native_reader_loads_sheet_view_metadata(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    path = tmp_path / "native-sheet-view.xlsx"
    _make_sheet_view_xlsx(path)

    monkeypatch.setenv("WOLFXL_NATIVE_READER", "1")
    wb = wolfxl.load_workbook(path)
    try:
        ws = wb["Views"]
        view = ws.sheet_view
        assert view.zoomScale == 150
        assert view.zoomScaleNormal == 120
        assert view.view == "pageLayout"
        assert view.showGridLines is False
        assert view.showRowColHeaders is False
        assert view.showZeros is False
        assert view.rightToLeft is True
        assert view.tabSelected is True
        assert view.topLeftCell == "C3"
        assert view.pane is not None
        assert view.pane.topLeftCell == "B2"
        assert view.pane.activePane == "bottomRight"
        assert view.selection[0].activeCell == "C3"
        assert view.selection[0].sqref == "C3:D4"
    finally:
        wb.close()


def test_native_reader_loads_sheet_properties(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    path = tmp_path / "native-sheet-properties.xlsx"
    _make_sheet_properties_xlsx(path)

    monkeypatch.setenv("WOLFXL_NATIVE_READER", "1")
    wb = wolfxl.load_workbook(path)
    try:
        ws = wb["Sheet Properties"]
        properties = ws.sheet_properties
        assert properties.codeName == "NativeProps"
        assert properties.enableFormatConditionsCalculation is False
        assert properties.filterMode is True
        assert properties.published is False
        assert properties.syncHorizontal is True
        assert properties.syncRef == "B2"
        assert properties.syncVertical is True
        assert properties.transitionEvaluation is True
        assert properties.transitionEntry is True
        assert properties.tabColor == "33AA55"
        assert properties.outlinePr.summaryBelow is False
        assert properties.outlinePr.summaryRight is False
        assert properties.outlinePr.applyStyles is True
        assert properties.outlinePr.showOutlineSymbols is False
        assert properties.pageSetUpPr.autoPageBreaks is False
        assert properties.pageSetUpPr.fitToPage is True
    finally:
        wb.close()


def test_native_reader_loads_drawing_images(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    path = tmp_path / "native-images.xlsx"
    _make_image_xlsx(path)

    monkeypatch.setenv("WOLFXL_NATIVE_READER", "1")
    wb = wolfxl.load_workbook(path)
    try:
        ws = wb["Images"]
        images = ws._images  # noqa: SLF001
        assert len(images) == 1
        img = images[0]
        assert img.format == "png"
        assert img._data == PNG_PATH.read_bytes()  # noqa: SLF001

        from wolfxl.drawing.spreadsheet_drawing import OneCellAnchor

        assert isinstance(img.anchor, OneCellAnchor)
        assert img.anchor._from.col == 1
        assert img.anchor._from.row == 4
        assert img.anchor.ext is not None
        assert img.anchor.ext.cx > 0
        assert img.anchor.ext.cy > 0
    finally:
        wb.close()


def test_native_reader_loads_drawing_charts(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    path = tmp_path / "native-charts.xlsx"
    _make_chart_xlsx(path)

    monkeypatch.setenv("WOLFXL_NATIVE_READER", "1")
    wb = wolfxl.load_workbook(path)
    try:
        ws = wb["Charts"]
        charts = ws._charts  # noqa: SLF001
        assert len(charts) == 1
        chart = charts[0]
        assert chart.__class__.__name__ == "BarChart"
        assert chart.title.tx.rich.paragraphs[0].r[0].t == "Sales Trend"
        assert chart.x_axis.title.tx.rich.paragraphs[0].r[0].t == "Month"
        assert chart.y_axis.title.tx.rich.paragraphs[0].r[0].t == "Sales"
        assert chart.style == 10
        assert chart.type == "bar"
        assert chart.grouping == "stacked"
        assert chart.varyColors is True
        assert chart.legend.position == "t"
        assert len(chart.series) == 1
        series = chart.series[0]
        assert series.tx.strRef.f == "'Charts'!B1"
        assert series.cat.strRef.f == "'Charts'!$A$2:$A$4"
        assert series.val.numRef.f == "'Charts'!$B$2:$B$4"

        from wolfxl.drawing.spreadsheet_drawing import OneCellAnchor

        assert isinstance(chart._anchor, OneCellAnchor)  # noqa: SLF001
        assert chart._anchor._from.col == 3  # noqa: SLF001
        assert chart._anchor._from.row == 4  # noqa: SLF001
    finally:
        wb.close()


def test_native_reader_loads_common_chart_families(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    path = tmp_path / "native-chart-families.xlsx"
    _make_chart_family_xlsx(path)

    monkeypatch.setenv("WOLFXL_NATIVE_READER", "1")
    wb = wolfxl.load_workbook(path)
    try:
        ws = wb["Chart Families"]
        charts_by_title = {
            chart.title.tx.rich.paragraphs[0].r[0].t: chart
            for chart in ws._charts  # noqa: SLF001
        }
        expected_classes = {
            "Line Trend": "LineChart",
            "Pie Mix": "PieChart",
            "Area Trend": "AreaChart",
            "Radar Trend": "RadarChart",
            "Scatter Trend": "ScatterChart",
            "Bubble Trend": "BubbleChart",
        }
        assert set(charts_by_title) == set(expected_classes)

        for title, class_name in expected_classes.items():
            chart = charts_by_title[title]
            assert chart.__class__.__name__ == class_name
            assert len(chart.series) == 1

        for title in ["Line Trend", "Pie Mix", "Area Trend", "Radar Trend"]:
            series = charts_by_title[title].series[0]
            assert series.tx.strRef.f == "'Chart Families'!B1"
            assert series.cat.strRef.f == "'Chart Families'!$A$2:$A$4"
            assert series.val.numRef.f == "'Chart Families'!$B$2:$B$4"

        scatter_series = charts_by_title["Scatter Trend"].series[0]
        assert charts_by_title["Scatter Trend"].scatterStyle == "smoothMarker"
        assert scatter_series.tx.v == "Scatter Sales"
        assert scatter_series.xVal.numRef.f == "'Chart Families'!$A$2:$A$4"
        assert scatter_series.yVal.numRef.f == "'Chart Families'!$B$2:$B$4"

        bubble_series = charts_by_title["Bubble Trend"].series[0]
        assert bubble_series.tx.v == "Bubble Expenses"
        assert bubble_series.xVal.numRef.f == "'Chart Families'!$B$2:$B$4"
        assert bubble_series.yVal.numRef.f == "'Chart Families'!$C$2:$C$4"
        assert bubble_series.bubbleSize.numRef.f == "'Chart Families'!$D$2:$D$4"
    finally:
        wb.close()


def test_native_reader_preserves_image_anchor_types(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    path = tmp_path / "native-image-anchors.xlsx"
    _make_wolfxl_anchor_image_xlsx(path)

    monkeypatch.setenv("WOLFXL_NATIVE_READER", "1")
    wb = wolfxl.load_workbook(path)
    try:
        ws = wb["Anchors"]
        images = ws._images  # noqa: SLF001
        assert len(images) == 2

        from wolfxl.drawing.spreadsheet_drawing import AbsoluteAnchor, TwoCellAnchor

        two_cell = images[0].anchor
        assert isinstance(two_cell, TwoCellAnchor)
        assert two_cell._from.col == 1
        assert two_cell._from.row == 1
        assert two_cell._from.colOff == 10
        assert two_cell._from.rowOff == 20
        assert two_cell.to.col == 3
        assert two_cell.to.row == 4
        assert two_cell.to.colOff == 30
        assert two_cell.to.rowOff == 40
        assert two_cell.editAs == "twoCell"

        absolute = images[1].anchor
        assert isinstance(absolute, AbsoluteAnchor)
        assert absolute.pos.x == 1000
        assert absolute.pos.y == 2000
        assert absolute.ext.cx == 3000
        assert absolute.ext.cy == 4000
    finally:
        wb.close()


def test_native_reader_expands_shared_formula_children(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    path = tmp_path / "native-shared-formulas.xlsx"
    _make_shared_formula_xlsx(path)
    expected_book = openpyxl.load_workbook(path, data_only=False)
    try:
        expected = [expected_book.active[f"B{row}"].value for row in range(1, 4)]
    finally:
        expected_book.close()

    monkeypatch.setenv("WOLFXL_NATIVE_READER", "1")
    wb = wolfxl.load_workbook(path, data_only=False)
    try:
        ws = wb["Sheet1"]
        assert [ws[f"B{row}"].value for row in range(1, 4)] == expected
        assert wb._rust_reader.read_cell_formula("Sheet1", "B2") == {  # noqa: SLF001
            "type": "formula",
            "formula": "=A2*2",
            "value": "=A2*2",
        }
    finally:
        wb.close()


def test_native_reader_flag_loads_bytes(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    path = tmp_path / "native-bytes.xlsx"
    _make_basic_xlsx(path)

    monkeypatch.setenv("WOLFXL_NATIVE_READER", "1")
    wb = wolfxl.load_workbook(path.read_bytes())
    try:
        assert wb._rust_reader.__class__.__name__ == "NativeXlsxBook"  # noqa: SLF001
        assert wb._rust_reader.opened_from_bytes() is True  # noqa: SLF001
        assert wb["Data"].iter_rows(values_only=True).__next__() == (
            "Label",
            42,
            True,
            "Anchor",
            None,
        )
        assert wb["Data"]["A5"].hyperlink.target == "https://example.com/report"
        assert wb["Data"]["A6"].comment.text == "Native reader note"
        assert wb["Data"].column_dimensions["C"].width == 18
        assert len(list(wb["Data"].data_validations)) == 1
    finally:
        wb.close()
