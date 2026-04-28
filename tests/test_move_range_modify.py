"""RFC-034 — ``Worksheet.move_range`` round-trip in modify mode.

End-to-end coverage for the paste-style range-move path. Threads
three layers:

1. ``Worksheet.move_range(cell_range, rows=0, cols=0, translate=False)``
   (Python) parses the A1 range, validates dest bounds, and appends a
   tuple to ``wb._pending_range_moves``.
2. ``Workbook._flush_pending_range_moves_to_patcher`` (Python) drains
   each tuple into ``XlsxPatcher.queue_range_move``.
3. ``XlsxPatcher::do_save`` Phase 2.5j (Rust) reads each affected sheet
   XML, calls ``wolfxl_structural::apply_range_move`` for each queued
   op, and folds the result back into ``file_patches``.

Sister contract: ``test_no_dirty_save_is_byte_identical`` confirms the
empty-queue path remains a no-op identity.
"""
from __future__ import annotations

import zipfile
from pathlib import Path

import openpyxl
import pytest

import wolfxl


# pytest marker so verify_rfc.py can collect this test.
pytestmark = pytest.mark.rfc034


@pytest.fixture(autouse=True)
def _force_test_epoch(monkeypatch: pytest.MonkeyPatch) -> None:
    """Pin ZIP entry mtimes for byte-stable saves."""
    monkeypatch.setenv("WOLFXL_TEST_EPOCH", "0")


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


def _make_simple_fixture(path: Path) -> None:
    """Workbook with cells in C3:E5 (a 3x3 block)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for r in range(3, 6):
        for c in range(3, 6):
            ws.cell(row=r, column=c, value=f"r{r}c{c}")
    wb.save(path)


def _make_formula_fixture(path: Path) -> None:
    """Workbook where C3 has a relative ref =A1, D3 has an absolute ref
    =$A$1, and a cell outside the moved block (G7) has =C3 — used to
    exercise the translate=True external-rewrite path."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 100  # anchor for absolute refs
    ws["C3"] = "=A1"  # relative ref — should NOT shift on the cell move
    ws["D3"] = "=$A$1"  # absolute ref — should NOT shift
    ws["G7"] = "=C3"  # external formula pointing INTO src=C3:D3
    wb.save(path)


def _make_merge_fixture(path: Path) -> None:
    """Workbook with a merge fully inside the source rectangle and a
    merge that straddles the boundary."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["C3"] = "merged-inside"
    ws.merge_cells("C3:D4")  # fully inside future src=C3:E5
    ws["A1"] = "merged-straddle"
    ws.merge_cells("A1:D2")  # straddles src boundary (A1, A2 outside)
    wb.save(path)


def _read_zip_text(path: Path, entry: str) -> str:
    with zipfile.ZipFile(path) as zf:
        return zf.read(entry).decode("utf-8")


# ---------------------------------------------------------------------------
# Tests — argument validation
# ---------------------------------------------------------------------------


def test_rejects_non_int_rows(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    _make_simple_fixture(src)
    wb = wolfxl.load_workbook(src, modify=True)
    ws = wb.active
    with pytest.raises(TypeError, match="rows"):
        ws.move_range("C3:E5", rows="2", cols=0)  # type: ignore[arg-type]


def test_rejects_non_int_cols(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    _make_simple_fixture(src)
    wb = wolfxl.load_workbook(src, modify=True)
    ws = wb.active
    with pytest.raises(TypeError, match="cols"):
        ws.move_range("C3:E5", rows=0, cols=2.5)  # type: ignore[arg-type]


def test_rejects_invalid_range_string(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    _make_simple_fixture(src)
    wb = wolfxl.load_workbook(src, modify=True)
    ws = wb.active
    with pytest.raises(ValueError):
        ws.move_range("not-a-range", rows=1, cols=1)


def test_rejects_destination_out_of_bounds(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    _make_simple_fixture(src)
    wb = wolfxl.load_workbook(src, modify=True)
    ws = wb.active
    # Move C3 up by 5 rows → row -2, out of bounds.
    with pytest.raises(ValueError, match="out of bounds"):
        ws.move_range("C3:E5", rows=-10, cols=0)


def test_zero_delta_is_noop_no_save_diff(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_simple_fixture(src)
    wb = wolfxl.load_workbook(src, modify=True)
    ws = wb.active
    ws.move_range("C3:E5", rows=0, cols=0)  # explicit no-op
    wb.save(dst)
    # No queued op → byte-identical.
    assert src.read_bytes() == dst.read_bytes()


# ---------------------------------------------------------------------------
# Tests — basic moves
# ---------------------------------------------------------------------------


def test_move_block_down_relocates_cells(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_simple_fixture(src)
    wb = wolfxl.load_workbook(src, modify=True)
    ws = wb.active
    ws.move_range("C3:E5", rows=5, cols=0)
    wb.save(dst)
    op = openpyxl.load_workbook(dst)
    s = op["Sheet1"]
    # Original C3 → C8.
    assert s["C8"].value == "r3c3"
    assert s["E10"].value == "r5c5"
    # Source slots are now blank.
    assert s["C3"].value is None
    assert s["E5"].value is None


def test_move_block_with_negative_rows_relocates_up(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    # Place block at C10:E12 so moving up by 5 lands at C5:E7.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for r in range(10, 13):
        for c in range(3, 6):
            ws.cell(row=r, column=c, value=f"r{r}c{c}")
    wb.save(src)

    wb2 = wolfxl.load_workbook(src, modify=True)
    ws2 = wb2.active
    ws2.move_range("C10:E12", rows=-5, cols=0)
    wb2.save(dst)
    op = openpyxl.load_workbook(dst)
    s = op["Sheet1"]
    assert s["C5"].value == "r10c3"
    assert s["E7"].value == "r12c5"
    assert s["C10"].value is None


def test_move_block_with_negative_cols_relocates_left(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_simple_fixture(src)
    wb = wolfxl.load_workbook(src, modify=True)
    ws = wb.active
    ws.move_range("C3:E5", rows=0, cols=-2)
    wb.save(dst)
    op = openpyxl.load_workbook(dst)
    s = op["Sheet1"]
    # Original C3 → A3.
    assert s["A3"].value == "r3c3"
    assert s["C5"].value == "r5c5"
    assert s["E5"].value is None


# ---------------------------------------------------------------------------
# Tests — formula handling
# ---------------------------------------------------------------------------


def test_absolute_refs_do_not_shift_inside_moved_block(tmp_path: Path) -> None:
    """`$A$1` is paste-style absolute — it must NOT shift even when the
    cell containing it relocates."""
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_formula_fixture(src)
    wb = wolfxl.load_workbook(src, modify=True)
    ws = wb.active
    ws.move_range("C3:D3", rows=5, cols=0)
    wb.save(dst)
    op = openpyxl.load_workbook(dst)
    s = op["Sheet1"]
    # C3 → C8: relative formula =A1. Translator: A1 is OUTSIDE src=C3:D3,
    # so move_range leaves it alone (matches openpyxl's `translate=False`
    # contract — refs outside src don't move).
    assert s["C8"].value == "=A1"
    # D3 → D8: absolute formula =$A$1 must remain unchanged.
    assert s["D8"].value == "=$A$1"


def test_relative_refs_inside_src_re_anchor(tmp_path: Path) -> None:
    """When a formula in src references another cell ALSO inside src,
    the ref re-anchors so the relationship is preserved at the new
    location."""
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    # Both C3 and D3 are inside src.
    ws["C3"] = 1
    ws["D3"] = "=C3"  # ref inside src → must re-anchor
    wb.save(src)

    wb2 = wolfxl.load_workbook(src, modify=True)
    ws2 = wb2.active
    ws2.move_range("C3:D3", rows=5, cols=0)
    wb2.save(dst)
    op = openpyxl.load_workbook(dst)
    s = op["Sheet1"]
    # D3 → D8; its reference to C3 → C8 (paste-translated by (5, 0)).
    assert s["D8"].value == "=C8"


def test_external_formula_not_translated_by_default(tmp_path: Path) -> None:
    """`translate=False` (default) leaves external formulas pointing at
    the OLD source coordinates."""
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_formula_fixture(src)
    wb = wolfxl.load_workbook(src, modify=True)
    ws = wb.active
    ws.move_range("C3:D3", rows=5, cols=0, translate=False)
    wb.save(dst)
    op = openpyxl.load_workbook(dst)
    s = op["Sheet1"]
    # G7 was =C3 — leave alone (translate=False).
    assert s["G7"].value == "=C3"


def test_external_formula_translated_when_requested(tmp_path: Path) -> None:
    """`translate=True` rewrites external formulas to re-anchor to the
    moved cells."""
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_formula_fixture(src)
    wb = wolfxl.load_workbook(src, modify=True)
    ws = wb.active
    ws.move_range("C3:D3", rows=5, cols=0, translate=True)
    wb.save(dst)
    op = openpyxl.load_workbook(dst)
    s = op["Sheet1"]
    # G7's =C3 → =C8 (re-anchored).
    assert s["G7"].value == "=C8"


# ---------------------------------------------------------------------------
# Tests — merge cells + anchors
# ---------------------------------------------------------------------------


def test_fully_inside_merge_shifts_with_block(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_merge_fixture(src)
    wb = wolfxl.load_workbook(src, modify=True)
    ws = wb.active
    ws.move_range("C3:E5", rows=5, cols=0)
    wb.save(dst)
    sheet_xml = _read_zip_text(dst, "xl/worksheets/sheet1.xml")
    # C3:D4 (fully inside) → C8:D9.
    assert 'ref="C8:D9"' in sheet_xml
    # A1:D2 (straddles) — left in place.
    assert 'ref="A1:D2"' in sheet_xml


# ---------------------------------------------------------------------------
# Tests — empty queue is a no-op
# ---------------------------------------------------------------------------


def test_no_dirty_save_is_byte_identical(tmp_path: Path) -> None:
    """The empty-queue path must remain a no-op identity per the RFC's
    no-op invariant."""
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_simple_fixture(src)
    wb = wolfxl.load_workbook(src, modify=True)
    # No move_range call.
    wb.save(dst)
    assert src.read_bytes() == dst.read_bytes()


# ---------------------------------------------------------------------------
# Tests — multi-op sequencing
# ---------------------------------------------------------------------------


def test_two_moves_compose_in_order(tmp_path: Path) -> None:
    """Two move_range calls should compose: the second sees the
    coordinate space produced by the first."""
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_simple_fixture(src)
    wb = wolfxl.load_workbook(src, modify=True)
    ws = wb.active
    # Move 1: C3:C3 → C8:C8 (down 5).
    ws.move_range("C3:C3", rows=5, cols=0)
    # Move 2: now C8:C8 → E8:E8 (right 2). E8 is empty before this move
    # so no overwrite drama.
    ws.move_range("C8:C8", rows=0, cols=2)
    wb.save(dst)
    op = openpyxl.load_workbook(dst)
    s = op["Sheet1"]
    assert s["E8"].value == "r3c3"
    assert s["C3"].value is None
    assert s["C8"].value is None
