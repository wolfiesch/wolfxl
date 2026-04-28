"""RFC-035 Sprint Θ Pod-C1 — write-mode ``Workbook.copy_worksheet``.

Modify-mode coverage lives in ``test_copy_worksheet_modify.py``.

Test menu:

A. Basic clone of a 5x5 grid in pure write mode (round-trips through save).
B. Default name dedup ("Sheet Copy", "Sheet Copy 2", ...).
C. Name collision via explicit ``name=`` raises ``ValueError``.
D. Cross-workbook source raises ``ValueError``.
E. Source's pending append/bulk buffers materialize before clone.
F. Sheet-scope properties (merges, freeze panes) propagate.
G. Cell formats (number_format) propagate.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from wolfxl import Workbook


pytestmark = pytest.mark.rfc035


@pytest.fixture(autouse=True)
def _force_test_epoch(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("WOLFXL_TEST_EPOCH", "0")


# ---------------------------------------------------------------------------
# A — Basic 5x5 clone, write-mode round-trip.
# ---------------------------------------------------------------------------


def test_a_basic_grid_clone(tmp_path: Path) -> None:
    out = tmp_path / "out.xlsx"

    wb = Workbook()
    src = wb.active
    assert src is not None
    src.title = "Template"
    for r in range(1, 6):
        for c in range(1, 6):
            src.cell(row=r, column=c, value=(r * 10 + c))

    dst = wb.copy_worksheet(src)
    assert dst.title == "Template Copy"
    assert wb.sheetnames == ["Template", "Template Copy"]

    wb.save(out)

    # Round-trip through openpyxl to validate the file.
    rwb = openpyxl.load_workbook(out)
    assert rwb.sheetnames == ["Template", "Template Copy"]
    for sheet_name in rwb.sheetnames:
        ws = rwb[sheet_name]
        for r in range(1, 6):
            for c in range(1, 6):
                assert ws.cell(row=r, column=c).value == (r * 10 + c), (
                    f"sheet={sheet_name} r{r}c{c}"
                )


# ---------------------------------------------------------------------------
# B — Default name dedup.
# ---------------------------------------------------------------------------


def test_b_default_name_dedup(tmp_path: Path) -> None:
    wb = Workbook()
    src = wb.active
    assert src is not None
    src.title = "Sheet"
    src["A1"] = "x"

    a = wb.copy_worksheet(src)
    b = wb.copy_worksheet(src)
    c = wb.copy_worksheet(src)

    assert a.title == "Sheet Copy"
    assert b.title == "Sheet Copy 2"
    assert c.title == "Sheet Copy 3"
    assert wb.sheetnames == [
        "Sheet",
        "Sheet Copy",
        "Sheet Copy 2",
        "Sheet Copy 3",
    ]

    out = tmp_path / "out.xlsx"
    wb.save(out)
    rwb = openpyxl.load_workbook(out)
    assert rwb.sheetnames == [
        "Sheet",
        "Sheet Copy",
        "Sheet Copy 2",
        "Sheet Copy 3",
    ]


# ---------------------------------------------------------------------------
# C — Name collision via explicit ``name=`` raises ``ValueError``.
# ---------------------------------------------------------------------------


def test_c_explicit_name_collision_raises() -> None:
    wb = Workbook()
    src = wb.active
    assert src is not None
    wb.create_sheet("Already")

    with pytest.raises(ValueError, match="already exists"):
        wb.copy_worksheet(src, name="Already")


# ---------------------------------------------------------------------------
# D — Cross-workbook source raises ``ValueError``.
# ---------------------------------------------------------------------------


def test_d_cross_workbook_source_raises() -> None:
    wb1 = Workbook()
    wb2 = Workbook()
    src = wb2.active
    assert src is not None

    with pytest.raises(ValueError, match="must belong to this workbook"):
        wb1.copy_worksheet(src)


# ---------------------------------------------------------------------------
# E — Pending append buffer is materialised before clone.
# ---------------------------------------------------------------------------


def test_e_pending_append_buffer_materialised(tmp_path: Path) -> None:
    out = tmp_path / "out.xlsx"
    wb = Workbook()
    src = wb.active
    assert src is not None
    src.title = "Data"
    src.append([1, 2, 3])
    src.append(["a", "b", "c"])
    # Note: do NOT call src.cell() in between — that would materialise
    # the buffer eagerly. We want the clone path to do it.

    dst = wb.copy_worksheet(src)
    assert dst.title == "Data Copy"

    wb.save(out)
    rwb = openpyxl.load_workbook(out)
    for sheet in (rwb["Data"], rwb["Data Copy"]):
        assert sheet["A1"].value == 1
        assert sheet["B1"].value == 2
        assert sheet["C1"].value == 3
        assert sheet["A2"].value == "a"
        assert sheet["B2"].value == "b"
        assert sheet["C2"].value == "c"


# ---------------------------------------------------------------------------
# F — Sheet-scope properties (merges, freeze panes) propagate.
# ---------------------------------------------------------------------------


def test_f_sheet_scope_properties_propagate(tmp_path: Path) -> None:
    out = tmp_path / "out.xlsx"
    wb = Workbook()
    src = wb.active
    assert src is not None
    src.title = "Layout"
    src["A1"] = "header"
    src["B1"] = "header2"
    src.merge_cells("A1:B1")
    src.freeze_panes = "A2"

    dst = wb.copy_worksheet(src)
    assert "A1:B1" in dst._merged_ranges  # noqa: SLF001
    assert dst._freeze_panes == "A2"  # noqa: SLF001

    wb.save(out)
    rwb = openpyxl.load_workbook(out)
    rdst = rwb["Layout Copy"]
    # openpyxl exposes the merged ranges as a list/set of strings.
    assert "A1:B1" in [str(r) for r in rdst.merged_cells.ranges]


# ---------------------------------------------------------------------------
# G — number_format propagates (smoke check on write-mode format flush).
# ---------------------------------------------------------------------------


def test_g_number_format_propagates(tmp_path: Path) -> None:
    out = tmp_path / "out.xlsx"
    wb = Workbook()
    src = wb.active
    assert src is not None
    src.title = "Money"
    src["A1"] = 1234.5
    src["A1"].number_format = "$#,##0.00"

    wb.copy_worksheet(src)
    wb.save(out)

    rwb = openpyxl.load_workbook(out)
    rdst = rwb["Money Copy"]
    assert rdst["A1"].value == 1234.5
    # openpyxl normalises the format string; assert non-default.
    assert "0.00" in rdst["A1"].number_format
