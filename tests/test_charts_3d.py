"""Sprint Μ-prime — 3D / Stock / Surface / OfPie chart family tests.

Two sub-suites coexist in this file:

1. **Pod-β′ write-mode + dict-contract tests** (RFC-046 §11). Verify
   construction, validation, and ``to_rust_dict()`` flat-shape output
   for the 8 new families. Cover write-mode round-trips through
   ``ws.add_chart()`` and the workbook's writer path.
2. **Pod-γ′ modify-mode round-trip tests** (RFC-046 §10.12). Load an
   existing xlsx, ``ws.add_chart(<NewFamilyChart>(...), "D2")`` via the
   serialize_chart_dict bridge, save, and verify the chart part is
   present + readable by openpyxl.

Pod-γ′ tests are clearly suffixed ``_modify_mode_round_trip`` so the
two sets can be discovered independently.
"""

from __future__ import annotations

import zipfile
from pathlib import Path
from typing import Any

import pytest

import wolfxl
from wolfxl.chart import (
    AreaChart3D,
    BarChart3D,
    LineChart3D,
    Pie3D,
    PieChart3D,
    ProjectedPieChart,
    Reference,
    StockChart,
    SurfaceChart,
    SurfaceChart3D,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _seed_data(ws: Any, n_cols: int = 2) -> None:
    """Populate ws with a simple n-column data block."""
    headers = [""] + [f"S{c}" for c in range(1, n_cols + 1)]
    ws.append(headers)
    for i in range(1, 6):
        row = [f"row{i}"] + [i * (10 + c) for c in range(n_cols)]
        ws.append(row)


def _basic_chart(chart_cls: type, **kwargs: Any) -> Any:
    """Build a chart with one series of data attached."""
    wb = wolfxl.Workbook()
    ws = wb.active
    _seed_data(ws, n_cols=2)
    chart = chart_cls(**kwargs)
    chart.add_data(
        Reference(ws, min_col=2, min_row=1, max_col=3, max_row=6),
        titles_from_data=True,
    )
    return chart, ws, wb


# ---------------------------------------------------------------------------
# BarChart3D
# ---------------------------------------------------------------------------


def test_bar_chart_3d_kind_and_view_3d_defaults() -> None:
    """BarChart3D emits kind='bar3d' with §11.1 default view_3d."""
    chart, _ws, _wb = _basic_chart(BarChart3D)
    d = chart.to_rust_dict()
    assert d["kind"] == "bar3d"
    assert d["series_type"] == "bar"
    assert d["bar_dir"] == "col"
    assert d["grouping"] == "clustered"
    assert d["view_3d"] == {
        "rot_x": 15,
        "rot_y": 20,
        "right_angle_axes": True,
        "depth_percent": 100,
    }


def test_bar_chart_3d_emits_z_axis() -> None:
    """3D variants emit a z_axis (serAx)."""
    chart, _ws, _wb = _basic_chart(BarChart3D)
    d = chart.to_rust_dict()
    assert d["z_axis"] is not None
    assert d["z_axis"]["ax_type"] == "ser"


def test_bar_chart_3d_view_3d_override() -> None:
    chart, _ws, _wb = _basic_chart(BarChart3D, view_3d={"rot_x": 30, "perspective": 25})
    d = chart.to_rust_dict()
    assert d["view_3d"]["rot_x"] == 30
    assert d["view_3d"]["perspective"] == 25
    # untouched defaults survive
    assert d["view_3d"]["rot_y"] == 20


# ---------------------------------------------------------------------------
# LineChart3D
# ---------------------------------------------------------------------------


def test_line_chart_3d_view_3d_defaults_match_openpyxl() -> None:
    """LineChart3D defaults: rot_x=15, rot_y=20, perspective=30."""
    chart, _ws, _wb = _basic_chart(LineChart3D)
    d = chart.to_rust_dict()
    assert d["kind"] == "line3d"
    assert d["view_3d"]["rot_x"] == 15
    assert d["view_3d"]["rot_y"] == 20
    assert d["view_3d"]["perspective"] == 30
    assert d["view_3d"]["depth_percent"] == 100


# ---------------------------------------------------------------------------
# PieChart3D / Pie3D alias
# ---------------------------------------------------------------------------


def test_pie_chart_3d_pie3d_alias_works() -> None:
    """Pie3D is a public alias for PieChart3D matching openpyxl's name."""
    assert Pie3D is PieChart3D
    chart, _ws, _wb = _basic_chart(Pie3D)
    d = chart.to_rust_dict()
    assert d["kind"] == "pie3d"


def test_pie_chart_3d_view_3d_defaults() -> None:
    chart, _ws, _wb = _basic_chart(PieChart3D)
    d = chart.to_rust_dict()
    assert d["view_3d"] == {
        "rot_x": 30,
        "rot_y": 0,
        "perspective": 30,
        "right_angle_axes": False,
    }


# ---------------------------------------------------------------------------
# AreaChart3D
# ---------------------------------------------------------------------------


def test_area_chart_3d_round_trip() -> None:
    chart, _ws, _wb = _basic_chart(AreaChart3D)
    d = chart.to_rust_dict()
    assert d["kind"] == "area3d"
    assert d["grouping"] == "standard"
    assert d["view_3d"]["depth_percent"] == 100


# ---------------------------------------------------------------------------
# SurfaceChart (2D)
# ---------------------------------------------------------------------------


def test_surface_chart_2d_wireframe_default_true() -> None:
    chart, _ws, _wb = _basic_chart(SurfaceChart)
    d = chart.to_rust_dict()
    assert d["kind"] == "surface"
    assert d["wireframe"] is True


def test_surface_chart_2d_wireframe_false() -> None:
    chart, _ws, _wb = _basic_chart(SurfaceChart, wireframe=False)
    d = chart.to_rust_dict()
    assert d["wireframe"] is False


# ---------------------------------------------------------------------------
# SurfaceChart3D
# ---------------------------------------------------------------------------


def test_surface_chart_3d_emits_surface3d_kind() -> None:
    chart, _ws, _wb = _basic_chart(SurfaceChart3D)
    d = chart.to_rust_dict()
    assert d["kind"] == "surface3d"
    assert d["wireframe"] is True
    assert d["view_3d"]["perspective"] == 30


# ---------------------------------------------------------------------------
# StockChart
# ---------------------------------------------------------------------------


def test_stock_chart_requires_exactly_4_series_empty() -> None:
    """StockChart raises ValueError when emit'd with no series."""
    chart = StockChart()
    with pytest.raises(ValueError, match="StockChart requires exactly 4 series"):
        chart.to_rust_dict()


def test_stock_chart_requires_exactly_4_series_two() -> None:
    """StockChart raises ValueError when emit'd with 2 series."""
    wb = wolfxl.Workbook()
    ws = wb.active
    _seed_data(ws, n_cols=2)
    chart = StockChart()
    chart.add_data(
        Reference(ws, min_col=2, min_row=1, max_col=3, max_row=6),
        titles_from_data=True,
    )
    with pytest.raises(ValueError, match="StockChart requires exactly 4 series"):
        chart.to_rust_dict()


def test_stock_chart_emits_hi_low_lines_and_up_down_bars() -> None:
    """StockChart with exactly 4 series emits hi_low_lines + up_down_bars."""
    wb = wolfxl.Workbook()
    ws = wb.active
    _seed_data(ws, n_cols=4)
    chart = StockChart()
    chart.add_data(
        Reference(ws, min_col=2, min_row=1, max_col=5, max_row=6),
        titles_from_data=True,
    )
    d = chart.to_rust_dict()
    assert d["kind"] == "stock"
    assert d["hi_low_lines"] is True
    assert d["up_down_bars"] is True
    assert len(d["series"]) == 4


def test_stock_chart_can_disable_decorators() -> None:
    wb = wolfxl.Workbook()
    ws = wb.active
    _seed_data(ws, n_cols=4)
    chart = StockChart(hi_low_lines=False, up_down_bars=False)
    chart.add_data(
        Reference(ws, min_col=2, min_row=1, max_col=5, max_row=6),
        titles_from_data=True,
    )
    d = chart.to_rust_dict()
    assert d["hi_low_lines"] is False
    assert d["up_down_bars"] is False


# ---------------------------------------------------------------------------
# ProjectedPieChart (of_pie)
# ---------------------------------------------------------------------------


def test_projected_pie_chart_default_is_pie_of_pie() -> None:
    chart, _ws, _wb = _basic_chart(ProjectedPieChart)
    d = chart.to_rust_dict()
    assert d["kind"] == "of_pie"
    assert d["of_pie_type"] == "pie"
    assert d["split_type"] == "auto"


def test_projected_pie_chart_bar_of_pie_variant() -> None:
    chart, _ws, _wb = _basic_chart(ProjectedPieChart, of_pie_type="bar")
    d = chart.to_rust_dict()
    assert d["of_pie_type"] == "bar"


def test_projected_pie_chart_split_type_validation() -> None:
    with pytest.raises(ValueError, match="split_type"):
        ProjectedPieChart(split_type="bogus")


def test_projected_pie_chart_of_pie_type_validation() -> None:
    with pytest.raises(ValueError, match="of_pie_type"):
        ProjectedPieChart(of_pie_type="circle")


def test_projected_pie_chart_split_pos_passes_through() -> None:
    chart, _ws, _wb = _basic_chart(
        ProjectedPieChart,
        of_pie_type="bar",
        split_type="val",
        split_pos=10,
        second_pie_size=75,
    )
    d = chart.to_rust_dict()
    assert d["split_pos"] == 10
    assert d["second_pie_size"] == 75


def test_projected_pie_chart_second_pie_size_validation() -> None:
    with pytest.raises(ValueError, match="second_pie_size"):
        ProjectedPieChart(second_pie_size=500)


# ---------------------------------------------------------------------------
# Cross-cutting: empty series guard hits 3D variants too
# ---------------------------------------------------------------------------


def test_bar_chart_3d_empty_series_raises() -> None:
    chart = BarChart3D()
    with pytest.raises(ValueError, match="requires at least one series"):
        chart.to_rust_dict()


def test_surface_chart_empty_series_raises() -> None:
    chart = SurfaceChart()
    with pytest.raises(ValueError, match="requires at least one series"):
        chart.to_rust_dict()


# ---------------------------------------------------------------------------
# Pod-γ′ — Modify-mode round-trip tests (RFC-046 §10.12).
# These exercise the Worksheet.add_chart() → serialize_chart_dict →
# patcher.queue_chart_add bridge end-to-end.
# ---------------------------------------------------------------------------

import openpyxl


pytestmark_modify = (
    pytest.mark.rfc046 if hasattr(pytest.mark, "rfc046") else pytest.mark.usefixtures()
)


@pytest.fixture(autouse=True)
def _force_test_epoch(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("WOLFXL_TEST_EPOCH", "0")


# ---------------------------------------------------------------------------
# Fixture helpers
# ---------------------------------------------------------------------------


def _make_data_fixture(path: Path, sheet_title: str = "Data") -> None:
    """A1:B5 mini table — enough to seed any 2D-style chart."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    for r in range(1, 6):
        ws.cell(r, 1, f"l{r}")
        ws.cell(r, 2, r * 10)
    wb.save(path)


def _make_ohlc_fixture(path: Path, sheet_title: str = "Data") -> None:
    """5×5 OHLC matrix for a StockChart (4 OHLC series).

    Layout::

        | label | open | high | low | close |
        | l1    | 10   | 12   | 9   | 11    |
        | ...   | ...  | ...  | ... | ...   |
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.cell(1, 1, "label")
    ws.cell(1, 2, "open")
    ws.cell(1, 3, "high")
    ws.cell(1, 4, "low")
    ws.cell(1, 5, "close")
    for r in range(2, 6):
        ws.cell(r, 1, f"l{r-1}")
        ws.cell(r, 2, 10 + r)
        ws.cell(r, 3, 12 + r)
        ws.cell(r, 4, 9 + r)
        ws.cell(r, 5, 11 + r)
    wb.save(path)


def _make_surface_grid_fixture(path: Path, sheet_title: str = "Data") -> None:
    """4x4 numeric grid for a SurfaceChart (each column is a series)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    for r in range(1, 5):
        for c in range(1, 5):
            ws.cell(r, c, (r * c) + 1)
    wb.save(path)


def _zip_listing(path: Path) -> list[str]:
    with zipfile.ZipFile(path, "r") as z:
        return sorted(z.namelist())


def _zip_read(path: Path, member: str) -> bytes:
    with zipfile.ZipFile(path, "r") as z:
        return z.read(member)


def _try_construct(chart_cls, *args, **kwargs):
    """Construct ``chart_cls`` or skip if Pod-β′ hasn't landed."""
    try:
        return chart_cls(*args, **kwargs)
    except NotImplementedError as exc:
        pytest.skip(f"{chart_cls.__name__} not yet implemented: {exc}")


# ---------------------------------------------------------------------------
# Modify-mode round-trip — 3D bar / line / pie / area
# ---------------------------------------------------------------------------


def test_bar_chart_3d_modify_mode_round_trip(tmp_path: Path) -> None:
    from wolfxl import load_workbook
    from wolfxl.chart import BarChart3D, Reference

    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_data_fixture(src)
    wb = load_workbook(src, modify=True)
    ws = wb["Data"]
    chart = _try_construct(BarChart3D)
    chart.add_data(
        Reference(ws, min_col=2, min_row=1, max_row=5),
        titles_from_data=False,
    )
    ws.add_chart(chart, "D2")
    wb.save(dst)

    entries = _zip_listing(dst)
    chart_files = [e for e in entries if e.startswith("xl/charts/chart")]
    assert chart_files, f"BarChart3D not emitted; entries={entries}"
    op = openpyxl.load_workbook(dst, data_only=False)
    assert len(op["Data"]._charts) >= 1


def test_line_chart_3d_modify_mode_round_trip(tmp_path: Path) -> None:
    from wolfxl import load_workbook
    from wolfxl.chart import LineChart3D, Reference

    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_data_fixture(src)
    wb = load_workbook(src, modify=True)
    ws = wb["Data"]
    chart = _try_construct(LineChart3D)
    chart.add_data(
        Reference(ws, min_col=2, min_row=1, max_row=5),
        titles_from_data=False,
    )
    ws.add_chart(chart, "D2")
    wb.save(dst)

    chart_files = [
        e for e in _zip_listing(dst) if e.startswith("xl/charts/chart")
    ]
    assert chart_files, "LineChart3D not emitted"


def test_pie_chart_3d_modify_mode_round_trip(tmp_path: Path) -> None:
    from wolfxl import load_workbook
    from wolfxl.chart import PieChart3D, Reference

    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_data_fixture(src)
    wb = load_workbook(src, modify=True)
    ws = wb["Data"]
    chart = _try_construct(PieChart3D)
    chart.add_data(
        Reference(ws, min_col=2, min_row=1, max_row=5),
        titles_from_data=False,
    )
    ws.add_chart(chart, "D2")
    wb.save(dst)

    chart_files = [
        e for e in _zip_listing(dst) if e.startswith("xl/charts/chart")
    ]
    assert chart_files, "PieChart3D not emitted"


def test_area_chart_3d_modify_mode_round_trip(tmp_path: Path) -> None:
    from wolfxl import load_workbook
    from wolfxl.chart import AreaChart3D, Reference

    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_data_fixture(src)
    wb = load_workbook(src, modify=True)
    ws = wb["Data"]
    chart = _try_construct(AreaChart3D)
    chart.add_data(
        Reference(ws, min_col=2, min_row=1, max_row=5),
        titles_from_data=False,
    )
    ws.add_chart(chart, "D2")
    wb.save(dst)

    chart_files = [
        e for e in _zip_listing(dst) if e.startswith("xl/charts/chart")
    ]
    assert chart_files, "AreaChart3D not emitted"


# ---------------------------------------------------------------------------
# Modify-mode round-trip — Surface (2D) and Surface3D
# ---------------------------------------------------------------------------


def test_surface_chart_modify_mode_round_trip(tmp_path: Path) -> None:
    from wolfxl import load_workbook
    from wolfxl.chart import Reference, SurfaceChart

    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_surface_grid_fixture(src)
    wb = load_workbook(src, modify=True)
    ws = wb["Data"]
    chart = _try_construct(SurfaceChart)
    chart.add_data(
        Reference(ws, min_col=1, max_col=4, min_row=1, max_row=4),
        titles_from_data=False,
    )
    ws.add_chart(chart, "F2")
    wb.save(dst)

    chart_files = [
        e for e in _zip_listing(dst) if e.startswith("xl/charts/chart")
    ]
    assert chart_files, "SurfaceChart not emitted"


def test_surface_chart_3d_modify_mode_round_trip(tmp_path: Path) -> None:
    from wolfxl import load_workbook
    from wolfxl.chart import Reference, SurfaceChart3D

    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_surface_grid_fixture(src)
    wb = load_workbook(src, modify=True)
    ws = wb["Data"]
    chart = _try_construct(SurfaceChart3D)
    chart.add_data(
        Reference(ws, min_col=1, max_col=4, min_row=1, max_row=4),
        titles_from_data=False,
    )
    ws.add_chart(chart, "F2")
    wb.save(dst)

    chart_files = [
        e for e in _zip_listing(dst) if e.startswith("xl/charts/chart")
    ]
    assert chart_files, "SurfaceChart3D not emitted"


# ---------------------------------------------------------------------------
# Modify-mode round-trip — Stock chart (4 OHLC series)
# ---------------------------------------------------------------------------


def test_stock_chart_modify_mode_round_trip(tmp_path: Path) -> None:
    from wolfxl import load_workbook
    from wolfxl.chart import Reference, StockChart

    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_ohlc_fixture(src)
    wb = load_workbook(src, modify=True)
    ws = wb["Data"]
    chart = _try_construct(StockChart)
    # Seed all 4 OHLC series (open / high / low / close).
    chart.add_data(
        Reference(ws, min_col=2, max_col=5, min_row=1, max_row=5),
        titles_from_data=True,
    )
    chart.set_categories(
        Reference(ws, min_col=1, max_col=1, min_row=2, max_row=5),
    )
    ws.add_chart(chart, "G2")
    wb.save(dst)

    chart_files = [
        e for e in _zip_listing(dst) if e.startswith("xl/charts/chart")
    ]
    assert chart_files, "StockChart not emitted"
    op = openpyxl.load_workbook(dst, data_only=False)
    assert len(op["Data"]._charts) >= 1


# ---------------------------------------------------------------------------
# Modify-mode round-trip — ProjectedPie (Pie of Pie / Bar of Pie)
# ---------------------------------------------------------------------------


def test_projected_pie_chart_modify_mode_round_trip(tmp_path: Path) -> None:
    from wolfxl import load_workbook
    from wolfxl.chart import ProjectedPieChart, Reference

    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_data_fixture(src)
    wb = load_workbook(src, modify=True)
    ws = wb["Data"]
    chart = _try_construct(ProjectedPieChart)
    chart.add_data(
        Reference(ws, min_col=2, min_row=1, max_row=5),
        titles_from_data=False,
    )
    ws.add_chart(chart, "D2")
    wb.save(dst)

    chart_files = [
        e for e in _zip_listing(dst) if e.startswith("xl/charts/chart")
    ]
    assert chart_files, "ProjectedPieChart not emitted"
