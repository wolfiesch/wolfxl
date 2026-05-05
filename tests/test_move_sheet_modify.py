"""RFC-036 — ``Workbook.move_sheet`` round-trip in modify mode.

End-to-end coverage for ``wb.move_sheet(sheet, offset)`` on an existing
file. The save-time path threads three layers:

1. ``Workbook.move_sheet`` (Python) validates the sheet/offset arg
   types, locates the source index in ``self._sheet_names``, clamps
   the new position to ``[0, n-1]``, updates the in-memory tab list
   (so subsequent reads see the new order), and queues the move on
   the patcher via ``_flush_pending_sheet_moves_to_patcher``.
2. ``XlsxPatcher::queue_sheet_move`` (Rust) appends to
   ``queued_sheet_moves``.
3. ``XlsxPatcher::do_save`` Phase 2.5h (Rust) reads ``xl/workbook.xml``,
   applies each queued ``(name, offset)`` against the running tab
   list (clamped, composing), rewrites the ``<sheets>`` block in the
   new order, and re-points every ``<definedName localSheetId="N">``
   integer that maps to a moved position. The merged bytes are
   routed through ``file_patches``.

The empty-queue invariant lives at the Rust layer: a workbook in
modify mode that's saved without any ``move_sheet`` calls produces
byte-identical ``xl/workbook.xml``. ``test_rfc036_empty_queue_is_no_op``
guards this.
"""

from __future__ import annotations

import zipfile
from pathlib import Path

import openpyxl
import pytest

from wolfxl import Workbook, load_workbook

# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture(autouse=True)
def _force_test_epoch(monkeypatch: pytest.MonkeyPatch) -> None:
    """Pin ZIP entry mtimes for byte-stable saves."""
    monkeypatch.setenv("WOLFXL_TEST_EPOCH", "0")


def _make_four_sheet_fixture(path: Path) -> None:
    """Workbook with sheets [A, B, C, D] and three sheet-scoped + one
    workbook-scoped defined names. Tests the localSheetId remap path."""
    from openpyxl.workbook.defined_name import DefinedName as XDefinedName

    wb = openpyxl.Workbook()
    ws_a = wb.active
    ws_a.title = "A"
    ws_a["A1"] = 1
    wb.create_sheet("B")["A1"] = 2
    wb.create_sheet("C")["A1"] = 3
    wb.create_sheet("D")["A1"] = 4

    # Sheet-scoped: Print_Area on A (localSheetId=0), Region on C (=2),
    # SuffixD on D (=3). Workbook-scoped: Anywhere (no localSheetId).
    wb.defined_names["_xlnm.Print_Area"] = XDefinedName(
        "_xlnm.Print_Area", attr_text="A!$A$1:$D$10", localSheetId=0
    )
    wb.defined_names["Region"] = XDefinedName(
        "Region", attr_text="C!$A$1:$A$10", localSheetId=2
    )
    wb.defined_names["SuffixD"] = XDefinedName(
        "SuffixD", attr_text="D!$A$1", localSheetId=3
    )
    wb.defined_names["Anywhere"] = XDefinedName(
        "Anywhere", attr_text="A!$Z$99"
    )
    wb.save(path)


def _make_two_sheet_fixture(path: Path) -> None:
    """Plain two-sheet workbook with no defined names. Tests the
    no-defined-names branch."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "First"
    ws["A1"] = "first"
    ws2 = wb.create_sheet("Second")
    ws2["A1"] = "second"
    wb.save(path)


def _read_workbook_xml(path: Path) -> str:
    with zipfile.ZipFile(path) as zf:
        return zf.read("xl/workbook.xml").decode("utf-8")


# ---------------------------------------------------------------------------
# In-memory state — wb.sheetnames updates immediately, both modes.
# ---------------------------------------------------------------------------


def test_rfc036_in_memory_tab_list_updates_after_move() -> None:
    """move_sheet must update self._sheet_names so subsequent reads
    of wb.sheetnames see the new order — even in write mode where the
    patcher path is None."""
    wb = Workbook()
    wb.create_sheet("Second")
    wb.create_sheet("Third")
    assert wb.sheetnames == ["Sheet", "Second", "Third"]

    wb.move_sheet("Sheet", offset=2)
    assert wb.sheetnames == ["Second", "Third", "Sheet"]


def test_rfc036_write_mode_move_sheet_persists_on_save(tmp_path: Path) -> None:
    """Fresh ``Workbook()`` saves must use the post-move native-writer order."""
    out = tmp_path / "write_mode_move.xlsx"
    wb = Workbook()
    wb.active.title = "A"
    wb["A"]["A1"] = "alpha"
    wb.create_sheet("B")["A1"] = "bravo"
    wb.create_sheet("C")["A1"] = "charlie"

    wb.move_sheet("A", offset=2)
    assert wb.sheetnames == ["B", "C", "A"]
    wb.save(out)

    rt = openpyxl.load_workbook(out)
    assert rt.sheetnames == ["B", "C", "A"]
    assert rt["A"]["A1"].value == "alpha"
    assert rt["B"]["A1"].value == "bravo"
    assert rt["C"]["A1"].value == "charlie"


def test_modify_mode_worksheet_rename_persists_on_save(tmp_path: Path) -> None:
    """Renaming a loaded worksheet must update workbook.xml like openpyxl."""
    src = tmp_path / "rename_src.xlsx"
    _make_two_sheet_fixture(src)

    wb = load_workbook(src, modify=True)
    wb["First"].title = "Renamed"
    out = tmp_path / "renamed.xlsx"
    wb.save(out)
    wb.close()

    reloaded = openpyxl.load_workbook(out)
    assert reloaded.sheetnames == ["Renamed", "Second"]
    assert reloaded["Renamed"]["A1"].value == "first"


def test_modify_mode_rename_then_edit_uses_new_title(tmp_path: Path) -> None:
    """Queued mutations after rename should target the renamed worksheet."""
    src = tmp_path / "rename_edit_src.xlsx"
    _make_two_sheet_fixture(src)

    wb = load_workbook(src, modify=True)
    ws = wb["First"]
    ws.title = "Renamed"
    ws["B1"] = "after"
    out = tmp_path / "renamed_edited.xlsx"
    wb.save(out)
    wb.close()

    reloaded = openpyxl.load_workbook(out)
    assert reloaded.sheetnames == ["Renamed", "Second"]
    assert reloaded["Renamed"]["A1"].value == "first"
    assert reloaded["Renamed"]["B1"].value == "after"


def test_rfc036_move_sheet_accepts_worksheet_instance() -> None:
    wb = Workbook()
    wb.create_sheet("Second")
    ws = wb["Second"]
    wb.move_sheet(ws, offset=-1)
    assert wb.sheetnames == ["Second", "Sheet"]


def test_rfc036_move_sheet_offset_zero_is_in_memory_no_op() -> None:
    wb = Workbook()
    wb.create_sheet("Second")
    wb.create_sheet("Third")
    before = list(wb.sheetnames)
    wb.move_sheet("Second", offset=0)
    assert wb.sheetnames == before


# ---------------------------------------------------------------------------
# Validation — TypeError / KeyError fire at call time, not save time.
# ---------------------------------------------------------------------------


def test_rfc036_invalid_sheet_type_raises_type_error() -> None:
    wb = Workbook()
    with pytest.raises(TypeError, match="Worksheet or str"):
        wb.move_sheet(42, offset=1)  # type: ignore[arg-type]


def test_rfc036_unknown_sheet_name_raises_key_error() -> None:
    wb = Workbook()
    with pytest.raises(KeyError):
        wb.move_sheet("DOES_NOT_EXIST", offset=1)


def test_rfc036_non_int_offset_raises_type_error() -> None:
    wb = Workbook()
    wb.create_sheet("Other")
    with pytest.raises(TypeError, match="int"):
        wb.move_sheet("Sheet", offset=1.5)  # type: ignore[arg-type]


def test_rfc036_bool_offset_rejected() -> None:
    """``isinstance(True, int)`` is True in Python, but a user passing
    True almost certainly meant something else. Reject explicitly."""
    wb = Workbook()
    wb.create_sheet("Other")
    with pytest.raises(TypeError, match="int"):
        wb.move_sheet("Sheet", offset=True)  # type: ignore[arg-type]


# ---------------------------------------------------------------------------
# Modify-mode round-trip — positive offset.
# ---------------------------------------------------------------------------


def test_rfc036_positive_offset_round_trip(tmp_path: Path) -> None:
    """[A,B,C,D] → move_sheet('A', offset=2) → [B,C,A,D]."""
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_four_sheet_fixture(src)

    wb = load_workbook(src, modify=True)
    wb.move_sheet("A", offset=2)
    assert wb.sheetnames == ["B", "C", "A", "D"]
    wb.save(dst)

    rt = openpyxl.load_workbook(dst)
    assert rt.sheetnames == ["B", "C", "A", "D"]
    # Cells survive: A's A1=1, B's A1=2, etc.
    assert rt["A"]["A1"].value == 1
    assert rt["B"]["A1"].value == 2
    assert rt["C"]["A1"].value == 3
    assert rt["D"]["A1"].value == 4


def test_rfc036_positive_offset_remaps_local_sheet_ids(
    tmp_path: Path,
) -> None:
    """After [A,B,C,D] → [B,C,A,D], the localSheetId integers must
    re-point: Print_Area was 0 (A); A is now at 2 → Print_Area=2.
    Region was 2 (C); C is now at 1 → Region=1. SuffixD was 3 (D);
    D unchanged → SuffixD=3. Anywhere has no localSheetId; unchanged."""
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_four_sheet_fixture(src)

    wb = load_workbook(src, modify=True)
    wb.move_sheet("A", offset=2)
    wb.save(dst)

    xml = _read_workbook_xml(dst)
    assert (
        '<definedName name="_xlnm.Print_Area" localSheetId="2">' in xml
    ), f"Print_Area should now reference position 2 (A's new pos):\n{xml}"
    assert (
        '<definedName name="Region" localSheetId="1">' in xml
    ), f"Region should now reference position 1 (C's new pos):\n{xml}"
    assert (
        '<definedName name="SuffixD" localSheetId="3">' in xml
    ), f"SuffixD should remain at position 3 (D unchanged):\n{xml}"
    assert (
        '<definedName name="Anywhere">' in xml
    ), f"Anywhere has no localSheetId; preserve verbatim:\n{xml}"


# ---------------------------------------------------------------------------
# Modify-mode round-trip — negative offset.
# ---------------------------------------------------------------------------


def test_rfc036_negative_offset_round_trip(tmp_path: Path) -> None:
    """[A,B,C,D] → move_sheet('D', offset=-2) → [A,D,B,C].
    Remap: 1→2 (B), 2→3 (C), 3→1 (D). 0 unchanged.
    SuffixD was at 3, ends at 1. Region was at 2, ends at 3."""
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_four_sheet_fixture(src)

    wb = load_workbook(src, modify=True)
    wb.move_sheet("D", offset=-2)
    assert wb.sheetnames == ["A", "D", "B", "C"]
    wb.save(dst)

    rt = openpyxl.load_workbook(dst)
    assert rt.sheetnames == ["A", "D", "B", "C"]
    xml = _read_workbook_xml(dst)
    assert '<definedName name="SuffixD" localSheetId="1">' in xml
    assert '<definedName name="Region" localSheetId="3">' in xml
    # Print_Area's pos 0 (A) is unchanged.
    assert '<definedName name="_xlnm.Print_Area" localSheetId="0">' in xml


# ---------------------------------------------------------------------------
# Boundary clamping.
# ---------------------------------------------------------------------------


def test_rfc036_high_offset_clamps_to_last(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_two_sheet_fixture(src)

    wb = load_workbook(src, modify=True)
    wb.move_sheet("First", offset=100)
    wb.save(dst)

    rt = openpyxl.load_workbook(dst)
    assert rt.sheetnames == ["Second", "First"]


def test_rfc036_low_offset_clamps_to_first(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_two_sheet_fixture(src)

    wb = load_workbook(src, modify=True)
    wb.move_sheet("Second", offset=-100)
    wb.save(dst)

    rt = openpyxl.load_workbook(dst)
    assert rt.sheetnames == ["Second", "First"]


# ---------------------------------------------------------------------------
# Empty queue — no move_sheet call ⇒ workbook.xml unchanged.
# ---------------------------------------------------------------------------


def test_rfc036_empty_queue_is_no_op(tmp_path: Path) -> None:
    """A modify-mode workbook saved without any move_sheet call must
    produce byte-identical xl/workbook.xml. Sister contract for
    Phase 2.5h's empty-queue short-circuit."""
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_four_sheet_fixture(src)

    wb = load_workbook(src, modify=True)
    wb.save(dst)

    src_xml = _read_workbook_xml(src)
    dst_xml = _read_workbook_xml(dst)
    assert src_xml == dst_xml


# ---------------------------------------------------------------------------
# Compose: move + defined-names mutation in the same save.
# ---------------------------------------------------------------------------


def test_rfc036_compose_with_defined_names(tmp_path: Path) -> None:
    """Both Phase 2.5h (sheet move) and Phase 2.5f (defined names)
    mutate workbook.xml. They must compose: the defined-names merger
    runs against the post-move bytes (so localSheetId remap is in
    effect when a new defined name with localSheetId is added)."""
    from wolfxl.workbook.defined_name import DefinedName

    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_four_sheet_fixture(src)

    wb = load_workbook(src, modify=True)
    # Move first, then add a defined name. After the move:
    # [A,B,C,D] → move A by +2 → [B,C,A,D]. The new defined name we
    # add references A's NEW position (2) — that's the user's job.
    wb.move_sheet("A", offset=2)
    wb.defined_names["NewName"] = DefinedName(
        name="NewName", value="A!$Z$99", localSheetId=2
    )
    wb.save(dst)

    xml = _read_workbook_xml(dst)
    # Existing names re-pointed by Phase 2.5h.
    assert '<definedName name="_xlnm.Print_Area" localSheetId="2">' in xml
    assert '<definedName name="Region" localSheetId="1">' in xml
    # New name added by Phase 2.5f at the post-move position.
    assert '<definedName name="NewName" localSheetId="2">A!$Z$99</definedName>' in xml

    rt = openpyxl.load_workbook(dst)
    assert rt.sheetnames == ["B", "C", "A", "D"]
    # NewName has localSheetId=2 (post-move = sheet A).
    # openpyxl scopes sheet-scoped defined names under wb[sheet].defined_names.
    assert "NewName" in rt["A"].defined_names


# ---------------------------------------------------------------------------
# Cell data + macros / charts / pivots / comments survive.
# ---------------------------------------------------------------------------


def test_rfc036_cell_data_survives_move(tmp_path: Path) -> None:
    """The patcher only touches xl/workbook.xml — every sheet's
    own xl/worksheets/sheet*.xml content flows through untouched."""
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_four_sheet_fixture(src)

    wb = load_workbook(src, modify=True)
    wb.move_sheet("D", offset=-3)
    wb.save(dst)

    rt = openpyxl.load_workbook(dst)
    assert rt.sheetnames == ["D", "A", "B", "C"]
    assert rt["A"]["A1"].value == 1
    assert rt["B"]["A1"].value == 2
    assert rt["C"]["A1"].value == 3
    assert rt["D"]["A1"].value == 4


# ---------------------------------------------------------------------------
# Multi-move composition.
# ---------------------------------------------------------------------------


def test_rfc036_multiple_moves_compose(tmp_path: Path) -> None:
    """[A,B,C,D] → move A +2 → [B,C,A,D] → move B +2 → [C,A,B,D].

    The patcher applies queued moves in order, each against the
    running tab list — same model as openpyxl. The final remap is
    the composition of the two intermediate remaps."""
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_four_sheet_fixture(src)

    wb = load_workbook(src, modify=True)
    wb.move_sheet("A", offset=2)
    wb.move_sheet("B", offset=2)
    assert wb.sheetnames == ["C", "A", "B", "D"]
    wb.save(dst)

    rt = openpyxl.load_workbook(dst)
    assert rt.sheetnames == ["C", "A", "B", "D"]

    xml = _read_workbook_xml(dst)
    # Print_Area was at pos 0 (A); after the two moves, A is at 1.
    assert '<definedName name="_xlnm.Print_Area" localSheetId="1">' in xml
    # Region was at pos 2 (C); after the two moves, C is at 0.
    assert '<definedName name="Region" localSheetId="0">' in xml
    # SuffixD was at pos 3 (D); D unchanged → 3.
    assert '<definedName name="SuffixD" localSheetId="3">' in xml


# ---------------------------------------------------------------------------
# openpyxl interop: result file is fully readable.
# ---------------------------------------------------------------------------


def test_rfc036_result_loadable_by_openpyxl(tmp_path: Path) -> None:
    """A wolfxl-modified file must round-trip through openpyxl
    without errors."""
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_four_sheet_fixture(src)

    wb = load_workbook(src, modify=True)
    wb.move_sheet("C", offset=-1)
    wb.save(dst)

    rt = openpyxl.load_workbook(dst)
    assert rt.sheetnames == ["A", "C", "B", "D"]
    # Defined names also survive openpyxl's parse. Sheet-scoped non-builtin
    # names land on wb[sheet].defined_names; ``_xlnm.Print_Area`` is
    # special-cased into ``ws.print_area`` by openpyxl; workbook-scoped
    # ones land on wb.defined_names. After the C move:
    #   Print_Area (was @ A=0) → still @ A   → A.print_area is set
    #   Region     (was @ C=2) → still @ C   → C.defined_names
    #   SuffixD    (was @ D=3) → still @ D   → D.defined_names
    #   Anywhere   (workbook scope)         → wb.defined_names
    assert rt["A"].print_area  # truthy
    assert "Region" in rt["C"].defined_names
    assert "SuffixD" in rt["D"].defined_names
    assert "Anywhere" in rt.defined_names


# ---------------------------------------------------------------------------
# Single-sheet workbook is safe.
# ---------------------------------------------------------------------------


def test_rfc036_single_sheet_workbook_is_safe(tmp_path: Path) -> None:
    """One-sheet workbooks have no meaningful move; both Python and
    Rust layers clamp / no-op without raising."""
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"

    wb_o = openpyxl.Workbook()
    wb_o.active.title = "Only"
    wb_o.save(src)

    wb = load_workbook(src, modify=True)
    wb.move_sheet("Only", offset=5)
    assert wb.sheetnames == ["Only"]
    wb.save(dst)

    rt = openpyxl.load_workbook(dst)
    assert rt.sheetnames == ["Only"]
