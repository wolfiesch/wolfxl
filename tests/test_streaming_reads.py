"""Sprint Ι Pod-β — SAX streaming read tests.

Pin the contract added in Sprint Ι Pod-β: ``load_workbook(read_only=True)``
exposes ``Worksheet.iter_rows`` as a true streaming generator backed by
``wolfxl._rust.StreamingSheetReader``. Cells yielded in that path are
``StreamingCell`` proxies that surface the same value/style attributes as
the eager ``Cell`` API but reject mutation. The auto-trigger heuristic is
also covered: workbooks loaded without ``read_only=True`` but with
> ``AUTO_STREAM_ROW_THRESHOLD`` rows still flow through the streaming
path, transparently to the caller.
"""

from __future__ import annotations

import datetime as dt
import os
from pathlib import Path

import openpyxl
import pytest

import wolfxl
from wolfxl._streaming import (
    AUTO_STREAM_ROW_THRESHOLD,
    StreamingCell,
    should_auto_stream,
)


# ---------------------------------------------------------------------------
# Fixture builders — generated on the fly via openpyxl so the fixture content
# stays in lockstep with the parity reference implementation.
# ---------------------------------------------------------------------------


def _build_basic(path: Path, n_rows: int = 1000, n_cols: int = 5) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for r in range(1, n_rows + 1):
        for c in range(1, n_cols + 1):
            ws.cell(row=r, column=c, value=r * 10 + c)
    wb.save(path)
    return path


def _build_styled(path: Path) -> Path:
    """Small fixture exercising bold/red/center alignment + named formats."""
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Styled"
    ws.cell(row=1, column=1, value="bold")
    ws.cell(row=1, column=1).font = Font(bold=True)
    ws.cell(row=1, column=2, value="red-fill")
    ws.cell(row=1, column=2).fill = PatternFill(
        fill_type="solid", fgColor="FFFF0000"
    )
    ws.cell(row=2, column=1, value=3.14)
    ws.cell(row=2, column=1).number_format = "0.00"
    ws.cell(row=2, column=2, value="centered")
    ws.cell(row=2, column=2).alignment = Alignment(horizontal="center")
    ws.cell(row=3, column=1, value="plain")
    wb.save(path)
    return path


def _build_mixed_types(path: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mixed"
    ws.cell(row=1, column=1, value=42)
    ws.cell(row=1, column=2, value=3.14)
    ws.cell(row=1, column=3, value="hello")
    ws.cell(row=1, column=4, value=True)
    ws.cell(row=1, column=5, value=False)
    ws.cell(row=2, column=1, value=dt.datetime(2024, 1, 15, 12, 30))
    ws.cell(row=2, column=2, value=dt.date(2024, 6, 1))
    ws.cell(row=2, column=3, value="=A1*2")  # formula
    # Row 3 is intentionally absent → sparse row test.
    ws.cell(row=4, column=2, value="middle")  # only B4 populated
    wb.save(path)
    return path


def _build_sparse(path: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sparse"
    ws.cell(row=1, column=1, value="r1")
    ws.cell(row=5, column=3, value="r5c3")
    ws.cell(row=10, column=1, value="r10")
    wb.save(path)
    return path


def _build_synthetic_60k(path: Path) -> Path:
    """Build a sheet with > AUTO_STREAM_ROW_THRESHOLD rows to exercise the
    auto-trigger. We use openpyxl's write_only mode for speed.
    """
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet("Big")
    for r in range(1, AUTO_STREAM_ROW_THRESHOLD + 5_001):
        ws.append([r, r * 2, f"row{r}"])
    wb.save(path)
    return path


@pytest.fixture
def basic_xlsx(tmp_path: Path) -> Path:
    return _build_basic(tmp_path / "basic.xlsx")


@pytest.fixture
def styled_xlsx(tmp_path: Path) -> Path:
    return _build_styled(tmp_path / "styled.xlsx")


@pytest.fixture
def mixed_xlsx(tmp_path: Path) -> Path:
    return _build_mixed_types(tmp_path / "mixed.xlsx")


@pytest.fixture
def sparse_xlsx(tmp_path: Path) -> Path:
    return _build_sparse(tmp_path / "sparse.xlsx")


# ---------------------------------------------------------------------------
# 1. Basic — 1000 rows in order via values_only.
# ---------------------------------------------------------------------------


def test_streaming_values_only_yields_1000_rows(basic_xlsx: Path) -> None:
    wb = wolfxl.load_workbook(basic_xlsx, read_only=True)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) == 1000
    assert rows[0] == (11, 12, 13, 14, 15)
    assert rows[-1] == (10001, 10002, 10003, 10004, 10005)


# ---------------------------------------------------------------------------
# 2. Style access — read-mode StreamingCell exposes font.bold etc.
# ---------------------------------------------------------------------------


def test_streaming_cells_expose_styles(styled_xlsx: Path) -> None:
    wb = wolfxl.load_workbook(styled_xlsx, read_only=True)
    ws = wb["Styled"]
    rows = list(ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=2))
    # All cells are StreamingCell instances.
    for r in rows:
        for c in r:
            assert isinstance(c, StreamingCell)
    # A1 = "bold" with bold font.
    assert rows[0][0].value == "bold"
    assert rows[0][0].font.bold is True
    # A2 = 3.14, number_format "0.00".
    assert rows[1][0].value == 3.14
    assert rows[1][0].number_format == "0.00"
    # B2 alignment horizontal=center.
    assert rows[1][1].alignment.horizontal == "center"


# ---------------------------------------------------------------------------
# 3. Bounded range — min_row=10, max_row=20 yields 11 rows.
# ---------------------------------------------------------------------------


def test_streaming_bounded_range(basic_xlsx: Path) -> None:
    wb = wolfxl.load_workbook(basic_xlsx, read_only=True)
    ws = wb["Sheet1"]
    rows = list(
        ws.iter_rows(min_row=10, max_row=20, min_col=1, max_col=3, values_only=True)
    )
    assert len(rows) == 11
    assert rows[0] == (101, 102, 103)
    assert rows[-1] == (201, 202, 203)


# ---------------------------------------------------------------------------
# 4. Mixed types — number, str, bool, date, formula all surface correctly.
# ---------------------------------------------------------------------------


def test_streaming_mixed_types(mixed_xlsx: Path) -> None:
    wb = wolfxl.load_workbook(mixed_xlsx, read_only=True)
    ws = wb["Mixed"]
    rows = list(
        ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=5, values_only=True)
    )
    # Row 1: int, float, str, True, False.
    r1 = rows[0]
    assert r1[0] == 42
    assert r1[1] == pytest.approx(3.14)
    assert r1[2] == "hello"
    assert r1[3] is True
    assert r1[4] is False
    # Row 2 col 3 — a formula cached as a string.
    r2 = rows[1]
    val_a = r2[2]
    # Streaming surfaces the formula text from `Cell.value` to match
    # openpyxl read_only=True semantics.
    assert val_a in ("=A1*2", "84")  # cached value may be either


# ---------------------------------------------------------------------------
# 5. Empty rows / sparse cells — yields tuples of Nones for missing cells.
# ---------------------------------------------------------------------------


def test_streaming_sparse_rows(sparse_xlsx: Path) -> None:
    wb = wolfxl.load_workbook(sparse_xlsx, read_only=True)
    ws = wb["Sparse"]
    rows = list(
        ws.iter_rows(min_row=1, max_row=10, min_col=1, max_col=3, values_only=True)
    )
    # Explicit bounds match openpyxl: missing rows are padded with empty tuples.
    assert len(rows) == 10
    # Sparse rows should appear with the right shape:
    found_first_cells = [r[0] for r in rows]
    assert "r1" in found_first_cells
    assert "r10" in found_first_cells


# ---------------------------------------------------------------------------
# 6. SST reference — a `t="s"` cell resolves to its shared-strings entry.
# ---------------------------------------------------------------------------


def test_streaming_sst_resolves(tmp_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    # Repeat strings → shared in SST.
    for i in range(10):
        ws.cell(row=i + 1, column=1, value="repeated")
        ws.cell(row=i + 1, column=2, value=f"unique-{i}")
    path = tmp_path / "sst.xlsx"
    wb.save(path)

    wb2 = wolfxl.load_workbook(path, read_only=True)
    ws2 = wb2["Sheet"]
    rows = list(
        ws2.iter_rows(min_row=1, max_row=10, min_col=1, max_col=2, values_only=True)
    )
    assert len(rows) == 10
    for i, r in enumerate(rows):
        assert r == ("repeated", f"unique-{i}")


# ---------------------------------------------------------------------------
# 7. Style index — `<c s="N">` surfaces non-zero style_id when set.
# ---------------------------------------------------------------------------


def test_streaming_style_id_is_set(styled_xlsx: Path) -> None:
    wb = wolfxl.load_workbook(styled_xlsx, read_only=True)
    ws = wb["Styled"]
    rows = list(ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=1))
    cell = rows[0][0]
    # Style id is internal — we expose it via the public font property,
    # which must have bold=True since the underlying `s=` index points
    # at a styled xf.
    assert cell.font.bold is True


# ---------------------------------------------------------------------------
# 8. Mutation rejection — every setter raises a clear RuntimeError.
# ---------------------------------------------------------------------------


def test_streaming_cell_value_setter_raises(basic_xlsx: Path) -> None:
    wb = wolfxl.load_workbook(basic_xlsx, read_only=True)
    ws = wb["Sheet1"]
    cell = next(iter(ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=1)))[0]
    with pytest.raises(RuntimeError, match="read_only=True"):
        cell.value = "X"


def test_streaming_cell_font_setter_raises(basic_xlsx: Path) -> None:
    from wolfxl import Font

    wb = wolfxl.load_workbook(basic_xlsx, read_only=True)
    ws = wb["Sheet1"]
    cell = next(iter(ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=1)))[0]
    with pytest.raises(RuntimeError, match="read_only=True"):
        cell.font = Font(bold=True)


def test_streaming_cell_typo_attr_raises(basic_xlsx: Path) -> None:
    wb = wolfxl.load_workbook(basic_xlsx, read_only=True)
    ws = wb["Sheet1"]
    cell = next(iter(ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=1)))[0]
    with pytest.raises(RuntimeError, match="read_only=True"):
        cell.bogus_attr = 123


# ---------------------------------------------------------------------------
# 9. Mode independence — read_only=False still works (no regression).
# ---------------------------------------------------------------------------


def test_eager_mode_unchanged(basic_xlsx: Path) -> None:
    wb = wolfxl.load_workbook(basic_xlsx)
    ws = wb["Sheet1"]
    # Direct cell access still returns a regular Cell, not StreamingCell.
    cell = ws["A1"]
    assert not isinstance(cell, StreamingCell)
    assert cell.value == 11
    # iter_rows in eager mode also returns regular Cell-tuple rows.
    rows = list(ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=3))
    assert all(not isinstance(c, StreamingCell) for c in rows[0])


# ---------------------------------------------------------------------------
# 10. Auto-trigger — a > 50k-row workbook streams even without read_only=True.
# ---------------------------------------------------------------------------


@pytest.mark.slow
def test_auto_trigger_above_threshold(tmp_path: Path) -> None:
    # Skip locally if generation is too slow; pinned slow.
    path = _build_synthetic_60k(tmp_path / "huge.xlsx")
    wb = wolfxl.load_workbook(path)  # read_only=False
    ws = wb.active
    assert should_auto_stream(ws)
    # Reading via iter_rows doesn't OOM and yields ALL rows.
    n = sum(1 for _ in ws.iter_rows(values_only=True))
    assert n > AUTO_STREAM_ROW_THRESHOLD


# ---------------------------------------------------------------------------
# 11. close() / generator cleanup — file handle is released after iteration.
# ---------------------------------------------------------------------------


def test_streaming_iterator_cleanup(basic_xlsx: Path) -> None:
    wb = wolfxl.load_workbook(basic_xlsx, read_only=True)
    ws = wb["Sheet1"]
    gen = ws.iter_rows(values_only=True)
    # Pull a single row, then drop the generator.
    next(gen)
    del gen
    # File is no longer locked; we can rename / delete it.
    new_path = basic_xlsx.parent / "renamed.xlsx"
    os.rename(basic_xlsx, new_path)
    assert new_path.exists()


# ---------------------------------------------------------------------------
# 12. Auto-trigger — should_auto_stream returns False for small sheets.
# ---------------------------------------------------------------------------


def test_no_auto_trigger_below_threshold(basic_xlsx: Path) -> None:
    wb = wolfxl.load_workbook(basic_xlsx)  # read_only=False
    ws = wb["Sheet1"]
    assert should_auto_stream(ws) is False


# ---------------------------------------------------------------------------
# 13. read_only=True surfaces the explicit flag on the workbook.
# ---------------------------------------------------------------------------


def test_workbook_read_only_property(basic_xlsx: Path) -> None:
    wb_r = wolfxl.load_workbook(basic_xlsx, read_only=True)
    assert wb_r.read_only is True
    wb_e = wolfxl.load_workbook(basic_xlsx)
    assert wb_e.read_only is False


# ---------------------------------------------------------------------------
# 14. min_col / max_col bounds are honored.
# ---------------------------------------------------------------------------


def test_streaming_column_bounds(basic_xlsx: Path) -> None:
    wb = wolfxl.load_workbook(basic_xlsx, read_only=True)
    ws = wb["Sheet1"]
    rows = list(
        ws.iter_rows(min_row=1, max_row=2, min_col=2, max_col=4, values_only=True)
    )
    assert rows == [(12, 13, 14), (22, 23, 24)]


# ---------------------------------------------------------------------------
# 15. Coordinate / row / column on streaming cells match openpyxl.
# ---------------------------------------------------------------------------


def test_streaming_cell_coords(basic_xlsx: Path) -> None:
    wb = wolfxl.load_workbook(basic_xlsx, read_only=True)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(min_row=2, max_row=2, min_col=3, max_col=3))
    cell = rows[0][0]
    assert cell.row == 2
    assert cell.column == 3
    assert cell.column_letter == "C"
    assert cell.coordinate == "C2"


# ---------------------------------------------------------------------------
# 16. Iteration yields exactly one tuple per `<row>` element present.
# ---------------------------------------------------------------------------


def test_streaming_yields_per_row_element(basic_xlsx: Path) -> None:
    wb = wolfxl.load_workbook(basic_xlsx, read_only=True)
    ws = wb["Sheet1"]
    count = 0
    for _ in ws.iter_rows(values_only=True):
        count += 1
    assert count == 1000


# ---------------------------------------------------------------------------
# 17. Sprint Λ Pod-γ — datetime cells convert via the styles table.
# A `<c s="N">` whose number format passes ``is_date_format`` must surface
# as a ``datetime`` (or ``date``/``time``) — not as the raw Excel serial.
# ---------------------------------------------------------------------------


def test_streaming_datetime_yields_datetime_values_only(mixed_xlsx: Path) -> None:
    wb = wolfxl.load_workbook(mixed_xlsx, read_only=True)
    ws = wb["Mixed"]
    rows = list(
        ws.iter_rows(min_row=2, max_row=2, min_col=1, max_col=2, values_only=True)
    )
    a2, b2 = rows[0]
    assert isinstance(a2, dt.datetime), f"A2 datetime divergence: {type(a2).__name__} {a2!r}"
    assert a2 == dt.datetime(2024, 1, 15, 12, 30)
    # B2 is a date — openpyxl read_only path returns a datetime at midnight.
    assert isinstance(b2, dt.datetime), f"B2 date divergence: {type(b2).__name__} {b2!r}"
    assert b2 == dt.datetime(2024, 6, 1, 0, 0)


def test_streaming_datetime_yields_datetime_via_streaming_cell(mixed_xlsx: Path) -> None:
    wb = wolfxl.load_workbook(mixed_xlsx, read_only=True)
    ws = wb["Mixed"]
    rows = list(ws.iter_rows(min_row=2, max_row=2, min_col=1, max_col=2))
    a2_cell, b2_cell = rows[0]
    assert isinstance(a2_cell.value, dt.datetime), (
        f"StreamingCell.value (A2) divergence: {type(a2_cell.value).__name__} {a2_cell.value!r}"
    )
    assert a2_cell.value == dt.datetime(2024, 1, 15, 12, 30)
    assert isinstance(b2_cell.value, dt.datetime), (
        f"StreamingCell.value (B2) divergence: {type(b2_cell.value).__name__} {b2_cell.value!r}"
    )
    assert b2_cell.value == dt.datetime(2024, 6, 1, 0, 0)
