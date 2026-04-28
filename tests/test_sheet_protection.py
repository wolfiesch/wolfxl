"""RFC-055 §2.6 — SheetProtection tests (Sprint Ο Pod 1A).

Covers password hash round-trip parity with openpyxl for ``"hunter2"``
(the canonical example from the RFC).
"""

from __future__ import annotations


from wolfxl import Workbook
from wolfxl.utils.protection import check_password, hash_password
from wolfxl.worksheet.protection import SheetProtection


class TestHashPassword:
    def test_hunter2_matches_openpyxl(self):
        # Pinned by RFC-055 §7. openpyxl emits 'C258' for the input
        # "hunter2"; this test guards against algorithm drift.
        from openpyxl.utils.protection import hash_password as openpyxl_hash
        assert hash_password("hunter2") == openpyxl_hash("hunter2")
        assert hash_password("hunter2") == "C258"

    def test_empty_returns_ce4b(self):
        # openpyxl-parity: empty input runs through the algorithm and
        # XORs against the magic 0xCE4B at the end, yielding "CE4B".
        assert hash_password("") == "CE4B"

    def test_distinct_passwords_distinct_hashes(self):
        assert hash_password("a") != hash_password("b")

    def test_case_sensitive(self):
        assert hash_password("Password") != hash_password("password")

    def test_check_password_round_trip(self):
        h = hash_password("hunter2")
        assert check_password("hunter2", h)
        assert not check_password("wrong", h)
        assert not check_password("", h)


class TestSheetProtectionDefaults:
    def test_defaults(self):
        sp = SheetProtection()
        assert sp.sheet is False
        assert sp.password is None
        # "Allow these actions" toggles default to True (Excel UX).
        assert sp.formatCells is True
        assert sp.formatColumns is True
        assert sp.sort is True
        assert sp.autoFilter is True
        assert sp.pivotTables is True
        # Selection toggles default to allowed too:
        assert sp.selectLockedCells is False
        assert sp.selectUnlockedCells is False

    def test_is_default(self):
        assert SheetProtection().is_default()
        sp = SheetProtection()
        sp.sheet = True
        assert not sp.is_default()


class TestSheetProtectionPassword:
    def test_set_password_hashes(self):
        sp = SheetProtection()
        sp.set_password("hunter2")
        assert sp.password == "C258"

    def test_set_password_empty_clears(self):
        sp = SheetProtection()
        sp.set_password("hunter2")
        sp.set_password("")
        assert sp.password is None

    def test_check_password_returns_true_on_match(self):
        sp = SheetProtection()
        sp.set_password("hunter2")
        assert sp.check_password("hunter2")

    def test_check_password_returns_false_on_mismatch(self):
        sp = SheetProtection()
        sp.set_password("hunter2")
        assert not sp.check_password("wrong")

    def test_check_password_with_no_password_set(self):
        sp = SheetProtection()
        assert not sp.check_password("anything")


class TestSheetProtectionAliases:
    def test_format_cells_alias(self):
        sp = SheetProtection()
        sp.format_cells = False
        assert sp.formatCells is False

    def test_auto_filter_alias(self):
        sp = SheetProtection()
        sp.auto_filter = False
        assert sp.autoFilter is False

    def test_pivot_tables_alias(self):
        sp = SheetProtection()
        sp.pivot_tables = False
        assert sp.pivotTables is False


class TestSheetProtectionEnable:
    def test_enable_turns_sheet_on(self):
        sp = SheetProtection()
        sp.enable()
        assert sp.sheet is True

    def test_disable_clears_password(self):
        sp = SheetProtection()
        sp.set_password("hunter2")
        sp.enable()
        sp.disable()
        assert sp.sheet is False
        assert sp.password is None


class TestSheetProtectionRustDict:
    def test_default_dict(self):
        d = SheetProtection().to_rust_dict()
        assert d["sheet"] is False
        assert d["password_hash"] is None

    def test_dict_with_password(self):
        sp = SheetProtection()
        sp.set_password("hunter2")
        sp.enable()
        d = sp.to_rust_dict()
        assert d["sheet"] is True
        assert d["password_hash"] == "C258"


class TestWorksheetProtection:
    def test_lazy_access(self):
        wb = Workbook()
        ws = wb.active
        assert isinstance(ws.protection, SheetProtection)
        assert ws.protection.is_default()

    def test_set_password_on_worksheet(self):
        wb = Workbook()
        ws = wb.active
        ws.protection.set_password("hunter2")
        ws.protection.enable()
        assert ws.protection.sheet is True
        assert ws.protection.password == "C258"

    def test_assignment_replaces(self):
        wb = Workbook()
        ws = wb.active
        new_sp = SheetProtection(sheet=True)
        ws.protection = new_sp
        assert ws.protection is new_sp
