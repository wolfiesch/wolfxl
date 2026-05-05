from __future__ import annotations

from zipfile import ZipFile

import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.worksheet.table import Table, TableStyleInfo

import wolfxl


def test_modify_save_to_source_path_is_atomic_and_valid(tmp_path):
    path = tmp_path / "source.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["Region", "Sales"])
    sheet.append(["West", 120])
    sheet.append(["East", 95])
    table = Table(displayName="SalesTable", ref="A1:B3")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    sheet.add_table(table)
    chart = BarChart()
    chart.add_data(Reference(sheet, min_col=2, min_row=1, max_row=3), titles_from_data=True)
    chart.set_categories(Reference(sheet, min_col=1, min_row=2, max_row=3))
    sheet.add_chart(chart, "D2")
    workbook.save(path)

    before_parts = set(ZipFile(path).namelist())
    assert "xl/tables/table1.xml" in before_parts
    assert "xl/charts/chart1.xml" in before_parts

    wolf_workbook = wolfxl.load_workbook(path, modify=True)
    wolf_workbook["Data"]["C1"] = "wolfxl_modify_smoke"
    wolf_workbook.save(path)
    wolf_workbook.close()

    after_parts = set(ZipFile(path).namelist())
    assert "xl/tables/table1.xml" in after_parts
    assert "xl/charts/chart1.xml" in after_parts

    roundtrip = openpyxl.load_workbook(path)
    assert roundtrip["Data"]["C1"].value == "wolfxl_modify_smoke"
    roundtrip.close()

