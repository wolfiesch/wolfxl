"""RFC-069 / G15 — combination-chart coverage.

Targets the four scenarios called out in RFC-069 §7.2:

1. Bar + Line on shared axes (no secondary value axis).
2. Bar + Line with the line on a secondary value axis (the
   compat-oracle ``charts_combination`` probe).
3. Bar + Line + Area three-family combo. Documented partial-coverage
   note where wolfxl's emit path falls short of full §6 compliance.
4. Validation: refuse two families with identical kind + identical
   axIds (likely copy-paste bug — fail loudly).

Read-side assertions exercise the dual-emission strategy: each chart
family is *also* emitted as its own standalone chartspace so
``openpyxl.load_workbook`` exposes every family on
``ws._charts`` (closes the gap noted in RFC-069 §2). Write-side
assertions sanity-check the multi-family ``<plotArea>`` shape per
RFC-069 §6 by inspecting ``xl/charts/chart1.xml`` directly.
"""

from __future__ import annotations

import zipfile
from pathlib import Path

import pytest

import wolfxl
from wolfxl.chart import AreaChart, BarChart, LineChart, Reference


def _seed_sheet(ws) -> None:
    """Populate three numeric series so Bar/Line/Area pick distinct columns."""
    rows = [
        ["x", "y", "z", "w"],
        [1, 10, 100, 1000],
        [2, 20, 200, 2000],
        [3, 30, 300, 3000],
    ]
    for row in rows:
        ws.append(row)


def _read_primary_chart_xml(out: Path) -> str:
    """Return the bytes of ``xl/charts/chart1.xml`` as decoded utf-8."""
    with zipfile.ZipFile(out) as z:
        return z.read("xl/charts/chart1.xml").decode("utf-8")


def test_bar_plus_line_shared_axes_no_secondary(tmp_path: Path) -> None:
    """Bar + Line sharing both x and y axes (RFC-069 §7.2 case 1).

    The line chart inherits the primary's value-axis ``ax_id`` so no
    secondary ``<c:valAx>`` is emitted. openpyxl's reload still sees
    both families because each is also written as a standalone
    chartspace at the same anchor (dual-emission strategy).
    """
    import openpyxl as _opx

    wb = wolfxl.Workbook()
    ws = wb.active
    _seed_sheet(ws)

    bar = BarChart()
    line = LineChart()
    bar.add_data(Reference(ws, min_col=2, min_row=1, max_row=4), titles_from_data=True)
    line.add_data(Reference(ws, min_col=3, min_row=1, max_row=4), titles_from_data=True)
    # Default y_axis.axId == 100 on both — secondary shares the primary.
    bar += line
    ws.add_chart(bar, "E2")
    out = tmp_path / "shared.xlsx"
    wb.save(out)

    chart_xml = _read_primary_chart_xml(out)
    # Multi-family plotArea: both `<c:barChart>` and `<c:lineChart>` exist.
    assert "<c:barChart>" in chart_xml
    assert "<c:lineChart>" in chart_xml
    # Secondary value axis is NOT emitted when ax_ids match — exactly
    # one `<c:valAx>` element.
    assert chart_xml.count("<c:valAx>") == 1, (
        f"shared y-axis should yield exactly one valAx, got "
        f"{chart_xml.count('<c:valAx>')}\n{chart_xml}"
    )

    # openpyxl read: both families are visible as distinct ws._charts
    # entries thanks to the dual-emission fallback.
    ref_wb = _opx.load_workbook(out)
    ref_ws = ref_wb[ref_wb.sheetnames[0]]
    types = {type(c).__name__ for c in ref_ws._charts}
    assert "BarChart" in types and "LineChart" in types, (
        f"openpyxl saw {types}, expected both BarChart and LineChart"
    )


def test_bar_plus_line_with_secondary_value_axis(tmp_path: Path) -> None:
    """The probe scenario: Bar + Line with line on a secondary axis
    (RFC-069 §7.2 case 2 / compat-oracle ``charts_combination``).
    """
    import openpyxl as _opx

    wb = wolfxl.Workbook()
    ws = wb.active
    _seed_sheet(ws)

    bar = BarChart()
    line = LineChart()
    bar.add_data(Reference(ws, min_col=2, min_row=1, max_row=4), titles_from_data=True)
    line.add_data(Reference(ws, min_col=3, min_row=1, max_row=4), titles_from_data=True)
    # Mark the line's value axis as secondary.
    line.y_axis.crosses = "max"
    line.y_axis.axId = 200
    bar += line
    ws.add_chart(bar, "E2")
    out = tmp_path / "secondary.xlsx"
    wb.save(out)

    chart_xml = _read_primary_chart_xml(out)
    # Multi-family plotArea.
    assert "<c:barChart>" in chart_xml
    assert "<c:lineChart>" in chart_xml
    # Two value axes total: primary + secondary.
    assert chart_xml.count("<c:valAx>") == 2, (
        f"secondary axis should yield two valAx siblings, got "
        f"{chart_xml.count('<c:valAx>')}"
    )
    # Secondary axis renders on the right with crosses=max (RFC-069 §6).
    assert '<c:axId val="200"/>' in chart_xml
    assert '<c:axPos val="r"/>' in chart_xml
    assert '<c:crosses val="max"/>' in chart_xml

    # openpyxl reload sees both families.
    ref_wb = _opx.load_workbook(out)
    ref_ws = ref_wb[ref_wb.sheetnames[0]]
    types = {type(c).__name__ for c in ref_ws._charts}
    assert "BarChart" in types and "LineChart" in types


def test_bar_plus_line_plus_area_three_family(tmp_path: Path) -> None:
    """Three-family combo (RFC-069 §7.2 case 3).

    Bar + Line + Area in one ``__iadd__`` chain. Verifies the multi-family
    plotArea handles N>2 secondaries and that openpyxl's reload sees all
    three types via dual-emission.

    *Partial-coverage note*: wolfxl's combo emit path supports an
    arbitrary number of axis-bearing 2D secondaries. 3D combos and Pie
    secondaries remain out of scope per RFC-069 §8.
    """
    import openpyxl as _opx

    wb = wolfxl.Workbook()
    ws = wb.active
    _seed_sheet(ws)

    bar = BarChart()
    line = LineChart()
    area = AreaChart()
    bar.add_data(Reference(ws, min_col=2, min_row=1, max_row=4), titles_from_data=True)
    line.add_data(Reference(ws, min_col=3, min_row=1, max_row=4), titles_from_data=True)
    area.add_data(Reference(ws, min_col=4, min_row=1, max_row=4), titles_from_data=True)
    bar += line
    bar += area
    ws.add_chart(bar, "E2")
    out = tmp_path / "three.xlsx"
    wb.save(out)

    chart_xml = _read_primary_chart_xml(out)
    assert "<c:barChart>" in chart_xml
    assert "<c:lineChart>" in chart_xml
    assert "<c:areaChart>" in chart_xml

    ref_wb = _opx.load_workbook(out)
    ref_ws = ref_wb[ref_wb.sheetnames[0]]
    types = {type(c).__name__ for c in ref_ws._charts}
    assert {"BarChart", "LineChart", "AreaChart"} <= types, (
        f"openpyxl saw {types}, expected all three families"
    )


def test_secondary_with_same_kind_and_axids_raises(tmp_path: Path) -> None:
    """Validation: refuse a copy-paste-shaped secondary (RFC-069 §5.3 / §7.2 case 4)."""
    wb = wolfxl.Workbook()
    ws = wb.active
    _seed_sheet(ws)

    primary = BarChart()
    duplicate = BarChart()
    primary.add_data(
        Reference(ws, min_col=2, min_row=1, max_row=4), titles_from_data=True
    )
    duplicate.add_data(
        Reference(ws, min_col=3, min_row=1, max_row=4), titles_from_data=True
    )
    # Both default to x=10, y=100. Same kind. This is the copy-paste bug
    # the validator must catch; surfacing it here keeps the user from
    # silently emitting a degenerate plotArea.
    primary += duplicate
    ws.add_chart(primary, "E2")

    with pytest.raises(ValueError, match="copy-paste"):
        wb.save(tmp_path / "dup.xlsx")


def test_pie_secondary_rejected(tmp_path: Path) -> None:
    """Pie/Doughnut combos are out of scope (RFC-069 §2 / §10)."""
    from wolfxl.chart import PieChart

    wb = wolfxl.Workbook()
    ws = wb.active
    _seed_sheet(ws)

    bar = BarChart()
    pie = PieChart()
    bar.add_data(Reference(ws, min_col=2, min_row=1, max_row=4), titles_from_data=True)
    pie.add_data(Reference(ws, min_col=3, min_row=1, max_row=4), titles_from_data=True)
    bar += pie
    ws.add_chart(bar, "E2")

    with pytest.raises(ValueError, match="Pie/Doughnut"):
        wb.save(tmp_path / "pie_combo.xlsx")


def test_secondary_with_empty_series_rejected(tmp_path: Path) -> None:
    """Validation: secondary chart with no series fails fast (RFC-069 §4.3)."""
    wb = wolfxl.Workbook()
    ws = wb.active
    _seed_sheet(ws)

    bar = BarChart()
    line = LineChart()  # No add_data — empty series.
    bar.add_data(Reference(ws, min_col=2, min_row=1, max_row=4), titles_from_data=True)
    bar += line
    ws.add_chart(bar, "E2")

    with pytest.raises(ValueError, match="at least one series"):
        wb.save(tmp_path / "empty_secondary.xlsx")
