"""RFC-023 — Cell comments + VML drawings round-trip in modify mode.

End-to-end coverage for ``cell.comment = Comment(...)`` on an existing
file.  The save-time path threads three layers:

1. ``Cell.comment`` setter (Python) drops the value into
   ``ws._pending_comments[coord]``.  ``None`` is the explicit-delete
   sentinel.
2. ``Workbook._flush_pending_comments_to_patcher`` (Python) drains
   each sheet's pending dict into ``XlsxPatcher.queue_comment`` /
   ``queue_comment_delete``.
3. ``XlsxPatcher::do_save`` Phase 2.5g (Rust) calls
   ``comments::build_comments`` against the source ``commentsN.xml`` +
   ``vmlDrawingN.vml`` (if any), merges with queued ops, mutates the
   sheet rels graph, pushes a ``SheetBlock::LegacyDrawing`` (slot 31)
   to the merger, and writes/adds/deletes the comments+vml ZIP parts
   plus the ``[Content_Types].xml`` Override / Default entries.
"""
from __future__ import annotations

import zipfile
from pathlib import Path

import openpyxl
import pytest

from wolfxl import load_workbook
from wolfxl.comments import Comment


@pytest.fixture(autouse=True)
def _force_test_epoch(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("WOLFXL_TEST_EPOCH", "0")


def _make_clean_fixture(path: Path) -> None:
    """Workbook with no comments — for testing the add-from-empty path."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "header"
    wb.save(path)


def _make_one_comment_fixture(path: Path) -> None:
    """Workbook with one existing comment via openpyxl."""
    from openpyxl.comments import Comment as OpyComment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "x"
    ws["A1"].comment = OpyComment("hello world", "alice")
    wb.save(path)


def test_add_comment_to_clean_file(tmp_path: Path) -> None:
    """Smoke: open a file with no comments, add one, openpyxl reads it back."""
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_clean_fixture(src)

    wb = load_workbook(src, modify=True)
    ws = wb["Sheet1"]
    ws["B2"].comment = Comment(text="from wolfxl", author="bob")
    wb.save(dst)

    # Verify with openpyxl.
    wb2 = openpyxl.load_workbook(dst)
    ws2 = wb2["Sheet1"]
    assert ws2["B2"].comment is not None
    assert ws2["B2"].comment.text == "from wolfxl"
    # Verify the comments part landed in the ZIP.
    with zipfile.ZipFile(dst) as zf:
        names = zf.namelist()
        assert any(n.startswith("xl/comments") and n.endswith(".xml") for n in names), names
        assert any(n.startswith("xl/drawings/vmlDrawing") for n in names), names


def test_modify_existing_comment(tmp_path: Path) -> None:
    """File with one comment → modify text → openpyxl sees new text."""
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_one_comment_fixture(src)

    wb = load_workbook(src, modify=True)
    ws = wb["Sheet1"]
    ws["A1"].comment = Comment(text="updated", author="bob")
    wb.save(dst)

    wb2 = openpyxl.load_workbook(dst)
    assert wb2["Sheet1"]["A1"].comment is not None
    assert "updated" in wb2["Sheet1"]["A1"].comment.text


def test_delete_comment(tmp_path: Path) -> None:
    """File with one comment → setting to None removes it."""
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_one_comment_fixture(src)

    wb = load_workbook(src, modify=True)
    ws = wb["Sheet1"]
    ws["A1"].comment = None
    wb.save(dst)

    wb2 = openpyxl.load_workbook(dst)
    assert wb2["Sheet1"]["A1"].comment is None


def test_no_dirty_save_is_byte_identical(tmp_path: Path) -> None:
    """Short-circuit guard: opening + saving without queueing comments
    must produce a byte-identical file (no spurious commentsN.xml regen)."""
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_one_comment_fixture(src)

    wb = load_workbook(src, modify=True)
    wb.save(dst)
    assert src.read_bytes() == dst.read_bytes()
