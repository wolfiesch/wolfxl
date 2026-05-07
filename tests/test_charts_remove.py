"""Sprint Ξ (RFC-050) — ``Worksheet.remove_chart`` / ``replace_chart``.

Coverage matrix:

* ``remove_chart`` removes a not-yet-flushed chart from the pending list
  and produces a workbook with no chart parts.
* ``remove_chart`` raises ``ValueError`` on an unknown chart.
* ``replace_chart`` swaps one chart for another in the pending list,
  preserving the anchor and the list position.
* ``replace_chart`` raises ``TypeError`` if the new value isn't a
  :class:`ChartBase` subclass.
* ``replace_chart`` raises ``ValueError`` if the old chart is unknown.
* Anchor inheritance: replacement chart with no anchor of its own
  inherits the anchor from the chart it replaces.
"""

from __future__ import annotations

import zipfile
from pathlib import Path
from typing import Any

import openpyxl
import pytest


pytestmark = pytest.mark.skipif(
    not pytest.importorskip("wolfxl", reason="wolfxl not installed"),
    reason="wolfxl required",
)


def _build_workbook_with_pending_chart() -> tuple[Any, Any, Any]:
    """Helper — build a write-mode workbook with one pending chart."""
    import wolfxl
    from wolfxl.chart import BarChart, Reference

    wb = wolfxl.Workbook()
    ws = wb.active
    ws.append(["Region", "Q1", "Q2"])
    ws.append(["NA", 100, 110])
    ws.append(["EU", 80, 95])

    bar = BarChart()
    bar.title = "Bar"
    data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=3)
    bar.add_data(data, titles_from_data=True)
    ws.add_chart(bar, "E2")

    return wb, ws, bar


def _build_workbook_with_two_openpyxl_charts(path: Path) -> Path:
    openpyxl = pytest.importorskip("openpyxl")
    from openpyxl.chart import BarChart, LineChart, Reference

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"
    ws.append(["Region", "Q1", "Q2"])
    ws.append(["NA", 100, 110])
    ws.append(["EU", 80, 95])
    data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=3)
    bar = BarChart()
    bar.add_data(data, titles_from_data=True)
    ws.add_chart(bar, "E2")
    line = LineChart()
    line.add_data(data, titles_from_data=True)
    ws.add_chart(line, "L2")
    wb.save(path)
    return path


def _inject_shared_chart_dependencies(path: Path) -> None:
    rels_xml = (
        b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        b'<Relationship Id="rId1" '
        b'Type="http://schemas.microsoft.com/office/2011/relationships/chartStyle" '
        b'Target="style1.xml"/>'
        b'<Relationship Id="rId2" '
        b'Type="http://schemas.microsoft.com/office/2011/relationships/chartColorStyle" '
        b'Target="colors1.xml"/>'
        b"</Relationships>"
    )
    shared_parts = {
        "xl/charts/_rels/chart1.xml.rels": rels_xml,
        "xl/charts/_rels/chart2.xml.rels": rels_xml,
        "xl/charts/style1.xml": b"<c:styleSheet/>",
        "xl/charts/colors1.xml": b"<c:colorStyle/>",
    }
    parts: dict[str, bytes] = {}
    with zipfile.ZipFile(path, "r") as zf:
        for name in zf.namelist():
            parts[name] = zf.read(name)
    parts.update(shared_parts)
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zf:
        for name, data in parts.items():
            zf.writestr(name, data)


def _build_workbook_with_image_and_chart(path: Path) -> Path:
    openpyxl = pytest.importorskip("openpyxl")
    from openpyxl.chart import BarChart, Reference
    from openpyxl.drawing.image import Image

    png_path = Path(__file__).parent / "fixtures" / "images" / "tiny_red_dot.png"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"
    ws.append(["Region", "Q1"])
    ws.append(["NA", 100])
    ws.append(["EU", 80])
    ws.add_image(Image(str(png_path)), "B5")
    data = Reference(ws, min_col=2, min_row=1, max_row=3)
    bar = BarChart()
    bar.add_data(data, titles_from_data=True)
    ws.add_chart(bar, "E2")
    wb.save(path)
    return path


def _build_openpyxl_chart_fixture(path: Path) -> None:
    from openpyxl.chart import BarChart, Reference

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    rows = [
        ["Region", "Q1", "Q2"],
        ["NA", 100, 110],
        ["EU", 80, 95],
    ]
    for row in rows:
        ws.append(row)
    chart = BarChart()
    chart.title = "Original"
    chart.add_data(
        Reference(ws, min_col=2, min_row=1, max_col=3, max_row=3),
        titles_from_data=True,
    )
    ws.add_chart(chart, "E2")
    wb.save(path)


def _chart_and_drawing_parts(path: Path) -> list[str]:
    with zipfile.ZipFile(path) as zf:
        return sorted(
            name
            for name in zf.namelist()
            if name.startswith("xl/charts/") or name.startswith("xl/drawings/")
        )


def test_remove_chart_drops_pending(tmp_path: Path) -> None:
    """Removing a pending chart yields a workbook with no chart parts."""
    wb, ws, bar = _build_workbook_with_pending_chart()
    assert len(ws._pending_charts) == 1

    ws.remove_chart(bar)
    assert len(ws._pending_charts) == 0

    out = tmp_path / "no_charts.xlsx"
    wb.save(out)

    with zipfile.ZipFile(out) as zf:
        names = zf.namelist()
    assert not any(n.startswith("xl/charts/") for n in names), (
        f"expected no chart parts after remove_chart, got {names!r}"
    )


def test_remove_chart_unknown_raises() -> None:
    """Removing a chart that was never added raises ``ValueError``."""
    import wolfxl
    from wolfxl.chart import BarChart

    wb = wolfxl.Workbook()
    ws = wb.active
    ghost = BarChart()
    with pytest.raises(ValueError, match="not added"):
        ws.remove_chart(ghost)


def test_remove_loaded_source_chart_persists(tmp_path: Path) -> None:
    import wolfxl

    src = tmp_path / "source_chart.xlsx"
    _build_openpyxl_chart_fixture(src)

    wb = wolfxl.load_workbook(src, modify=True)
    ws = wb["Data"]
    chart = ws._charts[0]
    ws.remove_chart(chart)
    out = tmp_path / "removed_source_chart.xlsx"
    wb.save(out)

    reloaded = openpyxl.load_workbook(out)
    assert reloaded["Data"]._charts == []
    with zipfile.ZipFile(out) as zf:
        assert not any(name.startswith("xl/charts/chart") for name in zf.namelist())


def test_replace_loaded_source_chart_persists(tmp_path: Path) -> None:
    import wolfxl
    from wolfxl.chart import LineChart, Reference

    src = tmp_path / "source_chart_replace.xlsx"
    _build_openpyxl_chart_fixture(src)

    wb = wolfxl.load_workbook(src, modify=True)
    ws = wb["Data"]
    old = ws._charts[0]
    line = LineChart()
    line.title = "Replacement"
    line.add_data(Reference(ws, min_col=2, min_row=1, max_col=3, max_row=3), titles_from_data=True)
    ws.replace_chart(old, line)
    out = tmp_path / "replaced_source_chart.xlsx"
    wb.save(out)

    with zipfile.ZipFile(out) as zf:
        chart_xml = zf.read("xl/charts/chart1.xml").decode("utf-8")
    assert "<c:lineChart>" in chart_xml
    assert "<c:barChart>" not in chart_xml


def test_same_save_scratch_chart_remove_and_sheet_delete_preserves_source_drawings(
    tmp_path: Path,
) -> None:
    """Scratch chart cleanup must not delete unrelated source drawing parts."""
    import shutil

    import wolfxl
    from wolfxl.chart import BarChart, Reference

    fixture_dir = Path(__file__).parent / "fixtures" / "external_oracle"
    for fixture in sorted(fixture_dir.glob("*.xlsx")):
        path = tmp_path / fixture.name
        shutil.copy2(fixture, path)
        before = _chart_and_drawing_parts(path)

        wb = wolfxl.load_workbook(path, modify=True)
        scratch = wb.create_sheet("WolfXL Chart Scratch")
        scratch["A1"] = "value"
        scratch["A2"] = 1
        chart = BarChart()
        chart.add_data(
            Reference(scratch, min_col=1, min_row=1, max_row=2),
            titles_from_data=True,
        )
        scratch.add_chart(chart, "C2")
        wb.save(path)
        wb.close()

        wb = wolfxl.load_workbook(path, modify=True)
        scratch = wb["WolfXL Chart Scratch"]
        scratch.remove_chart(scratch._charts[-1])
        wb.remove(scratch)
        wb.save(path)
        wb.close()

        assert _chart_and_drawing_parts(path) == before, fixture.name


def test_loaded_source_chart_title_edit_persists(tmp_path: Path) -> None:
    import wolfxl

    src = tmp_path / "source_chart_title.xlsx"
    _build_openpyxl_chart_fixture(src)

    wb = wolfxl.load_workbook(src, modify=True)
    wb["Data"]._charts[0].title = "Changed"
    out = tmp_path / "title_changed.xlsx"
    wb.save(out)

    with zipfile.ZipFile(out) as zf:
        chart_xml = zf.read("xl/charts/chart1.xml").decode("utf-8")
    assert "Changed" in chart_xml
    assert "Original" not in chart_xml


def test_replace_chart_swaps_in_place(tmp_path: Path) -> None:
    """Replacing keeps the chart count, position, and inherited anchor."""
    from wolfxl.chart import LineChart, Reference

    wb, ws, bar = _build_workbook_with_pending_chart()
    assert ws._pending_charts == [bar]
    assert bar._anchor == "E2"

    line = LineChart()
    line.title = "Line"
    data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=3)
    line.add_data(data, titles_from_data=True)

    ws.replace_chart(bar, line)
    assert ws._pending_charts == [line]
    # Anchor inherited from the chart it replaces.
    assert line._anchor == "E2"

    out = tmp_path / "replaced.xlsx"
    wb.save(out)
    with zipfile.ZipFile(out) as zf:
        chart_names = [n for n in zf.namelist() if n.startswith("xl/charts/")]
        # Exactly one chart part (the replacement, not both).
        chart_xmls = [n for n in chart_names if n.endswith(".xml")]
        assert len(chart_xmls) == 1, chart_names
        # The line chart emits <c:lineChart>, not <c:barChart>.
        xml = zf.read(chart_xmls[0]).decode()
        assert "<c:lineChart>" in xml
        assert "<c:barChart>" not in xml


def test_replace_chart_keeps_explicit_anchor(tmp_path: Path) -> None:
    """If new._anchor is set explicitly, it wins over inheritance."""
    from wolfxl.chart import LineChart, Reference

    wb, ws, bar = _build_workbook_with_pending_chart()

    line = LineChart()
    line._anchor = "G10"
    data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=3)
    line.add_data(data, titles_from_data=True)

    ws.replace_chart(bar, line)
    assert line._anchor == "G10"


def test_replace_chart_unknown_raises() -> None:
    """Replacing a chart that was never added raises ``ValueError``."""
    import wolfxl
    from wolfxl.chart import BarChart, LineChart

    wb = wolfxl.Workbook()
    ws = wb.active
    ghost_bar = BarChart()
    new_line = LineChart()
    with pytest.raises(ValueError, match="not added"):
        ws.replace_chart(ghost_bar, new_line)


def test_replace_chart_wrong_type_raises() -> None:
    """``replace_chart`` rejects non-ChartBase replacements."""
    wb, ws, bar = _build_workbook_with_pending_chart()
    with pytest.raises(TypeError, match="ChartBase"):
        ws.replace_chart(bar, "not a chart")


def test_remove_chart_then_add_chart_works(tmp_path: Path) -> None:
    """remove_chart followed by a fresh add_chart yields a 1-chart workbook."""
    from wolfxl.chart import LineChart, Reference

    wb, ws, bar = _build_workbook_with_pending_chart()
    ws.remove_chart(bar)

    line = LineChart()
    data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=3)
    line.add_data(data, titles_from_data=True)
    ws.add_chart(line, "G5")

    out = tmp_path / "remove_then_add.xlsx"
    wb.save(out)
    with zipfile.ZipFile(out) as zf:
        chart_xmls = [
            n for n in zf.namelist()
            if n.startswith("xl/charts/") and n.endswith(".xml")
        ]
        assert len(chart_xmls) == 1
        xml = zf.read(chart_xmls[0]).decode()
        assert "<c:lineChart>" in xml


def test_modify_remove_loaded_chart_round_trip(tmp_path: Path) -> None:
    """Removing a loaded source-workbook chart removes its OOXML parts."""
    import wolfxl

    base_wb, _, _ = _build_workbook_with_pending_chart()
    base = tmp_path / "with_chart.xlsx"
    base_wb.save(base)

    wb = wolfxl.load_workbook(base, modify=True)
    ws = wb.active
    assert len(ws._charts) == 1
    ws.remove_chart(ws._charts[0])
    out = tmp_path / "removed_loaded_chart.xlsx"
    wb.save(out)

    with zipfile.ZipFile(out) as zf:
        names = zf.namelist()
        assert not any(n.startswith("xl/charts/") for n in names), names
        assert "xl/drawings/drawing1.xml" not in names
        sheet_xml = zf.read("xl/worksheets/sheet1.xml").decode()
        content_types = zf.read("[Content_Types].xml").decode()
    assert "<drawing" not in sheet_xml
    assert "drawingml.chart+xml" not in content_types


def test_modify_replace_loaded_chart_round_trip(tmp_path: Path) -> None:
    """Replacing a loaded chart removes the old part and adds the new chart."""
    import wolfxl
    from wolfxl.chart import LineChart, Reference

    base_wb, _, _ = _build_workbook_with_pending_chart()
    base = tmp_path / "with_chart.xlsx"
    base_wb.save(base)

    wb = wolfxl.load_workbook(base, modify=True)
    ws = wb.active
    assert len(ws._charts) == 1
    line = LineChart()
    data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=3)
    line.add_data(data, titles_from_data=True)
    ws.replace_chart(ws._charts[0], line)
    out = tmp_path / "replaced_loaded_chart.xlsx"
    wb.save(out)

    with zipfile.ZipFile(out) as zf:
        chart_xmls = [
            n for n in zf.namelist()
            if n.startswith("xl/charts/") and n.endswith(".xml")
        ]
        assert len(chart_xmls) == 1
        xml = zf.read(chart_xmls[0]).decode()
    assert "<c:lineChart>" in xml
    assert "<c:barChart>" not in xml


def test_modify_remove_one_loaded_chart_keeps_second(tmp_path: Path) -> None:
    """Removing one loaded chart from a shared drawing keeps the sibling chart."""
    import wolfxl

    base = _build_workbook_with_two_openpyxl_charts(tmp_path / "two_charts.xlsx")
    wb = wolfxl.load_workbook(base, modify=True)
    ws = wb.active
    assert len(ws._charts) == 2
    ws.remove_chart(ws._charts[0])
    out = tmp_path / "one_chart_left.xlsx"
    wb.save(out)

    openpyxl = pytest.importorskip("openpyxl")
    wb2 = openpyxl.load_workbook(out)
    assert len(wb2.active._charts) == 1
    with zipfile.ZipFile(out) as zf:
        chart_xmls = [
            n for n in zf.namelist()
            if n.startswith("xl/charts/") and n.endswith(".xml")
        ]
        drawing_xml = zf.read("xl/drawings/drawing1.xml").decode()
    assert len(chart_xmls) == 1
    assert drawing_xml.count("<c:chart") == 1


def test_modify_remove_one_loaded_chart_preserves_shared_chart_dependencies(
    tmp_path: Path,
) -> None:
    import wolfxl

    base = _build_workbook_with_two_openpyxl_charts(tmp_path / "shared_chart_deps.xlsx")
    _inject_shared_chart_dependencies(base)

    wb = wolfxl.load_workbook(base, modify=True)
    ws = wb.active
    assert len(ws._charts) == 2
    ws.remove_chart(ws._charts[0])
    out = tmp_path / "shared_chart_deps_removed.xlsx"
    wb.save(out)

    with zipfile.ZipFile(out) as zf:
        names = set(zf.namelist())
        chart2_rels = zf.read("xl/charts/_rels/chart2.xml.rels").decode("utf-8")

    assert "xl/charts/chart1.xml" not in names
    assert "xl/charts/_rels/chart1.xml.rels" not in names
    assert "xl/charts/chart2.xml" in names
    assert "xl/charts/_rels/chart2.xml.rels" in names
    assert "xl/charts/style1.xml" in names
    assert "xl/charts/colors1.xml" in names
    assert "chartStyle" in chart2_rels
    assert "chartColorStyle" in chart2_rels


def test_modify_remove_loaded_chart_from_mixed_drawing_keeps_image(tmp_path: Path) -> None:
    """Removing a chart from an image+chart drawing preserves the image anchor."""
    import wolfxl

    base = _build_workbook_with_image_and_chart(tmp_path / "image_chart.xlsx")
    wb = wolfxl.load_workbook(base, modify=True)
    ws = wb.active
    assert len(ws._charts) == 1
    assert len(ws.images) == 1
    ws.remove_chart(ws._charts[0])
    out = tmp_path / "image_only.xlsx"
    wb.save(out)

    openpyxl = pytest.importorskip("openpyxl")
    wb2 = openpyxl.load_workbook(out)
    assert len(wb2.active._charts) == 0
    assert len(wb2.active._images) == 1
    with zipfile.ZipFile(out) as zf:
        names = zf.namelist()
        sheet_xml = zf.read("xl/worksheets/sheet1.xml").decode()
    assert "xl/drawings/drawing1.xml" in names
    assert "<drawing" in sheet_xml


def test_modify_remove_loaded_chart_then_add_new_chart(tmp_path: Path) -> None:
    """Explicit remove-then-add leaves exactly the newly queued chart."""
    import wolfxl
    from wolfxl.chart import LineChart, Reference

    base_wb, _, _ = _build_workbook_with_pending_chart()
    base = tmp_path / "with_chart.xlsx"
    base_wb.save(base)

    wb = wolfxl.load_workbook(base, modify=True)
    ws = wb.active
    ws.remove_chart(ws._charts[0])
    line = LineChart()
    data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=3)
    line.add_data(data, titles_from_data=True)
    ws.add_chart(line, "H4")
    out = tmp_path / "remove_then_add_loaded.xlsx"
    wb.save(out)

    with zipfile.ZipFile(out) as zf:
        chart_xmls = [
            n for n in zf.namelist()
            if n.startswith("xl/charts/") and n.endswith(".xml")
        ]
        assert len(chart_xmls) == 1
        xml = zf.read(chart_xmls[0]).decode()
    assert "<c:lineChart>" in xml
    assert "<c:barChart>" not in xml
