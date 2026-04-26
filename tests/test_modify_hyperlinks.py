"""RFC-022 — Cell hyperlinks round-trip in modify mode.

End-to-end coverage for ``cell.hyperlink = ...`` on an existing file.
The save-time path threads three layers:

1. ``Cell.hyperlink`` setter (Python) drops the value into
   ``ws._pending_hyperlinks[coord]``. ``None`` is the explicit-delete
   sentinel (INDEX decision #5 — never use ``pop()``).
2. ``Workbook._flush_pending_hyperlinks_to_patcher`` (Python) drains
   each sheet's pending dict into ``XlsxPatcher.queue_hyperlink`` /
   ``queue_hyperlink_delete``.
3. ``XlsxPatcher::do_save`` Phase 2.5e (Rust) calls
   ``hyperlinks::extract_hyperlinks`` against the source sheet XML +
   rels graph, merges with queued ops via ``build_hyperlinks_block``,
   pushes a ``SheetBlock::Hyperlinks`` (slot 19) to the merger, and
   emits the mutated rels graph through Phase 3's serializer.

Sister contract: ``test_no_dirty_save_is_byte_identical`` is the
regression guard for the short-circuit predicate. If a future refactor
forgets to require ``queued_hyperlinks.is_empty()``, this test fires.
"""
from __future__ import annotations

import zipfile
from pathlib import Path

import openpyxl
import pytest

from wolfxl import Workbook
from wolfxl.worksheet.hyperlink import Hyperlink

# ---------------------------------------------------------------------------
# Fixtures and helpers
# ---------------------------------------------------------------------------


@pytest.fixture(autouse=True)
def _force_test_epoch(monkeypatch: pytest.MonkeyPatch) -> None:
    """Pin ZIP entry mtimes for byte-stable saves."""
    monkeypatch.setenv("WOLFXL_TEST_EPOCH", "0")


def _make_clean_fixture(path: Path) -> None:
    """Workbook with no hyperlinks — for testing the add-from-empty path."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "header"
    wb.save(path)


def _make_two_link_fixture(path: Path) -> Path:
    """Workbook with two existing external hyperlinks. Used to test the
    "add a third — preserve the first two" and "delete one — keep the
    other" paths."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "first"
    ws["A1"].hyperlink = "https://example.com/one"
    ws["A2"] = "second"
    ws["A2"].hyperlink = "https://example.com/two"
    wb.save(path)
    return path


def _read_zip_text(path: Path, entry: str) -> str:
    with zipfile.ZipFile(path) as zf:
        return zf.read(entry).decode("utf-8")


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


def test_set_external_hyperlink_round_trip(tmp_path: Path) -> None:
    """Smoke test: clean file → set one external hyperlink → openpyxl reads."""
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    _make_clean_fixture(src)

    wb = Workbook._from_patcher(str(src))
    ws = wb.active
    ws["A1"].hyperlink = "https://example.com/added"
    wb.save(out)
    wb.close()

    re_wb = openpyxl.load_workbook(out)
    cell = re_wb.active["A1"]
    assert cell.hyperlink is not None
    assert cell.hyperlink.target == "https://example.com/added"


def test_set_internal_hyperlink_round_trip(tmp_path: Path) -> None:
    """Internal links use ``location`` (no rId allocated)."""
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    # Pre-create the Targets sheet via openpyxl since create_sheet is
    # write-mode only (modify mode mutates but doesn't add structure).
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "header"
    wb.create_sheet("Targets")
    wb.save(src)

    wb2 = Workbook._from_patcher(str(src))
    wb2["Sheet1"]["A1"].hyperlink = Hyperlink(location="'Targets'!A1")
    wb2.save(out)
    wb2.close()

    re_wb = openpyxl.load_workbook(out)
    cell = re_wb["Sheet1"]["A1"]
    assert cell.hyperlink is not None
    assert cell.hyperlink.location == "'Targets'!A1"
    assert cell.hyperlink.target is None


def test_delete_external_hyperlink_removes_rid(tmp_path: Path) -> None:
    """``cell.hyperlink = None`` drops the rId from sheet rels."""
    src = _make_two_link_fixture(tmp_path / "src.xlsx")
    out = tmp_path / "out.xlsx"

    wb = Workbook._from_patcher(str(src))
    wb["Sheet1"]["A1"].hyperlink = None
    wb.save(out)
    wb.close()

    rels = _read_zip_text(out, "xl/worksheets/_rels/sheet1.xml.rels")
    assert "https://example.com/one" not in rels, "deleted rId still in rels"
    assert "https://example.com/two" in rels, "remaining rId got dropped"


def test_no_dirty_save_is_byte_identical(tmp_path: Path) -> None:
    """Open + save without any hyperlink edit → byte-identical to source.

    Short-circuit-predicate regression guard for RFC-022. If a future
    refactor forgets to include ``queued_hyperlinks.is_empty()``, this
    test fires.
    """
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    _make_clean_fixture(src)
    src_bytes = src.read_bytes()

    wb = Workbook._from_patcher(str(src))
    wb.save(out)
    wb.close()

    assert out.read_bytes() == src_bytes, (
        "no-op modify-mode save changed bytes — short-circuit predicate "
        "likely missing queued_hyperlinks.is_empty()."
    )


def test_existing_hyperlinks_preserved_when_adding_new(tmp_path: Path) -> None:
    """Fixture with 2 hyperlinks; add a 3rd; all 3 present after save."""
    src = _make_two_link_fixture(tmp_path / "src.xlsx")
    out = tmp_path / "out.xlsx"

    wb = Workbook._from_patcher(str(src))
    wb["Sheet1"]["A3"].hyperlink = "https://example.com/three"
    wb.save(out)
    wb.close()

    re_wb = openpyxl.load_workbook(out)
    ws = re_wb.active
    assert ws["A1"].hyperlink.target == "https://example.com/one"
    assert ws["A2"].hyperlink.target == "https://example.com/two"
    assert ws["A3"].hyperlink.target == "https://example.com/three"


def test_hyperlink_with_tooltip_and_display(tmp_path: Path) -> None:
    """Both ``tooltip`` and ``display`` round-trip through openpyxl."""
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    _make_clean_fixture(src)

    wb = Workbook._from_patcher(str(src))
    wb.active["A1"].hyperlink = Hyperlink(
        target="https://example.com",
        tooltip="Click here",
        display="Example",
    )
    wb.save(out)
    wb.close()

    re_wb = openpyxl.load_workbook(out)
    h = re_wb.active["A1"].hyperlink
    assert h.target == "https://example.com"
    assert h.tooltip == "Click here"
    assert h.display == "Example"


def test_url_with_ampersand_xml_escaped(tmp_path: Path) -> None:
    """``&`` in target serializes as ``&amp;`` in the rels file (XML escape)."""
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    _make_clean_fixture(src)

    wb = Workbook._from_patcher(str(src))
    wb.active["A1"].hyperlink = "https://example.com/?q=excel&sort=desc"
    wb.save(out)
    wb.close()

    rels = _read_zip_text(out, "xl/worksheets/_rels/sheet1.xml.rels")
    assert "&amp;" in rels, f"ampersand not XML-escaped in rels: {rels}"
    assert "q=excel&sort" not in rels, "raw ampersand emitted (XML invalid)"
    # Round-trip through openpyxl recovers the literal:
    re_wb = openpyxl.load_workbook(out)
    assert re_wb.active["A1"].hyperlink.target == "https://example.com/?q=excel&sort=desc"


def test_set_then_delete_same_cell_in_one_session(tmp_path: Path) -> None:
    """Set + delete on the same cell in one save() → no hyperlink in output."""
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    _make_clean_fixture(src)

    wb = Workbook._from_patcher(str(src))
    wb.active["A1"].hyperlink = "https://example.com/transient"
    wb.active["A1"].hyperlink = None
    wb.save(out)
    wb.close()

    re_wb = openpyxl.load_workbook(out)
    assert re_wb.active["A1"].hyperlink is None
    # Rels file (if it exists) must not carry the transient URL.
    with zipfile.ZipFile(out) as zf:
        if "xl/worksheets/_rels/sheet1.xml.rels" in zf.namelist():
            rels = zf.read("xl/worksheets/_rels/sheet1.xml.rels").decode("utf-8")
            assert "transient" not in rels
