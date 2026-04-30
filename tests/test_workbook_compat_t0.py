"""T0 compat for Workbook API parity with openpyxl + utils exports."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import openpyxl.utils.cell as ouc
import pytest
import wolfxl.utils as wu

import wolfxl


def test_worksheets_list() -> None:
    wb = wolfxl.Workbook()
    ws1 = wb.active
    ws2 = wb.create_sheet("Sheet2")
    assert wb.worksheets == [ws1, ws2]


def test_get_sheet_by_name() -> None:
    wb = wolfxl.Workbook()
    ws = wb.active
    assert wb.get_sheet_by_name("Sheet") is ws


def test_index_returns_sheet_position() -> None:
    wb = wolfxl.Workbook()
    ws2 = wb.create_sheet("Sheet2")
    ws3 = wb.create_sheet("Sheet3")
    assert wb.index(ws2) == 1
    assert wb.index(ws3) == 2
    assert wb.get_index(ws3) == 2
    assert wb.get_sheet_names() == wb.sheetnames


def test_create_named_range_alias() -> None:
    wb = wolfxl.Workbook()
    ws = wb.active
    wb.create_named_range("Region", ws, "$A$1:$A$5")
    assert wb.defined_names["Region"].value == "'Sheet'!$A$1:$A$5"


def test_add_named_style_registers_name_and_binds() -> None:
    from wolfxl.styles import NamedStyle

    wb = wolfxl.Workbook()
    style = NamedStyle(name="Metric")
    wb.add_named_style(style)
    assert "Metric" in wb.named_styles
    assert "Metric" in wb.style_names
    assert style._wb is wb


def test_style_names_read_existing_named_styles(tmp_path: Path) -> None:
    from openpyxl.styles import NamedStyle

    path = tmp_path / "named-style.xlsx"
    op_wb = openpyxl.Workbook()
    op_wb.add_named_style(NamedStyle(name="Metric"))
    op_wb.save(path)

    wb = wolfxl.load_workbook(path)
    assert wb.style_names == op_wb.style_names == ["Normal", "Metric"]
    assert wb.named_styles == ["Normal", "Metric"]


def test_create_chartsheet_raises_clear_error() -> None:
    wb = wolfxl.Workbook()
    with pytest.raises(NotImplementedError, match="create_chartsheet"):
        wb.create_chartsheet("Chart")


def test_read_only_false_for_write_mode() -> None:
    wb = wolfxl.Workbook()
    assert wb.read_only is False


def test_read_only_true_for_read_mode(tmp_path: Path) -> None:
    # Sprint Ι Pod-β: ``Workbook.read_only`` now reflects the *explicit*
    # ``read_only=True`` opt-in passed to ``load_workbook`` (matching
    # openpyxl's contract), not the historic "no writer, no patcher"
    # inference. Plain read mode (the default) is no longer
    # automatically tagged as read_only — callers who want the
    # streaming fast path opt in by passing ``read_only=True``.
    path = tmp_path / "t.xlsx"
    wolfxl.Workbook().save(path)
    wb_default = wolfxl.load_workbook(path)
    assert wb_default.read_only is False
    wb_explicit = wolfxl.load_workbook(path, read_only=True)
    assert wb_explicit.read_only is True


def test_read_only_false_for_modify_mode(tmp_path: Path) -> None:
    path = tmp_path / "t.xlsx"
    wolfxl.Workbook().save(path)
    wb = wolfxl.load_workbook(path, modify=True)
    assert wb.read_only is False


def test_remove_sheet() -> None:
    wb = wolfxl.Workbook()
    ws2 = wb.create_sheet("Sheet2")
    wb.remove(ws2)
    assert "Sheet2" not in wb.sheetnames
    assert len(wb.worksheets) == 1


def test_remove_sheet_alias() -> None:
    wb = wolfxl.Workbook()
    ws2 = wb.create_sheet("Sheet2")
    wb.remove_sheet(ws2)
    assert "Sheet2" not in wb.sheetnames


def test_remove_read_mode_raises(tmp_path: Path) -> None:
    path = tmp_path / "t.xlsx"
    wolfxl.Workbook().save(path)
    wb = wolfxl.load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    with pytest.raises(RuntimeError, match="write mode"):
        wb.remove(ws)


# ---------------- utils parity ----------------


@pytest.mark.parametrize(
    "coord",
    ["A1", "B2", "Z100", "AA1", "A1:B2", "Z1:AA2"],
)
def test_absolute_coordinate(coord: str) -> None:
    assert wu.absolute_coordinate(coord) == ouc.absolute_coordinate(coord)


@pytest.mark.parametrize(
    "sheet",
    ["Sheet1", "My Sheet", "sheet with 'quote", "Data_2024", "Sheet!Bang"],
)
def test_quote_sheetname(sheet: str) -> None:
    assert wu.quote_sheetname(sheet) == ouc.quote_sheetname(sheet)


@pytest.mark.parametrize(
    "ref",
    ["Sheet1!A1:B2", "'My Sheet'!A1:C3", "'Weird Sheet'!Z1:AA10"],
)
def test_range_to_tuple(ref: str) -> None:
    assert wu.range_to_tuple(ref) == ouc.range_to_tuple(ref)


@pytest.mark.parametrize(
    "rng",
    ["A1:C2", "A1:A5", "B2:C3", "Z1:AA2"],
)
def test_rows_from_range(rng: str) -> None:
    assert list(wu.rows_from_range(rng)) == list(ouc.rows_from_range(rng))


@pytest.mark.parametrize(
    "rng",
    ["A1:C2", "A1:A5", "B2:D3", "Z1:AA2"],
)
def test_cols_from_range(rng: str) -> None:
    assert list(wu.cols_from_range(rng)) == list(ouc.cols_from_range(rng))


@pytest.mark.parametrize(
    "start,end",
    [("A", "C"), ("A", "A"), ("B", "D"), (1, 3), (5, 10), (10, 5)],
)
def test_get_column_interval(start: int | str, end: int | str) -> None:
    assert wu.get_column_interval(start, end) == ouc.get_column_interval(start, end)


def test_range_to_tuple_without_sheet_raises() -> None:
    with pytest.raises(ValueError):
        wu.range_to_tuple("A1:B2")
