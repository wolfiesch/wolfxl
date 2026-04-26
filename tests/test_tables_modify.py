"""RFC-024 — Sheet-scoped tables round-trip in modify mode.

End-to-end coverage for ``ws.add_table(t)`` on an existing file. The
save-time path threads four layers:

1. ``Worksheet.add_table`` (Python) appends the ``Table`` instance to
   ``self._pending_tables`` regardless of write/modify mode.
2. ``Workbook._flush_pending_tables_to_patcher`` (Python) drains each
   sheet's pending list into ``XlsxPatcher.queue_table``.
3. ``XlsxPatcher::do_save`` Phase 2.5f (Rust) scans the source ZIP for
   the workbook-wide existing-table inventory (id + name + count),
   calls ``tables::build_tables`` to allocate fresh ids, mutate the
   sheet rels graph, queue ``[Content_Types].xml`` Override entries,
   and emit a merged ``<tableParts>`` block (slot 37 in the merger).
4. The patcher's Phase-2.5c content-types aggregator and Phase-3
   merger absorb those ops into a single rewrite of
   ``[Content_Types].xml`` and the sheet XML respectively.

Manual / out-of-band check (documented but not asserted in CI):
   LibreOffice ``--headless --convert-to xlsx`` round-trip preserves
   the table style and column names. Run locally via:

       soffice --headless --convert-to xlsx out.xlsx --outdir /tmp/lo

   then re-open ``/tmp/lo/out.xlsx`` in openpyxl and inspect
   ``ws.tables[name]``. Excel-strict validation is covered by
   openpyxl reading the file back, which is asserted here.
"""

from __future__ import annotations

import zipfile
from pathlib import Path

import openpyxl
import pytest
from wolfxl.worksheet.table import Table, TableColumn, TableStyleInfo

from wolfxl import Workbook, load_workbook

# ---------------------------------------------------------------------------
# Fixtures and helpers
# ---------------------------------------------------------------------------


@pytest.fixture(autouse=True)
def _force_test_epoch(monkeypatch: pytest.MonkeyPatch) -> None:
    """Pin ZIP entry mtimes for byte-stable saves."""
    monkeypatch.setenv("WOLFXL_TEST_EPOCH", "0")


def _make_clean_fixture(path: Path, headers: list[str], rows: int = 4) -> None:
    """Workbook with no tables — header row + a few data rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)
    for r in range(2, 2 + rows):
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=r, column=col_idx, value=r * col_idx)
    wb.save(path)


def _make_one_table_fixture(path: Path) -> None:
    """Workbook that already has one table (id=1) on Sheet1."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    headers = ["A", "B", "C"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    for r in range(2, 6):
        for i in range(1, 4):
            ws.cell(row=r, column=i, value=r + i)
    t = openpyxl.worksheet.table.Table(displayName="Existing", ref="A1:C5")
    ws.add_table(t)
    wb.save(path)


def _make_two_sheet_fixture(path: Path) -> None:
    """Workbook with two sheets, no tables."""
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1["A1"] = "x"
    ws1["A2"] = 1
    ws2 = wb.create_sheet("Sheet2")
    ws2["A1"] = "y"
    ws2["A2"] = 2
    wb.save(path)


def _read_zip_text(path: Path, entry: str) -> str:
    with zipfile.ZipFile(path) as zf:
        return zf.read(entry).decode("utf-8")


def _zip_namelist(path: Path) -> list[str]:
    with zipfile.ZipFile(path) as zf:
        return list(zf.namelist())


def _simple_table(name: str, ref: str, columns: list[str]) -> Table:
    return Table(
        name=name,
        displayName=name,
        ref=ref,
        tableColumns=[TableColumn(id=i + 1, name=c) for i, c in enumerate(columns)],
    )


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


def test_add_table_to_clean_file(tmp_path: Path) -> None:
    """Clean file → one new table. New part exists, content-type override
    is present, sheet rels is created, sheet XML carries <tableParts>."""
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    _make_clean_fixture(src, ["Region", "Q1", "Q2"], rows=5)

    wb = load_workbook(src, modify=True)
    ws = wb.active
    ws.add_table(_simple_table("SalesTable", "A1:C6", ["Region", "Q1", "Q2"]))
    wb.save(out)
    wb.close()

    names = _zip_namelist(out)
    assert "xl/tables/table1.xml" in names

    table_xml = _read_zip_text(out, "xl/tables/table1.xml")
    assert 'id="1"' in table_xml
    assert 'name="SalesTable"' in table_xml
    assert 'displayName="SalesTable"' in table_xml
    assert 'ref="A1:C6"' in table_xml

    ct_xml = _read_zip_text(out, "[Content_Types].xml")
    assert "/xl/tables/table1.xml" in ct_xml
    assert "spreadsheetml.table+xml" in ct_xml

    rels_xml = _read_zip_text(out, "xl/worksheets/_rels/sheet1.xml.rels")
    assert "../tables/table1.xml" in rels_xml
    assert "/relationships/table" in rels_xml

    sheet_xml = _read_zip_text(out, "xl/worksheets/sheet1.xml")
    assert "<tableParts" in sheet_xml
    assert sheet_xml.count("<tablePart ") == 1


def test_add_table_to_file_with_one_existing_table(tmp_path: Path) -> None:
    """File already has table id=1 → new table gets id=2 (next available)
    and its part is named table2.xml."""
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    _make_one_table_fixture(src)

    wb = load_workbook(src, modify=True)
    ws = wb.active
    ws.add_table(_simple_table("Added", "E1:F4", ["X", "Y"]))
    wb.save(out)
    wb.close()

    names = _zip_namelist(out)
    assert "xl/tables/table1.xml" in names, "existing table preserved"
    assert "xl/tables/table2.xml" in names, "new table emitted"

    new_xml = _read_zip_text(out, "xl/tables/table2.xml")
    assert 'id="2"' in new_xml, new_xml
    assert 'name="Added"' in new_xml

    # Sheet XML must reference both tables
    sheet_xml = _read_zip_text(out, "xl/worksheets/sheet1.xml")
    assert sheet_xml.count("<tablePart ") == 2

    # openpyxl can read the result and sees both tables
    re_wb = openpyxl.load_workbook(out)
    re_ws = re_wb["Sheet1"]
    assert "Existing" in re_ws.tables
    assert "Added" in re_ws.tables


def test_table_style_preserved(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    _make_clean_fixture(src, ["A", "B"], rows=3)

    wb = load_workbook(src, modify=True)
    ws = wb.active
    t = _simple_table("Styled", "A1:B4", ["A", "B"])
    t.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight9",
        showFirstColumn=True,
        showRowStripes=True,
    )
    ws.add_table(t)
    wb.save(out)
    wb.close()

    table_xml = _read_zip_text(out, "xl/tables/table1.xml")
    assert 'name="TableStyleLight9"' in table_xml
    assert 'showFirstColumn="1"' in table_xml
    assert 'showRowStripes="1"' in table_xml


def test_openpyxl_can_read_the_result(tmp_path: Path) -> None:
    """End-to-end interop: write a table via wolfxl modify mode, read
    via openpyxl, verify range and column names round-trip."""
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    _make_clean_fixture(src, ["Region", "Q1", "Q2", "Q3"], rows=8)

    wb = load_workbook(src, modify=True)
    ws = wb.active
    cols = ["Region", "Q1", "Q2", "Q3"]
    ws.add_table(_simple_table("Interop", "A1:D9", cols))
    wb.save(out)
    wb.close()

    re_wb = openpyxl.load_workbook(out)
    re_ws = re_wb["Sheet1"]
    assert "Interop" in re_ws.tables, list(re_ws.tables.keys())
    re_t = re_ws.tables["Interop"]
    assert re_t.ref == "A1:D9"
    assert [c.name for c in re_t.tableColumns] == cols


def test_name_collision_raises(tmp_path: Path) -> None:
    """Adding a table whose name matches an existing table raises
    ValueError, not NotImplementedError."""
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    _make_one_table_fixture(src)  # creates table named "Existing"

    wb = load_workbook(src, modify=True)
    ws = wb.active
    ws.add_table(_simple_table("Existing", "E1:F4", ["X", "Y"]))
    with pytest.raises(ValueError, match="Existing"):
        wb.save(out)
    wb.close()


def test_no_pending_tables_is_byte_identical(tmp_path: Path) -> None:
    """Modify mode with no add_table calls must produce a byte-identical
    save (the Phase-2.5f short-circuit guard)."""
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    _make_one_table_fixture(src)

    wb = load_workbook(src, modify=True)
    wb.save(out)
    wb.close()

    assert src.read_bytes() == out.read_bytes()


def test_cross_mode_table_xml_byte_equivalent(tmp_path: Path) -> None:
    """RFC-024 §4.3 cross-mode parity: ``Workbook() + add_table + save``
    and ``load_workbook + add_table + save`` produce equivalent
    ``xl/tables/table1.xml`` bytes for the same Table input under
    ``WOLFXL_TEST_EPOCH=0``."""
    write_path = tmp_path / "write.xlsx"
    modify_src = tmp_path / "src.xlsx"
    modify_out = tmp_path / "modify.xlsx"

    cols = ["Region", "Q1", "Q2"]

    # Write mode
    wb_w = Workbook()
    ws_w = wb_w.active
    ws_w.title = "Sheet1"
    for i, h in enumerate(cols, start=1):
        ws_w.cell(row=1, column=i, value=h)
    for r in range(2, 6):
        for i in range(1, 4):
            ws_w.cell(row=r, column=i, value=r * i)
    ws_w.add_table(_simple_table("CrossMode", "A1:C5", cols))
    wb_w.save(write_path)
    wb_w.close()

    # Modify mode (start from a no-tables fixture, then add)
    _make_clean_fixture(modify_src, cols, rows=4)
    wb_m = load_workbook(modify_src, modify=True)
    ws_m = wb_m.active
    ws_m.add_table(_simple_table("CrossMode", "A1:C5", cols))
    wb_m.save(modify_out)
    wb_m.close()

    write_table_xml = _read_zip_text(write_path, "xl/tables/table1.xml")
    modify_table_xml = _read_zip_text(modify_out, "xl/tables/table1.xml")
    assert write_table_xml == modify_table_xml, (
        "table xml drifted between write and modify modes\n"
        f"write:  {write_table_xml}\n"
        f"modify: {modify_table_xml}"
    )


def test_two_sheet_id_allocation_is_global(tmp_path: Path) -> None:
    """Adding a table to each of two sheets in one save: ids must be
    workbook-unique (1 and 2), not per-sheet (both 1)."""
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    _make_two_sheet_fixture(src)

    wb = load_workbook(src, modify=True)
    wb["Sheet1"].add_table(_simple_table("T1", "A1:A3", ["X"]))
    wb["Sheet2"].add_table(_simple_table("T2", "A1:A3", ["Y"]))
    wb.save(out)
    wb.close()

    names = _zip_namelist(out)
    assert "xl/tables/table1.xml" in names
    assert "xl/tables/table2.xml" in names

    t1_xml = _read_zip_text(out, "xl/tables/table1.xml")
    t2_xml = _read_zip_text(out, "xl/tables/table2.xml")
    assert 'id="1"' in t1_xml
    assert 'id="2"' in t2_xml
    # Each sheet's rels file references its own table
    s1_rels = _read_zip_text(out, "xl/worksheets/_rels/sheet1.xml.rels")
    s2_rels = _read_zip_text(out, "xl/worksheets/_rels/sheet2.xml.rels")
    assert "../tables/table1.xml" in s1_rels
    assert "../tables/table2.xml" in s2_rels
    assert "../tables/table2.xml" not in s1_rels
    assert "../tables/table1.xml" not in s2_rels
