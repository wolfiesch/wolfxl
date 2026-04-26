"""Sprint Ι Pod-γ — password-protected reads via msoffcrypto-tool.

These tests cover the ``password=`` kwarg on
:func:`wolfxl.load_workbook`. The encrypted fixture is synthesized
once per session by encrypting ``tests/fixtures/minimal.xlsx`` with a
hardcoded test password (``wolfxl-test-pw``). The fixture lives in
the pytest tmp dir so we don't commit encrypted bytes to the repo.

Test password rationale: hardcoded in source, no real-world value.
The fixture is regenerated from a plaintext one every session, so
rotating the password is trivial.
"""

from __future__ import annotations

import io
from pathlib import Path

import pytest

import wolfxl

TEST_PASSWORD = "wolfxl-test-pw"


# ---------------------------------------------------------------------------
# Fixture synthesis — encrypts minimal.xlsx in-place (per session).
# ---------------------------------------------------------------------------


@pytest.fixture(scope="session")
def plaintext_xlsx(tmp_path_factory: pytest.TempPathFactory) -> Path:
    """A small but >4 KiB plaintext xlsx, used as the encryption input.

    Synthesised via ``openpyxl`` so we don't depend on the wolfxl
    writer for this fixture's correctness. msoffcrypto-tool 5.4.x has
    a known bug where round-tripping an xlsx whose payload fits in a
    single OLE mini-FAT sector (≤4 KiB) fails to decrypt — the OLE
    writer threads the EncryptedPackage stream into the mini-FAT
    incorrectly. Our fixture is sized just past that threshold so the
    EncryptedPackage lands in the regular FAT and round-trips cleanly.
    """
    import openpyxl

    out_dir = tmp_path_factory.mktemp("plaintext")
    out = out_dir / "fixture.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    # Pack enough cells to push the zip past 4 KiB.
    ws["A1"] = "secret-cell"
    ws["B1"] = 42
    ws["C1"] = 3.14
    for r in range(1, 200):
        for c in range(1, 10):
            ws.cell(row=r, column=c, value=f"r{r}c{c}")
    wb.save(out)
    return out


@pytest.fixture(scope="session")
def encrypted_xlsx(
    plaintext_xlsx: Path, tmp_path_factory: pytest.TempPathFactory
) -> Path:
    """A password-protected copy of the plaintext fixture.

    Synthesized at collection time via ``msoffcrypto-tool``'s
    ``OOXMLFile.encrypt`` API. Requires the optional ``encrypted``
    extra to be installed in the test venv (``uv add
    msoffcrypto-tool``).
    """
    pytest.importorskip("msoffcrypto")
    from msoffcrypto.format.ooxml import OOXMLFile

    out_dir = tmp_path_factory.mktemp("encrypted")
    encrypted = out_dir / "encrypted_secret.xlsx"
    with open(plaintext_xlsx, "rb") as src:
        of = OOXMLFile(src)
        with open(encrypted, "wb") as dst:
            of.encrypt(TEST_PASSWORD, dst)
    return encrypted


# ---------------------------------------------------------------------------
# Happy path
# ---------------------------------------------------------------------------


def test_password_read_unlocks_encrypted_workbook(encrypted_xlsx: Path) -> None:
    """Encrypted file + correct password → values surface."""
    wb = wolfxl.load_workbook(encrypted_xlsx, password=TEST_PASSWORD)
    try:
        # minimal.xlsx has at least one sheet — confirm we can access it.
        assert len(wb.sheetnames) >= 1
        ws = wb[wb.sheetnames[0]]
        # Probe a cell — mostly just asserting the read pipeline works.
        # minimal.xlsx is a basic fixture; we don't care which cells
        # have values, only that the API doesn't blow up on a
        # post-decrypt read.
        _ = ws["A1"].value
    finally:
        wb.close()


def test_password_accepts_bytes(encrypted_xlsx: Path) -> None:
    """Password may be passed as bytes; matches openpyxl signature."""
    wb = wolfxl.load_workbook(encrypted_xlsx, password=TEST_PASSWORD.encode("utf-8"))
    try:
        assert len(wb.sheetnames) >= 1
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Error paths
# ---------------------------------------------------------------------------


def test_wrong_password_raises_value_error(encrypted_xlsx: Path) -> None:
    """Wrong password surfaces as ``ValueError`` with a clear message."""
    with pytest.raises(ValueError) as excinfo:
        wolfxl.load_workbook(encrypted_xlsx, password="not-the-password")
    msg = str(excinfo.value).lower()
    assert "decrypt" in msg or "key" in msg


def test_missing_password_on_encrypted_file_raises(encrypted_xlsx: Path) -> None:
    """Encrypted file without ``password=`` produces a useful error.

    Without the password kwarg, wolfxl tries the normal reader path,
    which fails because the file is an OOXML CFB-wrapped blob, not a
    plain xlsx zip. The error need not be friendly here — we just want
    it to fail rather than silently produce a 0-sheet workbook.
    """
    with pytest.raises(Exception):
        wolfxl.load_workbook(encrypted_xlsx)


def test_password_on_plain_file_is_ignored(plaintext_xlsx: Path) -> None:
    """A password on a non-encrypted file silently passes through.

    Matches openpyxl's behaviour — a defensive caller can always pass
    ``password=`` and have it work whether or not the file actually
    needs decryption.
    """
    wb = wolfxl.load_workbook(plaintext_xlsx, password="ignored")
    try:
        assert len(wb.sheetnames) >= 1
    finally:
        wb.close()


def test_missing_msoffcrypto_raises_friendly_import_error(
    plaintext_xlsx: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """When ``msoffcrypto-tool`` isn't installed, ``ImportError`` points
    at the install hint."""
    import builtins

    real_import = builtins.__import__

    def fake_import(name: str, *a: object, **kw: object) -> object:
        if name == "msoffcrypto":
            raise ImportError("msoffcrypto not available (simulated)")
        return real_import(name, *a, **kw)  # type: ignore[arg-type]

    monkeypatch.setattr(builtins, "__import__", fake_import)
    with pytest.raises(ImportError) as excinfo:
        wolfxl.load_workbook(plaintext_xlsx, password="anything")
    assert "wolfxl[encrypted]" in str(excinfo.value)


# ---------------------------------------------------------------------------
# Mode interactions
# ---------------------------------------------------------------------------


def test_password_with_data_only(encrypted_xlsx: Path) -> None:
    """``data_only=True`` + ``password=`` plumbs through unchanged."""
    wb = wolfxl.load_workbook(encrypted_xlsx, password=TEST_PASSWORD, data_only=True)
    try:
        assert wb._data_only is True  # noqa: SLF001
        assert len(wb.sheetnames) >= 1
    finally:
        wb.close()


def test_password_with_modify_round_trip(
    encrypted_xlsx: Path, tmp_path: Path
) -> None:
    """``modify=True`` + ``password=`` works; saved output is plaintext.

    Sprint Ι Pod-γ ships modify-mode + password support because the
    decrypted bytes are materialised to a tempfile that the patcher
    can reopen. Saving produces a plaintext xlsx (write-side
    encryption is T3 out-of-scope).
    """
    wb = wolfxl.load_workbook(encrypted_xlsx, password=TEST_PASSWORD, modify=True)
    try:
        out = tmp_path / "round_trip.xlsx"
        wb.save(out)
    finally:
        wb.close()

    # Output is a plain xlsx (zipfile-readable, not OOXML-encrypted).
    import zipfile

    assert zipfile.is_zipfile(out)
    # Should re-open as a plain workbook (no password needed).
    wb2 = wolfxl.load_workbook(out)
    try:
        assert len(wb2.sheetnames) >= 1
    finally:
        wb2.close()


def test_save_with_password_kwarg_raises_not_implemented() -> None:
    """``wb.save(path, password=...)`` is reserved; today raises."""
    wb = wolfxl.Workbook()
    try:
        with pytest.raises(NotImplementedError) as excinfo:
            wb.save("ignored.xlsx", password="x")
        assert "encryption" in str(excinfo.value).lower()
    finally:
        wb.close()
