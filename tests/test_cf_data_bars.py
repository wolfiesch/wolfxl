"""G12 (Sprint 3) — DataBarRule cfvo edge cases.

Exercises the non-min/max thresholds (``num`` / ``percent`` / ``percentile``
/ ``formula``) plus ``showValue=False`` round-tripping through wolfxl's
write path. Each test saves with wolfxl, reloads with openpyxl, and asserts
the cfvo types and values landed in the OOXML correctly.

Companion to the ``cf_data_bars_advanced`` oracle probe in
``tests/test_openpyxl_compat_oracle.py``.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

import wolfxl
from wolfxl.formatting.rule import DataBarRule


def _databars_after_openpyxl(path: Path) -> list:
    """Return all dataBar rules found by openpyxl in the first sheet."""
    wb = openpyxl.load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    rules: list = []
    for cf_range in ws.conditional_formatting:
        rules.extend(ws.conditional_formatting[cf_range])
    return [r for r in rules if getattr(r, "type", "") == "dataBar"]


def _populate_column(ws, n: int = 10) -> None:
    for i in range(1, n + 1):
        ws.cell(row=i, column=1, value=i * 10)


def test_databar_min_max_still_works(tmp_path: Path) -> None:
    """Regression guard: the basic min/max DataBarRule still round-trips."""
    wb = wolfxl.Workbook()
    ws = wb.active
    _populate_column(ws)
    rule = DataBarRule(start_type="min", end_type="max", color="FF638EC6")
    ws.conditional_formatting.add("A1:A10", rule)
    out = tmp_path / "databar_min_max.xlsx"
    wb.save(out)

    bars = _databars_after_openpyxl(out)
    assert bars, "openpyxl saw no dataBar rule"
    cfvo = bars[0].dataBar.cfvo
    types = {c.type for c in cfvo}
    assert types == {"min", "max"}, f"cfvo types lost: {types}"


def test_databar_percent_thresholds(tmp_path: Path) -> None:
    """percent/percent cfvo round-trip through wolfxl write + openpyxl reload."""
    wb = wolfxl.Workbook()
    ws = wb.active
    _populate_column(ws)
    rule = DataBarRule(
        start_type="percent",
        start_value=20,
        end_type="percent",
        end_value=80,
        color="FF638EC6",
    )
    ws.conditional_formatting.add("A1:A10", rule)
    out = tmp_path / "databar_percent.xlsx"
    wb.save(out)

    bars = _databars_after_openpyxl(out)
    assert bars
    cfvo = bars[0].dataBar.cfvo
    assert {c.type for c in cfvo} == {"percent"}
    assert float(cfvo[0].val) == 20.0
    assert float(cfvo[1].val) == 80.0
    # Default showValue is True; the attribute is omitted on the dataBar
    # element, so openpyxl reports it as None (its sentinel for the OOXML
    # default).
    assert bars[0].dataBar.showValue in (None, True)


def test_databar_num_thresholds(tmp_path: Path) -> None:
    """Numeric (num) cfvo round-trip."""
    wb = wolfxl.Workbook()
    ws = wb.active
    _populate_column(ws)
    rule = DataBarRule(
        start_type="num",
        start_value=15,
        end_type="num",
        end_value=85,
        color="FFFFB300",
    )
    ws.conditional_formatting.add("A1:A10", rule)
    out = tmp_path / "databar_num.xlsx"
    wb.save(out)

    bars = _databars_after_openpyxl(out)
    assert bars
    cfvo = bars[0].dataBar.cfvo
    assert {c.type for c in cfvo} == {"num"}
    assert float(cfvo[0].val) == 15.0
    assert float(cfvo[1].val) == 85.0


def test_databar_formula_thresholds(tmp_path: Path) -> None:
    """Formula cfvo (start) paired with a numeric cfvo (end)."""
    wb = wolfxl.Workbook()
    ws = wb.active
    _populate_column(ws)
    # Seed Z1/Z2 so the formulas have a target.
    ws["Z1"] = 5
    ws["Z2"] = 95
    rule = DataBarRule(
        start_type="formula",
        start_value="$Z$1",
        end_type="formula",
        end_value="$Z$2",
        color="FF63BE7B",
    )
    ws.conditional_formatting.add("A1:A10", rule)
    out = tmp_path / "databar_formula.xlsx"
    wb.save(out)

    bars = _databars_after_openpyxl(out)
    assert bars
    cfvo = bars[0].dataBar.cfvo
    assert {c.type for c in cfvo} == {"formula"}
    assert cfvo[0].val == "$Z$1"
    assert cfvo[1].val == "$Z$2"


def test_databar_show_value_false(tmp_path: Path) -> None:
    """showValue=False emits showValue=\"0\" on the dataBar element."""
    wb = wolfxl.Workbook()
    ws = wb.active
    _populate_column(ws)
    rule = DataBarRule(
        start_type="min",
        end_type="max",
        color="FF638EC6",
        showValue=False,
    )
    ws.conditional_formatting.add("A1:A10", rule)
    out = tmp_path / "databar_no_value.xlsx"
    wb.save(out)

    bars = _databars_after_openpyxl(out)
    assert bars
    bar = bars[0].dataBar
    assert bar.showValue is False
    # cfvo types preserved alongside the showValue toggle.
    assert {c.type for c in bar.cfvo} == {"min", "max"}


def test_databar_mixed_percent_num_with_show_value(tmp_path: Path) -> None:
    """Mixed start/end types + showValue=False — the canonical G12 fixture."""
    wb = wolfxl.Workbook()
    ws = wb.active
    _populate_column(ws)
    rule = DataBarRule(
        start_type="percent",
        start_value=10,
        end_type="num",
        end_value=90,
        color="FF638EC6",
        showValue=False,
    )
    ws.conditional_formatting.add("A1:A10", rule)
    out = tmp_path / "databar_advanced.xlsx"
    wb.save(out)

    bars = _databars_after_openpyxl(out)
    assert bars
    bar = bars[0].dataBar
    assert {c.type for c in bar.cfvo} == {"percent", "num"}
    cfvo_min, cfvo_max = bar.cfvo[0], bar.cfvo[1]
    assert float(cfvo_min.val) == 10.0
    assert float(cfvo_max.val) == 90.0
    assert bar.showValue is False


def test_databar_wolfxl_roundtrip_preserves_type(tmp_path: Path) -> None:
    """wolfxl save → wolfxl load also preserves the dataBar rule type."""
    wb = wolfxl.Workbook()
    ws = wb.active
    _populate_column(ws)
    rule = DataBarRule(
        start_type="percent",
        start_value=25,
        end_type="num",
        end_value=75,
        color="FFAA0088",
        showValue=False,
    )
    ws.conditional_formatting.add("A1:A10", rule)
    out = tmp_path / "databar_wolfxl_rt.xlsx"
    wb.save(out)

    wb2 = wolfxl.load_workbook(out)
    rules = []
    for cf_range in wb2.active.conditional_formatting:
        rules.extend(cf_range.rules if hasattr(cf_range, "rules") else [])
    assert any(getattr(r, "type", "") == "dataBar" for r in rules)


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
