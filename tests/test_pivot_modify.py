"""Sprint Ν Pod-γ (RFC-047 / RFC-048) — modify-mode pivot table coverage.

The patcher's Phase 2.5m drains pending pivot caches + tables into:
- ``xl/pivotCache/pivotCacheDefinitionN.xml`` + records
- ``xl/pivotTables/pivotTableN.xml``
- Per-cache rels (definition → records)
- Per-table rels (table → cache)
- Workbook rels (workbook → cache definition)
- Sheet rels (sheet → pivot table)
- ``[Content_Types].xml`` overrides for all four part types
- ``<pivotCaches>`` block spliced into ``xl/workbook.xml``

These tests exercise the public API
``Workbook.add_pivot_cache(cache) + Worksheet.add_pivot_table(pt)``
through ``Workbook.save()`` and verify the resulting ZIP shape.
"""
from __future__ import annotations

import re
import zipfile
from pathlib import Path

import openpyxl
import pytest

from wolfxl import load_workbook
from wolfxl.chart.reference import Reference
from wolfxl.pivot import PivotCache, PivotTable


@pytest.fixture(autouse=True)
def _force_test_epoch(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("WOLFXL_TEST_EPOCH", "0")


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


def _make_pivot_source_fixture(
    path: Path, sheet_title: str = "Data"
) -> None:
    """4 cols × 5 rows mini source (1 header + 4 data)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    rows = [
        ("region", "quarter", "customer", "revenue"),
        ("North", "Q1", "Acme", 100.0),
        ("South", "Q1", "Acme", 200.0),
        ("North", "Q2", "Globex", 150.0),
        ("South", "Q2", "Globex", 250.0),
    ]
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(r_idx, c_idx, val)
    wb.save(path)


def _zip_listing(path: Path) -> list[str]:
    with zipfile.ZipFile(path, "r") as z:
        return sorted(z.namelist())


def _zip_read(path: Path, member: str) -> bytes:
    with zipfile.ZipFile(path, "r") as z:
        return z.read(member)


def _build_cache_and_table(wb_workbook, sheet_title: str = "Data"):
    ws = wb_workbook[sheet_title]
    src = Reference(
        worksheet=ws, min_col=1, min_row=1, max_col=4, max_row=5
    )
    cache = PivotCache(source=src)
    pt = PivotTable(
        cache=cache,
        location="F2",
        rows=["region"],
        cols=["quarter"],
        data=["revenue"],
    )
    return cache, pt


# ---------------------------------------------------------------------------
# Case 1 — End-to-end: add cache + table, save, verify ZIP shape.
# ---------------------------------------------------------------------------


def test_add_pivot_emits_all_parts(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_pivot_source_fixture(src)
    wb = load_workbook(src, modify=True)
    cache, pt = _build_cache_and_table(wb)
    wb.add_pivot_cache(cache)
    wb["Data"].add_pivot_table(pt)
    wb.save(dst)

    entries = _zip_listing(dst)
    cache_defs = [e for e in entries if e.startswith("xl/pivotCache/pivotCacheDefinition")]
    cache_recs = [e for e in entries if e.startswith("xl/pivotCache/pivotCacheRecords")]
    tables = [
        e for e in entries
        if re.match(r"^xl/pivotTables/pivotTable\d+\.xml$", e)
    ]
    assert cache_defs, f"no pivotCacheDefinition emitted: {entries}"
    assert cache_recs, f"no pivotCacheRecords emitted: {entries}"
    assert tables, f"no pivotTable emitted: {entries}"


def test_add_pivot_emits_content_type_overrides(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_pivot_source_fixture(src)
    wb = load_workbook(src, modify=True)
    cache, pt = _build_cache_and_table(wb)
    wb.add_pivot_cache(cache)
    wb["Data"].add_pivot_table(pt)
    wb.save(dst)

    ct = _zip_read(dst, "[Content_Types].xml").decode("utf-8")
    assert "pivotCacheDefinition+xml" in ct
    assert "pivotCacheRecords+xml" in ct
    assert "pivotTable+xml" in ct


def test_add_pivot_emits_workbook_rel_and_pivotcaches_block(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_pivot_source_fixture(src)
    wb = load_workbook(src, modify=True)
    cache, pt = _build_cache_and_table(wb)
    wb.add_pivot_cache(cache)
    wb["Data"].add_pivot_table(pt)
    wb.save(dst)

    wb_rels = _zip_read(dst, "xl/_rels/workbook.xml.rels").decode("utf-8")
    assert "pivotCacheDefinition" in wb_rels
    wb_xml = _zip_read(dst, "xl/workbook.xml").decode("utf-8")
    assert "<pivotCaches>" in wb_xml
    assert "<pivotCache " in wb_xml
    assert 'cacheId="0"' in wb_xml


def test_add_pivot_emits_sheet_rel_to_table(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_pivot_source_fixture(src)
    wb = load_workbook(src, modify=True)
    cache, pt = _build_cache_and_table(wb)
    wb.add_pivot_cache(cache)
    wb["Data"].add_pivot_table(pt)
    wb.save(dst)

    sheet_rels = _zip_read(
        dst, "xl/worksheets/_rels/sheet1.xml.rels"
    ).decode("utf-8")
    assert "pivotTable" in sheet_rels


def test_add_pivot_emits_per_cache_and_per_table_rels(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_pivot_source_fixture(src)
    wb = load_workbook(src, modify=True)
    cache, pt = _build_cache_and_table(wb)
    wb.add_pivot_cache(cache)
    wb["Data"].add_pivot_table(pt)
    wb.save(dst)

    entries = _zip_listing(dst)
    cache_rels = [
        e for e in entries
        if e.startswith("xl/pivotCache/_rels/pivotCacheDefinition")
        and e.endswith(".xml.rels")
    ]
    table_rels = [
        e for e in entries
        if e.startswith("xl/pivotTables/_rels/pivotTable")
        and e.endswith(".xml.rels")
    ]
    assert cache_rels, f"no per-cache rels: {entries}"
    assert table_rels, f"no per-table rels: {entries}"

    # Cache rels should point at records.
    cr = _zip_read(dst, cache_rels[0]).decode("utf-8")
    assert "pivotCacheRecords" in cr

    # Table rels should point at cache definition.
    tr = _zip_read(dst, table_rels[0]).decode("utf-8")
    assert "pivotCacheDefinition" in tr


# ---------------------------------------------------------------------------
# Case 2 — Multiple tables sharing one cache.
# ---------------------------------------------------------------------------


def test_two_tables_share_one_cache(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_pivot_source_fixture(src)
    wb = load_workbook(src, modify=True)
    ws = wb["Data"]
    refsrc = Reference(worksheet=ws, min_col=1, min_row=1, max_col=4, max_row=5)
    cache = PivotCache(source=refsrc)
    pt1 = PivotTable(
        cache=cache, location="F2", rows=["region"], data=["revenue"],
        name="P1",
    )
    pt2 = PivotTable(
        cache=cache, location="F12", rows=["customer"], data=["revenue"],
        name="P2",
    )
    wb.add_pivot_cache(cache)
    ws.add_pivot_table(pt1)
    ws.add_pivot_table(pt2)
    wb.save(dst)

    entries = _zip_listing(dst)
    cache_defs = [
        e for e in entries
        if re.match(r"^xl/pivotCache/pivotCacheDefinition\d+\.xml$", e)
    ]
    tables = [
        e for e in entries
        if re.match(r"^xl/pivotTables/pivotTable\d+\.xml$", e)
    ]
    assert len(cache_defs) == 1, f"expected 1 cache, got {cache_defs}"
    assert len(tables) == 2, f"expected 2 tables, got {tables}"


# ---------------------------------------------------------------------------
# Case 3 — Idempotency: re-saving an emitted file is a no-op (modify
# mode passes pivot parts through verbatim when nothing's queued).
# ---------------------------------------------------------------------------


def test_resave_pivot_workbook_is_stable(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    mid = tmp_path / "mid.xlsx"
    out = tmp_path / "out.xlsx"
    _make_pivot_source_fixture(src)
    wb = load_workbook(src, modify=True)
    cache, pt = _build_cache_and_table(wb)
    wb.add_pivot_cache(cache)
    wb["Data"].add_pivot_table(pt)
    wb.save(mid)

    # Reopen the just-saved workbook in modify mode and re-save
    # without queuing anything new.
    wb2 = load_workbook(mid, modify=True)
    wb2.save(out)

    # Pivot parts must survive the no-op save.
    entries = _zip_listing(out)
    assert any(
        e.startswith("xl/pivotCache/pivotCacheDefinition")
        for e in entries
    ), f"pivot cache lost on re-save: {entries}"
    assert any(
        e.startswith("xl/pivotTables/pivotTable") for e in entries
    ), f"pivot table lost on re-save: {entries}"


# ---------------------------------------------------------------------------
# Case 4 — Negative paths: validate guard rails on the public API.
# ---------------------------------------------------------------------------


def test_add_pivot_table_without_registered_cache_raises(
    tmp_path: Path,
) -> None:
    src = tmp_path / "src.xlsx"
    _make_pivot_source_fixture(src)
    wb = load_workbook(src, modify=True)
    cache, pt = _build_cache_and_table(wb)
    # Skip wb.add_pivot_cache(cache) on purpose.
    with pytest.raises(ValueError, match="not been registered"):
        wb["Data"].add_pivot_table(pt)


def test_add_pivot_cache_twice_raises(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    _make_pivot_source_fixture(src)
    wb = load_workbook(src, modify=True)
    cache, _pt = _build_cache_and_table(wb)
    wb.add_pivot_cache(cache)
    with pytest.raises(ValueError, match="already registered"):
        wb.add_pivot_cache(cache)


def test_add_pivot_cache_in_write_mode_raises(tmp_path: Path) -> None:
    from wolfxl import Workbook

    wb = Workbook()
    # Add a worksheet with the data so the Reference resolves.
    ws = wb.active
    ws.title = "Data"
    for r_idx, row in enumerate(
        [
            ("region", "revenue"),
            ("N", 1.0),
            ("S", 2.0),
        ],
        start=1,
    ):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(r_idx, c_idx, val)
    refsrc = Reference(worksheet=ws, min_col=1, min_row=1, max_col=2, max_row=3)
    cache = PivotCache(source=refsrc)
    with pytest.raises(RuntimeError, match="modify mode"):
        wb.add_pivot_cache(cache)


def test_add_pivot_table_in_write_mode_raises(tmp_path: Path) -> None:
    from wolfxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    for r_idx, row in enumerate(
        [
            ("region", "revenue"),
            ("N", 1.0),
            ("S", 2.0),
        ],
        start=1,
    ):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(r_idx, c_idx, val)
    refsrc = Reference(worksheet=ws, min_col=1, min_row=1, max_col=2, max_row=3)
    cache = PivotCache(source=refsrc)
    cache._cache_id = 0
    cache._materialize(ws)
    pt = PivotTable(
        cache=cache, location="F2", rows=["region"], data=["revenue"],
    )
    with pytest.raises(RuntimeError, match="modify mode"):
        ws.add_pivot_table(pt)


# ---------------------------------------------------------------------------
# Case 5 — Phase ordering: verify charts are added BEFORE pivots
# (Phase 2.5l → 2.5m). This is a behavioural test; a workbook with
# both a chart-add and a pivot-add should produce both parts cleanly.
# ---------------------------------------------------------------------------


_CHART_TMPL = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<c:chartSpace xmlns:c="http://schemas.openxmlformats.org/drawingml/2006/chart" xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">
<c:chart><c:plotArea>
<c:barChart><c:barDir val="col"/><c:grouping val="clustered"/>
<c:ser><c:idx val="0"/><c:order val="0"/>
<c:cat><c:numRef><c:f>'Data'!$A$1:$A$4</c:f></c:numRef></c:cat>
<c:val><c:numRef><c:f>'Data'!$D$1:$D$4</c:f></c:numRef></c:val>
</c:ser>
<c:axId val="1"/><c:axId val="2"/>
</c:barChart>
<c:catAx><c:axId val="1"/><c:scaling><c:orientation val="minMax"/></c:scaling><c:crossAx val="2"/></c:catAx>
<c:valAx><c:axId val="2"/><c:scaling><c:orientation val="minMax"/></c:scaling><c:crossAx val="1"/></c:valAx>
</c:plotArea></c:chart></c:chartSpace>"""


def test_chart_and_pivot_in_same_save(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_pivot_source_fixture(src)
    wb = load_workbook(src, modify=True)
    cache, pt = _build_cache_and_table(wb)
    wb.add_pivot_cache(cache)
    wb["Data"].add_pivot_table(pt)
    wb.add_chart_modify_mode("Data", _CHART_TMPL.encode("utf-8"), "K2")
    wb.save(dst)

    entries = _zip_listing(dst)
    assert any(e.startswith("xl/charts/chart") for e in entries), entries
    assert any(
        e.startswith("xl/pivotCache/pivotCacheDefinition")
        for e in entries
    ), entries
    assert any(
        e.startswith("xl/pivotTables/pivotTable") for e in entries
    ), entries
