"""Sprint Μ Pod-δ (RFC-046 §8) — write-mode chart coverage.

Verifies that ``Worksheet.add_chart(BarChart(...))`` and friends emit
structurally-valid xlsx with:

* ``xl/charts/chartN.xml`` (per-type root element)
* ``xl/charts/_rels/chartN.xml.rels`` (when needed)
* ``xl/drawings/drawingN.xml`` (twoCellAnchor / oneCellAnchor)
* ``xl/drawings/_rels/drawingN.xml.rels`` (chart rel)
* sheet rels carry the drawing rel
* ``[Content_Types].xml`` carries the chart override

Most assertions are XML-shape only (presence of named elements at the
right path) so the suite stays robust against attribute-order /
whitespace shifts in the writer crate.

These tests **will fail until Pods α + β land**: at that point
``wolfxl.chart`` exposes real classes (currently stubs) and
``Worksheet.add_chart`` is a real pymethod (currently absent).
"""

from __future__ import annotations

import zipfile
from pathlib import Path
from typing import Any

import pytest

import wolfxl

# All chart construction depends on Pod-β's API. Skip the entire module
# if the chart shim is still in stub mode (pre-merge).
try:
    from wolfxl.chart import (  # type: ignore[attr-defined]
        AreaChart,
        BarChart,
        BubbleChart,
        DoughnutChart,
        LineChart,
        PieChart,
        RadarChart,
        Reference,
        ScatterChart,
    )
    _CHART_API_AVAILABLE = True
except (ImportError, AttributeError, NotImplementedError):
    _CHART_API_AVAILABLE = False
    BarChart = LineChart = PieChart = DoughnutChart = AreaChart = None  # type: ignore[assignment]
    ScatterChart = BubbleChart = RadarChart = Reference = None  # type: ignore[assignment]


pytestmark = pytest.mark.skipif(
    not _CHART_API_AVAILABLE,
    reason="wolfxl.chart construction API ships in Sprint Μ Pods α+β; "
           "tests run once the pods integrate.",
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _seed_data(ws: Any) -> None:
    """Populate ws with the canonical 6×3 data block used by all tests."""
    ws.append(["", "Series A", "Series B"])
    for i in range(1, 6):
        ws.append([f"row{i}", i * 10, i * 5])


def _save_with_chart(
    chart_cls: type,
    *,
    tmp_path: Path,
    title: str | None = None,
    anchor: str = "E2",
    extra: Any = None,
) -> Path:
    """Build a workbook with one chart of the given class and save it."""
    wb = wolfxl.Workbook()
    ws = wb.active
    _seed_data(ws)
    chart = chart_cls()
    if title is not None:
        chart.title = title
    chart.add_data(
        Reference(ws, min_col=2, min_row=1, max_col=3, max_row=6),
        titles_from_data=True,
    )
    if hasattr(chart, "set_categories"):
        try:
            chart.set_categories(
                Reference(ws, min_col=1, min_row=2, max_row=6)
            )
        except Exception:
            pass
    if extra is not None:
        extra(chart)
    ws.add_chart(chart, anchor)
    out = tmp_path / "out.xlsx"
    wb.save(out)
    return out


def _assert_chart_part_present(
    out: Path, *, chart_count: int = 1, sheet: str = "sheet1"
) -> None:
    with zipfile.ZipFile(out) as z:
        names = set(z.namelist())
        for i in range(1, chart_count + 1):
            assert f"xl/charts/chart{i}.xml" in names, (
                f"missing xl/charts/chart{i}.xml in {sorted(names)}"
            )
        # Drawing wrapper
        assert "xl/drawings/drawing1.xml" in names
        # Sheet refers to drawing
        sheet_xml = z.read(f"xl/worksheets/{sheet}.xml").decode()
        assert "<drawing r:id=" in sheet_xml
        # Content types override
        ct = z.read("[Content_Types].xml").decode()
        assert "/xl/charts/chart1.xml" in ct


# ---------------------------------------------------------------------------
# Per-type basic round-trip (8 cases)
# ---------------------------------------------------------------------------


def test_bar_chart_round_trip(tmp_path: Path) -> None:
    out = _save_with_chart(BarChart, tmp_path=tmp_path, title="Bar")
    _assert_chart_part_present(out)
    xml = zipfile.ZipFile(out).read("xl/charts/chart1.xml").decode()
    assert "<c:barChart>" in xml or "barChart" in xml


def test_line_chart_round_trip(tmp_path: Path) -> None:
    out = _save_with_chart(LineChart, tmp_path=tmp_path, title="Line")
    _assert_chart_part_present(out)
    xml = zipfile.ZipFile(out).read("xl/charts/chart1.xml").decode()
    assert "lineChart" in xml


def test_pie_chart_round_trip(tmp_path: Path) -> None:
    out = _save_with_chart(PieChart, tmp_path=tmp_path, title="Pie")
    _assert_chart_part_present(out)
    xml = zipfile.ZipFile(out).read("xl/charts/chart1.xml").decode()
    assert "pieChart" in xml


def test_doughnut_chart_round_trip(tmp_path: Path) -> None:
    out = _save_with_chart(DoughnutChart, tmp_path=tmp_path, title="Donut")
    _assert_chart_part_present(out)
    xml = zipfile.ZipFile(out).read("xl/charts/chart1.xml").decode()
    assert "doughnutChart" in xml


def test_area_chart_round_trip(tmp_path: Path) -> None:
    out = _save_with_chart(AreaChart, tmp_path=tmp_path, title="Area")
    _assert_chart_part_present(out)
    xml = zipfile.ZipFile(out).read("xl/charts/chart1.xml").decode()
    assert "areaChart" in xml


def test_scatter_chart_round_trip(tmp_path: Path) -> None:
    out = _save_with_chart(ScatterChart, tmp_path=tmp_path, title="Scatter")
    _assert_chart_part_present(out)
    xml = zipfile.ZipFile(out).read("xl/charts/chart1.xml").decode()
    assert "scatterChart" in xml


def test_bubble_chart_round_trip(tmp_path: Path) -> None:
    out = _save_with_chart(BubbleChart, tmp_path=tmp_path, title="Bubble")
    _assert_chart_part_present(out)
    xml = zipfile.ZipFile(out).read("xl/charts/chart1.xml").decode()
    assert "bubbleChart" in xml


def test_radar_chart_round_trip(tmp_path: Path) -> None:
    out = _save_with_chart(RadarChart, tmp_path=tmp_path, title="Radar")
    _assert_chart_part_present(out)
    xml = zipfile.ZipFile(out).read("xl/charts/chart1.xml").decode()
    assert "radarChart" in xml


# ---------------------------------------------------------------------------
# Title — string vs rich text
# ---------------------------------------------------------------------------


def test_bar_chart_title_string(tmp_path: Path) -> None:
    out = _save_with_chart(BarChart, tmp_path=tmp_path, title="Q4 Sales")
    xml = zipfile.ZipFile(out).read("xl/charts/chart1.xml").decode()
    assert "Q4 Sales" in xml


def test_line_chart_title_rich_text(tmp_path: Path) -> None:
    """Rich-text title with bold + colored runs round-trips."""
    openpyxl = pytest.importorskip("openpyxl")
    from openpyxl.chart.text import RichText
    from openpyxl.drawing.text import (
        CharacterProperties,
        Paragraph,
        ParagraphProperties,
        RichTextProperties,
        Run,
        SolidColorChoice,
    )

    def _set_rich(c: Any) -> None:
        # Build via openpyxl types — Pod-β re-uses these.
        rp = ParagraphProperties()
        run_a = Run(t="Bold", rPr=CharacterProperties(b=True))
        run_b = Run(t=" Red", rPr=CharacterProperties(
            solidFill=SolidColorChoice(srgbClr="FF0000")
        ))
        para = Paragraph(pPr=rp, r=[run_a, run_b])
        c.title = RichText(bodyPr=RichTextProperties(), p=[para])

    out = _save_with_chart(LineChart, tmp_path=tmp_path, extra=_set_rich)
    xml = zipfile.ZipFile(out).read("xl/charts/chart1.xml").decode()
    assert "Bold" in xml and "Red" in xml
    assert "FF0000" in xml.upper() or "ff0000" in xml


# ---------------------------------------------------------------------------
# Legend position — parametrized over r/l/t/b/tr
# ---------------------------------------------------------------------------


@pytest.mark.parametrize("pos", ["r", "l", "t", "b", "tr"])
def test_bar_chart_legend_position(tmp_path: Path, pos: str) -> None:
    def _set_legend(c: Any) -> None:
        c.legend.position = pos

    out = _save_with_chart(BarChart, tmp_path=tmp_path, extra=_set_legend)
    xml = zipfile.ZipFile(out).read("xl/charts/chart1.xml").decode()
    # Legend block emits <c:legendPos val="r"/> etc.
    assert f'val="{pos}"' in xml


# ---------------------------------------------------------------------------
# Data labels
# ---------------------------------------------------------------------------


def test_pie_chart_data_labels_show_val(tmp_path: Path) -> None:
    from wolfxl.chart.label import DataLabelList  # type: ignore[import]

    def _set_labels(c: Any) -> None:
        c.dataLabels = DataLabelList(showVal=True)

    out = _save_with_chart(PieChart, tmp_path=tmp_path, extra=_set_labels)
    xml = zipfile.ZipFile(out).read("xl/charts/chart1.xml").decode()
    assert "showVal" in xml


def test_pie_chart_data_labels_show_cat_name(tmp_path: Path) -> None:
    from wolfxl.chart.label import DataLabelList  # type: ignore[import]

    def _set_labels(c: Any) -> None:
        c.dataLabels = DataLabelList(showCatName=True)

    out = _save_with_chart(PieChart, tmp_path=tmp_path, extra=_set_labels)
    xml = zipfile.ZipFile(out).read("xl/charts/chart1.xml").decode()
    assert "showCatName" in xml


def test_bar_chart_data_label_position(tmp_path: Path) -> None:
    from wolfxl.chart.label import DataLabelList  # type: ignore[import]

    def _set_labels(c: Any) -> None:
        c.dataLabels = DataLabelList(showVal=True, position="outEnd")

    out = _save_with_chart(BarChart, tmp_path=tmp_path, extra=_set_labels)
    xml = zipfile.ZipFile(out).read("xl/charts/chart1.xml").decode()
    assert "outEnd" in xml


# ---------------------------------------------------------------------------
# Multiple series
# ---------------------------------------------------------------------------


def test_bar_chart_multiple_series_count(tmp_path: Path) -> None:
    """Two-column data block becomes two series."""
    out = _save_with_chart(BarChart, tmp_path=tmp_path)
    # openpyxl reads back; we trust its parser
    openpyxl = pytest.importorskip("openpyxl")
    wb2 = openpyxl.load_workbook(out)
    ws2 = wb2.active
    assert len(ws2._charts) == 1
    assert len(ws2._charts[0].series) == 2


def test_bar_chart_each_series_points_at_correct_range(tmp_path: Path) -> None:
    out = _save_with_chart(BarChart, tmp_path=tmp_path)
    openpyxl = pytest.importorskip("openpyxl")
    wb2 = openpyxl.load_workbook(out)
    chart = wb2.active._charts[0]
    refs = []
    for s in chart.series:
        # Each series carries a numeric data Reference / formula string.
        refs.append(str(s.val.numRef.f) if s.val and s.val.numRef else "")
    # Ranges should reference cols B and C respectively.
    assert any("B" in r for r in refs)
    assert any("C" in r for r in refs)


# ---------------------------------------------------------------------------
# Layout (manual EMU coords)
# ---------------------------------------------------------------------------


def test_bar_chart_manual_layout_emu(tmp_path: Path) -> None:
    from wolfxl.chart.layout import Layout, ManualLayout  # type: ignore[import]

    def _set_layout(c: Any) -> None:
        c.layout = Layout(
            manualLayout=ManualLayout(
                x=0.1, y=0.1, w=0.8, h=0.8,
                xMode="edge", yMode="edge",
            )
        )

    out = _save_with_chart(BarChart, tmp_path=tmp_path, extra=_set_layout)
    xml = zipfile.ZipFile(out).read("xl/charts/chart1.xml").decode()
    assert "manualLayout" in xml


# ---------------------------------------------------------------------------
# Gridlines
# ---------------------------------------------------------------------------


def test_line_chart_major_gridlines(tmp_path: Path) -> None:
    from wolfxl.chart.axis import ChartLines  # type: ignore[import]

    def _set_grid(c: Any) -> None:
        c.y_axis.majorGridlines = ChartLines()

    out = _save_with_chart(LineChart, tmp_path=tmp_path, extra=_set_grid)
    xml = zipfile.ZipFile(out).read("xl/charts/chart1.xml").decode()
    assert "majorGridlines" in xml


def test_line_chart_minor_gridlines(tmp_path: Path) -> None:
    from wolfxl.chart.axis import ChartLines  # type: ignore[import]

    def _set_grid(c: Any) -> None:
        c.y_axis.minorGridlines = ChartLines()

    out = _save_with_chart(LineChart, tmp_path=tmp_path, extra=_set_grid)
    xml = zipfile.ZipFile(out).read("xl/charts/chart1.xml").decode()
    assert "minorGridlines" in xml


# ---------------------------------------------------------------------------
# Marker (Line / Scatter only)
# ---------------------------------------------------------------------------


def test_line_chart_marker_symbol(tmp_path: Path) -> None:
    from wolfxl.chart.marker import Marker  # type: ignore[import]

    def _set_marker(c: Any) -> None:
        for s in c.series:
            s.marker = Marker(symbol="circle", size=7)

    out = _save_with_chart(LineChart, tmp_path=tmp_path, extra=_set_marker)
    xml = zipfile.ZipFile(out).read("xl/charts/chart1.xml").decode()
    assert "circle" in xml
    assert 'val="7"' in xml


def test_scatter_chart_marker_symbol(tmp_path: Path) -> None:
    from wolfxl.chart.marker import Marker  # type: ignore[import]

    def _set_marker(c: Any) -> None:
        for s in c.series:
            s.marker = Marker(symbol="square", size=5)

    out = _save_with_chart(ScatterChart, tmp_path=tmp_path, extra=_set_marker)
    xml = zipfile.ZipFile(out).read("xl/charts/chart1.xml").decode()
    assert "square" in xml


# ---------------------------------------------------------------------------
# Graphical properties on series (line color, fill, dash)
# ---------------------------------------------------------------------------


def test_bar_chart_series_fill_color(tmp_path: Path) -> None:
    from wolfxl.chart.shapes import GraphicalProperties  # type: ignore[import]
    from wolfxl.drawing.fill import ColorChoice  # type: ignore[import]

    def _set_props(c: Any) -> None:
        gp = GraphicalProperties(solidFill=ColorChoice(srgbClr="00FF00"))
        c.series[0].graphicalProperties = gp

    out = _save_with_chart(BarChart, tmp_path=tmp_path, extra=_set_props)
    xml = zipfile.ZipFile(out).read("xl/charts/chart1.xml").decode()
    assert "00FF00" in xml.upper()


def test_line_chart_series_line_dash(tmp_path: Path) -> None:
    from wolfxl.chart.shapes import GraphicalProperties  # type: ignore[import]
    from wolfxl.drawing.line import LineProperties  # type: ignore[import]

    def _set_props(c: Any) -> None:
        gp = GraphicalProperties(ln=LineProperties(prstDash="dash"))
        c.series[0].graphicalProperties = gp

    out = _save_with_chart(LineChart, tmp_path=tmp_path, extra=_set_props)
    xml = zipfile.ZipFile(out).read("xl/charts/chart1.xml").decode()
    assert "dash" in xml


# ---------------------------------------------------------------------------
# Error bars (Bar / Line only)
# ---------------------------------------------------------------------------


def test_bar_chart_error_bars_fixed_val(tmp_path: Path) -> None:
    from wolfxl.chart.error_bar import ErrorBars  # type: ignore[import]

    def _set_eb(c: Any) -> None:
        c.series[0].errBars = ErrorBars(errBarType="both", errValType="fixedVal", val=2)

    out = _save_with_chart(BarChart, tmp_path=tmp_path, extra=_set_eb)
    xml = zipfile.ZipFile(out).read("xl/charts/chart1.xml").decode()
    assert "errBars" in xml
    assert "fixedVal" in xml


def test_line_chart_error_bars_std_err(tmp_path: Path) -> None:
    from wolfxl.chart.error_bar import ErrorBars  # type: ignore[import]

    def _set_eb(c: Any) -> None:
        c.series[0].errBars = ErrorBars(errBarType="plus", errValType="stdErr")

    out = _save_with_chart(LineChart, tmp_path=tmp_path, extra=_set_eb)
    xml = zipfile.ZipFile(out).read("xl/charts/chart1.xml").decode()
    assert "stdErr" in xml


# ---------------------------------------------------------------------------
# Trendline (Line / Scatter only)
# ---------------------------------------------------------------------------


def test_line_chart_trendline_linear(tmp_path: Path) -> None:
    from wolfxl.chart.trendline import Trendline  # type: ignore[import]

    def _set_tl(c: Any) -> None:
        c.series[0].trendline = Trendline(trendlineType="linear")

    out = _save_with_chart(LineChart, tmp_path=tmp_path, extra=_set_tl)
    xml = zipfile.ZipFile(out).read("xl/charts/chart1.xml").decode()
    assert "trendline" in xml.lower()
    assert "linear" in xml


def test_scatter_chart_trendline_polynomial_order_3(tmp_path: Path) -> None:
    from wolfxl.chart.trendline import Trendline  # type: ignore[import]

    def _set_tl(c: Any) -> None:
        c.series[0].trendline = Trendline(trendlineType="poly", order=3)

    out = _save_with_chart(ScatterChart, tmp_path=tmp_path, extra=_set_tl)
    xml = zipfile.ZipFile(out).read("xl/charts/chart1.xml").decode()
    assert "poly" in xml
    assert 'val="3"' in xml


# ---------------------------------------------------------------------------
# Vary colors / smoothing / grouping / scatter style / hole size
# ---------------------------------------------------------------------------


def test_pie_chart_vary_colors_true(tmp_path: Path) -> None:
    def _set_vc(c: Any) -> None:
        c.varyColors = True

    out = _save_with_chart(PieChart, tmp_path=tmp_path, extra=_set_vc)
    xml = zipfile.ZipFile(out).read("xl/charts/chart1.xml").decode()
    assert "varyColors" in xml


def test_doughnut_chart_vary_colors_true(tmp_path: Path) -> None:
    def _set_vc(c: Any) -> None:
        c.varyColors = True

    out = _save_with_chart(DoughnutChart, tmp_path=tmp_path, extra=_set_vc)
    xml = zipfile.ZipFile(out).read("xl/charts/chart1.xml").decode()
    assert "varyColors" in xml


def test_line_chart_smoothing(tmp_path: Path) -> None:
    def _set_smooth(c: Any) -> None:
        for s in c.series:
            s.smooth = True

    out = _save_with_chart(LineChart, tmp_path=tmp_path, extra=_set_smooth)
    xml = zipfile.ZipFile(out).read("xl/charts/chart1.xml").decode()
    assert "smooth" in xml


@pytest.mark.parametrize(
    "grouping", ["clustered", "stacked", "percentStacked"]
)
def test_bar_chart_grouping(tmp_path: Path, grouping: str) -> None:
    def _set_group(c: Any) -> None:
        c.grouping = grouping

    out = _save_with_chart(BarChart, tmp_path=tmp_path, extra=_set_group)
    xml = zipfile.ZipFile(out).read("xl/charts/chart1.xml").decode()
    assert grouping in xml


@pytest.mark.parametrize("style", ["lineMarker", "marker", "smooth"])
def test_scatter_chart_scatter_style(tmp_path: Path, style: str) -> None:
    def _set_style(c: Any) -> None:
        c.scatterStyle = style

    out = _save_with_chart(ScatterChart, tmp_path=tmp_path, extra=_set_style)
    xml = zipfile.ZipFile(out).read("xl/charts/chart1.xml").decode()
    assert style in xml


def test_doughnut_chart_hole_size(tmp_path: Path) -> None:
    def _set_hole(c: Any) -> None:
        c.holeSize = 50

    out = _save_with_chart(DoughnutChart, tmp_path=tmp_path, extra=_set_hole)
    xml = zipfile.ZipFile(out).read("xl/charts/chart1.xml").decode()
    assert "holeSize" in xml


# ---------------------------------------------------------------------------
# Error cases
# ---------------------------------------------------------------------------


def test_invalid_reference_raises() -> None:
    """Reference with an invalid worksheet handle should raise."""
    with pytest.raises((TypeError, ValueError, AttributeError)):
        Reference(None, min_col=1, min_row=1)  # type: ignore[arg-type]


def test_chart_with_no_series_emits_or_raises(tmp_path: Path) -> None:
    """Chart with no series should match openpyxl's behaviour.

    openpyxl emits an empty chart container; wolfxl is allowed to do the
    same OR raise — both are acceptable, just match somebody.
    """
    wb = wolfxl.Workbook()
    ws = wb.active
    _seed_data(ws)
    chart = BarChart()
    out = tmp_path / "out.xlsx"

    # add_chart of an empty chart may raise — that's fine. If it succeeds
    # the saved file should still be openable.
    try:
        ws.add_chart(chart, "E2")
        wb.save(out)
    except (ValueError, RuntimeError):
        pytest.skip("wolfxl raises on empty-series charts (acceptable)")

    openpyxl = pytest.importorskip("openpyxl")
    wb2 = openpyxl.load_workbook(out)  # must not corrupt
    assert wb2.active is not None


def test_unsupported_chart_type_3d_raises() -> None:
    """3D charts deferred to v1.6.1 — must raise NotImplementedError."""
    with pytest.raises((NotImplementedError, ImportError, AttributeError)):
        from wolfxl.chart import BarChart3D  # type: ignore[attr-defined]

        BarChart3D()  # noqa: F841


def test_chart_anchor_invalid_a1_raises(tmp_path: Path) -> None:
    """Anchor strings that aren't valid A1 cell refs should raise."""
    wb = wolfxl.Workbook()
    ws = wb.active
    _seed_data(ws)
    chart = BarChart()
    chart.add_data(
        Reference(ws, min_col=2, min_row=1, max_col=3, max_row=6),
        titles_from_data=True,
    )
    with pytest.raises((ValueError, TypeError)):
        ws.add_chart(chart, "ZZZZZZ123456789")


def test_chart_negative_layout_emu_passes_through(tmp_path: Path) -> None:
    """Excel allows negative layout offsets — we should not block them."""
    from wolfxl.chart.layout import Layout, ManualLayout  # type: ignore[import]

    def _set_layout(c: Any) -> None:
        c.layout = Layout(
            manualLayout=ManualLayout(
                x=-0.05, y=-0.05, w=0.5, h=0.5,
                xMode="edge", yMode="edge",
            )
        )

    out = _save_with_chart(BarChart, tmp_path=tmp_path, extra=_set_layout)
    xml = zipfile.ZipFile(out).read("xl/charts/chart1.xml").decode()
    assert "manualLayout" in xml
    # Negative value is preserved as-is.
    assert "-0.05" in xml or "-5" in xml
