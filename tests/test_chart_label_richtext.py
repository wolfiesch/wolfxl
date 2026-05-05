"""G10 - rich text in chart data labels and axis labels.

Covers the openpyxl-shaped public API for run-level formatting on:

* per-series and per-point ``DataLabelList`` blocks (``txPr`` /
  ``rich`` kwarg accepting a :class:`CellRichText`).
* category- and value-axis titles (``Title.tx.rich`` paragraphs).

Each test does a real wolfxl save and re-reads the workbook with
openpyxl to confirm the run-level properties survive the OOXML round
trip.
"""

from __future__ import annotations

import zipfile
from pathlib import Path

import openpyxl
import pytest

import wolfxl
from wolfxl.cell.rich_text import CellRichText, TextBlock
from wolfxl.cell.text import InlineFont
from wolfxl.chart import BarChart, Reference
from wolfxl.chart.label import DataLabel, DataLabelList
from wolfxl.chart.text import (
    CharacterProperties,
    Paragraph,
    ParagraphProperties,
    RegularTextRun,
    RichText,
    Text,
)
from wolfxl.chart.title import Title


def _seed_workbook() -> wolfxl.Workbook:
    wb = wolfxl.Workbook()
    ws = wb.active
    for row in [["x", "y"], [1, 10], [2, 20], [3, 30]]:
        ws.append(row)
    return wb


def _read_chart_xml(path: Path) -> str:
    with zipfile.ZipFile(path) as z:
        for name in z.namelist():
            if name.startswith("xl/charts/") and name.endswith(".xml") and "rels" not in name:
                return z.read(name).decode()
    raise AssertionError(f"no chart xml in {path}")


def test_data_labels_rich_text_via_kwarg(tmp_path: Path) -> None:
    wb = _seed_workbook()
    ws = wb.active
    chart = BarChart()
    chart.add_data(
        Reference(ws, min_col=2, min_row=1, max_row=4),
        titles_from_data=True,
    )
    rich = CellRichText(
        [
            TextBlock(InlineFont(b=True, sz=14, color="FFFF0000"), "bold-red"),
            "+plain",
        ]
    )
    chart.dataLabels = DataLabelList(rich=rich, showVal=True)
    ws.add_chart(chart, "D2")

    out = tmp_path / "labels.xlsx"
    wb.save(out)

    xml = _read_chart_xml(out)
    assert "<c:dLbls>" in xml
    assert "<c:txPr>" in xml
    assert "<a:t>bold-red</a:t>" in xml
    assert "<a:t>+plain</a:t>" in xml
    assert 'b="1"' in xml
    assert 'sz="1400"' in xml
    assert 'srgbClr val="FF0000"' in xml

    wb2 = openpyxl.load_workbook(out)
    ch = wb2.active._charts[0]
    assert ch.dLbls is not None
    assert ch.dLbls.showVal is True
    runs = ch.dLbls.txPr.p[0].r
    assert [r.t for r in runs] == ["bold-red", "+plain"]
    assert runs[0].rPr.b is True
    assert runs[0].rPr.sz == 1400.0
    assert runs[0].rPr.solidFill.srgbClr == "FF0000"


def test_per_point_data_label_rich_text(tmp_path: Path) -> None:
    wb = _seed_workbook()
    ws = wb.active
    chart = BarChart()
    chart.add_data(
        Reference(ws, min_col=2, min_row=1, max_row=4),
        titles_from_data=True,
    )
    rich = CellRichText([TextBlock(InlineFont(i=True, sz=11), "italic-pt")])
    chart.ser[0].dLbls = DataLabelList(
        dLbl=[DataLabel(idx=0, rich=rich)],
        rich=CellRichText([TextBlock(InlineFont(b=True), "ser-bold")]),
    )
    ws.add_chart(chart, "D2")

    out = tmp_path / "perpoint.xlsx"
    wb.save(out)

    wb2 = openpyxl.load_workbook(out)
    ch = wb2.active._charts[0]
    series = ch.series[0]
    assert series.dLbls is not None
    series_runs = series.dLbls.txPr.p[0].r
    assert series_runs[0].t == "ser-bold"
    assert series_runs[0].rPr.b is True


def test_axis_title_rich_text(tmp_path: Path) -> None:
    wb = _seed_workbook()
    ws = wb.active
    chart = BarChart()
    chart.add_data(
        Reference(ws, min_col=2, min_row=1, max_row=4),
        titles_from_data=True,
    )
    cat_title = Title(
        tx=Text(
            rich=RichText(
                p=[
                    Paragraph(
                        pPr=ParagraphProperties(defRPr=CharacterProperties()),
                        r=[
                            RegularTextRun(
                                rPr=CharacterProperties(
                                    b=True,
                                    sz=1300,
                                    latin="Calibri",
                                    solidFill="0000FF",
                                ),
                                t="Cat Axis",
                            )
                        ],
                    )
                ]
            )
        )
    )
    chart.x_axis.title = cat_title
    val_title = Title(
        tx=Text(
            rich=RichText(
                p=[
                    Paragraph(
                        r=[
                            RegularTextRun(
                                rPr=CharacterProperties(i=True, sz=1100),
                                t="Val Axis",
                            )
                        ]
                    )
                ]
            )
        )
    )
    chart.y_axis.title = val_title
    ws.add_chart(chart, "D2")

    out = tmp_path / "axes.xlsx"
    wb.save(out)

    xml = _read_chart_xml(out)
    assert "<c:catAx>" in xml
    assert "<a:t>Cat Axis</a:t>" in xml
    assert 'sz="1300"' in xml
    assert "<a:t>Val Axis</a:t>" in xml
    assert 'sz="1100"' in xml

    wb2 = openpyxl.load_workbook(out)
    ch = wb2.active._charts[0]
    cat_runs = ch.x_axis.title.tx.rich.p[0].r
    assert cat_runs[0].t == "Cat Axis"
    assert cat_runs[0].rPr.b is True
    assert cat_runs[0].rPr.sz == 1300.0
    assert cat_runs[0].rPr.solidFill.srgbClr == "0000FF"
    assert cat_runs[0].rPr.latin.typeface == "Calibri"

    val_runs = ch.y_axis.title.tx.rich.p[0].r
    assert val_runs[0].rPr.i is True
    assert val_runs[0].t == "Val Axis"


def test_axis_title_string_assignment_still_works(tmp_path: Path) -> None:
    """A bare string assigned to ``axis.title`` keeps the v1.7 behaviour."""
    wb = _seed_workbook()
    ws = wb.active
    chart = BarChart()
    chart.add_data(
        Reference(ws, min_col=2, min_row=1, max_row=4),
        titles_from_data=True,
    )
    chart.x_axis.title = "Plain X"
    chart.y_axis.title = "Plain Y"
    ws.add_chart(chart, "D2")

    out = tmp_path / "plain.xlsx"
    wb.save(out)

    wb2 = openpyxl.load_workbook(out)
    ch = wb2.active._charts[0]
    assert ch.x_axis.title.tx.rich.p[0].r[0].t == "Plain X"
    assert ch.y_axis.title.tx.rich.p[0].r[0].t == "Plain Y"


def test_openpyxl_saved_chart_with_rich_labels_preserved_through_modify(
    tmp_path: Path,
) -> None:
    """An openpyxl workbook with rich-text data labels survives a wolfxl
    modify-mode read+write."""
    from openpyxl.chart import BarChart as OBar
    from openpyxl.chart import Reference as ORef
    from openpyxl.chart.label import DataLabelList as ODll
    from openpyxl.chart.text import RichText as ORichText
    from openpyxl.chart.text import RichTextProperties as OBodyPr
    from openpyxl.drawing.text import (
        CharacterProperties as OCharProps,
    )
    from openpyxl.drawing.text import (
        Paragraph as OParagraph,
    )
    from openpyxl.drawing.text import (
        ParagraphProperties as OParaProps,
    )
    from openpyxl.drawing.text import (
        RegularTextRun as ORun,
    )

    src = tmp_path / "src.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in [["x", "y"], [1, 10], [2, 20]]:
        ws.append(row)
    ch = OBar()
    ch.add_data(ORef(ws, min_col=2, min_row=1, max_row=3), titles_from_data=True)
    para = OParagraph(
        pPr=OParaProps(defRPr=OCharProps()),
        r=[ORun(rPr=OCharProps(b=True, sz=1500), t="OPYxlLbl")],
    )
    ch.dLbls = ODll(txPr=ORichText(bodyPr=OBodyPr(), p=[para]), showVal=True)
    ws.add_chart(ch, "D2")
    wb.save(src)

    dst = tmp_path / "dst.xlsx"
    wb2 = wolfxl.load_workbook(src, modify=True)
    wb2["Sheet"]["A20"] = "modify-marker"
    wb2.save(dst)

    wb3 = openpyxl.load_workbook(dst)
    ch3 = wb3.active._charts[0]
    assert ch3.dLbls is not None
    runs = ch3.dLbls.txPr.p[0].r
    assert runs[0].t == "OPYxlLbl"
    assert runs[0].rPr.b is True
    assert runs[0].rPr.sz == 1500.0


@pytest.mark.parametrize(
    "font_kwargs,expected_sz",
    [
        ({"b": True}, None),
        ({"i": True, "sz": 12}, 1200.0),
        ({"sz": 18, "color": "00FF00"}, 1800.0),
    ],
)
def test_inline_font_field_translation(
    tmp_path: Path, font_kwargs: dict, expected_sz: float | None
) -> None:
    wb = _seed_workbook()
    ws = wb.active
    chart = BarChart()
    chart.add_data(
        Reference(ws, min_col=2, min_row=1, max_row=4),
        titles_from_data=True,
    )
    rich = CellRichText([TextBlock(InlineFont(**font_kwargs), "x")])
    chart.dataLabels = DataLabelList(rich=rich)
    ws.add_chart(chart, "D2")

    out = tmp_path / "p.xlsx"
    wb.save(out)
    wb2 = openpyxl.load_workbook(out)
    ch = wb2.active._charts[0]
    rpr = ch.dLbls.txPr.p[0].r[0].rPr
    if "b" in font_kwargs:
        assert rpr.b is True
    if "i" in font_kwargs:
        assert rpr.i is True
    if expected_sz is not None:
        assert rpr.sz == expected_sz
    if "color" in font_kwargs:
        assert rpr.solidFill.srgbClr == font_kwargs["color"]
