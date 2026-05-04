"""RFC-021 — Defined names mutation round-trip in modify mode.

End-to-end coverage for ``wb.defined_names["X"] = DefinedName(...)`` on
an existing file. The save-time path threads three layers:

1. ``DefinedNameDict.__setitem__`` (Python) routes the entry to
   ``wb._pending_defined_names``. The same machinery already shipped
   for write mode in T1 PR6.
2. ``Workbook._flush_defined_names_to_patcher`` (Python) drains the
   pending dict into ``XlsxPatcher.queue_defined_name``, filtering
   ``None``-valued optional fields so the Rust extractors see a clean
   missing-key signal (mirrors ``_flush_properties_to_patcher``).
3. ``XlsxPatcher::do_save`` Phase 2.5f (Rust) calls
   ``defined_names::merge_defined_names`` on ``xl/workbook.xml``, which
   either splices a fresh ``<definedNames>`` block after ``</sheets>``
   or upserts entries by ``(name, local_sheet_id)`` against the
   existing block. The merged bytes are routed through
   ``file_patches``.

Sister contract: ``test_rfc021_empty_queue_is_no_op`` is the
regression guard for the short-circuit predicate. If a future refactor
forgets to require ``queued_defined_names.is_empty()``, this test
fires. ``test_rfc021_existing_print_area_round_trip`` guards that
``_xlnm.Print_Area`` (and any other built-in name) survives unchanged.
"""

from __future__ import annotations

import zipfile
from pathlib import Path

import openpyxl
import pytest
from wolfxl.workbook.defined_name import DefinedName

from wolfxl import load_workbook

# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture(autouse=True)
def _force_test_epoch(monkeypatch: pytest.MonkeyPatch) -> None:
    """Pin ZIP entry mtimes for byte-stable saves."""
    monkeypatch.setenv("WOLFXL_TEST_EPOCH", "0")


def _make_plain_fixture(path: Path) -> None:
    """Two-sheet workbook with no defined names. Tests the inject-block path."""
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1["A1"] = "header"
    wb.create_sheet("Sheet2")
    wb.save(path)


def _make_fixture_with_names(path: Path) -> None:
    """Workbook with one workbook-scope name + one built-in print area.
    Tests upsert-into-existing-block + built-in preservation."""
    from openpyxl.workbook.defined_name import DefinedName as XDefinedName

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for i in range(1, 11):
        ws[f"A{i}"] = i
    wb.defined_names["Region"] = XDefinedName(
        "Region", attr_text="Sheet1!$A$1:$A$10"
    )
    # Built-in print area (sheet-scope, localSheetId=0).
    wb.defined_names["_xlnm.Print_Area"] = XDefinedName(
        "_xlnm.Print_Area", attr_text="Sheet1!$A$1:$D$20", localSheetId=0
    )
    wb.save(path)


def _read_workbook_xml(path: Path) -> str:
    with zipfile.ZipFile(path) as zf:
        return zf.read("xl/workbook.xml").decode("utf-8")


# ---------------------------------------------------------------------------
# DefinedName Python class — attr_text alias contract (RFC-021 §4.1)
# ---------------------------------------------------------------------------


def test_rfc021_defined_name_accepts_attr_text_alias() -> None:
    dn = DefinedName(name="X", attr_text="Sheet1!$A$1")
    assert dn.value == "Sheet1!$A$1"
    assert dn.attr_text == "Sheet1!$A$1"


def test_rfc021_defined_name_value_and_attr_text_equivalent() -> None:
    a = DefinedName(name="X", value="Sheet1!$A$1")
    b = DefinedName(name="X", attr_text="Sheet1!$A$1")
    assert a == b


def test_rfc021_defined_name_requires_value_or_attr_text() -> None:
    with pytest.raises(TypeError, match="value"):
        DefinedName(name="X")


def test_rfc021_defined_name_rejects_conflicting_value_attr_text() -> None:
    with pytest.raises(TypeError):
        DefinedName(name="X", value="A", attr_text="B")


# ---------------------------------------------------------------------------
# Add a new defined name to a file that had none.
# ---------------------------------------------------------------------------


def test_rfc021_add_new_defined_name_round_trip(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_plain_fixture(src)

    wb = load_workbook(src, modify=True)
    wb.defined_names["Budget"] = DefinedName(
        name="Budget", value="Sheet1!$A$1:$A$100"
    )
    wb.save(dst)

    xml = _read_workbook_xml(dst)
    assert "<definedNames>" in xml
    assert (
        '<definedName name="Budget">Sheet1!$A$1:$A$100</definedName>' in xml
    )
    assert xml.find("</sheets>") < xml.find("<definedNames>")

    rt = openpyxl.load_workbook(dst)
    assert "Budget" in rt.defined_names
    assert "$A$1:$A$100" in rt.defined_names["Budget"].attr_text


def test_rfc021_add_via_attr_text_kwarg(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_plain_fixture(src)

    wb = load_workbook(src, modify=True)
    wb.defined_names["Header"] = DefinedName(
        name="Header", attr_text="Sheet1!$A$1"
    )
    wb.save(dst)

    rt = openpyxl.load_workbook(dst)
    assert rt.defined_names["Header"].attr_text == "Sheet1!$A$1"


# ---------------------------------------------------------------------------
# Update an existing defined name.
# ---------------------------------------------------------------------------


def test_rfc021_update_existing_defined_name(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_fixture_with_names(src)

    wb = load_workbook(src, modify=True)
    wb.defined_names["Region"] = DefinedName(
        name="Region", value="Sheet1!$Z$99"
    )
    wb.save(dst)

    xml = _read_workbook_xml(dst)
    assert "Sheet1!$Z$99" in xml
    assert "Sheet1!$A$1:$A$10" not in xml
    assert xml.count('name="Region"') == 1


# ---------------------------------------------------------------------------
# Round-trip preservation of unrelated existing names.
# ---------------------------------------------------------------------------


def test_rfc021_preserve_existing_names_on_round_trip(tmp_path: Path) -> None:
    """Add a brand-new name; existing Region + Print_Area must survive."""
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_fixture_with_names(src)

    wb = load_workbook(src, modify=True)
    wb.defined_names["Margin"] = DefinedName(name="Margin", value="0.5")
    wb.save(dst)

    xml = _read_workbook_xml(dst)
    assert 'name="Region"' in xml
    assert "Sheet1!$A$1:$A$10" in xml
    assert 'name="_xlnm.Print_Area"' in xml
    assert 'localSheetId="0"' in xml
    assert "Sheet1!$A$1:$D$20" in xml
    assert 'name="Margin"' in xml
    assert xml.count("<definedNames>") == 1
    assert xml.count("</definedNames>") == 1


def test_rfc021_existing_print_area_round_trip(tmp_path: Path) -> None:
    """Built-in name byte-preservation: open + save with no name edits is
    a no-op for workbook.xml — the empty-queue guard kicks in."""
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_fixture_with_names(src)

    wb = load_workbook(src, modify=True)
    wb.save(dst)

    src_xml = _read_workbook_xml(src)
    dst_xml = _read_workbook_xml(dst)
    assert src_xml == dst_xml


# ---------------------------------------------------------------------------
# Sheet-scope (localSheetId) — workbook-scope and sheet-scope coexist.
# ---------------------------------------------------------------------------


def test_rfc021_local_sheet_id_scoping(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_plain_fixture(src)

    wb = load_workbook(src, modify=True)
    wb.defined_names["AllSheets"] = DefinedName(
        name="AllSheets", value="Sheet1!$A$1"
    )
    wb.defined_names["Sheet2Local"] = DefinedName(
        name="Sheet2Local", value="Sheet2!$B$2", localSheetId=1
    )
    wb.save(dst)

    xml = _read_workbook_xml(dst)
    assert (
        '<definedName name="AllSheets">Sheet1!$A$1</definedName>' in xml
    )
    assert (
        '<definedName name="Sheet2Local" localSheetId="1">Sheet2!$B$2</definedName>'
        in xml
    )

    rt = openpyxl.load_workbook(dst)
    assert rt.defined_names["AllSheets"].localSheetId is None
    # openpyxl scopes sheet-local names under the per-sheet defined_names
    # dict, not the workbook-scope one. The Rust side already verified
    # that ``localSheetId="1"`` is in the XML; here we confirm openpyxl
    # routes it to the right sheet.
    sheet1 = rt[rt.sheetnames[1]]
    assert "Sheet2Local" in sheet1.defined_names
    assert "$B$2" in sheet1.defined_names["Sheet2Local"].attr_text


# ---------------------------------------------------------------------------
# Empty queue is a no-op (RFC-021 §6 case 6 + §8 idempotency).
# ---------------------------------------------------------------------------


def test_rfc021_empty_queue_is_no_op(tmp_path: Path) -> None:
    """Open + save with no defined-name edits leaves workbook.xml byte-
    identical to the source."""
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_plain_fixture(src)

    wb = load_workbook(src, modify=True)
    _ = wb.defined_names
    wb.save(dst)

    assert _read_workbook_xml(src) == _read_workbook_xml(dst)


# ---------------------------------------------------------------------------
# Hidden flag emission.
# ---------------------------------------------------------------------------


def test_rfc021_hidden_attribute_emitted(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_plain_fixture(src)

    wb = load_workbook(src, modify=True)
    wb.defined_names["Internal"] = DefinedName(
        name="Internal", value="Sheet1!$A$1", hidden=True
    )
    wb.save(dst)

    xml = _read_workbook_xml(dst)
    assert 'hidden="1"' in xml


# ---------------------------------------------------------------------------
# Mixed save: cell edit + defined-name add in one save.
# ---------------------------------------------------------------------------


def test_rfc021_compose_with_cell_edit(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_plain_fixture(src)

    wb = load_workbook(src, modify=True)
    ws = wb["Sheet1"]
    ws["B2"] = "edited"
    wb.defined_names["Total"] = DefinedName(
        name="Total", value="Sheet1!$A$1:$B$2"
    )
    wb.save(dst)

    rt = load_workbook(dst)
    assert rt["Sheet1"]["B2"].value == "edited"
    rt2 = openpyxl.load_workbook(dst)
    assert "Total" in rt2.defined_names


# ---------------------------------------------------------------------------
# Pending queue is cleared after save.
# ---------------------------------------------------------------------------


def test_rfc021_pending_queue_cleared_after_save(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst1 = tmp_path / "dst1.xlsx"
    dst2 = tmp_path / "dst2.xlsx"
    _make_plain_fixture(src)

    wb = load_workbook(src, modify=True)
    wb.defined_names["Once"] = DefinedName(name="Once", value="Sheet1!$A$1")
    wb.save(dst1)

    assert wb._pending_defined_names == {}  # noqa: SLF001
    wb.save(dst2)

    src_xml = _read_workbook_xml(src)
    dst1_xml = _read_workbook_xml(dst1)
    dst2_xml = _read_workbook_xml(dst2)
    # Both saves write the same Once entry exactly once.
    assert dst2_xml.count('name="Once"') == dst1_xml.count('name="Once"')
    assert dst1_xml.count('name="Once"') == 1
    # Source remains untouched.
    assert src_xml == _read_workbook_xml(src)


# ---------------------------------------------------------------------------
# Phase 2 (G22) — `hidden=False` in modify mode clears an existing
# `hidden="1"`. Regression guard for the bug Codex flagged: the Python
# patcher payload used to omit the key on False, which the Rust upsert
# treated as "preserve source" instead of "remove the attr".
# ---------------------------------------------------------------------------


def test_g22_modify_mode_clears_hidden_when_set_false(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"

    # Seed a workbook with a hidden defined name (open + save in
    # modify mode so the source file genuinely contains hidden="1").
    wb_seed = openpyxl.Workbook()
    ws = wb_seed.active
    ws.title = "Sheet1"
    ws["A1"] = "x"
    wb_seed.save(src)

    seed = load_workbook(src, modify=True)
    seed.defined_names["Helper"] = DefinedName(
        name="Helper", value="Sheet1!$A$1", hidden=True
    )
    seeded = tmp_path / "seeded.xlsx"
    seed.save(seeded)
    assert 'hidden="1"' in _read_workbook_xml(seeded)

    # Now clear hidden via modify mode.
    wb = load_workbook(seeded, modify=True)
    wb.defined_names["Helper"] = DefinedName(
        name="Helper", value="Sheet1!$A$1", hidden=False
    )
    wb.save(dst)

    xml = _read_workbook_xml(dst)
    assert 'name="Helper"' in xml
    # The whole point of this test: hidden="1" must be gone.
    assert 'hidden="1"' not in xml
