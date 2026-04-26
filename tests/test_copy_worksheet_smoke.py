"""RFC-035 — ``Workbook.copy_worksheet`` smoke test.

Minimal end-to-end coverage for the modify-mode sheet-copy path.
Pod-γ replaces this with the full harness described in
``Plans/rfcs/035-copy-worksheet.md`` §6 — these three cases just
prove the wiring works:

1. **Happy path**: load → ``wb.copy_worksheet(wb.active)`` → save →
   reload → assert the destination sheet exists with the same cell
   values as the source.
2. **Write-mode rejection**: ``Workbook().copy_worksheet(...)``
   raises ``NotImplementedError`` per RFC-035 §3 OQ-a.
3. **Cross-workbook rejection**: copying a sheet from a different
   workbook raises ``ValueError``.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

import wolfxl
from wolfxl import Workbook, load_workbook


# pytest marker so verify_rfc.py can collect this test.
pytestmark = pytest.mark.rfc035


@pytest.fixture(autouse=True)
def _force_test_epoch(monkeypatch: pytest.MonkeyPatch) -> None:
    """Pin ZIP entry mtimes for byte-stable saves."""
    monkeypatch.setenv("WOLFXL_TEST_EPOCH", "0")


def _make_simple_fixture(path: Path) -> None:
    """Single-sheet workbook with a small block of cell values."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"
    ws["A1"] = "Header"
    ws["B1"] = "Value"
    ws["A2"] = "alpha"
    ws["B2"] = 10
    ws["A3"] = "beta"
    ws["B3"] = 20
    wb.save(path)


def test_copy_worksheet_happy_path_modify_mode(tmp_path: Path) -> None:
    """Modify-mode ``copy_worksheet`` clones cell values into the new sheet."""
    src = tmp_path / "source.xlsx"
    out = tmp_path / "out.xlsx"
    _make_simple_fixture(src)

    wb = load_workbook(str(src), modify=True)
    new_ws = wb.copy_worksheet(wb.active)
    assert new_ws.title == "Template Copy"
    wb.save(str(out))

    # Re-read with openpyxl: the destination sheet must exist and
    # carry the source's cell values.
    re = openpyxl.load_workbook(out)
    assert "Template" in re.sheetnames
    assert "Template Copy" in re.sheetnames
    src_sheet = re["Template"]
    dst_sheet = re["Template Copy"]
    for coord in ("A1", "B1", "A2", "B2", "A3", "B3"):
        assert dst_sheet[coord].value == src_sheet[coord].value, (
            f"cell {coord} mismatch: src={src_sheet[coord].value!r} "
            f"dst={dst_sheet[coord].value!r}"
        )

    # The output must be a valid xlsx that re-opens without raising.
    assert out.stat().st_size > 0


def test_copy_worksheet_write_mode_raises() -> None:
    """Write-mode workbooks reject copy_worksheet (RFC-035 §3 OQ-a)."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    with pytest.raises(NotImplementedError, match="modify-mode-only"):
        wb.copy_worksheet(ws)


def test_copy_worksheet_cross_workbook_raises(tmp_path: Path) -> None:
    """A sheet from another workbook is rejected with ValueError."""
    src = tmp_path / "source.xlsx"
    _make_simple_fixture(src)

    wb_a = load_workbook(str(src), modify=True)
    wb_b = load_workbook(str(src), modify=True)
    foreign_ws = wb_b.active
    assert foreign_ws is not None
    with pytest.raises(ValueError, match="must belong to this workbook"):
        wb_a.copy_worksheet(foreign_ws)
