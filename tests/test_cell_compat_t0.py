"""T0 compat for Cell attribute parity with openpyxl."""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

import openpyxl
import pytest

import wolfxl
from wolfxl.utils.exceptions import IllegalCharacterError


@pytest.fixture
def mixed_xlsx(tmp_path: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "hello"
    ws["B1"] = 42
    ws["C1"] = True
    ws["D1"] = date(2024, 1, 1)
    ws["E1"] = datetime(2024, 1, 1, 12, 30)
    ws["F1"] = "=SUM(A1:E1)"
    path = tmp_path / "mixed.xlsx"
    wb.save(path)
    return path


def test_column_letter(mixed_xlsx: Path) -> None:
    op_ws = openpyxl.load_workbook(mixed_xlsx)["Sheet1"]
    wx_ws = wolfxl.load_workbook(mixed_xlsx)["Sheet1"]
    for coord in ("A1", "B1", "F1"):
        assert wx_ws[coord].column_letter == op_ws[coord].column_letter


def test_openpyxl_metadata_accessors(mixed_xlsx: Path) -> None:
    op_ws = openpyxl.load_workbook(mixed_xlsx)["Sheet1"]
    wx_ws = wolfxl.load_workbook(mixed_xlsx)["Sheet1"]
    op_cell = op_ws["B1"]
    wx_cell = wx_ws["B1"]
    assert wx_cell.base_date == op_cell.base_date
    assert wx_cell.col_idx == op_cell.col_idx
    assert wx_cell.encoding == op_cell.encoding
    assert wx_cell.internal_value == op_cell.internal_value
    assert wx_cell.pivotButton == op_cell.pivotButton
    assert wx_cell.quotePrefix == op_cell.quotePrefix


def test_openpyxl_string_and_error_helpers() -> None:
    wb = wolfxl.Workbook()
    cell = wb.active["A1"]
    assert cell.check_string(b"abc") == "abc"
    assert len(cell.check_string("x" * 40000)) == 32767
    assert cell.check_error(ValueError("bad")) == "bad"
    with pytest.raises(IllegalCharacterError):
        cell.check_string("bad\x01")


def test_parent_backref(mixed_xlsx: Path) -> None:
    wx_ws = wolfxl.load_workbook(mixed_xlsx)["Sheet1"]
    assert wx_ws["A1"].parent is wx_ws


def test_offset_parity(mixed_xlsx: Path) -> None:
    op_ws = openpyxl.load_workbook(mixed_xlsx)["Sheet1"]
    wx_ws = wolfxl.load_workbook(mixed_xlsx)["Sheet1"]
    for r, c in [(0, 0), (1, 1), (0, 2), (2, 0)]:
        assert (
            wx_ws["A1"].offset(row=r, column=c).coordinate
            == op_ws["A1"].offset(row=r, column=c).coordinate
        )


def test_offset_positional_args() -> None:
    """openpyxl accepts ``offset(1, 1)`` positional — verify signature compat."""
    wb = wolfxl.Workbook()
    ws = wb.active
    assert ws["A1"].offset(1, 1).coordinate == "B2"


def test_data_type_string(mixed_xlsx: Path) -> None:
    wx_ws = wolfxl.load_workbook(mixed_xlsx)["Sheet1"]
    assert wx_ws["A1"].data_type == "s"


def test_data_type_number(mixed_xlsx: Path) -> None:
    wx_ws = wolfxl.load_workbook(mixed_xlsx)["Sheet1"]
    assert wx_ws["B1"].data_type == "n"


def test_data_type_boolean(mixed_xlsx: Path) -> None:
    wx_ws = wolfxl.load_workbook(mixed_xlsx)["Sheet1"]
    assert wx_ws["C1"].data_type == "b"


def test_data_type_date(mixed_xlsx: Path) -> None:
    wx_ws = wolfxl.load_workbook(mixed_xlsx)["Sheet1"]
    assert wx_ws["D1"].data_type == "d"
    assert wx_ws["E1"].data_type == "d"


def test_data_type_formula(mixed_xlsx: Path) -> None:
    wx_ws = wolfxl.load_workbook(mixed_xlsx)["Sheet1"]
    # In read mode without data_only, the cell value is the formula string
    # ("=SUM...") so data_type is "f". This matches openpyxl's convention.
    assert wx_ws["F1"].data_type == "f"


def test_is_date_true_for_date_value(mixed_xlsx: Path) -> None:
    wx_ws = wolfxl.load_workbook(mixed_xlsx)["Sheet1"]
    assert wx_ws["D1"].is_date is True
    assert wx_ws["E1"].is_date is True


def test_is_date_false_for_number(mixed_xlsx: Path) -> None:
    wx_ws = wolfxl.load_workbook(mixed_xlsx)["Sheet1"]
    assert wx_ws["B1"].is_date is False


def test_has_style_unset() -> None:
    wb = wolfxl.Workbook()
    ws = wb.active
    ws["A1"] = "hello"
    assert ws["A1"].has_style is False


def test_has_style_after_font_set() -> None:
    from wolfxl import Font

    wb = wolfxl.Workbook()
    ws = wb.active
    ws["A1"] = "hello"
    ws["A1"].font = Font(bold=True)
    assert ws["A1"].has_style is True
    assert ws["A1"].style_id == 1


def test_style_id_reads_default_and_styled_cells(tmp_path: Path) -> None:
    from openpyxl.styles import Font

    path = tmp_path / "styled.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "plain"
    ws["B1"] = "bold"
    ws["B1"].font = Font(bold=True)
    wb.save(path)

    op_ws = openpyxl.load_workbook(path).active
    wx_ws = wolfxl.load_workbook(path).active
    assert wx_ws["A1"].style_id == op_ws["A1"].style_id == 0
    assert wx_ws["B1"].style_id == op_ws["B1"].style_id


def test_style_getter_returns_none() -> None:
    wb = wolfxl.Workbook()
    ws = wb.active
    ws["A1"] = "hello"
    assert ws["A1"].style is None


def test_style_setter_rejects_unregistered_names() -> None:
    """Assigning a style name that wasn't registered raises ValueError.

    Named-style support landed in the G05 follow-up; the previous
    NotImplementedError was replaced with a registration check so the
    user sees an actionable error pointing them at add_named_style.
    """
    wb = wolfxl.Workbook()
    ws = wb.active
    assert ws is not None
    with pytest.raises(ValueError, match="not registered"):
        ws["A1"].style = "Good"
