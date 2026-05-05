# pyright: reportOptionalSubscript=false, reportOptionalMemberAccess=false
"""Openpyxl-API compatibility oracle (Sprint 0 / Gap G01).

This harness is the program-level scoreboard for ``Plans/openpyxl-parity-program.md``.
It owns one parametrised test (``test_compat_oracle_probe``) that runs a
named probe for every entry in ``docs/migration/_compat_spec.py`` that
carries a ``probe`` field.

Probe-status semantics (mirrors the spec's ``status`` field):

* ``supported``    - probe must pass; failure is a regression.
* ``partial``      - probe is xfail today; flips ``XPASS`` once the gap closes.
* ``not_yet``      - probe is xfail today; flips ``XPASS`` once the gap closes.
* ``out_of_scope`` - probe is skipped with a stable reason.

When a probe flips from xfail to xpass, the workflow is:

1. Confirm the underlying gap is actually closed (read the relevant code).
2. Update ``docs/migration/_compat_spec.py``: change ``status`` from
   ``not_yet`` (or ``partial``) to ``supported``; remove ``gap_id``.
3. Run ``python scripts/render_compat_matrix.py``.
4. Update ``Plans/openpyxl-parity-program.md`` status row (gap → ``landed``).
5. Re-run this oracle. The probe should now pass cleanly (no xfail mark).

The test prints a one-line summary at session end so the per-sprint gate can
read the rolling pass count without parsing pytest output.

S1+ may extend this harness by vendoring openpyxl's source-distribution test
files via ``scripts/fetch_openpyxl_corpus.py`` and running them under a
``sys.modules`` shim. For S0 we deliberately stop at the curated probe set so
the baseline number is meaningful and reproducible without network.
"""

from __future__ import annotations

import importlib.util
import json
import os
import re
import shutil
import zipfile
from collections import Counter
from datetime import date
from pathlib import Path
from typing import Any, Callable

import pytest


REPO_ROOT = Path(__file__).resolve().parents[1]
SPEC_PATH = REPO_ROOT / "docs" / "migration" / "_compat_spec.py"


def _load_spec_module():
    spec = importlib.util.spec_from_file_location("_compat_spec", SPEC_PATH)
    if spec is None or spec.loader is None:  # pragma: no cover - import guard
        raise RuntimeError(f"failed to import {SPEC_PATH}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


_SPEC = _load_spec_module()


# --------------------------------------------------------------------------
# Probe registry
# --------------------------------------------------------------------------
#
# Every probe is a function taking ``tmp_path`` and returning ``None`` on
# success. Failure is signalled by raising any exception (typically
# ``AssertionError``). Probes deliberately stay small - they exercise one
# openpyxl-API contract apiece, not full feature suites - so the harness
# stays cheap to run and the failure messages stay precise.

_ProbeFn = Callable[[Path], None]
PROBES: dict[str, _ProbeFn] = {}


def _register(name: str) -> Callable[[_ProbeFn], _ProbeFn]:
    def decorator(fn: _ProbeFn) -> _ProbeFn:
        PROBES[name] = fn
        return fn

    return decorator


def _zip_listing(path: Path) -> list[str]:
    with zipfile.ZipFile(path, "r") as zf:
        return sorted(zf.namelist())


def _zip_read_text(path: Path, member: str) -> str:
    with zipfile.ZipFile(path, "r") as zf:
        return zf.read(member).decode("utf-8")


# --------------------------------------------------------------------------
# Workbook + Worksheet probes
# --------------------------------------------------------------------------


@_register("workbook_open_basic")
def _probe_workbook_open_basic(tmp_path: Path) -> None:
    import wolfxl

    wb = wolfxl.Workbook()
    assert wb.active is not None
    assert "Sheet" in wb.sheetnames or len(wb.sheetnames) == 1


@_register("workbook_load_basic")
def _probe_workbook_load_basic(tmp_path: Path) -> None:
    import wolfxl

    wb = wolfxl.Workbook()
    ws = wb.active
    ws["A1"] = "hello"
    out = tmp_path / "load_basic.xlsx"
    wb.save(out)

    wb2 = wolfxl.load_workbook(out)
    assert wb2.active["A1"].value == "hello"


@_register("workbook_load_data_only")
def _probe_workbook_load_data_only(tmp_path: Path) -> None:
    import openpyxl
    import wolfxl

    src = tmp_path / "data_only.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = 1
    ws["A2"] = 2
    ws["A3"] = "=SUM(A1:A2)"
    wb.save(src)

    wb2 = wolfxl.load_workbook(src, data_only=False)
    assert wb2.active["A3"].value == "=SUM(A1:A2)"


@_register("workbook_load_modify")
def _probe_workbook_load_modify(tmp_path: Path) -> None:
    import wolfxl

    src = tmp_path / "modify.xlsx"
    wb = wolfxl.Workbook()
    wb.active["A1"] = "before"
    wb.save(src)

    wb2 = wolfxl.load_workbook(src, modify=True)
    wb2.active["A1"] = "after"
    wb2.save(src)
    wb2.close()

    wb3 = wolfxl.load_workbook(src)
    assert wb3.active["A1"].value == "after"


@_register("workbook_load_read_only")
def _probe_workbook_load_read_only(tmp_path: Path) -> None:
    import wolfxl

    src = tmp_path / "ro.xlsx"
    wb = wolfxl.Workbook()
    ws = wb.active
    for i in range(1, 11):
        ws.cell(row=i, column=1, value=i)
    wb.save(src)

    wb2 = wolfxl.load_workbook(src, read_only=True)
    values = [row[0].value for row in wb2.active.iter_rows(min_row=1, max_row=10)]
    assert values == list(range(1, 11))
    wb2.close()


@_register("workbook_save_basic")
def _probe_workbook_save_basic(tmp_path: Path) -> None:
    import wolfxl

    wb = wolfxl.Workbook()
    wb.active["A1"] = 42
    out = tmp_path / "save_basic.xlsx"
    wb.save(out)
    assert out.exists() and out.stat().st_size > 0


@_register("workbook_sheet_access")
def _probe_workbook_sheet_access(tmp_path: Path) -> None:
    import wolfxl

    wb = wolfxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    assert wb["Data"] is ws
    assert wb.active is ws
    assert "Data" in wb.sheetnames


@_register("workbook_create_sheet")
def _probe_workbook_create_sheet(tmp_path: Path) -> None:
    import wolfxl

    wb = wolfxl.Workbook()
    ws = wb.create_sheet(title="Extra")
    assert "Extra" in wb.sheetnames
    assert wb["Extra"] is ws


@_register("workbook_copy_worksheet")
def _probe_workbook_copy_worksheet(tmp_path: Path) -> None:
    import wolfxl

    wb = wolfxl.Workbook()
    ws = wb.active
    ws.title = "Source"
    ws["A1"] = "data"
    copy = wb.copy_worksheet(ws)
    assert copy["A1"].value == "data"
    assert copy.title != "Source"


@_register("workbook_write_only_streaming")
def _probe_workbook_write_only_streaming(tmp_path: Path) -> None:
    """Bounded-memory append-only path (G20 / RFC-073).

    The Rust-side `StreamingSheet` accumulates row XML into a per-sheet
    temp file; `sheet_xml::emit` splices it into the `<sheetData>` slot
    at save time. This probe confirms the end-to-end happy path: 1000
    appends, save, re-read recovers values.
    """
    import wolfxl

    wb = wolfxl.Workbook(write_only=True)
    ws = wb.create_sheet("stream")
    for i in range(1000):
        ws.append([i, i * 2])
    out = tmp_path / "stream.xlsx"
    wb.save(out)

    rb = wolfxl.load_workbook(out, read_only=True)
    rows = list(rb.active.iter_rows(values_only=True, max_row=5))
    assert rows[0] == (0, 0)
    assert rows[-1] == (4, 8)
    rb.close()


@_register("workbook_write_only_bounded_memory")
def _probe_workbook_write_only_bounded_memory(tmp_path: Path) -> None:
    """Peak RSS at 100k rows × 10 cols stays under 80 MiB.

    Runs the writer in a clean subprocess so pytest's accumulated heap
    (test fixtures, imports, prior probe state) doesn't pollute the
    sample. The 80 MiB ceiling is calibrated for the
    100k × 10 numeric-cells workload — SST stays empty (numbers don't
    intern), styles stays at the default 1 entry, and the only growth
    should be the BufWriter buffer (64 KiB) plus per-sheet emit String.

    If this probe fails it means streaming has regressed back to the
    eager BTreeMap accumulation. The actual bytes-on-disk size should
    be ~6-8 MiB; an 80 MiB peak is the slack budget for transient
    Python heap growth during the loop.

    Skips if `psutil` is not installed (calibrated only for posix).
    """
    import subprocess
    import sys

    try:
        import psutil  # noqa: F401  (probed in subprocess, but check first)
    except ImportError:
        pytest.skip("psutil not installed; bounded-memory probe needs it")

    out = tmp_path / "bounded.xlsx"
    script = f"""
import os, psutil, wolfxl
proc = psutil.Process(os.getpid())
baseline_rss = proc.memory_info().rss
wb = wolfxl.Workbook(write_only=True)
ws = wb.create_sheet("Stream")
peak = baseline_rss
for i in range(100_000):
    ws.append([i, i + 1, i + 2, i + 3, i + 4, i + 5, i + 6, i + 7, i + 8, i + 9])
    if i % 10_000 == 0:
        cur = proc.memory_info().rss
        if cur > peak:
            peak = cur
wb.save({str(out)!r})
final = proc.memory_info().rss
if final > peak:
    peak = final
delta_mib = (peak - baseline_rss) / (1024 * 1024)
print(f"BOUNDED_MEMORY_DELTA_MIB={{delta_mib:.2f}}")
"""
    result = subprocess.run(
        [sys.executable, "-c", script],
        capture_output=True,
        text=True,
        timeout=300,
    )
    assert result.returncode == 0, f"subprocess failed: {result.stderr}"
    delta_line = next(
        (line for line in result.stdout.splitlines() if line.startswith("BOUNDED_MEMORY_DELTA_MIB=")),
        None,
    )
    assert delta_line is not None, f"missing delta line in: {result.stdout}"
    delta_mib = float(delta_line.split("=", 1)[1])
    assert delta_mib < 80.0, (
        f"peak RSS delta {delta_mib:.2f} MiB exceeds 80 MiB budget — "
        f"streaming may have regressed to in-memory accumulation"
    )


# --------------------------------------------------------------------------
# Cell + style probes
# --------------------------------------------------------------------------


@_register("cell_basic_value")
def _probe_cell_basic_value(tmp_path: Path) -> None:
    import wolfxl

    wb = wolfxl.Workbook()
    ws = wb.active
    ws["A1"] = "string"
    ws["B1"] = 42
    ws["C1"] = 3.14
    ws["D1"] = True

    out = tmp_path / "cells.xlsx"
    wb.save(out)
    wb2 = wolfxl.load_workbook(out)
    assert wb2.active["A1"].value == "string"
    assert wb2.active["B1"].value == 42
    assert wb2.active["C1"].value == pytest.approx(3.14)
    assert wb2.active["D1"].value is True


@_register("cell_font_fill_border_alignment")
def _probe_cell_font_fill_border_alignment(tmp_path: Path) -> None:
    import wolfxl
    from wolfxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = wolfxl.Workbook()
    ws = wb.active
    ws["A1"] = "styled"
    ws["A1"].font = Font(name="Arial", size=12, bold=True, color="FF0000")
    ws["A1"].fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].border = Border(left=Side(style="thin"), right=Side(style="thin"))
    ws["A1"].number_format = "0.00"

    out = tmp_path / "style.xlsx"
    wb.save(out)
    wb2 = wolfxl.load_workbook(out)
    cell = wb2.active["A1"]
    assert cell.font.bold is True
    assert (cell.font.name or "").lower() in ("arial", "calibri", "")
    assert cell.fill.fill_type == "solid"
    assert cell.alignment.horizontal == "center"


@_register("cell_diagonal_borders")
def _probe_cell_diagonal_borders(tmp_path: Path) -> None:
    import wolfxl
    from wolfxl.styles import Border, Side

    wb = wolfxl.Workbook()
    ws = wb.active
    ws["A1"] = "diag"
    ws["A1"].border = Border(
        diagonal=Side(style="thin", color="000000"),
        diagonalUp=True,
        diagonalDown=False,
    )
    out = tmp_path / "diag.xlsx"
    wb.save(out)

    wb2 = wolfxl.load_workbook(out)
    border = wb2.active["A1"].border
    assert border.diagonalUp is True
    assert border.diagonal is not None
    assert border.diagonal.style == "thin"


@_register("cell_protection")
def _probe_cell_protection(tmp_path: Path) -> None:
    import wolfxl
    from wolfxl.styles import Protection

    wb = wolfxl.Workbook()
    ws = wb.active
    ws["A1"] = "locked"
    ws["A1"].protection = Protection(locked=True, hidden=True)
    out = tmp_path / "prot.xlsx"
    wb.save(out)

    wb2 = wolfxl.load_workbook(out)
    prot = wb2.active["A1"].protection
    assert prot.locked is True
    assert prot.hidden is True


@_register("cell_named_style")
def _probe_cell_named_style(tmp_path: Path) -> None:
    import wolfxl
    from wolfxl.styles import Font, NamedStyle

    wb = wolfxl.Workbook()
    style = NamedStyle(name="Highlight")
    style.font = Font(bold=True)
    wb.add_named_style(style)

    ws = wb.active
    ws["A1"] = "named"
    ws["A1"].style = "Highlight"

    out = tmp_path / "named_style.xlsx"
    wb.save(out)
    wb2 = wolfxl.load_workbook(out)
    assert wb2.active["A1"].style == "Highlight"


@_register("cell_gradient_fill")
def _probe_cell_gradient_fill(tmp_path: Path) -> None:
    import wolfxl
    from wolfxl.styles import Color, GradientFill

    wb = wolfxl.Workbook()
    ws = wb.active
    ws["A1"] = "grad"
    ws["A1"].fill = GradientFill(
        type="linear",
        degree=90,
        stop=(Color("FF0000"), Color("00FF00")),
    )
    out = tmp_path / "grad.xlsx"
    wb.save(out)

    wb2 = wolfxl.load_workbook(out)
    fill = wb2.active["A1"].fill
    assert fill.type == "linear" or fill.type == "gradient"


# --------------------------------------------------------------------------
# Charts probes
# --------------------------------------------------------------------------


@_register("charts_basic_2d")
def _probe_charts_basic_2d(tmp_path: Path) -> None:
    import wolfxl
    from wolfxl.chart import BarChart, Reference

    wb = wolfxl.Workbook()
    ws = wb.active
    for row in [["x", "y"], [1, 10], [2, 20], [3, 30]]:
        ws.append(row)

    chart = BarChart()
    chart.title = "Demo"
    data = Reference(ws, min_col=2, min_row=1, max_row=4)
    cats = Reference(ws, min_col=1, min_row=2, max_row=4)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "D2")

    out = tmp_path / "bar.xlsx"
    wb.save(out)
    assert out.exists() and out.stat().st_size > 0


@_register("charts_advanced_2d")
def _probe_charts_advanced_2d(tmp_path: Path) -> None:
    import wolfxl
    from wolfxl.chart import Reference, ScatterChart

    wb = wolfxl.Workbook()
    ws = wb.active
    for row in [["x", "y"], [1, 10], [2, 20], [3, 30]]:
        ws.append(row)

    chart = ScatterChart()
    xvals = Reference(ws, min_col=1, min_row=2, max_row=4)
    yvals = Reference(ws, min_col=2, min_row=1, max_row=4)
    chart.add_data(yvals, titles_from_data=True)
    chart.set_categories(xvals)
    ws.add_chart(chart, "D2")

    out = tmp_path / "scatter.xlsx"
    wb.save(out)
    assert out.exists() and out.stat().st_size > 0


@_register("charts_3d")
def _probe_charts_3d(tmp_path: Path) -> None:
    import openpyxl as _opx
    import wolfxl
    from wolfxl.chart import AreaChart3D, BarChart3D, LineChart3D, PieChart3D, Reference

    wb = wolfxl.Workbook()
    ws = wb.active
    for row in [["label", "value"], ["a", 10], ["b", 20], ["c", 30], ["d", 40]]:
        ws.append(row)

    families = [
        (BarChart3D, "D2"),
        (LineChart3D, "D18"),
        (PieChart3D, "L2"),
        (AreaChart3D, "L18"),
    ]
    for chart_cls, anchor in families:
        chart = chart_cls()
        chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=5), titles_from_data=True)
        ws.add_chart(chart, anchor)

    out = tmp_path / "charts_3d.xlsx"
    wb.save(out)

    ref_ws = _opx.load_workbook(out).active
    chart_types = {type(chart).__name__ for chart in ref_ws._charts}
    assert {"BarChart3D", "LineChart3D", "PieChart3D", "AreaChart3D"} <= chart_types


@_register("charts_surface_stock_projected")
def _probe_charts_surface_stock_projected(tmp_path: Path) -> None:
    import openpyxl as _opx
    import wolfxl
    from wolfxl.chart import ProjectedPieChart, Reference, StockChart, SurfaceChart, SurfaceChart3D

    wb = wolfxl.Workbook()
    ws = wb.active
    ws.append(["label", "open", "high", "low", "close"])
    for idx in range(1, 5):
        ws.append([f"q{idx}", 10 + idx, 14 + idx, 8 + idx, 12 + idx])

    surface = SurfaceChart()
    surface.add_data(Reference(ws, min_col=2, max_col=5, min_row=1, max_row=5), titles_from_data=True)
    ws.add_chart(surface, "G2")

    surface_3d = SurfaceChart3D()
    surface_3d.add_data(
        Reference(ws, min_col=2, max_col=5, min_row=1, max_row=5),
        titles_from_data=True,
    )
    ws.add_chart(surface_3d, "G18")

    stock = StockChart()
    stock.add_data(Reference(ws, min_col=2, max_col=5, min_row=1, max_row=5), titles_from_data=True)
    stock.set_categories(Reference(ws, min_col=1, min_row=2, max_row=5))
    ws.add_chart(stock, "O2")

    projected = ProjectedPieChart()
    projected.add_data(Reference(ws, min_col=5, min_row=1, max_row=5), titles_from_data=True)
    ws.add_chart(projected, "O18")

    out = tmp_path / "charts_surface_stock_projected.xlsx"
    wb.save(out)

    ref_ws = _opx.load_workbook(out).active
    chart_types = {type(chart).__name__ for chart in ref_ws._charts}
    assert {"SurfaceChart", "SurfaceChart3D", "StockChart", "ProjectedPieChart"} <= chart_types


@_register("charts_add_remove_replace")
def _probe_charts_add_remove_replace(tmp_path: Path) -> None:
    import wolfxl
    from wolfxl.chart import BarChart, Reference

    wb = wolfxl.Workbook()
    ws = wb.active
    for row in [["x", "y"], [1, 10], [2, 20]]:
        ws.append(row)

    chart = BarChart()
    data = Reference(ws, min_col=2, min_row=1, max_row=3)
    chart.add_data(data, titles_from_data=True)
    ws.add_chart(chart, "D2")
    assert hasattr(ws, "remove_chart")
    ws.remove_chart(chart)


@_register("charts_combination")
def _probe_charts_combination(tmp_path: Path) -> None:
    import openpyxl as _opx
    import wolfxl
    from wolfxl.chart import BarChart, LineChart, Reference

    wb = wolfxl.Workbook()
    ws = wb.active
    for row in [["x", "y", "z"], [1, 10, 100], [2, 20, 200], [3, 30, 300]]:
        ws.append(row)

    bar = BarChart()
    line = LineChart()
    bar.add_data(Reference(ws, min_col=2, min_row=1, max_row=4), titles_from_data=True)
    line.add_data(Reference(ws, min_col=3, min_row=1, max_row=4), titles_from_data=True)
    line.y_axis.crosses = "max"
    line.y_axis.axId = 200
    bar += line
    ws.add_chart(bar, "E2")
    out = tmp_path / "combo.xlsx"
    wb.save(out)

    ref_wb = _opx.load_workbook(out)
    ref_ws = ref_wb[ref_wb.sheetnames[0]]
    ref_charts = getattr(ref_ws, "_charts", [])
    assert ref_charts, "openpyxl found no charts in combo file"
    types = {type(c).__name__ for c in ref_charts}
    assert "BarChart" in types and "LineChart" in types, (
        f"openpyxl saw {types}, expected both BarChart and LineChart"
    )


@_register("charts_label_rich_text")
def _probe_charts_label_rich_text(tmp_path: Path) -> None:
    import wolfxl
    from wolfxl.cell.text import InlineFont
    from wolfxl.chart import BarChart, Reference
    from wolfxl.chart.label import DataLabelList
    from wolfxl.cell.rich_text import CellRichText, TextBlock

    wb = wolfxl.Workbook()
    ws = wb.active
    for row in [["x", "y"], [1, 10], [2, 20]]:
        ws.append(row)

    chart = BarChart()
    chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=3), titles_from_data=True)
    rich = CellRichText([TextBlock(InlineFont(b=True), "bold-label")])
    chart.dataLabels = DataLabelList(rich=rich)
    ws.add_chart(chart, "D2")
    wb.save(tmp_path / "rich_chart.xlsx")


@_register("charts_pivot_chart_per_point")
def _probe_charts_pivot_chart_per_point(tmp_path: Path) -> None:
    """G16 — per-data-point fill override on a pivot-source chart.

    The OOXML per-point override (`<c:dPt>` with `<c:spPr><a:solidFill>`)
    is identical for pivot and non-pivot charts; the only difference for
    a pivot chart is the chart-level `<c:pivotSource>` block plus a
    `<c:fmtId>` on each `<c:ser>`. This probe stamps a pivot-source on
    the chart (via the (name, fmt_id) tuple form so we don't have to
    materialise a real PivotTable) and then exercises the per-point
    override path. After save + openpyxl reload, we assert openpyxl
    sees the dPt entry on the series — proving the data_points dict
    bridge survives the pivot emit path.
    """
    import openpyxl as _opx
    import wolfxl
    from wolfxl.chart import BarChart, Reference
    from wolfxl.chart.marker import DataPoint
    from wolfxl.chart.shapes import GraphicalProperties

    wb = wolfxl.Workbook()
    ws = wb.active
    for row in [["x", "y"], [1, 10], [2, 20], [3, 30]]:
        ws.append(row)

    chart = BarChart()
    chart.add_data(
        Reference(ws, min_col=2, min_row=1, max_row=4),
        titles_from_data=True,
    )
    chart.pivot_source = ("MyPivot", 0)

    red = GraphicalProperties(solidFill="FF0000")
    chart.series[0].dPt = [DataPoint(idx=1, spPr=red)]

    ws.add_chart(chart, "D2")
    out = tmp_path / "pivot_per_point.xlsx"
    wb.save(out)

    ref_wb = _opx.load_workbook(out)
    ref_ws = ref_wb[ref_wb.sheetnames[0]]
    ref_chart = ref_ws._charts[0]
    ref_dpt = list(ref_chart.series[0].dPt)
    assert ref_dpt, "openpyxl saw no per-point overrides on pivot chart"
    assert any(dp.idx == 1 for dp in ref_dpt), "expected idx=1 override"


# --------------------------------------------------------------------------
# Pivot probes
# --------------------------------------------------------------------------


@_register("pivots_construction")
def _probe_pivots_construction(tmp_path: Path) -> None:
    """Pivot construction is a modify-mode operation (matches openpyxl
    workflow where pivots are stamped onto an existing workbook).
    """
    import wolfxl
    from wolfxl.chart import Reference
    from wolfxl.pivot import PivotCache, PivotTable

    src = tmp_path / "pivot_src.xlsx"
    wb = wolfxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["region", "amount"])
    for row in [("east", 10), ("west", 20), ("east", 30), ("west", 40)]:
        ws.append(row)
    wb.create_sheet("Pivot")
    wb.save(src)

    wb2 = wolfxl.load_workbook(src, modify=True)
    data_ws = wb2["Data"]
    cache = PivotCache(source=Reference(data_ws, min_col=1, min_row=1, max_col=2, max_row=5))
    wb2.add_pivot_cache(cache)
    pt = PivotTable(cache=cache, location="A1", rows=["region"], data=["amount"])
    wb2["Pivot"].add_pivot_table(pt)
    wb2.save(src)
    wb2.close()

    wb3 = wolfxl.load_workbook(src)
    assert "Pivot" in wb3.sheetnames


@_register("pivots_in_place_edit")
def _probe_pivots_in_place_edit(tmp_path: Path) -> None:
    """Edit an existing pivot's source range. Tracked under G17 (S5)."""
    import wolfxl
    from wolfxl.chart import Reference
    from wolfxl.pivot import PivotCache, PivotTable

    wb = wolfxl.Workbook()
    ws = wb.active
    ws.append(["region", "amount"])
    for row in [("east", 10), ("west", 20)]:
        ws.append(row)
    cache = PivotCache(source=Reference(ws, min_col=1, min_row=1, max_col=2, max_row=3))
    wb.add_pivot_cache(cache)
    pt = PivotTable(cache=cache, location="D2", rows=["region"], data=["amount"])
    target = wb.create_sheet("Pivot")
    target.add_pivot_table(pt, "A1")
    src = tmp_path / "pivot_edit.xlsx"
    wb.save(src)

    wb2 = wolfxl.load_workbook(src, modify=True)
    pivot = wb2["Pivot"].pivot_tables[0]
    pivot.source = Reference(wb2.active, min_col=1, min_row=1, max_col=2, max_row=3)
    wb2.save(src)


@_register("pivots_field_mutation")
def _probe_pivots_field_mutation(tmp_path: Path) -> None:
    import wolfxl
    from wolfxl.chart import Reference
    from wolfxl.pivot import DataField, DataFunction, PivotCache, PivotTable

    src = tmp_path / "pivot_field_mutation.xlsx"
    wb = wolfxl.Workbook()
    ws = wb.active
    for row in [("region", "amount", "scenario"), ("east", 10, "A"), ("west", 20, "B")]:
        ws.append(row)
    cache = PivotCache(source=Reference(ws, min_col=1, min_row=1, max_col=3, max_row=3))
    wb.add_pivot_cache(cache)
    pt = PivotTable(cache=cache, location="D2", rows=["region"], data=["amount"])
    wb.create_sheet("Pivot").add_pivot_table(pt, "A1")
    wb.save(src)

    wb2 = wolfxl.load_workbook(src, modify=True)
    pivot = wb2["Pivot"].pivot_tables[0]
    pivot.row_fields = ["region"]
    pivot.column_fields = ["scenario"]
    pivot.data_fields = [DataField("amount", function=DataFunction.SUM)]
    wb2.save(src)

    xml = _zip_read_text(src, "xl/pivotTables/pivotTable1.xml")
    assert '<colFields count="1"><field x="2"/></colFields>' in xml


@_register("pivots_filter_mutation")
def _probe_pivots_filter_mutation(tmp_path: Path) -> None:
    import wolfxl
    from wolfxl.chart import Reference
    from wolfxl.pivot import PageField, PivotCache, PivotTable

    src = tmp_path / "pivot_filter_mutation.xlsx"
    wb = wolfxl.Workbook()
    ws = wb.active
    for row in [("region", "amount"), ("east", 10), ("west", 20)]:
        ws.append(row)
    cache = PivotCache(source=Reference(ws, min_col=1, min_row=1, max_col=2, max_row=3))
    wb.add_pivot_cache(cache)
    pt = PivotTable(cache=cache, location="D2", rows=["region"], data=["amount"])
    wb.create_sheet("Pivot").add_pivot_table(pt, "A1")
    wb.save(src)

    wb2 = wolfxl.load_workbook(src, modify=True)
    wb2["Pivot"].pivot_tables[0].page_fields = [PageField("region", item_index=0)]
    wb2.save(src)

    xml = _zip_read_text(src, "xl/pivotTables/pivotTable1.xml")
    assert '<pageFields count="1"><pageField fld="0" item="0"/></pageFields>' in xml


@_register("pivots_aggregation_mutation")
def _probe_pivots_aggregation_mutation(tmp_path: Path) -> None:
    import wolfxl
    from wolfxl.chart import Reference
    from wolfxl.pivot import DataFunction, PivotCache, PivotTable

    src = tmp_path / "pivot_aggregation_mutation.xlsx"
    wb = wolfxl.Workbook()
    ws = wb.active
    for row in [("region", "amount"), ("east", 10), ("west", 20)]:
        ws.append(row)
    cache = PivotCache(source=Reference(ws, min_col=1, min_row=1, max_col=2, max_row=3))
    wb.add_pivot_cache(cache)
    pt = PivotTable(cache=cache, location="D2", rows=["region"], data=["amount"])
    wb.create_sheet("Pivot").add_pivot_table(pt, "A1")
    wb.save(src)

    wb2 = wolfxl.load_workbook(src, modify=True)
    wb2["Pivot"].pivot_tables[0].set_aggregation("amount", DataFunction.AVERAGE)
    wb2.save(src)

    xml = _zip_read_text(src, "xl/pivotTables/pivotTable1.xml")
    assert 'subtotal="average"' in xml


def _make_pivot_fixture(path: Path) -> None:
    import openpyxl
    import wolfxl
    from wolfxl.chart import Reference
    from wolfxl.pivot import PivotCache, PivotTable

    seed = path.parent / "_pivot_seed.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    for row in [
        ("region", "quarter", "revenue"),
        ("North", "Q1", 100.0),
        ("South", "Q1", 200.0),
        ("North", "Q2", 150.0),
        ("South", "Q2", 250.0),
    ]:
        ws.append(row)
    wb.save(seed)

    wbm = wolfxl.load_workbook(seed, modify=True)
    try:
        wsm = wbm["Data"]
        ref = Reference(wsm, min_col=1, min_row=1, max_col=3, max_row=5)
        cache = PivotCache(source=ref)
        table = PivotTable(
            cache=cache,
            location="F2",
            rows=["region"],
            cols=["quarter"],
            data=["revenue"],
            name="SalesPivot",
        )
        wbm.add_pivot_cache(cache)
        wsm.add_pivot_table(table)
        wbm.save(path)
    finally:
        wbm.close()
    seed.unlink(missing_ok=True)


@_register("pivots_linked_chart")
def _probe_pivots_linked_chart(tmp_path: Path) -> None:
    import wolfxl
    from wolfxl.chart import BarChart, Reference

    src = tmp_path / "pivot.xlsx"
    out = tmp_path / "pivot_chart.xlsx"
    _make_pivot_fixture(src)

    wb = wolfxl.load_workbook(src, modify=True)
    try:
        ws = wb["Data"]
        chart = BarChart()
        chart.add_data(Reference(ws, min_col=3, min_row=1, max_row=5), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=5))
        chart.pivot_source = ws.pivot_tables[0]
        ws.add_chart(chart, "F12")
        wb.save(out)
    finally:
        wb.close()

    chart_files = [name for name in _zip_listing(out) if re.match(r"^xl/charts/chart\d+\.xml$", name)]
    assert chart_files, "pivot-linked chart part missing"
    chart_xml = "\n".join(_zip_read_text(out, name) for name in chart_files)
    assert "<c:pivotSource>" in chart_xml
    assert "<c:name>SalesPivot</c:name>" in chart_xml


@_register("pivots_copy_worksheet")
def _probe_pivots_copy_worksheet(tmp_path: Path) -> None:
    import wolfxl

    src = tmp_path / "pivot.xlsx"
    out = tmp_path / "pivot_copy.xlsx"
    _make_pivot_fixture(src)

    wb = wolfxl.load_workbook(src, modify=True)
    try:
        wb.copy_worksheet(wb["Data"], name="DataCopy")
        wb.save(out)
    finally:
        wb.close()

    entries = _zip_listing(out)
    table_parts = [name for name in entries if re.match(r"^xl/pivotTables/pivotTable\d+\.xml$", name)]
    assert len(table_parts) == 2, f"expected source + copied pivot table parts, got {table_parts}"

    pivot_targets: set[str] = set()
    for rels in [name for name in entries if re.match(r"^xl/worksheets/_rels/sheet\d+\.xml\.rels$", name)]:
        rels_xml = _zip_read_text(out, rels)
        pivot_targets.update(re.findall(r'Target="\.\./pivotTables/(pivotTable\d+\.xml)"', rels_xml))
    assert len(pivot_targets) == 2, f"expected distinct pivot table rel targets, got {pivot_targets}"


# --------------------------------------------------------------------------
# Image probes
# --------------------------------------------------------------------------

_PNG_1X1 = (
    b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01\x08\x06\x00\x00\x00"
    b"\x1f\x15\xc4\x89\x00\x00\x00\rIDATx\x9cc\xfa\xcf\x00\x00\x00\x02\x00\x01\xe5'\xde\xfc"
    b"\x00\x00\x00\x00IEND\xaeB`\x82"
)


@_register("images_basic")
def _probe_images_basic(tmp_path: Path) -> None:
    import wolfxl
    from wolfxl.drawing.image import Image

    img_path = tmp_path / "tiny.png"
    img_path.write_bytes(_PNG_1X1)

    wb = wolfxl.Workbook()
    ws = wb.active
    ws.add_image(Image(str(img_path)), "B2")
    out = tmp_path / "with_image.xlsx"
    wb.save(out)
    assert out.exists() and out.stat().st_size > 0


@_register("images_replace_remove")
def _probe_images_replace_remove(tmp_path: Path) -> None:
    """Public replace/remove image API. Tracked under G06 (S1)."""
    import wolfxl
    from wolfxl.drawing.image import Image

    img_path = tmp_path / "tiny.png"
    img_path.write_bytes(_PNG_1X1)

    wb = wolfxl.Workbook()
    ws = wb.active
    img = Image(str(img_path))
    ws.add_image(img, "B2")
    assert hasattr(ws, "remove_image"), "Worksheet.remove_image missing"
    ws.remove_image(img)


# --------------------------------------------------------------------------
# Structural probes
# --------------------------------------------------------------------------


@_register("structural_insert_delete_rows")
def _probe_structural_insert_delete_rows(tmp_path: Path) -> None:
    """insert_rows / delete_rows shift existing data on save (RFC-030/031).

    wolfxl's modify-mode treats insert/delete as queued operations applied
    at flush time, so this probe asserts the post-save shift only and does
    not interleave writes at the moved rows.
    """
    import wolfxl

    src = tmp_path / "ins.xlsx"
    wb = wolfxl.Workbook()
    ws = wb.active
    for i in range(1, 6):
        ws.cell(row=i, column=1, value=i)
    wb.save(src)

    wb2 = wolfxl.load_workbook(src, modify=True)
    wb2.active.insert_rows(2, amount=1)
    wb2.save(src)
    wb2.close()

    wb3 = wolfxl.load_workbook(src)
    assert wb3.active["A1"].value == 1
    assert wb3.active["A2"].value is None  # inserted blank row
    assert wb3.active["A3"].value == 2

    wb4 = wolfxl.load_workbook(src, modify=True)
    wb4.active.delete_rows(2, amount=1)
    wb4.save(src)
    wb4.close()

    wb5 = wolfxl.load_workbook(src)
    assert wb5.active["A2"].value == 2


@_register("structural_insert_delete_cols")
def _probe_structural_insert_delete_cols(tmp_path: Path) -> None:
    import openpyxl as _opx
    import wolfxl

    src = tmp_path / "cols.xlsx"
    wb = _opx.Workbook()
    ws = wb.active
    for col, value in enumerate(["A", "B", "C", "D"], start=1):
        ws.cell(row=1, column=col, value=value)
    wb.save(src)

    wb2 = wolfxl.load_workbook(src, modify=True)
    wb2.active.insert_cols(2, amount=1)
    wb2.save(src)
    wb2.close()

    inserted = _opx.load_workbook(src).active
    assert inserted["A1"].value == "A"
    assert inserted["B1"].value is None
    assert inserted["C1"].value == "B"

    wb3 = wolfxl.load_workbook(src, modify=True)
    wb3.active.delete_cols(2, amount=1)
    wb3.save(src)
    wb3.close()

    deleted = _opx.load_workbook(src).active
    assert [deleted.cell(row=1, column=col).value for col in range(1, 5)] == ["A", "B", "C", "D"]


@_register("structural_move_range")
def _probe_structural_move_range(tmp_path: Path) -> None:
    import openpyxl as _opx
    import wolfxl

    src = tmp_path / "move.xlsx"
    wb = _opx.Workbook()
    ws = wb.active
    ws["B2"] = 10
    ws["C2"] = 20
    wb.save(src)

    wb2 = wolfxl.load_workbook(src, modify=True)
    wb2.active.move_range("B2:C2", rows=2, cols=1)
    wb2.save(src)
    wb2.close()

    moved = _opx.load_workbook(src).active
    assert moved["B2"].value is None
    assert moved["C2"].value is None
    assert moved["C4"].value == 10
    assert moved["D4"].value == 20


# --------------------------------------------------------------------------
# Modify-mode probes
# --------------------------------------------------------------------------


@_register("modify_defined_names")
def _probe_modify_defined_names(tmp_path: Path) -> None:
    import wolfxl
    from wolfxl.workbook.defined_name import DefinedName

    wb = wolfxl.Workbook()
    wb.active.title = "Data"
    wb.defined_names["TaxRate"] = DefinedName(name="TaxRate", value="Data!$A$1")
    out = tmp_path / "dn.xlsx"
    wb.save(out)

    wb2 = wolfxl.load_workbook(out)
    assert "TaxRate" in wb2.defined_names


@_register("modify_tables")
def _probe_modify_tables(tmp_path: Path) -> None:
    import wolfxl
    from wolfxl.worksheet.table import Table, TableStyleInfo

    wb = wolfxl.Workbook()
    ws = wb.active
    ws.append(["a", "b"])
    ws.append([1, 2])
    ws.append([3, 4])

    table = Table(name="MyTable", displayName="MyTable", ref="A1:B3")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)

    out = tmp_path / "tbl.xlsx"
    wb.save(out)

    wb2 = wolfxl.load_workbook(out)
    assert "MyTable" in [t.displayName for t in wb2.active.tables.values()]


@_register("modify_data_validations")
def _probe_modify_data_validations(tmp_path: Path) -> None:
    import wolfxl
    from wolfxl.worksheet.datavalidation import DataValidation

    wb = wolfxl.Workbook()
    ws = wb.active
    dv = DataValidation(type="list", formula1='"a,b,c"')
    dv.add("A1:A10")
    ws.data_validations.append(dv)

    out = tmp_path / "dv.xlsx"
    wb.save(out)

    wb2 = wolfxl.load_workbook(out)
    assert len(list(wb2.active.data_validations.dataValidation)) >= 1


@_register("modify_document_properties")
def _probe_modify_document_properties(tmp_path: Path) -> None:
    import openpyxl as _opx
    import wolfxl

    src = tmp_path / "props.xlsx"
    _opx.Workbook().save(src)

    wb = wolfxl.load_workbook(src, modify=True)
    wb.properties.title = "Phase 10 Oracle"
    wb.properties.creator = "wolfxl"
    wb.save(src)
    wb.close()

    ref = _opx.load_workbook(src)
    assert ref.properties.title == "Phase 10 Oracle"
    assert ref.properties.creator == "wolfxl"


# --------------------------------------------------------------------------
# Read-side probes
# --------------------------------------------------------------------------


@_register("read_xlsx")
def _probe_read_xlsx(tmp_path: Path) -> None:
    import openpyxl as _opx
    import wolfxl

    src = tmp_path / "read.xlsx"
    wb = _opx.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "hello"
    ws["B2"] = 42
    wb.save(src)

    got = wolfxl.load_workbook(src)
    assert got.sheetnames == ["Data"]
    assert got["Data"]["A1"].value == "hello"
    assert got["Data"]["B2"].value == 42


@_register("read_xlsb")
def _probe_read_xlsb(tmp_path: Path) -> None:
    import wolfxl

    fixture = Path(__file__).parent / "fixtures" / "sprint_kappa_smoke.xlsb"
    assert fixture.exists(), f"xlsb fixture missing: {fixture}"
    wb = wolfxl.load_workbook(fixture)
    assert wb.sheetnames, "xlsb fixture should expose sheet names"
    rows = list(wb.active.iter_rows(values_only=True, max_row=5))
    assert any(any(cell is not None for cell in row) for row in rows), "xlsb fixture read no values"


# --------------------------------------------------------------------------
# Utility probes
# --------------------------------------------------------------------------


@_register("utils_get_column_letter")
def _probe_utils_get_column_letter(tmp_path: Path) -> None:
    import openpyxl.utils
    import wolfxl.utils.cell as wc

    for col in (1, 26, 27, 52, 100, 1000, 16384):
        assert wc.get_column_letter(col) == openpyxl.utils.get_column_letter(col)


@_register("utils_column_index_from_string")
def _probe_utils_column_index_from_string(tmp_path: Path) -> None:
    import openpyxl.utils
    import wolfxl.utils.cell as wc

    for s in ("A", "Z", "AA", "AZ", "BA", "ZZ", "AAA", "XFD"):
        assert wc.column_index_from_string(s) == openpyxl.utils.column_index_from_string(s)


@_register("utils_range_boundaries")
def _probe_utils_range_boundaries(tmp_path: Path) -> None:
    import openpyxl.utils
    import wolfxl.utils.cell as wc

    for r in ("A1:B2", "C3:Z99", "AA1:AZ100"):
        assert wc.range_boundaries(r) == openpyxl.utils.range_boundaries(r)


@_register("utils_coordinate_to_tuple")
def _probe_utils_coordinate_to_tuple(tmp_path: Path) -> None:
    from openpyxl.utils.cell import coordinate_to_tuple as openpyxl_coordinate_to_tuple
    import wolfxl.utils.cell as wc

    for coordinate in ("A1", "B3", "C4", "XFD1048576"):
        assert wc.coordinate_to_tuple(coordinate) == openpyxl_coordinate_to_tuple(coordinate)

    with pytest.raises(Exception):
        openpyxl_coordinate_to_tuple("$C$4")

    for bad in ("not-a-cell", "A", "1", "AAAA1", "$C$4"):
        with pytest.raises(Exception):
            wc.coordinate_to_tuple(bad)


# --------------------------------------------------------------------------
# Protection probes
# --------------------------------------------------------------------------


@_register("protection_sheet")
def _probe_protection_sheet(tmp_path: Path) -> None:
    import openpyxl as _opx
    import wolfxl
    from wolfxl.worksheet.protection import SheetProtection

    wb = wolfxl.Workbook()
    ws = wb.active
    ws.protection = SheetProtection(sheet=True, password="secret", formatCells=False)
    out = tmp_path / "sheet_prot.xlsx"
    wb.save(out)

    wb2 = wolfxl.load_workbook(out)
    assert wb2.active.protection.sheet is True

    ref_wb = _opx.load_workbook(out)
    ref_prot = ref_wb[ref_wb.sheetnames[0]].protection
    assert ref_prot.sheet is True, "openpyxl saw protection.sheet=False"
    assert ref_prot.formatCells is False, (
        "openpyxl lost the formatCells=False override"
    )
    assert ref_prot.password is not None, "openpyxl saw no password hash"


@_register("protection_workbook")
def _probe_protection_workbook(tmp_path: Path) -> None:
    import wolfxl
    from wolfxl.workbook.protection import WorkbookProtection

    wb = wolfxl.Workbook()
    wb.security = WorkbookProtection(lockStructure=True, workbookPassword="secret")
    out = tmp_path / "wb_prot.xlsx"
    wb.save(out)

    wb2 = wolfxl.load_workbook(out)
    assert wb2.security is not None
    assert getattr(wb2.security, "lockStructure", False) is True


# --------------------------------------------------------------------------
# External links probes
# --------------------------------------------------------------------------


@_register("external_links_collection")
def _probe_external_links_collection(tmp_path: Path) -> None:
    """Workbook-level external link collection. Tracked under G18 (S6).

    The probe asserts that round-tripping a workbook with an external-link
    formula preserves the ``xl/externalLinks/`` parts. Today wolfxl preserves
    the parts on modify-save but does not expose a Python collection;
    ``wb._external_links`` (or equivalent public surface) is what S6 ships.
    """
    import wolfxl

    wb = wolfxl.Workbook()
    wb.active["A1"] = "='[ext.xlsx]Sheet1'!$A$1"
    out = tmp_path / "ext.xlsx"
    wb.save(out)

    wb2 = wolfxl.load_workbook(out)
    links = getattr(wb2, "_external_links", None) or getattr(wb2, "external_links", None)
    assert links is not None and len(links) >= 0  # surface must exist


@_register("external_links_authoring")
def _probe_external_links_authoring(tmp_path: Path) -> None:
    import wolfxl
    from wolfxl import ExternalLink

    out = tmp_path / "authored_external_link.xlsx"
    wb = wolfxl.Workbook()
    wb.active["A1"] = "='[linked.xlsx]Sheet1'!$A$1"
    wb._external_links.append(ExternalLink(target="linked.xlsx", sheet_names=["Sheet1"]))
    wb.save(out)

    entries = _zip_listing(out)
    assert "xl/externalLinks/externalLink1.xml" in entries
    assert "xl/externalLinks/_rels/externalLink1.xml.rels" in entries
    rels = _zip_read_text(out, "xl/externalLinks/_rels/externalLink1.xml.rels")
    assert "linked.xlsx" in rels


# --------------------------------------------------------------------------
# VBA inspection probes
# --------------------------------------------------------------------------


@_register("vba_inspect")
def _probe_vba_inspect(tmp_path: Path) -> None:
    """Read-only VBA archive inspection. Tracked under G19 (S6, RFC-072).

    Loads a vendored ``.xlsm`` fixture in modify-mode and asserts that
    ``wb.vba_archive`` surfaces the underlying ``xl/vbaProject.bin``
    bytes. Authoring is out of scope for v1.0 (G28).
    """
    import wolfxl

    fixture = Path(__file__).parent / "fixtures" / "macro_basic.xlsm"
    if not fixture.exists():
        pytest.skip("vba inspection fixture not vendored")
    wb = wolfxl.load_workbook(str(fixture), modify=True)
    archive = wb.vba_archive
    assert archive is not None, "wb.vba_archive must surface bytes for .xlsm"
    assert isinstance(archive, (bytes, bytearray, memoryview))
    assert len(archive) > 0


@_register("vba_preserve")
def _probe_vba_preserve(tmp_path: Path) -> None:
    import wolfxl

    fixture = Path(__file__).parent / "fixtures" / "macro_basic.xlsm"
    if not fixture.exists():
        pytest.skip("vba preservation fixture not vendored")
    work = tmp_path / "macro_preserve.xlsm"
    shutil.copy(fixture, work)

    wb = wolfxl.load_workbook(work, modify=True)
    before = wb.vba_archive
    assert before is not None and len(before) > 0
    wb.active["A1"] = "preserve"
    wb.save(work)
    wb.close()

    wb2 = wolfxl.load_workbook(work, modify=True)
    after = wb2.vba_archive
    assert after is not None
    assert bytes(after) == bytes(before)
    assert "xl/vbaProject.bin" in _zip_listing(work)


# --------------------------------------------------------------------------
# Comments probes
# --------------------------------------------------------------------------


@_register("comments_basic")
def _probe_comments_basic(tmp_path: Path) -> None:
    import wolfxl
    from wolfxl.comments import Comment

    wb = wolfxl.Workbook()
    ws = wb.active
    ws["A1"] = "anchor"
    ws["A1"].comment = Comment("hello", "wolfie")
    out = tmp_path / "cmt.xlsx"
    wb.save(out)

    wb2 = wolfxl.load_workbook(out)
    assert wb2.active["A1"].comment is not None
    assert wb2.active["A1"].comment.text == "hello"


@_register("comments_threaded")
def _probe_comments_threaded(tmp_path: Path) -> None:
    """Threaded comments (xl/threadedComments). Tracked under G08 (S2).

    Probe asserts the full round-trip: add a top-level threaded comment +
    one reply, save, reload, and confirm both texts survive. Class-existence
    alone is not enough — that lets the probe XPASS on partial work.
    """
    import wolfxl
    from wolfxl.comments import ThreadedComment

    wb = wolfxl.Workbook()
    ws = wb.active
    ws["A1"] = "anchor"
    alice = wb.persons.add(name="Alice", user_id="alice@example.com")
    top = ThreadedComment(text="Looks wrong", person=alice)
    top.replies.append(
        ThreadedComment(text="Agreed; investigating", person=alice, parent=top)
    )
    ws["A1"].threaded_comment = top
    out = tmp_path / "tc.xlsx"
    wb.save(out)

    wb2 = wolfxl.load_workbook(out)
    got = wb2.active["A1"].threaded_comment
    assert got is not None, "threaded comment lost across save+load"
    assert got.text == "Looks wrong"
    assert len(got.replies) == 1
    assert got.replies[0].text == "Agreed; investigating"


# --------------------------------------------------------------------------
# Rich text probes
# --------------------------------------------------------------------------


@_register("rich_text_cell")
def _probe_rich_text_cell(tmp_path: Path) -> None:
    import wolfxl
    from wolfxl.cell.rich_text import CellRichText, TextBlock
    from wolfxl.cell.text import InlineFont

    wb = wolfxl.Workbook()
    ws = wb.active
    ws["A1"] = CellRichText(
        [
            TextBlock(InlineFont(b=True), "bold "),
            TextBlock(InlineFont(i=True), "italic"),
        ]
    )
    out = tmp_path / "rt.xlsx"
    wb.save(out)

    wb2 = wolfxl.load_workbook(out, rich_text=True)
    val = wb2.active["A1"].value
    assert isinstance(val, CellRichText)


@_register("rich_text_headers_footers")
def _probe_rich_text_headers_footers(tmp_path: Path) -> None:
    """Rich text in headers/footers. Tracked under G09 (S2).

    Today ``ws.oddHeader`` is a HeaderFooter object but rich-text runs
    inside it are not authored from the Python side. The probe asserts a
    multi-run header round-trips its formatting metadata.
    """
    import wolfxl

    wb = wolfxl.Workbook()
    ws = wb.active
    ws["A1"] = "data"
    ws.oddHeader.center.text = "Title"
    ws.oddHeader.center.font = "Arial,Bold"
    ws.oddHeader.center.size = 14
    out = tmp_path / "hdr.xlsx"
    wb.save(out)

    wb2 = wolfxl.load_workbook(out)
    hf = wb2.active.oddHeader.center
    assert hf.text == "Title"
    assert hf.size == 14


# --------------------------------------------------------------------------
# Conditional formatting probes
# --------------------------------------------------------------------------


@_register("cf_basic_rules")
def _probe_cf_basic_rules(tmp_path: Path) -> None:
    """Basic CF rule round-trip, including openpyxl's generic text rule."""
    import wolfxl
    from wolfxl.formatting.rule import CellIsRule, Rule

    wb = wolfxl.Workbook()
    ws = wb.active
    for i in range(1, 6):
        ws.cell(row=i, column=1, value=i)
    ws.conditional_formatting.add(
        "A1:A5",
        CellIsRule(operator="greaterThan", formula=["3"]),
    )
    ws.conditional_formatting.add(
        "B1:B5",
        Rule(
            type="containsText",
            operator="containsText",
            text="foo",
            formula=['NOT(ISERROR(SEARCH("foo",B1)))'],
        ),
    )
    out = tmp_path / "cf.xlsx"
    wb.save(out)

    wb2 = wolfxl.load_workbook(out)
    rules = list(wb2.active.conditional_formatting)
    assert len(rules) >= 1
    import openpyxl as _opx

    ref_wb = _opx.load_workbook(out)
    seen_types = {
        rule.type
        for cf_range in ref_wb.active.conditional_formatting
        for rule in ref_wb.active.conditional_formatting[cf_range]
    }
    assert {"cellIs", "containsText"} <= seen_types


@_register("cf_cellis_operator_matrix")
def _probe_cf_cellis_operator_matrix(tmp_path: Path) -> None:
    """CellIsRule operator matrix, including two-formula operators."""
    import openpyxl as _opx
    import wolfxl
    from wolfxl.formatting.rule import CellIsRule

    cases = [
        ("A1:A5", "equal", ["3"]),
        ("B1:B5", "notEqual", ["3"]),
        ("C1:C5", "greaterThan", ["3"]),
        ("D1:D5", "greaterThanOrEqual", ["3"]),
        ("E1:E5", "lessThan", ["3"]),
        ("F1:F5", "lessThanOrEqual", ["3"]),
        ("G1:G5", "between", ["2", "4"]),
        ("H1:H5", "notBetween", ["2", "4"]),
        ("I1:I5", "between", ["SUM(A1,A2)", "10"]),
    ]
    wb = wolfxl.Workbook()
    ws = wb.active
    for row in range(1, 6):
        for col in range(1, 10):
            ws.cell(row=row, column=col, value=row)
    for sqref, operator, formula in cases:
        ws.conditional_formatting.add(sqref, CellIsRule(operator=operator, formula=formula))
    out = tmp_path / "cf_cellis_ops.xlsx"
    wb.save(out)

    ref_ws = _opx.load_workbook(out).active
    by_range = {
        str(cf_range.sqref): ref_ws.conditional_formatting[cf_range][0]
        for cf_range in ref_ws.conditional_formatting
    }
    for sqref, operator, formula in cases:
        rule = by_range[sqref]
        assert rule.type == "cellIs"
        assert rule.operator == operator
        assert [str(part) for part in rule.formula] == formula


@_register("cf_icon_sets")
def _probe_cf_icon_sets(tmp_path: Path) -> None:
    """Icon set rule. Tracked under G11 (S3)."""
    import wolfxl
    from wolfxl.formatting.rule import IconSetRule

    wb = wolfxl.Workbook()
    ws = wb.active
    for i in range(1, 6):
        ws.cell(row=i, column=1, value=i)
    rule = IconSetRule("3TrafficLights1", "percent", [0, 33, 67])
    ws.conditional_formatting.add("A1:A5", rule)
    out = tmp_path / "iconset.xlsx"
    wb.save(out)

    wb2 = wolfxl.load_workbook(out)
    rules = []
    for cf_range in wb2.active.conditional_formatting:
        rules.extend(cf_range.rules if hasattr(cf_range, "rules") else [])
    assert any(getattr(r, "type", "") == "iconSet" for r in rules)


@_register("cf_iconset_extended_attrs")
def _probe_cf_iconset_extended_attrs(tmp_path: Path) -> None:
    """4-icon, number-threshold icon set with percent/reverse flags."""
    import openpyxl as _opx
    import wolfxl
    from wolfxl.formatting.rule import IconSetRule

    wb = wolfxl.Workbook()
    ws = wb.active
    for i in range(1, 9):
        ws.cell(row=i, column=1, value=i)
    rule = IconSetRule(
        "4Rating",
        "num",
        [1, 3, 5, 7],
        showValue=False,
        percent=False,
        reverse=True,
    )
    ws.conditional_formatting.add("A1:A8", rule)
    out = tmp_path / "iconset_extended.xlsx"
    wb.save(out)

    ref_ws = _opx.load_workbook(out).active
    ref_rules = []
    for cf_range in ref_ws.conditional_formatting:
        ref_rules.extend(ref_ws.conditional_formatting[cf_range])
    icon_rules = [r for r in ref_rules if getattr(r, "type", "") == "iconSet"]
    assert icon_rules, "openpyxl saw no iconSet rule"
    icon_set = icon_rules[0].iconSet
    assert icon_set.iconSet == "4Rating"
    assert icon_set.showValue is False
    assert icon_set.percent is False
    assert icon_set.reverse is True
    assert [cfvo.type for cfvo in icon_set.cfvo] == ["num", "num", "num", "num"]
    assert [float(cfvo.val) for cfvo in icon_set.cfvo] == [1.0, 3.0, 5.0, 7.0]


@_register("cf_data_bars")
def _probe_cf_data_bars(tmp_path: Path) -> None:
    """Data-bar rule. Tracked under G12 (S3)."""
    import openpyxl as _opx
    import wolfxl
    from wolfxl.formatting.rule import DataBarRule

    wb = wolfxl.Workbook()
    ws = wb.active
    for i in range(1, 6):
        ws.cell(row=i, column=1, value=i)
    rule = DataBarRule(start_type="min", end_type="max", color="FF0000")
    ws.conditional_formatting.add("A1:A5", rule)
    out = tmp_path / "databar.xlsx"
    wb.save(out)

    wb2 = wolfxl.load_workbook(out)
    rules = []
    for cf_range in wb2.active.conditional_formatting:
        rules.extend(cf_range.rules if hasattr(cf_range, "rules") else [])
    assert any(getattr(r, "type", "") == "dataBar" for r in rules)

    ref_wb = _opx.load_workbook(out)
    ref_ws = ref_wb[ref_wb.sheetnames[0]]
    ref_rules = []
    for cf_range in ref_ws.conditional_formatting:
        ref_rules.extend(ref_ws.conditional_formatting[cf_range])
    bar_rules = [r for r in ref_rules if getattr(r, "type", "") == "dataBar"]
    assert bar_rules, "openpyxl saw no dataBar rule in our output"
    bar = bar_rules[0].dataBar
    assert bar is not None and bar.cfvo, "dataBar payload empty after openpyxl reload"
    assert {c.type for c in bar.cfvo} == {"min", "max"}, (
        f"cfvo types lost: {[c.type for c in bar.cfvo]}"
    )


@_register("cf_data_bars_advanced")
def _probe_cf_data_bars_advanced(tmp_path: Path) -> None:
    """Data-bar with percent / formula cfvo + showValue=False. G12 (S3)."""
    import openpyxl as _opx
    import wolfxl
    from wolfxl.formatting.rule import DataBarRule

    wb = wolfxl.Workbook()
    ws = wb.active
    for i in range(1, 11):
        ws.cell(row=i, column=1, value=i * 10)
    rule = DataBarRule(
        start_type="percent",
        start_value=10,
        end_type="num",
        end_value=90,
        color="FF638EC6",
        showValue=False,
    )
    ws.conditional_formatting.add("A1:A10", rule)
    out = tmp_path / "databar_adv.xlsx"
    wb.save(out)

    ref_wb = _opx.load_workbook(out)
    ref_ws = ref_wb[ref_wb.sheetnames[0]]
    ref_rules = []
    for cf_range in ref_ws.conditional_formatting:
        ref_rules.extend(ref_ws.conditional_formatting[cf_range])
    bar_rules = [r for r in ref_rules if getattr(r, "type", "") == "dataBar"]
    assert bar_rules, "openpyxl saw no dataBar rule"
    bar = bar_rules[0].dataBar
    assert bar is not None and bar.cfvo
    types = {c.type for c in bar.cfvo}
    assert types == {"percent", "num"}, f"cfvo types lost: {types}"
    # Locate min/max by index — openpyxl reads val as a float, so 10 -> 10.0.
    cfvo_min, cfvo_max = bar.cfvo[0], bar.cfvo[1]
    assert float(cfvo_min.val) == 10.0
    assert float(cfvo_max.val) == 90.0
    # showValue=False round-trip
    assert bar.showValue is False


@_register("cf_databar_length_attrs")
def _probe_cf_databar_length_attrs(tmp_path: Path) -> None:
    """DataBarRule minLength/maxLength flags survive openpyxl reload."""
    import openpyxl as _opx
    import wolfxl
    from wolfxl.formatting.rule import DataBarRule

    wb = wolfxl.Workbook()
    ws = wb.active
    for i in range(1, 11):
        ws.cell(row=i, column=1, value=i)
    rule = DataBarRule(
        start_type="formula",
        start_value="$A$1",
        end_type="formula",
        end_value="$A$10",
        color="FF4472C4",
        minLength=5,
        maxLength=90,
    )
    ws.conditional_formatting.add("A1:A10", rule)
    out = tmp_path / "databar_lengths.xlsx"
    wb.save(out)

    ref_ws = _opx.load_workbook(out).active
    ref_rules = []
    for cf_range in ref_ws.conditional_formatting:
        ref_rules.extend(ref_ws.conditional_formatting[cf_range])
    bar_rules = [r for r in ref_rules if getattr(r, "type", "") == "dataBar"]
    assert bar_rules, "openpyxl saw no dataBar rule"
    bar = bar_rules[0].dataBar
    assert bar.minLength == 5
    assert bar.maxLength == 90
    assert [cfvo.type for cfvo in bar.cfvo] == ["formula", "formula"]
    assert [cfvo.val for cfvo in bar.cfvo] == ["$A$1", "$A$10"]


@_register("cf_color_scales_advanced")
def _probe_cf_color_scales_advanced(tmp_path: Path) -> None:
    """3-stop color scale. Tracked under G13 (S3)."""
    import wolfxl
    from wolfxl.formatting.rule import ColorScaleRule

    wb = wolfxl.Workbook()
    ws = wb.active
    for i in range(1, 6):
        ws.cell(row=i, column=1, value=i)
    rule = ColorScaleRule(
        start_type="min",
        start_color="FF0000",
        mid_type="percentile",
        mid_value=50,
        mid_color="FFFF00",
        end_type="max",
        end_color="00FF00",
    )
    ws.conditional_formatting.add("A1:A5", rule)
    out = tmp_path / "cs.xlsx"
    wb.save(out)

    wb2 = wolfxl.load_workbook(out)
    rules = []
    for cf_range in wb2.active.conditional_formatting:
        rules.extend(cf_range.rules if hasattr(cf_range, "rules") else [])
    color_scales = [r for r in rules if getattr(r, "type", "") == "colorScale"]
    assert color_scales, "3-stop color scale not preserved"
    cs = color_scales[0]
    assert hasattr(cs, "colorScale")
    assert len(cs.colorScale.cfvo) == 3


@_register("cf_stop_if_true_priority")
def _probe_cf_stop_if_true_priority(tmp_path: Path) -> None:
    """stopIfTrue + priority + dxf. Tracked under G14 (S3)."""
    import wolfxl
    from wolfxl.formatting.rule import CellIsRule
    from wolfxl.styles import PatternFill

    wb = wolfxl.Workbook()
    ws = wb.active
    for i in range(1, 6):
        ws.cell(row=i, column=1, value=i)
    rule = CellIsRule(
        operator="greaterThan",
        formula=["3"],
        fill=PatternFill(fill_type="solid", fgColor="FFFF00"),
        stopIfTrue=True,
    )
    rule.priority = 1
    ws.conditional_formatting.add("A1:A5", rule)
    out = tmp_path / "cf_stop.xlsx"
    wb.save(out)

    wb2 = wolfxl.load_workbook(out)
    rules = []
    for cf_range in wb2.active.conditional_formatting:
        rules.extend(cf_range.rules if hasattr(cf_range, "rules") else [])
    assert rules, "no rules preserved"
    assert getattr(rules[0], "stopIfTrue", False) is True


# --------------------------------------------------------------------------
# Defined-name probes
# --------------------------------------------------------------------------


@_register("defined_names_basic")
def _probe_defined_names_basic(tmp_path: Path) -> None:
    import wolfxl
    from wolfxl.workbook.defined_name import DefinedName

    wb = wolfxl.Workbook()
    wb.active.title = "Data"
    wb.defined_names["X"] = DefinedName(name="X", value="Data!$A$1:$A$10")

    out = tmp_path / "dn.xlsx"
    wb.save(out)
    wb2 = wolfxl.load_workbook(out)
    assert "X" in wb2.defined_names


@_register("defined_names_edge_cases")
def _probe_defined_names_edge_cases(tmp_path: Path) -> None:
    """Full ECMA-376 §18.2.5 ``definedName`` attribute surface. G22 (Phase 2).

    Round-trips all 13 attributes openpyxl exposes (``hidden``, ``comment``,
    plus the 11 G22 additions) through wolfxl's reader, and additionally
    cross-validates that openpyxl reads what wolfxl wrote.
    """
    import wolfxl
    from wolfxl.workbook.defined_name import DefinedName

    wb = wolfxl.Workbook()
    wb.active.title = "Data"
    dn = DefinedName(
        name="Hidden",
        value="Data!$A$1",
        hidden=True,
        comment="hidden helper",
        customMenu="Custom Menu Text",
        description="A defined name with the full attr surface.",
        help="Press F1.",
        statusBar="Status bar prompt",
        shortcutKey="A",
        function=True,
        functionGroupId=2,
        vbProcedure=True,
        xlm=True,
        publishToServer=True,
        workbookParameter=True,
    )
    wb.defined_names["Hidden"] = dn

    out = tmp_path / "dn_edge.xlsx"
    wb.save(out)

    # Sub-probe 1: wolfxl writes -> wolfxl reads.
    wb2 = wolfxl.load_workbook(out)
    rt = wb2.defined_names["Hidden"]
    assert rt.hidden is True
    assert rt.comment == "hidden helper"
    assert rt.custom_menu == "Custom Menu Text"
    assert rt.description == "A defined name with the full attr surface."
    assert rt.help == "Press F1."
    assert rt.status_bar == "Status bar prompt"
    assert rt.shortcut_key == "A"
    assert rt.function is True
    assert rt.function_group_id == 2
    assert rt.vb_procedure is True
    assert rt.xlm is True
    assert rt.publish_to_server is True
    assert rt.workbook_parameter is True

    # Sub-probe 2: wolfxl writes -> openpyxl reads (cross-tool oracle).
    try:
        import openpyxl as opxl
    except ImportError:
        return
    opwb = opxl.load_workbook(out)
    op_dn = opwb.defined_names.get("Hidden")
    assert op_dn is not None
    assert op_dn.hidden is True or op_dn.hidden == 1
    assert op_dn.comment == "hidden helper"
    assert op_dn.customMenu == "Custom Menu Text"
    assert op_dn.description == "A defined name with the full attr surface."
    assert op_dn.help == "Press F1."
    assert op_dn.statusBar == "Status bar prompt"
    assert op_dn.shortcutKey == "A"
    assert op_dn.function is True or op_dn.function == 1
    assert op_dn.functionGroupId == 2
    assert op_dn.vbProcedure is True or op_dn.vbProcedure == 1
    assert op_dn.xlm is True or op_dn.xlm == 1
    assert op_dn.publishToServer is True or op_dn.publishToServer == 1
    assert op_dn.workbookParameter is True or op_dn.workbookParameter == 1


# --------------------------------------------------------------------------
# Print-settings probes
# --------------------------------------------------------------------------


@_register("print_settings_basic")
def _probe_print_settings_basic(tmp_path: Path) -> None:
    import wolfxl

    wb = wolfxl.Workbook()
    ws = wb.active
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_options.horizontalCentered = True
    out = tmp_path / "page.xlsx"
    wb.save(out)

    wb2 = wolfxl.load_workbook(out)
    assert wb2.active.page_setup.orientation == "landscape"


@_register("print_settings_depth")
def _probe_print_settings_depth(tmp_path: Path) -> None:
    """Deep PageSetup / PrintOptions / PageMargins surface. G24 (S8)."""
    import openpyxl as _opx
    import wolfxl

    wb = wolfxl.Workbook()
    ws = wb.active
    ws.title = "Report"
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.scale = 80
    ws.page_setup.firstPageNumber = 7
    ws.page_setup.fitToWidth = 2
    ws.page_setup.fitToHeight = 3
    ws.page_setup.orientation = "landscape"
    ws.page_setup.horizontalDpi = 300
    ws.page_setup.verticalDpi = 301
    ws.page_setup.cellComments = "atEnd"
    ws.page_setup.useFirstPageNumber = True
    ws.page_setup.errors = "blank"
    ws.page_setup.paperHeight = "297mm"
    ws.page_setup.paperWidth = "210mm"
    ws.page_setup.pageOrder = "overThenDown"
    ws.page_setup.usePrinterDefaults = False
    ws.page_setup.blackAndWhite = True
    ws.page_setup.draft = True
    ws.page_setup.copies = 4
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = False
    ws.print_options.headings = True
    ws.print_options.gridLines = True
    ws.print_options.gridLinesSet = False
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.35
    ws.page_margins.top = 0.45
    ws.page_margins.bottom = 0.55
    ws.page_margins.header = 0.65
    ws.page_margins.footer = 0.75
    ws.print_area = "B2:C4"
    ws.print_title_rows = "1:1"
    ws.print_title_cols = "A:A"
    out = tmp_path / "page_deep.xlsx"
    wb.save(out)

    wb2 = wolfxl.load_workbook(out)
    ws2 = wb2.active
    assert ws2.page_setup.paperSize == 9
    assert ws2.page_setup.scale == 80
    assert ws2.page_setup.firstPageNumber == 7
    assert ws2.page_setup.fitToWidth == 2
    assert ws2.page_setup.fitToHeight == 3
    assert ws2.page_setup.orientation == "landscape"
    assert ws2.page_setup.horizontalDpi == 300
    assert ws2.page_setup.verticalDpi == 301
    assert ws2.page_setup.cellComments == "atEnd"
    assert ws2.page_setup.useFirstPageNumber is True
    assert ws2.page_setup.errors == "blank"
    assert ws2.page_setup.paperHeight == "297mm"
    assert ws2.page_setup.paperWidth == "210mm"
    assert ws2.page_setup.pageOrder == "overThenDown"
    assert ws2.page_setup.usePrinterDefaults is False
    assert ws2.page_setup.blackAndWhite is True
    assert ws2.page_setup.draft is True
    assert ws2.page_setup.copies == 4
    assert ws2.print_options.horizontalCentered is True
    assert ws2.print_options.verticalCentered is False
    assert ws2.print_options.headings is True
    assert ws2.print_options.gridLines is True
    assert ws2.print_options.gridLinesSet is False
    assert ws2.page_margins.left == 0.25
    assert ws2.page_margins.right == 0.35
    assert ws2.page_margins.top == 0.45
    assert ws2.page_margins.bottom == 0.55
    assert ws2.page_margins.header == 0.65
    assert ws2.page_margins.footer == 0.75
    assert ws2.print_area == "Report!B2:C4"
    assert ws2.print_title_rows == "1:1"
    assert ws2.print_title_cols == "A:A"

    op_ws = _opx.load_workbook(out)["Report"]
    assert op_ws.page_setup.paperSize == 9
    assert op_ws.page_setup.scale == 80
    assert op_ws.page_setup.firstPageNumber == 7
    assert op_ws.page_setup.fitToWidth == 2
    assert op_ws.page_setup.fitToHeight == 3
    assert op_ws.page_setup.orientation == "landscape"
    assert op_ws.page_setup.horizontalDpi == 300
    assert op_ws.page_setup.verticalDpi == 301
    assert op_ws.page_setup.cellComments == "atEnd"
    assert op_ws.page_setup.useFirstPageNumber is True
    assert op_ws.page_setup.errors == "blank"
    assert op_ws.page_setup.paperHeight == "297mm"
    assert op_ws.page_setup.paperWidth == "210mm"
    assert op_ws.page_setup.pageOrder == "overThenDown"
    assert op_ws.page_setup.usePrinterDefaults is False
    assert op_ws.page_setup.blackAndWhite is True
    assert op_ws.page_setup.draft is True
    assert op_ws.page_setup.copies == 4
    assert op_ws.print_options.horizontalCentered is True
    assert op_ws.print_options.verticalCentered is False
    assert op_ws.print_options.headings is True
    assert op_ws.print_options.gridLines is True
    assert op_ws.print_options.gridLinesSet is False
    assert op_ws.page_margins.left == 0.25
    assert op_ws.page_margins.right == 0.35
    assert op_ws.page_margins.top == 0.45
    assert op_ws.page_margins.bottom == 0.55
    assert op_ws.page_margins.header == 0.65
    assert op_ws.page_margins.footer == 0.75
    assert op_ws.print_area == "'Report'!$B$2:$C$4"
    assert op_ws.print_title_rows == "$1:$1"
    assert op_ws.print_title_cols == "$A:$A"


# --------------------------------------------------------------------------
# Array / data-table formula probes
# --------------------------------------------------------------------------


@_register("array_formula_basic")
def _probe_array_formula_basic(tmp_path: Path) -> None:
    """Basic ArrayFormula round-trip. G07 (S1)."""
    import openpyxl as _opx
    from openpyxl.worksheet.formula import ArrayFormula as _OpxArrayFormula
    import wolfxl
    from wolfxl.worksheet.formula import ArrayFormula

    wb = wolfxl.Workbook()
    ws = wb.active
    ws["A1"] = 1
    ws["A2"] = 2
    ws["B1"] = ArrayFormula(ref="B1:B2", text="=A1:A2*2")
    out = tmp_path / "arr.xlsx"
    wb.save(out)

    wb2 = wolfxl.load_workbook(out)
    val = wb2.active["B1"].value
    assert isinstance(val, ArrayFormula) or "A1:A2" in str(val)

    ref_wb = _opx.load_workbook(out)
    ref_val = ref_wb[ref_wb.sheetnames[0]]["B1"].value
    assert isinstance(ref_val, _OpxArrayFormula), (
        f"openpyxl saw {type(ref_val).__name__}, expected ArrayFormula"
    )
    assert ref_val.ref == "B1:B2", f"ref lost: {ref_val.ref!r}"
    assert "A1:A2" in (ref_val.text or ""), f"formula lost: {ref_val.text!r}"


@_register("array_formula_data_table")
def _probe_array_formula_data_table(tmp_path: Path) -> None:
    """DataTableFormula round-trip. G07 (S1)."""
    import wolfxl
    from wolfxl.worksheet.formula import DataTableFormula

    wb = wolfxl.Workbook()
    ws = wb.active
    ws["A1"] = 5
    ws["B1"] = 10
    ws["C1"] = DataTableFormula(ref="C1:C3", t="dataTable", r1="A1", dt2D=False)
    out = tmp_path / "dt.xlsx"
    wb.save(out)

    wb2 = wolfxl.load_workbook(out)
    val = wb2.active["C1"].value
    assert val is not None


@_register("array_formula_spill_metadata")
def _probe_array_formula_spill_metadata(tmp_path: Path) -> None:
    """openpyxl 3.1.x array/spill metadata surface parity."""
    import openpyxl as _opx
    from openpyxl.worksheet.formula import DataTableFormula as _OpxDataTableFormula
    import wolfxl
    from wolfxl.worksheet.formula import ArrayFormula, DataTableFormula

    wb = wolfxl.Workbook()
    ws = wb.active
    ws["A1"] = ArrayFormula(ref="A1:A3")
    ws["B2"] = DataTableFormula(ref="B2:B4", del1=True, del2=True, aca=True)
    out = tmp_path / "array_spill_metadata.xlsx"
    wb.save(out)

    wolf_rt = wolfxl.load_workbook(out)
    wolf_ws = wolf_rt.active
    assert wolf_ws.array_formulae == {"A1": "A1:A3"}
    wolf_dt = wolf_ws["B2"].value
    assert isinstance(wolf_dt, DataTableFormula)
    assert wolf_dt.del1 is True
    assert wolf_dt.del2 is True

    op_rt = _opx.load_workbook(out)
    op_ws = op_rt.active
    assert op_ws.array_formulae == {"A1": "A1:A3"}
    op_dt = op_ws["B2"].value
    assert isinstance(op_dt, _OpxDataTableFormula)
    assert str(op_dt.del1).lower() in {"1", "true"}
    assert str(op_dt.del2).lower() in {"1", "true"}


# --------------------------------------------------------------------------
# Calc-chain probes
# --------------------------------------------------------------------------


@_register("calc_chain_basic")
def _probe_calc_chain_basic(tmp_path: Path) -> None:
    import openpyxl as _opx
    import wolfxl

    src = tmp_path / "calc_seed.xlsx"
    out = tmp_path / "calc_out.xlsx"
    wb = _opx.Workbook()
    ws = wb.active
    ws["A1"] = 1
    ws["A2"] = 2
    ws["A3"] = "=SUM(A1:A2)"
    wb.save(src)

    wb2 = wolfxl.load_workbook(src, modify=True)
    wb2.active["B1"] = "dirty"
    wb2.save(out)
    wb2.close()

    entries = _zip_listing(out)
    assert "xl/calcChain.xml" in entries
    calc_chain = _zip_read_text(out, "xl/calcChain.xml")
    assert 'r="A3"' in calc_chain
    _opx.load_workbook(out, data_only=False)


@_register("calc_chain_edge_cases")
def _probe_calc_chain_edge_cases(tmp_path: Path) -> None:
    import openpyxl as _opx
    import wolfxl

    src = tmp_path / "calc_edge_seed.xlsx"
    out = tmp_path / "calc_edge_out.xlsx"
    wb = _opx.Workbook()
    first = wb.active
    first.title = "First"
    first["A1"] = 1
    first["A2"] = 2
    first["B1"] = "=SUM(A1:A2)"
    first["B4"] = "=Second!A1"
    second = wb.create_sheet("Second")
    second["A1"] = 10
    second["B2"] = "=First!B1+A1"
    wb.save(src)

    rewritten = tmp_path / "calc_edge_seed_rewritten.xlsx"
    with zipfile.ZipFile(src, "r") as zsrc, zipfile.ZipFile(
        rewritten, "w", compression=zipfile.ZIP_DEFLATED
    ) as zdst:
        for info in zsrc.infolist():
            if info.filename != "xl/calcChain.xml":
                zdst.writestr(info, zsrc.read(info.filename))
        zdst.writestr(
            "xl/calcChain.xml",
            """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<calcChain xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <c r="B1" i="1"/>
  <c r="B4" i="1"/>
  <c r="X99" i="2"/>
  <extLst><ext uri="{wolfxl-test-calcchain-ext}"><x:test xmlns:x="urn:wolfxl:test">keep</x:test></ext></extLst>
</calcChain>""",
        )
    shutil.move(rewritten, src)

    wb2 = wolfxl.load_workbook(src, modify=True)
    wb2["First"].delete_rows(4)
    wb2.save(out)
    wb2.close()

    calc_chain = _zip_read_text(out, "xl/calcChain.xml")
    assert 'r="B1" i="1"' in calc_chain
    assert 'r="B2" i="2"' in calc_chain
    assert 'r="B4"' not in calc_chain
    assert 'r="X99"' not in calc_chain
    assert "{wolfxl-test-calcchain-ext}" in calc_chain
    assert "urn:wolfxl:test" in calc_chain
    _opx.load_workbook(out, data_only=False)


# --------------------------------------------------------------------------
# Slicer probes
# --------------------------------------------------------------------------


@_register("slicers_with_pivot")
def _probe_slicers_with_pivot(tmp_path: Path) -> None:
    import openpyxl
    import wolfxl
    from wolfxl.chart import Reference
    from wolfxl.pivot import Slicer, SlicerCache
    from wolfxl.pivot import PivotCache, PivotTable

    seed = tmp_path / "slicer_seed.xlsx"
    out = tmp_path / "pivot_slicer.xlsx"
    seed_wb = openpyxl.Workbook()
    seed_ws = seed_wb.active
    seed_ws.title = "Data"
    for row in [
        ("region", "quarter", "revenue"),
        ("North", "Q1", 100.0),
        ("South", "Q1", 200.0),
        ("North", "Q2", 150.0),
        ("South", "Q2", 250.0),
    ]:
        seed_ws.append(row)
    seed_wb.save(seed)

    wb = wolfxl.load_workbook(seed, modify=True)
    try:
        ws = wb["Data"]
        ref = Reference(ws, min_col=1, min_row=1, max_col=3, max_row=5)
        cache = PivotCache(source=ref)
        pivot = PivotTable(cache=cache, location="F2", rows=["region"], data=["revenue"])
        wb.add_pivot_cache(cache)
        ws.add_pivot_table(pivot)
        slicer_cache = SlicerCache(name="Slicer_region", source_pivot_cache=cache, field="region")
        slicer = Slicer(name="Slicer_region1", cache=slicer_cache, caption="Region")
        wb.add_slicer_cache(slicer_cache)
        ws.add_slicer(slicer, anchor="H2")
        wb.save(out)
    finally:
        wb.close()

    entries = _zip_listing(out)
    assert any(re.match(r"^xl/slicerCaches/slicerCache\d+\.xml$", name) for name in entries)
    assert any(re.match(r"^xl/slicers/slicer\d+\.xml$", name) for name in entries)
    rels_xml = "\n".join(
        _zip_read_text(out, name)
        for name in entries
        if re.match(r"^xl/worksheets/_rels/sheet\d+\.xml\.rels$", name)
    )
    assert "office/2007/relationships/slicer" in rels_xml


# --------------------------------------------------------------------------
# Test runner / parametrisation
# --------------------------------------------------------------------------


def _expand_probe_entries() -> list[dict[str, Any]]:
    """Yield one parametrise entry per `probe` plus per `secondary_probes`.

    Secondary probes share the parent's status / gap_id but parametrise
    under a synthesised id (`{parent}.{secondary}`) so the harness can
    flip them independently when a follow-up gap closes (e.g. the
    bounded-memory check arrives one sprint after the basic streaming
    probe).
    """
    expanded: list[dict[str, Any]] = []
    for entry in _SPEC.ENTRIES:
        if entry.get("probe"):
            expanded.append(entry)
            for secondary in entry.get("secondary_probes", []):
                expanded.append(
                    {
                        **entry,
                        "id": f"{entry['id']}.{secondary}",
                        "probe": secondary,
                    }
                )
    return expanded


_PROBE_ENTRIES = _expand_probe_entries()


def _id_for(entry: dict[str, Any]) -> str:
    return f"{entry['id']}[{entry['status']}]"


@pytest.mark.parametrize("entry", _PROBE_ENTRIES, ids=_id_for)
def test_compat_oracle_probe(
    entry: dict[str, Any], tmp_path: Path, request: pytest.FixtureRequest
) -> None:
    """Run one compat-oracle probe per spec entry that carries a `probe` field.

    Outcome by status:
    * ``supported``    → must pass; failure is a regression.
    * ``partial``      → xfail today; flips xpass once the gap closes.
    * ``not_yet``      → xfail today; flips xpass once the gap closes.
    * ``out_of_scope`` → skipped.
    """
    probe_name = entry["probe"]
    probe_fn = PROBES.get(probe_name)
    if probe_fn is None:
        pytest.skip(f"probe '{probe_name}' not registered yet")

    if entry["status"] == "out_of_scope":
        pytest.skip(f"out of scope: {entry['id']}")
    if entry["status"] in ("not_yet", "partial"):
        gap = entry.get("gap_id", "?")
        request.node.add_marker(
            pytest.mark.xfail(
                reason=(
                    f"compat-oracle baseline gap: {entry['id']} "
                    f"({entry['status']}, {gap}); flip when the gap closes"
                ),
                strict=False,
            )
        )

    probe_fn(tmp_path)


# --------------------------------------------------------------------------
# Session summary
# --------------------------------------------------------------------------


def _baseline_path() -> Path:
    return REPO_ROOT / ".pytest_cache" / "compat_oracle_baseline.json"


@pytest.fixture(scope="session", autouse=True)
def _emit_oracle_summary(request: pytest.FixtureRequest) -> Any:
    yield
    terminal = request.config.pluginmanager.get_plugin("terminalreporter")
    if terminal is None:  # pragma: no cover - pytest internals
        return

    counts: Counter[str] = Counter()
    for outcome in ("passed", "failed", "skipped", "xfailed", "xpassed", "error"):
        for report in terminal.stats.get(outcome, []):
            nodeid = getattr(report, "nodeid", "")
            if "test_compat_oracle_probe" not in nodeid:
                continue
            counts[outcome] += 1

    if not counts:
        return

    relevant = (
        counts["passed"]
        + counts["failed"]
        + counts["xfailed"]
        + counts["xpassed"]
        + counts["error"]
    )
    if relevant == 0:
        return

    pass_total = counts["passed"] + counts["xpassed"]
    pct = (pass_total / relevant) * 100

    terminal.write_sep("=", "openpyxl-compat oracle (Sprint 0 baseline)")
    terminal.write_line(
        f"  total probes:      {relevant}",
    )
    terminal.write_line(
        f"  passed (green):    {counts['passed']}",
    )
    terminal.write_line(
        f"  xpassed:           {counts['xpassed']} (gap closed; flip status to 'supported')",
    )
    terminal.write_line(
        f"  xfailed (gap):     {counts['xfailed']}",
    )
    terminal.write_line(
        f"  failed:            {counts['failed']}",
    )
    terminal.write_line(
        f"  pass rate:         {pct:.1f}% ({pass_total}/{relevant})",
    )

    if os.environ.get("WOLFXL_COMPAT_ORACLE_WRITE_BASELINE") == "1":
        baseline = {
            "date": date.today().isoformat(),
            "total": relevant,
            "passed": counts["passed"],
            "xpassed": counts["xpassed"],
            "xfailed": counts["xfailed"],
            "failed": counts["failed"],
            "pass_rate_pct": round(pct, 2),
        }
        path = _baseline_path()
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(json.dumps(baseline, indent=2) + "\n")
        terminal.write_line(f"  wrote baseline:    {path.relative_to(REPO_ROOT)}")
