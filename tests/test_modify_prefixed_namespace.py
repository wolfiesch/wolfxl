from __future__ import annotations

import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

import openpyxl

import wolfxl

MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"


def test_modify_inserted_cell_uses_prefixed_worksheet_namespace(tmp_path: Path) -> None:
    src = tmp_path / "prefixed.xlsx"
    _make_prefixed_namespace_workbook(src)

    workbook = wolfxl.load_workbook(src, modify=True)
    workbook["Data"]["J1"] = "wolfxl_modify_smoke"
    workbook.save(src)
    workbook.close()

    roundtrip = openpyxl.load_workbook(src)
    assert roundtrip["Data"]["J1"].value == "wolfxl_modify_smoke"
    roundtrip.close()

    with zipfile.ZipFile(src) as archive:
        sheet_xml = archive.read("xl/worksheets/sheet1.xml").decode()
    assert '<x:c r="J1" t="str"><x:v>wolfxl_modify_smoke</x:v></x:c>' in sheet_xml
    assert '<c r="J1"' not in sheet_xml


def test_insert_rows_preserves_prefixed_worksheet_end_tags(tmp_path: Path) -> None:
    src = tmp_path / "prefixed.xlsx"
    _make_prefixed_namespace_workbook(src)

    workbook = wolfxl.load_workbook(src, modify=True)
    workbook["Data"].insert_rows(1, amount=1)
    workbook.save(src)
    workbook.close()

    roundtrip = openpyxl.load_workbook(src)
    assert roundtrip["Data"]["A2"].value == "Region"
    roundtrip.close()

    with zipfile.ZipFile(src) as archive:
        sheet_xml = archive.read("xl/worksheets/sheet1.xml").decode()
    assert "<x:worksheet" in sheet_xml
    assert "</x:worksheet>" in sheet_xml
    assert "</worksheet>" not in sheet_xml


def _make_prefixed_namespace_workbook(path: Path) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet["A1"] = "Region"
    workbook.save(path)

    ET.register_namespace("x", MAIN_NS)
    with zipfile.ZipFile(path, "r") as source:
        entries = {name: source.read(name) for name in source.namelist()}

    root = ET.fromstring(entries["xl/worksheets/sheet1.xml"])
    entries["xl/worksheets/sheet1.xml"] = ET.tostring(
        root,
        encoding="utf-8",
        xml_declaration=True,
    )

    rewritten = path.with_suffix(".prefixed.xlsx")
    with zipfile.ZipFile(rewritten, "w", zipfile.ZIP_DEFLATED) as target:
        for name, data in entries.items():
            target.writestr(name, data)
    rewritten.replace(path)
