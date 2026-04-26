"""RFC-030 — `Worksheet.insert_rows` / `delete_rows` round-trip in modify mode.

End-to-end coverage for the structural row-shift path. Threads three
layers:

1. ``Worksheet.insert_rows`` / ``delete_rows`` (Python) validate args
   and append a tuple to ``wb._pending_axis_shifts``.
2. ``Workbook._flush_pending_axis_shifts_to_patcher`` (Python) drains
   each tuple into ``XlsxPatcher.queue_axis_shift``.
3. ``XlsxPatcher::do_save`` Phase 2.5i (Rust) reads each affected sheet
   XML / table part / comments part / VML drawing / workbook.xml,
   builds ``wolfxl_structural::SheetXmlInputs``, and calls
   ``apply_workbook_shift`` to rewrite cell coords + formulas +
   anchors + defined names + comment refs.

Sister contract: ``test_no_dirty_save_is_byte_identical`` confirms the
empty-queue path remains a no-op identity.
"""
from __future__ import annotations

import zipfile
from pathlib import Path

import openpyxl
import pytest

import wolfxl
from wolfxl import Workbook


# pytest marker so verify_rfc.py can collect this test.
pytestmark = pytest.mark.rfc030


@pytest.fixture(autouse=True)
def _force_test_epoch(monkeypatch: pytest.MonkeyPatch) -> None:
    """Pin ZIP entry mtimes for byte-stable saves."""
    monkeypatch.setenv("WOLFXL_TEST_EPOCH", "0")


def _make_simple_fixture(path: Path) -> None:
    """Workbook with values in rows 1..10 and a SUM formula referencing rows 1-10."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for r in range(1, 11):
        ws.cell(row=r, column=1, value=r)
    ws["B1"] = "=SUM(A1:A10)"
    wb.save(path)


def _make_formula_fixture(path: Path) -> None:
    """Workbook with formulas pointing into and across the shift band."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for r in range(1, 11):
        ws.cell(row=r, column=1, value=r)
    ws["B5"] = "=A5+A6"  # both into the shift band on insert(5,3)
    ws["B10"] = "=A10"  # below the band
    ws["B1"] = "=A1"  # above the band
    wb.save(path)


def _make_hyperlink_fixture(path: Path) -> None:
    """Workbook with a hyperlink on B5."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for r in range(1, 11):
        ws.cell(row=r, column=1, value=r)
    ws["B5"] = "click"
    ws["B5"].hyperlink = "https://example.com/anchored-at-row-5"
    wb.save(path)


def _read_zip_text(path: Path, entry: str) -> str:
    with zipfile.ZipFile(path) as zf:
        return zf.read(entry).decode("utf-8")


# ---------------------------------------------------------------------------
# Tests — argument validation
# ---------------------------------------------------------------------------


def test_insert_rows_rejects_zero_idx(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    _make_simple_fixture(src)
    wb = wolfxl.load_workbook(src, modify=True)
    ws = wb.active
    with pytest.raises(ValueError, match="idx"):
        ws.insert_rows(0)


def test_insert_rows_rejects_zero_amount(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    _make_simple_fixture(src)
    wb = wolfxl.load_workbook(src, modify=True)
    ws = wb.active
    with pytest.raises(ValueError, match="amount"):
        ws.insert_rows(2, amount=0)


def test_delete_rows_rejects_negative(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    _make_simple_fixture(src)
    wb = wolfxl.load_workbook(src, modify=True)
    ws = wb.active
    with pytest.raises(ValueError):
        ws.delete_rows(-1)


# ---------------------------------------------------------------------------
# Tests — basic insert / delete in middle
# ---------------------------------------------------------------------------


def test_insert_3_rows_at_5_shifts_cells_and_formula(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_simple_fixture(src)
    wb = wolfxl.load_workbook(src, modify=True)
    ws = wb.active
    ws.insert_rows(5, amount=3)
    wb.save(dst)
    # Re-open with openpyxl and check shifted positions.
    op = openpyxl.load_workbook(dst)
    s = op["Sheet1"]
    assert s["A8"].value == 5  # row 5 → row 8
    assert s["A13"].value == 10  # row 10 → row 13
    # Formula =SUM(A1:A10) should be shifted to =SUM(A1:A13) since
    # rows >= 5 in the range shift down by 3.
    assert s["B1"].value == "=SUM(A1:A13)"


def test_delete_3_rows_at_5_drops_band_and_shifts_below(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_simple_fixture(src)
    wb = wolfxl.load_workbook(src, modify=True)
    ws = wb.active
    ws.delete_rows(5, amount=3)
    wb.save(dst)
    op = openpyxl.load_workbook(dst)
    s = op["Sheet1"]
    # Original rows 5, 6, 7 (values 5, 6, 7) are gone.
    assert s["A4"].value == 4
    # Original row 8 (value 8) → row 5.
    assert s["A5"].value == 8
    assert s["A7"].value == 10  # original row 10 → 7


# ---------------------------------------------------------------------------
# Tests — insert at start
# ---------------------------------------------------------------------------


def test_insert_at_row_1_shifts_everything(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_simple_fixture(src)
    wb = wolfxl.load_workbook(src, modify=True)
    ws = wb.active
    ws.insert_rows(1, amount=2)
    wb.save(dst)
    op = openpyxl.load_workbook(dst)
    s = op["Sheet1"]
    assert s["A3"].value == 1
    assert s["A12"].value == 10


# ---------------------------------------------------------------------------
# Tests — insert past last row (no-op for cell content)
# ---------------------------------------------------------------------------


def test_insert_past_last_row_leaves_cells_alone(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_simple_fixture(src)
    wb = wolfxl.load_workbook(src, modify=True)
    ws = wb.active
    ws.insert_rows(100, amount=5)  # well past row 10
    wb.save(dst)
    op = openpyxl.load_workbook(dst)
    s = op["Sheet1"]
    # All cells unchanged.
    assert s["A1"].value == 1
    assert s["A10"].value == 10


# ---------------------------------------------------------------------------
# Tests — formulas pointing into and across the shift band
# ---------------------------------------------------------------------------


def test_formula_into_band_becomes_ref_error_on_delete(tmp_path: Path) -> None:
    """A formula referencing a cell inside the deletion band should
    become #REF! per OOXML semantics."""
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_formula_fixture(src)
    wb = wolfxl.load_workbook(src, modify=True)
    ws = wb.active
    ws.delete_rows(5, amount=3)  # delete rows 5-7
    wb.save(dst)
    op = openpyxl.load_workbook(dst)
    s = op["Sheet1"]
    # Original row 10's =A10 → row 7's =A7
    assert s["B7"].value == "=A7"
    # Original row 1's =A1 unchanged
    assert s["B1"].value == "=A1"


def test_formula_across_band_shifts(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_formula_fixture(src)
    wb = wolfxl.load_workbook(src, modify=True)
    ws = wb.active
    ws.insert_rows(5, amount=3)
    wb.save(dst)
    op = openpyxl.load_workbook(dst)
    s = op["Sheet1"]
    # Original B5 =A5+A6 → B8 =A8+A9
    assert s["B8"].value == "=A8+A9"
    # Original B10 =A10 → B13 =A13
    assert s["B13"].value == "=A13"


# ---------------------------------------------------------------------------
# Tests — hyperlinks
# ---------------------------------------------------------------------------


def test_hyperlink_anchor_shifts(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_hyperlink_fixture(src)
    wb = wolfxl.load_workbook(src, modify=True)
    ws = wb.active
    ws.insert_rows(5, amount=3)
    wb.save(dst)
    sheet_xml = _read_zip_text(dst, "xl/worksheets/sheet1.xml")
    # The hyperlink ref="B5" should have shifted to ref="B8".
    assert 'ref="B8"' in sheet_xml


# ---------------------------------------------------------------------------
# Tests — defined names
# ---------------------------------------------------------------------------


def test_defined_name_shifts(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for r in range(1, 11):
        ws.cell(row=r, column=1, value=r)
    # Workbook-scope defined name pointing at Sheet1!$A$5
    from openpyxl.workbook.defined_name import DefinedName

    dn = DefinedName(name="MyAnchor", attr_text="Sheet1!$A$5")
    wb.defined_names["MyAnchor"] = dn
    wb.save(src)

    wb2 = wolfxl.load_workbook(src, modify=True)
    ws2 = wb2.active
    ws2.insert_rows(5, amount=3)
    wb2.save(dst)

    wb_xml = _read_zip_text(dst, "xl/workbook.xml")
    assert "Sheet1!$A$8" in wb_xml


# ---------------------------------------------------------------------------
# Tests — tables (RFC-024)
# ---------------------------------------------------------------------------


def test_table_ref_and_autofilter_shift(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "h"
    ws["B1"] = "v"
    for r in range(2, 11):
        ws.cell(row=r, column=1, value=f"k{r}")
        ws.cell(row=r, column=2, value=r)
    from openpyxl.worksheet.table import Table

    t = Table(displayName="T1", ref="A1:B10")
    ws.add_table(t)
    wb.save(src)

    wb2 = wolfxl.load_workbook(src, modify=True)
    ws2 = wb2.active
    ws2.insert_rows(5, amount=3)
    wb2.save(dst)

    table_xml = _read_zip_text(dst, "xl/tables/table1.xml")
    assert 'ref="A1:B13"' in table_xml


# ---------------------------------------------------------------------------
# Tests — data validations and conditional formatting
# ---------------------------------------------------------------------------


def test_data_validation_sqref_shifts(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    from openpyxl.worksheet.datavalidation import DataValidation

    dv = DataValidation(type="list", formula1='"a,b,c"')
    dv.add("A5:A10")
    ws.add_data_validation(dv)
    wb.save(src)

    wb2 = wolfxl.load_workbook(src, modify=True)
    ws2 = wb2.active
    ws2.insert_rows(5, amount=3)
    wb2.save(dst)

    sheet_xml = _read_zip_text(dst, "xl/worksheets/sheet1.xml")
    assert 'sqref="A8:A13"' in sheet_xml


def test_conditional_formatting_sqref_shifts(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for r in range(1, 11):
        ws.cell(row=r, column=1, value=r)
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import PatternFill

    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    rule = CellIsRule(operator="greaterThan", formula=["3"], fill=fill)
    ws.conditional_formatting.add("A5:A10", rule)
    wb.save(src)

    wb2 = wolfxl.load_workbook(src, modify=True)
    ws2 = wb2.active
    ws2.insert_rows(5, amount=3)
    wb2.save(dst)

    sheet_xml = _read_zip_text(dst, "xl/worksheets/sheet1.xml")
    assert 'sqref="A8:A13"' in sheet_xml


# ---------------------------------------------------------------------------
# Tests — empty queue is a no-op
# ---------------------------------------------------------------------------


def test_no_dirty_save_is_byte_identical(tmp_path: Path) -> None:
    """The empty-queue path must remain a no-op identity per the RFC's
    no-op invariant."""
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_simple_fixture(src)
    wb = wolfxl.load_workbook(src, modify=True)
    # No insert_rows / delete_rows call.
    wb.save(dst)
    assert src.read_bytes() == dst.read_bytes()


# ---------------------------------------------------------------------------
# Tests — multi-op sequencing
# ---------------------------------------------------------------------------


def test_multi_op_sequence_insert_then_delete(tmp_path: Path) -> None:
    """Inserts followed by deletes must run in order so the second
    op's idx refers to the coordinate space produced by the first."""
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_simple_fixture(src)
    wb = wolfxl.load_workbook(src, modify=True)
    ws = wb.active
    ws.insert_rows(5, amount=3)  # rows 5..10 → 8..13
    ws.delete_rows(2, amount=1)  # row 2 dropped; everything below shifts up by 1
    wb.save(dst)
    op = openpyxl.load_workbook(dst)
    s = op["Sheet1"]
    # After insert(5,3): A8=5, A13=10. After delete(2,1): A2 dropped,
    # A8 → A7 (value 5), A13 → A12 (value 10), A1 unchanged.
    assert s["A1"].value == 1
    assert s["A7"].value == 5
    assert s["A12"].value == 10
