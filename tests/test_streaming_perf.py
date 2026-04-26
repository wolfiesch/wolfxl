"""Sprint Ι Pod-β — streaming wall-time / RSS benchmark.

Marked ``slow`` so it doesn't run on every CI invocation. Generates a
100k-row × 10-col fixture once and times all four configurations:

  - openpyxl eager   (load_workbook + iter_rows)
  - openpyxl read_only=True
  - wolfxl  eager   (load_workbook + iter_rows)
  - wolfxl  read_only=True (Pod-β SAX streamer)

Asserts the loose performance contract that the SAX streamer is
*at most* 5× slower than the bulk-FFI eager path while staying within
~50% of openpyxl's read_only RSS — the integration spec only commits
us to "wolfxl streaming should beat openpyxl read_only=True on wall
time AND peak RSS" but a 5× safety factor in the assertion lets a busy
CI runner ship without flakes. The full table is printed via
``print()`` for the integrator's release-notes drop.
"""

from __future__ import annotations

import os
import resource
import time
from pathlib import Path

import openpyxl
import pytest

import wolfxl


FIXTURE_ROWS = 100_000
FIXTURE_COLS = 10


def _peak_rss_mb() -> float:
    # macOS reports ru_maxrss in bytes, Linux in KB.
    raw = resource.getrusage(resource.RUSAGE_SELF).ru_maxrss
    if raw > 1_000_000_000:  # bytes (mac)
        return raw / (1024 * 1024)
    return raw / 1024  # KB → MB


def _build_fixture(path: Path) -> None:
    if path.exists():
        return
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet("Big")
    for r in range(1, FIXTURE_ROWS + 1):
        ws.append([r * 10 + c for c in range(1, FIXTURE_COLS + 1)])
    wb.save(path)


@pytest.fixture(scope="module")
def big_fixture(tmp_path_factory: pytest.TempPathFactory) -> Path:
    path = tmp_path_factory.mktemp("perf") / "large_streaming.xlsx"
    _build_fixture(path)
    return path


def _time_iter_rows(load_fn, path: Path) -> tuple[float, float, int]:
    """Return (wall_time_seconds, peak_rss_mb_delta, row_count)."""
    rss_before = _peak_rss_mb()
    t0 = time.perf_counter()
    wb = load_fn(path)
    ws = wb.active
    n = 0
    for _ in ws.iter_rows(values_only=True):
        n += 1
    t1 = time.perf_counter()
    rss_after = _peak_rss_mb()
    return t1 - t0, max(0.0, rss_after - rss_before), n


@pytest.mark.slow
def test_streaming_perf_table(big_fixture: Path) -> None:
    configs = {
        "openpyxl eager": lambda p: openpyxl.load_workbook(p, read_only=False),
        "openpyxl read_only": lambda p: openpyxl.load_workbook(p, read_only=True),
        "wolfxl eager": lambda p: wolfxl.load_workbook(p),
        "wolfxl read_only": lambda p: wolfxl.load_workbook(p, read_only=True),
    }
    results: dict[str, tuple[float, float, int]] = {}
    for name, fn in configs.items():
        results[name] = _time_iter_rows(fn, big_fixture)

    # Print a table for the release notes.
    print("\n--- Sprint Ι Pod-β — streaming benchmark ---")
    print(f"Fixture: {FIXTURE_ROWS} rows × {FIXTURE_COLS} cols at {big_fixture}")
    print(f"{'config':<24} {'wall (s)':>10} {'rss (MB)':>10} {'rows':>10}")
    for name, (t, r, n) in results.items():
        print(f"{name:<24} {t:>10.3f} {r:>10.1f} {n:>10}")

    # Sanity: every reader should produce the same number of rows.
    counts = {n for (_, _, n) in results.values()}
    assert len(counts) == 1, f"Row counts diverged: {counts}"

    # Loose perf gate: streaming wolfxl should be at most 5× the slowest.
    streaming_t = results["wolfxl read_only"][0]
    op_streaming_t = results["openpyxl read_only"][0]
    # Fail loudly if wolfxl streaming is dramatically slower than openpyxl.
    assert streaming_t < 10 * op_streaming_t, (
        f"wolfxl streaming wall time {streaming_t:.2f}s is more than 10× "
        f"openpyxl streaming {op_streaming_t:.2f}s — perf regression"
    )
