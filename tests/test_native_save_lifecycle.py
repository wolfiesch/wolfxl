"""Native write-mode save lifecycle parity.

Openpyxl eager workbooks can be edited and saved repeatedly. WolfXL's native
writer must match that behavior; only ``Workbook(write_only=True)`` is
consumed-on-save.
"""
from __future__ import annotations

from pathlib import Path


def test_eager_workbook_can_save_again_after_more_edits(tmp_path: Path) -> None:
    """A successful eager save does not consume the workbook."""
    import wolfxl
    from openpyxl import load_workbook

    wb = wolfxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.cell(row=1, column=1, value="hello")
    out1 = tmp_path / "first.xlsx"
    wb.save(str(out1))
    assert out1.exists() and out1.stat().st_size > 0

    ws.cell(row=1, column=2, value="again")
    out2 = tmp_path / "second.xlsx"
    wb.save(str(out2))

    reloaded = load_workbook(out2)
    assert reloaded.active["A1"].value == "hello"
    assert reloaded.active["B1"].value == "again"


def test_eager_workbook_can_retry_after_failed_save(tmp_path: Path) -> None:
    """A failed eager save should not make openpyxl-style retry impossible."""
    import wolfxl
    from openpyxl import load_workbook

    wb = wolfxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.cell(row=1, column=1, value="hello")

    # /dev/full on Linux always-fails-write; on macOS we use a directory
    # path that doesn't exist. Either yields a write error from fs::write.
    bad_path = str(tmp_path / "no_such_dir" / "bad.xlsx")
    try:
        wb.save(bad_path)
    except Exception:
        pass
    else:  # pragma: no cover - defensive; the path should fail on all platforms
        raise AssertionError("expected save to nonexistent directory to fail")

    good_path = tmp_path / "good.xlsx"
    wb.save(str(good_path))
    assert load_workbook(good_path).active["A1"].value == "hello"
