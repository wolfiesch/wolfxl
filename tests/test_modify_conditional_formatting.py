"""Modify-mode conditional formatting (RFC-026) — end-to-end coverage.

Mirror of ``tests/test_modify_data_validations.py``. The headline gates:

- ``test_add_cf_preserves_existing`` (RFC §8 risk #1) — the merger uses
  replace-all CF semantics (RFC-011 §5.5), so the patcher must capture
  the source's existing ``<conditionalFormatting>`` blocks via
  ``extract_existing_cf_blocks`` and re-emit them verbatim. This test
  catches any byte-range bug in that capture before users do.
- ``test_dxf_id_monotonic_across_sheets`` — the only new architecture
  this slice introduces: a workbook-wide ``running_dxf_count`` threaded
  through every sheet's CF flush in deterministic sorted-name order, so
  ``xl/styles.xml`` carries one consistent set of ``<dxf>`` entries and
  every ``dxfId`` reference is unique.
"""
from __future__ import annotations

import re
import zipfile
from pathlib import Path

import openpyxl
import pytest
from wolfxl.formatting.rule import (
    CellIsRule,
    ColorScaleRule,
    DataBarRule,
    FormulaRule,
    IconSetRule,
)

from wolfxl import Workbook

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _make_clean_fixture(path: Path, sheet_titles: tuple[str, ...] = ("Sheet1",)) -> None:
    """Workbook with no existing CF rules. Creates the requested sheets."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_titles[0]
    ws["A1"] = "header"
    for title in sheet_titles[1:]:
        wb.create_sheet(title)
    wb.save(path)


def _make_fixture_with_one_cf(path: Path) -> None:
    """Workbook with one existing CellIs CF rule (priority=1, dxfId=0)."""
    from openpyxl.formatting.rule import CellIsRule as OpyCellIs
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "header"
    rule = OpyCellIs(
        operator="greaterThan",
        formula=["100"],
        font=Font(bold=True),
        fill=PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid"),
    )
    ws.conditional_formatting.add("A2:A10", rule)
    wb.save(path)


def _read_zip_text(path: Path, member: str) -> str:
    with zipfile.ZipFile(path) as zf:
        return zf.read(member).decode("utf-8")


def _read_sheet_xml(path: Path, sheet_path: str = "xl/worksheets/sheet1.xml") -> str:
    return _read_zip_text(path, sheet_path)


def _read_styles_xml(path: Path) -> str:
    return _read_zip_text(path, "xl/styles.xml")


def _all_cf_blocks(xml: str) -> list[str]:
    """Return every ``<conditionalFormatting …>…</conditionalFormatting>`` element."""
    return re.findall(
        r"<conditionalFormatting[^>]*?(?:/>|>.*?</conditionalFormatting>)", xml, re.DOTALL
    )


def _cf_block_for_sqref(xml: str, sqref: str) -> str | None:
    """First CF block whose sqref attribute equals ``sqref`` (exact match)."""
    pattern = (
        r'<conditionalFormatting[^>]*?sqref="' + re.escape(sqref) + r'"[^>]*?'
        r"(?:/>|>.*?</conditionalFormatting>)"
    )
    m = re.search(pattern, xml, re.DOTALL)
    return m.group(0) if m else None


def _count_dxfs_in_styles(styles_xml: str) -> int:
    """Count actual ``<dxf>`` children of ``<dxfs>`` (NOT the count attribute)."""
    m = re.search(r"<dxfs[^>]*?>(.*?)</dxfs>", styles_xml, re.DOTALL)
    if m is None:
        return 0
    return len(re.findall(r"<dxf[\s/>]", m.group(1)))


# ---------------------------------------------------------------------------
# Tests — RFC §6 coverage + cross-sheet allocator gate
# ---------------------------------------------------------------------------


def test_add_cellis_rule_to_clean_file(tmp_path: Path) -> None:
    """No existing CF + add one CellIs rule → block emitted with priority=1
    + dxfId=0 + a single ``<dxf>`` appended to styles.xml."""
    src = tmp_path / "src.xlsx"
    _make_clean_fixture(src)

    wb = Workbook._from_patcher(str(src))
    ws = wb["Sheet1"]
    ws.conditional_formatting.add(
        "A1:A10",
        CellIsRule(operator="equal", formula=["1"], extra={"font_bold": True}),
    )
    out = tmp_path / "out.xlsx"
    wb.save(out)
    wb.close()

    sheet_xml = _read_sheet_xml(out)
    block = _cf_block_for_sqref(sheet_xml, "A1:A10")
    assert block is not None, f"missing CF block in: {sheet_xml[:500]}"
    assert 'type="cellIs"' in block
    assert 'priority="1"' in block
    assert 'operator="equal"' in block
    assert 'dxfId="0"' in block

    styles_xml = _read_styles_xml(out)
    assert _count_dxfs_in_styles(styles_xml) == 1, (
        f"expected exactly 1 <dxf>, got: {styles_xml}"
    )


def test_add_cf_preserves_existing(tmp_path: Path) -> None:
    """HEADLINE BYTE-PRESERVATION GATE (RFC §8 risk #1).

    Merger Q4 drops every existing ``<conditionalFormatting>`` block
    when any CF block is supplied. The patcher must therefore re-emit
    them verbatim from byte slices captured by
    ``extract_existing_cf_blocks``. If that capture has a byte-range
    bug, this assertion fails.
    """
    src = tmp_path / "src.xlsx"
    _make_fixture_with_one_cf(src)

    src_xml = _read_sheet_xml(src)
    existing_blocks = _all_cf_blocks(src_xml)
    assert len(existing_blocks) == 1, (
        f"fixture should have exactly 1 CF block, got {len(existing_blocks)}"
    )
    existing_block = existing_blocks[0]
    src_dxf_count = _count_dxfs_in_styles(_read_styles_xml(src))
    assert src_dxf_count >= 1, "fixture's CF rule should have allocated a dxf"

    wb = Workbook._from_patcher(str(src))
    ws = wb["Sheet1"]
    ws.conditional_formatting.add(
        "B2:B10",
        CellIsRule(
            operator="lessThan",
            formula=["50"],
            extra={"font_italic": True},
        ),
    )
    out = tmp_path / "out.xlsx"
    wb.save(out)
    wb.close()

    out_xml = _read_sheet_xml(out)
    out_blocks = _all_cf_blocks(out_xml)
    assert len(out_blocks) == 2, (
        f"expected 2 CF blocks (existing + new), got {len(out_blocks)}: {out_blocks}"
    )

    # The existing block's bytes must appear verbatim in the saved file.
    assert existing_block in out_xml, (
        "existing CF block was not preserved verbatim. "
        "If extract_existing_cf_blocks captured the wrong byte range, the "
        "block may have been parsed-and-re-emitted (which loses unknown "
        "attributes / escape forms / self-closing variants).\n"
        f"  source:\n{existing_block}\n"
        f"  output (full sheet):\n{out_xml}"
    )

    new_block = _cf_block_for_sqref(out_xml, "B2:B10")
    assert new_block is not None, f"new block missing: {out_xml}"
    # Priority must continue past the existing rule's priority (which is 1).
    assert 'priority="2"' in new_block, f"expected priority=2: {new_block}"
    # dxfId must continue past the existing rule's dxfId (which is 0).
    assert f'dxfId="{src_dxf_count}"' in new_block, (
        f"expected dxfId={src_dxf_count}, got: {new_block}"
    )

    out_dxf_count = _count_dxfs_in_styles(_read_styles_xml(out))
    assert out_dxf_count == src_dxf_count + 1, (
        f"expected {src_dxf_count + 1} dxfs after add, got {out_dxf_count}"
    )


def test_dxf_id_monotonic_across_rules(tmp_path: Path) -> None:
    """Two CellIs rules in one ``add()`` cycle to the same sheet → adjacent dxfIds."""
    src = tmp_path / "src.xlsx"
    _make_clean_fixture(src)

    wb = Workbook._from_patcher(str(src))
    ws = wb["Sheet1"]
    ws.conditional_formatting.add(
        "A1:A10",
        CellIsRule(operator="equal", formula=["1"], extra={"font_bold": True}),
    )
    ws.conditional_formatting.add(
        "A1:A10",
        CellIsRule(operator="equal", formula=["2"], extra={"font_italic": True}),
    )
    out = tmp_path / "out.xlsx"
    wb.save(out)
    wb.close()

    sheet_xml = _read_sheet_xml(out)
    block = _cf_block_for_sqref(sheet_xml, "A1:A10")
    assert block is not None
    assert 'dxfId="0"' in block
    assert 'dxfId="1"' in block
    assert _count_dxfs_in_styles(_read_styles_xml(out)) == 2


def test_priority_monotonic_across_sqrefs(tmp_path: Path) -> None:
    """Two ``add()`` calls with different sqrefs → second's priority > first's."""
    src = tmp_path / "src.xlsx"
    _make_clean_fixture(src)

    wb = Workbook._from_patcher(str(src))
    ws = wb["Sheet1"]
    ws.conditional_formatting.add(
        "A1:A10", CellIsRule(operator="equal", formula=["1"], extra={"font_bold": True})
    )
    ws.conditional_formatting.add(
        "B1:B10", CellIsRule(operator="equal", formula=["2"], extra={"font_italic": True})
    )
    out = tmp_path / "out.xlsx"
    wb.save(out)
    wb.close()

    sheet_xml = _read_sheet_xml(out)
    a_block = _cf_block_for_sqref(sheet_xml, "A1:A10")
    b_block = _cf_block_for_sqref(sheet_xml, "B1:B10")
    assert a_block is not None and b_block is not None
    a_pri = int(re.search(r'priority="(\d+)"', a_block).group(1))  # type: ignore[union-attr]
    b_pri = int(re.search(r'priority="(\d+)"', b_block).group(1))  # type: ignore[union-attr]
    assert b_pri > a_pri, f"expected B's priority > A's: A={a_pri}, B={b_pri}"


def test_dxf_id_monotonic_across_sheets(tmp_path: Path) -> None:
    """CROSS-SHEET ALLOCATOR GATE.

    CF on Sheet1 + CF on Sheet2 in one ``save()`` → second sheet's dxfIds
    continue from where Sheet1 left off. The Rust patcher's Phase-2.5b
    threads a single ``running_dxf_count`` through both sheets in
    deterministic sorted-name order.
    """
    src = tmp_path / "src.xlsx"
    _make_clean_fixture(src, sheet_titles=("Sheet1", "Sheet2"))

    wb = Workbook._from_patcher(str(src))
    wb["Sheet1"].conditional_formatting.add(
        "A1:A10", CellIsRule(operator="equal", formula=["1"], extra={"font_bold": True})
    )
    wb["Sheet2"].conditional_formatting.add(
        "A1:A10", CellIsRule(operator="equal", formula=["2"], extra={"font_italic": True})
    )
    out = tmp_path / "out.xlsx"
    wb.save(out)
    wb.close()

    s1 = _read_sheet_xml(out, "xl/worksheets/sheet1.xml")
    s2 = _read_sheet_xml(out, "xl/worksheets/sheet2.xml")
    s1_block = _cf_block_for_sqref(s1, "A1:A10")
    s2_block = _cf_block_for_sqref(s2, "A1:A10")
    assert s1_block is not None and s2_block is not None

    s1_dxf = int(re.search(r'dxfId="(\d+)"', s1_block).group(1))  # type: ignore[union-attr]
    s2_dxf = int(re.search(r'dxfId="(\d+)"', s2_block).group(1))  # type: ignore[union-attr]
    # Both 0 and 1 must be allocated, distinct, with sheet1 going first
    # (deterministic sort by sheet name).
    assert {s1_dxf, s2_dxf} == {0, 1}, f"got s1={s1_dxf}, s2={s2_dxf}"
    assert s1_dxf == 0, f"deterministic order: Sheet1 should win dxfId=0, got {s1_dxf}"

    assert _count_dxfs_in_styles(_read_styles_xml(out)) == 2


def test_colorscale_no_dxf(tmp_path: Path) -> None:
    """ColorScale rules carry no ``dxfId`` and don't grow ``<dxfs>``."""
    src = tmp_path / "src.xlsx"
    _make_clean_fixture(src)

    wb = Workbook._from_patcher(str(src))
    ws = wb["Sheet1"]
    ws.conditional_formatting.add(
        "A1:A10",
        ColorScaleRule(
            start_type="min",
            start_color="FF63BE7B",
            end_type="max",
            end_color="FFF8696B",
        ),
    )
    out = tmp_path / "out.xlsx"
    wb.save(out)
    wb.close()

    sheet_xml = _read_sheet_xml(out)
    block = _cf_block_for_sqref(sheet_xml, "A1:A10")
    assert block is not None
    assert 'type="colorScale"' in block
    assert "dxfId" not in block, f"colorScale must not carry dxfId: {block}"

    # styles.xml must not have grown a <dxfs> from this slice.
    src_dxf = _count_dxfs_in_styles(_read_styles_xml(src))
    out_dxf = _count_dxfs_in_styles(_read_styles_xml(out))
    assert out_dxf == src_dxf, f"colorScale must not allocate dxf: {src_dxf} → {out_dxf}"


def test_databar_no_dxf(tmp_path: Path) -> None:
    """DataBar rules carry no ``dxfId`` and don't grow ``<dxfs>``."""
    src = tmp_path / "src.xlsx"
    _make_clean_fixture(src)

    wb = Workbook._from_patcher(str(src))
    ws = wb["Sheet1"]
    ws.conditional_formatting.add(
        "A1:A10",
        DataBarRule(start_type="min", end_type="max", color="FF638EC6"),
    )
    out = tmp_path / "out.xlsx"
    wb.save(out)
    wb.close()

    sheet_xml = _read_sheet_xml(out)
    block = _cf_block_for_sqref(sheet_xml, "A1:A10")
    assert block is not None
    assert 'type="dataBar"' in block
    assert "dxfId" not in block, f"dataBar must not carry dxfId: {block}"

    src_dxf = _count_dxfs_in_styles(_read_styles_xml(src))
    out_dxf = _count_dxfs_in_styles(_read_styles_xml(out))
    assert out_dxf == src_dxf, f"dataBar must not allocate dxf: {src_dxf} → {out_dxf}"


def test_expression_rule_escapes_formula(tmp_path: Path) -> None:
    """formula="A1>B1" must serialize as ``<formula>A1&gt;B1</formula>``."""
    src = tmp_path / "src.xlsx"
    _make_clean_fixture(src)

    wb = Workbook._from_patcher(str(src))
    ws = wb["Sheet1"]
    ws.conditional_formatting.add(
        "C1:C10",
        FormulaRule(formula=["A1>B1"], extra={"font_bold": True}),
    )
    out = tmp_path / "out.xlsx"
    wb.save(out)
    wb.close()

    sheet_xml = _read_sheet_xml(out)
    block = _cf_block_for_sqref(sheet_xml, "C1:C10")
    assert block is not None
    assert 'type="expression"' in block
    assert "<formula>A1&gt;B1</formula>" in block, f"'>' not escaped: {block}"
    # Raw '>' inside text content would break OOXML.
    assert "<formula>A1>B1</formula>" not in block


def test_cellis_between_two_formulas(tmp_path: Path) -> None:
    """``operator='between'`` emits both ``<formula>`` children."""
    src = tmp_path / "src.xlsx"
    _make_clean_fixture(src)

    wb = Workbook._from_patcher(str(src))
    ws = wb["Sheet1"]
    ws.conditional_formatting.add(
        "A1:A10",
        CellIsRule(
            operator="between", formula=["10", "20"], extra={"font_bold": True}
        ),
    )
    out = tmp_path / "out.xlsx"
    wb.save(out)
    wb.close()

    sheet_xml = _read_sheet_xml(out)
    block = _cf_block_for_sqref(sheet_xml, "A1:A10")
    assert block is not None
    assert 'operator="between"' in block
    assert "<formula>10</formula>" in block
    assert "<formula>20</formula>" in block


def test_create_dxfs_section_when_absent(tmp_path: Path) -> None:
    """A clean fixture has no ``<dxfs>``; first CF rule with a dxf creates it."""
    src = tmp_path / "src.xlsx"
    _make_clean_fixture(src)

    src_styles = _read_styles_xml(src)
    # Either no <dxfs> at all, or an empty one — confirm the count is 0.
    assert _count_dxfs_in_styles(src_styles) == 0, (
        f"clean fixture should have no <dxf> children: {src_styles}"
    )

    wb = Workbook._from_patcher(str(src))
    ws = wb["Sheet1"]
    ws.conditional_formatting.add(
        "A1:A10",
        CellIsRule(operator="equal", formula=["1"], extra={"font_bold": True}),
    )
    out = tmp_path / "out.xlsx"
    wb.save(out)
    wb.close()

    out_styles = _read_styles_xml(out)
    assert "<dxfs" in out_styles, (
        f"<dxfs> section was not created: {out_styles[-500:]}"
    )
    assert _count_dxfs_in_styles(out_styles) == 1


def test_cf_round_trip_via_load_workbook(tmp_path: Path) -> None:
    """openpyxl reads back the rules we wrote, matching kind+sqref."""
    src = tmp_path / "src.xlsx"
    _make_clean_fixture(src)

    wb = Workbook._from_patcher(str(src))
    ws = wb["Sheet1"]
    ws.conditional_formatting.add(
        "A2:A10",
        CellIsRule(operator="greaterThan", formula=["50"], extra={"font_bold": True}),
    )
    out = tmp_path / "out.xlsx"
    wb.save(out)
    wb.close()

    wb2 = openpyxl.load_workbook(out)
    ws2 = wb2["Sheet1"]
    entries = list(ws2.conditional_formatting)
    found = False
    for entry in entries:
        # openpyxl exposes range as e.sqref.ranges or as e.cells; the
        # public bridge is iterating + str(sqref).
        sqref_text = str(getattr(entry, "sqref", ""))
        if "A2:A10" in sqref_text:
            for r in entry.rules:
                if getattr(r, "type", None) == "cellIs":
                    found = True
                    break
    assert found, f"CellIs rule on A2:A10 missing on re-open: {entries}"


def test_cf_no_pending_no_op(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    """Open + save with no CF append must be byte-identical (short-circuit guard).

    If ``do_save``'s short-circuit predicate did not include
    ``queued_cf_patches.is_empty()``, even a no-op save would re-zip.
    """
    monkeypatch.setenv("WOLFXL_TEST_EPOCH", "0")
    src = tmp_path / "src.xlsx"
    _make_clean_fixture(src)

    wb = Workbook._from_patcher(str(src))
    out = tmp_path / "out.xlsx"
    wb.save(out)
    wb.close()

    assert src.read_bytes() == out.read_bytes(), (
        "no-op save (no CF queued) must be byte-identical — "
        "the do_save short-circuit predicate must check queued_cf_patches"
    )


def test_unsupported_rule_kind_raises(tmp_path: Path) -> None:
    """IconSet (and other stubbed kinds) raise NotImplementedError pointing at §10."""
    src = tmp_path / "src.xlsx"
    _make_clean_fixture(src)

    wb = Workbook._from_patcher(str(src))
    ws = wb["Sheet1"]
    with pytest.raises(NotImplementedError, match="026-conditional-formatting"):
        ws.conditional_formatting.add("A1:A10", IconSetRule(icon_style="3Arrows"))
    wb.close()


# ---------------------------------------------------------------------------
# Pin epoch so any test that re-runs gets stable bytes.
# ---------------------------------------------------------------------------


@pytest.fixture(autouse=True)
def _force_test_epoch(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("WOLFXL_TEST_EPOCH", "0")
