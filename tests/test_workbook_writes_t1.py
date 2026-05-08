"""T1 PR6 — Workbook writes: DocumentProperties + DefinedName.

Write mode accepts ``wb.properties.title = ...`` and ``wb.defined_names[k] = DefinedName(...)``.
Modify mode now ships properties round-trip (RFC-020); ``defined_names`` is
still T1.5 (RFC-021) so its raise-contract test stays.
"""

from __future__ import annotations

from pathlib import Path
from zipfile import ZipFile

import pytest
from wolfxl.packaging.core import DocumentProperties
from wolfxl.workbook.defined_name import DefinedName

from wolfxl import Workbook, load_workbook

openpyxl = pytest.importorskip("openpyxl")


def test_properties_write_round_trip_via_openpyxl(tmp_path: Path) -> None:
    """wolfxl write → openpyxl read on the fields rust_xlsxwriter supports."""
    path = tmp_path / "wolfxl_props.xlsx"
    wb = Workbook()
    wb.properties.title = "Q3 Report"
    wb.properties.creator = "Alice"
    wb.properties.subject = "Revenue"
    wb.properties.keywords = "revenue, q3"
    wb.properties.description = "Quarterly revenue snapshot"
    wb.properties.category = "Finance"
    wb.save(str(path))

    op_wb = openpyxl.load_workbook(path)
    assert op_wb.properties.title == "Q3 Report"
    assert op_wb.properties.creator == "Alice"
    assert op_wb.properties.subject == "Revenue"
    assert op_wb.properties.keywords == "revenue, q3"
    assert op_wb.properties.description == "Quarterly revenue snapshot"
    assert op_wb.properties.category == "Finance"


def test_properties_setter_marks_dirty() -> None:
    wb = Workbook()
    wb.properties = DocumentProperties(title="replaced", creator="bob")
    assert wb._properties_dirty is True
    assert wb.properties.title == "replaced"
    assert wb.properties.creator == "bob"


def test_properties_write_round_trip_via_wolfxl(tmp_path: Path) -> None:
    """wolfxl write → wolfxl read (covers our own reader parity)."""
    path = tmp_path / "wolfxl_props_self.xlsx"
    wb = Workbook()
    wb.properties.title = "My Report"
    wb.properties.creator = "Me"
    wb.save(str(path))

    rb = Workbook._from_reader(str(path))
    assert rb.properties.title == "My Report"
    assert rb.properties.creator == "Me"


def test_defined_name_write_round_trip(tmp_path: Path) -> None:
    path = tmp_path / "wolfxl_names.xlsx"
    wb = Workbook()
    ws = wb.active
    for i in range(1, 6):
        ws[f"A{i}"] = i
    wb.defined_names["Region"] = DefinedName(
        name="Region",
        value="Sheet!$A$1:$A$5",
    )
    # In-session access works immediately.
    assert wb.defined_names["Region"].value == "Sheet!$A$1:$A$5"
    wb.save(str(path))

    op_wb = openpyxl.load_workbook(path)
    assert "Region" in op_wb.defined_names
    refers = op_wb.defined_names["Region"].value
    # rust_xlsxwriter may or may not add a leading "=" to refers_to — accept both.
    assert refers.lstrip("=").replace("'", "") in {
        "Sheet!$A$1:$A$5",
        "Sheet1!$A$1:$A$5",
    } or "A1:A5" in refers.replace("'", "")


def test_defined_name_add_helper_round_trip(tmp_path: Path) -> None:
    path = tmp_path / "wolfxl_names_add.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["B1"] = 99
    wb.defined_names.add(DefinedName(name="Top", value="Sheet!$B$1"))
    wb.save(str(path))

    op_wb = openpyxl.load_workbook(path)
    assert "Top" in op_wb.defined_names


def test_defined_names_empty_workbook_saves(tmp_path: Path) -> None:
    """A workbook with no defined names should save without trouble."""
    path = tmp_path / "no_names.xlsx"
    wb = Workbook()
    wb.active["A1"] = 1
    wb.save(str(path))
    # Reopen — defined_names is empty dict-like.
    rb = Workbook._from_reader(str(path))
    assert len(rb.defined_names) == 0


def test_properties_modify_mode_round_trips(tmp_path: Path) -> None:
    """RFC-020: the patcher rewrites docProps/core.xml in modify mode.

    Replaces the prior T1.5-raise contract. The full positive-coverage
    surface lives in tests/test_modify_properties.py; this one just
    pins the contract that save() does NOT raise.
    """
    src = tmp_path / "exists_props.xlsx"
    out = tmp_path / "out_props.xlsx"
    openpyxl.Workbook().save(src)
    wb = Workbook._from_patcher(str(src))
    wb.properties.title = "new"
    wb.save(str(out))
    wb.close()
    # Re-open via openpyxl to verify the title round-trips.
    wb_check = openpyxl.load_workbook(str(out))
    assert wb_check.properties.title == "new"


def test_defined_names_modify_mode_round_trip(tmp_path: Path) -> None:
    """RFC-021 — modify mode now supports defined-name mutation. The
    former NotImplementedError("T1.5") guard at this site is replaced
    by a real round-trip via the patcher's Phase 2.5f. See
    ``tests/test_defined_names_modify.py`` for the broader matrix."""
    src = tmp_path / "exists_names.xlsx"
    dst = tmp_path / "out.xlsx"
    openpyxl.Workbook().save(src)
    wb = Workbook._from_patcher(str(src))
    wb.defined_names["New"] = DefinedName(name="New", value="Sheet!$A$1")
    wb.save(str(dst))
    wb.close()

    rt = openpyxl.load_workbook(dst)
    assert "New" in rt.defined_names
    assert rt.defined_names["New"].attr_text == "Sheet!$A$1"


def test_sheet_rename_retargets_sheet_scoped_defined_name(tmp_path: Path) -> None:
    """Modify-mode sheet rename keeps hidden external-data names coherent."""
    from openpyxl.workbook.defined_name import DefinedName as XDefinedName

    src = tmp_path / "external_data_name.xlsx"
    dst = tmp_path / "renamed.xlsx"

    op = openpyxl.Workbook()
    ws = op.active
    ws.title = "Sales Order_data"
    for row in range(1, 5):
        for col in range(1, 5):
            ws.cell(row=row, column=col, value=row * col)
    defined_name = XDefinedName(
        "ExternalData_1",
        attr_text="'Sales Order_data'!$A$1:$D$4",
        localSheetId=0,
        hidden=True,
    )
    if hasattr(op.defined_names, "add"):
        op.defined_names.add(defined_name)
    else:
        op.defined_names["ExternalData_1"] = defined_name
    op.save(src)

    wb = load_workbook(src, modify=True)
    wb["Sales Order_data"].title = "WolfXL Fidelity Rename"
    wb.save(dst)
    wb.close()

    with ZipFile(dst) as zf:
        workbook_xml = zf.read("xl/workbook.xml").decode("utf-8")
    assert "'WolfXL Fidelity Rename'!$A$1:$D$4" in workbook_xml
    assert "'Sales Order_data'!$A$1:$D$4" not in workbook_xml
    assert 'name="ExternalData_1"' in workbook_xml
    assert 'localSheetId="0"' in workbook_xml
    assert 'hidden="1"' in workbook_xml
