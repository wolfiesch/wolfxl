"""RFC-020 — Document Properties round-trip in modify mode.

End-to-end coverage for ``wb.properties.<field> = ...`` on an existing
file. The save-time path traverses three layers:

1. ``Workbook._flush_properties_to_patcher`` (Python) builds a flat
   snake_case dict and calls ``XlsxPatcher.queue_properties``.
2. ``XlsxPatcher::do_save`` Phase 2.5d (Rust) routes the payload through
   ``properties::rewrite_core_props`` / ``rewrite_app_props`` and
   writes via ``file_patches`` if the source already has the entry,
   else via RFC-013's ``file_adds`` (the optional-``docProps/core.xml``
   case from RFC-020 §8 risk #3).
3. ``properties::rewrite_*`` (Rust) emits the OOXML strict-element
   order, escapes user text via ``xml_text_escape``, defaults
   ``creator``/``lastModifiedBy`` to ``"wolfxl"`` per OOXML convention,
   and honors ``WOLFXL_TEST_EPOCH=0`` for deterministic timestamps.

Sister contracts:

* ``test_no_dirty_save_is_byte_identical`` is the regression guard for
  the short-circuit predicate on ``do_save``. If a future refactor
  forgets to require ``queued_props.is_none()``, this test fires.
* ``test_app_xml_drops_company_manager_known_loss`` is a *negative*
  guard: RFC-020 §7 documents that ``<Company>`` and ``<Manager>`` from
  the source ``app.xml`` are dropped on dirty save (Python's
  ``DocumentProperties`` doesn't expose them). The test pins the loss
  so a future patch that accidentally fixes the round-trip surfaces
  here, not in user reports.
"""
from __future__ import annotations

import re
import zipfile
from datetime import datetime
from pathlib import Path

import openpyxl
import pytest

from wolfxl import Workbook

# ---------------------------------------------------------------------------
# Fixtures and helpers
# ---------------------------------------------------------------------------


@pytest.fixture(autouse=True)
def _force_test_epoch(monkeypatch: pytest.MonkeyPatch) -> None:
    """Pin `dcterms:modified` to the Unix epoch for byte-stable saves."""
    monkeypatch.setenv("WOLFXL_TEST_EPOCH", "0")


def _make_clean_fixture(path: Path, sheet_titles: tuple[str, ...] = ("Sheet1",)) -> None:
    """Workbook with the openpyxl default ``docProps/{core,app}.xml`` shape."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_titles[0]
    ws["A1"] = "header"
    for title in sheet_titles[1:]:
        wb.create_sheet(title)
    wb.save(path)


def _make_fixture_with_subject(path: Path, subject: str = "QuarterlyData") -> None:
    """Fixture pre-populated with ``subject`` so we can prove other fields
    survive when the user only mutates ``creator``."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "header"
    wb.properties.subject = subject
    wb.save(path)


def _make_fixture_with_company_manager(path: Path) -> None:
    """openpyxl writes Company + Manager in app.xml when set on the source.
    We need this to assert RFC-020 §7's known regression."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    # openpyxl's DocumentProperties doesn't expose company/manager,
    # but its packaging emits them from the writer's defaults if set
    # on the workbook's properties shim. Easier to inject directly.
    wb.save(path)
    # Inject Company/Manager into app.xml by rewriting the entry.
    import shutil
    tmp = path.with_suffix(".tmp.xlsx")
    with zipfile.ZipFile(path) as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == "docProps/app.xml":
                text = data.decode("utf-8")
                # Insert <Company> and <Manager> just before </Properties>.
                injection = "<Company>Acme Corp</Company><Manager>Alice</Manager>"
                text = text.replace("</Properties>", f"{injection}</Properties>")
                data = text.encode("utf-8")
            zout.writestr(item, data)
    shutil.move(str(tmp), str(path))


def _read_zip_text(path: Path, entry: str) -> str:
    with zipfile.ZipFile(path) as zf:
        return zf.read(entry).decode("utf-8")


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


def test_set_title_round_trip(tmp_path: Path) -> None:
    """The smoke test: set one field, save, openpyxl reads it back."""
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    _make_clean_fixture(src)

    wb = Workbook._from_patcher(str(src))
    wb.properties.title = "Quarterly Report"
    wb.save(out)
    wb.close()

    re_wb = openpyxl.load_workbook(out)
    assert re_wb.properties.title == "Quarterly Report"


def test_set_creator_preserves_existing_other_fields(tmp_path: Path) -> None:
    """Mutating one field must NOT zero out the other fields the source
    already had. Modify mode reads the cache from the source's
    ``docProps/core.xml`` at first ``wb.properties`` access."""
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    _make_fixture_with_subject(src, subject="QuarterlyData")

    wb = Workbook._from_patcher(str(src))
    # Touching .creator triggers __setattr__ which marks dirty; reading
    # .subject first ensures the cache is hydrated from the source.
    assert wb.properties.subject == "QuarterlyData"
    wb.properties.creator = "Alice"
    wb.save(out)
    wb.close()

    re_wb = openpyxl.load_workbook(out)
    assert re_wb.properties.subject == "QuarterlyData", "subject lost on dirty save"
    assert re_wb.properties.creator == "Alice"


def test_no_dirty_save_is_byte_identical(tmp_path: Path) -> None:
    """Open + save without touching properties → byte-identical to source.

    This is the short-circuit-predicate regression guard. If
    ``do_save`` accidentally rewrites ``docProps/*`` for an untouched
    workbook, this test fires.
    """
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    _make_clean_fixture(src)
    src_bytes = src.read_bytes()

    wb = Workbook._from_patcher(str(src))
    wb.save(out)
    wb.close()

    assert out.read_bytes() == src_bytes, (
        "no-op modify-mode save changed bytes — short-circuit predicate "
        "likely missing queued_props.is_none() or properties got eagerly "
        "marked dirty during cache hydration."
    )


def test_xml_special_chars_escaped(tmp_path: Path) -> None:
    """``"A & B < C"`` must serialize as ``A &amp; B &lt; C``. The Rust
    serializer uses ``xml_text_escape`` on every text-content field per
    RFC-020 §6 risk #1."""
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    _make_clean_fixture(src)

    wb = Workbook._from_patcher(str(src))
    wb.properties.title = "A & B < C"
    wb.save(out)
    wb.close()

    core_xml = _read_zip_text(out, "docProps/core.xml")
    assert "A &amp; B &lt; C" in core_xml, core_xml
    # The unescaped form must NOT leak through.
    assert "A & B < C" not in core_xml


def test_modified_timestamp_uses_test_epoch(tmp_path: Path) -> None:
    """``WOLFXL_TEST_EPOCH=0`` → ``dcterms:modified`` is the Unix epoch.

    The autouse fixture sets the env var; the Rust ``current_timestamp_iso8601``
    short-circuits to ``1970-01-01T00:00:00Z`` when it sees ``=0``.
    Without this hook, two saves produce different bytes and golden-file
    tests can't exist.
    """
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    _make_clean_fixture(src)

    wb = Workbook._from_patcher(str(src))
    wb.properties.title = "anything"
    wb.save(out)
    wb.close()

    core_xml = _read_zip_text(out, "docProps/core.xml")
    assert "<dcterms:modified" in core_xml
    assert "1970-01-01T00:00:00Z" in core_xml, core_xml


def test_user_explicit_modified_wins(tmp_path: Path) -> None:
    """If the user sets ``props.modified`` to a specific ``datetime``,
    that value reaches the XML — not ``current_timestamp_iso8601``.

    Confirms the flush path threads ``modified_iso`` through and the
    Rust serializer prefers ``payload.modified_iso`` over its own clock.
    """
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    _make_clean_fixture(src)

    wb = Workbook._from_patcher(str(src))
    wb.properties.title = "anything"  # also dirties; required for flush
    wb.properties.modified = datetime(2020, 6, 15, 12, 0, 0)
    wb.save(out)
    wb.close()

    core_xml = _read_zip_text(out, "docProps/core.xml")
    assert "2020-06-15T12:00:00" in core_xml, core_xml
    assert "1970-01-01T00:00:00Z" not in core_xml, "test-epoch leaked despite explicit modified"


def test_app_xml_titles_of_parts_in_sheet_order(tmp_path: Path) -> None:
    """``<TitlesOfParts>`` must list sheets in the workbook's source
    order. The patcher's ``sheet_order`` field (RFC-013 commit 1)
    captures this from ``xl/workbook.xml`` at open time; the Python
    flush passes ``list(self._sheet_names)`` through, which mirrors the
    same ordering.
    """
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    _make_clean_fixture(src, sheet_titles=("Alpha", "Bravo", "Charlie"))

    wb = Workbook._from_patcher(str(src))
    wb.properties.title = "ordered"
    wb.save(out)
    wb.close()

    app_xml = _read_zip_text(out, "docProps/app.xml")
    # Capture every <vt:lpstr> in <TitlesOfParts>.
    m = re.search(r"<TitlesOfParts>.*?</TitlesOfParts>", app_xml, re.DOTALL)
    assert m is not None, app_xml
    titles = re.findall(r"<vt:lpstr>([^<]*)</vt:lpstr>", m.group(0))
    assert titles == ["Alpha", "Bravo", "Charlie"], titles


def test_app_xml_drops_company_manager_known_loss(tmp_path: Path) -> None:
    """RFC-020 §7 documented regression: a dirty save rewrites app.xml
    in full from the Python ``DocumentProperties`` payload, which has
    no ``company`` / ``manager`` fields. So a source ``<Company>`` and
    ``<Manager>`` are silently dropped.

    This test pins the loss. If a future patch starts preserving these
    fields, this test will fail and the regression note in
    Plans/rfcs/020-document-properties.md should be cleared.
    """
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    _make_fixture_with_company_manager(src)

    # Sanity: source has Company + Manager.
    src_app = _read_zip_text(src, "docProps/app.xml")
    assert "<Company>Acme Corp</Company>" in src_app
    assert "<Manager>Alice</Manager>" in src_app

    wb = Workbook._from_patcher(str(src))
    wb.properties.title = "force_rewrite"
    wb.save(out)
    wb.close()

    out_app = _read_zip_text(out, "docProps/app.xml")
    assert "<Company>" not in out_app, (
        "Company unexpectedly preserved — RFC-020 §7 regression closed?"
    )
    assert "<Manager>" not in out_app, (
        "Manager unexpectedly preserved — RFC-020 §7 regression closed?"
    )


def test_creator_falls_back_to_wolfxl_when_unset(tmp_path: Path) -> None:
    """Source has no ``creator`` (openpyxl puts an empty/default), the
    user mutates only ``title``, the Rust serializer fills in
    ``"wolfxl"`` per OOXML convention via ``DEFAULT_CREATOR``."""
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    _make_clean_fixture(src)

    wb = Workbook._from_patcher(str(src))
    # Explicitly clear creator so the flush filters it out and Rust
    # has to apply the default.
    wb.properties.creator = None
    wb.properties.title = "needs default creator"
    wb.save(out)
    wb.close()

    core_xml = _read_zip_text(out, "docProps/core.xml")
    assert "<dc:creator>wolfxl</dc:creator>" in core_xml, core_xml


def test_round_trip_via_load_workbook_modify_chain(tmp_path: Path) -> None:
    """Modify → save → re-open in modify mode → mutate again → save →
    final read sees both writes. Catches a class of bugs where the
    second open's ``_properties_cache`` hydration shadows what the
    first save wrote."""
    src = tmp_path / "src.xlsx"
    mid = tmp_path / "mid.xlsx"
    final = tmp_path / "final.xlsx"
    _make_clean_fixture(src)

    wb1 = Workbook._from_patcher(str(src))
    wb1.properties.title = "first"
    wb1.properties.creator = "Alice"
    wb1.save(mid)
    wb1.close()

    wb2 = Workbook._from_patcher(str(mid))
    # Verify the first round-trip is visible on re-open.
    assert wb2.properties.title == "first"
    assert wb2.properties.creator == "Alice"
    # Mutate only the subject; title + creator should persist.
    wb2.properties.subject = "Q3"
    wb2.save(final)
    wb2.close()

    re_wb = openpyxl.load_workbook(final)
    assert re_wb.properties.title == "first"
    assert re_wb.properties.creator == "Alice"
    assert re_wb.properties.subject == "Q3"
