"""Modify-mode data validations (RFC-025) — end-to-end coverage.

These tests build a fixture xlsx with openpyxl, open it with wolfxl in
modify mode, append data validations, save, then assert on the saved
sheet XML. The byte-level assertions read the sheet XML directly out of
the ZIP — going back through openpyxl on read would re-serialize and
mask any preservation regression we care about.

The headline gate is ``test_add_dv_preserves_existing`` (RFC §8 risk #1):
if the byte-slice capture in ``extract_existing_dv_block`` had a bug, the
existing DV's bytes would shift and the contains-substring check would
fail.
"""
from __future__ import annotations

import os
import re
import zipfile
from pathlib import Path

import openpyxl
import pytest
from wolfxl.worksheet.datavalidation import DataValidation

from wolfxl import Workbook

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_clean_fixture(path: Path) -> None:
    """Empty workbook — no DVs."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "header"
    wb.save(path)


def _make_fixture_with_one_dv(path: Path) -> None:
    """Workbook with exactly one existing list DV (openpyxl-built)."""
    from openpyxl.worksheet.datavalidation import DataValidation as OpyDV

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "header"
    dv = OpyDV(
        type="list",
        formula1='"Existing,Values"',
        allow_blank=True,
        showErrorMessage=True,
    )
    dv.add("A2:A100")
    ws.add_data_validation(dv)
    wb.save(path)


def _read_sheet_xml(path: Path, sheet_path: str = "xl/worksheets/sheet1.xml") -> str:
    with zipfile.ZipFile(path) as zf:
        return zf.read(sheet_path).decode("utf-8")


def _read_dv_block(path: Path) -> str | None:
    """Return the entire ``<dataValidations …>…</dataValidations>`` element
    or None if missing."""
    xml = _read_sheet_xml(path)
    m = re.search(r"<dataValidations[^>]*?(?:/>|>.*?</dataValidations>)", xml, re.DOTALL)
    return m.group(0) if m else None


# ---------------------------------------------------------------------------
# Tests — RFC §6 coverage
# ---------------------------------------------------------------------------

def test_add_dv_to_clean_file(tmp_path: Path) -> None:
    """File with no existing DVs. After save, sheet has count=1."""
    src = tmp_path / "src.xlsx"
    _make_clean_fixture(src)

    wb = Workbook._from_patcher(str(src))
    ws = wb["Sheet1"]
    ws.data_validations.append(
        DataValidation(type="list", formula1='"A,B,C"', sqref="B2:B10", allowBlank=True)
    )
    out = tmp_path / "out.xlsx"
    wb.save(out)
    wb.close()

    block = _read_dv_block(out)
    assert block is not None, "expected a <dataValidations> block in saved file"
    assert 'count="1"' in block, f"expected count=1, got: {block}"
    assert 'type="list"' in block
    assert 'sqref="B2:B10"' in block
    assert "<formula1>\"A,B,C\"</formula1>" in block


def test_add_dv_preserves_existing(tmp_path: Path) -> None:
    """Headline RFC §8 risk #1 gate: existing DV bytes survive verbatim."""
    src = tmp_path / "src.xlsx"
    _make_fixture_with_one_dv(src)

    # Capture the existing DV's element text BEFORE our save, so we can
    # confirm it survives byte-for-byte.
    src_block = _read_dv_block(src)
    assert src_block is not None, "fixture missing existing DV block"
    # Match `<dataValidation ` (with the trailing space, so we don't also
    # match the wrapper `<dataValidations`).
    m = re.search(r"<dataValidation\s[^>]*?>.*?</dataValidation>", src_block, re.DOTALL)
    assert m is not None, "fixture's DV missing children?"
    existing_dv_xml = m.group(0)
    assert '"Existing,Values"' in existing_dv_xml

    wb = Workbook._from_patcher(str(src))
    ws = wb["Sheet1"]
    ws.data_validations.append(
        DataValidation(
            type="whole",
            operator="between",
            formula1="1",
            formula2="100",
            sqref="C2:C100",
            showErrorMessage=True,
        )
    )
    out = tmp_path / "out.xlsx"
    wb.save(out)
    wb.close()

    out_block = _read_dv_block(out)
    assert out_block is not None
    assert 'count="2"' in out_block, f"expected count=2 after append, got: {out_block}"
    # The existing DV element must appear byte-for-byte in the new block.
    assert existing_dv_xml in out_block, (
        f"existing DV bytes were not preserved verbatim.\n"
        f"  source: {existing_dv_xml}\n"
        f"  output: {out_block}"
    )
    # The new DV is also there.
    assert 'type="whole"' in out_block
    assert 'operator="between"' in out_block
    assert "<formula2>100</formula2>" in out_block


def test_dv_list_inline_values(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    _make_clean_fixture(src)
    wb = Workbook._from_patcher(str(src))
    wb["Sheet1"].data_validations.append(
        DataValidation(type="list", formula1='"Apple,Banana,Cherry"', sqref="A2:A10")
    )
    out = tmp_path / "out.xlsx"
    wb.save(out)
    block = _read_dv_block(out)
    assert block is not None
    # The double quotes are part of the OOXML inline-list grammar — they MUST
    # survive into the output text.
    assert "<formula1>\"Apple,Banana,Cherry\"</formula1>" in block


def test_dv_list_range_ref(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    _make_clean_fixture(src)
    wb = Workbook._from_patcher(str(src))
    wb["Sheet1"].data_validations.append(
        DataValidation(type="list", formula1="Sheet2!$A$1:$A$5", sqref="A2:A10")
    )
    out = tmp_path / "out.xlsx"
    wb.save(out)
    block = _read_dv_block(out)
    assert block is not None
    assert "<formula1>Sheet2!$A$1:$A$5</formula1>" in block
    # operator MUST be omitted for type=list per OOXML §18.3.1.32
    assert "operator=" not in block, f"operator unexpectedly emitted: {block}"


def test_dv_whole_between(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    _make_clean_fixture(src)
    wb = Workbook._from_patcher(str(src))
    wb["Sheet1"].data_validations.append(
        DataValidation(
            type="whole",
            operator="between",
            formula1="1",
            formula2="100",
            sqref="C2:C100",
        )
    )
    out = tmp_path / "out.xlsx"
    wb.save(out)
    block = _read_dv_block(out)
    assert block is not None
    assert "<formula1>1</formula1>" in block
    assert "<formula2>100</formula2>" in block


def test_dv_custom_type(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    _make_clean_fixture(src)
    wb = Workbook._from_patcher(str(src))
    wb["Sheet1"].data_validations.append(
        DataValidation(type="custom", formula1="=LEN(A1)>5", sqref="A1")
    )
    out = tmp_path / "out.xlsx"
    wb.save(out)
    block = _read_dv_block(out)
    assert block is not None
    assert 'type="custom"' in block
    # operator MUST be omitted for type=custom
    assert "operator=" not in block, f"operator unexpectedly emitted: {block}"
    # `>` inside text content must be escaped
    assert "&gt;" in block, f"expected escape of '>', got: {block}"


def test_dv_multi_sqref(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    _make_clean_fixture(src)
    wb = Workbook._from_patcher(str(src))
    wb["Sheet1"].data_validations.append(
        DataValidation(type="list", formula1='"X,Y"', sqref="A1:A10 C1:C10")
    )
    out = tmp_path / "out.xlsx"
    wb.save(out)
    block = _read_dv_block(out)
    assert block is not None
    # Space-separated multi-range MUST pass through unchanged.
    assert 'sqref="A1:A10 C1:C10"' in block, f"sqref mangled: {block}"


def test_dv_count_attribute_matches(tmp_path: Path) -> None:
    """count must equal the literal number of <dataValidation> child tags."""
    src = tmp_path / "src.xlsx"
    _make_fixture_with_one_dv(src)
    wb = Workbook._from_patcher(str(src))
    ws = wb["Sheet1"]
    ws.data_validations.append(
        DataValidation(type="list", formula1='"P,Q"', sqref="D1:D10")
    )
    ws.data_validations.append(
        DataValidation(type="whole", operator="equal", formula1="42", sqref="E1")
    )
    out = tmp_path / "out.xlsx"
    wb.save(out)
    block = _read_dv_block(out)
    assert block is not None
    # 1 existing + 2 new
    assert 'count="3"' in block, f"expected count=3, got: {block}"
    assert block.count("<dataValidation ") == 3 or block.count("<dataValidation") == 3, (
        f"expected 3 child tags, got: {block}"
    )


def test_dv_round_trip_via_load_workbook(tmp_path: Path) -> None:
    """After modify+save, opening the file again surfaces the new DVs."""
    src = tmp_path / "src.xlsx"
    _make_clean_fixture(src)

    wb = Workbook._from_patcher(str(src))
    wb["Sheet1"].data_validations.append(
        DataValidation(
            type="whole", operator="between", formula1="1", formula2="9", sqref="A2:A10"
        )
    )
    out = tmp_path / "out.xlsx"
    wb.save(out)
    wb.close()

    # Re-open via openpyxl (independent reader) and confirm the DV survived.
    wb2 = openpyxl.load_workbook(out)
    ws2 = wb2["Sheet1"]
    sqrefs = [str(dv.sqref) for dv in ws2.data_validations.dataValidation]
    assert any("A2:A10" in s for s in sqrefs), f"DV missing on re-open, found: {sqrefs}"


def test_dv_no_pending_no_op(tmp_path: Path) -> None:
    """Open + save with no DV append must produce byte-identical output.

    Regression guard for the do_save short-circuit predicate: if the
    queued_dv_patches check were missing, even a no-op save would
    re-zip every entry and risk drifting bytes.
    """
    os.environ["WOLFXL_TEST_EPOCH"] = "0"
    src = tmp_path / "src.xlsx"
    _make_clean_fixture(src)

    wb = Workbook._from_patcher(str(src))
    out = tmp_path / "out.xlsx"
    wb.save(out)
    wb.close()

    # The short-circuit path uses fs::copy, so the output is the source
    # file byte-for-byte (no re-zip, no re-deflation).
    assert src.read_bytes() == out.read_bytes(), (
        "no-op save must produce a byte-identical file (short-circuit broken?)"
    )


def test_dv_error_style_warning(tmp_path: Path) -> None:
    """When errorStyle='warning' is supplied, it appears in the output. The
    default 'stop' is omitted.

    Note: ``DataValidation`` doesn't expose ``errorStyle`` as a field today
    (see datavalidation.py:35-46 — only error/errorTitle), so we set it via
    setattr to exercise the patcher pathway. If a future expansion adds the
    field to the dataclass, this test still passes.
    """
    src = tmp_path / "src.xlsx"
    _make_clean_fixture(src)
    wb = Workbook._from_patcher(str(src))
    dv = DataValidation(type="whole", operator="equal", formula1="5", sqref="A1")
    setattr(dv, "errorStyle", "warning")  # noqa: B010
    wb["Sheet1"].data_validations.append(dv)
    out = tmp_path / "out.xlsx"
    wb.save(out)
    block = _read_dv_block(out)
    assert block is not None
    assert 'errorStyle="warning"' in block, f"expected errorStyle=warning: {block}"


# ---------------------------------------------------------------------------
# Sanity: skip everything if maturin extension isn't built.
# ---------------------------------------------------------------------------

@pytest.fixture(autouse=True)
def _force_test_epoch(monkeypatch: pytest.MonkeyPatch) -> None:
    """Pin ZIP entry mtimes to the Unix epoch so the no-op-save test gets
    byte-identical output regardless of when fixtures were created."""
    monkeypatch.setenv("WOLFXL_TEST_EPOCH", "0")
