"""Sprint Μ Pod-γ (RFC-046 §6) — modify-mode ``add_chart`` coverage.

The patcher's Phase 2.5l drains pending chart adds into:
- ``xl/charts/chartN.xml``  (caller-supplied bytes routed verbatim)
- ``xl/drawings/drawingN.xml`` + nested rels  (synthesized OR merged)
- The sheet's rels graph   (drawing rel re-used or freshly added)
- ``[Content_Types].xml``  (chart + drawing overrides)

These tests cover the integrator-facing public API
``Workbook.add_chart_modify_mode(sheet, chart_xml, anchor_a1)``. Pod-β
is expected to wire ``Worksheet.add_chart`` on top of this stub
(routing through Pod-α's ``emit_chart_xml``); when those land they
must keep the same downstream contract.
"""
from __future__ import annotations

import re
import zipfile
from pathlib import Path

import openpyxl
import pytest

from wolfxl import load_workbook


pytestmark = pytest.mark.rfc046 if hasattr(pytest.mark, "rfc046") else pytest.mark.usefixtures()


@pytest.fixture(autouse=True)
def _force_test_epoch(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("WOLFXL_TEST_EPOCH", "0")


# ---------------------------------------------------------------------------
# Fixture helpers
# ---------------------------------------------------------------------------

# Minimal `<chartSpace>` body with one BarChart series referencing
# A1:B4 on the embedding sheet. Caller passes ``sheet_title`` so the
# `<f>` formula uses the right sheet ref.
_CHART_TMPL_BAR = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<c:chartSpace xmlns:c="http://schemas.openxmlformats.org/drawingml/2006/chart" xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">
<c:chart><c:plotArea>
<c:barChart><c:barDir val="col"/><c:grouping val="clustered"/>
<c:ser><c:idx val="0"/><c:order val="0"/>
<c:cat><c:numRef><c:f>'{sheet_title}'!$A$1:$A$4</c:f></c:numRef></c:cat>
<c:val><c:numRef><c:f>'{sheet_title}'!$B$1:$B$4</c:f></c:numRef></c:val>
</c:ser>
<c:axId val="1"/><c:axId val="2"/>
</c:barChart>
<c:catAx><c:axId val="1"/><c:scaling><c:orientation val="minMax"/></c:scaling><c:crossAx val="2"/></c:catAx>
<c:valAx><c:axId val="2"/><c:scaling><c:orientation val="minMax"/></c:scaling><c:crossAx val="1"/></c:valAx>
</c:plotArea></c:chart></c:chartSpace>"""


def _bar_chart_xml(sheet_title: str = "Data") -> bytes:
    return _CHART_TMPL_BAR.format(sheet_title=sheet_title).encode("utf-8")


def _make_data_fixture(path: Path, sheet_title: str = "Data") -> None:
    """A1:B4 mini table so the chart's data range is real."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    for r in range(1, 5):
        ws.cell(r, 1, f"l{r}")
        ws.cell(r, 2, r * 10)
    wb.save(path)


def _make_data_with_image_fixture(path: Path, sheet_title: str = "Data") -> None:
    """A1:B4 mini table + one PNG image so the sheet has an existing
    drawing rel — exercises the SAX-merge into existing drawing path.
    """
    from openpyxl.drawing.image import Image

    # Build a 1x1 PNG.
    img_bytes = (
        b"\x89PNG\r\n\x1a\n"
        b"\x00\x00\x00\rIHDR"
        b"\x00\x00\x00\x01\x00\x00\x00\x01\x08\x06\x00\x00\x00\x1f\x15\xc4\x89"
        b"\x00\x00\x00\rIDATx\x9cc\xfc\x0f\x00\x00\x01\x01\x01\x00\x18\xdd\x8d\xb4"
        b"\x00\x00\x00\x00IEND\xaeB`\x82"
    )
    tmp_png = path.parent / "_tiny.png"
    tmp_png.write_bytes(img_bytes)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    for r in range(1, 5):
        ws.cell(r, 1, f"l{r}")
        ws.cell(r, 2, r * 10)
    img = Image(str(tmp_png))
    img.anchor = "F2"
    ws.add_image(img)
    wb.save(path)
    tmp_png.unlink(missing_ok=True)


def _make_data_with_chart_fixture(path: Path, sheet_title: str = "Data") -> None:
    """A1:B4 + a pre-existing chart so we can layer on a SECOND chart
    via modify mode.
    """
    from openpyxl.chart import BarChart, Reference

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    for r in range(1, 5):
        ws.cell(r, 1, f"l{r}")
        ws.cell(r, 2, r * 10)
    ch = BarChart()
    ch.title = "first"
    data = Reference(ws, min_col=2, min_row=1, max_row=4)
    ch.add_data(data, titles_from_data=False)
    ws.add_chart(ch, "D2")
    wb.save(path)


def _make_data_with_table_fixture(path: Path, sheet_title: str = "Data") -> None:
    """A1:B4 + an Excel table — tests RFC-024 / RFC-046 composition
    (chart-add doesn't disturb table parts).
    """
    from openpyxl.worksheet.table import Table, TableStyleInfo

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.cell(1, 1, "Label")
    ws.cell(1, 2, "Value")
    for r in range(2, 5):
        ws.cell(r, 1, f"l{r}")
        ws.cell(r, 2, r * 10)
    tbl = Table(displayName="DataTbl", ref="A1:B4")
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2")
    ws.add_table(tbl)
    wb.save(path)


def _zip_listing(path: Path) -> list[str]:
    with zipfile.ZipFile(path, "r") as z:
        return sorted(z.namelist())


def _zip_read(path: Path, member: str) -> bytes:
    with zipfile.ZipFile(path, "r") as z:
        return z.read(member)


# ---------------------------------------------------------------------------
# Case 1 — Add a chart to a workbook with NO existing drawing.
# ---------------------------------------------------------------------------


def test_add_chart_to_workbook_without_existing_drawings(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_data_fixture(src)
    wb = load_workbook(src, modify=True)
    wb.add_chart_modify_mode("Data", _bar_chart_xml("Data"), "D2")
    wb.save(dst)

    entries = _zip_listing(dst)
    chart_files = [e for e in entries if e.startswith("xl/charts/chart")]
    drawing_files = [
        e for e in entries
        if re.match(r"^xl/drawings/drawing\d+\.xml$", e)
    ]
    assert chart_files, f"no chart part emitted; entries={entries}"
    assert drawing_files, f"no drawing part emitted; entries={entries}"
    # Sheet rels point at the new drawing.
    sheet_rels = _zip_read(dst, "xl/worksheets/_rels/sheet1.xml.rels")
    assert b"drawings/drawing" in sheet_rels
    # Drawing rels point at our chart.
    drawing_rels = _zip_read(
        dst, "xl/drawings/_rels/" + drawing_files[0].rsplit("/", 1)[1] + ".rels"
    )
    assert b"charts/chart" in drawing_rels
    # Content-Types has chart override.
    ct = _zip_read(dst, "[Content_Types].xml")
    assert b"drawingml.chart+xml" in ct
    # Sheet XML has a <drawing r:id="..."/> ref.
    sheet_xml = _zip_read(dst, "xl/worksheets/sheet1.xml")
    assert b"<drawing" in sheet_xml


# ---------------------------------------------------------------------------
# Case 2 — Workbook ALREADY has a drawing (image), add a chart on top.
#          Phase 2.5l SAX-merges the new graphicFrame into the
#          existing drawing XML.
# ---------------------------------------------------------------------------


def test_add_chart_to_workbook_with_existing_drawing(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_data_with_image_fixture(src)
    wb = load_workbook(src, modify=True)
    wb.add_chart_modify_mode("Data", _bar_chart_xml("Data"), "D2")
    wb.save(dst)

    entries = _zip_listing(dst)
    chart_files = [e for e in entries if e.startswith("xl/charts/chart")]
    drawing_files = [
        e for e in entries
        if re.match(r"^xl/drawings/drawing\d+\.xml$", e)
    ]
    assert len(chart_files) == 1
    # Existing image drawing should still exist (count = 1; we
    # appended into it instead of creating a fresh one).
    assert len(drawing_files) == 1, (
        f"expected SAX-merge into existing drawing; got {drawing_files}"
    )
    # Original image still present.
    assert any(e.startswith("xl/media/image") for e in entries), (
        "image media should be preserved after chart-add"
    )
    # Drawing XML now contains BOTH the original <xdr:pic> AND a
    # newly-spliced <xdr:graphicFrame>.
    drawing_body = _zip_read(dst, drawing_files[0]).decode()
    # openpyxl emits default-namespaced (no `xdr:` prefix); accept either.
    assert "<pic" in drawing_body or "<xdr:pic" in drawing_body, (
        f"original image anchor must survive; got {drawing_body[:600]}"
    )
    assert "graphicFrame" in drawing_body, (
        f"chart frame must be appended; got {drawing_body[:600]}"
    )


# ---------------------------------------------------------------------------
# Case 3 — Workbook already has a chart, add ANOTHER one. Both
#          should be visible after round-trip.
# ---------------------------------------------------------------------------


def test_add_chart_to_workbook_with_existing_chart(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_data_with_chart_fixture(src)
    wb = load_workbook(src, modify=True)
    wb.add_chart_modify_mode("Data", _bar_chart_xml("Data"), "H2")
    wb.save(dst)

    entries = _zip_listing(dst)
    chart_files = sorted(e for e in entries if e.startswith("xl/charts/chart"))
    assert len(chart_files) == 2, (
        f"expected 2 chart parts (1 original + 1 added); got {chart_files}"
    )
    # Drawing XML now references both charts.
    drawing_files = [
        e for e in entries
        if re.match(r"^xl/drawings/drawing\d+\.xml$", e)
    ]
    assert len(drawing_files) == 1
    drawing_body = _zip_read(dst, drawing_files[0]).decode()
    # At least two graphicFrames after merge — accept either prefix style.
    n_frames = (
        drawing_body.count("<xdr:graphicFrame")
        + drawing_body.count("<graphicFrame")
    )
    assert n_frames >= 2, f"expected >= 2 graphicFrames; body={drawing_body[:500]}"


# ---------------------------------------------------------------------------
# Case 4 — Add a chart and modify a data cell IN THE SAME save.
#          Phase 2.5l drains BEFORE Phase 3 cell patches; both must
#          land cleanly.
# ---------------------------------------------------------------------------


def test_add_chart_then_modify_data_cells_same_save(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_data_fixture(src)
    wb = load_workbook(src, modify=True)
    # First, queue a chart.
    wb.add_chart_modify_mode("Data", _bar_chart_xml("Data"), "D2")
    # Then mutate a cell that is part of the chart's data range.
    ws = wb["Data"]
    ws["B1"] = 999
    wb.save(dst)

    # Cell rewrite landed.
    op = openpyxl.load_workbook(dst, data_only=False)
    assert op["Data"]["B1"].value == 999
    # Chart present in zip.
    chart_files = [
        e for e in _zip_listing(dst) if e.startswith("xl/charts/chart")
    ]
    assert chart_files, "chart not emitted alongside cell rewrite"


# ---------------------------------------------------------------------------
# Case 5 — Cross-RFC composition: existing table + new chart.
# ---------------------------------------------------------------------------


def test_add_chart_to_workbook_with_existing_table_drawing(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_data_with_table_fixture(src)
    wb = load_workbook(src, modify=True)
    wb.add_chart_modify_mode("Data", _bar_chart_xml("Data"), "D2")
    wb.save(dst)

    entries = _zip_listing(dst)
    # Table + chart + drawing all present.
    assert any(e.startswith("xl/tables/") for e in entries), entries
    assert any(e.startswith("xl/charts/") for e in entries), entries
    assert any(
        re.match(r"^xl/drawings/drawing\d+\.xml$", e) for e in entries
    ), entries


# ---------------------------------------------------------------------------
# Case 6 — Round-trip via openpyxl as parser. Validates the emitted
#          OOXML is loadable.
# ---------------------------------------------------------------------------


def test_add_chart_via_modify_then_load_via_openpyxl(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_data_fixture(src)
    wb = load_workbook(src, modify=True)
    wb.add_chart_modify_mode("Data", _bar_chart_xml("Data"), "D2")
    wb.save(dst)
    # openpyxl must accept the file.
    op = openpyxl.load_workbook(dst, data_only=False)
    ws = op["Data"]
    assert ws.cell(1, 1).value == "l1"
    # And the chart should round-trip into openpyxl's chart list.
    assert len(ws._charts) >= 1, ws._charts


# ---------------------------------------------------------------------------
# Case 7 — Multiple charts queued in one call site, single drawing.
# ---------------------------------------------------------------------------


def test_add_multiple_charts_to_same_sheet(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_data_fixture(src)
    wb = load_workbook(src, modify=True)
    wb.add_chart_modify_mode("Data", _bar_chart_xml("Data"), "D2")
    wb.add_chart_modify_mode("Data", _bar_chart_xml("Data"), "L2")
    wb.save(dst)
    entries = _zip_listing(dst)
    chart_files = sorted(e for e in entries if e.startswith("xl/charts/chart"))
    drawing_files = [
        e for e in entries if re.match(r"^xl/drawings/drawing\d+\.xml$", e)
    ]
    assert len(chart_files) == 2, chart_files
    # Both charts share ONE fresh drawing (same-call adds collapse).
    assert len(drawing_files) == 1, drawing_files
    body = _zip_read(dst, drawing_files[0]).decode()
    n_frames = body.count("<xdr:graphicFrame") + body.count("<graphicFrame")
    assert n_frames == 2, f"expected 2 graphicFrames; body={body[:500]}"


# ---------------------------------------------------------------------------
# Case 8 — Validation: bad sheet title, bad chart bytes, bad anchor.
# ---------------------------------------------------------------------------


def test_add_chart_validation_errors(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    _make_data_fixture(src)
    wb = load_workbook(src, modify=True)
    with pytest.raises(ValueError):
        wb.add_chart_modify_mode("NoSuchSheet", _bar_chart_xml("Data"), "D2")
    with pytest.raises(TypeError):
        wb.add_chart_modify_mode("Data", "not-bytes", "D2")  # type: ignore[arg-type]
    with pytest.raises(ValueError):
        wb.add_chart_modify_mode("Data", _bar_chart_xml("Data"), "")


# ===========================================================================
# Sprint Μ-prime Pod-γ′ — modify-mode HIGH-LEVEL chart bridge
# (Worksheet.add_chart(BarChart()) wired through serialize_chart_dict).
#
# These tests exercise the dict→bytes bridge added in
# ``_flush_pending_charts_to_patcher`` (RFC-046 §10.12). They depend on
# Pod-α′'s ``serialize_chart_dict`` and Pod-β′'s ``to_rust_dict``
# flat-shape — when Pod-α′/β′ haven't merged yet they will raise
# NotImplementedError at save time. The integrator validates end-to-end
# post-merge.
# ===========================================================================


def test_modify_mode_add_bar_chart_high_level(tmp_path: Path) -> None:
    """Load existing xlsx, ws.add_chart(BarChart()), save, reload,
    verify chart present."""
    from wolfxl.chart import BarChart, Reference

    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_data_fixture(src)
    wb = load_workbook(src, modify=True)
    ws = wb["Data"]
    chart = BarChart()
    chart.title = "From-modify"
    chart.add_data(
        Reference(ws, min_col=2, min_row=1, max_row=4),
        titles_from_data=False,
    )
    ws.add_chart(chart, "D2")
    wb.save(dst)

    entries = _zip_listing(dst)
    chart_files = [e for e in entries if e.startswith("xl/charts/chart")]
    assert chart_files, f"high-level chart not emitted; entries={entries}"
    drawing_files = [
        e for e in entries
        if re.match(r"^xl/drawings/drawing\d+\.xml$", e)
    ]
    assert drawing_files, f"no drawing part emitted; entries={entries}"
    # openpyxl must accept the resulting file.
    op = openpyxl.load_workbook(dst, data_only=False)
    assert len(op["Data"]._charts) >= 1


def test_modify_mode_add_line_chart_high_level(tmp_path: Path) -> None:
    """Same path, with LineChart instead of BarChart."""
    from wolfxl.chart import LineChart, Reference

    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_data_fixture(src)
    wb = load_workbook(src, modify=True)
    ws = wb["Data"]
    chart = LineChart()
    chart.add_data(
        Reference(ws, min_col=2, min_row=1, max_row=4),
        titles_from_data=False,
    )
    ws.add_chart(chart, "D2")
    wb.save(dst)

    entries = _zip_listing(dst)
    chart_files = [e for e in entries if e.startswith("xl/charts/chart")]
    assert chart_files, f"line chart not emitted; entries={entries}"
    op = openpyxl.load_workbook(dst, data_only=False)
    assert len(op["Data"]._charts) >= 1


def test_modify_mode_add_chart_with_existing_drawing(tmp_path: Path) -> None:
    """Sheet has an image, add a high-level chart, both survive
    (drawing SAX-merge)."""
    from wolfxl.chart import BarChart, Reference

    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_data_with_image_fixture(src)
    wb = load_workbook(src, modify=True)
    ws = wb["Data"]
    chart = BarChart()
    chart.add_data(
        Reference(ws, min_col=2, min_row=1, max_row=4),
        titles_from_data=False,
    )
    ws.add_chart(chart, "D2")
    wb.save(dst)

    entries = _zip_listing(dst)
    chart_files = [e for e in entries if e.startswith("xl/charts/chart")]
    drawing_files = [
        e for e in entries
        if re.match(r"^xl/drawings/drawing\d+\.xml$", e)
    ]
    assert len(chart_files) == 1, chart_files
    # Drawing should be SAX-merged (still 1 drawing part).
    assert len(drawing_files) == 1, drawing_files
    # Original image media must still be present.
    assert any(e.startswith("xl/media/image") for e in entries), entries
    drawing_body = _zip_read(dst, drawing_files[0]).decode()
    assert "<pic" in drawing_body or "<xdr:pic" in drawing_body, (
        f"original image must survive; got {drawing_body[:600]}"
    )
    assert "graphicFrame" in drawing_body, (
        f"chart frame must be appended; got {drawing_body[:600]}"
    )


def test_modify_mode_add_chart_then_copy_worksheet_deep_clones(
    tmp_path: Path,
) -> None:
    """RFC-035 §10 composition: add a high-level chart in modify mode
    AND copy the same sheet in the SAME save. Both the new chart and
    the copy's deep-cloned chart must land cleanly."""
    from wolfxl.chart import BarChart, Reference

    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_data_fixture(src, sheet_title="Data")
    wb = load_workbook(src, modify=True)
    ws = wb["Data"]
    chart = BarChart()
    chart.add_data(
        Reference(ws, min_col=2, min_row=1, max_row=4),
        titles_from_data=False,
    )
    ws.add_chart(chart, "D2")
    # And copy the worksheet in the same save.
    wb.copy_worksheet(ws)
    wb.save(dst)

    entries = _zip_listing(dst)
    chart_files = [e for e in entries if e.startswith("xl/charts/chart")]
    # Source had no charts but we add 1 high-level then copy → the
    # high-level chart goes onto the source, and the copy deep-clones
    # the pending chart in the same save.
    assert len(chart_files) == 2, f"expected source+copy charts; entries={entries}"


def test_modify_mode_add_3d_chart_works(tmp_path: Path) -> None:
    """End-to-end via the bridge with a 3D chart family (Pod-β′
    BarChart3D). Verifies the bridge is chart-kind-agnostic."""
    pytest.importorskip("wolfxl.chart")
    from wolfxl.chart import BarChart3D, Reference

    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_data_fixture(src)
    wb = load_workbook(src, modify=True)
    ws = wb["Data"]
    chart = BarChart3D()
    chart.add_data(
        Reference(ws, min_col=2, min_row=1, max_row=4),
        titles_from_data=False,
    )
    ws.add_chart(chart, "D2")
    wb.save(dst)

    entries = _zip_listing(dst)
    chart_files = [e for e in entries if e.startswith("xl/charts/chart")]
    assert chart_files, f"3D chart not emitted; entries={entries}"


def test_modify_mode_high_level_no_longer_warns(tmp_path: Path) -> None:
    """Sprint Μ-prime: the v1.6.0 'warn-and-drop' path is gone — adding
    a high-level chart in modify mode should NOT emit a RuntimeWarning."""
    import warnings as _warnings

    from wolfxl.chart import BarChart, Reference

    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_data_fixture(src)
    wb = load_workbook(src, modify=True)
    ws = wb["Data"]
    chart = BarChart()
    chart.add_data(
        Reference(ws, min_col=2, min_row=1, max_row=4),
        titles_from_data=False,
    )
    ws.add_chart(chart, "D2")
    with _warnings.catch_warnings(record=True) as caught:
        _warnings.simplefilter("always")
        wb.save(dst)
    relevant = [
        w for w in caught
        if issubclass(w.category, RuntimeWarning)
        and "high-level Worksheet.add_chart" in str(w.message)
    ]
    assert not relevant, (
        f"v1.6.0 warn-and-drop path should be gone; got {[str(w.message) for w in relevant]}"
    )


# ---------------------------------------------------------------------------
# Sprint Μ-prime Pod-γ′ — bytes-level escape hatch regression.
#
# The high-level bridge must NOT break the v1.6.0
# Workbook.add_chart_modify_mode(sheet, chart_xml: bytes, anchor, ...)
# escape hatch; both code paths drain in the same flush call.
# ---------------------------------------------------------------------------


def test_bytes_escape_hatch_still_works_after_high_level_bridge(
    tmp_path: Path,
) -> None:
    """Use both pathways in the same save — bytes escape hatch and
    high-level Worksheet.add_chart — verify both charts land."""
    from wolfxl.chart import BarChart, Reference

    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_data_fixture(src)
    wb = load_workbook(src, modify=True)
    # Path A: bytes escape hatch.
    wb.add_chart_modify_mode("Data", _bar_chart_xml("Data"), "D2")
    # Path B: high-level bridge.
    ws = wb["Data"]
    chart = BarChart()
    chart.add_data(
        Reference(ws, min_col=2, min_row=1, max_row=4),
        titles_from_data=False,
    )
    ws.add_chart(chart, "L2")
    wb.save(dst)

    entries = _zip_listing(dst)
    chart_files = sorted(
        e for e in entries if e.startswith("xl/charts/chart")
    )
    assert len(chart_files) == 2, (
        f"expected 2 chart parts (escape hatch + bridge); got {chart_files}"
    )
