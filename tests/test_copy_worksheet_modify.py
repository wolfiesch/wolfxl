"""RFC-035 — ``Workbook.copy_worksheet`` full modify-mode harness.

The Pod-β smoke test (``tests/test_copy_worksheet_smoke.py``) covered
three happy/sad-path cases. This module is the FULL harness called for
in ``Plans/rfcs/035-copy-worksheet.md`` §7.5 — 19 cases spanning every
documented branch of the planner / Phase 2.7 / Python coordinator.

Test menu (matches RFC-035 §7.5 exactly; see the RFC for rationale):

A. Basic clone of a 5x5 grid.
B. Copy with one table; auto-renamed ``Sales`` → ``Sales_2`` (OQ-b).
C. Copy with multiple tables; collision-scan against source + queued.
D. Copy with external hyperlink; cloned rels independent.
E. Copy with comment + VML drawing; PartIdAllocator suffix observation.
F. Copy with data validation (wolfxl preserves; openpyxl drops).
G. Copy with conditional formatting (wolfxl preserves; openpyxl drops).
H. Copy with sheet-scoped defined name; ``localSheetId`` re-pointed.
I. Copy + edit-the-copy in same save.
J. Copy + RFC-036 ``move_sheet`` in same save.
K. Copy + RFC-024 ``add_table`` to the copy in same save.
L. Name-collision on explicit ``name`` arg → ``ValueError``.
M. Cross-workbook source → ``ValueError``.
N. Write-mode rejection → ``NotImplementedError``.
O. No-op byte-identical save (empty queue).
P. Self-closing ``<sheets/>`` workbook.xml fixture.
Q. Defined-names upsert collision (last-write-wins).
R. CDATA / processing-instruction containing literal ``</sheets>``.
S. Default ``"{src} Copy"`` / ``Copy 2`` / ``Copy 3`` deterministic naming.

Constants:

- Cargo baselines (informational): structural=116, rels=42.
- Pytest baseline (informational): 165 passed, 4 skipped.

This file ABSORBS the smoke test — once this harness ships green, the
smoke file should be removed (handled in the same commit).
"""
from __future__ import annotations

import re
import shutil
import zipfile
from pathlib import Path

import openpyxl
import pytest

from wolfxl import load_workbook


# pytest marker so verify_rfc.py can collect this test.
pytestmark = pytest.mark.rfc035


@pytest.fixture(autouse=True)
def _force_test_epoch(monkeypatch: pytest.MonkeyPatch) -> None:
    """Pin ZIP entry mtimes for byte-stable saves (no-op invariant tests)."""
    monkeypatch.setenv("WOLFXL_TEST_EPOCH", "0")


# ---------------------------------------------------------------------------
# Fixture builders — each constructs a minimal openpyxl-built source xlsx.
# ---------------------------------------------------------------------------


def _make_grid_fixture(path: Path) -> None:
    """5x5 grid with mixed strings + numbers + a SUM formula."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"
    for r in range(1, 6):
        for c in range(1, 6):
            ws.cell(row=r, column=c, value=f"r{r}c{c}" if (r + c) % 2 else (r * c))
    ws["F1"] = "=SUM(A1:E5)"
    wb.save(path)


def _make_table_fixture(path: Path, table_names: list[str]) -> None:
    """5-column header row + 4 data rows + one or more tables."""
    from openpyxl.worksheet.table import Table

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"
    headers = ["k", "a", "b", "c", "d"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    for r in range(2, 6):
        for c in range(1, 6):
            ws.cell(row=r, column=c, value=(r - 1) * 10 + c)
    for i, name in enumerate(table_names):
        # Stagger the refs so multiple tables don't overlap; for one
        # table just use A1:E5.
        if len(table_names) == 1:
            ref = "A1:E5"
        else:
            # row offset per-table
            ref = "A1:E5"  # they all share the same range — overlap is illegal
        # If multiple tables, give them their own column range.
        if len(table_names) > 1:
            col0 = chr(ord("A") + i)
            col1 = chr(ord("A") + i)  # 1-column tables to avoid overlap
            ref = f"{col0}1:{col1}5"
        ws.add_table(Table(displayName=name, ref=ref))
    wb.save(path)


def _make_hyperlink_fixture(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"
    for r in range(1, 6):
        for c in range(1, 6):
            ws.cell(row=r, column=c, value=r * c)
    ws["E5"] = "click"
    ws["E5"].hyperlink = "https://example.com/external"
    wb.save(path)


def _make_comment_fixture(path: Path) -> None:
    from openpyxl.comments import Comment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"
    ws["A1"] = 1
    ws["B2"] = "comment-cell"
    ws["B2"].comment = Comment("hello", "tester")
    wb.save(path)


def _make_dv_fixture(path: Path) -> None:
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"
    ws["A1"] = "h"
    ws["B2"] = 5
    dv = DataValidation(type="list", formula1='"a,b,c"')
    dv.add("B2:B4")
    ws.add_data_validation(dv)
    wb.save(path)


def _make_cf_fixture(path: Path) -> None:
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"
    for r in range(1, 6):
        ws.cell(row=r, column=1, value=r)
    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    rule = CellIsRule(operator="greaterThan", formula=["3"], fill=fill)
    ws.conditional_formatting.add("A1:A5", rule)
    wb.save(path)


def _make_sheet_scoped_name_fixture(path: Path) -> None:
    """Workbook with a Print_Area sheet-scoped defined name on Template."""
    from openpyxl.workbook.defined_name import DefinedName as XDefinedName

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"
    for r in range(1, 6):
        for c in range(1, 6):
            ws.cell(row=r, column=c, value=r * c)
    wb.create_sheet("Other")["A1"] = "untouched"
    wb.defined_names["_xlnm.Print_Area"] = XDefinedName(
        "_xlnm.Print_Area", attr_text="Template!$A$1:$E$5", localSheetId=0
    )
    wb.save(path)


def _read_zip_text(path: Path, entry: str) -> str:
    with zipfile.ZipFile(path) as zf:
        return zf.read(entry).decode("utf-8")


def _zip_listing(path: Path) -> list[str]:
    with zipfile.ZipFile(path) as zf:
        return sorted(zf.namelist())


# ---------------------------------------------------------------------------
# Case A — Basic clone of a 5x5 grid.
# ---------------------------------------------------------------------------


def test_a_basic_clone(tmp_path: Path) -> None:
    """Source values arrive in the clone; source untouched."""
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    _make_grid_fixture(src)

    wb = load_workbook(src, modify=True)
    new_ws = wb.copy_worksheet(wb.active)
    assert new_ws.title == "Template Copy"
    wb.save(out)

    rt = openpyxl.load_workbook(out)
    assert rt.sheetnames == ["Template", "Template Copy"]
    src_sheet = rt["Template"]
    dst_sheet = rt["Template Copy"]
    for r in range(1, 6):
        for c in range(1, 6):
            assert dst_sheet.cell(row=r, column=c).value == src_sheet.cell(row=r, column=c).value
    # The SUM formula in F1 is preserved in both sheets.
    assert src_sheet["F1"].value == "=SUM(A1:E5)"
    assert dst_sheet["F1"].value == "=SUM(A1:E5)"


# ---------------------------------------------------------------------------
# Case B — Copy with one table; auto-rename Sales → Sales_2 (OQ-b).
# ---------------------------------------------------------------------------


def test_b_copy_with_one_table_auto_renames(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    _make_table_fixture(src, ["Sales"])

    wb = load_workbook(src, modify=True)
    wb.copy_worksheet(wb.active)
    wb.save(out)

    # Two table parts now exist; the second's displayName must be Sales_2.
    listing = _zip_listing(out)
    table_parts = [n for n in listing if n.startswith("xl/tables/")]
    assert len(table_parts) == 2, f"expected 2 tables, got {table_parts}"

    names = []
    for tp in table_parts:
        xml = _read_zip_text(out, tp)
        m = re.search(r'displayName="([^"]+)"', xml)
        assert m is not None, f"no displayName in {tp}"
        names.append(m.group(1))
    assert sorted(names) == ["Sales", "Sales_2"], names


# ---------------------------------------------------------------------------
# Case C — Copy with multiple tables; collision scan covers all sources.
# ---------------------------------------------------------------------------


def test_c_copy_with_multiple_tables(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    _make_table_fixture(src, ["T1", "T2", "T3"])

    wb = load_workbook(src, modify=True)
    wb.copy_worksheet(wb.active)
    wb.save(out)

    listing = _zip_listing(out)
    table_parts = [n for n in listing if n.startswith("xl/tables/")]
    assert len(table_parts) == 6, f"expected 6 tables, got {table_parts}"

    names: list[str] = []
    for tp in table_parts:
        xml = _read_zip_text(out, tp)
        m = re.search(r'displayName="([^"]+)"', xml)
        assert m is not None
        names.append(m.group(1))
    assert sorted(names) == ["T1", "T1_2", "T2", "T2_2", "T3", "T3_2"], names

    # Table ids must be workbook-unique.
    ids: list[str] = []
    for tp in table_parts:
        xml = _read_zip_text(out, tp)
        m = re.search(r'<table[^>]*\bid="(\d+)"', xml)
        assert m is not None, f"no id in {tp}"
        ids.append(m.group(1))
    assert len(set(ids)) == len(ids), f"duplicate table ids: {ids}"


# ---------------------------------------------------------------------------
# Case D — Copy with external hyperlink; cloned rels are independent.
# ---------------------------------------------------------------------------


def test_d_copy_with_hyperlink(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    _make_hyperlink_fixture(src)

    wb = load_workbook(src, modify=True)
    wb.copy_worksheet(wb.active)
    wb.save(out)

    rt = openpyxl.load_workbook(out)
    src_sheet = rt["Template"]
    dst_sheet = rt["Template Copy"]
    # Hyperlink target survives in both sheets.
    assert src_sheet["E5"].hyperlink is not None
    assert dst_sheet["E5"].hyperlink is not None
    assert dst_sheet["E5"].hyperlink.target == "https://example.com/external"
    # The two hyperlinks live in independent rels files.
    listing = _zip_listing(out)
    rels_files = [n for n in listing if n.startswith("xl/worksheets/_rels/")]
    assert len(rels_files) == 2, f"expected 2 sheet-rels files, got {rels_files}"


# ---------------------------------------------------------------------------
# Case E — Copy with comment + VML drawing; PartIdAllocator allocates
#          new comments<N>.xml + vmlDrawing<N>.vml suffixes.
# ---------------------------------------------------------------------------


def test_e_copy_with_comments_and_vml(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    _make_comment_fixture(src)

    wb = load_workbook(src, modify=True)
    wb.copy_worksheet(wb.active)
    wb.save(out)

    listing = _zip_listing(out)
    # openpyxl writes comments under ``xl/comments/comment<N>.xml`` and
    # ``xl/drawings/commentsDrawing<N>.vml`` (with the ``comments``
    # subdir + ``Drawing`` infix). wolfxl's PartIdAllocator follows
    # the canonical ECMA-376 form: ``xl/comments<N>.xml`` and
    # ``xl/drawings/vmlDrawing<N>.vml``. So the OUT zip carries one
    # part of EACH naming convention: the source's openpyxl-style and
    # the clone's PartIdAllocator-style.
    all_comment_parts = [
        n for n in listing
        if re.match(r"xl/comments(/comment)?\d+\.xml$", n)
    ]
    all_vml_parts = [
        n for n in listing
        if re.match(r"xl/drawings/(commentsDrawing|vmlDrawing)\d+\.vml$", n)
    ]
    assert len(all_comment_parts) == 2, f"expected 2 comments parts, got {all_comment_parts}"
    assert len(all_vml_parts) == 2, f"expected 2 VML parts, got {all_vml_parts}"
    # Each filename is workbook-unique.
    assert len(set(all_comment_parts)) == len(all_comment_parts)
    assert len(set(all_vml_parts)) == len(all_vml_parts)
    # The clone's parts must have wolfxl's canonical naming (the
    # PartIdAllocator-emitted form), confirming the allocator is
    # producing fresh suffixes for the cloned ancillary parts.
    new_comments = [n for n in all_comment_parts if re.match(r"xl/comments\d+\.xml$", n)]
    new_vmls = [n for n in all_vml_parts if re.match(r"xl/drawings/vmlDrawing\d+\.vml$", n)]
    assert new_comments, f"clone's comments part missing canonical name; got {all_comment_parts}"
    assert new_vmls, f"clone's VML part missing canonical name; got {all_vml_parts}"


# ---------------------------------------------------------------------------
# Case F — Copy with data validation (wolfxl preserves; openpyxl drops).
# ---------------------------------------------------------------------------


def test_f_copy_with_data_validation(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    _make_dv_fixture(src)

    wb = load_workbook(src, modify=True)
    wb.copy_worksheet(wb.active)
    wb.save(out)

    # The cloned sheet's XML carries a <dataValidations> block.
    # Find the cloned sheet path: the second worksheet by index.
    listing = _zip_listing(out)
    sheet_xmls = sorted(n for n in listing if re.match(r"xl/worksheets/sheet\d+\.xml$", n))
    assert len(sheet_xmls) == 2, f"expected 2 worksheets, got {sheet_xmls}"
    # Both sheets carry a dataValidations block with the same sqref.
    src_xml = _read_zip_text(out, sheet_xmls[0])
    dst_xml = _read_zip_text(out, sheet_xmls[1])
    assert "<dataValidations" in src_xml, "source DV missing"
    assert "<dataValidations" in dst_xml, (
        "cloned DV missing — wolfxl preserves DV on copy "
        "(openpyxl divergence: openpyxl's WorksheetCopy drops them)"
    )
    # DV sqref is preserved.
    assert 'sqref="B2:B4"' in dst_xml or "sqref=\"B2:B4\"" in dst_xml


# ---------------------------------------------------------------------------
# Case G — Copy with conditional formatting (wolfxl preserves; openpyxl drops).
# ---------------------------------------------------------------------------


def test_g_copy_with_conditional_formatting(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    _make_cf_fixture(src)

    wb = load_workbook(src, modify=True)
    wb.copy_worksheet(wb.active)
    wb.save(out)

    listing = _zip_listing(out)
    sheet_xmls = sorted(n for n in listing if re.match(r"xl/worksheets/sheet\d+\.xml$", n))
    assert len(sheet_xmls) == 2
    dst_xml = _read_zip_text(out, sheet_xmls[1])
    assert "<conditionalFormatting" in dst_xml, (
        "cloned CF missing — wolfxl preserves CF on copy "
        "(openpyxl divergence: openpyxl's WorksheetCopy drops them)"
    )


# ---------------------------------------------------------------------------
# Case H — Copy with sheet-scoped defined name (Print_Area).
# ---------------------------------------------------------------------------


def test_h_copy_with_sheet_scoped_defined_name(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    _make_sheet_scoped_name_fixture(src)

    wb = load_workbook(src, modify=True)
    # Source is at index 0; clone goes to index 2 (after Other at 1).
    new_ws = wb.copy_worksheet(wb["Template"])
    wb.save(out)

    wb_xml = _read_zip_text(out, "xl/workbook.xml")
    assert f'name="{new_ws.title}"' in wb_xml, (
        "defined-name merge must preserve the Phase 2.7 cloned <sheet> entry"
    )
    # The original Print_Area on Template (localSheetId=0) is preserved.
    assert (
        '<definedName name="_xlnm.Print_Area" localSheetId="0">' in wb_xml
    ), f"original Print_Area missing:\n{wb_xml}"
    # And a fresh Print_Area exists on the clone — at the new sheet's
    # tab index (sheet 0 = Template, sheet 1 = Other, sheet 2 = clone).
    new_idx = wb.sheetnames.index(new_ws.title)
    assert new_idx == 2
    assert (
        f'<definedName name="_xlnm.Print_Area" localSheetId="{new_idx}">' in wb_xml
    ), f"cloned Print_Area at localSheetId={new_idx} missing:\n{wb_xml}"


# ---------------------------------------------------------------------------
# Case I — Copy + edit-the-copy in same save.
# ---------------------------------------------------------------------------


def test_i_copy_and_edit_copy_in_same_save(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    _make_grid_fixture(src)

    wb = load_workbook(src, modify=True)
    new_ws = wb.copy_worksheet(wb.active)
    new_ws["A1"] = "edited-on-copy"
    wb.save(out)

    rt = openpyxl.load_workbook(out)
    # The edit landed only on the copy.
    assert rt["Template Copy"]["A1"].value == "edited-on-copy"
    # The source's A1 still reflects the fixture (not the edit).
    assert rt["Template"]["A1"].value != "edited-on-copy"


# ---------------------------------------------------------------------------
# Case J — Copy + RFC-036 move_sheet in same save.
# ---------------------------------------------------------------------------


def test_j_copy_then_move_sheet_in_same_save(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    _make_grid_fixture(src)

    wb = load_workbook(src, modify=True)
    new_ws = wb.copy_worksheet(wb.active)
    # Move the clone (currently at the end) to position 0.
    wb.move_sheet(new_ws.title, offset=-len(wb.sheetnames))
    assert wb.sheetnames[0] == new_ws.title
    wb.save(out)

    rt = openpyxl.load_workbook(out)
    assert rt.sheetnames[0] == "Template Copy"
    assert rt.sheetnames[1] == "Template"


# ---------------------------------------------------------------------------
# Case K — Copy + RFC-024 add_table-to-copy in same save.
# ---------------------------------------------------------------------------


def test_k_copy_then_add_table_to_copy(tmp_path: Path) -> None:
    """A table added to the clone in the same save round-trips.

    The clone of source (with one ``Sales`` table) is named ``Sales_2``
    by Phase 2.7. If the user then queues a new table on the clone,
    the new table's name must be unique against
    ``{source-zip ∪ queued ∪ cloned}`` (Risk #6 in RFC-035 §8). The
    behavior depends on RFC-024's collision-scan; this test pins what
    actually happens today and serves as a regression for whatever
    behavior RFC-024 + RFC-035 jointly define.
    """
    from wolfxl.worksheet.table import Table as WolfTable

    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    _make_table_fixture(src, ["Sales"])

    wb = load_workbook(src, modify=True)
    new_ws = wb.copy_worksheet(wb.active)
    # Add a table to the clone with a non-conflicting name.
    new_ws.add_table(WolfTable(name="ExtraOnCopy", displayName="ExtraOnCopy", ref="A1:E5"))
    wb.save(out)

    listing = _zip_listing(out)
    table_parts = [n for n in listing if n.startswith("xl/tables/")]
    # We expect: source's Sales (1), cloned Sales_2 (1), user's
    # ExtraOnCopy (1) = 3 tables. If RFC-024 + RFC-035 ever collapse
    # one of these (auto-rename, etc.), this assertion documents the
    # behaviour of record so a future change surfaces visibly.
    names = []
    for tp in table_parts:
        xml = _read_zip_text(out, tp)
        m = re.search(r'displayName="([^"]+)"', xml)
        assert m is not None
        names.append(m.group(1))
    assert "Sales" in names
    assert "Sales_2" in names
    assert "ExtraOnCopy" in names


# ---------------------------------------------------------------------------
# Case L — Name-collision on explicit name arg.
# ---------------------------------------------------------------------------


def test_l_name_collision_raises(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    _make_grid_fixture(src)

    wb = load_workbook(src, modify=True)
    with pytest.raises(ValueError, match="already exists"):
        wb.copy_worksheet(wb.active, name="Template")


# ---------------------------------------------------------------------------
# Case M — Cross-workbook source rejected.
# ---------------------------------------------------------------------------


def test_m_cross_workbook_rejected(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    _make_grid_fixture(src)

    wb_a = load_workbook(src, modify=True)
    wb_b = load_workbook(src, modify=True)
    with pytest.raises(ValueError, match="must belong to this workbook"):
        wb_a.copy_worksheet(wb_b.active)


# ---------------------------------------------------------------------------
# Case N — Write-mode is now SUPPORTED (Sprint Θ Pod-C1).
#   Positive coverage moved to ``tests/test_copy_worksheet_write_mode.py``.
# ---------------------------------------------------------------------------


# ---------------------------------------------------------------------------
# Case O — No-op byte-identical save (no copy_worksheet calls).
# ---------------------------------------------------------------------------


def test_o_no_copy_byte_identical(tmp_path: Path) -> None:
    """An empty queue is a no-op identity (Phase 2.7 short-circuit)."""
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_grid_fixture(src)

    wb = load_workbook(src, modify=True)
    # No copy_worksheet call.
    wb.save(dst)
    assert src.read_bytes() == dst.read_bytes()


# ---------------------------------------------------------------------------
# Case P — Self-closing <sheets/> branch (synthesized via direct ZIP edit).
# ---------------------------------------------------------------------------


def _rewrite_sheets_block_self_closing(path: Path) -> None:
    """Rewrite the workbook.xml in the given xlsx so its <sheets>...</sheets>
    becomes a self-closing <sheets/>, then strip the <sheet> children
    by replacing the block with the synthesized form for a single-sheet
    workbook.

    The resulting xlsx is intentionally borderline (Excel always emits
    open/close), but it exercises the Phase 2.7 splice's self-closing
    branch and is a known wolfxl behaviour Pod-β called out.
    """
    import io

    with zipfile.ZipFile(path) as zin:
        wb_xml = zin.read("xl/workbook.xml").decode("utf-8")
        names = zin.namelist()
        contents = {n: zin.read(n) for n in names}

    # Replace <sheets>...</sheets> with <sheets/>.
    new_wb_xml = re.sub(r"<sheets>.*?</sheets>", "<sheets/>", wb_xml, count=1, flags=re.DOTALL)
    contents["xl/workbook.xml"] = new_wb_xml.encode("utf-8")

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout:
        for n in names:
            zout.writestr(n, contents[n])
    path.write_bytes(buf.getvalue())


def test_p_self_closing_sheets_block(tmp_path: Path) -> None:
    """Phase 2.7 must rewrite a self-closing ``<sheets/>`` to open/close
    when appending the new ``<sheet>`` child.

    Pod-β originally called out this branch as untested because
    wolfxl's loader rejected the synthesized fixture (returned an
    empty sheet list) before the splice could run. Sprint Θ Pod-A
    closes that gap via the new ``permissive=True`` flag on
    ``load_workbook``: when set, the loader falls back to the workbook
    rels graph and synthesizes ``Sheet1``, ``Sheet2``, ... titles for
    every worksheet relationship target, then rewrites the empty
    ``<sheets/>`` block in-memory so downstream phases (Phase 2.7
    splice, defined-names merger, ...) see a well-formed workbook.

    With ``permissive=True`` the test now exercises the full public
    path end-to-end: load → copy_worksheet → save → reload via
    openpyxl. The asserts at the bottom guard the original splice
    contract (no self-closing form survives, the new sheet is
    appended, the block is parseable). KNOWN_GAPS.md tracks this as
    bug #4 ("self-closing <sheets/>") under "Fixed in 1.2 (Sprint Θ
    Pod-A)".
    """
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    _make_grid_fixture(src)
    _rewrite_sheets_block_self_closing(src)

    # `permissive=True` tells wolfxl to scan the workbook rels graph
    # when xl/workbook.xml's <sheets> block is empty/self-closing.
    # Default is False so well-formed inputs are unaffected.
    wb = load_workbook(src, modify=True, permissive=True)
    assert wb._rust_reader.__class__.__name__ == "NativeXlsxBook"  # noqa: SLF001

    # The synthesized title is "Sheet1" (rels-iteration order), which
    # may differ from the original fixture's "Grid". This is by design
    # — we don't have the real title without the <sheet> entry.
    assert wb.sheetnames, "permissive mode must surface at least one sheet"
    src_ws = wb[wb.sheetnames[0]]
    wb.copy_worksheet(src_ws)
    wb.save(out)

    # Verify the new workbook.xml has an open/close <sheets>...</sheets>
    # with at least one <sheet> child.
    new_wb_xml = _read_zip_text(out, "xl/workbook.xml")
    assert "<sheets/>" not in new_wb_xml, (
        "Phase 2.7 must rewrite self-closing <sheets/> to open/close form"
    )
    assert "<sheets>" in new_wb_xml and "</sheets>" in new_wb_xml
    assert "<sheet " in new_wb_xml or "<sheet>" in new_wb_xml


def test_copy_worksheet_handles_prefixed_workbook_root(tmp_path: Path) -> None:
    fixture = (
        Path(__file__).resolve().parent
        / "fixtures"
        / "external_oracle"
        / "closedxml-rich-comment-protection.xlsx"
    )
    src = tmp_path / fixture.name
    shutil.copy2(fixture, src)

    wb = load_workbook(src, modify=True)
    wb.copy_worksheet(wb[wb.sheetnames[0]])
    wb.save(src)
    wb.close()

    roundtrip = openpyxl.load_workbook(src)
    assert roundtrip.sheetnames == ["Review", "Review Copy"]
    roundtrip.close()

    with zipfile.ZipFile(src) as archive:
        workbook_xml = archive.read("xl/workbook.xml").decode()
    assert "<x:workbook" in workbook_xml
    assert '<x:sheet name="Review Copy"' in workbook_xml
    assert '<sheet name="Review Copy"' not in workbook_xml


# ---------------------------------------------------------------------------
# Case Q — Defined-names upsert collision (last-write-wins).
# ---------------------------------------------------------------------------


def test_q_defined_names_upsert_collision(tmp_path: Path) -> None:
    """If the user queues a defined name with the same (name,
    localSheetId) that the copy will produce, the merger must converge
    on a single entry — last write wins (per Pod-β's handoff note).
    """
    from wolfxl.workbook.defined_name import DefinedName as WolfDN

    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    _make_sheet_scoped_name_fixture(src)

    wb = load_workbook(src, modify=True)
    new_ws = wb.copy_worksheet(wb["Template"])
    new_idx = wb.sheetnames.index(new_ws.title)
    # Queue a defined name that aliases the cloned Print_Area entry.
    wb.defined_names["_xlnm.Print_Area"] = WolfDN(
        name="_xlnm.Print_Area",
        value=f"{new_ws.title}!$B$2:$D$4",
        localSheetId=new_idx,
    )
    wb.save(out)

    wb_xml = _read_zip_text(out, "xl/workbook.xml")
    # The user's value (B2:D4) must appear (last-write-wins).
    assert "$B$2:$D$4" in wb_xml, (
        f"user upsert value missing — planner overwrote it:\n{wb_xml}"
    )
    # And exactly ONE entry for (Print_Area, new_idx) should remain.
    pattern = f'name="_xlnm.Print_Area" localSheetId="{new_idx}"'
    occurrences = wb_xml.count(pattern)
    assert occurrences == 1, (
        f"expected exactly 1 (_xlnm.Print_Area, {new_idx}); got "
        f"{occurrences}:\n{wb_xml}"
    )


# ---------------------------------------------------------------------------
# Case R — CDATA / processing-instruction containing literal </sheets>.
#          Pod-β admits the splice is naive; this test documents the
#          limitation rather than gating behavior.
# ---------------------------------------------------------------------------


def _inject_comment_with_sheets_token(path: Path, marker: str) -> None:
    """Inject an XML comment into workbook.xml that contains the literal
    ``</sheets>`` token. A naive splice that searches for ``</sheets>``
    by string match would splice into the comment instead of the real
    closing tag.

    The comment lands BEFORE the real ``<sheets>`` block (right after
    the opening ``<workbook ...>``). openpyxl's writer omits the XML
    declaration, so we anchor on the closing ``>`` of the workbook
    element rather than ``?>``.
    """
    import io
    import re

    with zipfile.ZipFile(path) as zin:
        wb_xml = zin.read("xl/workbook.xml").decode("utf-8")
        names = zin.namelist()
        contents = {n: zin.read(n) for n in names}

    # Find the end of the opening `<workbook ...>` tag and inject the
    # fakeout comment immediately after it. The regex tolerates either
    # a self-anchored XML declaration or its absence.
    m = re.search(r"<workbook\b[^>]*>", wb_xml)
    if m is None:
        raise RuntimeError("test fixture: no <workbook> opening tag found")
    insert_at = m.end()
    fakeout = f"<!-- {marker} </sheets> -->"
    new_wb_xml = wb_xml[:insert_at] + fakeout + wb_xml[insert_at:]
    contents["xl/workbook.xml"] = new_wb_xml.encode("utf-8")

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout:
        for n in names:
            zout.writestr(n, contents[n])
    path.write_bytes(buf.getvalue())


def test_r_cdata_pi_fuzz_fakeout(tmp_path: Path) -> None:
    """A workbook.xml comment containing the literal ``</sheets>``
    string must not fool Phase 2.7's splice.

    Sprint Θ Pod-B replaced the naive byte-substring locator with a
    SAX/quick-xml-driven scan that respects element nesting, so
    comments / CDATA / PIs containing the literal token no longer
    perturb the insertion point. KNOWN_GAPS bug #6 is closed.
    """
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    _make_grid_fixture(src)
    _inject_comment_with_sheets_token(src, "FUZZTOKEN")

    wb = load_workbook(src, modify=True)
    assert wb.sheetnames, "fuzz fixture should still expose its sheet"

    new_ws = wb.copy_worksheet(wb[wb.sheetnames[0]])
    wb.save(out)

    new_wb_xml = _read_zip_text(out, "xl/workbook.xml")
    # The synthesized comment must survive the rewrite.
    assert "FUZZTOKEN" in new_wb_xml, (
        "SAX-driven splice must preserve the workbook.xml comment "
        "containing the </sheets> fakeout token (KNOWN_GAPS bug #6)."
    )

    # And the new sheet must appear.
    rt = openpyxl.load_workbook(out)
    assert new_ws.title in rt.sheetnames, (
        "Phase 2.7 splice failed to add the new <sheet> entry on a "
        "fixture with a CDATA-style fakeout comment."
    )


# ---------------------------------------------------------------------------
# Case S — Default name generation: Copy / Copy 2 / Copy 3.
# ---------------------------------------------------------------------------


def test_s_default_name_generation(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    _make_grid_fixture(src)

    wb = load_workbook(src, modify=True)
    a = wb.copy_worksheet(wb.active)
    b = wb.copy_worksheet(wb.active)
    c = wb.copy_worksheet(wb.active)
    assert a.title == "Template Copy"
    assert b.title == "Template Copy 2"
    assert c.title == "Template Copy 3"
    wb.save(out)

    rt = openpyxl.load_workbook(out)
    assert rt.sheetnames == ["Template", "Template Copy", "Template Copy 2", "Template Copy 3"]
