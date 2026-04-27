"""Sprint Ν Pod-γ (RFC-035 §10) — pivot table deep-clone on
``Workbook.copy_worksheet``.

Pre-Sprint Ν behavior: pivot tables on a copied sheet were aliased
— the cloned sheet's pivot rel pointed at the SOURCE pivot table
XML part, so any post-copy edit to the source's table layout would
bleed through to the copy.

Post-Sprint Ν behavior: ``Workbook.copy_worksheet(src)`` deep-clones
every pivot table referenced by the source:

* fresh ``pivotTableN`` part is allocated;
* table XML is bytes-rewritten to update any
  ``<worksheetSource sheet="<src>"/>`` (no-op for the common case
  where the cache's source range is on a different sheet);
* per-table rels file is copied verbatim (table → cache target
  preserved — the cache itself stays workbook-shared);
* sheet rel + content-type override are added for the new part.

The cache is intentionally NOT deep-cloned (workbook-scope per
RFC-047 §6).
"""
from __future__ import annotations

import re
import zipfile
from pathlib import Path

import openpyxl
import pytest

from wolfxl import load_workbook
from wolfxl.chart.reference import Reference
from wolfxl.pivot import PivotCache, PivotTable


@pytest.fixture(autouse=True)
def _force_test_epoch(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("WOLFXL_TEST_EPOCH", "0")


# ---------------------------------------------------------------------------
# Fixture: workbook with a worksheet that already carries a pivot.
# ---------------------------------------------------------------------------


def _make_pivot_workbook_fixture(path: Path) -> None:
    """4 cols × 5 rows source with one pivot table on the same sheet."""
    src_seed = path.parent / "_seed.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    rows = [
        ("region", "quarter", "customer", "revenue"),
        ("North", "Q1", "Acme", 100.0),
        ("South", "Q1", "Acme", 200.0),
        ("North", "Q2", "Globex", 150.0),
        ("South", "Q2", "Globex", 250.0),
    ]
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(r_idx, c_idx, val)
    wb.save(src_seed)

    # Re-open in modify mode and add a pivot.
    wbm = load_workbook(src_seed, modify=True)
    wsm = wbm["Data"]
    refsrc = Reference(
        worksheet=wsm, min_col=1, min_row=1, max_col=4, max_row=5
    )
    cache = PivotCache(source=refsrc)
    pt = PivotTable(
        cache=cache,
        location="F2",
        rows=["region"],
        cols=["quarter"],
        data=["revenue"],
    )
    wbm.add_pivot_cache(cache)
    wsm.add_pivot_table(pt)
    wbm.save(path)
    src_seed.unlink(missing_ok=True)


def _zip_listing(path: Path) -> list[str]:
    with zipfile.ZipFile(path, "r") as z:
        return sorted(z.namelist())


def _zip_read(path: Path, member: str) -> bytes:
    with zipfile.ZipFile(path, "r") as z:
        return z.read(member)


# ---------------------------------------------------------------------------
# Case 1 — copy_worksheet allocates a fresh pivotTableN part.
# ---------------------------------------------------------------------------


def test_copy_worksheet_with_pivot_allocates_new_table_part(
    tmp_path: Path,
) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_pivot_workbook_fixture(src)

    wb = load_workbook(src, modify=True)
    wb.copy_worksheet(wb["Data"], name="DataCopy")
    wb.save(dst)

    entries = _zip_listing(dst)
    table_parts = [
        e for e in entries
        if re.match(r"^xl/pivotTables/pivotTable\d+\.xml$", e)
    ]
    assert len(table_parts) == 2, (
        f"expected 2 pivot tables (orig + clone), got {table_parts}"
    )


# ---------------------------------------------------------------------------
# Case 2 — Cache stays SHARED (workbook-scope, RFC-047 §6).
# ---------------------------------------------------------------------------


def test_copy_worksheet_with_pivot_keeps_cache_shared(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_pivot_workbook_fixture(src)

    wb = load_workbook(src, modify=True)
    wb.copy_worksheet(wb["Data"], name="DataCopy")
    wb.save(dst)

    entries = _zip_listing(dst)
    cache_defs = [
        e for e in entries
        if re.match(r"^xl/pivotCache/pivotCacheDefinition\d+\.xml$", e)
    ]
    cache_recs = [
        e for e in entries
        if re.match(r"^xl/pivotCache/pivotCacheRecords\d+\.xml$", e)
    ]
    assert len(cache_defs) == 1, (
        f"cache must stay workbook-shared; got {cache_defs}"
    )
    assert len(cache_recs) == 1, (
        f"records must stay workbook-shared; got {cache_recs}"
    )


# ---------------------------------------------------------------------------
# Case 3 — Cloned sheet has its own pivot rel pointing at the new
# table part (not the original).
# ---------------------------------------------------------------------------


def test_copy_worksheet_with_pivot_emits_distinct_sheet_rel(
    tmp_path: Path,
) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_pivot_workbook_fixture(src)

    wb = load_workbook(src, modify=True)
    wb.copy_worksheet(wb["Data"], name="DataCopy")
    wb.save(dst)

    entries = _zip_listing(dst)
    sheet_rels_files = [
        e for e in entries
        if re.match(r"^xl/worksheets/_rels/sheet\d+\.xml\.rels$", e)
    ]
    # Both source and clone sheets must have a rels file referencing
    # a pivot table part.
    pivot_targets: set[str] = set()
    for srf in sheet_rels_files:
        body = _zip_read(dst, srf).decode("utf-8")
        for m in re.finditer(
            r'Target="\.\./pivotTables/(pivotTable\d+\.xml)"', body
        ):
            pivot_targets.add(m.group(1))
    assert len(pivot_targets) >= 2, (
        f"expected source + clone to point at distinct pivot parts; "
        f"got {pivot_targets} from {sheet_rels_files}"
    )


# ---------------------------------------------------------------------------
# Case 4 — Cloned table part has its own per-table rels pointing at
# the (shared) cache definition.
# ---------------------------------------------------------------------------


def test_copy_worksheet_with_pivot_emits_per_table_rels(
    tmp_path: Path,
) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_pivot_workbook_fixture(src)

    wb = load_workbook(src, modify=True)
    wb.copy_worksheet(wb["Data"], name="DataCopy")
    wb.save(dst)

    entries = _zip_listing(dst)
    table_rels = [
        e for e in entries
        if e.startswith("xl/pivotTables/_rels/pivotTable")
        and e.endswith(".xml.rels")
    ]
    assert len(table_rels) == 2, (
        f"each table needs its own rels file; got {table_rels}"
    )
    # Both rels files must point at a pivotCacheDefinition target.
    for trf in table_rels:
        body = _zip_read(dst, trf).decode("utf-8")
        assert "pivotCacheDefinition" in body, (
            f"per-table rels must point at cache; {trf} = {body!r}"
        )


# ---------------------------------------------------------------------------
# Case 5 — Content-types has overrides for both the original and the
# cloned pivotTableN parts.
# ---------------------------------------------------------------------------


def test_copy_worksheet_with_pivot_content_types_has_both_overrides(
    tmp_path: Path,
) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_pivot_workbook_fixture(src)

    wb = load_workbook(src, modify=True)
    wb.copy_worksheet(wb["Data"], name="DataCopy")
    wb.save(dst)

    ct = _zip_read(dst, "[Content_Types].xml").decode("utf-8")
    pt_overrides = re.findall(
        r'PartName="/xl/pivotTables/pivotTable(\d+)\.xml"', ct
    )
    assert len(pt_overrides) == 2, (
        f"expected 2 pivotTable overrides, got {pt_overrides}"
    )


# ---------------------------------------------------------------------------
# Case 6 — Idempotency: re-saving the deep-cloned workbook is stable
# (both pivot tables survive, no extra parts spawn).
# ---------------------------------------------------------------------------


def test_resave_after_pivot_deep_clone_is_stable(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    mid = tmp_path / "mid.xlsx"
    out = tmp_path / "out.xlsx"
    _make_pivot_workbook_fixture(src)

    wb = load_workbook(src, modify=True)
    wb.copy_worksheet(wb["Data"], name="DataCopy")
    wb.save(mid)

    wb2 = load_workbook(mid, modify=True)
    wb2.save(out)

    entries_mid = [
        e for e in _zip_listing(mid)
        if e.startswith("xl/pivotTables/pivotTable")
    ]
    entries_out = [
        e for e in _zip_listing(out)
        if e.startswith("xl/pivotTables/pivotTable")
    ]
    assert entries_out == entries_mid, (
        f"re-save changed pivot part set: mid={entries_mid} "
        f"out={entries_out}"
    )
