"""Tests for the wolfxl openpyxl-compatible API."""

from __future__ import annotations

import re
import zipfile
from datetime import datetime
from pathlib import Path

import pytest


def _require_rust() -> None:
    pytest.importorskip("wolfxl._rust")


# ======================================================================
# Pure-Python unit tests (no Rust needed)
# ======================================================================


class TestUtils:
    """Coordinate conversion helpers."""

    def test_column_letter(self) -> None:
        from wolfxl._utils import column_letter

        assert column_letter(1) == "A"
        assert column_letter(26) == "Z"
        assert column_letter(27) == "AA"
        assert column_letter(702) == "ZZ"

    def test_column_index(self) -> None:
        from wolfxl._utils import column_index

        assert column_index("A") == 1
        assert column_index("Z") == 26
        assert column_index("AA") == 27
        assert column_index("ZZ") == 702

    def test_a1_roundtrip(self) -> None:
        from wolfxl._utils import a1_to_rowcol, rowcol_to_a1

        assert a1_to_rowcol("B3") == (3, 2)
        assert rowcol_to_a1(3, 2) == "B3"
        assert a1_to_rowcol("AA100") == (100, 27)
        assert rowcol_to_a1(100, 27) == "AA100"

    def test_invalid_a1_raises(self) -> None:
        from wolfxl._utils import a1_to_rowcol

        with pytest.raises(ValueError, match="Invalid A1 reference"):
            a1_to_rowcol("123")


class TestStyles:
    """Frozen style dataclasses."""

    def test_font_defaults(self) -> None:
        from wolfxl._styles import Font

        f = Font()
        assert f.bold is False
        assert f.name is None
        assert f.size is None

    def test_font_is_frozen(self) -> None:
        from wolfxl._styles import Font

        f = Font(bold=True)
        with pytest.raises(AttributeError):
            f.bold = False  # type: ignore[misc]

    def test_color_hex_conversion(self) -> None:
        from wolfxl._styles import Color

        c = Color(rgb="FFFF0000")
        assert c.to_hex() == "#FF0000"
        assert Color.from_hex("#00FF00").rgb == "FF00FF00"

    def test_pattern_fill(self) -> None:
        from wolfxl._styles import PatternFill

        fill = PatternFill(patternType="solid", fgColor="#FF0000")
        assert fill._fg_hex() == "#FF0000"

    def test_border_defaults(self) -> None:
        from wolfxl._styles import Border, Side

        b = Border()
        assert b.left == Side()
        assert b.top.style is None

    def test_alignment_defaults(self) -> None:
        from wolfxl._styles import Alignment

        a = Alignment()
        assert a.horizontal is None
        assert a.wrap_text is False
        assert a.indent == 0


# ======================================================================
# Read tests (require wolfxl._rust + fixtures)
# ======================================================================

REPO_ROOT = Path(__file__).resolve().parents[1]
FIXTURES = REPO_ROOT / "tests" / "fixtures"
PARITY_FIXTURES = REPO_ROOT / "tests" / "parity" / "fixtures" / "synthgl_snapshot"


class TestReadMode:
    """Read an existing Excel fixture via CalamineStyledBook."""

    def setup_method(self) -> None:
        _require_rust()

    def test_load_workbook_basic(self) -> None:
        from wolfxl import load_workbook

        path = FIXTURES / "tier1" / "01_cell_values.xlsx"
        if not path.exists():
            pytest.skip("fixture not found")
        wb = load_workbook(str(path))
        assert "Sheet1" in wb.sheetnames or len(wb.sheetnames) > 0
        ws = wb[wb.sheetnames[0]]
        # Column B has test values, A has labels. Row 2 = "Hello World" per manifest.
        val = ws["B2"].value
        assert val == "Hello World"
        wb.close()

    def test_read_number(self) -> None:
        from wolfxl import load_workbook

        path = FIXTURES / "tier1" / "01_cell_values.xlsx"
        if not path.exists():
            pytest.skip("fixture not found")
        wb = load_workbook(str(path))
        ws = wb[wb.sheetnames[0]]
        # Row 7 col B = integer 42
        val = ws["B7"].value
        assert val == 42 or val == 42.0
        wb.close()

    def test_read_font_bold(self) -> None:
        from wolfxl import load_workbook

        path = FIXTURES / "tier1" / "03_text_formatting.xlsx"
        if not path.exists():
            pytest.skip("fixture not found")
        wb = load_workbook(str(path))
        ws = wb[wb.sheetnames[0]]
        # Row 2, col B = bold text per manifest
        cell = ws["B2"]
        assert cell.font.bold is True
        wb.close()

    def test_read_background_color(self) -> None:
        from wolfxl import load_workbook

        path = FIXTURES / "tier1" / "04_background_colors.xlsx"
        if not path.exists():
            pytest.skip("fixture not found")
        wb = load_workbook(str(path))
        ws = wb[wb.sheetnames[0]]
        # Row 2, col B should have a background color
        fill = ws["B2"].fill
        # Just verify it parsed without error — exact color varies by fixture
        assert fill is not None
        wb.close()

    def test_context_manager(self) -> None:
        from wolfxl import load_workbook

        path = FIXTURES / "tier1" / "01_cell_values.xlsx"
        if not path.exists():
            pytest.skip("fixture not found")
        with load_workbook(str(path)) as wb:
            assert len(wb.sheetnames) > 0

    def test_iter_rows_read_mode(self) -> None:
        from wolfxl import load_workbook

        path = FIXTURES / "tier1" / "01_cell_values.xlsx"
        if not path.exists():
            pytest.skip("fixture not found")
        wb = load_workbook(str(path))
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(min_row=1, max_row=3, values_only=True))
        assert len(rows) == 3
        assert rows[1][1] == "Hello World"  # B2
        wb.close()

    def test_iter_rows_auto_dimensions(self) -> None:
        from wolfxl import load_workbook

        path = FIXTURES / "tier1" / "01_cell_values.xlsx"
        if not path.exists():
            pytest.skip("fixture not found")
        wb = load_workbook(str(path))
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) > 10  # fixture has ~20 rows
        wb.close()

    def test_cell_records_exposes_values_and_compact_format_metadata(self, tmp_path: Path) -> None:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, Side

        from wolfxl import load_workbook

        path = tmp_path / "records.xlsx"
        op_wb = Workbook()
        ws = op_wb.active
        ws.title = "Records"
        ws["A1"] = "Header"
        ws["A1"].font = Font(bold=True, italic=True, size=12)
        ws["A2"] = "Child"
        ws["A2"].alignment = Alignment(indent=2, horizontal="left")
        ws["B2"] = 1500
        ws["B2"].number_format = "$#,##0.00_);($#,##0.00)"
        ws["B3"] = "Total"
        ws["B3"].border = Border(bottom=Side(style="double"))
        ws["C4"] = "=SUM(B2:B2)"
        op_wb.save(path)
        op_wb.close()

        with load_workbook(str(path), data_only=False) as wb:
            records = {record["coordinate"]: record for record in wb["Records"].cell_records()}

        assert records["A1"]["value"] == "Header"
        assert records["A1"]["bold"] is True
        assert records["A1"]["italic"] is True
        assert records["A1"]["font_size"] == 12.0
        assert records["A2"]["indent"] == 2
        assert records["A2"]["h_align"] == "left"
        assert records["B2"]["value"] == 1500
        assert records["B2"]["number_format"] == "$#,##0.00_);($#,##0.00)"
        assert records["B3"]["has_bottom_border"] is True
        assert records["B3"]["is_double_underline"] is True
        assert records["C4"]["data_type"] == "formula"
        assert records["C4"]["formula"] == "=SUM(B2:B2)"

    def test_cell_records_can_emit_dense_empty_range(self, tmp_path: Path) -> None:
        from openpyxl import Workbook

        from wolfxl import load_workbook

        path = tmp_path / "dense-records.xlsx"
        op_wb = Workbook()
        ws = op_wb.active
        ws.title = "Dense"
        ws["B2"] = "value"
        op_wb.save(path)
        op_wb.close()

        with load_workbook(str(path)) as wb:
            records = wb["Dense"].cell_records(
                min_row=1,
                max_row=2,
                min_col=1,
                max_col=2,
                include_format=False,
                include_empty=True,
            )

        assert [record["coordinate"] for record in records] == ["A1", "B1", "A2", "B2"]
        assert records[0]["data_type"] == "blank"
        assert records[-1]["value"] == "value"

    def test_cell_records_can_skip_coordinate_strings(self, tmp_path: Path) -> None:
        from openpyxl import Workbook

        from wolfxl import load_workbook

        path = tmp_path / "no-coordinate-records.xlsx"
        op_wb = Workbook()
        ws = op_wb.active
        ws.title = "Records"
        ws["B2"] = "value"
        op_wb.save(path)
        op_wb.close()

        with load_workbook(str(path)) as wb:
            records = wb["Records"].cell_records(include_coordinate=False)

        assert records == [{"row": 2, "column": 2, "data_type": "string", "value": "value"}]

    def test_cell_records_data_only_skips_uncached_formulas_by_default(
        self,
        tmp_path: Path,
    ) -> None:
        from openpyxl import Workbook

        from wolfxl import load_workbook

        path = tmp_path / "uncached-formula.xlsx"
        op_wb = Workbook()
        ws = op_wb.active
        ws.title = "Formula"
        ws["A1"] = 10
        ws["B2"] = "=SUM(A1:A1)"
        op_wb.save(path)
        op_wb.close()

        with load_workbook(str(path), data_only=False) as wb:
            records = {record["coordinate"]: record for record in wb["Formula"].cell_records()}
        assert records["B2"]["data_type"] == "formula"
        assert records["B2"]["formula"] == "=SUM(A1:A1)"

        with load_workbook(str(path), data_only=True) as wb:
            records = {record["coordinate"]: record for record in wb["Formula"].cell_records()}
            dense = wb["Formula"].cell_records(
                min_row=2,
                max_row=2,
                min_col=2,
                max_col=2,
                include_empty=True,
            )

        assert "B2" not in records
        assert dense[0]["coordinate"] == "B2"
        assert dense[0]["data_type"] == "blank"
        assert dense[0]["value"] is None
        assert dense[0]["formula"] == "=SUM(A1:A1)"

    def test_cell_records_include_formula_blanks_false_skips_uncached_inside_range(
        self,
        tmp_path: Path,
    ) -> None:
        # When a template formula sits inside the populated rectangular range,
        # calamine returns Some(Data::Empty) rather than None for that cell.
        # `include_formula_blanks=False` must still suppress the formula record;
        # treating Some(Data::Empty) as a backing entry was the regression.
        from openpyxl import Workbook

        from wolfxl import load_workbook

        path = tmp_path / "uncached-formula-in-range.xlsx"
        op_wb = Workbook()
        ws = op_wb.active
        ws.title = "Formula"
        ws["A1"] = 10
        ws["B2"] = "=SUM(A1:A1)"
        op_wb.save(path)
        op_wb.close()

        with load_workbook(str(path), data_only=False) as wb:
            records = wb["Formula"].cell_records(include_formula_blanks=False)

        coords = [r["coordinate"] for r in records]
        assert "B2" not in coords, (
            f"uncached formula B2 should be suppressed when "
            f"include_formula_blanks=False, got coords={coords}"
        )

    def test_max_row_max_column_reflect_pending_writes_in_write_mode(
        self,
    ) -> None:
        # Pure write mode: ``ws.append()`` parks rows in ``_append_buffer``
        # without materializing Cell objects. ``max_row``/``max_column`` must
        # reflect the buffer extents — otherwise downstream range derivations
        # (e.g. iter_rows bounds) miss the appended data before save.
        from wolfxl import Workbook

        wb = Workbook()
        ws = wb.active
        assert ws.max_row == 1
        assert ws.max_column == 1

        ws.append([1, 2, 3])
        ws.append([4, 5, 6, 7])
        assert ws.max_row == 2, f"expected 2 after two appends, got {ws.max_row}"
        assert ws.max_column == 4, (
            f"expected 4 (widest appended row), got {ws.max_column}"
        )

    def test_cell_records_overlays_modify_mode_pending_edits(
        self,
        tmp_path: Path,
    ) -> None:
        # In modify mode the Rust reader serves on-disk values while pending
        # edits live in Python-side ``_dirty``/``_append_buffer``. The
        # iterator must overlay the edits — otherwise callers see stale
        # values for cells they just modified.
        from openpyxl import Workbook as OpWorkbook

        from wolfxl import load_workbook

        path = tmp_path / "modify-overlay.xlsx"
        op_wb = OpWorkbook()
        ws = op_wb.active
        ws.title = "Overlay"
        ws["A1"] = "old-a1"
        ws["B2"] = "old-b2"
        op_wb.save(path)
        op_wb.close()

        with load_workbook(str(path), modify=True) as wb:
            ws = wb["Overlay"]
            ws["A1"] = "new-a1"  # overwrite existing on-disk cell
            ws["E10"] = "added"  # write outside on-disk range

            records = ws.cell_records()
            by_coord = {r["coordinate"]: r["value"] for r in records}

        assert by_coord.get("A1") == "new-a1", (
            f"expected pending edit on A1 to overlay on-disk value, got {by_coord}"
        )
        assert by_coord.get("B2") == "old-b2", (
            f"expected unmodified B2 to keep on-disk value, got {by_coord}"
        )
        assert by_coord.get("E10") == "added", (
            f"expected new edit at E10 to appear in records, got {by_coord}"
        )

    def test_read_cell_value_data_only_blanks_uncached_formula(
        self,
        tmp_path: Path,
    ) -> None:
        # `data_only=True` per-cell access should match openpyxl: an uncached
        # formula reads as None, not as the placeholder empty string calamine
        # parks in `Some(Data::String(""))`. `read_sheet_records` already does
        # this; the per-cell and bulk readers must mirror it or template-heavy
        # sheets corrupt downstream ingestion.
        from openpyxl import Workbook as OpWorkbook

        from wolfxl import load_workbook

        path = tmp_path / "uncached-formula-cell.xlsx"
        op_wb = OpWorkbook()
        ws = op_wb.active
        ws.title = "Formula"
        ws["A1"] = 10
        ws["B2"] = "=SUM(A1:A1)"
        op_wb.save(path)
        op_wb.close()

        with load_workbook(str(path), data_only=True) as wb:
            sheet = wb["Formula"]
            assert sheet["B2"].value is None, (
                f"expected None for uncached formula in data_only mode, got {sheet['B2'].value!r}"
            )

    def test_iter_rows_values_only_blanks_uncached_formulas(
        self,
        tmp_path: Path,
    ) -> None:
        # Same uncached-formula normalization for the bulk path
        # (`read_sheet_values` / `read_sheet_values_plain`).
        from openpyxl import Workbook as OpWorkbook

        from wolfxl import load_workbook

        path = tmp_path / "uncached-formula-bulk.xlsx"
        op_wb = OpWorkbook()
        ws = op_wb.active
        ws.title = "Formula"
        ws["A1"] = 10
        ws["B2"] = "=SUM(A1:A1)"
        op_wb.save(path)
        op_wb.close()

        with load_workbook(str(path), data_only=True) as wb:
            rows = list(
                wb["Formula"].iter_rows(
                    min_row=1, max_row=2, min_col=1, max_col=2, values_only=True,
                )
            )

        assert rows[1][1] is None, (
            f"expected None at B2 in values_only bulk read, got {rows[1][1]!r}"
        )

    def test_max_row_does_not_inflate_from_read_only_cell_access(
        self,
        tmp_path: Path,
    ) -> None:
        # `_cells` is a read-cache: ws['Z999'].value populates it without
        # ever marking a cell dirty. The pending-bounds helper must iterate
        # `_dirty` (actually-modified cells), not `_cells` — otherwise read
        # access at far coordinates inflates max_row/max_column and
        # downstream range derivations scan empty space.
        from openpyxl import Workbook as OpWorkbook

        from wolfxl import load_workbook

        path = tmp_path / "no-inflate.xlsx"
        op_wb = OpWorkbook()
        ws = op_wb.active
        ws.title = "Tight"
        ws["A1"] = "only-cell"
        op_wb.save(path)
        op_wb.close()

        with load_workbook(str(path), modify=True) as wb:
            ws = wb["Tight"]
            assert ws.max_row == 1
            assert ws.max_column == 1

            # Mere read access at a far coordinate must not be treated as
            # a pending write.
            _ = ws["Z999"].value
            assert ws.max_row == 1, (
                f"reading Z999 inflated max_row to {ws.max_row}"
            )
            assert ws.max_column == 1, (
                f"reading Z999 inflated max_column to {ws.max_column}"
            )
            assert ws.calculate_dimension() == "A1:A1", (
                f"reading Z999 inflated dimension to {ws.calculate_dimension()}"
            )

            # Now actually write — bounds should grow.
            ws["Z999"] = "now-dirty"
            assert ws.max_row == 999
            assert ws.max_column == 26

    def test_cell_records_overlay_labels_pending_formulas_as_formula(
        self,
        tmp_path: Path,
    ) -> None:
        # Strings starting with '=' are formulas in openpyxl's convention
        # (and match Rust's formula_map_cache path). Overlay records must
        # label them data_type='formula' or any consumer counting/filtering
        # formula records misses pending edits.
        from openpyxl import Workbook as OpWorkbook

        from wolfxl import load_workbook

        path = tmp_path / "overlay-formula.xlsx"
        op_wb = OpWorkbook()
        ws = op_wb.active
        ws.title = "Form"
        ws["A1"] = 10
        op_wb.save(path)
        op_wb.close()

        with load_workbook(str(path), modify=True) as wb:
            ws = wb["Form"]
            ws["B2"] = "=SUM(A1:A1)"  # extra-overlay formula
            by_coord = {
                r["coordinate"]: r["data_type"]
                for r in ws.cell_records()
            }

        assert by_coord.get("B2") == "formula", (
            f"expected pending formula at B2 to label 'formula', got {by_coord.get('B2')!r}"
        )

    def test_cell_records_overlay_clears_stale_formula_metadata_on_literal_swap(
        self,
        tmp_path: Path,
    ) -> None:
        # Replacing a formula cell with a literal must drop the on-disk
        # `formula` field from the overlay record. Leaving it stale lets
        # downstream consumers see `data_type='number'` next to a leftover
        # `formula='SUM(...)'`, which mis-classifies a now-literal cell as
        # a formula cell.
        from openpyxl import Workbook as OpWorkbook

        from wolfxl import load_workbook

        path = tmp_path / "overlay-formula-cleared.xlsx"
        op_wb = OpWorkbook()
        ws = op_wb.active
        ws.title = "F"
        ws["A1"] = 1
        ws["A2"] = 2
        ws["B1"] = "=SUM(A1:A2)"  # on-disk formula cell
        op_wb.save(path)
        op_wb.close()

        with load_workbook(str(path), modify=True) as wb:
            ws = wb["F"]
            ws["B1"] = 99  # literal overwrites the formula
            by_coord = {r["coordinate"]: r for r in ws.cell_records()}

        b1 = by_coord["B1"]
        assert b1["value"] == 99
        assert b1["data_type"] == "number"
        assert "formula" not in b1, (
            f"stale formula leaked into literal-overwrite record: {b1!r}"
        )

    def test_cell_records_overlay_replaces_formula_metadata_on_formula_swap(
        self,
        tmp_path: Path,
    ) -> None:
        # Replacing one formula with a different formula (or a literal with
        # a formula) must update the `formula` field too — not just `value`
        # / `data_type` — so consumers don't see contradictory metadata.
        # Stripped of leading "=" to match the Rust reader's convention.
        from openpyxl import Workbook as OpWorkbook

        from wolfxl import load_workbook

        path = tmp_path / "overlay-formula-replaced.xlsx"
        op_wb = OpWorkbook()
        ws = op_wb.active
        ws.title = "F"
        ws["A1"] = 1
        ws["B1"] = "=A1+1"  # original formula
        op_wb.save(path)
        op_wb.close()

        with load_workbook(str(path), modify=True) as wb:
            ws = wb["F"]
            ws["B1"] = "=A1*10"  # different formula
            by_coord = {r["coordinate"]: r for r in ws.cell_records()}

        b1 = by_coord["B1"]
        assert b1["data_type"] == "formula"
        assert b1.get("formula") == "A1*10", (
            f"formula field not refreshed: {b1!r}"
        )

    def test_iter_cell_records_write_mode_honors_include_coordinate(self) -> None:
        # Write mode (no Rust reader) goes through the Python fallback.
        # Earlier the fallback always emitted `coordinate`, ignoring
        # `include_coordinate=False` and forcing per-cell A1 string
        # allocation that the API explicitly promises to skip.
        from wolfxl import Workbook as WolfxlWorkbook

        # Fresh in-memory workbook → _rust_reader is None → fallback path.
        wb = WolfxlWorkbook()
        ws = wb.active
        ws["A1"] = 1
        ws["B2"] = 2

        records_with = ws.cell_records(include_coordinate=True)
        records_without = ws.cell_records(include_coordinate=False)

        wb.close()

        assert all("coordinate" in r for r in records_with), records_with
        assert all("coordinate" not in r for r in records_without), records_without
        # Sanity: the rest of the schema is unchanged.
        assert {r["value"] for r in records_without} == {1, 2}

    def test_cell_records_overlay_uses_canonical_data_type_labels(
        self,
        tmp_path: Path,
    ) -> None:
        # Overlay records must use Rust's canonical labels (`string`,
        # `number`, `boolean`) — not Python type names (`str`, `int`,
        # `bool`). Mixed schemas in one `cell_records()` result break
        # consumers that filter/group by the documented tokens.
        from openpyxl import Workbook as OpWorkbook

        from wolfxl import load_workbook

        path = tmp_path / "overlay-types.xlsx"
        op_wb = OpWorkbook()
        ws = op_wb.active
        ws.title = "Types"
        ws["A1"] = "old"
        op_wb.save(path)
        op_wb.close()

        with load_workbook(str(path), modify=True) as wb:
            ws = wb["Types"]
            ws["A1"] = "patched"  # patches a Rust-returned record
            ws["B1"] = 42         # int → "number"
            ws["C1"] = 3.14       # float → "number"
            ws["D1"] = True       # bool → "boolean" (NOT "number")
            ws["E1"] = "text"     # str → "string"

            by_coord = {r["coordinate"]: r["data_type"] for r in ws.cell_records()}

        # Patched-overlay path
        assert by_coord["A1"] == "string", f"A1 should be 'string', got {by_coord['A1']!r}"
        # Extra-overlay path (cells outside on-disk range)
        assert by_coord["B1"] == "number", f"B1 (int) should be 'number', got {by_coord['B1']!r}"
        assert by_coord["C1"] == "number", f"C1 (float) should be 'number', got {by_coord['C1']!r}"
        assert by_coord["D1"] == "boolean", f"D1 (bool) should be 'boolean', got {by_coord['D1']!r}"
        assert by_coord["E1"] == "string", f"E1 (str) should be 'string', got {by_coord['E1']!r}"

    def test_calculate_dimension_includes_modify_mode_pending_edits(
        self,
        tmp_path: Path,
    ) -> None:
        # Modify mode has both a Rust reader (on-disk extents) AND Python-side
        # pending writes. ``calculate_dimension()`` must union them, otherwise
        # callers that derive ranges from it omit unsaved cells.
        from openpyxl import Workbook as OpWorkbook

        from wolfxl import load_workbook

        path = tmp_path / "modify-pending.xlsx"
        op_wb = OpWorkbook()
        ws = op_wb.active
        ws.title = "Pending"
        ws["A1"] = "head"
        ws["B2"] = "tail"
        op_wb.save(path)
        op_wb.close()

        with load_workbook(str(path), modify=True) as wb:
            ws = wb["Pending"]
            on_disk = ws.calculate_dimension()
            assert on_disk == "A1:B2", on_disk
            # Touch a cell well outside the on-disk bounds.
            ws["E10"] = "added"
            after = ws.calculate_dimension()
            assert after == "A1:E10", (
                f"expected union of on-disk bounds and pending edit at E10, "
                f"got {after}"
            )

    def test_calculate_dimension_uses_cell_storage_when_dimension_tag_is_stale(
        self,
        tmp_path: Path,
    ) -> None:
        from openpyxl import Workbook

        from wolfxl import load_workbook

        source = tmp_path / "source-dimension.xlsx"
        stale = tmp_path / "stale-dimension.xlsx"
        op_wb = Workbook()
        ws = op_wb.active
        ws.title = "Stale"
        ws["A1"] = "head"
        ws["C7"] = "tail"
        op_wb.save(source)
        op_wb.close()

        with zipfile.ZipFile(source, "r") as zin, zipfile.ZipFile(stale, "w") as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == "xl/worksheets/sheet1.xml":
                    text = data.decode("utf-8")
                    text = re.sub(r'<dimension ref="[^"]+"', '<dimension ref="A1"', text, count=1)
                    data = text.encode("utf-8")
                zout.writestr(item, data)

        with load_workbook(str(stale)) as wb:
            ws = wb["Stale"]
            assert ws.calculate_dimension() == "A1:C7"
            assert ws.max_row == 7
            assert ws.max_column == 3

    def test_calculate_dimension_preserves_offset_used_range(self, tmp_path: Path) -> None:
        from openpyxl import Workbook

        from wolfxl import load_workbook

        path = tmp_path / "offset-dimension.xlsx"
        op_wb = Workbook()
        ws = op_wb.active
        ws.title = "Offset"
        ws["C4"] = "tail"
        op_wb.save(path)
        op_wb.close()

        with load_workbook(str(path)) as wb:
            ws = wb["Offset"]
            assert ws.calculate_dimension() == "C4:C4"
            assert ws.max_row == 4
            assert ws.max_column == 3

    def test_calculate_dimension_includes_buffered_write_apis(self) -> None:
        from wolfxl import Workbook

        wb = Workbook()
        ws = wb.active

        assert ws.calculate_dimension() == "A1:A1"
        ws.append(["Name", "Amount"])
        assert ws.calculate_dimension() == "A1:B1"
        ws.write_rows([["tail"]], start_row=4, start_col=3)
        assert ws.calculate_dimension() == "A1:C4"

    def test_date_cells_read_as_midnight_datetime(self) -> None:
        from wolfxl import load_workbook

        path = PARITY_FIXTURES / "flat_register" / "excelx_accounts_payable.xlsx"
        if not path.exists():
            pytest.skip("parity fixture not found")

        wb = load_workbook(str(path))
        value = wb["Sheet1"]["C2"].value
        wb.close()

        assert value == datetime(2024, 12, 16, 0, 0)
        assert isinstance(value, datetime)

    def test_data_only_returns_cached_formula_values(self) -> None:
        from wolfxl import load_workbook

        path = PARITY_FIXTURES / "time_series" / "ilpa_pe_fund_reporting_v1.1.xlsx"
        if not path.exists():
            pytest.skip("parity fixture not found")

        with load_workbook(str(path), data_only=False) as formula_wb:
            assert formula_wb["Reporting Template"]["E4"].value == "=P4"

        with load_workbook(str(path), data_only=True) as cached_wb:
            cell_value = cached_wb["Reporting Template"]["E4"].value
            row_value = next(
                cached_wb["Reporting Template"].iter_rows(
                    min_row=4,
                    max_row=4,
                    min_col=5,
                    max_col=5,
                    values_only=True,
                )
            )[0]

        expected = datetime(2015, 10, 1, 0, 0)
        assert cell_value == expected
        assert row_value == expected

    def test_sheet_dimensions_follow_worksheet_dimension_ref(self) -> None:
        from wolfxl import load_workbook

        path = PARITY_FIXTURES / "time_series" / "ilpa_pe_fund_reporting_v1.1.xlsx"
        if not path.exists():
            pytest.skip("parity fixture not found")

        with load_workbook(str(path)) as wb:
            ws = wb["Suggested Guidance"]
            assert ws.max_row == 201
            assert ws.max_column == 3

    def test_number_format_preserves_excel_escape_backslashes(self) -> None:
        from wolfxl import load_workbook

        path = PARITY_FIXTURES / "time_series" / "ilpa_pe_fund_reporting_v1.1.xlsx"
        if not path.exists():
            pytest.skip("parity fixture not found")

        with load_workbook(str(path)) as wb:
            assert wb["Reporting Template"]["E4"].number_format == r"\([$-409]mmm\-yy\ \-"

    def test_merged_subordinate_cells_do_not_expose_anchor_number_formats(self) -> None:
        from wolfxl import load_workbook

        path = PARITY_FIXTURES / "time_series" / "ilpa_pe_fund_reporting_v1.1.xlsx"
        if not path.exists():
            pytest.skip("parity fixture not found")

        with load_workbook(str(path)) as wb:
            ws = wb["Reporting Template"]
            assert ws["E13"].number_format == "#,##0_);(#,##0)"
            assert ws["F13"].number_format is None
            assert ws["G13"].number_format is None
            assert ws["K71"].number_format == r'"$"#,##0_);\("$"#,##0\)'
            assert ws["L71"].number_format is None
            assert ws["M71"].number_format is None
            assert ws["F67"].number_format is None
            assert ws["G67"].number_format is None

    def test_workbook_contains(self) -> None:
        from wolfxl import load_workbook

        path = FIXTURES / "tier1" / "01_cell_values.xlsx"
        if not path.exists():
            pytest.skip("fixture not found")
        wb = load_workbook(str(path))
        first = wb.sheetnames[0]
        assert first in wb
        assert "NonexistentSheet" not in wb
        wb.close()


# ======================================================================
# Write tests (require wolfxl._rust)
# ======================================================================


class TestWriteMode:
    """Write a new Excel file via RustXlsxWriterBook."""

    def setup_method(self) -> None:
        _require_rust()

    def test_write_basic(self, tmp_path: Path) -> None:
        from wolfxl import Workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "Hello"
        ws["B1"] = 42
        ws["C1"] = True
        out = tmp_path / "basic.xlsx"
        wb.save(str(out))
        assert out.exists()
        assert out.stat().st_size > 0

    def test_write_with_font(self, tmp_path: Path) -> None:
        from wolfxl import Font, Workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "Bold"
        ws["A1"].font = Font(bold=True, size=14, name="Arial")
        out = tmp_path / "font.xlsx"
        wb.save(str(out))
        assert out.exists()

    def test_write_with_fill(self, tmp_path: Path) -> None:
        from wolfxl import PatternFill, Workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "Colored"
        ws["A1"].fill = PatternFill(patternType="solid", fgColor="#FF0000")
        out = tmp_path / "fill.xlsx"
        wb.save(str(out))
        assert out.exists()

    def test_write_with_border(self, tmp_path: Path) -> None:
        from wolfxl import Border, Side, Workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "Bordered"
        ws["A1"].border = Border(
            left=Side(style="thin", color="#000000"),
            right=Side(style="thin", color="#000000"),
            top=Side(style="thin", color="#000000"),
            bottom=Side(style="thin", color="#000000"),
        )
        out = tmp_path / "border.xlsx"
        wb.save(str(out))
        assert out.exists()

    def test_write_with_alignment(self, tmp_path: Path) -> None:
        from wolfxl import Alignment, Workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "Centered"
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        out = tmp_path / "align.xlsx"
        wb.save(str(out))
        assert out.exists()

    def test_write_multiple_sheets(self, tmp_path: Path) -> None:
        from wolfxl import Workbook

        wb = Workbook()
        ws1 = wb.active
        assert ws1 is not None
        ws1["A1"] = "Sheet1 data"
        ws2 = wb.create_sheet("Data")
        ws2["A1"] = "Sheet2 data"
        assert wb.sheetnames == ["Sheet", "Data"]
        out = tmp_path / "multi.xlsx"
        wb.save(str(out))
        assert out.exists()

    def test_write_number_format(self, tmp_path: Path) -> None:
        from wolfxl import Workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = 42000
        ws["A1"].number_format = "$#,##0"
        out = tmp_path / "numfmt.xlsx"
        wb.save(str(out))
        assert out.exists()

    def test_cell_method(self, tmp_path: Path) -> None:
        from wolfxl import Workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        c = ws.cell(row=1, column=1, value="Via cell()")
        assert c.value == "Via cell()"
        assert c.coordinate == "A1"
        out = tmp_path / "cell_method.xlsx"
        wb.save(str(out))
        assert out.exists()


# ======================================================================
# Round-trip tests (write with wolfxl, read back)
# ======================================================================


class TestRoundTrip:
    """Write with wolfxl, read back with wolfxl."""

    def setup_method(self) -> None:
        _require_rust()

    def test_roundtrip_values(self, tmp_path: Path) -> None:
        from wolfxl import Workbook, load_workbook

        # Write
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "text"
        ws["A2"] = 123
        ws["A3"] = 3.14
        ws["A4"] = True
        out = tmp_path / "roundtrip.xlsx"
        wb.save(str(out))

        # Read back
        wb2 = load_workbook(str(out))
        ws2 = wb2[wb2.sheetnames[0]]
        assert ws2["A1"].value == "text"
        assert ws2["A2"].value == 123 or ws2["A2"].value == 123.0
        assert abs(ws2["A3"].value - 3.14) < 0.001
        assert ws2["A4"].value is True
        wb2.close()

    def test_roundtrip_font(self, tmp_path: Path) -> None:
        from wolfxl import Font, Workbook, load_workbook

        # Write
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "Bold"
        ws["A1"].font = Font(bold=True)
        out = tmp_path / "roundtrip_font.xlsx"
        wb.save(str(out))

        # Read back
        wb2 = load_workbook(str(out))
        ws2 = wb2[wb2.sheetnames[0]]
        assert ws2["A1"].value == "Bold"
        assert ws2["A1"].font.bold is True
        wb2.close()

    def test_roundtrip_fill(self, tmp_path: Path) -> None:
        from wolfxl import PatternFill, Workbook, load_workbook

        # Write
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "Red"
        ws["A1"].fill = PatternFill(patternType="solid", fgColor="#FF0000")
        out = tmp_path / "roundtrip_fill.xlsx"
        wb.save(str(out))

        # Read back
        wb2 = load_workbook(str(out))
        ws2 = wb2[wb2.sheetnames[0]]
        fill = ws2["A1"].fill
        assert fill.fgColor is not None
        # The color should contain FF0000 (exact format may vary)
        fg = str(fill.fgColor).upper()
        assert "FF0000" in fg
        wb2.close()

    def test_roundtrip_formula(self, tmp_path: Path) -> None:
        from wolfxl import Workbook, load_workbook

        # Write
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = 10
        ws["A2"] = 20
        ws["A3"] = "=SUM(A1:A2)"
        out = tmp_path / "roundtrip_formula.xlsx"
        wb.save(str(out))

        # Read back — formula should be preserved as string
        wb2 = load_workbook(str(out))
        ws2 = wb2[wb2.sheetnames[0]]
        val = ws2["A3"].value
        assert val is not None
        assert "SUM" in str(val).upper()
        wb2.close()


# ======================================================================
# Modify mode tests (load existing, modify, save, verify)
# ======================================================================


FIXTURE = FIXTURES / "tier1" / "01_cell_values.xlsx"


class TestModifyMode:
    """Test the read-modify-write path via WolfXL (XlsxPatcher)."""

    def setup_method(self) -> None:
        _require_rust()
        if not FIXTURE.exists():
            pytest.skip("tier1 fixture not available")

    def test_modify_repr(self) -> None:
        from wolfxl import load_workbook

        wb = load_workbook(str(FIXTURE), modify=True)
        assert "modify" in repr(wb)
        wb.close()

    def test_modify_string_value(self, tmp_path: Path) -> None:
        from wolfxl import load_workbook

        wb = load_workbook(str(FIXTURE), modify=True)
        ws = wb.active
        assert ws is not None
        ws["A1"] = "Modified"
        out = tmp_path / "mod_string.xlsx"
        wb.save(str(out))
        wb.close()

        # Verify with wolfxl read
        wb2 = load_workbook(str(out))
        assert wb2.active is not None
        assert wb2.active["A1"].value == "Modified"
        wb2.close()

    def test_modify_number_value(self, tmp_path: Path) -> None:
        from wolfxl import load_workbook

        wb = load_workbook(str(FIXTURE), modify=True)
        ws = wb.active
        assert ws is not None
        ws["B2"] = 99.5
        out = tmp_path / "mod_number.xlsx"
        wb.save(str(out))
        wb.close()

        wb2 = load_workbook(str(out))
        assert wb2.active is not None
        assert abs(wb2.active["B2"].value - 99.5) < 0.001
        wb2.close()

    def test_modify_boolean_value(self, tmp_path: Path) -> None:
        from wolfxl import load_workbook

        wb = load_workbook(str(FIXTURE), modify=True)
        ws = wb.active
        assert ws is not None
        ws["C3"] = True
        out = tmp_path / "mod_bool.xlsx"
        wb.save(str(out))
        wb.close()

        wb2 = load_workbook(str(out))
        assert wb2.active is not None
        assert wb2.active["C3"].value is True
        wb2.close()

    def test_modify_formula(self, tmp_path: Path) -> None:
        from wolfxl import load_workbook

        wb = load_workbook(str(FIXTURE), modify=True)
        ws = wb.active
        assert ws is not None
        ws["D4"] = "=SUM(1,2,3)"
        out = tmp_path / "mod_formula.xlsx"
        wb.save(str(out))
        wb.close()

        # Verify formula preserved (openpyxl reads with = prefix)
        import openpyxl

        wb2 = openpyxl.load_workbook(str(out))
        assert wb2.active is not None
        assert wb2.active["D4"].value == "=SUM(1,2,3)"
        wb2.close()

    def test_modify_preserves_unchanged(self, tmp_path: Path) -> None:
        """Cells not touched should remain unchanged after save."""
        from wolfxl import load_workbook

        # Read original B1
        wb_orig = load_workbook(str(FIXTURE))
        assert wb_orig.active is not None
        orig_b1 = wb_orig.active["B1"].value
        wb_orig.close()

        # Modify only A1
        wb = load_workbook(str(FIXTURE), modify=True)
        ws = wb.active
        assert ws is not None
        ws["A1"] = "Changed"
        out = tmp_path / "mod_preserve.xlsx"
        wb.save(str(out))
        wb.close()

        # B1 should still have its original value
        wb2 = load_workbook(str(out))
        assert wb2.active is not None
        assert wb2.active["B1"].value == orig_b1
        assert wb2.active["A1"].value == "Changed"
        wb2.close()

    def test_modify_read_then_write(self, tmp_path: Path) -> None:
        """Read a value, modify it, save — the classic read-modify-write cycle."""
        from wolfxl import load_workbook

        wb = load_workbook(str(FIXTURE), modify=True)
        ws = wb.active
        assert ws is not None
        original = ws["A1"].value  # read via calamine
        ws["A1"] = f"WAS: {original}"  # write via patcher
        out = tmp_path / "mod_rmw.xlsx"
        wb.save(str(out))
        wb.close()

        wb2 = load_workbook(str(out))
        assert wb2.active is not None
        assert wb2.active["A1"].value == f"WAS: {original}"
        wb2.close()

    def test_modify_insert_new_cell(self, tmp_path: Path) -> None:
        """Insert a cell at a position that didn't exist in the original."""
        from wolfxl import load_workbook

        wb = load_workbook(str(FIXTURE), modify=True)
        ws = wb.active
        assert ws is not None
        ws["Z99"] = "New cell"
        out = tmp_path / "mod_insert.xlsx"
        wb.save(str(out))
        wb.close()

        wb2 = load_workbook(str(out))
        assert wb2.active is not None
        assert wb2.active["Z99"].value == "New cell"
        wb2.close()

    def test_modify_font(self, tmp_path: Path) -> None:
        """Modify mode: set font on a cell."""
        import openpyxl

        from wolfxl import Font, load_workbook

        wb = load_workbook(str(FIXTURE), modify=True)
        ws = wb.active
        assert ws is not None
        ws["A1"] = "Bold"
        ws["A1"].font = Font(bold=True, size=14, name="Arial", color="#FF0000")
        out = tmp_path / "mod_font.xlsx"
        wb.save(str(out))
        wb.close()

        wb2 = openpyxl.load_workbook(str(out))
        assert wb2.active is not None
        f = wb2.active["A1"].font
        assert f.bold is True
        assert f.size == 14.0
        assert f.name == "Arial"
        assert "FF0000" in str(f.color.rgb)
        wb2.close()

    def test_modify_format_only_preserves_value(self, tmp_path: Path) -> None:
        """Modify mode: format-only edits must preserve existing cell values."""
        import openpyxl

        from wolfxl import Font, load_workbook

        # Read original value via openpyxl.
        wb0 = openpyxl.load_workbook(str(FIXTURE))
        assert wb0.active is not None
        original = wb0.active["B2"].value
        wb0.close()

        # Modify only formatting (no value assignment).
        wb = load_workbook(str(FIXTURE), modify=True)
        ws = wb.active
        assert ws is not None
        ws["B2"].font = Font(bold=True)
        out = tmp_path / "mod_format_only.xlsx"
        wb.save(str(out))
        wb.close()

        wb2 = openpyxl.load_workbook(str(out))
        assert wb2.active is not None
        cell = wb2.active["B2"]
        assert cell.value == original
        assert cell.font.bold is True
        wb2.close()

    def test_modify_fill(self, tmp_path: Path) -> None:
        """Modify mode: set fill on a cell."""
        import openpyxl

        from wolfxl import PatternFill, load_workbook

        wb = load_workbook(str(FIXTURE), modify=True)
        ws = wb.active
        assert ws is not None
        ws["A1"] = "Yellow"
        ws["A1"].fill = PatternFill(patternType="solid", fgColor="#FFFF00")
        out = tmp_path / "mod_fill.xlsx"
        wb.save(str(out))
        wb.close()

        wb2 = openpyxl.load_workbook(str(out))
        assert wb2.active is not None
        fill = wb2.active["A1"].fill
        assert fill.patternType == "solid"
        assert "FFFF00" in str(fill.fgColor.rgb)
        wb2.close()

    def test_modify_alignment(self, tmp_path: Path) -> None:
        """Modify mode: set alignment on a cell."""
        import openpyxl

        from wolfxl import Alignment, load_workbook

        wb = load_workbook(str(FIXTURE), modify=True)
        ws = wb.active
        assert ws is not None
        ws["A1"] = "Centered"
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        out = tmp_path / "mod_align.xlsx"
        wb.save(str(out))
        wb.close()

        wb2 = openpyxl.load_workbook(str(out))
        assert wb2.active is not None
        al = wb2.active["A1"].alignment
        assert al.horizontal == "center"
        assert al.vertical == "center"
        assert al.wrapText is True
        wb2.close()

    def test_modify_border(self, tmp_path: Path) -> None:
        """Modify mode: set border on a cell."""
        import openpyxl

        from wolfxl import Border, Side, load_workbook

        wb = load_workbook(str(FIXTURE), modify=True)
        ws = wb.active
        assert ws is not None
        ws["A1"] = "Bordered"
        ws["A1"].border = Border(
            left=Side(style="thin", color="#000000"),
            right=Side(style="medium", color="#FF0000"),
            top=Side(style="thin", color="#000000"),
            bottom=Side(style="thin", color="#000000"),
        )
        out = tmp_path / "mod_border.xlsx"
        wb.save(str(out))
        wb.close()

        wb2 = openpyxl.load_workbook(str(out))
        assert wb2.active is not None
        b = wb2.active["A1"].border
        assert b.left.style == "thin"
        assert b.right.style == "medium"
        wb2.close()

    def test_modify_number_format(self, tmp_path: Path) -> None:
        """Modify mode: set number format on a cell."""
        import openpyxl

        from wolfxl import load_workbook

        wb = load_workbook(str(FIXTURE), modify=True)
        ws = wb.active
        assert ws is not None
        ws["A1"] = 42000
        ws["A1"].number_format = "$#,##0"
        out = tmp_path / "mod_numfmt.xlsx"
        wb.save(str(out))
        wb.close()

        wb2 = openpyxl.load_workbook(str(out))
        assert wb2.active is not None
        assert wb2.active["A1"].number_format == "$#,##0"
        wb2.close()

    def test_modify_combined_value_and_format(self, tmp_path: Path) -> None:
        """Modify mode: set both value and format on the same cell."""
        import openpyxl

        from wolfxl import Font, PatternFill, load_workbook

        wb = load_workbook(str(FIXTURE), modify=True)
        ws = wb.active
        assert ws is not None
        ws["A1"] = "Styled"
        ws["A1"].font = Font(bold=True, italic=True)
        ws["A1"].fill = PatternFill(patternType="solid", fgColor="#00FF00")
        out = tmp_path / "mod_combined.xlsx"
        wb.save(str(out))
        wb.close()

        wb2 = openpyxl.load_workbook(str(out))
        assert wb2.active is not None
        cell = wb2.active["A1"]
        assert cell.value == "Styled"
        assert cell.font.bold is True
        assert cell.font.italic is True
        assert "00FF00" in str(cell.fill.fgColor.rgb)
        wb2.close()

    def test_modify_multiple_sheets(self, tmp_path: Path) -> None:
        """Modify mode: patch cells across multiple sheets."""
        multi_fixture = FIXTURES /"tier1" / "09_multiple_sheets.xlsx"
        if not multi_fixture.exists():
            pytest.skip("multi-sheet fixture not available")

        import openpyxl

        from wolfxl import load_workbook

        wb = load_workbook(str(multi_fixture), modify=True)
        wb["Alpha"]["A1"] = "Patched Alpha"
        wb["Beta"]["B2"] = 999
        wb["Gamma"]["C3"] = True
        out = tmp_path / "mod_multi.xlsx"
        wb.save(str(out))
        wb.close()

        wb2 = openpyxl.load_workbook(str(out))
        assert wb2["Alpha"]["A1"].value == "Patched Alpha"
        assert wb2["Beta"]["B2"].value == 999
        assert wb2["Gamma"]["C3"].value is True
        # Originals preserved
        assert wb2["Alpha"]["B1"].value is not None  # original data still there
        wb2.close()

    def test_modify_preserves_images(self, tmp_path: Path) -> None:
        """Modify mode: files with images should preserve them after patching."""
        img_fixture = FIXTURES /"tier2" / "14_images.xlsx"
        if not img_fixture.exists():
            pytest.skip("images fixture not available")

        import openpyxl

        from wolfxl import load_workbook

        orig_size = img_fixture.stat().st_size

        wb = load_workbook(str(img_fixture), modify=True)
        ws = wb.active
        assert ws is not None
        ws["A1"] = "Has images"
        out = tmp_path / "mod_images.xlsx"
        wb.save(str(out))
        wb.close()

        # File should still be similar size (images preserved)
        new_size = out.stat().st_size
        ratio = new_size / orig_size
        assert 0.5 < ratio < 2.0, f"Size ratio {ratio:.2f} suggests corruption"

        # Should still open fine
        wb2 = openpyxl.load_workbook(str(out))
        assert wb2.active is not None
        assert wb2.active["A1"].value == "Has images"
        wb2.close()

    def test_modify_preserves_hyperlinks(self, tmp_path: Path) -> None:
        """Modify mode: files with hyperlinks should preserve them."""
        link_fixture = FIXTURES /"tier2" / "13_hyperlinks.xlsx"
        if not link_fixture.exists():
            pytest.skip("hyperlinks fixture not available")

        import openpyxl

        from wolfxl import load_workbook

        wb = load_workbook(str(link_fixture), modify=True)
        ws = wb.active
        assert ws is not None
        ws["A1"] = "Has links"
        out = tmp_path / "mod_links.xlsx"
        wb.save(str(out))
        wb.close()

        wb2 = openpyxl.load_workbook(str(out))
        assert wb2.active is not None
        assert wb2.active["A1"].value == "Has links"
        wb2.close()


# ======================================================================
# Batch flush tests (write_sheet_values optimization)
# ======================================================================


class TestBatchFlush:
    """Verify the batch flush path produces correct output."""

    def setup_method(self) -> None:
        _require_rust()

    def test_batch_numeric_roundtrip(self, tmp_path: Path) -> None:
        """Bulk numeric writes should batch into write_sheet_values."""
        from wolfxl import Workbook, load_workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        # Write a 100×10 grid of numbers — all batchable types.
        for r in range(1, 101):
            for c in range(1, 11):
                ws.cell(row=r, column=c, value=r * 100 + c)
        out = tmp_path / "batch_numeric.xlsx"
        wb.save(str(out))

        wb2 = load_workbook(str(out))
        ws2 = wb2[wb2.sheetnames[0]]
        assert ws2["A1"].value == 101
        assert ws2["J100"].value == 10010
        assert ws2["E50"].value == 5005
        wb2.close()

    def test_batch_mixed_types(self, tmp_path: Path) -> None:
        """Mixed types: ints/strs batch, bools/formulas go per-cell."""
        from wolfxl import Workbook, load_workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = 42          # batchable
        ws["B1"] = "hello"     # batchable
        ws["C1"] = 3.14        # batchable
        ws["D1"] = True        # NOT batchable (bool)
        ws["E1"] = "=SUM(1,2)" # NOT batchable (formula)
        ws["F1"] = None         # batchable (skip)
        out = tmp_path / "batch_mixed.xlsx"
        wb.save(str(out))

        wb2 = load_workbook(str(out))
        ws2 = wb2[wb2.sheetnames[0]]
        assert ws2["A1"].value == 42 or ws2["A1"].value == 42.0
        assert ws2["B1"].value == "hello"
        assert abs(ws2["C1"].value - 3.14) < 0.001
        assert ws2["D1"].value is True
        val_e = ws2["E1"].value
        assert val_e is not None and "SUM" in str(val_e).upper()
        wb2.close()

    def test_batch_with_format(self, tmp_path: Path) -> None:
        """Cells with both values and formats: value batches, format per-cell."""
        from wolfxl import Font, Workbook, load_workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        # A1: value + format (value goes to batch, format goes per-cell)
        ws["A1"] = "styled"
        ws["A1"].font = Font(bold=True)
        # B1: value only (pure batch)
        ws["B1"] = "plain"
        out = tmp_path / "batch_format.xlsx"
        wb.save(str(out))

        wb2 = load_workbook(str(out))
        ws2 = wb2[wb2.sheetnames[0]]
        assert ws2["A1"].value == "styled"
        assert ws2["A1"].font.bold is True
        assert ws2["B1"].value == "plain"
        wb2.close()

    def test_batch_large_grid(self, tmp_path: Path) -> None:
        """10K cells — the batch path should handle this without issues."""
        from wolfxl import Workbook, load_workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        rows, cols = 1000, 10
        for r in range(1, rows + 1):
            for c in range(1, cols + 1):
                ws.cell(row=r, column=c, value=float(r * cols + c))
        out = tmp_path / "batch_10k.xlsx"
        wb.save(str(out))

        wb2 = load_workbook(str(out))
        ws2 = wb2[wb2.sheetnames[0]]
        # Spot-check corners
        assert ws2["A1"].value == 11.0
        assert ws2["J1000"].value == 10010.0
        wb2.close()


# ======================================================================
# Append tests (openpyxl-compatible ws.append)
# ======================================================================


class TestAppend:
    """Test ws.append() — openpyxl-compatible row insertion."""

    def setup_method(self) -> None:
        _require_rust()

    def test_append_basic(self, tmp_path: Path) -> None:
        from wolfxl import Workbook, load_workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["Name", "Age", "City"])
        ws.append(["Alice", 30, "NYC"])
        ws.append(["Bob", 25, "LA"])
        out = tmp_path / "append_basic.xlsx"
        wb.save(str(out))

        wb2 = load_workbook(str(out))
        ws2 = wb2[wb2.sheetnames[0]]
        assert ws2["A1"].value == "Name"
        assert ws2["B2"].value == 30 or ws2["B2"].value == 30.0
        assert ws2["C3"].value == "LA"
        wb2.close()

    def test_append_many_rows(self, tmp_path: Path) -> None:
        """Append 5000 rows — exercises the batch flush path."""
        from wolfxl import Workbook, load_workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        for i in range(5000):
            ws.append([i, f"row_{i}", i * 1.1])
        out = tmp_path / "append_5k.xlsx"
        wb.save(str(out))

        wb2 = load_workbook(str(out))
        ws2 = wb2[wb2.sheetnames[0]]
        assert ws2["A1"].value == 0 or ws2["A1"].value == 0.0
        assert ws2["B5000"].value == "row_4999"
        wb2.close()

    def test_append_mixed_with_cell(self, tmp_path: Path) -> None:
        """Mix append() with direct cell() writes.

        Like openpyxl, direct cell() writes do NOT advance the append counter.
        So append() after cell(row=2) still writes to the next append row (2),
        overwriting the cell() value. This matches openpyxl semantics.
        """
        from wolfxl import Workbook, load_workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["Header1", "Header2"])   # row 1, next_append = 2
        ws.cell(row=5, column=1, value="manual")  # does NOT advance counter
        ws.append(["Row2_A", "Row2_B"])     # row 2 (matches openpyxl)
        out = tmp_path / "append_mixed.xlsx"
        wb.save(str(out))

        wb2 = load_workbook(str(out))
        ws2 = wb2[wb2.sheetnames[0]]
        assert ws2["A1"].value == "Header1"
        assert ws2["A2"].value == "Row2_A"
        assert ws2["A5"].value == "manual"
        wb2.close()

    def test_append_empty_row(self, tmp_path: Path) -> None:
        """Appending an empty iterable should still advance the row counter."""
        from wolfxl import Workbook, load_workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["row1"])
        ws.append([])  # empty
        ws.append(["row3"])
        out = tmp_path / "append_empty.xlsx"
        wb.save(str(out))

        wb2 = load_workbook(str(out))
        ws2 = wb2[wb2.sheetnames[0]]
        assert ws2["A1"].value == "row1"
        assert ws2["A3"].value == "row3"
        wb2.close()


# ======================================================================
# PathLike support tests
# ======================================================================


class TestPathLike:
    """Test os.PathLike support for save() and load_workbook()."""

    def setup_method(self) -> None:
        _require_rust()

    def test_save_with_path(self, tmp_path: Path) -> None:
        from wolfxl import Workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "pathlib"
        out = tmp_path / "pathlike.xlsx"
        wb.save(out)  # Pass Path object, not str
        assert out.exists()

    def test_load_with_path(self, tmp_path: Path) -> None:
        from wolfxl import Workbook, load_workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "pathlib"
        out = tmp_path / "pathlike_load.xlsx"
        wb.save(out)

        wb2 = load_workbook(out)  # Pass Path object, not str
        assert wb2.active is not None
        assert wb2.active["A1"].value == "pathlib"
        wb2.close()

    def test_save_and_load_modify_with_path(self) -> None:
        fixture = FIXTURES / "tier1" / "01_cell_values.xlsx"
        if not fixture.exists():
            pytest.skip("fixture not found")
        from wolfxl import load_workbook

        wb = load_workbook(fixture, modify=True)  # Path object
        assert wb.active is not None
        wb.close()


# ======================================================================
# Title setter tests
# ======================================================================


class TestTitleSetter:
    """Test ws.title setter for renaming worksheets."""

    def setup_method(self) -> None:
        _require_rust()

    def test_rename_sheet(self) -> None:
        from wolfxl import Workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        assert ws.title == "Sheet"
        ws.title = "Data"
        assert ws.title == "Data"
        assert wb.sheetnames == ["Data"]
        assert "Data" in wb
        assert "Sheet" not in wb

    def test_rename_duplicate_raises(self) -> None:
        from wolfxl import Workbook

        wb = Workbook()
        wb.create_sheet("Other")
        ws = wb.active
        assert ws is not None
        with pytest.raises(ValueError, match="already exists"):
            ws.title = "Other"

    def test_rename_noop(self) -> None:
        from wolfxl import Workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Sheet"  # Same name — should be a no-op
        assert ws.title == "Sheet"


# ======================================================================
# Opt 4: Title setter Rust sync tests
# ======================================================================


class TestTitleSetterRust:
    """Test that ws.title setter syncs with Rust writer (Opt 4)."""

    def setup_method(self) -> None:
        _require_rust()

    def test_rename_then_save(self, tmp_path: Path) -> None:
        """ws.title = 'X' -> save -> load -> sheet name is 'X'."""
        from wolfxl import Workbook, load_workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "hello"
        ws.title = "Renamed"
        out = tmp_path / "rename_save.xlsx"
        wb.save(str(out))

        wb2 = load_workbook(str(out))
        assert wb2.sheetnames == ["Renamed"]
        assert wb2["Renamed"]["A1"].value == "hello"
        wb2.close()

    def test_rename_preserves_data(self, tmp_path: Path) -> None:
        """Data written before rename survives save/load."""
        from wolfxl import Workbook, load_workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "before"
        ws["B2"] = 42
        ws.title = "NewName"
        ws["C3"] = "after"
        out = tmp_path / "rename_data.xlsx"
        wb.save(str(out))

        wb2 = load_workbook(str(out))
        ws2 = wb2["NewName"]
        assert ws2["A1"].value == "before"
        assert ws2["B2"].value == 42 or ws2["B2"].value == 42.0
        assert ws2["C3"].value == "after"
        wb2.close()

    def test_rename_multiple_sheets(self, tmp_path: Path) -> None:
        """Rename one of several sheets, others unaffected."""
        from wolfxl import Workbook, load_workbook

        wb = Workbook()
        ws1 = wb.active
        assert ws1 is not None
        ws1["A1"] = "sheet1"
        ws2 = wb.create_sheet("Other")
        ws2["A1"] = "sheet2"
        ws1.title = "First"
        out = tmp_path / "rename_multi.xlsx"
        wb.save(str(out))

        wb2 = load_workbook(str(out))
        assert set(wb2.sheetnames) == {"First", "Other"}
        assert wb2["First"]["A1"].value == "sheet1"
        assert wb2["Other"]["A1"].value == "sheet2"
        wb2.close()

    def test_rename_then_append(self, tmp_path: Path) -> None:
        """Rename then append rows - both paths sync correctly."""
        from wolfxl import Workbook, load_workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Data"
        ws.append(["col1", "col2"])
        ws.append([1, 2])
        ws.append([3, 4])
        out = tmp_path / "rename_append.xlsx"
        wb.save(str(out))

        wb2 = load_workbook(str(out))
        assert wb2.sheetnames == ["Data"]
        ws2 = wb2["Data"]
        assert ws2["A1"].value == "col1"
        assert ws2["B3"].value == 4 or ws2["B3"].value == 4.0
        wb2.close()


# ======================================================================
# Opt 1: Bulk read tests
# ======================================================================


class TestBulkRead:
    """Test iter_rows bulk path via read_sheet_values_plain (Opt 1)."""

    def setup_method(self) -> None:
        _require_rust()

    def test_values_only_matches_cell_by_cell(self, tmp_path: Path) -> None:
        """Bulk path produces identical output to per-cell path."""
        from wolfxl import Workbook, load_workbook

        # Write known data.
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["Name", "Age", "Score"])
        ws.append(["Alice", 30, 95.5])
        ws.append(["Bob", 25, 88.0])
        out = tmp_path / "bulk_read.xlsx"
        wb.save(str(out))

        # Read with values_only (exercises bulk path).
        wb2 = load_workbook(str(out))
        ws2 = wb2[wb2.sheetnames[0]]
        rows = list(ws2.iter_rows(values_only=True))
        assert len(rows) == 3
        assert rows[0] == ("Name", "Age", "Score")
        # Numbers may be float.
        assert rows[1][0] == "Alice"
        assert rows[1][1] == 30 or rows[1][1] == 30.0
        assert abs(rows[1][2] - 95.5) < 0.01
        wb2.close()

    def test_values_only_with_range_limits(self, tmp_path: Path) -> None:
        """min_row/max_row/min_col/max_col respected."""
        from wolfxl import Workbook, load_workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        for r in range(1, 6):
            ws.append([r * 10 + c for c in range(1, 6)])
        out = tmp_path / "bulk_range.xlsx"
        wb.save(str(out))

        wb2 = load_workbook(str(out))
        ws2 = wb2[wb2.sheetnames[0]]
        rows = list(ws2.iter_rows(min_row=2, max_row=4, min_col=2, max_col=4,
                                  values_only=True))
        assert len(rows) == 3
        # Row 2, cols 2-4 should be [22, 23, 24].
        assert len(rows[0]) == 3
        r2_vals = [v if isinstance(v, int) else int(v) for v in rows[0]]
        assert r2_vals == [22, 23, 24]
        wb2.close()

    def test_values_only_empty_sheet(self, tmp_path: Path) -> None:
        """Empty sheet yields no rows."""
        from wolfxl import Workbook, load_workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        # Write nothing, just save.
        out = tmp_path / "bulk_empty.xlsx"
        wb.save(str(out))

        wb2 = load_workbook(str(out))
        ws2 = wb2[wb2.sheetnames[0]]
        rows = list(ws2.iter_rows(values_only=True))
        # Should produce at most 1 row (dimension-detected as 1x1).
        assert len(rows) <= 1
        wb2.close()

    def test_values_only_mixed_types(self, tmp_path: Path) -> None:
        """str, int, float, bool, None all correct."""
        from wolfxl import Workbook, load_workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "text"
        ws["B1"] = 42
        ws["C1"] = 3.14
        ws["D1"] = True
        # E1 left empty (None).
        out = tmp_path / "bulk_types.xlsx"
        wb.save(str(out))

        wb2 = load_workbook(str(out))
        ws2 = wb2[wb2.sheetnames[0]]
        rows = list(ws2.iter_rows(min_row=1, max_row=1, min_col=1, max_col=5,
                                  values_only=True))
        assert len(rows) == 1
        row = rows[0]
        assert row[0] == "text"
        assert row[1] == 42 or row[1] == 42.0
        assert abs(row[2] - 3.14) < 0.01
        assert row[3] is True
        assert row[4] is None
        wb2.close()


# ======================================================================
# Opt 3: write_rows tests
# ======================================================================


class TestWriteRows:
    """Test write_rows bulk API (Opt 3)."""

    def setup_method(self) -> None:
        _require_rust()

    def test_write_rows_basic(self, tmp_path: Path) -> None:
        """write_rows produces same file as cell-by-cell."""
        from wolfxl import Workbook, load_workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        grid = [
            ["Name", "Age"],
            ["Alice", 30],
            ["Bob", 25],
        ]
        ws.write_rows(grid)
        out = tmp_path / "write_rows.xlsx"
        wb.save(str(out))

        wb2 = load_workbook(str(out))
        ws2 = wb2[wb2.sheetnames[0]]
        assert ws2["A1"].value == "Name"
        assert ws2["B1"].value == "Age"
        assert ws2["A2"].value == "Alice"
        assert ws2["B2"].value == 30 or ws2["B2"].value == 30.0
        assert ws2["A3"].value == "Bob"
        wb2.close()

    def test_write_rows_offset(self, tmp_path: Path) -> None:
        """start_row/start_col positioning correct."""
        from wolfxl import Workbook, load_workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.write_rows([["offset"]], start_row=3, start_col=2)
        out = tmp_path / "write_rows_offset.xlsx"
        wb.save(str(out))

        wb2 = load_workbook(str(out))
        ws2 = wb2[wb2.sheetnames[0]]
        assert ws2["B3"].value == "offset"
        wb2.close()

    def test_write_rows_roundtrip(self, tmp_path: Path) -> None:
        """write_rows -> save -> load -> read matches input."""
        from wolfxl import Workbook, load_workbook

        grid = [[i * 10 + j for j in range(5)] for i in range(100)]
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.write_rows(grid)
        out = tmp_path / "write_rows_rt.xlsx"
        wb.save(str(out))

        wb2 = load_workbook(str(out))
        ws2 = wb2[wb2.sheetnames[0]]
        rows = list(ws2.iter_rows(values_only=True))
        assert len(rows) == 100
        for i, row in enumerate(rows):
            for j, val in enumerate(row):
                expected = i * 10 + j
                assert val == expected or val == float(expected)
        wb2.close()

    def test_write_rows_with_booleans(self, tmp_path: Path) -> None:
        """write_rows handles non-batchable types (bools, formulas)."""
        from wolfxl import Workbook, load_workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.write_rows([[True, False, "=1+2"]], start_row=1, start_col=1)
        out = tmp_path / "write_rows_bool.xlsx"
        wb.save(str(out))

        wb2 = load_workbook(str(out))
        ws2 = wb2[wb2.sheetnames[0]]
        assert ws2["A1"].value is True
        assert ws2["B1"].value is False
        val_c = ws2["C1"].value
        assert val_c is not None and "1" in str(val_c)
        wb2.close()

    def test_write_rows_empty(self, tmp_path: Path) -> None:
        """write_rows with empty list is a no-op."""
        from wolfxl import Workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.write_rows([])  # should not raise
        out = tmp_path / "write_rows_empty.xlsx"
        wb.save(str(out))
        assert out.exists()


# ======================================================================
# Opt 5: Plain read tests
# ======================================================================


class TestPlainRead:
    """Test read_sheet_values_plain Rust method (Opt 5)."""

    def setup_method(self) -> None:
        _require_rust()

    def test_plain_matches_payload(self, tmp_path: Path) -> None:
        """Plain values match _payload_to_python(dict) values."""
        from wolfxl._cell import _payload_to_python

        from wolfxl import Workbook, load_workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "text"
        ws["B1"] = 42
        ws["C1"] = 3.14
        out = tmp_path / "plain_match.xlsx"
        wb.save(str(out))

        wb2 = load_workbook(str(out))
        reader = wb2._rust_reader
        sheet = wb2.sheetnames[0]

        # Get dict-based values.
        dict_rows = reader.read_sheet_values(sheet, "A1:C1")
        dict_vals = [_payload_to_python(cell) for cell in dict_rows[0]]

        # Get plain values.
        plain_rows = reader.read_sheet_values_plain(sheet, "A1:C1")
        plain_vals = list(plain_rows[0])

        # Compare.
        assert len(dict_vals) == len(plain_vals)
        for dv, pv in zip(dict_vals, plain_vals):
            if isinstance(dv, float) and isinstance(pv, float):
                assert abs(dv - pv) < 0.001
            else:
                assert dv == pv, f"Mismatch: dict={dv!r} plain={pv!r}"
        wb2.close()

    def test_plain_all_types(self, tmp_path: Path) -> None:
        """All basic calamine Data types correctly converted."""
        from wolfxl import Workbook, load_workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "hello"    # string
        ws["B1"] = 42         # int -> number
        ws["C1"] = 3.14       # float
        ws["D1"] = True       # bool
        # E1 empty -> None
        out = tmp_path / "plain_types.xlsx"
        wb.save(str(out))

        wb2 = load_workbook(str(out))
        reader = wb2._rust_reader
        sheet = wb2.sheetnames[0]
        rows = reader.read_sheet_values_plain(sheet, "A1:E1")
        assert len(rows) == 1
        row = rows[0]
        assert row[0] == "hello"
        assert row[1] == 42 or row[1] == 42.0
        assert abs(row[2] - 3.14) < 0.01
        assert row[3] is True
        assert row[4] is None
        wb2.close()


# ======================================================================
# openpyxl Compat Expansion (Phase 7)
# ======================================================================


class TestFreezePanes:
    """freeze_panes property on write-mode worksheets."""

    def test_set_and_get(self) -> None:
        from wolfxl import Workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        assert ws.freeze_panes is None
        ws.freeze_panes = "B2"
        assert ws.freeze_panes == "B2"

    def test_roundtrip(self, tmp_path: Path) -> None:
        """Freeze panes survive write -> read cycle."""
        from wolfxl import Workbook, load_workbook

        p = tmp_path / "freeze.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "header"
        ws.freeze_panes = "A2"
        wb.save(str(p))

        wb2 = load_workbook(str(p))
        ws2 = wb2.active
        assert ws2 is not None
        assert ws2.freeze_panes == "A2"
        wb2.close()


class TestRowDimensions:
    """row_dimensions proxy on worksheets."""

    def test_set_height(self) -> None:
        from wolfxl import Workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.row_dimensions[1].height = 30.0
        assert ws.row_dimensions[1].height == 30.0

    def test_roundtrip(self, tmp_path: Path) -> None:
        from wolfxl import Workbook, load_workbook

        p = tmp_path / "rowheight.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "data"
        ws.row_dimensions[1].height = 25.0
        wb.save(str(p))

        wb2 = load_workbook(str(p))
        ws2 = wb2.active
        assert ws2 is not None
        h = ws2.row_dimensions[1].height
        assert h is not None
        assert abs(h - 25.0) < 1.0
        wb2.close()


class TestColumnDimensions:
    """column_dimensions proxy on worksheets."""

    def test_set_width(self) -> None:
        from wolfxl import Workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.column_dimensions["A"].width = 20.0
        assert ws.column_dimensions["A"].width == 20.0

    def test_roundtrip(self, tmp_path: Path) -> None:
        from wolfxl import Workbook, load_workbook

        p = tmp_path / "colwidth.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "data"
        ws.column_dimensions["A"].width = 18.0
        wb.save(str(p))

        wb2 = load_workbook(str(p))
        ws2 = wb2.active
        assert ws2 is not None
        w = ws2.column_dimensions["A"].width
        assert w is not None
        assert abs(w - 18.0) < 1.0
        wb2.close()


class TestAutoFilter:
    """auto_filter proxy on worksheets."""

    def test_set_ref(self) -> None:
        from wolfxl import Workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        assert ws.auto_filter.ref is None
        ws.auto_filter.ref = "A1:D10"
        assert ws.auto_filter.ref == "A1:D10"


class TestUnmergeCells:
    """unmerge_cells method on worksheets."""

    def setup_method(self) -> None:
        _require_rust()

    def test_basic_unmerge(self, tmp_path: Path) -> None:
        """Merge then unmerge removes the range from internal tracking."""
        from wolfxl import Workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.merge_cells("A1:B2")
        assert "A1:B2" in ws._merged_ranges
        ws.unmerge_cells("A1:B2")
        assert "A1:B2" not in ws._merged_ranges

    def test_unmerge_unknown_range_no_error(self) -> None:
        """Unmerging a range that was never merged should not raise."""
        from wolfxl import Workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        # Should silently do nothing.
        ws.unmerge_cells("C3:D4")
        assert "C3:D4" not in ws._merged_ranges

    def test_cells_accessible_after_unmerge(self, tmp_path: Path) -> None:
        """After unmerge, individual cells remain writable and readable."""
        from wolfxl import Workbook, load_workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.merge_cells("A1:B2")
        ws.unmerge_cells("A1:B2")
        ws["A1"] = "top-left"
        ws["B2"] = "bottom-right"
        out = tmp_path / "unmerge_access.xlsx"
        wb.save(str(out))

        wb2 = load_workbook(str(out))
        ws2 = wb2[wb2.sheetnames[0]]
        assert ws2["A1"].value == "top-left"
        assert ws2["B2"].value == "bottom-right"
        wb2.close()


class TestPrintArea:
    """print_area property on worksheets."""

    def test_default_is_none(self) -> None:
        from wolfxl import Workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        assert ws.print_area is None

    def test_set_and_get(self) -> None:
        from wolfxl import Workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.print_area = "A1:D10"
        assert ws.print_area == "A1:D10"

    def test_set_to_none(self) -> None:
        """Setting print_area back to None clears it."""
        from wolfxl import Workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.print_area = "A1:D10"
        ws.print_area = None
        assert ws.print_area is None

    def test_roundtrip(self, tmp_path: Path) -> None:
        """print_area survives save to xlsx (verified via XML inspection)."""
        import zipfile

        from wolfxl import Workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "data"
        ws.print_area = "A1:D10"
        out = str(tmp_path / "pa.xlsx")
        wb.save(out)

        # Verify print area is in the workbook XML
        with zipfile.ZipFile(out) as zf:
            # Print areas are defined as named ranges in xl/workbook.xml
            workbook_xml = zf.read("xl/workbook.xml").decode("utf-8")
            assert "Print_Area" in workbook_xml or "print_area" in workbook_xml.lower()
