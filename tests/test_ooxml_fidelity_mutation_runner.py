from __future__ import annotations

import importlib.util
import shutil
import sys
import zipfile
from pathlib import Path
from types import ModuleType

import openpyxl


def _load_runner_module() -> ModuleType:
    script = Path(__file__).resolve().parents[1] / "scripts" / "run_ooxml_fidelity_mutations.py"
    spec = importlib.util.spec_from_file_location("run_ooxml_fidelity_mutations", script)
    assert spec is not None
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


runner_module = _load_runner_module()


def _make_fixture(path: Path) -> None:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"
    worksheet["A1"] = "revenue"
    worksheet["A2"] = 100
    workbook.save(path)


def _make_empty_workbook_without_sheets(path: Path) -> None:
    parts = {
        "[Content_Types].xml": """<?xml version="1.0" encoding="UTF-8"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
</Types>""",
        "_rels/.rels": """<?xml version="1.0" encoding="UTF-8"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
</Relationships>""",
        "xl/workbook.xml": """<?xml version="1.0" encoding="UTF-8"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"
          xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <sheets/>
</workbook>""",
        "xl/_rels/workbook.xml.rels": """<?xml version="1.0" encoding="UTF-8"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"/>""",
    }
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as archive:
        for name, content in parts.items():
            archive.writestr(name, content)


def test_safe_tail_row_uses_package_level_feature_refs(tmp_path: Path) -> None:
    fixture = tmp_path / "data-validation-tail.xlsx"
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet["A1"] = "status"
    validation = openpyxl.worksheet.datavalidation.DataValidation(
        type="list",
        formula1='"Open,Closed"',
        showErrorMessage=True,
    )
    validation.add("C2:C20")
    worksheet.add_data_validation(validation)
    workbook.save(fixture)

    loaded = runner_module.wolfxl.load_workbook(fixture, modify=True)
    try:
        assert loaded.active.max_row == 1
        assert runner_module._safe_tail_row_index(fixture, loaded.active) == 21
    finally:
        loaded.close()


def test_safe_tail_col_uses_package_level_feature_refs(tmp_path: Path) -> None:
    fixture = tmp_path / "data-validation-tail-col.xlsx"
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet["A1"] = "status"
    validation = openpyxl.worksheet.datavalidation.DataValidation(
        type="list",
        formula1='"Open,Closed"',
        showErrorMessage=True,
    )
    validation.add("AA1:AA2")
    worksheet.add_data_validation(validation)
    workbook.save(fixture)

    loaded = runner_module.wolfxl.load_workbook(fixture, modify=True)
    try:
        assert loaded.active.max_column == 1
        assert runner_module._safe_tail_col_index(fixture, loaded.active) == 28
    finally:
        loaded.close()


def test_runner_writes_report_for_safe_mutations(tmp_path: Path) -> None:
    fixture_dir = tmp_path / "fixtures"
    output_dir = tmp_path / "out"
    fixture_dir.mkdir()
    fixture = fixture_dir / "simple.xlsx"
    _make_fixture(fixture)

    report = runner_module.run_sweep(fixture_dir, output_dir)

    assert report["result_count"] == 9
    assert report["failure_count"] == 0
    assert (output_dir / "report.json").is_file()
    statuses = {result["mutation"]: result["status"] for result in report["results"]}
    assert statuses == {
        "no_op": "passed",
        "marker_cell": "passed",
        "style_cell": "passed_with_expected_drift",
        "insert_tail_row": "passed",
        "insert_tail_col": "passed",
        "delete_marker_tail_row": "passed",
        "delete_marker_tail_col": "passed",
        "copy_remove_sheet": "passed",
        "move_marker_range": "passed",
    }


def test_runner_treats_existing_sheet_mutation_as_noop_when_no_sheets(
    tmp_path: Path,
) -> None:
    fixture_dir = tmp_path / "fixtures"
    output_dir = tmp_path / "out"
    fixture_dir.mkdir()
    _make_empty_workbook_without_sheets(fixture_dir / "empty_worksheet.xlsx")

    report = runner_module.run_sweep(
        fixture_dir,
        output_dir,
        mutations=("marker_cell", "style_cell"),
        skip_invalid_source=True,
    )

    assert report["failure_count"] == 0
    assert {result["status"] for result in report["results"]} == {"passed"}


def test_apply_mutation_targets_first_worksheet_when_chartsheet_is_first(
    tmp_path: Path, monkeypatch
) -> None:
    class DummyChartsheet:
        title = "Chart"

    class DummyWorksheet:
        title = "Data"
        max_row = 1
        max_column = 1

        def __init__(self) -> None:
            self.values: dict[str, object] = {}

        def __setitem__(self, key: str, value: object) -> None:
            self.values[key] = value

        def __getitem__(self, key: str):
            return self.values[key]

        def cell(self, row: int, column: int):
            return self.values.setdefault(f"{row}:{column}", object())

    class DummyWorkbook:
        sheetnames = ["Chart", "Data"]

        def __init__(self) -> None:
            self.chart = DummyChartsheet()
            self.worksheet = DummyWorksheet()
            self.saved = False
            self.closed = False

        def __getitem__(self, name: str):
            return {"Chart": self.chart, "Data": self.worksheet}[name]

        def save(self, _path: Path) -> None:
            self.saved = True

        def close(self) -> None:
            self.closed = True

    workbook = DummyWorkbook()

    monkeypatch.setattr(
        runner_module.wolfxl,
        "load_workbook",
        lambda _path, modify: workbook,
    )

    runner_module._apply_mutation(tmp_path / "chartsheet-first.xlsx", "marker_cell")

    assert workbook.worksheet.values[runner_module.MARKER_CELL] == runner_module.MARKER_VALUE
    assert workbook.saved is True
    assert workbook.closed is True


def test_runner_can_discover_recursive_fixture_trees(tmp_path: Path) -> None:
    fixture_dir = tmp_path / "fixtures"
    output_dir = tmp_path / "out"
    nested_dir = fixture_dir / "nested" / "deep"
    nested_dir.mkdir(parents=True)
    _make_fixture(nested_dir / "simple.xlsx")

    report = runner_module.run_sweep(
        fixture_dir,
        output_dir,
        mutations=("no_op",),
        recursive=True,
    )

    assert report["recursive"] is True
    assert report["result_count"] == 1
    result = report["results"][0]
    assert result["fixture"] == "nested/deep/simple.xlsx"
    assert result["status"] == "passed"
    assert (output_dir / "nested_deep_simple" / "no_op" / "after-simple.xlsx").is_file()


def test_discover_fixtures_includes_macro_enabled_ooxml_without_manifest(
    tmp_path: Path,
) -> None:
    fixture_dir = tmp_path / "fixtures"
    fixture_dir.mkdir()
    _make_fixture(fixture_dir / "plain.xlsx")
    _make_fixture(fixture_dir / "macro.xlsm")
    _make_fixture(fixture_dir / "~$lock.xlsx")
    fixture_dir.joinpath("notes.txt").write_text("ignore me")

    entries = runner_module.discover_fixtures(fixture_dir)

    assert [entry.filename for entry in entries] == ["macro.xlsm", "plain.xlsx"]


def test_discover_fixtures_supports_excelbench_files_manifest(tmp_path: Path) -> None:
    fixture_dir = tmp_path / "fixtures"
    tier_dir = fixture_dir / "tier1"
    tier_dir.mkdir(parents=True)
    _make_fixture(tier_dir / "plain.xlsx")
    fixture_dir.joinpath("manifest.json").write_text(
        """{
  "excel_version": "16.105.3",
  "files": [
    {
      "path": "tier1/plain.xlsx",
      "feature": "cell_values",
      "sha256": "abc123"
    }
  ]
}"""
    )

    entries = runner_module.discover_fixtures(fixture_dir, recursive=True)

    assert entries == [
        runner_module.FixtureEntry(
            filename="tier1/plain.xlsx",
            sha256="abc123",
            fixture_id="cell_values",
            tool="excel",
            app_unsupported_features=None,
        )
    ]


def test_discover_fixtures_supports_curated_workbooks_manifest(tmp_path: Path) -> None:
    fixture_dir = tmp_path / "fixtures"
    workbook_dir = fixture_dir / "published"
    workbook_dir.mkdir(parents=True)
    _make_fixture(workbook_dir / "table.xlsx")
    fixture_dir.joinpath("manifest.json").write_text(
        """{
  "workbooks": [
    {
      "workbook_id": "tables_filters",
      "path": "published/table.xlsx",
      "source_type": "synthetic_external",
      "app_unsupported_features": ["powerview"]
    }
  ]
}"""
    )

    entries = runner_module.discover_fixtures(fixture_dir)

    assert entries == [
        runner_module.FixtureEntry(
            filename="published/table.xlsx",
            sha256=None,
            fixture_id="tables_filters",
            tool="synthetic_external",
            app_unsupported_features=["powerview"],
        )
    ]


def test_runner_can_exclude_fixtures_by_glob(tmp_path: Path) -> None:
    fixture_dir = tmp_path / "fixtures"
    output_dir = tmp_path / "out"
    nested_dir = fixture_dir / "nested"
    nested_dir.mkdir(parents=True)
    _make_fixture(fixture_dir / "keep.xlsx")
    _make_fixture(nested_dir / "skip-heavy.xlsx")

    report = runner_module.run_sweep(
        fixture_dir,
        output_dir,
        mutations=("no_op",),
        recursive=True,
        exclude_fixture_patterns=("*skip-heavy.xlsx",),
    )

    assert report["exclude_fixture_patterns"] == ["*skip-heavy.xlsx"]
    assert report["skipped_fixture_count"] == 1
    assert report["skipped_fixtures"] == ["nested/skip-heavy.xlsx"]
    assert report["result_count"] == 1
    assert report["results"][0]["fixture"] == "keep.xlsx"


def test_runner_can_skip_invalid_source_workbooks_for_exploratory_corpora(
    tmp_path: Path,
) -> None:
    fixture_dir = tmp_path / "fixtures"
    output_dir = tmp_path / "out"
    fixture_dir.mkdir()
    invalid = fixture_dir / "bad.xlsx"
    invalid.write_text("not an OOXML zip")

    report = runner_module.run_sweep(
        fixture_dir,
        output_dir,
        mutations=("no_op",),
        skip_invalid_source=True,
    )

    assert report["skip_invalid_source"] is True
    assert report["failure_count"] == 0
    assert report["result_count"] == 1
    result = report["results"][0]
    assert result["fixture"] == "bad.xlsx"
    assert result["status"] == "skipped_source_invalid"
    assert "could not determine file format" in result["error"]


def test_runner_treats_native_invalid_zip_as_skippable_source_error() -> None:
    error = RuntimeError("native xlsx open failed: invalid Zip archive: Could not find EOCD")

    assert runner_module._is_skippable_source_error(error) is True


def test_runner_can_skip_missing_manifest_fixtures_for_exploratory_corpora(
    tmp_path: Path,
) -> None:
    fixture_dir = tmp_path / "fixtures"
    output_dir = tmp_path / "out"
    fixture_dir.mkdir()
    fixture_dir.joinpath("manifest.json").write_text(
        """{
  "workbooks": [
    {
      "workbook_id": "missing_placeholder",
      "path": "missing/placeholder.xlsx"
    }
  ]
}"""
    )

    report = runner_module.run_sweep(
        fixture_dir,
        output_dir,
        mutations=("no_op",),
        skip_invalid_source=True,
    )

    assert report["failure_count"] == 0
    assert report["result_count"] == 1
    result = report["results"][0]
    assert result["fixture"] == "missing/placeholder.xlsx"
    assert result["status"] == "skipped_source_missing"
    assert "fixture missing:" in result["error"]


def test_runner_supports_add_remove_chart_mutation(tmp_path: Path) -> None:
    fixture_dir = tmp_path / "fixtures"
    output_dir = tmp_path / "out"
    fixture_dir.mkdir()
    fixture = fixture_dir / "simple.xlsx"
    _make_fixture(fixture)

    report = runner_module.run_sweep(
        fixture_dir,
        output_dir,
        mutations=("add_remove_chart",),
    )

    assert report["failure_count"] == 0
    assert report["results"][0]["status"] == "passed"


def test_runner_requires_formula_move_translation_drift(tmp_path: Path) -> None:
    fixture_dir = tmp_path / "fixtures"
    output_dir = tmp_path / "out"
    fixture_dir.mkdir()
    fixture = fixture_dir / "simple.xlsx"
    _make_fixture(fixture)

    report = runner_module.run_sweep(
        fixture_dir,
        output_dir,
        mutations=("move_formula_range",),
    )

    assert report["failure_count"] == 0
    result = report["results"][0]
    assert result["status"] == "passed_with_expected_drift"
    assert result["issue_count"] == 0
    assert result["expected_issue_count"] == 1
    assert result["expected_issues"][0]["kind"] == "worksheet_formulas_semantic_drift"
    assert "Z2" in result["expected_issues"][0]["message"]


def test_runner_retargets_external_links_with_required_drift(tmp_path: Path) -> None:
    fixture_dir = tmp_path / "fixtures"
    output_dir = tmp_path / "out"
    fixture_dir.mkdir()
    fixture = fixture_dir / "external_links_basic.xlsx"
    fixture.write_bytes(
        (Path(__file__).parent / "fixtures" / "external_links_basic.xlsx").read_bytes()
    )

    report = runner_module.run_sweep(
        fixture_dir,
        output_dir,
        mutations=("retarget_external_links",),
    )

    assert report["failure_count"] == 0
    result = report["results"][0]
    assert result["status"] == "passed_with_expected_drift"
    assert result["issue_count"] == 0
    assert result["expected_issue_count"] == 3
    assert {issue["kind"] for issue in result["expected_issues"]} == {
        "external_links_semantic_drift",
        "missing_relationship",
        "worksheet_formulas_semantic_drift",
    }
    semantic_drift = [
        issue
        for issue in result["expected_issues"]
        if issue["kind"] == "external_links_semantic_drift"
    ][0]
    assert "wolfxl-retargeted-external-link.xlsx" in semantic_drift["message"]

    with zipfile.ZipFile(result["after"]) as archive:
        rels = archive.read("xl/externalLinks/_rels/externalLink1.xml.rels").decode("utf-8")
    assert "wolfxl-retargeted-external-link.xlsx" in rels


def test_runner_requires_external_link_retarget_drift_when_links_exist(
    tmp_path: Path, monkeypatch
) -> None:
    fixture_dir = tmp_path / "fixtures"
    output_dir = tmp_path / "out"
    fixture_dir.mkdir()
    fixture = fixture_dir / "external_links_basic.xlsx"
    fixture.write_bytes(
        (Path(__file__).parent / "fixtures" / "external_links_basic.xlsx").read_bytes()
    )

    def fake_audit(_before: Path, _after: Path) -> dict:
        return {"issues": []}

    monkeypatch.setattr(runner_module.audit_ooxml_fidelity, "audit", fake_audit)

    report = runner_module.run_sweep(
        fixture_dir,
        output_dir,
        mutations=("retarget_external_links",),
    )

    assert report["failure_count"] == 1
    result = report["results"][0]
    assert result["status"] == "failed"
    assert result["issues"][0]["kind"] == "missing_required_expected_drift"


def test_runner_rejects_unrelated_external_link_relationship_loss_for_retarget(
    tmp_path: Path, monkeypatch
) -> None:
    fixture_dir = tmp_path / "fixtures"
    output_dir = tmp_path / "out"
    fixture_dir.mkdir()
    fixture = fixture_dir / "external_links_basic.xlsx"
    fixture.write_bytes(
        (Path(__file__).parent / "fixtures" / "external_links_basic.xlsx").read_bytes()
    )

    def fake_audit(_before: Path, _after: Path) -> dict:
        return {
            "issues": [
                {
                    "kind": "external_links_semantic_drift",
                    "severity": "error",
                    "part": "external_links",
                    "message": (
                        "external_links semantic fingerprint changed after save: "
                        "wolfxl-retargeted-external-link.xlsx"
                    ),
                },
                {
                    "kind": "worksheet_formulas_semantic_drift",
                    "severity": "error",
                    "part": "worksheet_formulas",
                    "message": (
                        "worksheet_formulas semantic fingerprint changed after save: "
                        "wolfxl-retargeted-external-link.xlsx"
                    ),
                },
                {
                    "kind": "missing_relationship",
                    "severity": "error",
                    "part": "xl/externalLinks/_rels/externalLink1.xml.rels",
                    "message": (
                        "Relationship existed before save and is missing after save: "
                        "xl/externalLinks/_rels/externalLink1.xml.rels rId1 "
                        "http://schemas.openxmlformats.org/officeDocument/2006/"
                        "relationships/externalLinkPath -> ext.xlsx"
                    ),
                },
                {
                    "kind": "missing_relationship",
                    "severity": "error",
                    "part": "xl/externalLinks/_rels/externalLink1.xml.rels",
                    "message": (
                        "Relationship existed before save and is missing after save: "
                        "xl/externalLinks/_rels/externalLink1.xml.rels rId2 "
                        "http://schemas.openxmlformats.org/officeDocument/2006/"
                        "relationships/externalLinkPath -> unrelated.xlsx"
                    ),
                },
            ]
        }

    monkeypatch.setattr(runner_module.audit_ooxml_fidelity, "audit", fake_audit)

    report = runner_module.run_sweep(
        fixture_dir,
        output_dir,
        mutations=("retarget_external_links",),
    )

    assert report["failure_count"] == 1
    result = report["results"][0]
    assert result["status"] == "failed"
    assert [issue["kind"] for issue in result["expected_issues"]] == [
        "external_links_semantic_drift",
        "worksheet_formulas_semantic_drift",
        "missing_relationship",
    ]
    assert [issue["message"] for issue in result["issues"]] == [
        (
            "Relationship existed before save and is missing after save: "
            "xl/externalLinks/_rels/externalLink1.xml.rels rId2 "
            "http://schemas.openxmlformats.org/officeDocument/2006/"
            "relationships/externalLinkPath -> unrelated.xlsx"
        )
    ]


def test_runner_reports_manifest_hash_mismatch(tmp_path: Path) -> None:
    fixture_dir = tmp_path / "fixtures"
    output_dir = tmp_path / "out"
    fixture_dir.mkdir()
    fixture = fixture_dir / "simple.xlsx"
    _make_fixture(fixture)
    (fixture_dir / "manifest.json").write_text(
        """
{
  "fixtures": [
    {"filename": "simple.xlsx", "sha256": "not-the-real-hash"}
  ]
}
""".strip()
    )

    report = runner_module.run_sweep(
        fixture_dir,
        output_dir,
        mutations=("no_op",),
    )

    assert report["failure_count"] == 1
    assert report["results"][0]["status"] == "hash_mismatch"


def test_runner_separates_expected_rename_drift(tmp_path: Path, monkeypatch) -> None:
    fixture_dir = tmp_path / "fixtures"
    output_dir = tmp_path / "out"
    fixture_dir.mkdir()
    fixture = fixture_dir / "simple.xlsx"
    _make_fixture(fixture)

    def fake_audit(_before: Path, _after: Path) -> dict:
        return {
            "issues": [
                {
                    "kind": "charts_semantic_drift",
                    "severity": "error",
                    "part": "charts",
                    "message": "expected formula change after sheet rename",
                },
                {
                    "kind": "workbook_globals_semantic_drift",
                    "severity": "error",
                    "part": "workbook_globals",
                    "message": "expected defined-name formula change after sheet rename",
                }
            ]
        }

    monkeypatch.setattr(runner_module.audit_ooxml_fidelity, "audit", fake_audit)

    report = runner_module.run_sweep(
        fixture_dir,
        output_dir,
        mutations=("rename_first_sheet",),
    )

    assert report["failure_count"] == 0
    result = report["results"][0]
    assert result["status"] == "passed_with_expected_drift"
    assert result["issue_count"] == 0
    assert result["expected_issue_count"] == 2
    assert {issue["kind"] for issue in result["expected_issues"]} == {
        "charts_semantic_drift",
        "workbook_globals_semantic_drift",
    }


def test_rename_first_sheet_rewrites_existing_chart_formula_refs(tmp_path: Path) -> None:
    fixture_dir = tmp_path / "fixtures"
    output_dir = tmp_path / "out"
    fixture_dir.mkdir()
    fixture = fixture_dir / "excelize-chart-points-formula-cf.xlsx"
    source = (
        Path(__file__).resolve().parents[1]
        / "tests"
        / "fixtures"
        / "external_oracle"
        / fixture.name
    )
    shutil.copy2(source, fixture)

    report = runner_module.run_sweep(
        fixture_dir,
        output_dir,
        mutations=("rename_first_sheet",),
    )

    assert report["failure_count"] == 0
    after = (
        output_dir
        / "excelize-chart-points-formula-cf"
        / "rename_first_sheet"
        / "after-excelize-chart-points-formula-cf.xlsx"
    )
    with zipfile.ZipFile(after) as archive:
        chart_xml = archive.read("xl/charts/chart1.xml").decode()

    assert "Metrics!" not in chart_xml
    assert "&apos;WolfXL Fidelity Rename&apos;!$B$1" in chart_xml
    assert "&apos;WolfXL Fidelity Rename&apos;!$A$2:$A$4" in chart_xml


def test_rename_first_sheet_rewrites_chart_pivot_source_sheet_refs(
    tmp_path: Path,
) -> None:
    fixture_dir = tmp_path / "fixtures"
    output_dir = tmp_path / "out"
    fixture_dir.mkdir()
    fixture = fixture_dir / "real-excel-pivot-chart-slicers.xlsx"
    source = (
        Path(__file__).resolve().parents[1]
        / "tests"
        / "fixtures"
        / "external_oracle"
        / fixture.name
    )
    shutil.copy2(source, fixture)

    report = runner_module.run_sweep(
        fixture_dir,
        output_dir,
        mutations=("rename_first_sheet",),
    )

    assert report["failure_count"] == 0
    after = (
        output_dir
        / "real-excel-pivot-chart-slicers"
        / "rename_first_sheet"
        / "after-real-excel-pivot-chart-slicers.xlsx"
    )
    with zipfile.ZipFile(after) as archive:
        chart_xml = archive.read("xl/charts/chart1.xml").decode()

    assert "Pivot Table!PivotTable9" not in chart_xml
    assert "WolfXL Fidelity Rename!PivotTable9" in chart_xml


def test_rename_first_sheet_rewrites_workbook_defined_name_sheet_refs(
    tmp_path: Path,
) -> None:
    fixture_dir = tmp_path / "fixtures"
    output_dir = tmp_path / "out"
    fixture_dir.mkdir()
    fixture = fixture_dir / "openpyxl-p1-structured-validation-protection-customxml.xlsx"
    source = (
        Path(__file__).resolve().parents[1]
        / "tests"
        / "fixtures"
        / "external_oracle"
        / fixture.name
    )
    shutil.copy2(source, fixture)

    report = runner_module.run_sweep(
        fixture_dir,
        output_dir,
        mutations=("rename_first_sheet",),
    )

    assert report["failure_count"] == 0
    after = (
        output_dir
        / "openpyxl-p1-structured-validation-protection-customxml"
        / "rename_first_sheet"
        / "after-openpyxl-p1-structured-validation-protection-customxml.xlsx"
    )
    with zipfile.ZipFile(after) as archive:
        workbook_xml = archive.read("xl/workbook.xml").decode()

    assert "'P1 Surface'!$A$1:$D$4" not in workbook_xml
    assert "&apos;WolfXL Fidelity Rename&apos;!$A$1:$D$4" in workbook_xml


def test_rename_first_sheet_rewrites_pivot_cache_worksheet_source_sheet_refs(
    tmp_path: Path,
) -> None:
    fixture_dir = tmp_path / "fixtures"
    output_dir = tmp_path / "out"
    fixture_dir.mkdir()
    fixture = fixture_dir / "excelize-2.10-pivot-slicers.xlsx"
    source = (
        Path(__file__).resolve().parents[1]
        / "tests"
        / "fixtures"
        / "external_oracle"
        / fixture.name
    )
    shutil.copy2(source, fixture)

    report = runner_module.run_sweep(
        fixture_dir,
        output_dir,
        mutations=("rename_first_sheet",),
    )

    assert report["failure_count"] == 0
    after = (
        output_dir
        / "excelize-2.10-pivot-slicers"
        / "rename_first_sheet"
        / "after-excelize-2.10-pivot-slicers.xlsx"
    )
    with zipfile.ZipFile(after) as archive:
        cache_xml = archive.read("xl/pivotCache/pivotCacheDefinition1.xml").decode()

    assert 'sheet="Data"' not in cache_xml
    assert 'sheet="WolfXL Fidelity Rename"' in cache_xml


def test_rename_first_sheet_rewrites_existing_worksheet_formula_sheet_refs(
    tmp_path: Path,
) -> None:
    fixture_dir = tmp_path / "fixtures"
    output_dir = tmp_path / "out"
    fixture_dir.mkdir()
    fixture = fixture_dir / "worksheet-formulas.xlsx"
    workbook = openpyxl.Workbook()
    overview = workbook.active
    overview.title = "Overview"
    overview["A1"] = 42
    calc = workbook.create_sheet("Calc")
    calc["A1"] = "=Overview!$A$1"
    calc["A2"] = '="Overview!"'
    calc["A3"] = "=VLOOKUP(B2,[1]Overview!A:D,2,0)"
    workbook.save(fixture)

    report = runner_module.run_sweep(
        fixture_dir,
        output_dir,
        mutations=("rename_first_sheet",),
    )

    assert report["failure_count"] == 0
    after = (
        output_dir
        / "worksheet-formulas"
        / "rename_first_sheet"
        / "after-worksheet-formulas.xlsx"
    )
    with zipfile.ZipFile(after) as archive:
        sheet_xml = archive.read("xl/worksheets/sheet2.xml").decode()

    assert "Overview!$A$1" not in sheet_xml
    assert "&apos;WolfXL Fidelity Rename&apos;!$A$1" in sheet_xml
    assert "Overview!" in sheet_xml
    assert "[1]Overview!A:D" in sheet_xml
    assert "[1]&apos;WolfXL Fidelity Rename&apos;!A:D" not in sheet_xml


def test_runner_separates_expected_interior_delete_drift(tmp_path: Path, monkeypatch) -> None:
    fixture_dir = tmp_path / "fixtures"
    output_dir = tmp_path / "out"
    fixture_dir.mkdir()
    fixture = fixture_dir / "simple.xlsx"
    _make_fixture(fixture)

    def fake_audit(_before: Path, _after: Path) -> dict:
        return {
            "issues": [
                {
                    "kind": "conditional_formatting_semantic_drift",
                    "severity": "error",
                    "part": "conditional_formatting",
                    "message": "expected range change after row delete",
                },
                {
                    "kind": "data_validations_semantic_drift",
                    "severity": "error",
                    "part": "data_validations",
                    "message": "expected validation range change after row delete",
                },
            ]
        }

    monkeypatch.setattr(runner_module.audit_ooxml_fidelity, "audit", fake_audit)

    report = runner_module.run_sweep(
        fixture_dir,
        output_dir,
        mutations=("delete_first_row",),
    )

    assert report["failure_count"] == 0
    result = report["results"][0]
    assert result["status"] == "passed_with_expected_drift"
    assert result["issue_count"] == 0
    assert result["expected_issue_count"] == 2
    assert {issue["kind"] for issue in result["expected_issues"]} == {
        "conditional_formatting_semantic_drift",
        "data_validations_semantic_drift",
    }


def test_runner_separates_expected_sheet_copy_drift(tmp_path: Path, monkeypatch) -> None:
    fixture_dir = tmp_path / "fixtures"
    output_dir = tmp_path / "out"
    fixture_dir.mkdir()
    fixture = fixture_dir / "simple.xlsx"
    _make_fixture(fixture)

    def fake_audit(_before: Path, _after: Path) -> dict:
        return {
            "issues": [
                {
                    "kind": "charts_semantic_drift",
                    "severity": "error",
                    "part": "charts",
                    "message": "expected copied chart part",
                },
                {
                    "kind": "chart_styles_semantic_drift",
                    "severity": "error",
                    "part": "chart_styles",
                    "message": "expected copied chart style sidecar",
                },
                {
                    "kind": "slicers_semantic_drift",
                    "severity": "error",
                    "part": "slicers",
                    "message": "expected copied slicer part",
                },
                {
                    "kind": "timelines_semantic_drift",
                    "severity": "error",
                    "part": "timelines",
                    "message": "expected copied timeline part",
                },
                {
                    "kind": "data_validations_semantic_drift",
                    "severity": "error",
                    "part": "data_validations",
                    "message": "expected copied data validation part",
                },
                {
                    "kind": "page_setup_semantic_drift",
                    "severity": "error",
                    "part": "page_setup",
                    "message": "expected copied page setup",
                },
                {
                    "kind": "structured_references_semantic_drift",
                    "severity": "error",
                    "part": "structured_references",
                    "message": "expected copied structured references",
                },
                {
                    "kind": "workbook_globals_semantic_drift",
                    "severity": "error",
                    "part": "workbook_globals",
                    "message": "expected copied workbook globals",
                },
            ]
        }

    monkeypatch.setattr(runner_module.audit_ooxml_fidelity, "audit", fake_audit)

    report = runner_module.run_sweep(
        fixture_dir,
        output_dir,
        mutations=("copy_first_sheet",),
    )

    assert report["failure_count"] == 0
    result = report["results"][0]
    assert result["status"] == "passed_with_expected_drift"
    assert result["issue_count"] == 0
    assert result["expected_issue_count"] == 8
    assert {issue["kind"] for issue in result["expected_issues"]} == {
        "chart_styles_semantic_drift",
        "charts_semantic_drift",
        "data_validations_semantic_drift",
        "page_setup_semantic_drift",
        "slicers_semantic_drift",
        "structured_references_semantic_drift",
        "timelines_semantic_drift",
        "workbook_globals_semantic_drift",
    }


def test_runner_accepts_structural_external_link_formula_drift_only(
    tmp_path: Path, monkeypatch
) -> None:
    fixture_dir = tmp_path / "fixtures"
    output_dir = tmp_path / "out"
    fixture_dir.mkdir()
    fixture = fixture_dir / "simple.xlsx"
    _make_fixture(fixture)

    def fake_audit(_before: Path, _after: Path) -> dict:
        return {
            "issues": [
                {
                    "kind": "external_links_semantic_drift",
                    "severity": "error",
                    "part": "external_links",
                    "message": "expected worksheet_formulas drift after sheet copy",
                },
                {
                    "kind": "external_links_semantic_drift",
                    "severity": "error",
                    "part": "external_links",
                    "message": "unexpected target changed from a.xlsx to b.xlsx",
                },
            ]
        }

    monkeypatch.setattr(runner_module.audit_ooxml_fidelity, "audit", fake_audit)

    report = runner_module.run_sweep(
        fixture_dir,
        output_dir,
        mutations=("copy_first_sheet",),
    )

    assert report["failure_count"] == 1
    result = report["results"][0]
    assert result["status"] == "failed"
    assert [issue["message"] for issue in result["expected_issues"]] == [
        "expected worksheet_formulas drift after sheet copy",
    ]
    assert [issue["message"] for issue in result["issues"]] == [
        "unexpected target changed from a.xlsx to b.xlsx",
    ]


def test_runner_accepts_structural_delete_calc_chain_volatility_only(
    tmp_path: Path, monkeypatch
) -> None:
    fixture_dir = tmp_path / "fixtures"
    output_dir = tmp_path / "out"
    fixture_dir.mkdir()
    fixture = fixture_dir / "simple.xlsx"
    _make_fixture(fixture)

    def fake_audit(_before: Path, _after: Path) -> dict:
        return {
            "issues": [
                {
                    "kind": "missing_part",
                    "severity": "error",
                    "part": "xl/calcChain.xml",
                    "message": "missing part after save: xl/calcChain.xml",
                },
                {
                    "kind": "missing_relationship",
                    "severity": "error",
                    "part": "xl/_rels/workbook.xml.rels",
                    "message": (
                        "missing relationship after save: "
                        "xl/_rels/workbook.xml.rels rId6 "
                        "http://schemas.openxmlformats.org/officeDocument/2006/"
                        "relationships/calcChain -> calcChain.xml"
                    ),
                },
                {
                    "kind": "feature_part_loss",
                    "severity": "error",
                    "part": "calc_chain",
                    "message": "calc_chain parts disappeared after save: ['xl/calcChain.xml']",
                },
                {
                    "kind": "missing_part",
                    "severity": "error",
                    "part": "xl/charts/chart1.xml",
                    "message": "missing part after save: xl/charts/chart1.xml",
                },
            ]
        }

    monkeypatch.setattr(runner_module.audit_ooxml_fidelity, "audit", fake_audit)

    report = runner_module.run_sweep(
        fixture_dir,
        output_dir,
        mutations=("delete_first_row",),
    )

    assert report["failure_count"] == 1
    result = report["results"][0]
    assert result["status"] == "failed"
    assert [issue["message"] for issue in result["expected_issues"]] == [
        "missing part after save: xl/calcChain.xml",
        (
            "missing relationship after save: xl/_rels/workbook.xml.rels rId6 "
            "http://schemas.openxmlformats.org/officeDocument/2006/"
            "relationships/calcChain -> calcChain.xml"
        ),
        "calc_chain parts disappeared after save: ['xl/calcChain.xml']",
    ]
    assert [issue["message"] for issue in result["issues"]] == [
        "missing part after save: xl/charts/chart1.xml",
    ]


def test_runner_accepts_structural_delete_semantic_drifts(tmp_path: Path, monkeypatch) -> None:
    fixture_dir = tmp_path / "fixtures"
    output_dir = tmp_path / "out"
    fixture_dir.mkdir()
    fixture = fixture_dir / "simple.xlsx"
    _make_fixture(fixture)

    def fake_audit(_before: Path, _after: Path) -> dict:
        return {
            "issues": [
                {
                    "kind": "extensions_semantic_drift",
                    "severity": "error",
                    "part": "extensions",
                    "message": "extensions semantic fingerprint changed: before={'sqref':'C2:C4'} after={'sqref':'C1:C3'}",
                },
                {
                    "kind": "drawing_objects_semantic_drift",
                    "severity": "error",
                    "part": "drawing_objects",
                    "message": "comment anchor changed: before={'ref':'C2'} after={'ref':'C1'}",
                },
                {
                    "kind": "structured_references_semantic_drift",
                    "severity": "error",
                    "part": "structured_references",
                    "message": "formula moved: before={'r':'F2'} after={'r':'E1'}",
                },
                {
                    "kind": "workbook_globals_semantic_drift",
                    "severity": "error",
                    "part": "workbook_globals",
                    "message": "defined name shifted: before={'A1:D4'} after={'A1:C3'}",
                },
            ]
        }

    monkeypatch.setattr(runner_module.audit_ooxml_fidelity, "audit", fake_audit)

    report = runner_module.run_sweep(
        fixture_dir,
        output_dir,
        mutations=("delete_first_row",),
    )

    assert report["failure_count"] == 0
    result = report["results"][0]
    assert result["status"] == "passed_with_expected_drift"
    assert result["issue_count"] == 0
    assert {issue["kind"] for issue in result["expected_issues"]} == {
        "drawing_objects_semantic_drift",
        "extensions_semantic_drift",
        "structured_references_semantic_drift",
        "workbook_globals_semantic_drift",
    }


def test_runner_accepts_tail_mutation_structural_drawing_anchor_drift(
    tmp_path: Path, monkeypatch
) -> None:
    fixture_dir = tmp_path / "fixtures"
    output_dir = tmp_path / "out"
    fixture_dir.mkdir()
    fixture = fixture_dir / "simple.xlsx"
    _make_fixture(fixture)

    def fake_audit(_before: Path, _after: Path) -> dict:
        return {
            "issues": [
                {
                    "kind": "drawing_objects_semantic_drift",
                    "severity": "error",
                    "part": "drawing_objects",
                    "message": "drawing anchor row/col shifted after tail mutation",
                }
            ]
        }

    monkeypatch.setattr(runner_module.audit_ooxml_fidelity, "audit", fake_audit)
    monkeypatch.setattr(
        runner_module,
        "_drawing_objects_match_except_structural_anchor_text",
        lambda _before, _after: True,
    )

    report = runner_module.run_sweep(
        fixture_dir,
        output_dir,
        mutations=("insert_tail_col",),
    )

    assert report["failure_count"] == 0
    result = report["results"][0]
    assert result["status"] == "passed_with_expected_drift"
    assert result["issue_count"] == 0
    assert result["expected_issues"][0]["kind"] == "drawing_objects_semantic_drift"


def test_runner_does_not_hide_tail_mutation_non_anchor_drawing_drift(
    tmp_path: Path, monkeypatch
) -> None:
    fixture_dir = tmp_path / "fixtures"
    output_dir = tmp_path / "out"
    fixture_dir.mkdir()
    fixture = fixture_dir / "simple.xlsx"
    _make_fixture(fixture)

    def fake_audit(_before: Path, _after: Path) -> dict:
        return {
            "issues": [
                {
                    "kind": "drawing_objects_semantic_drift",
                    "severity": "error",
                    "part": "drawing_objects",
                    "message": "drawing relationship disappeared after tail mutation",
                }
            ]
        }

    monkeypatch.setattr(runner_module.audit_ooxml_fidelity, "audit", fake_audit)
    monkeypatch.setattr(
        runner_module,
        "_drawing_objects_match_except_structural_anchor_text",
        lambda _before, _after: False,
    )

    report = runner_module.run_sweep(
        fixture_dir,
        output_dir,
        mutations=("insert_tail_col",),
    )

    assert report["failure_count"] == 1
    result = report["results"][0]
    assert result["status"] == "failed"
    assert result["issue_count"] == 1
    assert result["issues"][0]["kind"] == "drawing_objects_semantic_drift"


def test_runner_does_not_hide_structural_delete_total_feature_loss(
    tmp_path: Path, monkeypatch
) -> None:
    fixture_dir = tmp_path / "fixtures"
    output_dir = tmp_path / "out"
    fixture_dir.mkdir()
    fixture = fixture_dir / "simple.xlsx"
    _make_fixture(fixture)

    def fake_audit(_before: Path, _after: Path) -> dict:
        return {
            "issues": [
                {
                    "kind": "extensions_semantic_drift",
                    "severity": "error",
                    "part": "extensions",
                    "message": "extensions semantic fingerprint changed: before={'ext':'present'} after={}",
                }
            ]
        }

    monkeypatch.setattr(runner_module.audit_ooxml_fidelity, "audit", fake_audit)

    report = runner_module.run_sweep(
        fixture_dir,
        output_dir,
        mutations=("delete_first_row",),
    )

    assert report["failure_count"] == 1
    result = report["results"][0]
    assert result["status"] == "failed"
    assert result["expected_issue_count"] == 0
    assert result["issues"][0]["kind"] == "extensions_semantic_drift"


def test_runner_accepts_deleted_first_axis_formula_loss(tmp_path: Path, monkeypatch) -> None:
    fixture_dir = tmp_path / "fixtures"
    output_dir = tmp_path / "out"
    fixture_dir.mkdir()
    fixture = fixture_dir / "simple.xlsx"
    _make_fixture(fixture)

    def fake_audit(_before: Path, _after: Path) -> dict:
        return {
            "issues": [
                {
                    "kind": "worksheet_formulas_semantic_drift",
                    "severity": "error",
                    "part": "worksheet_formulas",
                    "message": (
                        "worksheet_formulas semantic fingerprint changed after save: "
                        "before={'xl/worksheets/sheet1.xml': [((('r', 'B1'),), "
                        "(('t', None),), '[1]Sheet1!$A$1')]} after={}"
                    ),
                }
            ]
        }

    monkeypatch.setattr(runner_module.audit_ooxml_fidelity, "audit", fake_audit)

    report = runner_module.run_sweep(
        fixture_dir,
        output_dir,
        mutations=("delete_first_row",),
    )

    assert report["failure_count"] == 0
    result = report["results"][0]
    assert result["status"] == "passed_with_expected_drift"
    assert result["issue_count"] == 0
    assert result["expected_issues"][0]["kind"] == "worksheet_formulas_semantic_drift"


def test_runner_does_not_hide_non_deleted_axis_formula_loss(tmp_path: Path, monkeypatch) -> None:
    fixture_dir = tmp_path / "fixtures"
    output_dir = tmp_path / "out"
    fixture_dir.mkdir()
    fixture = fixture_dir / "simple.xlsx"
    _make_fixture(fixture)

    def fake_audit(_before: Path, _after: Path) -> dict:
        return {
            "issues": [
                {
                    "kind": "worksheet_formulas_semantic_drift",
                    "severity": "error",
                    "part": "worksheet_formulas",
                    "message": (
                        "worksheet_formulas semantic fingerprint changed after save: "
                        "before={'xl/worksheets/sheet1.xml': [((('r', 'B2'),), "
                        "(('t', None),), 'SUM(A2:A3)')]} after={}"
                    ),
                }
            ]
        }

    monkeypatch.setattr(runner_module.audit_ooxml_fidelity, "audit", fake_audit)

    report = runner_module.run_sweep(
        fixture_dir,
        output_dir,
        mutations=("delete_first_row",),
    )

    assert report["failure_count"] == 1
    result = report["results"][0]
    assert result["status"] == "failed"
    assert result["expected_issue_count"] == 0
    assert result["issues"][0]["kind"] == "worksheet_formulas_semantic_drift"


def test_runner_separates_expected_feature_add_drift(tmp_path: Path, monkeypatch) -> None:
    fixture_dir = tmp_path / "fixtures"
    output_dir = tmp_path / "out"
    fixture_dir.mkdir()
    fixture = fixture_dir / "simple.xlsx"
    _make_fixture(fixture)

    def fake_audit(_before: Path, after: Path) -> dict:
        if "add_data_validation" in str(after):
            issue = {
                "kind": "data_validations_semantic_drift",
                "severity": "error",
                "part": "data_validations",
                "message": "expected added data validation at AB2:AB10",
            }
        else:
            issue = {
                "kind": "conditional_formatting_semantic_drift",
                "severity": "error",
                "part": "conditional_formatting",
                "message": "expected added conditional format at AC2:AC10",
            }
        return {"issues": [issue]}

    monkeypatch.setattr(runner_module.audit_ooxml_fidelity, "audit", fake_audit)

    report = runner_module.run_sweep(
        fixture_dir,
        output_dir,
        mutations=("add_data_validation", "add_conditional_formatting"),
    )

    assert report["failure_count"] == 0
    statuses = {result["mutation"]: result["status"] for result in report["results"]}
    assert statuses == {
        "add_data_validation": "passed_with_expected_drift",
        "add_conditional_formatting": "passed_with_expected_drift",
    }


def test_runner_separates_expected_style_cell_drift(tmp_path: Path, monkeypatch) -> None:
    fixture_dir = tmp_path / "fixtures"
    output_dir = tmp_path / "out"
    fixture_dir.mkdir()
    fixture = fixture_dir / "simple.xlsx"
    _make_fixture(fixture)

    def fake_audit(_before: Path, _after: Path) -> dict:
        return {
            "issues": [
                {
                    "kind": "style_theme_semantic_drift",
                    "severity": "error",
                    "part": "style_theme",
                    "message": "expected style mutation contains color FF1F4E79",
                }
            ]
        }

    monkeypatch.setattr(runner_module.audit_ooxml_fidelity, "audit", fake_audit)

    report = runner_module.run_sweep(
        fixture_dir,
        output_dir,
        mutations=("style_cell",),
    )

    assert report["failure_count"] == 0
    result = report["results"][0]
    assert result["status"] == "passed_with_expected_drift"
    assert result["issue_count"] == 0
    assert result["expected_issue_count"] == 1
    assert result["expected_issues"][0]["kind"] == "style_theme_semantic_drift"


def test_runner_does_not_hide_style_loss_drift(tmp_path: Path, monkeypatch) -> None:
    fixture_dir = tmp_path / "fixtures"
    output_dir = tmp_path / "out"
    fixture_dir.mkdir()
    fixture = fixture_dir / "simple.xlsx"
    _make_fixture(fixture)

    def fake_audit(_before: Path, _after: Path) -> dict:
        return {
            "issues": [
                {
                    "kind": "style_theme_semantic_drift",
                    "severity": "error",
                    "part": "style_theme",
                    "message": "before had custom style after={}",
                }
            ]
        }

    monkeypatch.setattr(runner_module.audit_ooxml_fidelity, "audit", fake_audit)

    report = runner_module.run_sweep(
        fixture_dir,
        output_dir,
        mutations=("style_cell",),
    )

    assert report["failure_count"] == 1
    result = report["results"][0]
    assert result["status"] == "failed"
    assert result["expected_issue_count"] == 0
    assert result["issues"][0]["kind"] == "style_theme_semantic_drift"


def test_runner_does_not_hide_feature_add_loss_drift(tmp_path: Path, monkeypatch) -> None:
    fixture_dir = tmp_path / "fixtures"
    output_dir = tmp_path / "out"
    fixture_dir.mkdir()
    fixture = fixture_dir / "simple.xlsx"
    _make_fixture(fixture)

    def fake_audit(_before: Path, _after: Path) -> dict:
        return {
            "issues": [
                {
                    "kind": "conditional_formatting_semantic_drift",
                    "severity": "error",
                    "part": "conditional_formatting",
                    "message": "before had conditional formatting after={}",
                }
            ]
        }

    monkeypatch.setattr(runner_module.audit_ooxml_fidelity, "audit", fake_audit)

    report = runner_module.run_sweep(
        fixture_dir,
        output_dir,
        mutations=("add_conditional_formatting",),
    )

    assert report["failure_count"] == 1
    assert report["results"][0]["status"] == "failed"


def test_runner_allows_add_conditional_formatting_dxf_style_drift(
    tmp_path: Path, monkeypatch
) -> None:
    fixture_dir = tmp_path / "fixtures"
    output_dir = tmp_path / "out"
    fixture_dir.mkdir()
    fixture = fixture_dir / "simple.xlsx"
    _make_fixture(fixture)

    def fake_audit(_before: Path, _after: Path) -> dict:
        return {
            "issues": [
                {
                    "kind": "conditional_formatting_semantic_drift",
                    "severity": "error",
                    "part": "conditional_formatting",
                    "message": "expected added conditional format at AC2:AC10",
                },
                {
                    "kind": "style_theme_semantic_drift",
                    "severity": "error",
                    "part": "style_theme",
                    "message": "expected additive dxfs drift",
                },
            ]
        }

    monkeypatch.setattr(runner_module.audit_ooxml_fidelity, "audit", fake_audit)

    report = runner_module.run_sweep(
        fixture_dir,
        output_dir,
        mutations=("add_conditional_formatting",),
    )

    assert report["failure_count"] == 0
    result = report["results"][0]
    assert result["status"] == "passed_with_expected_drift"
    assert result["issue_count"] == 0
    assert {issue["kind"] for issue in result["expected_issues"]} == {
        "conditional_formatting_semantic_drift",
        "style_theme_semantic_drift",
    }


def test_runner_does_not_hide_add_conditional_formatting_style_loss(
    tmp_path: Path, monkeypatch
) -> None:
    fixture_dir = tmp_path / "fixtures"
    output_dir = tmp_path / "out"
    fixture_dir.mkdir()
    fixture = fixture_dir / "simple.xlsx"
    _make_fixture(fixture)

    def fake_audit(_before: Path, after: Path) -> dict:
        _remove_style_font(after)
        return {
            "issues": [
                {
                    "kind": "style_theme_semantic_drift",
                    "severity": "error",
                    "part": "style_theme",
                    "message": "before had custom style after={}",
                },
            ]
        }

    def _remove_style_font(path: Path) -> None:
        tmp_path = path.with_suffix(".rewritten.xlsx")
        with zipfile.ZipFile(path) as source_archive:
            root = runner_module._read_styles_root(source_archive)
            assert root is not None
            fonts = runner_module._first_child_by_local_name(root, "fonts")
            assert fonts is not None
            fonts.clear()
            with zipfile.ZipFile(tmp_path, "w") as target_archive:
                for entry in source_archive.infolist():
                    if entry.filename == "xl/styles.xml":
                        target_archive.writestr(entry, runner_module.ElementTree.tostring(root))
                    else:
                        target_archive.writestr(entry, source_archive.read(entry.filename))
        tmp_path.replace(path)

    monkeypatch.setattr(runner_module.audit_ooxml_fidelity, "audit", fake_audit)

    report = runner_module.run_sweep(
        fixture_dir,
        output_dir,
        mutations=("add_conditional_formatting",),
    )

    assert report["failure_count"] == 1
    result = report["results"][0]
    assert result["status"] == "failed"
    assert result["expected_issue_count"] == 0
    assert result["issues"][0]["kind"] == "style_theme_semantic_drift"


def test_runner_does_not_hide_missing_formula_translation_drift(
    tmp_path: Path, monkeypatch
) -> None:
    fixture_dir = tmp_path / "fixtures"
    output_dir = tmp_path / "out"
    fixture_dir.mkdir()
    fixture = fixture_dir / "simple.xlsx"
    _make_fixture(fixture)

    def fake_audit(_before: Path, _after: Path) -> dict:
        return {"issues": []}

    monkeypatch.setattr(runner_module.audit_ooxml_fidelity, "audit", fake_audit)

    report = runner_module.run_sweep(
        fixture_dir,
        output_dir,
        mutations=("move_formula_range",),
    )

    assert report["failure_count"] == 1
    result = report["results"][0]
    assert result["status"] == "failed"
    assert result["issues"][0]["kind"] == "missing_required_expected_drift"
