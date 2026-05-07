"""Pre-release modify-save preservation test over the pinned external-oracle fixture pack.

The pack lives in `tests/fixtures/external_oracle/` and contains workbooks
authored by Excelize, ClosedXML, NPOI, ExcelJS, and Apache POI — features
that openpyxl rarely *constructs* deeply (pivots, slicers, charts, complex
conditional formatting, drawings, comments, tables, validations). For each
fixture, the test:

1. Verifies the on-disk SHA256 matches the pinned manifest entry — guards
   against accidental modification of the in-tree pack.
2. Asserts every entry in the fixture's `expected_parts` list survives
   representative wolfxl modify-save cycles. This is stronger than "no
   parts lost" because it catches drift on the parts that openpyxl normalizes
   away or that would silently get dropped.
3. Confirms the marker mutation round-trips through openpyxl, the ZIP CRC
   checks pass, and each fixture continues to open cleanly.

The `WOLFXL_EXTERNAL_FIXTURES_DIR` env var still overrides the in-tree
path so a freshly-regenerated ExcelBench pack can be tested before being
pinned. When the env var is set, the SHA256 check is skipped because the
hashes there are not pinned.
"""

from __future__ import annotations

import hashlib
import importlib.util
import json
import os
import shutil
import sys
import zipfile
from pathlib import Path
from types import ModuleType

import openpyxl
import pytest

import wolfxl

_FIXTURE_ENV = "WOLFXL_EXTERNAL_FIXTURES_DIR"
_PINNED_DIR = Path(__file__).resolve().parent / "fixtures" / "external_oracle"
_MANIFEST_NAME = "manifest.json"
_MARKER_CELL = "Z1"
_MARKER_VALUE = "wolfxl_external_fixture_smoke"
_STYLE_CELL = "AA1"
_MUTATIONS = (
    "no_op",
    "marker_cell",
    "style_cell",
    "insert_tail_row",
    "insert_tail_col",
    "move_marker_range",
)


def _load_ooxml_audit_module() -> ModuleType:
    script = Path(__file__).resolve().parents[1] / "scripts" / "audit_ooxml_fidelity.py"
    spec = importlib.util.spec_from_file_location("audit_ooxml_fidelity", script)
    assert spec is not None
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


_OOXML_AUDIT = _load_ooxml_audit_module()


def _fixture_dir() -> Path:
    if env_path := os.environ.get(_FIXTURE_ENV):
        return Path(env_path).expanduser()
    return _PINNED_DIR


def _load_manifest(fixture_dir: Path) -> list[dict]:
    """Return the manifest fixture list, or [] if no manifest exists.

    The pinned in-tree pack always has a manifest. The opt-in env override
    may point at an unpinned ExcelBench output dir (no manifest) — in that
    case we fall back to discovering xlsx files and skip the SHA256 check.
    """
    manifest_path = fixture_dir / _MANIFEST_NAME
    if not manifest_path.is_file():
        return []
    payload = json.loads(manifest_path.read_text())
    return list(payload.get("fixtures", []))


def _discover_unpinned(fixture_dir: Path) -> list[dict]:
    """Build a manifest-shaped list for an unpinned (env-override) directory."""
    return [
        {"filename": p.name, "expected_parts": [], "sha256": None}
        for p in sorted(fixture_dir.glob("*.xlsx"))
        if p.is_file() and not p.name.startswith("~$")
    ]


def _fixture_entries() -> tuple[Path, list[dict]]:
    fixture_dir = _fixture_dir()
    if not fixture_dir.is_dir():
        return fixture_dir, []
    entries = _load_manifest(fixture_dir)
    if not entries:
        entries = _discover_unpinned(fixture_dir)
    return fixture_dir, entries


_FIXTURE_DIR, _FIXTURE_ENTRIES = _fixture_entries()


def _entry_id(entry: dict) -> str:
    return entry["filename"]


@pytest.mark.skipif(
    not _FIXTURE_ENTRIES,
    reason=(
        "external oracle fixture pack not found; expected at "
        f"{_PINNED_DIR} or override via {_FIXTURE_ENV}"
    ),
)
@pytest.mark.parametrize("entry", _FIXTURE_ENTRIES, ids=_entry_id)
@pytest.mark.parametrize("mutation", _MUTATIONS)
def test_external_oracle_fixture_modify_save_preserves_expected_parts(
    entry: dict, mutation: str, tmp_path: Path
) -> None:
    """Each pinned fixture's expected_parts must survive wolfxl modify-save."""
    fixture_path = _FIXTURE_DIR / entry["filename"]
    assert fixture_path.is_file(), f"fixture missing: {fixture_path}"

    # Pin guard: in-tree pack has SHA256s pinned. If the manifest entry has
    # one, verify it. Env-override packs (no manifest) skip this.
    if expected_hash := entry.get("sha256"):
        actual_hash = hashlib.sha256(fixture_path.read_bytes()).hexdigest()
        assert actual_hash == expected_hash, (
            f"{fixture_path.name} on-disk SHA256 mismatch — fixture has been "
            "modified or replaced. Re-pin via "
            "scripts/refresh_external_oracle_fixtures.py if intentional, "
            "otherwise restore from git."
        )

    before_audit_path = tmp_path / f"before-{fixture_path.name}"
    work_path = tmp_path / fixture_path.name
    shutil.copy2(fixture_path, before_audit_path)
    shutil.copy2(fixture_path, work_path)

    before_parts = _zip_parts(work_path)

    workbook = wolfxl.load_workbook(work_path, modify=True)
    sheet_name = workbook.sheetnames[0]
    if mutation == "marker_cell":
        workbook[sheet_name][_MARKER_CELL] = _MARKER_VALUE
    elif mutation == "style_cell":
        from wolfxl.styles import Font, PatternFill

        cell = workbook[sheet_name][_STYLE_CELL]
        cell.value = _MARKER_VALUE
        cell.font = Font(bold=True, color="FF1F4E79")
        cell.fill = PatternFill(fill_type="solid", fgColor="FFEAF2F8")
    elif mutation == "insert_tail_row":
        worksheet = workbook[sheet_name]
        row_idx = int(getattr(worksheet, "max_row", 1) or 1) + 1
        worksheet.insert_rows(row_idx, amount=1)
        worksheet.cell(row=row_idx, column=1).value = _MARKER_VALUE
    elif mutation == "insert_tail_col":
        worksheet = workbook[sheet_name]
        col_idx = int(getattr(worksheet, "max_column", 1) or 1) + 1
        worksheet.insert_cols(col_idx, amount=1)
        worksheet.cell(row=1, column=col_idx).value = _MARKER_VALUE
    elif mutation == "move_marker_range":
        worksheet = workbook[sheet_name]
        worksheet["Z1"] = _MARKER_VALUE
        worksheet["AA1"] = f"{_MARKER_VALUE}_right"
        worksheet.move_range("Z1:AA1", rows=1, cols=0)
    workbook.save(work_path)
    workbook.close()

    after_parts = _zip_parts(work_path)

    audit_report = _OOXML_AUDIT.audit(before_audit_path, work_path)
    assert not audit_report["issues"], (
        f"{fixture_path.name} failed OOXML fidelity audit after {mutation}: "
        f"{json.dumps(audit_report['issues'], indent=2, sort_keys=True)}"
    )

    # Stronger gate than "no parts lost": every entry the fixture's
    # originating tool authored must still be present.
    expected_parts = set(entry.get("expected_parts") or [])
    missing_expected = expected_parts - after_parts
    assert not missing_expected, (
        f"{fixture_path.name} lost EXPECTED OOXML parts during modify-save: "
        f"{sorted(missing_expected)}. These parts were authored by "
        f"{entry.get('tool', 'the originating tool')} and are part of the "
        "pinned preservation contract."
    )

    # Belt-and-suspenders: also assert no other parts disappeared.
    missing_any = before_parts - after_parts
    assert not missing_any, (
        f"{fixture_path.name} lost incidental OOXML parts during modify-save: "
        f"{sorted(missing_any)}"
    )

    # Openpyxl load validates the wolfxl save is readable by the canonical
    # OOXML reader, not just by wolfxl itself.
    roundtrip = openpyxl.load_workbook(work_path, data_only=False)
    try:
        if mutation == "marker_cell":
            assert roundtrip[sheet_name][_MARKER_CELL].value == _MARKER_VALUE, (
                f"marker write to {fixture_path.name}!{_MARKER_CELL} did not "
                "round-trip through openpyxl"
            )
    finally:
        roundtrip.close()

    # ZIP CRC integrity.
    with zipfile.ZipFile(work_path) as archive:
        bad = archive.testzip()
        assert bad is None, f"{fixture_path.name} ZIP integrity failure: {bad}"


def _zip_parts(path: Path) -> set[str]:
    with zipfile.ZipFile(path) as archive:
        return set(archive.namelist())
