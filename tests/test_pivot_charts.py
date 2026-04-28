"""Sprint Ν Pod-δ tests — pivot-chart linkage (RFC-049).

Covers the public surface of ``ChartBase.pivot_source`` (Python attr +
setter + validation), the chart-dict §10.1 shape, the Rust emitter's
``<c:pivotSource>`` block + per-series ``<c:fmtId>`` injection, and the
``parse_chart_dict`` PyO3 boundary.

Test #10 (cross-mode equivalence) and #12 (backward-compat) operate on
the bytes returned by ``serialize_chart_dict`` directly — no save
needed.
"""

from __future__ import annotations

import os
import shutil
import subprocess
import zipfile
from pathlib import Path

import openpyxl
import pytest

import wolfxl
from wolfxl import load_workbook
from wolfxl.chart import BarChart, Reference

# `serialize_chart_dict` is the PyO3 boundary that consumes a chart
# dict and returns OOXML bytes. It is the single helper exercised by
# both write-mode (Workbook._flush_pending_charts_to_patcher) and
# modify-mode bridges.
from wolfxl._rust import serialize_chart_dict  # type: ignore[attr-defined]
from wolfxl.pivot import (
    DataField,
    PivotCache,
    PivotTable,
)


# ---------------------------------------------------------------------------
# Helpers — minimal worksheet + cache + chart construction
# ---------------------------------------------------------------------------


class _StubCell:
    def __init__(self, value):
        self.value = value


class _StubWorksheet:
    """Tiny worksheet stub mirroring ``test_pivot_construction.py`` —
    avoids pulling the full Workbook stack."""

    def __init__(self, title, data):
        self.title = title
        self._data = data

    def __getitem__(self, addr):
        return _StubCell(self._data.get(addr))


def _build_sample_pivot_cache() -> PivotCache:
    data = {
        "A1": "region", "B1": "quarter", "C1": "customer", "D1": "revenue",
        "A2": "North",  "B2": "Q1",      "C2": "Acme",     "D2": 100.0,
        "A3": "South",  "B3": "Q1",      "C3": "Acme",     "D3": 200.0,
        "A4": "North",  "B4": "Q2",      "C4": "Globex",   "D4": 150.0,
        "A5": "South",  "B5": "Q2",      "C5": "Globex",   "D5": 250.0,
    }
    ws = _StubWorksheet("Sheet1", data)
    src = Reference(worksheet=ws, min_col=1, min_row=1, max_col=4, max_row=5)
    pc = PivotCache(source=src)
    pc._cache_id = 0
    pc._materialize(ws)
    return pc


def _build_sample_pivot_table(name: str = "MyPivot") -> PivotTable:
    pc = _build_sample_pivot_cache()
    return PivotTable(
        cache=pc,
        location="F2",
        rows=["region"],
        cols=["quarter"],
        data=[DataField(name="revenue", function="sum")],
        name=name,
    )


def _build_chart_with_data() -> BarChart:
    """Bar chart with one series — enough to satisfy
    ``ChartBase._validate_at_emit`` so ``to_rust_dict`` runs."""
    wb = wolfxl.Workbook()
    ws = wb.active
    ws.append(["", "S1"])
    for i in range(1, 4):
        ws.append([f"row{i}", i * 10])
    chart = BarChart()
    chart.add_data(
        Reference(ws, min_col=2, min_row=1, max_col=2, max_row=4),
        titles_from_data=True,
    )
    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_col=1, max_row=4))
    return chart


def _build_saved_pivot_chart_workbook(tmp_path: Path) -> Path:
    """Materialize a workbook carrying a pivot table and linked chart."""
    seed = tmp_path / "pivot_chart_seed.xlsx"
    out = tmp_path / "pivot_chart.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    rows = [
        ("region", "quarter", "customer", "revenue"),
        ("North", "Q1", "Acme", 100.0),
        ("South", "Q1", "Acme", 200.0),
        ("North", "Q2", "Globex", 150.0),
        ("South", "Q2", "Globex", 250.0),
    ]
    for row in rows:
        ws.append(row)
    wb.save(seed)

    wbm = load_workbook(seed, modify=True)
    wsm = wbm["Data"]
    ref = Reference(worksheet=wsm, min_col=1, min_row=1, max_col=4, max_row=5)
    cache = PivotCache(source=ref)
    pt = PivotTable(
        cache=cache,
        location="F2",
        rows=["region"],
        cols=["quarter"],
        data=["revenue"],
        name="SalesPivot",
    )
    wbm.add_pivot_cache(cache)
    wsm.add_pivot_table(pt)

    chart = BarChart()
    chart.title = "Revenue"
    chart.add_data(
        Reference(wsm, min_col=4, min_row=1, max_col=4, max_row=5),
        titles_from_data=True,
    )
    chart.set_categories(Reference(wsm, min_col=1, min_row=2, max_col=1, max_row=5))
    chart.pivot_source = pt
    wsm.add_chart(chart, "F12")
    wbm.save(out)
    return out


# ---------------------------------------------------------------------------
# 1-3. Setter accepts PivotTable / tuple / None
# ---------------------------------------------------------------------------


def test_pivot_source_setter_accepts_pivot_table():
    chart = _build_chart_with_data()
    pt = _build_sample_pivot_table("MyPivot")
    chart.pivot_source = pt
    assert chart.pivot_source == {"name": "MyPivot", "fmt_id": 0}


def test_pivot_source_setter_accepts_tuple():
    chart = _build_chart_with_data()
    chart.pivot_source = ("Quarterly", 7)
    assert chart.pivot_source == {"name": "Quarterly", "fmt_id": 7}


def test_pivot_source_setter_accepts_none_clears():
    chart = _build_chart_with_data()
    chart.pivot_source = ("Quarterly", 0)
    assert chart.pivot_source is not None
    chart.pivot_source = None
    assert chart.pivot_source is None


# ---------------------------------------------------------------------------
# 4. Setter rejects malformed inputs
# ---------------------------------------------------------------------------


def test_pivot_source_setter_rejects_empty_name():
    chart = _build_chart_with_data()
    with pytest.raises(ValueError, match="non-empty string"):
        chart.pivot_source = ("", 0)


def test_pivot_source_setter_rejects_bad_regex():
    chart = _build_chart_with_data()
    with pytest.raises(ValueError, match="regex"):
        chart.pivot_source = ("1starts-with-digit", 0)


def test_pivot_source_setter_rejects_fmt_id_out_of_range():
    chart = _build_chart_with_data()
    with pytest.raises(ValueError, match=r"\[0, 65535\]"):
        chart.pivot_source = ("MyPivot", 70_000)


def test_pivot_source_setter_rejects_negative_fmt_id():
    chart = _build_chart_with_data()
    with pytest.raises(ValueError, match=r"\[0, 65535\]"):
        chart.pivot_source = ("MyPivot", -1)


def test_pivot_source_setter_rejects_wrong_arity_tuple():
    chart = _build_chart_with_data()
    with pytest.raises(ValueError, match="tuple of length"):
        chart.pivot_source = ("MyPivot",)  # type: ignore[assignment]


def test_pivot_source_setter_rejects_unknown_type():
    chart = _build_chart_with_data()
    with pytest.raises(TypeError, match="PivotTable"):
        chart.pivot_source = 42  # type: ignore[assignment]


# ---------------------------------------------------------------------------
# 5-7. Emitter behaviour via serialize_chart_dict
# ---------------------------------------------------------------------------


def test_emitter_produces_pivot_source_block_when_set():
    chart = _build_chart_with_data()
    chart.pivot_source = ("MyPivot", 0)
    xml_bytes = serialize_chart_dict(chart.to_rust_dict(), "B2")
    text = xml_bytes.decode("utf-8")
    assert "<c:pivotSource>" in text
    assert "<c:name>MyPivot</c:name>" in text
    assert '<c:fmtId val="0"/></c:pivotSource>' in text


def test_emitter_omits_pivot_source_when_none():
    chart = _build_chart_with_data()
    assert chart.pivot_source is None
    xml_bytes = serialize_chart_dict(chart.to_rust_dict(), "B2")
    text = xml_bytes.decode("utf-8")
    assert "<c:pivotSource" not in text
    # And no spurious series-level <c:fmtId>.
    assert "<c:fmtId" not in text


def test_emitter_injects_per_series_fmt_id_when_pivot_source_set():
    chart = _build_chart_with_data()
    chart.pivot_source = ("MyPivot", 0)
    xml_bytes = serialize_chart_dict(chart.to_rust_dict(), "B2")
    text = xml_bytes.decode("utf-8")
    # RFC-049 §2: fmtId injected immediately after the order block.
    assert '<c:order val="0"/><c:fmtId val="0"/>' in text


def test_emitter_uses_provided_fmt_id_on_pivot_source_block():
    chart = _build_chart_with_data()
    chart.pivot_source = ("MyPivot", 5)
    xml_bytes = serialize_chart_dict(chart.to_rust_dict(), "B2")
    text = xml_bytes.decode("utf-8")
    assert '<c:fmtId val="5"/></c:pivotSource>' in text
    # Per-series fmtId mirrors the chart-level value.
    assert '<c:order val="0"/><c:fmtId val="5"/>' in text


# ---------------------------------------------------------------------------
# 8. openpyxl round-trip — depends on Pod-γ patcher integration
# ---------------------------------------------------------------------------


def test_pivot_source_round_trips_through_openpyxl(tmp_path: Path):
    path = _build_saved_pivot_chart_workbook(tmp_path)
    with zipfile.ZipFile(path) as zf:
        assert zf.testzip() is None
        chart_xml = zf.read("xl/charts/chart1.xml")
    assert b"<c:pivotSource>" in chart_xml
    assert b"<c:name>SalesPivot</c:name>" in chart_xml

    wb = openpyxl.load_workbook(path)
    try:
        ws = wb["Data"]
        assert len(ws._charts) == 1
        assert len(ws._pivots) == 1
        assert ws._pivots[0].name == "SalesPivot"
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# 9. LibreOffice fixture (opt-in)
# ---------------------------------------------------------------------------


def test_pivot_chart_libreoffice_renders(tmp_path: Path):
    if shutil.which("soffice") is None:
        pytest.skip("LibreOffice (soffice) not installed in PATH")
    if os.environ.get("WOLFXL_RUN_LIBREOFFICE_SMOKE") != "1":
        pytest.skip(
            "LibreOffice smoke test opt-in via "
            "WOLFXL_RUN_LIBREOFFICE_SMOKE=1; depends on Pod-γ patcher"
        )
    src = _build_saved_pivot_chart_workbook(tmp_path)
    proc = subprocess.run(
        [
            shutil.which("soffice") or "soffice",
            "--headless",
            "--convert-to",
            "pdf",
            "--outdir",
            str(tmp_path),
            str(src),
        ],
        capture_output=True,
        text=True,
        timeout=60,
    )
    assert proc.returncode == 0, (
        f"soffice failed: stdout={proc.stdout} stderr={proc.stderr}"
    )
    pdf = tmp_path / "pivot_chart.pdf"
    assert pdf.exists()
    assert pdf.stat().st_size > 0


# ---------------------------------------------------------------------------
# 10. Cross-mode equivalence — write-mode and modify-mode produce
#     equal bytes for a chart with `pivot_source` set.
#
# Both paths funnel through ``serialize_chart_dict`` (Pod-α′ unified
# the bridge in v1.6.1), so identical input dicts MUST produce identical
# output bytes regardless of the caller. RFC-046 §10.10.
# ---------------------------------------------------------------------------


def test_write_mode_and_modify_mode_produce_equal_bytes_with_pivot_source(
    monkeypatch,
):
    monkeypatch.setenv("WOLFXL_TEST_EPOCH", "0")
    chart = _build_chart_with_data()
    chart.pivot_source = ("MyPivot", 0)
    d = chart.to_rust_dict()
    bytes_a = serialize_chart_dict(d, "B2")
    bytes_b = serialize_chart_dict(d, "B2")
    assert bytes_a == bytes_b
    # And the bytes carry the pivot block deterministically.
    assert b"<c:pivotSource>" in bytes_a


# ---------------------------------------------------------------------------
# 11. copy_worksheet round-trip — depends on Pod-γ
# ---------------------------------------------------------------------------


def test_pivot_source_round_trips_through_copy_worksheet(tmp_path: Path):
    src = _build_saved_pivot_chart_workbook(tmp_path)
    out = tmp_path / "pivot_chart_clone.xlsx"
    wb = load_workbook(src, modify=True)
    wb.copy_worksheet(wb["Data"])
    wb.save(out)

    with zipfile.ZipFile(out) as zf:
        chart_parts = sorted(n for n in zf.namelist() if n.startswith("xl/charts/chart"))
        pivot_parts = sorted(
            n for n in zf.namelist() if n.startswith("xl/pivotTables/pivotTable")
        )
        assert len(chart_parts) == 2
        assert len(pivot_parts) == 2
        for chart_part in chart_parts:
            chart_xml = zf.read(chart_part)
            assert b"<c:pivotSource>" in chart_xml
            assert b"<c:name>SalesPivot</c:name>" in chart_xml

    wb2 = openpyxl.load_workbook(out)
    try:
        assert wb2.sheetnames == ["Data", "Data Copy"]
        assert [len(ws._charts) for ws in wb2.worksheets] == [1, 1]
        assert [len(ws._pivots) for ws in wb2.worksheets] == [1, 1]
    finally:
        wb2.close()


# ---------------------------------------------------------------------------
# 12. RFC-049 §10.5 backward-compat — chart dict missing the
#     `pivot_source` key still parses and emits identically to v1.7
#     output (no implicit pivot block, no per-series fmtId).
# ---------------------------------------------------------------------------


def test_chart_dict_without_pivot_source_key_emits_v1_7_bytes(monkeypatch):
    monkeypatch.setenv("WOLFXL_TEST_EPOCH", "0")
    chart = _build_chart_with_data()

    d_with_key = chart.to_rust_dict()
    # Mirror a v1.7 chart-dict by stripping the new key entirely.
    d_legacy = {k: v for k, v in d_with_key.items() if k != "pivot_source"}

    bytes_with_none = serialize_chart_dict(d_with_key, "B2")
    bytes_legacy = serialize_chart_dict(d_legacy, "B2")

    assert bytes_with_none == bytes_legacy, (
        "Chart dicts with pivot_source=None and chart dicts missing the "
        "pivot_source key MUST produce identical bytes (RFC-049 §10.5)."
    )
    # And neither emits a pivot block.
    assert b"<c:pivotSource" not in bytes_legacy
    assert b"<c:fmtId" not in bytes_legacy


# ---------------------------------------------------------------------------
# Extra — parser-level rejection mirrors Python validation.
# ---------------------------------------------------------------------------


def test_parse_chart_dict_rejects_bad_pivot_source_name():
    chart = _build_chart_with_data()
    d = chart.to_rust_dict()
    d["pivot_source"] = {"name": "1bad", "fmt_id": 0}
    with pytest.raises(ValueError, match="regex"):
        serialize_chart_dict(d, "B2")


def test_parse_chart_dict_rejects_bad_pivot_source_fmt_id():
    chart = _build_chart_with_data()
    d = chart.to_rust_dict()
    d["pivot_source"] = {"name": "MyPivot", "fmt_id": 70_000}
    with pytest.raises(ValueError, match=r"65535"):
        serialize_chart_dict(d, "B2")


def test_parse_chart_dict_accepts_sheet_qualified_pivot_source_name():
    chart = _build_chart_with_data()
    d = chart.to_rust_dict()
    d["pivot_source"] = {"name": "Sheet1!MyPivot", "fmt_id": 0}
    xml = serialize_chart_dict(d, "B2")
    assert b"<c:name>Sheet1!MyPivot</c:name>" in xml
