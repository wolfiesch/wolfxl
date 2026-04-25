"""T0 compat parity for Worksheet iteration/slicing/dimensions.

Each test builds a workbook with openpyxl, saves it, opens with wolfxl,
and asserts the return shape and values match openpyxl exactly. This pins
the contracts that make wolfxl a true drop-in for common openpyxl idioms.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

import wolfxl


@pytest.fixture
def small_xlsx(tmp_path: Path) -> Path:
    """Build a small .xlsx with known data for parity tests."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for r, row in enumerate([[1, 2, 3], [4, 5, 6], [7, 8, 9], [10, 11, 12]], start=1):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)
    path = tmp_path / "small.xlsx"
    wb.save(path)
    return path


def test_dimensions_matches_openpyxl(small_xlsx: Path) -> None:
    op = openpyxl.load_workbook(small_xlsx)["Sheet1"]
    wx = wolfxl.load_workbook(small_xlsx)["Sheet1"]
    assert wx.dimensions == op.dimensions


def test_min_row_min_column(small_xlsx: Path) -> None:
    op = openpyxl.load_workbook(small_xlsx)["Sheet1"]
    wx = wolfxl.load_workbook(small_xlsx)["Sheet1"]
    assert wx.min_row == op.min_row
    assert wx.min_column == op.min_column


def test_parent_backref(small_xlsx: Path) -> None:
    wb = wolfxl.load_workbook(small_xlsx)
    ws = wb["Sheet1"]
    assert ws.parent is wb


def test_range_2d(small_xlsx: Path) -> None:
    op = openpyxl.load_workbook(small_xlsx)["Sheet1"]
    wx = wolfxl.load_workbook(small_xlsx)["Sheet1"]

    op_block = [[c.value for c in row] for row in op["A1:C2"]]
    wx_block = [[c.value for c in row] for row in wx["A1:C2"]]
    assert wx_block == op_block
    assert isinstance(wx["A1:C2"], tuple)
    assert isinstance(wx["A1:C2"][0], tuple)


def test_column_range_bounded(small_xlsx: Path) -> None:
    op = openpyxl.load_workbook(small_xlsx)["Sheet1"]
    wx = wolfxl.load_workbook(small_xlsx)["Sheet1"]

    op_block = [[c.value for c in row] for row in op["A:B"]]
    wx_block = [[c.value for c in row] for row in wx["A:B"]]
    assert wx_block == op_block


def test_row_range(small_xlsx: Path) -> None:
    op = openpyxl.load_workbook(small_xlsx)["Sheet1"]
    wx = wolfxl.load_workbook(small_xlsx)["Sheet1"]

    op_block = [[c.value for c in row] for row in op["1:2"]]
    wx_block = [[c.value for c in row] for row in wx["1:2"]]
    assert wx_block == op_block


def test_single_column(small_xlsx: Path) -> None:
    op = openpyxl.load_workbook(small_xlsx)["Sheet1"]
    wx = wolfxl.load_workbook(small_xlsx)["Sheet1"]

    op_col = [c.value for c in op["A"]]
    wx_col = [c.value for c in wx["A"]]
    assert wx_col == op_col


def test_single_row_by_int(small_xlsx: Path) -> None:
    op = openpyxl.load_workbook(small_xlsx)["Sheet1"]
    wx = wolfxl.load_workbook(small_xlsx)["Sheet1"]

    op_row = [c.value for c in op[1]]
    wx_row = [c.value for c in wx[1]]
    assert wx_row == op_row


def test_single_row_by_str(small_xlsx: Path) -> None:
    op = openpyxl.load_workbook(small_xlsx)["Sheet1"]
    wx = wolfxl.load_workbook(small_xlsx)["Sheet1"]

    op_row = [c.value for c in op["2"]]
    wx_row = [c.value for c in wx["2"]]
    assert wx_row == op_row


def test_row_slice_int(small_xlsx: Path) -> None:
    op = openpyxl.load_workbook(small_xlsx)["Sheet1"]
    wx = wolfxl.load_workbook(small_xlsx)["Sheet1"]

    op_block = [[c.value for c in row] for row in op[1:3]]
    wx_block = [[c.value for c in row] for row in wx[1:3]]
    assert wx_block == op_block


def test_iter_cols_cells(small_xlsx: Path) -> None:
    op = openpyxl.load_workbook(small_xlsx)["Sheet1"]
    wx = wolfxl.load_workbook(small_xlsx)["Sheet1"]

    op_cols = [[c.value for c in col] for col in op.iter_cols()]
    wx_cols = [[c.value for c in col] for col in wx.iter_cols()]
    assert wx_cols == op_cols


def test_iter_cols_values_only(small_xlsx: Path) -> None:
    op = openpyxl.load_workbook(small_xlsx)["Sheet1"]
    wx = wolfxl.load_workbook(small_xlsx)["Sheet1"]

    op_cols = [tuple(col) for col in op.iter_cols(values_only=True)]
    wx_cols = [tuple(col) for col in wx.iter_cols(values_only=True)]
    assert wx_cols == op_cols


def test_iter_cols_partial_bounds(small_xlsx: Path) -> None:
    op = openpyxl.load_workbook(small_xlsx)["Sheet1"]
    wx = wolfxl.load_workbook(small_xlsx)["Sheet1"]

    op_cols = [
        tuple(col)
        for col in op.iter_cols(min_col=2, max_col=3, min_row=1, max_row=2, values_only=True)
    ]
    wx_cols = [
        tuple(col)
        for col in wx.iter_cols(min_col=2, max_col=3, min_row=1, max_row=2, values_only=True)
    ]
    assert wx_cols == op_cols


def test_rows_property(small_xlsx: Path) -> None:
    op = openpyxl.load_workbook(small_xlsx)["Sheet1"]
    wx = wolfxl.load_workbook(small_xlsx)["Sheet1"]

    op_rows = [[c.value for c in row] for row in op.rows]
    wx_rows = [[c.value for c in row] for row in wx.rows]
    assert wx_rows == op_rows


def test_columns_property(small_xlsx: Path) -> None:
    op = openpyxl.load_workbook(small_xlsx)["Sheet1"]
    wx = wolfxl.load_workbook(small_xlsx)["Sheet1"]

    op_cols = [[c.value for c in col] for col in op.columns]
    wx_cols = [[c.value for c in col] for col in wx.columns]
    assert wx_cols == op_cols


def test_values_property(small_xlsx: Path) -> None:
    op = openpyxl.load_workbook(small_xlsx)["Sheet1"]
    wx = wolfxl.load_workbook(small_xlsx)["Sheet1"]

    assert list(wx.values) == list(op.values)


def test_iter_cols_bulk_fast_path(small_xlsx: Path) -> None:
    """values_only + read mode should hit the single bulk Rust call.

    Verified by wrapping the reader to count ``read_sheet_values_plain`` calls.
    Rust-bound attributes can't be monkeypatched directly, so we swap the
    workbook's ``_rust_reader`` with a proxy object that delegates.
    """

    class CountingReader:
        def __init__(self, inner: object) -> None:
            self._inner = inner
            self.plain_calls = 0

        def __getattr__(self, name: str):
            return getattr(self._inner, name)

        def read_sheet_values_plain(self, *args, **kwargs):
            self.plain_calls += 1
            return self._inner.read_sheet_values_plain(*args, **kwargs)

    wb = wolfxl.load_workbook(small_xlsx)
    ws = wb["Sheet1"]
    counting = CountingReader(wb._rust_reader)  # noqa: SLF001
    wb._rust_reader = counting  # noqa: SLF001
    _ = list(ws.iter_cols(values_only=True))
    assert counting.plain_calls == 1


def test_write_mode_dimensions() -> None:
    wb = wolfxl.Workbook()
    ws = wb.active
    ws["A1"] = 1
    ws["C3"] = 2
    assert ws.dimensions == "A1:C3"
    assert ws.min_row == 1
    assert ws.min_column == 1
    assert ws.max_row == 3
    assert ws.max_column == 3


def test_write_mode_range_access() -> None:
    wb = wolfxl.Workbook()
    ws = wb.active
    for r in range(1, 4):
        for c in range(1, 4):
            ws.cell(row=r, column=c, value=r * 10 + c)
    block = ws["A1:B2"]
    assert [[c.value for c in row] for row in block] == [[11, 12], [21, 22]]


def test_round_trip_with_openpyxl(tmp_path: Path) -> None:
    """Write with wolfxl, read back with openpyxl, access via ranges."""
    wb = wolfxl.Workbook()
    ws = wb.active
    for r, row in enumerate([["a", "b", "c"], [1, 2, 3]], start=1):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)
    path = tmp_path / "out.xlsx"
    wb.save(path)

    # Re-open with wolfxl in read mode and verify slicing.
    wx = wolfxl.load_workbook(path)[wb.sheetnames[0]]
    assert [[c.value for c in row] for row in wx["A1:C2"]] == [["a", "b", "c"], [1, 2, 3]]
