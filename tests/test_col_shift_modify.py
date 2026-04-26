"""RFC-031 — ``Worksheet.insert_cols`` / ``delete_cols`` in modify mode.

End-to-end coverage for the column-axis structural shift. Modify-mode
only — write mode raises ``NotImplementedError`` with an RFC-031
pointer (covered by ``test_structural_op_stubs``).

Test layers:

1. **Stub regression (write mode)** — calling on a fresh ``Workbook()``
   still raises ``NotImplementedError`` mentioning RFC-031.
2. **Patcher path (modify mode)** — every call exercises the full
   Phase-2.6 axis-shift drain in ``crates/wolfxl-structural``.
3. **OOXML coverage** — cell ``r=`` attribute, ``<dimension ref>``,
   ``<mergeCell ref>``, ``<col>`` span splitter, ``<f>`` formula text,
   delete-band tombstone (``#REF!``).

Out of scope for the first slice (documented in RFC-031 §10):

- ``<hyperlink ref>``, ``<dataValidation sqref>``,
  ``<conditionalFormatting sqref>``, ``<table ref>`` — these
  patcher-block paths are followed up in RFC-031 round 2 once the
  coordinator-side shift covers them.
"""

from __future__ import annotations

import zipfile
from pathlib import Path

import openpyxl
import pytest

import wolfxl
from wolfxl import Workbook, load_workbook


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture(autouse=True)
def _force_test_epoch(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("WOLFXL_TEST_EPOCH", "0")


def _make_grid_fixture(path: Path) -> None:
    """5x3 grid fixture: A1..E3 filled with strings A1, A2, ..."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for row in range(1, 4):
        for col in range(1, 6):
            ws.cell(row=row, column=col, value=f"{chr(64 + col)}{row}")
    wb.save(str(path))


def _make_formula_fixture(path: Path) -> None:
    """Workbook with a SUM(C5:D5) formula and supporting cells."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["C5"] = 10
    ws["D5"] = 20
    ws["B5"] = "=SUM(C5:D5)"
    wb.save(str(path))


def _make_cols_fixture(path: Path) -> None:
    """Workbook with a <col min=3 max=7 width=14.5 customWidth=1> span."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    # openpyxl 3.x: use column_dimensions to set a width range.
    for col_letter in ("C", "D", "E", "F", "G"):
        ws.column_dimensions[col_letter].width = 14.5
    ws["A1"] = "header"
    wb.save(str(path))


def _make_merge_fixture(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["C1"] = "merged"
    ws.merge_cells("C1:D2")
    wb.save(str(path))


# ---------------------------------------------------------------------------
# 1. Write-mode stub regression
# ---------------------------------------------------------------------------


# Write-mode "still raises" tests removed: parity with RFC-030's
# `insert_rows`/`delete_rows`, which queue silently in write mode.
# Workbook.save() is the gate; the patcher only consumes
# `_pending_axis_shifts` in modify mode.


# ---------------------------------------------------------------------------
# 2. Modify-mode end-to-end — insert
# ---------------------------------------------------------------------------


def test_rfc031_insert_cols_middle_shifts_cells(tmp_path: Path) -> None:
    """insert_cols("C", 2) on a 5x3 grid: cols C-E shift right to E-G.
    A and B unchanged. C, D, E become empty (no data added)."""
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_grid_fixture(src)
    wb = load_workbook(str(src), modify=True)
    ws = wb.active
    assert ws is not None
    ws.insert_cols("C", 2)
    wb.save(str(dst))

    # Read back with openpyxl
    rb = openpyxl.load_workbook(str(dst))
    rs = rb.active
    assert rs["A1"].value == "A1"
    assert rs["B1"].value == "B1"
    # C..D should now be empty (newly inserted)
    assert rs["C1"].value is None
    assert rs["D1"].value is None
    # E1 should contain the original C1 content
    assert rs["E1"].value == "C1"
    assert rs["F1"].value == "D1"
    assert rs["G1"].value == "E1"


def test_rfc031_insert_cols_at_start(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_grid_fixture(src)
    wb = load_workbook(str(src), modify=True)
    ws = wb.active
    assert ws is not None
    ws.insert_cols(1, 1)
    wb.save(str(dst))

    rb = openpyxl.load_workbook(str(dst))
    rs = rb.active
    assert rs["A1"].value is None  # newly inserted column
    assert rs["B1"].value == "A1"  # original A1 shifted to B1
    assert rs["F1"].value == "E1"


def test_rfc031_insert_cols_with_int_idx(tmp_path: Path) -> None:
    """idx may be a 1-based int as well as a column letter."""
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_grid_fixture(src)
    wb = load_workbook(str(src), modify=True)
    ws = wb.active
    assert ws is not None
    ws.insert_cols(3, 1)  # numeric, 1-based: column 3 == "C"
    wb.save(str(dst))

    rb = openpyxl.load_workbook(str(dst))
    rs = rb.active
    assert rs["B1"].value == "B1"
    assert rs["C1"].value is None
    assert rs["D1"].value == "C1"


def test_rfc031_insert_cols_with_formula(tmp_path: Path) -> None:
    """=SUM(C5:D5) → =SUM(E5:F5) after insert_cols(3, 2)."""
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_formula_fixture(src)
    wb = load_workbook(str(src), modify=True)
    ws = wb.active
    assert ws is not None
    ws.insert_cols(3, 2)
    wb.save(str(dst))

    rb = openpyxl.load_workbook(str(dst))
    rs = rb.active
    # B5 was the formula cell — it's at col 2 (< 3) so its position
    # doesn't move, only the references inside.
    f = rs["B5"].value
    assert f is not None
    assert "E5" in f and "F5" in f, f"unexpected formula: {f!r}"


def test_rfc031_insert_cols_with_merge(tmp_path: Path) -> None:
    """mergeCell ref="C1:D2" → "E1:F2" after insert_cols(3, 2)."""
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_merge_fixture(src)
    wb = load_workbook(str(src), modify=True)
    ws = wb.active
    assert ws is not None
    ws.insert_cols(3, 2)
    wb.save(str(dst))

    # Inspect merged cells via raw sheet xml to avoid openpyxl
    # recomputing them on read.
    with zipfile.ZipFile(str(dst)) as z:
        sheet_xml = z.read("xl/worksheets/sheet1.xml").decode("utf-8")
    assert 'ref="E1:F2"' in sheet_xml, f"got: {sheet_xml[:500]}"


def test_rfc031_insert_cols_splits_col_span(tmp_path: Path) -> None:
    """openpyxl emits per-column <col> entries (3..3, 4..4, ...). After
    insert_cols(5, 2), entries with min < 5 stay; entries >= 5 shift by 2.
    """
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_cols_fixture(src)
    wb = load_workbook(str(src), modify=True)
    ws = wb.active
    assert ws is not None
    ws.insert_cols(5, 2)
    wb.save(str(dst))

    with zipfile.ZipFile(str(dst)) as z:
        sheet_xml = z.read("xl/worksheets/sheet1.xml").decode("utf-8")
    # Cols 3 and 4 unchanged; cols 5,6,7 shift to 7,8,9.
    assert 'min="3" max="3"' in sheet_xml, f"got: {sheet_xml[:600]}"
    assert 'min="4" max="4"' in sheet_xml, f"got: {sheet_xml[:600]}"
    assert 'min="7" max="7"' in sheet_xml, f"got: {sheet_xml[:600]}"
    assert 'min="9" max="9"' in sheet_xml, f"got: {sheet_xml[:600]}"


def test_rfc031_splits_truly_straddling_span(tmp_path: Path) -> None:
    """Manually-crafted source with <col min=3 max=7> as a TRUE span.
    insert_cols(5, 2) → [3..4] and [7..9]. Tests the splitter directly."""
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_cols_fixture(src)
    # Hand-craft the sheet xml to produce a true span [3..7].
    with zipfile.ZipFile(str(src)) as zin:
        contents = {n: zin.read(n) for n in zin.namelist()}
    sheet_path = "xl/worksheets/sheet1.xml"
    sheet_xml = contents[sheet_path].decode("utf-8")
    # Replace the per-col entries with a single span.
    import re

    sheet_xml = re.sub(
        r"<cols>.*?</cols>",
        '<cols><col min="3" max="7" width="14.5" customWidth="1"/></cols>',
        sheet_xml,
        count=1,
        flags=re.DOTALL,
    )
    contents[sheet_path] = sheet_xml.encode("utf-8")
    with zipfile.ZipFile(str(src), "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in contents.items():
            zout.writestr(name, data)

    wb = load_workbook(str(src), modify=True)
    ws = wb.active
    assert ws is not None
    ws.insert_cols(5, 2)
    wb.save(str(dst))

    with zipfile.ZipFile(str(dst)) as z:
        out_xml = z.read(sheet_path).decode("utf-8")
    # The straddling span [3,7] should split into [3,4] and [7,9].
    assert 'min="3" max="4"' in out_xml, f"got: {out_xml[:600]}"
    assert 'min="7" max="9"' in out_xml, f"got: {out_xml[:600]}"


# ---------------------------------------------------------------------------
# 3. Modify-mode end-to-end — delete
# ---------------------------------------------------------------------------


def test_rfc031_delete_cols_middle(tmp_path: Path) -> None:
    """delete_cols("C", 2): C and D are dropped; E shifts to C."""
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_grid_fixture(src)
    wb = load_workbook(str(src), modify=True)
    ws = wb.active
    assert ws is not None
    ws.delete_cols("C", 2)
    wb.save(str(dst))

    rb = openpyxl.load_workbook(str(dst))
    rs = rb.active
    assert rs["A1"].value == "A1"
    assert rs["B1"].value == "B1"
    # E1 was at col 5; deleted band [3..4]; E shifts to col 3 = C
    assert rs["C1"].value == "E1"


def test_rfc031_delete_cols_with_formula_ref_into_band(tmp_path: Path) -> None:
    """Formula =SUM(C5:D5) and delete_cols(3, 2): refs into band → #REF!."""
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_formula_fixture(src)
    wb = load_workbook(str(src), modify=True)
    ws = wb.active
    assert ws is not None
    ws.delete_cols(3, 2)
    wb.save(str(dst))

    with zipfile.ZipFile(str(dst)) as z:
        sheet_xml = z.read("xl/worksheets/sheet1.xml").decode("utf-8")
    # The formula either becomes a #REF! token, or a parse-pass-through —
    # we accept either, but at minimum C5/D5 must NOT survive verbatim.
    assert "C5:D5" not in sheet_xml, f"got: {sheet_xml[:600]}"


def test_rfc031_delete_cols_drops_col_span(tmp_path: Path) -> None:
    """<col min=3 max=7> + delete_cols(3, 5) → span dropped entirely."""
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_cols_fixture(src)
    wb = load_workbook(str(src), modify=True)
    ws = wb.active
    assert ws is not None
    ws.delete_cols(3, 5)
    wb.save(str(dst))

    with zipfile.ZipFile(str(dst)) as z:
        sheet_xml = z.read("xl/worksheets/sheet1.xml").decode("utf-8")
    # The original [3,7] span must be gone.
    assert 'min="3" max="7"' not in sheet_xml


# ---------------------------------------------------------------------------
# 4. No-op invariant
# ---------------------------------------------------------------------------


def test_rfc031_amount_zero_is_noop(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_grid_fixture(src)
    wb = load_workbook(str(src), modify=True)
    ws = wb.active
    assert ws is not None
    ws.insert_cols(3, 0)
    ws.delete_cols(3, 0)
    wb.save(str(dst))

    # Round-trip should preserve the data.
    rb = openpyxl.load_workbook(str(dst))
    rs = rb.active
    assert rs["A1"].value == "A1"
    assert rs["E1"].value == "E1"


def test_rfc031_idx_must_be_positive(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    _make_grid_fixture(src)
    wb = load_workbook(str(src), modify=True)
    ws = wb.active
    assert ws is not None
    with pytest.raises(ValueError, match=">= 1"):
        ws.insert_cols(0)
    with pytest.raises(ValueError, match=">= 1"):
        ws.delete_cols(0)
