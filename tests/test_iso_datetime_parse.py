"""W4E.P2 regression: parse_iso_datetime accepts tz-aware ISO 8601.

Previously, ``src/util.rs::parse_iso_datetime`` only stripped a
trailing ``Z``. Inputs like ``2025-01-15T10:30:45+05:30`` and
``2025-01-15T10:30:45-08:00`` (which Python's ``datetime.isoformat()``
emits when tzinfo is set) parsed to ``None`` — silent data loss in
``set_properties({"created": ...})`` for any tz-aware datetime.

The fix tries ``DateTime::parse_from_rfc3339`` first (handles ``Z``,
``+HH:MM``, ``-HH:MM``), then falls back to the naive parses. We
discard the offset (Excel's ``dcterms:created`` is tz-naive); only
the local datetime survives, which is the documented Excel behavior.

These tests round-trip via ``set_properties()`` + openpyxl reload to
confirm the parser path is wired through both backends.
"""
from __future__ import annotations

import datetime
from pathlib import Path

import openpyxl
import pytest


_FIXED_DT = datetime.datetime(2025, 1, 15, 10, 30, 45)


@pytest.mark.parametrize(
    "iso_string",
    [
        "2025-01-15T10:30:45",
        "2025-01-15T10:30:45Z",
        "2025-01-15T10:30:45+00:00",
        # P2 regression: positive offset must round-trip without dropping
        # the datetime entirely.
        "2025-01-15T10:30:45+05:30",
        # P2 regression: negative offset.
        "2025-01-15T10:30:45-08:00",
        # Microseconds + offset (Python isoformat() with tzinfo).
        # NB: dcterms:created emits whole seconds (per Excel's own
        # behavior in format_naive), so microseconds are lost at emit
        # — but the datetime itself must NOT be silently dropped.
        "2025-01-15T10:30:45.123456-08:00",
    ],
)
@pytest.mark.parametrize("backend", ["native", "oracle"])
def test_set_properties_created_accepts_tz_iso(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
    iso_string: str,
    backend: str,
) -> None:
    """Both backends should accept tz-aware ISO datetimes via
    ``set_properties()`` and emit a non-empty ``dcterms:created`` that
    openpyxl parses back to the local datetime."""
    monkeypatch.setenv("WOLFXL_WRITER", backend)
    import wolfxl

    wb = wolfxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.cell(row=1, column=1, value="probe")
    # Reach into the rust writer directly — set_properties() is the
    # 4B/oracle pymethod that exercises parse_iso_datetime.
    wb._rust_writer.set_properties({"created": iso_string})

    out = tmp_path / f"{backend}.xlsx"
    wb.save(str(out))
    assert out.exists() and out.stat().st_size > 0

    rb = openpyxl.load_workbook(out)
    got = rb.properties.created
    assert got is not None, (
        f"{backend}: dcterms:created was empty after iso={iso_string!r} — "
        f"parse_iso_datetime returned None and dropped the field"
    )
    # openpyxl returns a tz-naive datetime; compare local components
    # to whole-second precision (microseconds are dropped at emit per
    # Excel's dcterms:created spec).
    got_naive = got.replace(tzinfo=None, microsecond=0) if got.tzinfo else got.replace(microsecond=0)
    assert got_naive == _FIXED_DT, (
        f"{backend}: iso={iso_string!r} expected={_FIXED_DT} got={got_naive}"
    )
