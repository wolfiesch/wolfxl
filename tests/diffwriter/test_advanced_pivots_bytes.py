"""Sprint Ο Pod 3.5 (RFC-061) — byte-stable tests for slicer parts.

Pinned with ``WOLFXL_TEST_EPOCH=0`` so timestamps are deterministic.
Asserts:
  * Writing the same slicer-augmented workbook twice yields
    byte-identical ``xl/slicerCaches/slicerCache1.xml`` and
    ``xl/slicers/slicer1.xml``.
  * The workbook.xml `<extLst>` block carries `<x14:slicerCaches>`.
  * The owning sheet's `<extLst>` carries `<x14:slicerList>`.
"""
from __future__ import annotations

import zipfile
from pathlib import Path

import openpyxl
import pytest

from wolfxl import load_workbook
from wolfxl.chart.reference import Reference
from wolfxl.pivot import PivotCache, PivotTable, Slicer, SlicerCache


@pytest.fixture(autouse=True)
def _pin_epoch(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("WOLFXL_TEST_EPOCH", "0")


def _make_pivot_source(path: Path, title: str = "Data") -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    rows = [
        ("region", "quarter", "revenue"),
        ("North", "Q1", 100.0),
        ("South", "Q1", 200.0),
        ("North", "Q2", 150.0),
        ("South", "Q2", 250.0),
    ]
    for r, row in enumerate(rows, start=1):
        for c, v in enumerate(row, start=1):
            ws.cell(r, c, v)
    wb.save(path)


def _add_slicer(src: Path, dst: Path) -> None:
    wb = load_workbook(src, modify=True)
    sheet = wb["Data"]
    ref = Reference(worksheet=sheet, min_col=1, min_row=1, max_col=3, max_row=5)
    cache = PivotCache(source=ref)
    pt = PivotTable(cache=cache, location="F2", rows=["region"], data=["revenue"])
    wb.add_pivot_cache(cache)
    sheet.add_pivot_table(pt)
    sc = SlicerCache(name="Slicer_region", source_pivot_cache=cache, field="region")
    sl = Slicer(name="Slicer_region1", cache=sc, caption="Region")
    wb.add_slicer_cache(sc)
    sheet.add_slicer(sl, anchor="H2")
    wb.save(dst)


def _zip_read(p: Path, member: str) -> bytes:
    with zipfile.ZipFile(p, "r") as z:
        return z.read(member)


def test_slicer_cache_xml_byte_stable(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    a = tmp_path / "a.xlsx"
    b = tmp_path / "b.xlsx"
    _make_pivot_source(src)
    _add_slicer(src, a)
    _add_slicer(src, b)

    cache_a = _zip_read(a, "xl/slicerCaches/slicerCache1.xml")
    cache_b = _zip_read(b, "xl/slicerCaches/slicerCache1.xml")
    assert cache_a == cache_b, "slicerCache1.xml is not byte-stable"


def test_slicer_xml_byte_stable(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    a = tmp_path / "a.xlsx"
    b = tmp_path / "b.xlsx"
    _make_pivot_source(src)
    _add_slicer(src, a)
    _add_slicer(src, b)

    slicer_a = _zip_read(a, "xl/slicers/slicer1.xml")
    slicer_b = _zip_read(b, "xl/slicers/slicer1.xml")
    assert slicer_a == slicer_b, "slicer1.xml is not byte-stable"


def test_slicer_cache_xml_carries_pivot_linkage(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_pivot_source(src)
    _add_slicer(src, dst)

    cache_xml = _zip_read(dst, "xl/slicerCaches/slicerCache1.xml").decode()
    # RFC-061 §3.1 wire-format pins.
    assert "<slicerCacheDefinition" in cache_xml, cache_xml
    assert 'name="Slicer_region"' in cache_xml, cache_xml
    assert "<pivotTables>" in cache_xml, cache_xml
    assert "<tabular" in cache_xml, cache_xml


def test_slicer_xml_carries_anchor_and_cache_ref(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_pivot_source(src)
    _add_slicer(src, dst)

    slicer_xml = _zip_read(dst, "xl/slicers/slicer1.xml").decode()
    assert "<slicers" in slicer_xml, slicer_xml
    assert 'name="Slicer_region1"' in slicer_xml, slicer_xml
    assert 'cache="Slicer_region"' in slicer_xml, slicer_xml
    assert 'caption="Region"' in slicer_xml, slicer_xml


def test_workbook_xml_has_x14_slicer_caches_ext(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_pivot_source(src)
    _add_slicer(src, dst)

    wb_xml = _zip_read(dst, "xl/workbook.xml").decode()
    assert "<extLst>" in wb_xml, wb_xml[:1500]
    assert "<x14:slicerCaches" in wb_xml, wb_xml[:1500]
    # The extension URI should be the canonical slicer-caches URI.
    assert "{A8765BA9-456A-4DAB-B4F3-ACF838C121DE}" in wb_xml


def test_sheet_xml_has_x14_slicer_list_ext(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_pivot_source(src)
    _add_slicer(src, dst)

    # Source workbook from openpyxl writes one sheet → sheet1.xml.
    sheet_xml = _zip_read(dst, "xl/worksheets/sheet1.xml").decode()
    assert "<x14:slicerList" in sheet_xml, sheet_xml[-1500:]
    assert "{3A4CF648-6AED-40f4-86FF-DC5316D8AED3}" in sheet_xml


def test_content_types_overrides_present(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_pivot_source(src)
    _add_slicer(src, dst)

    ct = _zip_read(dst, "[Content_Types].xml").decode()
    assert "application/vnd.ms-excel.slicerCache+xml" in ct, ct
    assert "application/vnd.ms-excel.slicer+xml" in ct, ct


def test_workbook_rels_carries_slicer_cache_rel(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_pivot_source(src)
    _add_slicer(src, dst)

    rels = _zip_read(dst, "xl/_rels/workbook.xml.rels").decode()
    assert "office/2007/relationships/slicerCache" in rels, rels
    assert "slicerCaches/slicerCache1.xml" in rels, rels


def test_sheet_rels_carries_slicer_rel(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_pivot_source(src)
    _add_slicer(src, dst)

    sheet_rels = _zip_read(dst, "xl/worksheets/_rels/sheet1.xml.rels").decode()
    assert "office/2007/relationships/slicer" in sheet_rels, sheet_rels
    assert "slicers/slicer1.xml" in sheet_rels, sheet_rels


def test_slicer_cache_rels_points_at_pivot_cache(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    _make_pivot_source(src)
    _add_slicer(src, dst)

    cache_rels = _zip_read(
        dst, "xl/slicerCaches/_rels/slicerCache1.xml.rels"
    ).decode()
    assert "pivotCacheDefinition" in cache_rels, cache_rels
