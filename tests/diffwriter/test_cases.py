"""Pytest driver — one parametrized test per case across all case modules.

Discovers every ``CASES`` list in ``tests.diffwriter.cases.*`` at collection
time and exercises the build path on the native writer. Because the W5
rip-out removed the legacy ``rust_xlsxwriter`` oracle, the harness no
longer performs cross-backend byte / structural / semantic diffs. Instead
each case asserts:

  1. The build closure runs without raising on the public API.
  2. ``Workbook.save()`` produces a non-empty xlsx file.
  3. The written file opens cleanly under openpyxl (the soft secondary
     oracle): every active worksheet is iterable and the cell collection
     materializes without error.

Layer-1 / Layer-2 / Layer-3 byte-and-tree diffs against a committed
golden fixture are tracked separately as a RFC-002+ follow-up — until
those land, this driver is the minimum viable regression net for the
case builders themselves.
"""
from __future__ import annotations

import importlib
import pkgutil
from pathlib import Path
from typing import Any, Callable

import openpyxl
import pytest

from . import cases as cases_pkg


def _discover_cases() -> list[tuple[str, Callable[[Any], None]]]:
    """Walk ``tests/diffwriter/cases`` and return every (id, build) pair."""
    out: list[tuple[str, Callable[[Any], None]]] = []
    for mod_info in pkgutil.iter_modules(cases_pkg.__path__):
        mod = importlib.import_module(f"tests.diffwriter.cases.{mod_info.name}")
        cases = getattr(mod, "CASES", None)
        if cases is None:
            continue
        for case_id, build_fn in cases:
            out.append((case_id, build_fn))
    return out


_ALL_CASES = _discover_cases()


@pytest.mark.parametrize(
    "case_id,build_fn",
    _ALL_CASES,
    ids=[c[0] for c in _ALL_CASES],
)
def test_native_case_round_trips(
    tmp_path: Path,
    case_id: str,
    build_fn: Callable[[Any], None],
) -> None:
    """Build the case via the native writer; verify openpyxl can re-open it."""
    import wolfxl

    wb = wolfxl.Workbook()
    build_fn(wb)
    output_path = tmp_path / "out.xlsx"
    wb.save(str(output_path))
    assert output_path.exists(), f"{case_id}: save did not produce {output_path}"
    assert output_path.stat().st_size > 0, f"{case_id}: empty xlsx written"

    # Soft secondary oracle: openpyxl must parse the file. We don't
    # assert specific cell values here — case builders don't carry
    # expected-output metadata, and per-feature parity is covered by
    # tests/parity/. The point is to catch gross regressions where the
    # native writer emits something openpyxl can't decode.
    rb = openpyxl.load_workbook(str(output_path))
    for sheet_name in rb.sheetnames:
        sh = rb[sheet_name]
        # Force materialization of the cell grid; an emit bug that
        # produced unparseable XML would surface here.
        list(sh.iter_rows(values_only=True))
