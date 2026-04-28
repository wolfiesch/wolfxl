"""RFC-035 — byte-stability / determinism gate for ``copy_worksheet``.

Layer-2 of the §6 verification matrix. The repo's full diffwriter
golden infrastructure is tracked as a post-RFC-002 follow-up (see
``tests/diffwriter/README.md``); until those goldens land, this test
is the minimum-viable equivalent: load the same canonical fixture
twice, run the SAME copy_worksheet operation on both, save both, and
assert the output bytes are identical.

Determinism gate:

1. ``WOLFXL_TEST_EPOCH=0`` pins ZIP entry mtimes (already used across
   the harness via the autouse fixture).
2. Two independent ``load_workbook`` + ``copy_worksheet`` + ``save``
   sequences must produce byte-equal output.
3. The cloned sheet's ``xl/worksheets/sheetN.xml`` and rels file are
   byte-equal across the two runs.

Canonical fixture: 5×5 grid + one table + one external hyperlink + a
sheet-scoped Print_Area defined name. Mirrors the fixture surface
called for in RFC-035 §6 row 2 (golden round-trip).

If this test ever fails, the most likely culprit is non-deterministic
ordering in: (a) the rels graph allocator, (b) the content-types
override list, or (c) the zip entry order. Each of these has explicit
tests in the Rust crates; this test is the integration-level guard.
"""

from __future__ import annotations

import zipfile
from pathlib import Path

import openpyxl
import pytest

from wolfxl import load_workbook

pytestmark = pytest.mark.rfc035


@pytest.fixture(autouse=True)
def _force_test_epoch(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("WOLFXL_TEST_EPOCH", "0")


def _build_canonical_fixture(path: Path) -> None:
    """Multi-feature fixture: 5×5 grid + table + hyperlink + sheet-scope name."""
    from openpyxl.workbook.defined_name import DefinedName as XDefinedName
    from openpyxl.worksheet.table import Table

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"
    headers = ["k", "a", "b", "c", "d"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    for r in range(2, 6):
        for c in range(1, 6):
            ws.cell(row=r, column=c, value=(r - 1) * 10 + c)
    ws.add_table(Table(displayName="Sales", ref="A1:E5"))
    ws["E5"] = "click"
    ws["E5"].hyperlink = "https://example.com/golden"
    wb.defined_names["_xlnm.Print_Area"] = XDefinedName(
        "_xlnm.Print_Area", attr_text="Template!$A$1:$E$5", localSheetId=0
    )
    wb.save(path)


def _do_copy(src: Path, out: Path) -> None:
    wb = load_workbook(src, modify=True)
    wb.copy_worksheet(wb.active)
    wb.save(out)


def _zip_entry_bytes(path: Path, entry: str) -> bytes:
    with zipfile.ZipFile(path) as zf:
        return zf.read(entry)


def test_copy_worksheet_two_runs_byte_identical(tmp_path: Path) -> None:
    """Two independent runs must produce byte-equal xlsx files."""
    src1 = tmp_path / "src1.xlsx"
    src2 = tmp_path / "src2.xlsx"
    out1 = tmp_path / "out1.xlsx"
    out2 = tmp_path / "out2.xlsx"

    # Build the same fixture twice (independently, via openpyxl).
    _build_canonical_fixture(src1)
    _build_canonical_fixture(src2)

    # Sanity: the two source fixtures should already be byte-equal at
    # the input stage (with WOLFXL_TEST_EPOCH=0 + openpyxl). If they
    # diverge, the divergence isn't on the wolfxl side and the test
    # should xfail with the openpyxl-fixture culprit.
    if src1.read_bytes() != src2.read_bytes():
        pytest.xfail(
            "openpyxl source fixtures diverge byte-wise — non-determinism "
            "is on the openpyxl fixture side, not wolfxl. This test "
            "asserts wolfxl's output determinism only."
        )

    _do_copy(src1, out1)
    _do_copy(src2, out2)

    out1_bytes = out1.read_bytes()
    out2_bytes = out2.read_bytes()
    assert out1_bytes == out2_bytes, (
        f"copy_worksheet is non-deterministic: out1 size={len(out1_bytes)} "
        f"out2 size={len(out2_bytes)}; first byte diff at "
        f"{next((i for i, (a, b) in enumerate(zip(out1_bytes, out2_bytes)) if a != b), 'EOF')}"
    )


def test_clone_sheet_xml_byte_stable(tmp_path: Path) -> None:
    """The cloned sheet's worksheet XML must be byte-stable across runs.

    Specific byte-level snapshot of ``xl/worksheets/sheet2.xml`` (the
    clone) — narrows down a potential non-determinism to the sheet
    clone's bytes specifically.
    """
    src1 = tmp_path / "src1.xlsx"
    src2 = tmp_path / "src2.xlsx"
    out1 = tmp_path / "out1.xlsx"
    out2 = tmp_path / "out2.xlsx"

    _build_canonical_fixture(src1)
    _build_canonical_fixture(src2)

    _do_copy(src1, out1)
    _do_copy(src2, out2)

    sheet1_xml = _zip_entry_bytes(out1, "xl/worksheets/sheet2.xml")
    sheet2_xml = _zip_entry_bytes(out2, "xl/worksheets/sheet2.xml")
    assert sheet1_xml == sheet2_xml, (
        "Cloned sheet XML diverges across runs — non-deterministic "
        "rels-rId allocation or table-id allocation is the likely "
        "culprit."
    )


def test_clone_rels_byte_stable(tmp_path: Path) -> None:
    """The cloned sheet's rels file must be byte-stable across runs."""
    src1 = tmp_path / "src1.xlsx"
    src2 = tmp_path / "src2.xlsx"
    out1 = tmp_path / "out1.xlsx"
    out2 = tmp_path / "out2.xlsx"

    _build_canonical_fixture(src1)
    _build_canonical_fixture(src2)

    _do_copy(src1, out1)
    _do_copy(src2, out2)

    rels1 = _zip_entry_bytes(out1, "xl/worksheets/_rels/sheet2.xml.rels")
    rels2 = _zip_entry_bytes(out2, "xl/worksheets/_rels/sheet2.xml.rels")
    assert rels1 == rels2, (
        "Cloned sheet rels file diverges across runs — non-deterministic "
        "rId allocation."
    )


def test_workbook_xml_byte_stable_after_copy(tmp_path: Path) -> None:
    """workbook.xml must be byte-stable across runs after copy_worksheet
    (the <sheets> append + sheet-scope defined name re-point all happen
    inside Phase 2.7 + 2.5h).
    """
    src1 = tmp_path / "src1.xlsx"
    src2 = tmp_path / "src2.xlsx"
    out1 = tmp_path / "out1.xlsx"
    out2 = tmp_path / "out2.xlsx"

    _build_canonical_fixture(src1)
    _build_canonical_fixture(src2)

    _do_copy(src1, out1)
    _do_copy(src2, out2)

    wb1 = _zip_entry_bytes(out1, "xl/workbook.xml")
    wb2 = _zip_entry_bytes(out2, "xl/workbook.xml")
    assert wb1 == wb2, (
        "workbook.xml diverges across runs — Phase 2.7's <sheet> "
        "append or 2.5h's defined-name re-point is non-deterministic."
    )
