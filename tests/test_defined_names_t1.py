"""T1 PR3 — DefinedNameDict upgrade.

wb.defined_names is now a DefinedNameDict whose values are DefinedName
objects. This is a BREAKING change from T0 (it used to be ``dict[str, str]``).
Migration: ``wb.defined_names[name].value`` returns the old string.
"""

from __future__ import annotations

from pathlib import Path

import pytest
from wolfxl.workbook import DefinedNameDict
from wolfxl.workbook.defined_name import DefinedName

from wolfxl import Workbook

openpyxl = pytest.importorskip("openpyxl")


@pytest.fixture()
def workbook_with_names(tmp_path: Path) -> Path:
    from openpyxl.workbook.defined_name import DefinedName as XDefinedName

    path = tmp_path / "names.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    for i in range(1, 11):
        ws[f"A{i}"] = i
    dn = XDefinedName("Region", attr_text=f"'{ws.title}'!$A$1:$A$10")
    wb.defined_names["Region"] = dn
    wb.save(path)
    return path


def test_defined_names_is_dict_subclass(workbook_with_names: Path) -> None:
    wb = Workbook._from_reader(str(workbook_with_names))
    dns = wb.defined_names
    assert isinstance(dns, DefinedNameDict)
    assert isinstance(dns, dict)


def test_defined_name_entry_has_value_attribute(workbook_with_names: Path) -> None:
    wb = Workbook._from_reader(str(workbook_with_names))
    entry = wb.defined_names["Region"]
    assert isinstance(entry, DefinedName)
    # Leading '=' is stripped. Should include the sheet-qualified range.
    assert "A1:A10" in entry.value or "$A$1:$A$10" in entry.value
    assert entry.attr_text == entry.value  # openpyxl alias


def test_write_mode_defined_names_empty() -> None:
    wb = Workbook()
    dns = wb.defined_names
    assert isinstance(dns, DefinedNameDict)
    assert len(dns) == 0
    # ``== {}`` still holds because DefinedNameDict inherits from dict.
    assert dns == {}


def test_defined_names_cache_stable(workbook_with_names: Path) -> None:
    """wb.defined_names returns the same dict instance each call."""
    wb = Workbook._from_reader(str(workbook_with_names))
    a = wb.defined_names
    b = wb.defined_names
    assert a is b


def test_defined_name_dict_rejects_mismatched_key() -> None:
    wb = Workbook()
    dns = wb.defined_names
    dn = DefinedName(name="Foo", value="Sheet1!A1")
    with pytest.raises(ValueError, match="does not match"):
        dns["Bar"] = dn


def test_defined_name_dict_rejects_non_defined_name_value() -> None:
    wb = Workbook()
    dns = wb.defined_names
    with pytest.raises(TypeError, match="DefinedName"):
        dns["Foo"] = "not a DefinedName"  # type: ignore[arg-type]


def test_defined_name_dict_accepts_correct_value() -> None:
    wb = Workbook()
    dns = wb.defined_names
    dns["Region"] = DefinedName(name="Region", value="Sheet1!$A$1:$A$10")
    assert dns["Region"].value == "Sheet1!$A$1:$A$10"


def test_defined_name_dict_add_helper() -> None:
    wb = Workbook()
    dn = DefinedName(name="Totals", value="Sheet1!$B$1")
    wb.defined_names.add(dn)
    assert wb.defined_names["Totals"] is dn


def test_sheet_scoped_defined_names_live_on_worksheet(tmp_path: Path) -> None:
    from openpyxl.workbook.defined_name import DefinedName as XDefinedName

    path = tmp_path / "sheet-scoped-names.xlsx"
    op_wb = openpyxl.Workbook()
    data = op_wb.active
    data.title = "Data"
    other = op_wb.create_sheet("Other")
    op_wb.defined_names.add(XDefinedName("GlobalName", attr_text="Data!$A$1"))
    data.defined_names.add(XDefinedName("LocalName", attr_text="$B$2"))
    other.defined_names.add(XDefinedName("OtherLocal", attr_text="$C$3"))
    op_wb.save(path)

    wb = Workbook._from_reader(str(path))

    assert list(wb.defined_names) == ["GlobalName"]
    assert wb.defined_names["GlobalName"].attr_text == "Data!$A$1"
    assert list(wb["Data"].defined_names) == ["LocalName"]
    assert wb["Data"].defined_names["LocalName"].attr_text == "$B$2"
    assert list(wb["Other"].defined_names) == ["OtherLocal"]
    assert wb["Other"].defined_names["OtherLocal"].attr_text == "$C$3"
