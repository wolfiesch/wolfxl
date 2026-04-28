"""Sprint Μ Pod-δ — chart parity tests vs openpyxl (RFC-046 §8).

Three axes of coverage:

1. **Structural XML compare** — for each of the 8 chart types, build
   an identical chart with wolfxl and openpyxl, save, extract
   ``xl/charts/chart1.xml`` from each, canonicalise via
   ``_chart_helpers.canonicalize_chart_xml``, and assert structural
   equivalence (axis IDs zeroed; namespace prefixes collapsed; attribute
   order normalised by c14n).

2. **openpyxl reads wolfxl** — wolfxl writes a chart-bearing xlsx,
   openpyxl loads it, and we verify (a) the right chart class lands on
   ``ws._charts``, (b) the title round-trips, (c) the series count is
   correct, (d) each series's data Reference points at the same range.

3. **Cross-feature parity** — 10 hand-picked sub-feature tests that
   exercise representative gaps (data labels, error bars, trendlines,
   layout, legend, axis configuration).

4. **LibreOffice headless smoke** — gated by ``WOLFXL_RUN_LIBREOFFICE_SMOKE=1``;
   verifies a wolfxl-written chart-bearing xlsx renders cleanly via
   ``soffice --headless --convert-to pdf`` for every chart kind.

The tests are imports-skipped on missing deps (``openpyxl``, ``lxml``,
``Pillow``) and chart-API-skipped until Pods α + β land. Once those
land, the import skip flips off and the structural tests will reveal any
divergences for the integrator to triage.
"""

from __future__ import annotations

import os
import shutil
import subprocess
from pathlib import Path
from typing import Any

import pytest

openpyxl = pytest.importorskip("openpyxl")
lxml_etree = pytest.importorskip("lxml.etree")
pytest.importorskip("PIL")  # openpyxl image features need Pillow

import wolfxl  # noqa: E402

from ._chart_helpers import (  # noqa: E402
    CHART_KINDS,
    build_identical_chart,
    canonicalize_chart_xml,
    extract_chart_xml,
    first_chart_xml,
    structurally_equivalent,
)

# Skip the entire module if Pod-β's chart API is still in stub form.
try:
    from wolfxl.chart import (  # type: ignore[attr-defined] # noqa: F401
        BarChart,
        Reference,
    )
    _CHART_API_AVAILABLE = True
except (ImportError, AttributeError, NotImplementedError):
    _CHART_API_AVAILABLE = False


pytestmark = pytest.mark.skipif(
    not _CHART_API_AVAILABLE,
    reason="wolfxl.chart construction API ships in Sprint Μ Pods α+β.",
)


# ---------------------------------------------------------------------------
# Shared data-seeding helper
# ---------------------------------------------------------------------------


def _seed(ws: Any) -> None:
    ws.append(["", "S1", "S2"])
    for i in range(1, 6):
        ws.append([f"r{i}", i * 10, i * 5])


def _save_pair(kind: str, tmp_path: Path, **features: Any) -> tuple[Path, Path]:
    """Build *the same* chart with both libraries, save, return paths."""
    # wolfxl side
    wwb = wolfxl.Workbook()
    wws = wwb.active
    _seed(wws)
    wchart = build_identical_chart("wolfxl", kind, wws, features=features)
    wws.add_chart(wchart, "E2")
    wpath = tmp_path / "wolfxl.xlsx"
    wwb.save(wpath)

    # openpyxl side
    owb = openpyxl.Workbook()
    ows = owb.active
    _seed(ows)
    ochart = build_identical_chart("openpyxl", kind, ows, features=features)
    ows.add_chart(ochart, "E2")
    opath = tmp_path / "openpyxl.xlsx"
    owb.save(opath)

    return wpath, opath


# ---------------------------------------------------------------------------
# 1. Structural XML compare — per chart type (8 cases)
# ---------------------------------------------------------------------------


@pytest.mark.parametrize("kind", CHART_KINDS)
def test_chart_xml_structurally_matches_openpyxl(tmp_path: Path, kind: str) -> None:
    """For each chart type, canonicalised chart1.xml should match openpyxl.

    Accepted divergences (handled inside ``canonicalize_chart_xml``):

    * ``<c:axId val="..."/>`` — openpyxl uses random ints, wolfxl uses a
      counter. Both are zeroed before comparison.
    * ``<c:crossAx val="..."/>`` — references the auto-generated axis IDs;
      same treatment.
    * Attribute order — c14n already normalises lexicographically.
    """
    wpath, opath = _save_pair(kind, tmp_path, title=f"{kind} test")

    wxml = first_chart_xml(wpath)
    oxml = first_chart_xml(opath)

    ok, diff = structurally_equivalent(wxml, oxml)
    assert ok, f"chart kind={kind} XML diverged from openpyxl:\n{diff}"


# ---------------------------------------------------------------------------
# 2. openpyxl reads wolfxl — per chart type (8 cases)
# ---------------------------------------------------------------------------


_KIND_TO_OXC_CLS = {
    "bar": "BarChart",
    "line": "LineChart",
    "pie": "PieChart",
    "doughnut": "DoughnutChart",
    "area": "AreaChart",
    "scatter": "ScatterChart",
    "bubble": "BubbleChart",
    "radar": "RadarChart",
}


@pytest.mark.parametrize("kind", CHART_KINDS)
def test_openpyxl_reads_wolfxl_chart(tmp_path: Path, kind: str) -> None:
    """wolfxl-written chart loads back via openpyxl with class + title intact."""
    wb = wolfxl.Workbook()
    ws = wb.active
    _seed(ws)
    chart = build_identical_chart(
        "wolfxl", kind, ws, features={"title": f"{kind}-title"}
    )
    ws.add_chart(chart, "E2")
    out = tmp_path / "out.xlsx"
    wb.save(out)

    wb2 = openpyxl.load_workbook(out)
    ws2 = wb2.active
    charts = ws2._charts
    assert len(charts) == 1, f"expected 1 chart, got {len(charts)}"

    expected_cls = _KIND_TO_OXC_CLS[kind]
    actual_cls = type(charts[0]).__name__
    assert actual_cls == expected_cls, (
        f"openpyxl read wolfxl chart as {actual_cls}, expected {expected_cls}"
    )

    # Title round-trip — openpyxl exposes Title's text either as
    # ``.tx.rich.p[0].r[0].t`` or via the str helper. Just check the
    # canonical raw XML carries the title text.
    raw = first_chart_xml(out).decode()
    assert f"{kind}-title" in raw

    # Series count — every chart kind should carry exactly 2 series
    # since the seed block has 2 numeric columns.
    assert len(charts[0].series) == 2

    # Each series's numRef points at column B or C of the data block.
    refs = [str(s.val.numRef.f) for s in charts[0].series if s.val and s.val.numRef]
    assert any("B" in r for r in refs), f"no series points at column B: {refs}"
    assert any("C" in r for r in refs), f"no series points at column C: {refs}"


# ---------------------------------------------------------------------------
# 3. Cross-feature parity — 10 hand-picked cases
# ---------------------------------------------------------------------------


def test_bar_chart_with_categories_xml_matches_openpyxl(tmp_path: Path) -> None:
    wpath, opath = _save_pair("bar", tmp_path, title="With cats")
    ok, diff = structurally_equivalent(
        first_chart_xml(wpath), first_chart_xml(opath)
    )
    assert ok, f"bar+categories diverged:\n{diff}"


def test_line_chart_with_marker_xml_matches_openpyxl(tmp_path: Path) -> None:
    wpath, opath = _save_pair("line", tmp_path, marker=True)
    ok, diff = structurally_equivalent(
        first_chart_xml(wpath), first_chart_xml(opath)
    )
    assert ok, f"line+marker diverged:\n{diff}"


def test_pie_chart_with_data_labels_xml_matches_openpyxl(tmp_path: Path) -> None:
    wpath, opath = _save_pair("pie", tmp_path, data_labels=True)
    ok, diff = structurally_equivalent(
        first_chart_xml(wpath), first_chart_xml(opath)
    )
    assert ok, f"pie+labels diverged:\n{diff}"


def test_scatter_chart_with_trendline_xml_matches_openpyxl(tmp_path: Path) -> None:
    wpath, opath = _save_pair("scatter", tmp_path, trendline="linear")
    ok, diff = structurally_equivalent(
        first_chart_xml(wpath), first_chart_xml(opath)
    )
    assert ok, f"scatter+trendline diverged:\n{diff}"


def test_axis_categoryAxis_xml_matches_openpyxl(tmp_path: Path) -> None:
    """The categoryAxis block (``<c:catAx>``) round-trips identically."""
    wpath, opath = _save_pair("bar", tmp_path)
    wxml = first_chart_xml(wpath).decode()
    oxml = first_chart_xml(opath).decode()
    assert "catAx" in wxml and "catAx" in oxml


def test_axis_valueAxis_with_majorUnit_xml_matches_openpyxl(
    tmp_path: Path,
) -> None:
    """Setting ``y_axis.majorUnit`` should land in canonical XML."""

    def _attach(chart: Any) -> None:
        try:
            chart.y_axis.majorUnit = 5
        except AttributeError:
            pass

    wpath, opath = _save_pair("bar", tmp_path)
    # Apply majorUnit to both fixtures by rebuilding here — tighter
    # control than threading another feature through ``build_identical_chart``.
    for path, lib in ((wpath, "wolfxl"), (opath, "openpyxl")):
        # The structural diff after this is the gold check.
        pass

    # Re-build, this time setting majorUnit directly.
    wb = wolfxl.Workbook()
    ws = wb.active
    _seed(ws)
    wchart = build_identical_chart("wolfxl", "bar", ws)
    _attach(wchart)
    ws.add_chart(wchart, "E2")
    wpath2 = tmp_path / "wolfxl_majorunit.xlsx"
    wb.save(wpath2)

    owb = openpyxl.Workbook()
    ows = owb.active
    _seed(ows)
    ochart = build_identical_chart("openpyxl", "bar", ows)
    _attach(ochart)
    ows.add_chart(ochart, "E2")
    opath2 = tmp_path / "openpyxl_majorunit.xlsx"
    owb.save(opath2)

    ok, diff = structurally_equivalent(
        first_chart_xml(wpath2), first_chart_xml(opath2)
    )
    assert ok, f"valueAxis+majorUnit diverged:\n{diff}"


def test_legend_top_position_xml_matches_openpyxl(tmp_path: Path) -> None:
    wpath, opath = _save_pair("bar", tmp_path, legend_pos="t")
    ok, diff = structurally_equivalent(
        first_chart_xml(wpath), first_chart_xml(opath)
    )
    assert ok, f"legend top diverged:\n{diff}"


def test_data_labels_position_xml_matches_openpyxl(tmp_path: Path) -> None:
    """``DataLabelList(showVal=True, position="outEnd")`` should round-trip."""
    from openpyxl.chart.label import DataLabelList

    wb = wolfxl.Workbook()
    ws = wb.active
    _seed(ws)
    wchart = build_identical_chart("wolfxl", "bar", ws)
    from wolfxl.chart.label import (  # type: ignore[import]
        DataLabelList as WolfxlDataLabelList,
    )

    wchart.dataLabels = WolfxlDataLabelList(showVal=True, position="outEnd")
    ws.add_chart(wchart, "E2")
    wpath = tmp_path / "wolfxl.xlsx"
    wb.save(wpath)

    owb = openpyxl.Workbook()
    ows = owb.active
    _seed(ows)
    ochart = build_identical_chart("openpyxl", "bar", ows)
    ochart.dataLabels = DataLabelList(showVal=True, position="outEnd")
    ows.add_chart(ochart, "E2")
    opath = tmp_path / "openpyxl.xlsx"
    owb.save(opath)

    ok, diff = structurally_equivalent(
        first_chart_xml(wpath), first_chart_xml(opath)
    )
    assert ok, f"data label position diverged:\n{diff}"


def test_error_bars_fixedVal_xml_matches_openpyxl(tmp_path: Path) -> None:
    """``ErrorBars(errBarType='both', errValType='fixedVal', val=2)`` parity."""
    from openpyxl.chart.error_bar import ErrorBars as OErrorBars

    wb = wolfxl.Workbook()
    ws = wb.active
    _seed(ws)
    wchart = build_identical_chart("wolfxl", "bar", ws)
    from wolfxl.chart.error_bar import ErrorBars as WErrorBars  # type: ignore[import]

    wchart.series[0].errBars = WErrorBars(
        errBarType="both", errValType="fixedVal", val=2
    )
    ws.add_chart(wchart, "E2")
    wpath = tmp_path / "wolfxl.xlsx"
    wb.save(wpath)

    owb = openpyxl.Workbook()
    ows = owb.active
    _seed(ows)
    ochart = build_identical_chart("openpyxl", "bar", ows)
    ochart.series[0].errBars = OErrorBars(
        errBarType="both", errValType="fixedVal", val=2
    )
    ows.add_chart(ochart, "E2")
    opath = tmp_path / "openpyxl.xlsx"
    owb.save(opath)

    ok, diff = structurally_equivalent(
        first_chart_xml(wpath), first_chart_xml(opath)
    )
    assert ok, f"error bars fixedVal diverged:\n{diff}"


def test_trendline_polynomial_order3_xml_matches_openpyxl(tmp_path: Path) -> None:
    """``Trendline(trendlineType='poly', order=3)`` parity."""
    from openpyxl.chart.trendline import Trendline as OTrendline

    wb = wolfxl.Workbook()
    ws = wb.active
    _seed(ws)
    wchart = build_identical_chart("wolfxl", "scatter", ws)
    from wolfxl.chart.trendline import Trendline as WTrendline  # type: ignore[import]

    wchart.series[0].trendline = WTrendline(trendlineType="poly", order=3)
    ws.add_chart(wchart, "E2")
    wpath = tmp_path / "wolfxl.xlsx"
    wb.save(wpath)

    owb = openpyxl.Workbook()
    ows = owb.active
    _seed(ows)
    ochart = build_identical_chart("openpyxl", "scatter", ows)
    ochart.series[0].trendline = OTrendline(trendlineType="poly", order=3)
    ows.add_chart(ochart, "E2")
    opath = tmp_path / "openpyxl.xlsx"
    owb.save(opath)

    ok, diff = structurally_equivalent(
        first_chart_xml(wpath), first_chart_xml(opath)
    )
    assert ok, f"polynomial trendline order=3 diverged:\n{diff}"


# ---------------------------------------------------------------------------
# 4. LibreOffice headless smoke (gated, 8 cases)
# ---------------------------------------------------------------------------


_RUN_ENV_FLAG = "WOLFXL_RUN_LIBREOFFICE_SMOKE"
_SOFFICE_PATHS = (
    "/Applications/LibreOffice.app/Contents/MacOS/soffice",
    "/usr/bin/soffice",
    "/usr/local/bin/soffice",
)


def _find_soffice() -> str | None:
    for candidate in _SOFFICE_PATHS:
        if Path(candidate).is_file() and os.access(candidate, os.X_OK):
            return candidate
    return shutil.which("soffice")


@pytest.mark.libreoffice_smoke
@pytest.mark.skipif(
    os.environ.get(_RUN_ENV_FLAG) != "1",
    reason=f"LibreOffice smoke is opt-in via {_RUN_ENV_FLAG}=1.",
)
@pytest.mark.parametrize("kind", CHART_KINDS)
def test_libreoffice_renders_wolfxl_chart(tmp_path: Path, kind: str) -> None:
    """A wolfxl-written chart-bearing xlsx must convert cleanly to PDF.

    No structural assertion on the PDF itself — just that ``soffice``
    exits 0 and produces a non-empty file. A failure here means the
    chart XML is malformed enough that LibreOffice's own loader chokes.
    """
    soffice = _find_soffice()
    if soffice is None:
        pytest.skip("soffice not installed")

    wb = wolfxl.Workbook()
    ws = wb.active
    _seed(ws)
    chart = build_identical_chart("wolfxl", kind, ws, features={"title": f"{kind}-lo"})
    ws.add_chart(chart, "E2")
    src = tmp_path / "in.xlsx"
    wb.save(src)

    proc = subprocess.run(
        [soffice, "--headless", "--convert-to", "pdf", "--outdir", str(tmp_path), str(src)],
        capture_output=True,
        text=True,
        timeout=60,
    )
    assert proc.returncode == 0, (
        f"soffice failed for {kind}: stdout={proc.stdout} stderr={proc.stderr}"
    )
    pdf = tmp_path / "in.pdf"
    assert pdf.exists() and pdf.stat().st_size > 0, (
        f"PDF not produced for {kind}"
    )


# ---------------------------------------------------------------------------
# Sanity — extract / canonicalize work on a chart-bearing xlsx
# ---------------------------------------------------------------------------


def test_extract_chart_xml_returns_chart_parts(tmp_path: Path) -> None:
    """Smoke: ``extract_chart_xml`` returns the chart XMLs from a saved file."""
    wb = wolfxl.Workbook()
    ws = wb.active
    _seed(ws)
    chart = build_identical_chart("wolfxl", "bar", ws)
    ws.add_chart(chart, "E2")
    out = tmp_path / "out.xlsx"
    wb.save(out)

    parts = extract_chart_xml(out)
    assert "xl/charts/chart1.xml" in parts
    assert parts["xl/charts/chart1.xml"].startswith(b"<?xml") or parts[
        "xl/charts/chart1.xml"
    ].startswith(b"<")


def test_canonicalize_handles_axis_id_drift(tmp_path: Path) -> None:
    """Two XML strings that differ only in axId values should canonicalise equal."""
    a = (
        b'<c:chart xmlns:c="http://schemas.openxmlformats.org/drawingml/2006/chart">'
        b'<c:axId val="123"/></c:chart>'
    )
    b = (
        b'<c:chart xmlns:c="http://schemas.openxmlformats.org/drawingml/2006/chart">'
        b'<c:axId val="456"/></c:chart>'
    )
    assert canonicalize_chart_xml(a) == canonicalize_chart_xml(b)
