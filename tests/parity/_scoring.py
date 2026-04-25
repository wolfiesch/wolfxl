"""Scoring tiers for the parity harness.

Borrowed from ExcelBench's tiered fidelity model. Each dimension declares
the set of attributes that participate and the tier that governs them:

* ``HARD`` - any mismatch fails CI immediately.
* ``SOFT`` - count is tracked in ``ratchet.json``; the count is never
  allowed to *increase* over the committed baseline (regressions block CI).
* ``INFO`` - reported only, not gated.

See ``Plans`` section "Pass semantics" for the full table.

W4C extension: ``compare_two_workbooks(path_a, path_b)`` reuses the same
tier machinery for the differential writer harness — both args are read
through openpyxl, then walked dimension-by-dimension. Layer 3 of
``tests/diffwriter/`` calls this and asserts ``hard_failures() == []``.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from enum import Enum
from pathlib import Path
from typing import Any


class Tier(str, Enum):
    """``str`` mixin keeps ``Tier.HARD == "hard"`` true on Python 3.9-3.10
    (``enum.StrEnum`` is 3.11+)."""

    HARD = "hard"
    SOFT = "soft"
    INFO = "info"


# Dimension-to-tier map. When adding a dimension, classify it here or the
# harness will fail loud — there is no implicit default.
DIMENSION_TIERS: dict[str, Tier] = {
    # ---- HARD (any mismatch fails CI) ----------------------------------
    "value": Tier.HARD,
    "number_format": Tier.HARD,
    "merged_cells": Tier.HARD,
    "max_row": Tier.HARD,
    "max_col": Tier.HARD,
    "freeze_panes": Tier.HARD,
    "defined_name.refers_to": Tier.HARD,
    "column_width": Tier.HARD,
    "sheet_names": Tier.HARD,
    "utils.get_column_letter": Tier.HARD,
    "utils.column_index_from_string": Tier.HARD,
    "utils.range_boundaries": Tier.HARD,
    "utils.coordinate_to_tuple": Tier.HARD,
    "utils.is_date_format": Tier.HARD,
    "utils.from_excel": Tier.HARD,

    # ---- SOFT (tracked in ratchet, never allowed to drop) --------------
    "font.name": Tier.SOFT,
    "font.size": Tier.SOFT,
    "font.bold": Tier.SOFT,
    "font.italic": Tier.SOFT,
    "fill.fg_color": Tier.SOFT,
    "border.style": Tier.SOFT,
    "border.color": Tier.SOFT,
    "alignment.horizontal": Tier.SOFT,
    "alignment.vertical": Tier.SOFT,
    "alignment.wrap_text": Tier.SOFT,

    # ---- INFO (reported only) ------------------------------------------
    "rich_text_runs": Tier.INFO,
    "row_height": Tier.INFO,
    "tab_color": Tier.INFO,
    "font.underline": Tier.INFO,
    "font.strike": Tier.INFO,
}


@dataclass
class Mismatch:
    """A single point-wise difference between openpyxl and wolfxl."""

    dimension: str
    location: str
    """Human-readable locator, e.g. ``'Sheet1!A1'`` or ``'defined_name:Total'``."""
    openpyxl_value: Any
    wolfxl_value: Any

    @property
    def tier(self) -> Tier:
        return DIMENSION_TIERS.get(self.dimension, Tier.HARD)

    def __str__(self) -> str:  # pragma: no cover - diagnostic only
        return (
            f"[{self.tier.value.upper():4s}] {self.dimension} @ {self.location}: "
            f"openpyxl={self.openpyxl_value!r}  wolfxl={self.wolfxl_value!r}"
        )


@dataclass
class ParityReport:
    """Aggregate of all mismatches for a single fixture (or a test call)."""

    fixture_id: str
    mismatches: list[Mismatch] = field(default_factory=list)

    def record(
        self,
        dimension: str,
        location: str,
        openpyxl_value: Any,
        wolfxl_value: Any,
    ) -> None:
        if _values_equal(openpyxl_value, wolfxl_value):
            return
        self.mismatches.append(
            Mismatch(
                dimension=dimension,
                location=location,
                openpyxl_value=openpyxl_value,
                wolfxl_value=wolfxl_value,
            )
        )

    def hard_failures(self) -> list[Mismatch]:
        return [m for m in self.mismatches if m.tier is Tier.HARD]

    def soft_failures(self) -> list[Mismatch]:
        return [m for m in self.mismatches if m.tier is Tier.SOFT]

    def info_notes(self) -> list[Mismatch]:
        return [m for m in self.mismatches if m.tier is Tier.INFO]

    def summary(self) -> dict[str, int]:
        return {
            "hard": len(self.hard_failures()),
            "soft": len(self.soft_failures()),
            "info": len(self.info_notes()),
        }


def _values_equal(a: Any, b: Any) -> bool:
    """Fuzzy equality for floats, tolerant of None/empty-string ambiguity."""
    if a is None and b == "":
        return True
    if b is None and a == "":
        return True
    if isinstance(a, float) and isinstance(b, float):
        # NaN equality
        if a != a and b != b:
            return True
        return abs(a - b) <= max(1e-9, 1e-9 * max(abs(a), abs(b)))
    if isinstance(a, float) and isinstance(b, int):
        return _values_equal(a, float(b))
    if isinstance(b, float) and isinstance(a, int):
        return _values_equal(float(a), b)
    return a == b


def _normalize_number_format(v: Any) -> Any:
    """openpyxl returns ``'General'`` for unformatted cells; coerce ``None``
    to match so the dual-backend comparison doesn't trip on that quirk."""
    if v is None:
        return "General"
    return v


def _resolve_defined_name(defn: Any) -> str | None:
    """Strip the leading ``=`` from a ``DefinedName.value``-style ref string."""
    refers = getattr(defn, "value", defn)
    if isinstance(refers, str) and refers.startswith("="):
        refers = refers[1:]
    return refers


def compare_two_workbooks(path_a: Path, path_b: Path) -> ParityReport:
    """Open ``path_a`` and ``path_b`` via openpyxl and produce a tiered
    ``ParityReport``. Used by the differential writer harness Layer 3.

    Walks every common sheet, compares dimensions / freeze panes /
    merged cells / per-cell value+number_format / defined names, and
    feeds each comparison through ``ParityReport.record`` so the
    HARD/SOFT/INFO tiers from ``DIMENSION_TIERS`` apply.
    """
    import openpyxl  # local import: keeps ``_scoring.py`` importable when
                     # openpyxl isn't installed (e.g. minimal test runs).

    wb_a = openpyxl.load_workbook(str(path_a), data_only=False)
    wb_b = openpyxl.load_workbook(str(path_b), data_only=False)
    report = ParityReport(fixture_id=f"{path_a.name}_vs_{path_b.name}")
    try:
        a_names = list(wb_a.sheetnames)
        b_names = list(wb_b.sheetnames)
        report.record("sheet_names", "workbook", a_names, b_names)

        common = [n for n in a_names if n in b_names]
        for sheet in common:
            _compare_sheet(report, wb_a[sheet], wb_b[sheet], sheet)

        a_named = {
            name: _resolve_defined_name(defn) for name, defn in wb_a.defined_names.items()
        }
        b_named = {
            name: _resolve_defined_name(defn) for name, defn in wb_b.defined_names.items()
        }
        for name in set(a_named) | set(b_named):
            report.record(
                "defined_name.refers_to",
                f"defined_name:{name}",
                a_named.get(name),
                b_named.get(name),
            )
    finally:
        wb_a.close()
        wb_b.close()
    return report


_MAX_CELLS_PER_SHEET = 10_000


def _compare_sheet(
    report: ParityReport, ws_a: Any, ws_b: Any, sheet_name: str,
) -> None:
    report.record("max_row", f"{sheet_name}:dim", ws_a.max_row, ws_b.max_row)
    report.record("max_col", f"{sheet_name}:dim", ws_a.max_column, ws_b.max_column)
    report.record(
        "freeze_panes",
        f"{sheet_name}:freeze",
        ws_a.freeze_panes,
        ws_b.freeze_panes,
    )

    a_merged = {str(r) for r in ws_a.merged_cells.ranges}
    b_merged = {str(r) for r in ws_b.merged_cells.ranges}
    report.record("merged_cells", f"{sheet_name}:merged", a_merged, b_merged)

    # Column widths must be measured BEFORE the cell loop — the
    # ``_MAX_CELLS_PER_SHEET`` early return below skips anything that
    # follows on large sheets, and ``column_width`` is HARD-tier so we
    # cannot let it silently fall through. ``column_dimensions`` is a
    # dict-like that lazily creates entries on key access; iterate the
    # union of explicitly-set columns from both sides only.
    a_widths = {
        col: dim.width
        for col, dim in ws_a.column_dimensions.items()
        if dim.width is not None
    }
    b_widths = {
        col: dim.width
        for col, dim in ws_b.column_dimensions.items()
        if dim.width is not None
    }
    for col in sorted(set(a_widths) | set(b_widths)):
        report.record(
            "column_width",
            f"{sheet_name}:col:{col}",
            a_widths.get(col),
            b_widths.get(col),
        )

    cells_seen = 0
    for row in ws_a.iter_rows():
        for cell_a in row:
            if cell_a.value is None and getattr(cell_a, "style_id", 0) == 0:
                continue
            if cells_seen >= _MAX_CELLS_PER_SHEET:
                return
            cells_seen += 1
            coord = cell_a.coordinate
            cell_b = ws_b[coord]
            report.record("value", f"{sheet_name}!{coord}", cell_a.value, cell_b.value)
            report.record(
                "number_format",
                f"{sheet_name}!{coord}",
                _normalize_number_format(cell_a.number_format),
                _normalize_number_format(cell_b.number_format),
            )
