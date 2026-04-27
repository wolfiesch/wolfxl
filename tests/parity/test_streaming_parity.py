"""Parity: ``wolfxl.load_workbook(..., read_only=True)`` ↔ openpyxl read_only=True.

Sprint Ι Pod-β contract: the streaming reader must yield value tuples
that round-trip to the same Python representation as openpyxl's own
read-only generator. Style attributes (``font.bold``, ``fill.fgColor``,
``number_format``, ``alignment.horizontal``) must match for any cell
visited via ``iter_rows()``.

We deliberately keep the parity surface tight: cached formula values,
inline strings, and shared strings. Rich-text and nested ``<r>`` runs
are out of scope for Pod-β — see the Phase 3 KNOWN_GAPS row.
"""

from __future__ import annotations

import datetime as dt
from pathlib import Path

import openpyxl
import pytest
from openpyxl.styles import Alignment, Font, PatternFill

import wolfxl


def _make_fixture(path: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Parity"
    # Row 1 — strings (SST), numbers, bool.
    ws["A1"] = "alpha"
    ws["B1"] = "beta"
    ws["C1"] = 1
    ws["D1"] = 2.5
    ws["E1"] = True
    # Row 2 — repeated strings (forces SST hits).
    ws["A2"] = "alpha"
    ws["B2"] = "alpha"
    # Row 3 — date / datetime.
    ws["A3"] = dt.datetime(2024, 6, 1, 12, 0, 0)
    # Row 4 — formula with cached numeric value (note: openpyxl's read_only
    # doesn't evaluate; both layers should report the formula text).
    ws["A4"] = "=C1+D1"
    # Styled cells.
    ws["A1"].font = Font(bold=True)
    ws["B1"].fill = PatternFill(fill_type="solid", fgColor="FFFFCC00")
    ws["D1"].number_format = "0.00"
    ws["A2"].alignment = Alignment(horizontal="center")
    wb.save(path)
    return path


@pytest.fixture
def parity_xlsx(tmp_path: Path) -> Path:
    return _make_fixture(tmp_path / "parity.xlsx")


# ---------------------------------------------------------------------------


def test_values_only_tuples_match_openpyxl(parity_xlsx: Path) -> None:
    op_wb = openpyxl.load_workbook(parity_xlsx, read_only=True)
    op_ws = op_wb["Parity"]
    op_rows = list(
        op_ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=5, values_only=True)
    )

    wx_wb = wolfxl.load_workbook(parity_xlsx, read_only=True)
    wx_ws = wx_wb["Parity"]
    wx_rows = list(
        wx_ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=5, values_only=True)
    )

    assert len(wx_rows) == len(op_rows)
    # Rows 1+2 cover SST/repeat/numeric/bool — full parity.
    # Row 3 is a datetime cell (Excel serial number) and Row 4 is a
    # formula whose cached value diverges between openpyxl
    # (returns the formula text in read_only mode) and the wolfxl
    # streamer's Phase-1 representation. Both are tracked in
    # ``KNOWN_GAPS.md`` Phase-3 (rich-text + datetime conversion in
    # streaming) and are out-of-scope for Pod-β.
    for op_r, wx_r in zip(op_rows, wx_rows, strict=False):
        for op_v, wx_v in zip(op_r, wx_r, strict=False):
            if isinstance(op_v, float) and isinstance(wx_v, float):
                assert wx_v == pytest.approx(op_v)
            else:
                assert wx_v == op_v


def test_cell_font_bold_parity(parity_xlsx: Path) -> None:
    op_wb = openpyxl.load_workbook(parity_xlsx, read_only=True)
    op_ws = op_wb["Parity"]
    op_a1 = next(iter(op_ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=1)))[0]

    wx_wb = wolfxl.load_workbook(parity_xlsx, read_only=True)
    wx_ws = wx_wb["Parity"]
    wx_a1 = next(iter(wx_ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=1)))[0]

    assert wx_a1.font.bold == bool(op_a1.font.bold)


def test_cell_number_format_parity(parity_xlsx: Path) -> None:
    op_wb = openpyxl.load_workbook(parity_xlsx, read_only=True)
    op_ws = op_wb["Parity"]
    op_d1 = next(iter(op_ws.iter_rows(min_row=1, max_row=1, min_col=4, max_col=4)))[0]

    wx_wb = wolfxl.load_workbook(parity_xlsx, read_only=True)
    wx_ws = wx_wb["Parity"]
    wx_d1 = next(iter(wx_ws.iter_rows(min_row=1, max_row=1, min_col=4, max_col=4)))[0]

    assert wx_d1.number_format == op_d1.number_format


def test_cell_alignment_parity(parity_xlsx: Path) -> None:
    op_wb = openpyxl.load_workbook(parity_xlsx, read_only=True)
    op_ws = op_wb["Parity"]
    op_a2 = next(iter(op_ws.iter_rows(min_row=2, max_row=2, min_col=1, max_col=1)))[0]

    wx_wb = wolfxl.load_workbook(parity_xlsx, read_only=True)
    wx_ws = wx_wb["Parity"]
    wx_a2 = next(iter(wx_ws.iter_rows(min_row=2, max_row=2, min_col=1, max_col=1)))[0]

    assert wx_a2.alignment.horizontal == op_a2.alignment.horizontal


def test_row_count_matches(parity_xlsx: Path) -> None:
    op_wb = openpyxl.load_workbook(parity_xlsx, read_only=True)
    op_ws = op_wb["Parity"]
    op_count = sum(1 for _ in op_ws.iter_rows(values_only=True))

    wx_wb = wolfxl.load_workbook(parity_xlsx, read_only=True)
    wx_ws = wx_wb["Parity"]
    wx_count = sum(1 for _ in wx_ws.iter_rows(values_only=True))

    assert wx_count == op_count


# ---------------------------------------------------------------------------
# Sprint Λ Pod-γ — datetime divergence in streaming reads.
# Cells whose number format is a date format must surface as Python
# datetimes via both ``values_only=True`` and ``StreamingCell.value``.
# ---------------------------------------------------------------------------


def test_streaming_values_only_datetime_matches_openpyxl(parity_xlsx: Path) -> None:
    """Row 3 column A is ``datetime(2024, 6, 1, 12, 0, 0)``. openpyxl's
    read-only path returns a ``datetime``; wolfxl must too."""
    op_wb = openpyxl.load_workbook(parity_xlsx, read_only=True)
    op_ws = op_wb["Parity"]
    op_a3 = next(
        iter(op_ws.iter_rows(min_row=3, max_row=3, min_col=1, max_col=1, values_only=True))
    )[0]

    wx_wb = wolfxl.load_workbook(parity_xlsx, read_only=True)
    wx_ws = wx_wb["Parity"]
    wx_a3 = next(
        iter(wx_ws.iter_rows(min_row=3, max_row=3, min_col=1, max_col=1, values_only=True))
    )[0]

    assert isinstance(op_a3, dt.datetime), f"openpyxl baseline changed: {op_a3!r}"
    assert isinstance(wx_a3, dt.datetime), (
        f"wolfxl streaming values_only datetime divergence: got {type(wx_a3).__name__} "
        f"({wx_a3!r}), expected datetime to match openpyxl read_only=True"
    )
    assert wx_a3 == op_a3


def test_streaming_cell_value_datetime_matches_openpyxl(parity_xlsx: Path) -> None:
    """Same row via ``StreamingCell.value`` (values_only=False)."""
    op_wb = openpyxl.load_workbook(parity_xlsx, read_only=True)
    op_ws = op_wb["Parity"]
    op_a3 = next(
        iter(op_ws.iter_rows(min_row=3, max_row=3, min_col=1, max_col=1))
    )[0]

    wx_wb = wolfxl.load_workbook(parity_xlsx, read_only=True)
    wx_ws = wx_wb["Parity"]
    wx_a3 = next(
        iter(wx_ws.iter_rows(min_row=3, max_row=3, min_col=1, max_col=1))
    )[0]

    assert isinstance(op_a3.value, dt.datetime), f"openpyxl baseline changed: {op_a3.value!r}"
    assert isinstance(wx_a3.value, dt.datetime), (
        f"wolfxl streaming StreamingCell.value datetime divergence: got "
        f"{type(wx_a3.value).__name__} ({wx_a3.value!r})"
    )
    assert wx_a3.value == op_a3.value
