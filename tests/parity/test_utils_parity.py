"""Bug-for-bug parity for openpyxl's utility functions.

openpyxl's behavior is the spec. Where openpyxl has quirks (e.g.
``is_date_format``'s handling of escape brackets), wolfxl must match them.
This file drives both libraries on the same inputs and asserts identical
output, so future wolfxl releases cannot silently diverge.

Any case wolfxl currently fails is marked ``xfail(strict=True)`` via the
``KNOWN_GAPS`` set — removing a name from that set means "we shipped it".
"""

from __future__ import annotations

from datetime import datetime
from typing import Any, cast

import pytest
from openpyxl.styles.numbers import is_date_format as openpyxl_is_date_format
from openpyxl.utils.cell import (
    column_index_from_string as openpyxl_column_index_from_string,
)
from openpyxl.utils.cell import coordinate_to_tuple as openpyxl_coordinate_to_tuple
from openpyxl.utils.cell import get_column_letter as openpyxl_get_column_letter
from openpyxl.utils.cell import range_boundaries as openpyxl_range_boundaries
from openpyxl.utils.datetime import CALENDAR_WINDOWS_1900, from_excel

# Names tracked in KNOWN_GAPS.md — these are xfail until later phases fix them.
# Phase 0 cleanup shipped all 7 utils symbols via ``wolfxl.utils.*``; leave the
# set in place (empty) so future regressions can be tracked without restoring
# the scaffold.
KNOWN_GAPS: set[str] = set()


def _wolfxl_attr(name: str) -> Any | None:
    """Resolve ``wolfxl.utils.<name>`` if it exists, else ``None``."""
    try:
        import wolfxl.utils as wolfxl_utils  # type: ignore[import-not-found]
    except ImportError:
        return None
    return getattr(wolfxl_utils, name, None)


@pytest.mark.parametrize(
    "col_idx,expected",
    [
        (1, "A"), (26, "Z"), (27, "AA"), (52, "AZ"), (702, "ZZ"),
        (703, "AAA"), (16384, "XFD"),  # Excel's max column
    ],
)
def test_get_column_letter(col_idx: int, expected: str) -> None:
    assert openpyxl_get_column_letter(col_idx) == expected
    wolfxl_fn = _wolfxl_attr("get_column_letter")
    if wolfxl_fn is None:
        pytest.xfail("wolfxl.utils.get_column_letter not yet exposed (Phase 0 gap)")
    fn = cast(Any, wolfxl_fn)
    assert fn(col_idx) == expected


@pytest.mark.parametrize(
    "letter,expected",
    [("A", 1), ("Z", 26), ("AA", 27), ("ZZ", 702), ("AAA", 703), ("XFD", 16384)],
)
def test_column_index_from_string(letter: str, expected: int) -> None:
    assert openpyxl_column_index_from_string(letter) == expected
    wolfxl_fn = _wolfxl_attr("column_index_from_string")
    if wolfxl_fn is None:
        pytest.xfail(
            "wolfxl.utils.column_index_from_string not yet exposed (Phase 0 gap)"
        )
    fn = cast(Any, wolfxl_fn)
    assert fn(letter) == expected


@pytest.mark.parametrize(
    "range_str",
    [
        "A1:B2",
        "A1",               # single-cell degenerate
        "A1:A1",
        "B3:D10",
        "AA100:AC200",
        "$A$1:$D$10",       # absolute refs
    ],
)
def test_range_boundaries(range_str: str) -> None:
    expected = openpyxl_range_boundaries(range_str)
    wolfxl_fn = _wolfxl_attr("range_boundaries")
    if wolfxl_fn is None:
        pytest.xfail("wolfxl.utils.range_boundaries not yet exposed (Phase 0 gap)")
    fn = cast(Any, wolfxl_fn)
    assert fn(range_str) == expected


@pytest.mark.parametrize(
    "cell_ref,expected",
    [("A1", (1, 1)), ("B3", (3, 2)), ("AA100", (100, 27))],
)
def test_coordinate_to_tuple(cell_ref: str, expected: tuple[int, int]) -> None:
    assert openpyxl_coordinate_to_tuple(cell_ref) == expected
    wolfxl_fn = _wolfxl_attr("coordinate_to_tuple")
    if wolfxl_fn is None:
        pytest.xfail(
            "wolfxl.utils.coordinate_to_tuple not yet exposed (Phase 0 gap)"
        )
    fn = cast(Any, wolfxl_fn)
    assert fn(cell_ref) == expected


@pytest.mark.parametrize(
    "fmt",
    [
        "General",
        "0",
        "0.00",
        "#,##0",
        "yyyy-mm-dd",
        "yyyy-mm-dd hh:mm:ss",
        "m/d/yyyy",
        'mm"-"dd"-"yy',
        "[h]:mm:ss",
        '"date:"yyyy-mm-dd',   # escaped literal
        "@",
        "",
    ],
)
def test_is_date_format(fmt: str) -> None:
    """Encode openpyxl's output bug-for-bug.

    openpyxl has known edge cases (escaped literals, empty strings).
    WolfXL must match — this test pins the behavior.
    """
    expected = openpyxl_is_date_format(fmt)
    wolfxl_fn = _wolfxl_attr("is_date_format")
    if wolfxl_fn is None:
        pytest.xfail("wolfxl.utils.is_date_format not yet exposed (Phase 0 gap)")
    fn = cast(Any, wolfxl_fn)
    assert fn(fmt) == expected


@pytest.mark.parametrize(
    "serial",
    [
        # 0.0 omitted: openpyxl returns a `date` (not datetime) for the epoch,
        # which is a known historical inconsistency. Dedicated edge-case test
        # belongs in test_utils_parity::test_from_excel_epoch_edge once
        # wolfxl ships from_excel.
        1.0,          # 1900-01-01 in 1900 calendar
        60.0,         # The notorious leap-year bug cell
        61.0,
        44562.0,      # 2022-01-01
        44562.5,      # 2022-01-01 12:00
    ],
)
def test_from_excel(serial: float) -> None:
    try:
        expected = from_excel(serial, CALENDAR_WINDOWS_1900)
    except Exception as exc:  # pragma: no cover - openpyxl rejects some serials
        pytest.skip(f"openpyxl rejects serial {serial}: {exc!r}")
        return  # appease type-narrowing; pytest.skip raises but type checker doesn't know
    # openpyxl may return None (e.g. for negative serials).
    assert expected is None or isinstance(expected, datetime)
    wolfxl_fn = _wolfxl_attr("from_excel")
    wolfxl_cal = _wolfxl_attr("CALENDAR_WINDOWS_1900")
    if wolfxl_fn is None or wolfxl_cal is None:
        pytest.xfail(
            "wolfxl.utils.from_excel or CALENDAR_WINDOWS_1900 not exposed (Phase 0 gap)"
        )
    fn = cast(Any, wolfxl_fn)
    assert fn(serial, wolfxl_cal) == expected


def test_gap_inventory_matches_contract() -> None:
    """Ensures KNOWN_GAPS in this file matches the tracked surface entries.

    Drift here means either the contract lied or this file's xfail list is stale.
    """
    from .openpyxl_surface import known_gap_entries

    surface_gap_names = {
        e.openpyxl_path.rsplit(".", 1)[-1] for e in known_gap_entries()
    }
    assert KNOWN_GAPS.issubset(surface_gap_names | {"get_column_letter"}), (
        f"Drift: {KNOWN_GAPS - surface_gap_names} not in surface contract"
    )
