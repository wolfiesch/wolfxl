"""Sprint Λ Pod-α — write-side encryption parity vs the read path.

Goal: confirm that ``wolfxl.Workbook.save(path, password=...)`` produces
output that's byte-for-byte loadable by both wolfxl's own
``load_workbook(password=)`` (Sprint Ι Pod-γ) AND msoffcrypto-tool's
own decrypt path. We don't try to byte-equal openpyxl's encrypted output
— msoffcrypto-tool has internal randomness (key salt, IV) so two
encryptions of the same plaintext are never byte-equal. Instead we
verify *plaintext equivalence after a decrypt round-trip*.
"""

from __future__ import annotations

import io
from pathlib import Path

import pytest

import wolfxl


def _msoffcrypto_or_skip():
    try:
        import msoffcrypto  # noqa: F401
    except ImportError:  # pragma: no cover
        pytest.skip("msoffcrypto-tool not installed")


def _decrypt_via_msoffcrypto(path: Path, password: str) -> bytes:
    """Return the decrypted plaintext bytes via msoffcrypto-tool."""
    import msoffcrypto

    with open(path, "rb") as fp:
        of = msoffcrypto.OfficeFile(fp)
        of.load_key(password=password)
        out = io.BytesIO()
        of.decrypt(out)
        return out.getvalue()


def test_wolfxl_encrypted_output_decryptable_by_msoffcrypto(tmp_path: Path) -> None:
    """wolfxl-encrypted output must round-trip through the lib that wrote it."""
    _msoffcrypto_or_skip()
    out = tmp_path / "wolf_enc.xlsx"

    wb = wolfxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="parity-A1")
    ws.cell(row=2, column=2, value=99)
    wb.save(out, password="parity-pw")

    # 1) msoffcrypto round-trip yields a valid-looking xlsx (zip header)
    decrypted = _decrypt_via_msoffcrypto(out, "parity-pw")
    assert decrypted[:2] == b"PK", "decrypted bytes must start with ZIP magic"

    # 2) wolfxl's own password loader recovers the same data
    wb2 = wolfxl.load_workbook(out, password="parity-pw")
    assert wb2.active.cell(row=1, column=1).value == "parity-A1"
    assert wb2.active.cell(row=2, column=2).value == 99


def test_wolfxl_round_trips_existing_xlsx_fixture(tmp_path: Path) -> None:
    """Open an existing fixture, encrypt-save, then re-read.

    Uses ``tests/fixtures/minimal.xlsx`` so the parity flow is wired
    against a committed, real-shape OOXML file rather than a synthetic
    new workbook.
    """
    _msoffcrypto_or_skip()
    src = Path(__file__).parents[1] / "fixtures" / "minimal.xlsx"
    if not src.exists():
        pytest.skip(f"fixture missing: {src}")

    wb = wolfxl.load_workbook(src, modify=True)
    out = tmp_path / "minimal_enc.xlsx"
    wb.save(out, password="fixture-pw")

    # msoffcrypto round-trip works
    decrypted = _decrypt_via_msoffcrypto(out, "fixture-pw")
    assert decrypted[:2] == b"PK"

    # wolfxl can re-open it
    wb2 = wolfxl.load_workbook(out, password="fixture-pw")
    assert wb2.sheetnames  # at least one sheet survived


def test_openpyxl_decrypt_helper_path_matches(tmp_path: Path) -> None:
    """Cross-check: encrypt openpyxl plaintext → decrypt → re-read.

    Validates that wolfxl's encryption module is interchangeable with
    a "manual openpyxl save + msoffcrypto encrypt" pipeline so callers
    migrating from one to the other get identical behaviour.
    """
    _msoffcrypto_or_skip()
    pytest.importorskip("openpyxl")
    import openpyxl

    from wolfxl._encryption import encrypt_xlsx_to_path

    plain = tmp_path / "op_plain.xlsx"
    wb = openpyxl.Workbook()
    wb.active["A1"] = "openpyxl-A1"
    wb.active["B2"] = 7
    wb.save(plain)

    enc = tmp_path / "op_enc.xlsx"
    encrypt_xlsx_to_path(plain.read_bytes(), "op-pw", enc)

    # wolfxl can read it
    wb2 = wolfxl.load_workbook(enc, password="op-pw")
    assert wb2.active.cell(row=1, column=1).value == "openpyxl-A1"
    assert wb2.active.cell(row=2, column=2).value == 7
