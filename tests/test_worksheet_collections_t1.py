"""T1 PR2 — worksheet-level read parity for tables, DVs, and CF.

Same cross-library pattern as PR1: openpyxl builds the fixture, wolfxl
reads it. Each collection is also checked for the lazy-cache contract.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import pytest
from wolfxl.formatting import ConditionalFormatting, ConditionalFormattingList
from wolfxl.formatting.rule import CellIsRule, FormulaRule, Rule
from wolfxl.worksheet.datavalidation import DataValidation, DataValidationList
from wolfxl.worksheet.table import Table, TableStyleInfo

from wolfxl import Workbook

openpyxl = pytest.importorskip("openpyxl")


@pytest.fixture()
def rich_fixture(tmp_path: Path) -> Path:
    """Build a workbook with a table, a DV list, and a cellIs CF rule."""
    from openpyxl.formatting.rule import CellIsRule as XCellIsRule  # noqa: N814
    from openpyxl.formatting.rule import FormulaRule as XFormulaRule  # noqa: N814
    from openpyxl.worksheet.datavalidation import DataValidation as XDV  # noqa: N814
    from openpyxl.worksheet.table import Table as XTable  # noqa: N814
    from openpyxl.worksheet.table import TableStyleInfo as XStyle  # noqa: N814

    path = tmp_path / "rich.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    # Header row + data for the table.
    ws["A1"] = "Name"
    ws["B1"] = "Region"
    ws["C1"] = "Sales"
    ws["D1"] = "Notes"
    for i in range(2, 11):
        ws[f"A{i}"] = f"Row{i}"
        ws[f"B{i}"] = "North"
        ws[f"C{i}"] = i * 10
        ws[f"D{i}"] = ""

    # A named table on A1:D10.
    table = XTable(displayName="SalesTable", ref="A1:D10")
    table.tableStyleInfo = XStyle(
        name="TableStyleLight9",
        showRowStripes=True,
    )
    ws.add_table(table)

    # DV on A2:A10 — list of allowed values.
    dv = XDV(type="list", formula1='"Red,Blue,Green"', allow_blank=True)
    dv.add("A2:A10")
    ws.add_data_validation(dv)

    # CellIs CF on C2:C10 — highlight when > 50.
    ws.conditional_formatting.add(
        "C2:C10",
        XCellIsRule(operator="greaterThan", formula=["50"]),
    )
    # A second formula-based CF on the same range.
    ws.conditional_formatting.add(
        "C2:C10",
        XFormulaRule(formula=["$C2=100"]),
    )

    wb.save(path)
    return path


def test_tables_read(rich_fixture: Path) -> None:
    wb = Workbook._from_reader(str(rich_fixture))
    ws = wb["Data"]
    assert list(ws.tables.keys()) == ["SalesTable"]
    t = ws.tables["SalesTable"]
    assert isinstance(t, Table)
    assert t.name == "SalesTable"
    assert t.ref == "A1:D10"
    assert t.headerRowCount == 1
    assert t.totalsRowCount == 0
    # At least the column names should round-trip.
    assert len(t.tableColumns) == 4
    names = [c.name for c in t.tableColumns]
    assert "Name" in names
    assert "Sales" in names
    # TableStyleInfo present — style name round-trips.
    assert t.tableStyleInfo is not None
    assert isinstance(t.tableStyleInfo, TableStyleInfo)
    assert t.tableStyleInfo.name == "TableStyleLight9"


def test_data_validations_read(rich_fixture: Path) -> None:
    wb = Workbook._from_reader(str(rich_fixture))
    ws = wb["Data"]
    dvs = ws.data_validations
    assert isinstance(dvs, DataValidationList)
    assert len(dvs) == 1
    dv = dvs.dataValidation[0]
    assert isinstance(dv, DataValidation)
    assert dv.type == "list"
    # formula1 round-trips (with the leading ``=`` Rust appends).
    assert dv.formula1 is not None
    assert '"Red,Blue,Green"' in dv.formula1
    assert "A2:A10" in dv.sqref


def test_conditional_formatting_read(rich_fixture: Path) -> None:
    wb = Workbook._from_reader(str(rich_fixture))
    ws = wb["Data"]
    cf = ws.conditional_formatting
    assert isinstance(cf, ConditionalFormattingList)
    entries = list(cf)
    # Both rules attach to C2:C10 — openpyxl groups them into one entry.
    assert len(entries) == 1
    entry = entries[0]
    assert isinstance(entry, ConditionalFormatting)
    assert entry.sqref == "C2:C10"
    # Two rules: the CellIs comparison + the formula rule.
    assert len(entry.rules) == 2
    types = {r.type for r in entry.rules}
    assert types == {"cellIs", "expression"}

    # `.cfRule` is the openpyxl alias for `.rules`.
    assert entry.cfRule is entry.rules


def test_empty_collections_in_write_mode() -> None:
    """A fresh Workbook() has no tables/DVs/CF — collections read as empty."""
    wb = Workbook()
    ws = wb.active
    assert ws.tables == {}
    assert len(ws.data_validations) == 0
    assert len(ws.conditional_formatting) == 0
    assert list(ws.tables.keys()) == []


class _CountingReader:
    """Forwards attribute access and counts per-method calls we care about."""

    def __init__(self, inner: Any) -> None:
        self._inner = inner
        self.table_calls = 0
        self.dv_calls = 0
        self.cf_calls = 0

    def read_tables(self, sheet: str) -> Any:
        self.table_calls += 1
        return self._inner.read_tables(sheet)

    def read_data_validations(self, sheet: str) -> Any:
        self.dv_calls += 1
        return self._inner.read_data_validations(sheet)

    def read_conditional_formats(self, sheet: str) -> Any:
        self.cf_calls += 1
        return self._inner.read_conditional_formats(sheet)

    def __getattr__(self, name: str) -> Any:
        return getattr(self._inner, name)


def test_collection_reads_are_single_shot(rich_fixture: Path) -> None:
    """Repeatedly accessing each collection hits Rust exactly once per sheet."""
    wb = Workbook._from_reader(str(rich_fixture))
    counter = _CountingReader(wb._rust_reader)
    wb._rust_reader = counter

    ws = wb["Data"]
    for _ in range(10):
        _ = ws.tables
        _ = ws.data_validations
        _ = ws.conditional_formatting
        # Also touch the underlying lists/dicts
        _ = list(ws.tables.values())
        _ = list(ws.data_validations)
        _ = list(ws.conditional_formatting)

    assert counter.table_calls == 1
    assert counter.dv_calls == 1
    assert counter.cf_calls == 1


def test_cf_write_mode_add_queues(rich_fixture: Path) -> None:  # noqa: ARG001
    """In write mode, ``cf.add()`` queues but doesn't yet round-trip (PR5)."""
    wb = Workbook()
    ws = wb.active
    # Attempt to queue a rule — PR5 will make this fully round-trip.
    ws.conditional_formatting.add("A1:A10", CellIsRule(operator="greaterThan", formula=["5"]))
    entries = list(ws.conditional_formatting)
    assert len(entries) == 1
    assert entries[0].sqref == "A1:A10"
    assert len(entries[0].rules) == 1


def test_dv_list_append_queues_in_write_mode() -> None:
    wb = Workbook()
    ws = wb.active
    # Before the first call, DVL is lazy — accessing it creates the container.
    dvs = ws.data_validations
    assert len(dvs) == 0
    dvs.append(DataValidation(type="list", formula1='"a,b,c"', sqref="A1:A10"))
    assert len(dvs) == 1


def test_dv_list_append_works_in_modify_mode(rich_fixture: Path) -> None:
    """RFC-025 shipped: ``append`` queues onto ``_pending_data_validations``
    in modify mode, same as write mode. The patcher flushes them on
    ``save()``. This test only asserts the queue side; round-trip
    coverage lives in ``tests/test_modify_data_validations.py``.
    """
    wb = Workbook._from_patcher(str(rich_fixture))
    ws = wb["Data"]
    before = len(ws.data_validations)
    ws.data_validations.append(
        DataValidation(type="whole", operator="greaterThan", formula1="=10", sqref="B1:B10")
    )
    assert len(ws.data_validations) == before + 1
    assert len(ws._pending_data_validations) == 1  # noqa: SLF001


def test_cf_add_in_modify_mode_raises(rich_fixture: Path) -> None:
    wb = Workbook._from_patcher(str(rich_fixture))
    ws = wb["Data"]
    with pytest.raises(NotImplementedError, match="T1.5"):
        ws.conditional_formatting.add(
            "D2:D10", FormulaRule(formula=["$D2>0"])
        )


def test_rule_types_round_trip() -> None:
    """Each specific rule constructor sets the right ``type`` tag."""
    cr = CellIsRule(operator="equal", formula=["42"])
    assert cr.type == "cellIs"
    assert isinstance(cr, Rule)

    fr = FormulaRule(formula=["TRUE"])
    assert fr.type == "expression"
