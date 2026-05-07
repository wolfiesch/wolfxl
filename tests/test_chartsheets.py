"""Chartsheet authoring parity smoke tests."""

from __future__ import annotations

import zipfile
from pathlib import Path

import pytest

import wolfxl
from wolfxl.chart import BarChart, Reference


def _workbook_with_chart() -> tuple[wolfxl.Workbook, BarChart]:
    wb = wolfxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["month", "sales"])
    ws.append(["Jan", 10])
    ws.append(["Feb", 20])
    chart = BarChart()
    chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=3), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=3))
    return wb, chart


def test_create_chartsheet_with_chart_round_trips_through_openpyxl(tmp_path: Path) -> None:
    openpyxl = pytest.importorskip("openpyxl")

    wb, chart = _workbook_with_chart()
    cs = wb.create_chartsheet("Sales Chart")
    cs.add_chart(chart)
    out = tmp_path / "chartsheet.xlsx"
    wb.save(out)

    op = openpyxl.load_workbook(out)
    assert op.sheetnames == ["Sheet", "Sales Chart"]
    assert len(op.chartsheets) == 1
    assert op.chartsheets[0].title == "Sales Chart"
    assert len(op.chartsheets[0]._charts) == 1  # noqa: SLF001

    with zipfile.ZipFile(out) as z:
        names = set(z.namelist())
        assert "xl/chartsheets/sheet1.xml" in names
        assert "xl/chartsheets/_rels/sheet1.xml.rels" in names
        assert "xl/drawings/drawing1.xml" in names
        assert "xl/drawings/_rels/drawing1.xml.rels" in names
        assert "xl/charts/chart1.xml" in names
        workbook_rels = z.read("xl/_rels/workbook.xml.rels").decode()
        assert "/relationships/chartsheet" in workbook_rels
        content_types = z.read("[Content_Types].xml").decode()
        assert "/xl/chartsheets/sheet1.xml" in content_types


def test_eager_chartsheet_survives_second_save(tmp_path: Path) -> None:
    openpyxl = pytest.importorskip("openpyxl")

    wb, chart = _workbook_with_chart()
    cs = wb.create_chartsheet("Sales Chart")
    cs.add_chart(chart)
    first = tmp_path / "first.xlsx"
    second = tmp_path / "second.xlsx"
    wb.save(first)
    wb.active["C1"] = "resaved"
    wb.save(second)

    op = openpyxl.load_workbook(second)
    assert op.sheetnames == ["Sheet", "Sales Chart"]
    assert [cs.title for cs in op.chartsheets] == ["Sales Chart"]
    assert len(op.chartsheets[0]._charts) == 1  # noqa: SLF001


def test_create_chartsheet_in_modify_mode_round_trips(tmp_path: Path) -> None:
    openpyxl = pytest.importorskip("openpyxl")

    src = tmp_path / "source.xlsx"
    op = openpyxl.Workbook()
    ws = op.active
    ws.title = "Data"
    ws.append(["month", "sales"])
    ws.append(["Jan", 10])
    ws.append(["Feb", 20])
    op.save(src)

    wb = wolfxl.load_workbook(src, modify=True)
    data = wb["Data"]
    chart = BarChart()
    chart.add_data(Reference(data, min_col=2, min_row=1, max_row=3), titles_from_data=True)
    chart.set_categories(Reference(data, min_col=1, min_row=2, max_row=3))
    cs = wb.create_chartsheet("Sales Chart")
    cs.add_chart(chart)
    out = tmp_path / "modify_chartsheet.xlsx"
    wb.save(out)

    reloaded = openpyxl.load_workbook(out)
    assert reloaded.sheetnames == ["Data", "Sales Chart"]
    assert [cs.title for cs in reloaded.chartsheets] == ["Sales Chart"]
    assert len(reloaded.chartsheets[0]._charts) == 1  # noqa: SLF001


def test_chartsheet_can_be_inserted_between_worksheets(tmp_path: Path) -> None:
    openpyxl = pytest.importorskip("openpyxl")

    wb, chart = _workbook_with_chart()
    wb.create_sheet("After")
    cs = wb.create_chartsheet("Middle", index=1)
    cs.add_chart(chart)
    out = tmp_path / "chartsheet_middle.xlsx"
    wb.save(out)

    op = openpyxl.load_workbook(out)
    assert op.sheetnames == ["Sheet", "Middle", "After"]
    assert [ws.title for ws in op.worksheets] == ["Sheet", "After"]
    assert [cs.title for cs in op.chartsheets] == ["Middle"]


def test_chartsheet_insert_remaps_sheet_local_defined_names(tmp_path: Path) -> None:
    openpyxl = pytest.importorskip("openpyxl")
    from openpyxl.workbook.defined_name import DefinedName

    src = tmp_path / "defined_name_source.xlsx"
    op = openpyxl.Workbook()
    op.active.title = "First"
    op.create_sheet("Second")
    op.defined_names["SecondLocal"] = DefinedName(
        "SecondLocal",
        attr_text="Second!$A$1",
        localSheetId=1,
    )
    op.save(src)

    wb = wolfxl.load_workbook(src, modify=True)
    wb.create_chartsheet("Chart", index=1)
    out = tmp_path / "defined_name_with_chartsheet.xlsx"
    wb.save(out)

    reloaded = openpyxl.load_workbook(out)
    assert reloaded.sheetnames == ["First", "Chart", "Second"]
    with zipfile.ZipFile(out) as z:
        workbook_xml = z.read("xl/workbook.xml").decode()
    assert 'name="SecondLocal" localSheetId="2"' in workbook_xml


def test_empty_chartsheet_saves_like_openpyxl(tmp_path: Path) -> None:
    openpyxl = pytest.importorskip("openpyxl")

    wb = wolfxl.Workbook()
    wb.create_chartsheet("Empty Chart")
    out = tmp_path / "empty_chartsheet.xlsx"
    wb.save(out)

    op = openpyxl.load_workbook(out)
    assert op.sheetnames == ["Sheet", "Empty Chart"]
    assert [cs.title for cs in op.chartsheets] == ["Empty Chart"]

    with zipfile.ZipFile(out) as z:
        names = set(z.namelist())
        assert "xl/chartsheets/sheet1.xml" in names
        assert "xl/chartsheets/_rels/sheet1.xml.rels" in names
        rels = z.read("xl/chartsheets/_rels/sheet1.xml.rels").decode()
        assert "<Relationship " not in rels


def test_openpyxl_authored_chartsheet_loads_as_chartsheet(tmp_path: Path) -> None:
    openpyxl = pytest.importorskip("openpyxl")

    src = tmp_path / "openpyxl_chartsheet.xlsx"
    op = openpyxl.Workbook()
    ws = op.active
    ws.append(["month", "sales"])
    ws.append(["Jan", 10])
    ws.append(["Feb", 20])
    chart = openpyxl.chart.BarChart()
    chart.add_data(openpyxl.chart.Reference(ws, min_col=2, min_row=1, max_row=3))
    cs = op.create_chartsheet("ChartOnly")
    cs.add_chart(chart)
    op.save(src)

    wb = wolfxl.load_workbook(src)
    assert wb.sheetnames == ["Sheet", "ChartOnly"]
    assert [ws.title for ws in wb.worksheets] == ["Sheet"]
    assert [cs.title for cs in wb.chartsheets] == ["ChartOnly"]
    assert wb["ChartOnly"] is wb.chartsheets[0]
    assert len(wb.chartsheets[0]._charts) == 1  # noqa: SLF001


def test_openpyxl_authored_chartsheet_is_preserved_on_modify_save(tmp_path: Path) -> None:
    openpyxl = pytest.importorskip("openpyxl")

    src = tmp_path / "source.xlsx"
    op = openpyxl.Workbook()
    ws = op.active
    ws.append(["month", "sales"])
    ws.append(["Jan", 10])
    ws.append(["Feb", 20])
    chart = openpyxl.chart.BarChart()
    chart.add_data(openpyxl.chart.Reference(ws, min_col=2, min_row=1, max_row=3))
    cs = op.create_chartsheet("ChartOnly")
    cs.add_chart(chart)
    op.save(src)

    wb = wolfxl.load_workbook(src, modify=True)
    wb["Sheet"]["C1"] = "edited"
    out = tmp_path / "modified.xlsx"
    wb.save(out)

    reloaded = openpyxl.load_workbook(out)
    assert reloaded["Sheet"]["C1"].value == "edited"
    assert [cs.title for cs in reloaded.chartsheets] == ["ChartOnly"]
    assert len(reloaded.chartsheets[0]._charts) == 1  # noqa: SLF001
