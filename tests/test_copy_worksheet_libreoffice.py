"""RFC-035 — LibreOffice cross-renderer manual gate.

Layer-4 of the §6 verification matrix. LibreOffice rendering is a
**manual gate**: a wolfxl-saved xlsx (with copy_worksheet applied)
must open in LibreOffice without "Repair the workbook?" dialogs and
must render the cloned sheet identically to the source.

The test is gated behind the ``WOLFXL_RUN_LIBREOFFICE_SMOKE=1``
environment variable (mirroring ``tests/diffwriter/soffice_smoke.py``)
and skipped otherwise. When the env var is set:

1. A canonical copy_worksheet fixture is built (5×5 grid + table +
   external hyperlink + sheet-scope Print_Area).
2. WolfXL clones the active sheet and saves the result.
3. ``soffice --headless --convert-to xlsx`` round-trips the file.
4. ``soffice --headless --convert-to pdf`` produces a PDF (visual
   regression check by the developer).
5. The xlsx round-trip is asserted: exit code 0, no
   ``corrupt|repaired|error`` in stderr, valid output zip.

Manual rendering verification (the human gate):

  WOLFXL_RUN_LIBREOFFICE_SMOKE=1 pytest tests/test_copy_worksheet_libreoffice.py -v -s

  Then open the resulting PDF (printed path in test output) and
  visually inspect:

  - The cloned sheet renders identically to the source (cells,
    fonts, table styling, hyperlink underline).
  - The cloned sheet appears in the tab bar with the expected name
    (``Template Copy``).
  - The original sheet's Print_Area is unchanged; the clone's
    Print_Area covers the same range but is scoped to the clone.

If any visual regression is observed, capture a screenshot and
attach it to the PR description.

Setup (one-time):

  brew install --cask libreoffice    # macOS
  apt-get install libreoffice        # Linux

This test is informational — it does NOT gate CI. It runs only when
the developer explicitly opts in.
"""

from __future__ import annotations

import os
import shutil
import subprocess
import zipfile
from pathlib import Path

import openpyxl
import pytest

from wolfxl import load_workbook

pytestmark = pytest.mark.rfc035


_RUN_ENV_FLAG = "WOLFXL_RUN_LIBREOFFICE_SMOKE"
_SOFFICE_PATHS = (
    "/Applications/LibreOffice.app/Contents/MacOS/soffice",
    "/usr/bin/soffice",
    "/usr/local/bin/soffice",
)
_SMOKE_KEYWORDS = ("corrupt", "repaired", "error")
_SUBPROCESS_TIMEOUT_S = 60


def _find_soffice() -> str | None:
    for candidate in _SOFFICE_PATHS:
        if Path(candidate).is_file() and os.access(candidate, os.X_OK):
            return candidate
    return shutil.which("soffice")


@pytest.fixture(autouse=True)
def _force_test_epoch(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("WOLFXL_TEST_EPOCH", "0")


def _build_canonical_copy_fixture(path: Path) -> None:
    from openpyxl.workbook.defined_name import DefinedName as XDefinedName
    from openpyxl.worksheet.table import Table

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"
    headers = ["k", "a", "b", "c", "d"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    for r in range(2, 6):
        for c in range(1, 6):
            ws.cell(row=r, column=c, value=(r - 1) * 10 + c)
    ws.add_table(Table(displayName="Sales", ref="A1:E5"))
    ws["E5"] = "click"
    ws["E5"].hyperlink = "https://example.com/lo"
    wb.defined_names["_xlnm.Print_Area"] = XDefinedName(
        "_xlnm.Print_Area", attr_text="Template!$A$1:$E$5", localSheetId=0
    )
    wb.save(path)


@pytest.mark.manual
@pytest.mark.skipif(
    os.environ.get(_RUN_ENV_FLAG) != "1",
    reason=(
        f"LibreOffice cross-renderer test is opt-in. Set "
        f"{_RUN_ENV_FLAG}=1 and ensure ``soffice`` is on PATH (or "
        "installed at the platform default location) to run."
    ),
)
def test_copy_worksheet_libreoffice_round_trip(tmp_path: Path) -> None:
    """A wolfxl-cloned xlsx must round-trip through LibreOffice headless.

    Procedure (executed inline):
    1. Build the canonical fixture.
    2. ``wolfxl.load_workbook(modify=True).copy_worksheet(...).save(...)``
    3. ``soffice --headless --convert-to xlsx --outdir /tmp/lo source.xlsx``
    4. ``soffice --headless --convert-to pdf --outdir /tmp/lo source.xlsx``
    5. Assert: zero exit code, no corrupt/repaired/error in stderr,
       valid output zip, valid PDF (>0 bytes).

    The PDF path is printed for the developer to inspect manually.
    """
    soffice = _find_soffice()
    if not soffice:
        pytest.skip("soffice not found on this machine")

    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    _build_canonical_copy_fixture(src)

    # Step 2: wolfxl copy_worksheet + save
    wb = load_workbook(src, modify=True)
    wb.copy_worksheet(wb.active)
    wb.save(out)

    # Step 3: xlsx round-trip
    converted_dir = tmp_path / "lo_xlsx"
    converted_dir.mkdir()
    proc = subprocess.run(
        [
            soffice,
            "--headless",
            "--convert-to",
            "xlsx",
            "--outdir",
            str(converted_dir),
            str(out),
        ],
        capture_output=True,
        text=True,
        timeout=_SUBPROCESS_TIMEOUT_S,
    )
    assert proc.returncode == 0, (
        f"soffice xlsx round-trip exit {proc.returncode}: "
        f"{proc.stderr[:300]}"
    )
    stderr_lc = proc.stderr.lower()
    for kw in _SMOKE_KEYWORDS:
        assert kw not in stderr_lc, (
            f"soffice stderr contains {kw!r}: {proc.stderr[:300]}"
        )
    converted_xlsx = converted_dir / out.name
    assert converted_xlsx.is_file(), (
        f"soffice produced no output at {converted_xlsx}"
    )
    assert zipfile.is_zipfile(converted_xlsx), (
        "soffice output is not a valid zip"
    )

    # Step 4: PDF for visual inspection
    pdf_dir = tmp_path / "lo_pdf"
    pdf_dir.mkdir()
    proc_pdf = subprocess.run(
        [
            soffice,
            "--headless",
            "--convert-to",
            "pdf",
            "--outdir",
            str(pdf_dir),
            str(out),
        ],
        capture_output=True,
        text=True,
        timeout=_SUBPROCESS_TIMEOUT_S,
    )
    assert proc_pdf.returncode == 0, (
        f"soffice pdf conversion exit {proc_pdf.returncode}: "
        f"{proc_pdf.stderr[:300]}"
    )
    pdf_path = pdf_dir / out.with_suffix(".pdf").name
    assert pdf_path.is_file(), f"no PDF at {pdf_path}"
    assert pdf_path.stat().st_size > 0, "PDF is empty"

    # Print path for manual inspection.
    print(f"\n[RFC-035] LibreOffice PDF for visual review: {pdf_path}")
    print(
        "  Visual checks: (1) cloned sheet renders identically to source; "
        "(2) tab bar shows 'Template Copy'; (3) Print_Area survives on clone."
    )


@pytest.mark.manual
def test_libreoffice_procedure_documented() -> None:
    """The LibreOffice cross-renderer procedure is a documented part
    of RFC-035 §6 row 4. This test exists to surface the procedure
    via ``pytest -v`` so a developer running the harness sees what
    the manual gate involves.

    Procedure summary:
    1. ``brew install --cask libreoffice`` (or apt-get install).
    2. ``WOLFXL_RUN_LIBREOFFICE_SMOKE=1 pytest tests/test_copy_worksheet_libreoffice.py -v -s``
    3. Inspect the PDF path printed by the test for visual regressions.
    4. Capture a screenshot if any visual divergence is found and
       attach to the PR description.

    This test is always green; its purpose is to ensure the procedure
    appears in test output (via the docstring + assertion message).
    """
    procedure = (
        "RFC-035 LibreOffice cross-renderer: see the test_copy_worksheet_libreoffice "
        "module docstring. Run with WOLFXL_RUN_LIBREOFFICE_SMOKE=1 to engage the gate."
    )
    assert procedure  # always truthy; the docstring carries the contract.
